"""Add peg_ratios + cheap_on_growth tabs to screener_report.xlsx.

peg_ratios tab     — every ticker scored, sortable by any ratio
cheap_on_growth   — strictest filter: name has at least one PEG-style
                     ratio < 1.5 AND rev growth >= 10% AND 1y perf
                     in [-50%, +20%] AND market cap >= $200M.
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

XLSX = Path('screener_report.xlsx')

def load_meta():
    meta = {}
    for f in ['universe_us_wide.csv','universe_expanded.csv','universe_wider.csv',
              'universe_eu.csv','universe_eu_extra.csv','universe_canada.csv',
              'universe_us_large_mega.csv',
              'universe_japan.csv','universe_korea.csv','universe_hongkong.csv',
              'universe_australia.csv','universe_india.csv']:
        p = Path(f)
        if not p.exists(): continue
        try:
            d = pd.read_csv(p, usecols=lambda c: c in ['symbol','name','sector','industry','country'], low_memory=False)
            for _, r in d.iterrows():
                tk = r.get('symbol')
                if not isinstance(tk, str): continue
                k = tk.upper()
                if k in meta: continue
                meta[k] = {f: r.get(f) for f in ['name','sector','industry','country'] if f in d.columns}
                # also store with safe-name form (underscores)
                import re
                safe_k = re.sub(r'\.', '_', k)
                if safe_k != k and safe_k not in meta:
                    meta[safe_k] = meta[k]
        except Exception: pass
    return meta

META = load_meta()

def enrich(df, ticker_col='ticker'):
    for col in ('name','sector','industry','country'):
        df[col] = df[ticker_col].astype(str).str.upper().map(lambda t: META.get(t, {}).get(col))
    return df

# Full PEG-ratios tab (all 641 scored)
all_df = pd.read_csv('results_peg/all.csv')
all_df = enrich(all_df)

# Sanity filter: drop bogus rows
all_df = all_df[
    (all_df['gross_margin_now_pct'].fillna(0) <= 100) &
    (all_df['gross_margin_now_pct'].fillna(-99) > -50)
]

peg_cols = ['ticker','name','sector','industry','country',
            'perf_1y_pct','market_cap','gross_margin_now_pct',
            # Growth rates — LTM and latest-year side by side
            'rev_growth_ltm_pct','rev_growth_yr_pct',
            'gross_growth_ltm_pct','gross_growth_yr_pct',
            'ebitda_growth_ltm_pct','ebitda_growth_yr_pct',
            # Raw multiples
            'trailingPE','priceToSales','evToSales','evToEbitda','evToGrossProfit',
            # PEG-style ratios — LTM growth
            'PEG_ltm','PSG_ltm',
            'EV_Sales_over_revG_ltm','EV_GP_over_GPg_ltm','EV_EBITDA_over_EBg_ltm',
            # New: harder-to-manipulate divisor (gross profit growth)
            'EV_EBITDA_over_GPg_ltm','PE_over_GPg_ltm','PS_over_GPg_ltm',
            # Same set using latest-year growth
            'PSG_yr','EV_Sales_over_revG_yr','EV_GP_over_GPg_yr','EV_EBITDA_over_EBg_yr',
            'EV_EBITDA_over_GPg_yr','PE_over_GPg_yr','PS_over_GPg_yr',
            # Dollar amounts
            'rev_ltm_M','rev_yr_M','gross_ltm_M','gross_yr_M','ebitda_ltm_M','ebitda_yr_M']
peg_cols = [c for c in peg_cols if c in all_df.columns]
peg_tab = all_df[peg_cols].copy()

# cheap_on_growth: strict filter for the most attractive
cheap = peg_tab[
    (peg_tab[['rev_growth_ltm_pct','rev_growth_yr_pct']].max(axis=1).fillna(-99) >= 10) &
    (peg_tab['perf_1y_pct'].between(-50, 20)) &
    (peg_tab['market_cap'] >= 200e6) &
    (
        peg_tab['EV_GP_over_GPg_ltm'].fillna(99).between(0.01, 1.5) |
        peg_tab['EV_GP_over_GPg_yr'].fillna(99).between(0.01, 1.5) |
        peg_tab['EV_EBITDA_over_EBg_ltm'].fillna(99).between(0.01, 1.5) |
        peg_tab['EV_EBITDA_over_EBg_yr'].fillna(99).between(0.01, 1.5) |
        peg_tab['PSG_ltm'].fillna(99).between(0.01, 1.5) |
        peg_tab['PSG_yr'].fillna(99).between(0.01, 1.5)
    )
].copy()
# Composite = best (lowest) ratio across the six "value per growth" variants
ratio_cols = ['EV_GP_over_GPg_ltm','EV_GP_over_GPg_yr',
              'EV_EBITDA_over_EBg_ltm','EV_EBITDA_over_EBg_yr',
              'PSG_ltm','PSG_yr']
cheap['best_peg_score'] = cheap[ratio_cols].apply(
    lambda r: min([v for v in r if pd.notna(v) and v > 0], default=99), axis=1)
cheap = cheap.sort_values('best_peg_score', ascending=True).head(500)

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    peg_tab.sort_values('EV_GP_g', ascending=True).to_excel(xw, sheet_name='peg_ratios', index=False)
    cheap.to_excel(xw, sheet_name='cheap_on_growth', index=False)
    for sname, w in [('peg_ratios','B2'), ('cheap_on_growth','B2')]:
        ws = xw.sheets[sname]
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 26
        for c in 'EFGHIJKLMNOPQRSTU': ws.column_dimensions[c].width = 13
        ws.freeze_panes = w

# Reorder: cheap_on_growth at top, peg_ratios at end
wb = load_workbook(XLSX)
desired_order = ['forensic_audit','cheap_on_growth','clean_topline','unpriced_op_leverage',
                 'asymmetry_tier','stealth_breakout']
present = [s for s in desired_order if s in wb.sheetnames]
for i, s in enumerate(present):
    cur = wb.sheetnames.index(s)
    wb.move_sheet(s, offset=i - cur)
wb.save(XLSX)

print(f'Wrote peg_ratios ({len(peg_tab)} rows) and cheap_on_growth ({len(cheap)} rows)')
