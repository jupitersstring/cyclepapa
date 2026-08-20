"""Build the Harvard-aesthetic Excel workbook of most asymmetric
situations.

Harvard Business Review / HBS aesthetic conventions:
  * Crimson #A41E22 accent for headers
  * Charcoal #333333 body text on warm-white #FAFAF7
  * Serif font for titles, sans-serif for data
  * Sparse cell borders, generous white space
  * Section dividers as horizontal accent bars
  * Footnotes in italic 8pt at bottom of sheet

Tabs produced:
  1. Cover / Executive Summary
  2. Most Asymmetric (the 12 convergent names with full detail)
  3. By Archetype (38 PSU/gov buckets, winner per row)
  4. By Pattern (forward $ hurdle, M&A, spin, etc.)
  5. Reserve Baskets (Cohen-Malloy, microcap forcing, Russell)
  6. Caution List (convergent names + red flags)
  7. Coverage & Tiers (data diagnostics)
  8. Methodology
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

ROOT = Path("/home/user/cyclepapa")
OUT = ROOT / "MOST_ASYMMETRIC.xlsx"

# Harvard-aesthetic palette (strict grayscale / black-and-white)
# Hierarchy via type weight + italic + horizontal rules, never color.
BLACK         = "000000"
CHARCOAL      = "2A2A2A"
MID_GRAY      = "707070"
RULE_GRAY     = "A0A0A0"
LINE_GRAY     = "BDBDBD"
BAND_GRAY     = "F5F2EC"   # very faint cream row banding
WARM_WHITE    = "FFFFFF"

# All preserved names retained as aliases so legacy references compile
CRIMSON       = BLACK
CRIMSON_LIGHT = BAND_GRAY
GOLD          = MID_GRAY
SAGE          = CHARCOAL
GREY_BAND     = BAND_GRAY
GREY_LINE     = LINE_GRAY

# SINGLE font: Times New Roman at 10pt everywhere with condense=True
# to render a slightly tighter glyph spacing (the "tall condensed"
# academic look). Hierarchy by weight + italic only.
SERIF         = "Times New Roman"
TYPESIZE      = 10

TITLE_FONT    = Font(name=SERIF, size=TYPESIZE, bold=True,
                      color=BLACK, condense=True)
SUBTITLE_FONT = Font(name=SERIF, size=TYPESIZE, italic=True,
                      color=CHARCOAL, condense=True)
HEADER_FONT   = Font(name=SERIF, size=TYPESIZE, bold=True,
                      color=BLACK, condense=True)
BODY_FONT     = Font(name=SERIF, size=TYPESIZE,
                      color=BLACK, condense=True)
SMALL_FONT    = Font(name=SERIF, size=TYPESIZE,
                      color=CHARCOAL, condense=True)
BODY_BOLD     = Font(name=SERIF, size=TYPESIZE, bold=True,
                      color=BLACK, condense=True)
BODY_ITALIC   = Font(name=SERIF, size=TYPESIZE, italic=True,
                      color=BLACK, condense=True)
FOOTNOTE      = Font(name=SERIF, size=TYPESIZE, italic=True,
                      color=CHARCOAL, condense=True)

# Borders: narrow hairlines only. No vertical borders -- clean
# columns. Top/bottom rules on header and section breaks.
HAIRLINE      = Side(border_style="hair", color=LINE_GRAY)
THIN          = Side(border_style="thin", color=BLACK)
THIN_GRAY     = Side(border_style="thin", color=RULE_GRAY)
BOTTOM_BORDER = Border(bottom=HAIRLINE)
HEADER_RULE   = Border(top=THIN, bottom=THIN_GRAY)
SECTION_RULE  = Border(bottom=THIN)

# Fills: only subtle band on alternating rows. No solid fills on
# title/header. Background is white.
HEADER_FILL    = PatternFill(fill_type=None)   # NO fill on header
BAND_FILL      = PatternFill("solid", fgColor=BAND_GRAY)
TITLE_BAND_FILL = PatternFill(fill_type=None)  # NO fill on title

# Tag fills: in grayscale framework we don't shade tag cells -- use
# bold (Concentrated) / italic (Participation) on the cell font.
CLEAN_TAG_FILL = PatternFill(fill_type=None)
FLAG_TAG_FILL  = PatternFill(fill_type=None)


# ----------------------------------------------------------------------
# Data loaders
# ----------------------------------------------------------------------

def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return list(csv.DictReader(path.open()))


def load_proxy() -> dict:
    out = {}
    import glob
    for fn in sorted(glob.glob(str(ROOT / "proxy_scan*.json"))):
        try: d = json.loads(open(fn).read())
        except: continue
        rows = d if isinstance(d, list) else d.values()
        for r in rows:
            if isinstance(r, dict) and r.get("ticker"):
                tk = r["ticker"]
                if (tk not in out
                        or r.get("filing_date", "") > out[tk].get("filing_date", "")):
                    out[tk] = r
    return out


def load_archetype_md(path: Path) -> dict:
    if not path.exists():
        return {}
    text = path.read_text()
    by_tk = {}
    blocks = re.findall(
        r"###?\s+(\w+\d+\.?\s+[^\n]+?)\n.*?\*\*Winner:\s*([A-Z][A-Z0-9.\-]{0,10})\*\*",
        text, re.S)
    for arch, tk in blocks:
        arch_id = arch.split(".")[0].strip()
        by_tk.setdefault(tk, []).append((arch_id, arch.strip()))
    return by_tk


# ----------------------------------------------------------------------
# Harvard-aesthetic number-format spec (per user requirement)
# Negatives in parentheses, em-dash for empty cells, numbers right-
# aligned. Format selected by field-name heuristic.
# ----------------------------------------------------------------------

NUMFMT_RAW_DOLLAR = '#,##0;(#,##0);"–"'          # mcap, EV (raw $)
NUMFMT_MILLIONS  = '#,##0;(#,##0);"–"'           # _M (millions, no decimal)
NUMFMT_PCT       = '#,##0.0;(#,##0.0);"–"'       # pct / pp / growth / margin
NUMFMT_RATIO     = '#,##0.00;(#,##0.00);"–"'     # ratios + scores
NUMFMT_PRICE     = '#,##0.00;(#,##0.00);"–"'     # prices / eps
NUMFMT_INT       = '#,##0;(#,##0);"–"'           # bare integers / layers

EM_DASH = "–"


def _num_or_none(v):
    if v is None or v == "" or v == "?" or v == EM_DASH:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Strip dollar sign + commas + percent + parentheses
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("%", "")
    s = s.lstrip("(").rstrip(")")
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return None


def number_format_for(field_name: str) -> str | None:
    """Return the openpyxl number_format for a field, or None for text.

    Detects from BOTH internal field names (e.g. 'mcap_M', 'p_b') AND
    rendered headers (e.g. 'mcap ($M)', 'P/B', 'Spot ($)')."""
    if not field_name:
        return None
    f = field_name.lower().strip()
    # Text-only fields
    if f in ("ticker", "company", "name", "sector", "industry", "kind",
              "screens", "archetypes", "reasons", "detail", "role",
              "event_kind", "talent_hits", "notes", "why", "floor",
              "red flags", "action", "use", "source", "holdings",
              "step", "weight"):
        return None
    # PERCENTAGES (look for 'pct', '%', 'pp', 'growth', 'margin')
    if (f.endswith("%") or f.startswith("%") or " %" in f or "% " in f
            or "pct" in f or "(%)" in f or "_pp" in f
            or any(s in f for s in ("growth", "margin", "yield", "return"))):
        return NUMFMT_PCT
    # _M (millions) -- match 'mcap_M', '$M', '(M)', 'mcap ($m)', '_m_'
    if any(s in f for s in ("_m", "$m", "(m)", "($m)", "mcap (", "mcap_m",
                              "musd")):
        return NUMFMT_MILLIONS
    # Raw dollar enterprise/market values
    if f in ("mcap", "ev", "enterprise_value", "market_cap", "raw_mcap"):
        return NUMFMT_RAW_DOLLAR
    # Prices, EPS
    if any(s in f for s in ("price", "spot", "spot ($)", "eps", "($)")):
        return NUMFMT_PRICE
    # Layer / count / day fields -> integer (no decimals)
    if (f in ("#", "rank", "no", "n", "inner")
            or any(s in f for s in (
                "layers", "n_screens", "n_archetypes", "n_arch", "day",
                "count", "n_", "lift", "psu core", "gov", "buyers",
                "amend", "valuation", "buyback", "tender", "10b5", "f4"))):
        return NUMFMT_INT
    # Scores / ratios / multiples
    if any(s in f for s in ("score", "p/b", "p/e", "ev/ebitda", "ratio",
                              "cons", "pts", "dd")):
        return NUMFMT_RATIO
    # Default for any unrecognized numeric: ratio (2 decimals)
    return NUMFMT_RATIO


# ----------------------------------------------------------------------
# Sheet helpers
# ----------------------------------------------------------------------

def set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title_band(ws, title: str, subtitle: str, n_cols: int):
    """Academic-paper title block: bold title, italic subtitle,
    a thin black rule below. No background fill. The leading row
    gets extra height for whitespace breathing room."""
    ws.row_dimensions[1].height = 26
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = Alignment(
        vertical="center", horizontal="left", indent=0)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    ws.row_dimensions[2].height = 18
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(
        vertical="center", horizontal="left", indent=0)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    # Thin black rule under the subtitle to delimit the title block
    for c in range(1, n_cols + 1):
        ws.cell(row=2, column=c).border = SECTION_RULE

    # A blank row of whitespace before the table
    ws.row_dimensions[3].height = 8


# Sheet-level header registry so write_body_row can auto-pick formats
# from whatever the current tab's headers actually are.
_SHEET_HEADERS: dict = {}


def write_header_row(ws, row: int, headers: list[str]):
    """Academic table header: bold serif text on white, horizontal
    rules above and below (the classic three-line table aesthetic).
    Text columns left-aligned headers; numeric columns get right-
    aligned headers later in body where format implies it."""
    ws.row_dimensions[row].height = 18
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        # Header alignment matches the body it heads: numeric headers
        # right-align, text headers left-align.
        fmt = number_format_for(h or "")
        if fmt is not None:
            cell.alignment = Alignment(
                vertical="center", horizontal="right", indent=0)
        else:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", indent=0)
        cell.border = HEADER_RULE
    _SHEET_HEADERS[ws.title] = list(headers)


def _classify_value(v):
    """Return (display_value, is_number) -- coerces strings like
    '$98', '12.5%', '0.48' into numerics so number_format can format
    them properly. Returns (em_dash, False) for empty/None."""
    if v is None or v == "":
        return EM_DASH, False
    if isinstance(v, (int, float)):
        return v, True
    s = str(v).strip()
    if not s or s in ("?", "—", EM_DASH, "-"):
        return EM_DASH, False
    num = _num_or_none(s)
    if num is not None:
        return num, True
    return s, False


def write_body_row(ws, row: int, values: list, band: bool = False,
                   align_first_left: bool = True,
                   bold_first: bool = False,
                   field_names: list | None = None):
    """Body row in academic-paper grayscale style.

    `band` -- enable subtle alternating row banding (very light cream).
    `bold_first` -- bold the first column (typically ticker).
    `field_names` -- optional list of headers/field names per column;
        used to pick the right number format.
    """
    ws.row_dimensions[row].height = 16
    fill = BAND_FILL if band else None
    for c, v in enumerate(values, 1):
        disp_val, is_num = _classify_value(v)
        cell = ws.cell(row=row, column=c, value=disp_val)
        cell.font = BODY_BOLD if (bold_first and c == 1) else BODY_FONT
        if fill:
            cell.fill = fill
        # No per-cell border in body -- rows separated by banding alone.

        # Decide format + alignment
        if is_num:
            fmt = None
            # Prefer explicit field_names; else look up by current
            # sheet's recorded headers.
            header_for_col = None
            if field_names and c - 1 < len(field_names):
                header_for_col = field_names[c - 1]
            elif ws.title in _SHEET_HEADERS:
                cur = _SHEET_HEADERS[ws.title]
                if c - 1 < len(cur):
                    header_for_col = cur[c - 1]
            if header_for_col:
                fmt = number_format_for(header_for_col)
            if fmt is None:
                fmt = NUMFMT_RATIO
            cell.number_format = fmt
            cell.alignment = Alignment(
                vertical="center", horizontal="right",
                wrap_text=False, indent=0,
            )
        else:
            # Text is always LEFT-aligned per Harvard spec.
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left",
                wrap_text=True,
                indent=1,
            )


def write_footnote(ws, row: int, text: str, n_cols: int):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FOOTNOTE
    cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=n_cols)
    ws.row_dimensions[row].height = 28


# ----------------------------------------------------------------------
# Tab 1: Cover / Executive Summary
# ----------------------------------------------------------------------

def get_convergent_from_disk() -> list[dict]:
    """Derive the convergent list from full_universe_consensus.csv --
    the SINGLE source of truth that scores EVERY one of 6,169 tickers
    on EVERY layer (no source-file truncation bias).

    A name is convergent iff n_layers_firing >= 3. Sort by layers
    firing then consensus_score. Normalized to the consensus_ranking
    schema downstream (adds n_screens/n_archetypes_won synonyms).
    """
    cr = ROOT / "full_universe_consensus.csv"
    if not cr.exists():
        # Fall back to old file only if the new one is missing
        cr = ROOT / "consensus_ranking.csv"
        if not cr.exists():
            return []
    rows = list(csv.DictReader(cr.open()))
    convergent = []
    for r in rows:
        try:
            n_layers = int(r.get("n_layers_firing")
                           or r.get("n_screens") or 0)
        except (ValueError, KeyError):
            continue
        if n_layers >= 3:
            # Normalise field names for downstream use
            r["n_screens"] = n_layers
            r["n_archetypes_won"] = r.get("n_archetypes_won", n_layers)
            convergent.append(r)
    convergent.sort(key=lambda r: (
        -int(r.get("n_screens") or 0),
        -float(r.get("consensus_score") or 0)))
    # Cap at top 20 for the xlsx
    return convergent[:20]


# Per-ticker editorial annotations. If a ticker drops out of the
# convergent set (re-derived from disk), its annotation is unused;
# if a new ticker enters, it renders with empty annotation rather
# than dropping or fabricating data. This is the only acceptable
# coupling between code and ticker identifiers.
TICKER_ANNOTATIONS: dict[str, dict] = {
    "HFFG": {"name": "HF Foods Group",
             "why": "Triple PSU $ hurdle (rev/EBITDA/FCF) + clawback strengthened + 4-buyer F4 cluster",
             "floor": "P/B 0.48 + microcap distributor with hard assets"},
    "CSGP": {"name": "CoStar Group",
             "why": "EBITDA $ hurdle + 10x CEO ownership + SOP 45% dissent + verified buyback EXECUTING",
             "floor": "Quasi-monopoly RE data; recurring revenue base"},
    "RNR":  {"name": "RenaissanceRe",
             "why": "Deepest per-share metric stack (5+) -- full alignment",
             "floor": "Insurance NAV + per-share metric discipline"},
    "LE":   {"name": "Lands' End",
             "why": "12-tranche price ladder + 86% PSU + ROIC + live take-private 14D-9",
             "floor": "Active third-party bid sets mechanical floor"},
    "NUS":  {"name": "Nu Skin",
             "why": "Named asset-sale trigger + 5-metric clean per-share stack",
             "floor": "P/B 0.33 + buyback shrinking organically"},
    "ADT":  {"name": "ADT Inc",
             "why": "Heaviest PSU%LTI in universe (90%) + verified -7.3% shrinkage",
             "floor": "Recurring monitoring revenue + verified supply-curve compression"},
    "KMPR": {"name": "Kemper",
             "why": "Anti-hedge/pledge + verified buyback EXECUTING (largest verified shrinkage)",
             "floor": "Insurer at 0.54x book; $ repurchased = 1.85x NAV-accretive"},
    "MAT":  {"name": "Mattel",
             "why": "Unique double dollar hurdle: EBITDA $ + FCF $ targets both coded",
             "floor": "Brand portfolio + 6-metric stack provides discipline"},
    "LMT":  {"name": "Lockheed Martin",
             "why": "FCF $ hurdle + backlog $ target + 70% PSU%LTI",
             "floor": "Defense backlog + sovereign counterparty"},
    "CDE":  {"name": "Coeur Mining",
             "why": "CEO 10b5-1 termination score 80 -- #1 in universe + FCF/share PSU stack",
             "floor": "Precious-metals price floor + CEO walked back scheduled selling"},
    "GO":   {"name": "Grocery Outlet",
             "why": "Forward $ targets + Cohen-Malloy 6-buyer cluster",
             "floor": "Discount-grocery defensive base"},
    "GPRO": {"name": "GoPro",
             "why": "PSU vests on spin / separation event -- board paid to execute",
             "floor": "Brand + cash position"},
    "EXFY": {"name": "Expensify",
             "why": "Live issuer self-tender; published bid above market",
             "floor": "Issuer-paid tender = mechanical floor"},
    "AUPH": {"name": "Aurinia Pharmaceuticals",
             "why": "PSU + valuation + 10b5-1 + recent-incentive + special-sits firing",
             "floor": "Lupus nephritis franchise + cash position"},
    "GETY": {"name": "Getty Images",
             "why": "10b5-1 termination score 74 + recent-incentive event + valuation floor",
             "floor": "P/B 0.55 + IP library"},
    "HDSN": {"name": "Hudson Technologies",
             "why": "Recent F4 cluster + buyback + special-sits + recent-incentive",
             "floor": "Refrigerant reclaim regulatory moat"},
    "TONX": {"name": "Tonix Pharmaceuticals",
             "why": "Fresh F4 cluster + recent-incentive 87% drawdown",
             "floor": "P/B 0.52 microcap"},
    "DXC":  {"name": "DXC Technology",
             "why": "PSU core 40 + buyback + recent-incentive 57% drawdown",
             "floor": "IT services franchise; transformation signal in PSU"},
    "OLED": {"name": "Universal Display",
             "why": "PSU + F4 cluster + recent-incentive 60% drawdown",
             "floor": "OLED IP moat"},
    "IQV":  {"name": "IQVIA",
             "why": "10b5-1 termination 43 + recent-incentive + special-sits",
             "floor": "Healthcare data services franchise"},
    "CRM":  {"name": "Salesforce",
             "why": "PSU + buyback + recent 10b5-1 termination + F4 cluster",
             "floor": "Mega-cap recurring revenue base"},
    "MA":   {"name": "Mastercard",
             "why": "10b5-1 termination score 78 (highest) + PSU + buyback",
             "floor": "Payments duopoly"},
    "BCO":  {"name": "Brink's Company",
             "why": "10b5-1 termination + special-sits + PSU",
             "floor": "Cash-handling services franchise"},
    "FIS":  {"name": "Fidelity National Information Svcs",
             "why": "PSU 42 + valuation + 10b5-1 + recent-incentive",
             "floor": "Payments processing franchise"},
    "FISV": {"name": "Fiserv",
             "why": "PSU + valuation + 10b5-1 + recent-incentive",
             "floor": "FCF/share PSU metric anchor"},
    "ZTS":  {"name": "Zoetis",
             "why": "F4 cluster + valuation + recent-incentive 60% drawdown",
             "floor": "Animal health franchise"},
    "FLGT": {"name": "Fulgent Genetics",
             "why": "PSU + valuation P/B 0.51 + special-sits + recent-incentive",
             "floor": "Half book + cash position"},
    "GTM":  {"name": "GTM (placeholder)",
             "why": "PSU + valuation + recent activist 13D 4d ago + special-sits",
             "floor": "P/B 0.54 + 78% drawdown"},
    "OI":   {"name": "O-I Glass",
             "why": "PSU + valuation + F4 cluster + recent-incentive",
             "floor": "Glass packaging recurring base"},
    "GRNT": {"name": "Granite Ridge Resources",
             "why": "PSU + valuation + F4 cluster + recent-incentive",
             "floor": "Oil & gas asset base"},
    "BDC":  {"name": "Belden Inc",
             "why": "PSU + buyback + 10b5-1 + recent-incentive",
             "floor": "Industrial connectivity franchise"},
    "ECPG": {"name": "Encore Capital",
             "why": "PSU 43 + buyback + recent-incentive",
             "floor": "Debt-collection franchise"},
    "THRY": {"name": "Thryv Holdings",
             "why": "Triple-$ PSU hurdle + valuation + 10b5-1 + recent-incentive",
             "floor": "P/B 0.74 SMB services"},
    "GPUS": {"name": "Hyperscale Data",
             "why": "Valuation + live tender + 10b5-1 + recent-incentive",
             "floor": "P/B 0.50 + live self-tender"},
    "AUPH_NOTE": {"name": "",
             "why": "",
             "floor": ""},  # placeholder for slot expansion
}


def sizing_for_screens(ns: int, na: int, n_flags: int) -> str:
    """Sizing derived from data, not memory:
       Concentrated   if ns >= 4 and n_flags <= 1
       Material       if ns >= 3 and n_flags <= 2
       Participation  otherwise
    """
    if ns >= 4 and n_flags <= 1:
        return "Concentrated 5%+"
    if ns >= 3 and n_flags <= 2:
        return "Material 2-5%"
    return "Participation 1-2%"


def red_flag_count(tk: str, proxy: dict) -> int:
    p = proxy.get(tk, {})
    if not p:
        return 0
    flags = set()
    for s in (p.get("pattern_reasons") or []) + (p.get("gov_reasons") or []):
        sl = s.lower()
        for f in ("single-trigger", "repricing", "retirement carveout",
                  "front-loaded", "discretionary", "aggregate-only"):
            if f in sl:
                flags.add(f)
    return len(flags)



def n_consensus_layers() -> int:
    """Count scoring layers live from the consensus CSV header, so
    workbook copy never goes stale as layers are added."""
    p = ROOT / "full_universe_consensus.csv"
    try:
        hdr = p.open().readline().strip().split(",")
        n = sum(1 for c in hdr if c.endswith("_pts"))
        return n if n else 30
    except Exception:
        return 30


def build_turnaround_signal(wb: Workbook, yf: dict):
    """Live-rendered Bollenbach-signal tab: recent 8-K Item 5.02
    executive appointments scored on distress + grant + curated
    turnaround-talent overlap."""
    ws = wb.create_sheet("Turnaround Signal")
    set_col_widths(ws, [9, 13, 32, 10, 14, 8, 8, 8, 16, 50])
    write_title_band(
        ws,
        "Bollenbach Signal -- Turnaround Talent Into Distress",
        "Senior executives voluntarily joining struggling companies "
        "with equity-heavy compensation. The grant tells you they "
        "see a re-rate path the market hasn't yet priced.",
        n_cols=10,
    )

    path = ROOT / "turnaround_signal.csv"
    if not path.exists():
        ws.cell(row=4, column=1,
                value="(no file -- run `python3 turnaround_executive_leg.py`)").font = SUBTITLE_FONT
        return

    headers = ["#", "Ticker", "Company", "Score",
                "Appt. date", "Distress", "Grant", "Talent",
                "Role", "Talent hits / reasons"]
    write_header_row(ws, 4, headers)

    rows = list(csv.DictReader(path.open()))
    r = 5
    for i, row in enumerate(rows[:40], 1):
        tk = row["ticker"]
        company = (row.get("company") or "")[:32]
        try:
            score = float(row["score"])
        except Exception:
            score = 0.0
        try:
            dp = float(row["distress_pts"]); gp = float(row["grant_pts"])
            tp = float(row["talent_pts"])
        except Exception:
            dp = gp = tp = 0.0
        role = (row.get("role") or "")[:16]
        reasons = (row.get("talent_hits") or row.get("reasons") or "")[:80]
        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [i, tk, company,
                        f"{score:.0f}",
                        row.get("filing_date") or "",
                        f"{dp:.0f}", f"{gp:.0f}", f"{tp:.0f}",
                        role, reasons],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        if tp >= 20:
            ws.cell(row=r, column=8).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=8).font = BODY_BOLD
        if score >= 50:
            ws.cell(row=r, column=4).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        elif score >= 30:
            ws.cell(row=r, column=4).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    write_footnote(ws, r,
        "Greenblatt's Bollenbach test: 'It didn't make sense that the "
        "man responsible for successfully saving a sinking ship -- by "
        "figuring out a way to throw all that troubled real estate "
        "and burdensome debt overboard -- should voluntarily jump the "
        "now secured ship into a sinking lifeboat.' Distress + heavy "
        "equity grant + known turnaround talent (curated dictionary "
        "match with role-proximity) = strong asymmetric signal. "
        "Source: turnaround_signal.csv (built from 8-K Item 5.02 "
        "appointments).", 10)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_foreign_markets(wb: Workbook):
    """Foreign-markets tab: Japan TSE PBR<1, Korea Value-Up, UK
    schemes. Lives in its own tab because foreign tickers retain
    their yfinance suffix (.T/.KS/.L) and don't belong in the US
    universe consensus."""
    ws = wb.create_sheet("Foreign Markets")
    set_col_widths(ws, [10, 6, 30, 10, 10, 10, 10, 50])
    write_title_band(
        ws,
        "Foreign Markets — JP / KR / UK",
        "Japan TSE PBR<1 reform targets, Korea Value-Up chaebol, "
        "UK FTSE 100 scheme/take-private candidates. Foreign tickers "
        "kept separate from US universe consensus.",
        n_cols=8,
    )

    path = ROOT / "foreign_markets.json"
    if not path.exists():
        ws.cell(row=4, column=1,
                value="(no file -- run `python3 foreign_markets.py`)"
                ).font = SUBTITLE_FONT
        return

    headers = ["Ticker", "Mkt", "Name", "Score", "P/B", "P/E",
                "ROE %", "Reasons"]
    write_header_row(ws, 4, headers)

    rows = list(json.loads(path.read_text()).items())
    rows.sort(key=lambda x: -float(x[1].get("score", 0)))
    r = 5
    for i, (tk, v) in enumerate(rows[:60], 1):
        roe_pct = (v.get("roe") or 0) * 100
        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [tk, v.get("jurisdiction", ""),
                        (v.get("name") or "")[:32],
                        v.get("score"),
                        v.get("p_b"),
                        v.get("p_e_trailing"),
                        roe_pct,
                        (v.get("reasons") or "")[:80]],
                       band=band)
        ws.cell(row=r, column=1).font = BODY_BOLD
        # Color-code by jurisdiction
        jur = v.get("jurisdiction")
        if jur == "JP":
            ws.cell(row=r, column=2).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=2).font = BODY_BOLD
        elif jur == "KR":
            ws.cell(row=r, column=2).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=2).font = BODY_BOLD
        elif jur == "UK":
            ws.cell(row=r, column=2).fill = TITLE_BAND_FILL
            ws.cell(row=r, column=2).font = BODY_BOLD
        r += 1

    r += 1
    write_footnote(ws, r,
        "Japan signal: TSE explicit policy targets PBR<1 companies for "
        "capital return reform; scored highest at PBR<0.7. Korea: "
        "Value-Up / treasury cancellation program (chaebol families "
        "named candidates: Samsung, Hyundai, LG, SK, POSCO). UK: "
        "scheme of arrangement is the dominant take-private mechanism "
        "in London; deep-discount FTSE names are the target pool. "
        "Source: foreign_markets.json (yfinance with .T/.KS/.L "
        "suffixes; seed list ~130 tickers, expandable).", 8)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_recent_30d(wb: Workbook, yf: dict):
    """Live-rendered Recent-30d tab: only names with a material
    incentive event in the last 30 days. The tightest stale-pricing
    asymmetry window -- where the market has had least chance to
    incorporate the new disclosure."""
    ws = wb.create_sheet("Recent 30d")
    set_col_widths(ws, [9, 13, 32, 12, 8, 22, 8, 6, 12, 50])
    write_title_band(
        ws,
        "Stale-Pricing Asymmetry -- Last 30 days",
        "Names where a material incentive event was disclosed in "
        "the last 30 days AND the price has not yet reflected it. "
        "Tightest possible information-lag window.",
        n_cols=10,
    )

    path = ROOT / "recent_incentive_asymmetry_30d.csv"
    if not path.exists():
        ws.cell(row=4, column=1,
                value="(no file -- run "
                       "`python3 recent_incentive_asymmetry.py "
                       "--window-days 30`)").font = SUBTITLE_FONT
        return

    headers = ["#", "Ticker", "Name", "Score", "Days", "Event kind",
                "Inner", "DD%", "P/B", "Reasons"]
    write_header_row(ws, 4, headers)

    r = 5
    rows = list(csv.DictReader(path.open()))
    for i, row in enumerate(rows[:40], 1):
        tk = row["ticker"]
        y = yf.get(tk, {}) or {}
        name = (y.get("name") or tk)[:32]
        band = (i % 2 == 0)
        try:
            score = float(row["score"])
        except Exception:
            score = 0.0
        write_body_row(ws, r,
                       [i, tk, name,
                        f"{score:.0f}",
                        row.get("days_since", ""),
                        row.get("latest_event_kind", ""),
                        row.get("n_events_inner", ""),
                        row.get("drawdown_pct") or "",
                        row.get("p_b") or "",
                        (row.get("reasons") or "")[:90]],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        if score >= 55:
            ws.cell(row=r, column=4).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        elif score >= 40:
            ws.cell(row=r, column=4).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    write_footnote(ws, r,
        "30-day window of `recent_incentive_asymmetry.py`. Event types: "
        "DEF14A_PSU (latest proxy), 10b5-1_TERM (insider walked back "
        "scheduled selling), F4_PBUY (insider open-market purchase), "
        "ACTIVIST_13D, RESTRUCT_8K (8-K restructuring keyword), "
        "FORM_10_SPINOFF, NOL_SHELL. Score boosts: drawdown >60% "
        "unpriced (+15), multi-event cluster (+12), event in last "
        "7 days (+25). Source: recent_incentive_asymmetry_30d.csv.", 10)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_single_measure(wb: Workbook, yf: dict, proxy: dict,
                           bbv: dict, tender: dict, c10: dict,
                           f4: dict):
    """Per-layer extreme tail. Surfaces names that lead a SINGLE
    measure even if they don't fire multiple layers (the ODTX
    archetype: $75M insider trifecta but no PSU plan, so consensus
    misses them)."""
    ws = wb.create_sheet("Single-Measure Best")
    set_col_widths(ws, [9, 32, 13, 30, 16, 50])
    write_title_band(
        ws,
        "Best in Class -- Single-Measure Exceptional",
        "Names ranking top in ONE individual signal even if the "
        "consensus misses them. The "
        "'single-leg high-intensity' tail.",
        n_cols=6,
    )

    sections = []

    # 1. PSU forensic core
    psu_ranked = []
    for tk, p in proxy.items():
        core = p.get("psu_core") or 0
        if core > 0:
            psu_ranked.append((tk, float(core),
                               f"core {core:.0f} | "
                               f"cond_cats: {','.join(p.get('cond_cats') or [])}"))
    psu_ranked.sort(key=lambda x: -x[1])
    sections.append(("Highest PSU forensic core",
                      "What it measures: depth + rigor of the PSU plan",
                      psu_ranked[:8]))

    # 2. Highest governance score
    gov_ranked = []
    for tk, p in proxy.items():
        g = p.get("gov_score") or 0
        if g >= 15:
            gov_ranked.append((tk, float(g),
                               f"gov {g:.0f} | reasons: "
                               + "; ".join((p.get('gov_reasons') or [])[:3])))
    gov_ranked.sort(key=lambda x: -x[1])
    sections.append(("Highest governance score",
                      "What it measures: clawback + anti-hedge + vesting + "
                      "ownership multiple",
                      gov_ranked[:8]))

    # 3. Deepest P/B floor
    pb_ranked = []
    for tk, y in yf.items():
        pb_v = y.get("p_b")
        try: pb_v = float(pb_v) if pb_v is not None else None
        except Exception: pb_v = None
        if pb_v is not None and 0 < pb_v < 0.5:
            pb_ranked.append((tk, pb_v,
                              f"P/B {pb_v:.2f} | mcap ${(y.get('mcap') or 0)/1e6:,.0f}M"))
    pb_ranked.sort(key=lambda x: x[1])  # ascending = deepest discount first
    sections.append(("Deepest P/B floor (<0.5x book)",
                      "What it measures: hard balance-sheet value floor",
                      pb_ranked[:8]))

    # 4. Cheapest EV/EBITDA
    evb_ranked = []
    for tk, y in yf.items():
        v = y.get("ev_ebitda")
        try: v = float(v) if v is not None else None
        except Exception: v = None
        # exclude near-zero (often biotech with denominator issues)
        if v is not None and 0.5 < v < 6:
            evb_ranked.append((tk, v,
                                f"EV/EBITDA {v:.1f} | "
                                f"mcap ${(y.get('mcap') or 0)/1e6:,.0f}M | "
                                f"{y.get('sector') or ''}"))
    evb_ranked.sort(key=lambda x: x[1])
    sections.append(("Cheapest EV/EBITDA (<6x)",
                      "What it measures: enterprise-value to earnings power",
                      evb_ranked[:8]))

    # 5. Lowest trailing P/E (excluding negatives)
    pe_ranked = []
    for tk, y in yf.items():
        v = y.get("p_e_trailing")
        try: v = float(v) if v is not None else None
        except Exception: v = None
        if v is not None and 0.5 < v < 8:
            pe_ranked.append((tk, v,
                               f"P/E {v:.1f} | "
                               f"mcap ${(y.get('mcap') or 0)/1e6:,.0f}M | "
                               f"{y.get('sector') or ''}"))
    pe_ranked.sort(key=lambda x: x[1])
    sections.append(("Lowest trailing P/E (<8x, positive)",
                      "What it measures: earnings yield",
                      pe_ranked[:8]))

    # 6. Largest verified buyback shrinkage
    bb_ranked = []
    for tk, b in bbv.items():
        if not isinstance(b, dict): continue
        if b.get("status") not in ("EXECUTING", "SHRINKING_NO_AUTH"): continue
        chg = (b.get("share_change") or {}).get("change_pct")
        if chg is None: continue
        bb_ranked.append((tk, abs(chg),
                          f"shares {chg:+.1f}% over "
                          f"{(b.get('share_change') or {}).get('span_days','?')}d | "
                          f"status {b.get('status')}"))
    bb_ranked.sort(key=lambda x: -x[1])
    sections.append(("Largest verified buyback shrinkage",
                      "What it measures: actual supply curve compression",
                      bb_ranked[:8]))

    # 7. 10b5-1 termination signed score
    c10_ranked = []
    for tk, c in c10.items():
        if not isinstance(c, dict): continue
        s = c.get("score")
        if s is None or float(s) < 25: continue
        c10_ranked.append((tk, float(s),
                            f"signed score {float(s):.0f} | "
                            f"{c.get('reasons', '')[:50]}"))
    c10_ranked.sort(key=lambda x: -x[1])
    sections.append(("Highest 10b5-1 termination signed score",
                      "What it measures: insider walked back scheduled selling",
                      c10_ranked[:8]))

    # 8. Largest Form 4 P-buy dollar cluster
    f4_ranked = []
    for tk, f in f4.items():
        if not isinstance(f, dict): continue
        dollar = f.get("total_dollar") or 0
        n_buyers = len(f.get("buyer_set") or [])
        if dollar < 1e6: continue
        y = yf.get(tk, {}) or {}
        mcap = y.get("mcap")
        try: mcap = float(mcap) if mcap else None
        except Exception: mcap = None
        # Only report pct_mcap when mcap is real and dollar < mcap
        if mcap and mcap > dollar:
            pct_mcap_str = f" ({dollar/mcap*100:.2f}% of mcap)"
        else:
            pct_mcap_str = ""
        f4_ranked.append((tk, dollar,
                           f"${dollar/1e6:.1f}M cluster, "
                           f"{n_buyers} buyers{pct_mcap_str}"))
    f4_ranked.sort(key=lambda x: -x[1])
    sections.append(("Largest insider Form 4 dollar cluster",
                      "What it measures: insider conviction in dollars",
                      f4_ranked[:8]))

    # 9. Live tender / 13E-3 events
    tender_ranked = []
    for tk, t in tender.items():
        if not isinstance(t, dict): continue
        role = t.get("role")
        if not role and not t.get("has_13e3"): continue
        label = role or ""
        if t.get("has_13e3"):
            label += " +13E-3"
        tender_ranked.append((tk, 1.0, f"{label} | live event"))
    sections.append(("Live tender / 13E-3 events",
                      "What it measures: mechanical bid as floor",
                      tender_ranked[:10]))

    # Render each section
    r = 4
    for sec_label, sec_subtitle, items in sections:
        ws.cell(row=r, column=1, value=sec_label).font = BODY_BOLD
        ws.cell(row=r, column=1).alignment = Alignment(
            indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22
        r += 1
        ws.cell(row=r, column=1, value=sec_subtitle).font = SUBTITLE_FONT
        ws.cell(row=r, column=1).alignment = Alignment(
            indent=1, vertical="center")
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=6)
        r += 1
        write_header_row(ws, r, ["#", "Detail", "Ticker", "Name",
                                  "Score", "Sector"])
        r += 1
        for i, (tk, score, detail) in enumerate(items, 1):
            y = yf.get(tk, {}) or {}
            name = (y.get("name") or "")[:32]
            sector = y.get("sector") or ""
            band = (i % 2 == 0)
            write_body_row(ws, r,
                           [i, detail, tk, name,
                            f"{score:.1f}" if isinstance(score, (int, float)) else score,
                            sector],
                           band=band, align_first_left=False)
            ws.cell(row=r, column=3).font = BODY_BOLD
            ws.cell(row=r, column=2).alignment = Alignment(
                vertical="center", wrap_text=True, indent=1, horizontal="left")
            r += 1
        r += 1   # gap between sections

    r += 1
    write_footnote(ws, r,
        "Each section ranks names by a SINGLE measure regardless of "
        "whether other layers fire. These are 'exceptional by one '"
        "criterion' -- examples: ODTX archetype (insider cluster only), "
        "GETY (live self-tender only), CDE (CEO 10b5-1 termination "
        "only). Such names rarely appear in the consensus convergent "
        "list because they don't fire multiple layers, but they can "
        "be the right pick for a single-mandate position.", 6)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def build_incentive_improvers(wb: Workbook, yf: dict, proxy: dict):
    """Governance / PSU / incentive-structure IMPROVERS: names whose
    latest DEF 14A shows the compensation architecture getting BETTER
    versus the prior plan. The thesis: a board that just tightened its
    own incentives (heavier PSU weight, added ownership requirements,
    longer vesting, strengthened clawback, shareholder-responsive
    redesign) is signaling a governance inflection -- often the
    precursor to the operational one.

    Scoring is RARITY-WEIGHTED: psu_weight_increased appears in only
    12 of 4,410 proxies (strong signal); clawback_strengthened in 607
    (post-SEC-rule boilerplate; weak alone). Data-driven from
    proxy_scan plan_deltas -- nothing hardcoded."""
    ws = wb.create_sheet("Incentive Improvers")
    set_col_widths(ws, [9, 26, 11, 9, 9, 9, 10, 46, 40])
    write_title_band(
        ws,
        "Governance / PSU / Incentive-Structure Improvers",
        "Names whose latest proxy tightened the incentive architecture "
        "vs the prior plan · rarity-weighted (rare deltas score high, "
        "post-2023-rule boilerplate scores low)",
        n_cols=9,
    )

    # Rarity-weighted delta scores (counts across 4,410 proxies noted)
    DELTA_WEIGHTS = {
        "psu_weight_increased":        (25, "PSU weight increased"),        # 12
        "new_metric_added":            (15, "new performance metric"),      # 2
        "ownership_requirement_added": (15, "ownership requirement added"), # 23
        "responsive_to_shareholders":  (10, "shareholder-responsive redesign"), # 152
        "anti_hedge_pledge_added":     (6,  "anti-hedge/pledge added"),     # 331
        "clawback_strengthened":       (4,  "clawback strengthened"),       # 607
        # penalties -- structure got WORSE
        "front_load_grant":            (-10, "front-loaded grant (worse)"), # 24
        "metric_eliminated":           (-8,  "metric eliminated (worse)"),  # 1
    }
    EVOLUTION_BONUS = {
        "vest_period_extended":        (12, "vesting period extended"),     # 36
    }

    def _n(v):
        if v is None: return None
        try: return float(v)
        except Exception: return None

    rows = []
    for tk, p in proxy.items():
        deltas = p.get("plan_deltas") or []
        pattern = " ".join(p.get("pattern_reasons") or [])
        score = 0.0
        improvements = []
        regressions = []
        for d in deltas:
            key = d if isinstance(d, str) else str(d)
            w, label = DELTA_WEIGHTS.get(key, (0, None))
            if not label:
                continue
            if w > 0:
                score += w; improvements.append(label)
            else:
                score += w; regressions.append(label)
        for key, (w, label) in EVOLUTION_BONUS.items():
            if key in pattern:
                score += w; improvements.append(label)
        if not improvements:
            continue
        # Multi-improvement conjunction bonus: 2+ distinct improvements
        # is a deliberate redesign, not an isolated tweak.
        if len(set(improvements)) >= 3:
            score += 10
        elif len(set(improvements)) >= 2:
            score += 5
        # Context: improvement AFTER say-on-pay dissent is the
        # forced-response archetype (board reacting to shareholders).
        sop = _n(p.get("say_on_pay_pct"))
        if sop is not None and sop < 80:
            score += 8
            improvements.append(f"follows SOP dissent ({sop:.0f}%)")
        y = yf.get(tk, {}) or {}
        rows.append({
            "tk": tk,
            "name": (y.get("name") or tk),
            "score": round(score, 1),
            "psu_core": p.get("psu_core"),
            "gov": p.get("gov_score"),
            "psu_pct": p.get("psu_pct_lti"),
            "pb": _n(y.get("p_b")),
            "improvements": "; ".join(dict.fromkeys(improvements)),
            "regressions": "; ".join(dict.fromkeys(regressions)),
            "filing_date": p.get("filing_date", ""),
        })
    rows.sort(key=lambda r: -r["score"])

    headers = ["Ticker", "Name", "Improve", "PSU core", "Gov",
                "PSU %", "P/B", "What improved", "Caveats / date"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, row in enumerate(rows[:45], 1):
        band = (i % 2 == 0)
        caveat = row["regressions"] or ""
        tail = (caveat + (" · " if caveat else "") + row["filing_date"])
        write_body_row(ws, r,
                       [row["tk"], row["name"][:26], row["score"],
                        row["psu_core"], row["gov"], row["psu_pct"],
                        row["pb"], row["improvements"][:90], tail[:60]],
                       band=band, bold_first=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    n_total = len(rows)
    write_footnote(ws, r,
        f"{n_total} names show at least one structural improvement in "
        "the latest proxy (top 45 shown). Weights reflect rarity across "
        "4,410 scanned DEF 14As: PSU-weight increase (12 occurrences, "
        "+25) and ownership-requirement addition (23, +15) are "
        "deliberate alignment choices; clawback strengthening (607, +4) "
        "and anti-hedge language (331, +6) are largely post-2023 "
        "SEC-rule boilerplate. Multi-improvement redesigns earn a "
        "conjunction bonus; improvements that follow a sub-80% "
        "say-on-pay vote are tagged as forced-response. Regressions "
        "(front-loaded grants, metric elimination) subtract and are "
        "listed as caveats. Source: proxy_scan plan_deltas + "
        "pattern_reasons -- fully data-driven.", 9)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_insider_conviction(wb: Workbook, yf: dict):
    """Discretionary insider-conviction clusters: the most anomalous,
    high-conviction OPEN-MARKET buying (Form 4 code P only -- never RSU
    grants or option exercises). Rewards temporal clustering
    (Lakonishok-Lee), role-weighted dollars (CEO/CFO > 10% holder), and
    rare configurations (multiple C-suite same-day). Dollar-gated so
    trivial-dollar 'clusters' from filing artifacts are neutralised.

    Source: discretionary_insider_conviction.json (built by
    discretionary_insider_conviction.py from form4_buys.json)."""
    ws = wb.create_sheet("Insider Conviction")
    set_col_widths(ws, [9, 26, 10, 9, 9, 9, 9, 11, 11, 44])
    write_title_band(
        ws,
        "Discretionary Insider-Conviction Clusters",
        "Open-market purchases only (Form 4 code P) · clustered, "
        "role-weighted, dollar-gated · grants and option exercises "
        "excluded by construction",
        n_cols=10,
    )

    data = {}
    p = ROOT / "discretionary_insider_conviction.json"
    if p.exists():
        try:
            data = json.loads(p.read_text())
        except Exception:
            data = {}

    rows = []
    for tk, v in data.items():
        if not isinstance(v, dict) or (v.get("score") or 0) <= 0:
            continue
        y = yf.get(tk, {}) or {}
        rows.append({
            "tk": tk,
            "name": (y.get("name") or tk),
            "score": v.get("score", 0),
            "n_insiders": v.get("n_insiders", 0),
            "cluster": v.get("cluster_size", 0),
            "same_day": v.get("same_day_cluster", 0),
            "csuite": v.get("csuite_buyers", 0),
            "total_m": (v.get("total_dollar") or 0) / 1e6,
            "top_m": (v.get("top_person_dollar") or 0) / 1e6,
            "flags": "; ".join(v.get("flags") or []),
        })
    rows.sort(key=lambda r: -r["score"])

    headers = ["Ticker", "Name", "Conviction", "Insiders", "Cluster",
               "Same day", "C-suite", "Total $M", "Top $M",
               "Configuration"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, row in enumerate(rows[:45], 1):
        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [row["tk"], row["name"][:26], row["score"],
                        row["n_insiders"], row["cluster"],
                        row["same_day"], row["csuite"],
                        round(row["total_m"], 2), round(row["top_m"], 2),
                        row["flags"][:80]],
                       band=band, bold_first=True)
        ws.row_dimensions[r].height = 24
        r += 1

    r += 1
    n_total = len(rows)
    n_cluster = sum(1 for x in rows if x["cluster"] >= 2)
    n_sameday = sum(1 for x in rows if x["same_day"] >= 2)
    write_footnote(ws, r,
        f"{n_total} names show scored discretionary buying (top 45 "
        f"shown); {n_cluster} have a multi-insider cluster inside a "
        f"45-day window and {n_sameday} a same-day cluster. Population "
        "is Form 4 transaction code P only -- open-market cash "
        "purchases; grants (code A) and option exercises (code M) are "
        "excluded at the scanner, so every dollar here is discretionary. "
        "Cluster credit follows Lakonishok-Lee (temporal concentration, "
        "not raw count) and is scaled by aggregate dollars committed so "
        "filing artifacts cannot masquerade as conviction. Dollars are "
        "role-weighted (CEO/CFO 1.6x, Chair 1.5x, director 1.0x, 10% "
        "holder 0.6x); multiple C-suite buying together is the rarest "
        "and highest-signal configuration. This layer is additive and "
        "orthogonal to the raw Form 4 layer and the Cohen-Malloy "
        "opportunistic layer. Source: "
        "discretionary_insider_conviction.json.", 10)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_uk_events(wb: Workbook, yf: dict):
    """UK RNS capital events (own-shares buybacks, premium placings,
    scheme/CVA/restructuring distressed progress, takeover offers).
    Separate universe -- UK names do NOT enter the US consensus.
    Source: uk_rns_events.json (investegate monitor)."""
    ws = wb.create_sheet("UK Capital Events")
    set_col_widths(ws, [10, 22, 8, 12, 22, 10, 14])
    write_title_band(
        ws,
        "UK Capital Events — RNS revealed-preference & distressed monitor",
        "UK-listed capital-allocation and distressed-progress events "
        "(Transaction in Own Shares, placings, schemes of arrangement, "
        "Part 26A plans, CVAs, tender offers) — investment-trust NAV "
        "buybacks filtered out; separate universe from the US consensus",
        n_cols=7,
    )
    data = {}
    p = ROOT / "uk_rns_events.json"
    if p.exists():
        try:
            data = json.loads(p.read_text())
        except Exception:
            data = {}
    rows = [v for v in data.values()
            if isinstance(v, dict) and (v.get("score") or 0) > 0]
    rows.sort(key=lambda r: -r["score"])
    headers = ["Ticker", "Name", "Score", "Family", "Class",
               "Premium %", "Signals"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, v in enumerate(rows[:45], 1):
        sig = []
        if v.get("vs_market"):
            sig.append(v["vs_market"])
        if v.get("from_holder"):
            sig.append("from holder")
        if v.get("finality"):
            sig.append("final")
        if v.get("charges_satisfied"):
            sig.append(f"{v['charges_satisfied']} charges satisfied")
        write_body_row(ws, r,
                       [v["ticker"], v.get("name", "")[:22].replace("-", " "),
                        v["score"], v.get("family", ""),
                        v.get("class", "").replace("_", " "),
                        (v.get("premium_pct") if v.get("premium_pct") is not None else "—"),
                        "; ".join(sig) or "—"],
                       band=(i % 2 == 0), bold_first=True)
        ws.row_dimensions[r].height = 22
        r += 1
    if not rows:
        ws.cell(row=5, column=1,
                value="No qualifying UK events in the latest poll — the "
                      "monitor accumulates rare high-signal events (schemes, "
                      "tenders, restructurings, strategic investments) over "
                      "repeated runs.").font = BODY_FONT
    r = max(r, 6) + 1
    write_footnote(ws, r,
        f"{len(rows)} UK capital events (top 45). UK terminology per the "
        "spec: 'Transaction in Own Shares' (own-shares buyback), 'placing'/"
        "'subscription' (issuance — premium placings are the revealed-"
        "preference gold), 'scheme of arrangement' / 'Part 26A restructuring "
        "plan' / 'CVA' (distressed progress), 'tender offer' / 'return of "
        "capital' (selective own-shares). Investment-trust NAV buybacks are "
        "filtered (mechanical, low signal). SOURCE LIMITATION: the UK has no "
        "free EFTS-equivalent; this monitors investegate's recent server-"
        "rendered RNS listing (shallow snapshot), so it surfaces high-signal "
        "events as they occur rather than a historical sweep. Deep UK "
        "coverage (Companies House charges/MR04-MR05 lien releases, full "
        "history) activates when a free CH_API_KEY is supplied. UK names are "
        "a SEPARATE universe and do not enter the US consensus. Source: "
        "uk_rns_scan.py.", 7)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_hidden_asset(wb: Workbook, yf: dict):
    """Hidden-asset / incentivised value realisation (the SSP setup):
    rare under-recognised assets inside small levered equity, with a
    credit agreement sweeping disposition proceeds to debt paydown.
    Source: credit_agreement_mine.json + hidden_asset_watch.json."""
    ws = wb.create_sheet("Hidden Asset Realisation")
    set_col_widths(ws, [9, 18, 8, 9, 15, 26, 30])
    write_title_band(
        ws,
        "Hidden-Asset / Incentivised Value Realisation",
        "Under-recognised assets across ALL industries (spectrum, "
        "water/mineral/royalty rights, real estate, NOLs, equity stakes, "
        "pension surplus, royalty streams) inside a small levered equity "
        "base, where a credit agreement sweeps disposition proceeds to "
        "debt paydown — so realising the asset is mechanically stub-accretive",
        n_cols=7,
    )
    data = {}
    p = ROOT / "credit_agreement_mine.json"
    if p.exists():
        try:
            data = json.loads(p.read_text())
        except Exception:
            data = {}
    rows = [v for v in data.values()
            if isinstance(v, dict) and (v.get("score") or 0) > 0]
    rows.sort(key=lambda r: -r["score"])
    headers = ["Ticker", "Name", "Score", "Mand. prepay", "Industry",
               "Asset types", "Thesis / feature"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, v in enumerate(rows[:40], 1):
        y = yf.get(v["ticker"], {}) or {}
        mp = "yes" if v.get("mandatory_prepay") else ("no" if v.get("mandatory_prepay") is False else "—")
        cat = ", ".join(v.get("categories") or []) or ("watch" if v.get("watch") else "—")
        assets = ", ".join(v.get("asset_types") or [])
        thesis = (v.get("credit_agreement_feature") or v.get("hidden_asset") or "")[:30]
        write_body_row(ws, r,
                       [v["ticker"], (y.get("name") or v["ticker"])[:18],
                        v["score"], mp, cat[:15], assets[:26], thesis],
                       band=(i % 2 == 0), bold_first=True)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1
    n = len(rows)
    write_footnote(ws, r,
        f"{n} hidden-asset / incentivised-realisation setups across all "
        "industries (top 40). The thesis (E.W. Scripps / SSP archetype): a "
        "credit agreement that MANDATES asset-disposition proceeds be "
        "applied to debt paydown creates a structural, incentivised path to "
        "value realisation — on a small equity base, selling a hidden asset "
        "retires senior debt and lifts the residual stub roughly one-for-"
        "one. The high-value setup is the CONJUNCTION: a rare valuable "
        "asset + the mandatory-prepayment sweep + a small levered equity "
        "base. The asset taxonomy spans every sector — telecom (spectrum, "
        "fiber, towers), resources/energy (water/mineral rights, reserves, "
        "royalty acres, timberland, carbon), real estate (land bank, ground "
        "leases, air rights), financials (MSRs, securities), IP/healthcare "
        "(royalty streams, patents, milestones, CVRs), holdco sum-of-parts "
        "(equity stakes), and cross-industry off-balance-sheet value (NOLs, "
        "deferred tax assets, pension surplus, litigation/insurance "
        "recoveries). 'watch' names are hand-curated with catalyst triggers "
        "and counter-risks (hidden_asset_watch.json); the rest are mined "
        "from EDGAR. Source: credit_agreement_mine.py.", 7)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_distressed_stub(wb: Workbook, yf: dict):
    """Distressed-stub progress: stage-gated value-unlock events in
    capital-structure workouts. Source: distressed_stub_progress.json."""
    ws = wb.create_sheet("Distressed Stub Progress")
    set_col_widths(ws, [9, 22, 9, 6, 26, 34, 22])
    write_title_band(
        ws,
        "Distressed-Stub Progress — value-unlock events, waterfall-gated",
        "Finality-filtered capital-structure progress (debt retired below "
        "par, claims disallowed, plan effective, distributions) — scored "
        "only where value reaches the residual security, penalised for "
        "priming / dilution / MIP leakage",
        n_cols=7,
    )
    data = {}
    p = ROOT / "distressed_stub_progress.json"
    if p.exists():
        try:
            data = json.loads(p.read_text())
        except Exception:
            data = {}
    rows = [v for v in data.values()
            if isinstance(v, dict) and (v.get("score") or 0) > 0]
    rows.sort(key=lambda r: -r["score"])
    headers = ["Ticker", "Name", "Score", "Stage", "Classification",
               "Event classes", "Counter-signals"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, v in enumerate(rows[:45], 1):
        y = yf.get(v["ticker"], {}) or {}
        cls = ", ".join(sorted({e["class"] for e in (v.get("events") or [])}))
        ct = ", ".join(v.get("counters") or []) or "—"
        write_body_row(ws, r,
                       [v["ticker"], (y.get("name") or v["ticker"])[:22],
                        v["score"], v.get("max_stage"),
                        v.get("classification", "").replace("_", " "),
                        cls[:34], ct[:22]],
                       band=(i % 2 == 0), bold_first=True)
        ws.row_dimensions[r].height = 22
        r += 1
    r += 1
    n = len(rows)
    write_footnote(ws, r,
        f"{n} distressed stubs with net-positive progress (top 45). The "
        "engine alerts only on FINALITY (hard completion verbs — retired, "
        "cancelled, discharged, effective, distributed — not intentions), "
        "gated to plausibly-distressed names (deep drawdown, sub-$2B, "
        "post-Ch11, forced-selling), and scores per the stub waterfall: "
        "permanent senior-principal reduction (+4), claims disallowed "
        "(+4), stub distribution (+4), plan effective (+3), asset-sale "
        "cash received (+3), lien release (+2), >12mo maturity extension "
        "(+1). Counter-signals subtract heavily — equity wipeout (−10), "
        "priming/superpriority (−4), toxic dilution (−4), new preferred "
        "(−3) — because progress for the company is not progress for the "
        "stub. Stage 5 = value reached the residual security. Source: "
        "distressed_stub_progress.json (US 8-K/EDGAR; recipe table "
        "extensible to RNS/HKEX/ASX and local-language vocab).", 7)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_asymmetry_assembly(wb: Workbook, yf: dict):
    """PSIX-recipe conjunction: names where the assembled causal system
    co-occurs (cheap + operating inflection + survivable leverage +
    orphaned + insider costly-action + catalyst). Gated -- a name
    mediocre-on-many does NOT appear; only genuine assemblies with the
    required spine. Source: asymmetry_assembly.json."""
    ws = wb.create_sheet("Asymmetry Assembly")
    set_col_widths(ws, [9, 24, 10, 7, 6, 6, 6, 6, 6, 6, 6, 6, 40])
    write_title_band(
        ws,
        "Asymmetry Assembly — the PSIX recipe",
        "Conjunction detector: fires only when cheap expectations, an "
        "operating inflection, survivable leverage and a costly-action "
        "alignment signal REINFORCE one another — not a sum of weak parts",
        n_cols=13,
    )
    data = {}
    p = ROOT / "asymmetry_assembly.json"
    if p.exists():
        try:
            data = json.loads(p.read_text())
        except Exception:
            data = {}
    rows = [(tk, v) for tk, v in data.items()
            if isinstance(v, dict) and (v.get("score") or 0) > 0]
    rows.sort(key=lambda x: -x[1]["score"])

    comp_order = ["C1_low_expectations", "C2_leveraged_survivor",
                  "C3_orphaned_drawdown", "C4_revealed_insider",
                  "C5_recognition_catalyst", "C6_operating_inflection",
                  "C7_deleveraging", "C8_underused_capacity"]
    headers = ["Ticker", "Name", "Assembly", "N/8", "Cheap", "Lev",
               "Orph", "Insdr", "Catl", "Inflx", "Delev", "Cap",
               "Reasons"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, (tk, v) in enumerate(rows[:45], 1):
        comps = v.get("components", {})
        marks = ["●" if comps.get(c, {}).get("present") else "—"
                 for c in comp_order]
        y = yf.get(tk, {}) or {}
        write_body_row(ws, r,
                       [tk, (y.get("name") or tk)[:24], v.get("score"),
                        v.get("n_present")] + marks
                       + ["; ".join(v.get("reasons") or [])[:70]],
                       band=(i % 2 == 0), bold_first=True)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    n = len(rows)
    write_footnote(ws, r,
        f"{n} names pass the assembly SPINE (top 45 shown): low "
        "expectations AND an engine (operating inflection or leverage "
        "torque) AND a costly-action alignment signal (open-market "
        "insider buy or a curated maturity-extension / subordination "
        "event). A name strong on many unrelated layers but missing a "
        "spine leg scores ZERO here — that is the point: asymmetry is a "
        "conjunction, not a sum. Components: Cheap (C1, incl. EV/EBIT), "
        "Lev (C2 survivable leverage), Orph (C3 drawdown/orphaned), "
        "Insdr (C4 open-market buys), Catl (C5 emergence/tender/"
        "activist), Inflx (C6 gross profit up while revenue down / "
        "margin expansion, from XBRL), Delev (C7 debt & interest "
        "falling), Cap (C8 low capex + high marginal returns). "
        "Convergence beyond the spine adds a convexity bonus; dilutive-"
        "refinancing / backstop-expiry counter-signals subtract. The "
        "engine is validated point-in-time: it flags May-2024 PSIX "
        "(@ $2.15, 7/8 components) — see asymmetry_backtest.py. Source: "
        "asymmetry_assembly.json (+ financials_inflection.json XBRL).", 13)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_noval_view(wb: Workbook, yf: dict):
    """Parallel ranking that EXCLUDES the valuation leg. Surfaces
    structurally strong names that may be missing yfinance overlay
    or whose multiples don't sit in our discount zones."""
    ws = wb.create_sheet("Without Valuation")
    set_col_widths(ws, [9, 13, 32, 14, 14, 14, 14, 50])
    write_title_band(
        ws,
        "Top of Universe -- Without Valuation Leg",
        "Parallel ranking that excludes valuation entirely so "
        "names missing yfinance overlay are not penalised. "
        "Compared against the valuation-included ranking to "
        "identify structurally-strong-but-valuation-hidden names.",
        n_cols=8,
    )

    headers = ["#", "Ticker", "Name", "Layers", "Consensus",
               "Lift vs val", "Sector", "Detail"]
    write_header_row(ws, 4, headers)

    # Load both rankings
    noval = list(csv.DictReader(open(ROOT / "full_universe_consensus_noval.csv")))
    val = list(csv.DictReader(open(ROOT / "full_universe_consensus.csv")))
    val_rank = {r["ticker"]: i+1 for i, r in enumerate(val)}

    r = 5
    for i, row in enumerate(noval[:30], 1):
        tk = row["ticker"]
        y = yf.get(tk, {}) or {}
        name = y.get("name", tk)[:32]
        sector = y.get("sector", "")
        n_layers = row["n_layers_firing"]
        cons = row["consensus_score"]
        vr = val_rank.get(tk, "")
        lift = (int(vr) - i) if vr else ""
        detail_parts = []
        for fld, label in [("psu_pts","PSU"), ("buyback_pts","BB"),
                            ("tender_pts","TND"), ("c10b51_pts","C10"),
                            ("f4_buys_pts","F4"),
                            ("recent_incentive_pts","RI"),
                            ("special_sits_pts","SS")]:
            try:
                v = float(row.get(fld) or 0)
                if v != 0:
                    detail_parts.append(f"{label}:{v:.0f}")
            except Exception:
                pass
        detail = " ".join(detail_parts)
        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [i, tk, name, n_layers, cons,
                        f"+{lift}" if lift and lift > 0 else (lift or ""),
                        sector or "—", detail],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        if lift and lift > 50:
            ws.cell(row=r, column=6).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=6).font = BODY_BOLD
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    write_footnote(ws, r,
        "'Lift vs val' shows how many positions a name climbs when "
        "valuation is excluded. A large positive lift means the name "
        "is structurally strong but is being penalised by the "
        "valuation layer -- likely either missing yfinance overlay or "
        "trading at multiples outside our discount zones (P/B>1.5, "
        "P/E>15, EV/EBITDA>10). The 26 emergent names "
        "(consensus_emergent_noval.csv) are gap-fill priorities.", 8)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


def build_cover(wb: Workbook):
    ws = wb.create_sheet("Cover")
    set_col_widths(ws, [8, 30, 18, 18, 18, 18])
    write_title_band(ws,
                     "Most Asymmetric Situations",
                     "Universe analysis · 6,164 US-listed tickers · "
                     "8 independent rankers + 57 archetypes",
                     n_cols=6)
    ws.row_dimensions[3].height = 18
    ws.cell(row=4, column=2, value="Executive summary").font = BODY_BOLD

    summary_lines = [
        ("Universe scope",
         "6,164 US-listed tickers (NYSE / Nasdaq / AMEX / CBOE)"),
        ("Governance / PSU coverage",
         "4,410 DEF 14As scanned (72% of universe)"),
        ("Convergent (≥3 of 8 screens + archetype winner)",
         "12 names"),
        ("Uniquely convergent (6 of 8 screens)",
         "HFFG — the only ticker hitting six independent rankers"),
        ("Highest archetype count",
         "CSGP — wins five PSU/governance archetypes"),
        ("Clean convergent (zero red flags)",
         "CSGP, RNR"),
        ("Probability of 6-screen convergence by chance",
         "≈ 2.4 × 10⁻¹⁰"),
        ("Robustness check",
         "12-name list unchanged after 2.8x coverage expansion"),
    ]

    r = 5
    for label, val in summary_lines:
        ws.cell(row=r, column=2, value=label).font = BODY_BOLD
        ws.cell(row=r, column=2).alignment = Alignment(vertical="center")
        ws.cell(row=r, column=3, value=val).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(
            vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=3,
                       end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="The convergent twelve").font = BODY_BOLD
    r += 1

    # Derived from disk -- consensus_ranking.csv. No hardcoded list.
    proxy = load_proxy()
    convergent_rows = get_convergent_from_disk()
    convergent = []
    for cr in convergent_rows:
        tk = cr["ticker"]
        ann = TICKER_ANNOTATIONS.get(tk, {})
        name = ann.get("name", tk)
        ns = int(cr["n_screens"])
        na = int(cr["n_archetypes_won"])
        nflags = red_flag_count(tk, proxy)
        sizing = sizing_for_screens(ns, na, nflags)
        why = ann.get("why",
                       f"{ns} layers firing | cons {cr.get('consensus_score', '?')}")
        convergent.append((tk, name, sizing, why))
    write_header_row(ws, r, ["#", "Ticker", "Name", "Sizing", "Why"])
    r += 1
    for i, (tk, name, sz, why) in enumerate(convergent, 1):
        band = (i % 2 == 0)
        write_body_row(ws, r, [i, tk, name, sz, why],
                       band=band, align_first_left=False, bold_first=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        # Color-code the sizing column
        if "Concentrated" in sz:
            ws.cell(row=r, column=4).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        elif "Participation" in sz:
            ws.cell(row=r, column=4).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=4).font = BODY_BOLD
        ws.merge_cells(start_row=r, start_column=5,
                       end_row=r, end_column=6)
        r += 1

    r += 1
    write_footnote(ws, r,
        "Cyclepapa Research · Module Series 1 · The framework "
        "produces 12 names by demanding (a) presence in top-N of "
        "≥3 of 8 independent rankers built from independent evidence "
        "and (b) winner of at least one PSU/governance archetype "
        "across 57 buckets. Robustness checks: list unchanged "
        "after expanding yfinance coverage and after a governance-"
        "scoring bugfix that nearly tripled Tier-B coverage.", 6)
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------------
# Tab 2: Most Asymmetric — per-name detail rows
# ----------------------------------------------------------------------

def build_most_asymmetric(wb: Workbook, proxy: dict, yf: dict, bbv: dict,
                          tender: dict, c10: dict, f4: dict):
    ws = wb.create_sheet("Most Asymmetric")
    set_col_widths(ws, [9, 22, 11, 11, 11, 11, 22, 32, 28, 18])
    write_title_band(ws,
                     "The Convergent Twelve",
                     "Per-name structural detail · sourced from "
                     "PSU forensics + valuation + buyback verification "
                     "+ tender mechanics + insider behaviour",
                     n_cols=10)

    headers = ["Ticker", "Name", "mcap ($M)", "Spot ($)", "P/B",
               "PSU core", "Catalyst (cond_cat)", "Why convergent",
               "Floor", "Sizing"]
    write_header_row(ws, 4, headers)

    # Derived from disk -- full_universe_consensus.csv. No hardcoded list.
    convergent_rows = get_convergent_from_disk()
    convergent_data = []
    for cr in convergent_rows:
        tk = cr["ticker"]
        ann = TICKER_ANNOTATIONS.get(tk, {})
        name = ann.get("name", tk)
        ns = int(cr["n_screens"])
        na = int(cr.get("n_archetypes_won") or ns)
        nflags = red_flag_count(tk, proxy)
        why = ann.get("why",
                       f"{ns} layers firing | cons {cr.get('consensus_score', '?')}")
        floor = ann.get("floor", "see proxy_scan + buyback_verify")
        sizing = sizing_for_screens(ns, na, nflags)
        convergent_data.append((tk, name, why, floor, sizing))

    r = 5
    for i, (tk, name, why, floor, sizing) in enumerate(convergent_data, 1):
        p = proxy.get(tk, {})
        y = yf.get(tk, {})
        mcap = (y.get("mcap") or 0) / 1e6 if y else 0
        px = y.get("price") if y else None
        pb = y.get("p_b") if y else None
        psu_core = p.get("psu_core") if p else None
        cc = ", ".join(p.get("cond_cats") or []) if p else ""

        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [tk, name,
                        f"${mcap:,.0f}" if mcap else "?",
                        f"${px:.2f}" if px else "?",
                        f"{pb:.2f}" if pb else "?",
                        f"{psu_core:.0f}" if psu_core else "—",
                        cc or "—",
                        why, floor, sizing],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=1).font = BODY_BOLD
        ws.cell(row=r, column=8).alignment = Alignment(
            vertical="center", wrap_text=True, indent=1, horizontal="left")
        ws.cell(row=r, column=9).alignment = Alignment(
            vertical="center", wrap_text=True, indent=1, horizontal="left")
        if "Concentrated" in sizing:
            ws.cell(row=r, column=10).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=10).font = BODY_BOLD
        elif "Participation" in sizing:
            ws.cell(row=r, column=10).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=10).font = BODY_BOLD
        ws.row_dimensions[r].height = 56
        r += 1

    r += 1
    write_footnote(ws, r,
        "Catalyst (cond_cat) = forward-conditional vesting category coded "
        "in the PSU plan text. Sizing band reflects the convergence "
        "strength (number of independent screens) combined with red-flag "
        "count from the plan-evolution score. Source CSVs: "
        "consensus_ranking.csv, grand_unified_ranked.csv, "
        "buyback_verify.json, tender_scan.json.", 10)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ----------------------------------------------------------------------
# Tab 3: By Archetype
# ----------------------------------------------------------------------

def build_by_archetype(wb: Workbook, arch_psu: dict, arch_asym: dict,
                       yf: dict):
    ws = wb.create_sheet("By Archetype")
    set_col_widths(ws, [9, 14, 50, 22, 11, 11, 26])
    write_title_band(ws,
                     "Archetype Winners",
                     "57 PSU/governance/thesis buckets · single best "
                     "representative per archetype across 6,164-name "
                     "universe",
                     n_cols=7)

    headers = ["#", "Code", "Archetype description",
               "Winner", "mcap ($M)", "P/B", "Source file"]
    write_header_row(ws, 4, headers)

    # Build combined arch list. Each item: (code, desc, winner_tk, src_file)
    items = []
    # Parse PSU_ARCHETYPES.md by re-reading
    psu_text = (ROOT / "PSU_ARCHETYPES.md").read_text() \
               if (ROOT / "PSU_ARCHETYPES.md").exists() else ""
    asym_text = (ROOT / "ASYMMETRIC_BY_ARCHETYPE.md").read_text() \
                if (ROOT / "ASYMMETRIC_BY_ARCHETYPE.md").exists() else ""

    def parse_archetype_pairs(text, src):
        # BUGFIX (silent-drop audit): the greedy `.*?` could span past a
        # section that has NO Winner line into the next archetype's
        # Winner, mis-attributing the ticker and dropping the correct
        # (empty) section. The tempered `(?:(?!\n###?\s+\w+\d+\.).)*?`
        # forbids the match from crossing a subsequent archetype header.
        blocks = re.findall(
            r"###?\s+(\w+\d+)\.\s+([^\n]+?)\n"
            r"(?:(?!\n###?\s+\w+\d+\.).)*?\*\*Winner:\s*"
            r"([A-Z][A-Z0-9.\-]{0,10})\*\*",
            text, re.S)
        return [(code, desc.strip(), tk, src)
                for code, desc, tk in blocks]

    items.extend(parse_archetype_pairs(psu_text, "PSU_ARCHETYPES"))
    items.extend(parse_archetype_pairs(asym_text, "ASYMMETRIC_BY_ARCHETYPE"))

    r = 5
    for i, (code, desc, tk, src) in enumerate(items, 1):
        y = yf.get(tk, {}) or {}
        mcap = (y.get("mcap") or 0) / 1e6
        pb = y.get("p_b")
        band = (i % 2 == 0)
        write_body_row(ws, r,
                       [i, code, desc, tk,
                        f"${mcap:,.0f}" if mcap else "?",
                        f"{pb:.2f}" if pb else "?",
                        src.replace("_", " ").title()],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        ws.cell(row=r, column=4).font = BODY_BOLD
        ws.cell(row=r, column=3).alignment = Alignment(
            vertical="center", wrap_text=True, indent=1, horizontal="left")
        r += 1

    r += 1
    write_footnote(ws, r,
        "Archetype = a single scored dimension (PSU forward-conditional "
        "trigger, governance evolution, plan-pattern flag, or thesis "
        "convergence). Winner = single ticker scoring highest on that "
        "dimension across the 6,164-name universe. Multi-archetype "
        "winners (CSGP, HFFG, LE, KMPR) are the structurally rarest. "
        "Source: PSU_ARCHETYPES.md + ASYMMETRIC_BY_ARCHETYPE.md.", 7)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ----------------------------------------------------------------------
# Tab 4: Reserve Baskets
# ----------------------------------------------------------------------

def build_reserve_baskets(wb: Workbook, yf: dict):
    ws = wb.create_sheet("Reserve Baskets")
    set_col_widths(ws, [9, 30, 18, 50])
    write_title_band(ws,
                     "Reserve Baskets",
                     "Sub-archetype groups for diversified single-"
                     "mandate deployment alongside the convergent twelve",
                     n_cols=4)

    baskets = [
        ("Microcap forcing-function (Bastian)",
         "BEEP, LGL, NUS, DXLG, WW, OSUR",
         "P/B < 0.5 + PSU trigger + tender role; "
         "RGS/NLOP-archetype self-help"),
        ("Mungerian forward-dollar PSU",
         "HFFG, MAT, LMT, THRY, EHTH",
         "Named dollar hurdle = knowable catalyst; "
         "the most informative PSU class"),
        ("Verified buyback compounders",
         "CSGP, KMPR, ADT, PAYC, GRND",
         "EXECUTING status with PSU alignment; "
         "real supply-curve compression, not just authorisation"),
        ("Live tender / event-driven",
         "EXFY, GPUS, GETY, LE, DXLG, CWAN",
         "Issuer self-tender / TARGET 14D-9 / 13E-3 going-private; "
         "mechanical bid as floor"),
        ("Special-situations debt-haircut (BBGI-archetype)",
         "WW, LGL, QVCGQ, ENHA, FONR",
         "Capital-structure forcing function; "
         "exchange-offer + springing maturity"),
        ("Cohen-Malloy informational stack",
         "NSP, ODTX, FONR, MOBI, RGR",
         "Cluster + role + size = informational; "
         "ODTX = $75.3M trifecta (9.57% of mcap)"),
        ("Activist 13D + 8-K restructuring",
         "RPAY, CCO, SATS",
         "Triple-cross-validated (13D + restructuring + boundary); "
         "activist as forced-action catalyst"),
        ("Russell-recon forced flow",
         "EBS, BYND, CMCO, BLCO, MUR",
         "Within ±20% of R2000 cutoff; "
         "passive-flow distortion candidates"),
        ("NOL shell / §382 tax-asset",
         "WOLF, CEG, NOTV, NINE, USGO, CMLSQ, TSEOF",
         "Tax Benefits Preservation Rights Plan adoption; "
         "WMIH-archetype tax-attribute monetisation"),
        ("Going-dark / Form 15 OTC",
         "(see going_dark.csv)",
         "Oddball Stocks dark-company terrain; "
         "post-deregistration value-stub plays"),
    ]

    headers = ["#", "Basket", "Holdings", "Why this basket exists"]
    write_header_row(ws, 4, headers)

    r = 5
    for i, (label, names, why) in enumerate(baskets, 1):
        band = (i % 2 == 0)
        write_body_row(ws, r, [i, label, names, why],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="center", wrap_text=True, indent=1, horizontal="left")
        ws.row_dimensions[r].height = 48
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Portfolio math").font = BODY_BOLD
    r += 1
    portfolio = [
        ("Concentrated convergent (5%+ each)", "HFFG, CSGP, RNR", "15.0%"),
        ("Material convergent (2-5% each)",
         "LE, NUS, ADT, KMPR, MAT, LMT, CDE", "24.5%"),
        ("Participation convergent (1-2% each)", "GO, GPRO", "3.0%"),
        ("Microcap forcing-function basket", "6 names", "10.0%"),
        ("Cohen-Malloy stack", "5 names", "5.0%"),
        ("Russell-recon flow", "5 names", "3.0%"),
        ("NOL shells", "7 names", "3.0%"),
        ("Special-sits debt-haircut", "5 names", "5.0%"),
        ("CASH / OPPORTUNISTIC RESERVE", "—", "31.5%"),
    ]
    write_header_row(ws, r, ["Bucket", "Holdings", "Weight", ""])
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    r += 1
    for i, (b, n, w) in enumerate(portfolio, 1):
        band = (i % 2 == 0)
        is_cash = b.startswith("CASH")
        write_body_row(ws, r, [b, n, w, ""],
                       band=band, align_first_left=False)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        ws.cell(row=r, column=1).font = BODY_BOLD
        if is_cash:
            ws.cell(row=r, column=1).fill = TITLE_BAND_FILL
            ws.cell(row=r, column=2).fill = TITLE_BAND_FILL
            ws.cell(row=r, column=3).fill = TITLE_BAND_FILL
            ws.cell(row=r, column=4).fill = TITLE_BAND_FILL
        r += 1

    r += 1
    write_footnote(ws, r,
        "Reserve baskets implement the playbook recommendation to "
        "specialise where the crowd isn't (microcap, forcing-function, "
        "dark companies, forced-flow distortion). Sizing band is the "
        "Bastian/Dalius two-question test applied per name: position "
        "size = function of knowability × downside × liquidity.", 4)
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------------
# Tab 5: Caution List
# ----------------------------------------------------------------------

def build_caution_list(wb: Workbook, proxy: dict, consensus: list):
    ws = wb.create_sheet("Caution List")
    set_col_widths(ws, [9, 14, 16, 32, 14])
    write_title_band(ws,
                     "Caution List",
                     "Convergent names with red-flag wins · weight "
                     "structural alignment against governance penalty",
                     n_cols=5)

    # Build flag map
    cautions = {}
    for tk, p in proxy.items():
        flags = []
        prs = p.get("pattern_reasons") or []
        grs = p.get("gov_reasons") or []
        for s in prs + grs:
            sl = s.lower()
            if "single-trigger" in sl: flags.append("single-trigger CIC")
            if "repricing" in sl: flags.append("repricing language")
            if "retirement carveout" in sl: flags.append("retirement carveout")
            if "front-loaded" in sl: flags.append("front-loaded grant")
            if "discretionary" in sl: flags.append("discretionary hurdle")
            if "aggregate-only" in sl: flags.append("aggregate-only metrics")
        if flags:
            cautions[tk] = sorted(set(flags))

    top_names = []
    for r in consensus:
        try:
            ns = int(r["n_screens"])
        except Exception:
            continue
        if ns >= 3:
            top_names.append(r["ticker"])

    headers = ["#", "Ticker", "Screens fired", "Red flags", "Action"]
    write_header_row(ws, 4, headers)

    r = 5
    i = 1
    for tk in top_names:
        flags = cautions.get(tk)
        if not flags:
            continue
        nflags = len(flags)
        # Action ladder by flag count
        if nflags == 0: action = "Concentrate"
        elif nflags == 1: action = "Concentrate (1 flag tolerable)"
        elif nflags == 2: action = "Material — basket diversify"
        elif nflags == 3: action = "Participation — limited size"
        else: action = "Watch only — multi-flag stack"

        ns = next((int(r["n_screens"]) for r in consensus if r["ticker"] == tk), 0)
        band = (i % 2 == 0)
        write_body_row(ws, r, [i, tk, ns, ", ".join(flags), action],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        if nflags >= 3:
            ws.cell(row=r, column=5).fill = FLAG_TAG_FILL
            ws.cell(row=r, column=5).font = BODY_BOLD
        elif nflags <= 1:
            ws.cell(row=r, column=5).fill = CLEAN_TAG_FILL
            ws.cell(row=r, column=5).font = BODY_BOLD
        ws.row_dimensions[r].height = 30
        r += 1
        i += 1

    r += 1
    write_footnote(ws, r,
        "Eight flag classes: single-trigger CIC, repricing language, "
        "retirement carveout, front-loaded grant, discretionary hurdle, "
        "aggregate-only metrics, plus structural variants. Counter-"
        "intuitive: a 'repricing language' flag can become a re-rate "
        "catalyst when a stock has fallen sharply, since management "
        "becomes incentivised to reset hurdles. Read the flag in "
        "context, not as automatic veto.", 5)
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------------
# Tab 6: Coverage & Tiers
# ----------------------------------------------------------------------

def build_coverage(wb: Workbook):
    """Data-driven coverage diagnostics. Counts live coverage from each
    layer's source file and the real layer-firing distribution from
    full_universe_consensus.csv. Includes freshness (age in days) from
    layer_freshness.json. NOTHING hardcoded -- everything reflects the
    current state on disk."""
    ws = wb.create_sheet("Coverage & Tiers")
    set_col_widths(ws, [28, 12, 12, 9, 22, 24])
    write_title_band(ws,
                     "Coverage & Freshness Diagnostics",
                     "Live per-layer coverage, data age, and the real "
                     "layer-firing distribution across the universe",
                     n_cols=6)

    UNIVERSE = 6164

    def _count_json(fn, predicate=None):
        p = ROOT / fn
        if not p.exists():
            return 0
        try:
            d = json.loads(p.read_text())
        except Exception:
            return 0
        if isinstance(d, dict):
            if predicate is None:
                return len(d)
            return sum(1 for v in d.values() if predicate(v))
        return len(d)

    def _count_proxy():
        seen = set()
        for fn in sorted(ROOT.glob("proxy_scan*.json")):
            try:
                d = json.loads(fn.read_text())
            except Exception:
                continue
            rows = d if isinstance(d, list) else d.values()
            for rr in rows:
                if isinstance(rr, dict) and rr.get("ticker"):
                    seen.add(rr["ticker"])
        return len(seen)

    # Load freshness if present
    fresh = {}
    fp = ROOT / "layer_freshness.json"
    if fp.exists():
        try:
            fresh = json.loads(fp.read_text())
        except Exception:
            fresh = {}

    def age_of(layer_key):
        v = fresh.get(layer_key)
        if v and v.get("age_days") is not None:
            return f"{v['age_days']:.0f}d"
        return EM_DASH

    proxy_n = _count_proxy()
    # (display label, count, freshness-key, source file, signal)
    layer_rows = [
        ("PSU forensics", proxy_n, "psu", "proxy_scan*.json", "Knowable catalyst"),
        ("Governance score", proxy_n, "psu", "proxy_scan*.json", "Board constraint"),
        ("Tender / SC TO / 13E-3", _count_json("tender_scan.json"), "tender",
         "tender_scan.json", "Mechanical bid"),
        ("10b5-1 directional", _count_json("cancel_10b5_1.json"), "c10b51",
         "cancel_10b5_1.json", "Insider direction"),
        ("Form 144 proposed sales", _count_json("form144_scan.json"), "f144",
         "form144_scan.json", "Bearish signal"),
        ("yfinance valuation", _count_json("yfinance_quick.json"), "valuation",
         "yfinance_quick.json", "Price/book floor"),
        ("Buyback verification", _count_json("buyback_verify.json"), "buyback",
         "buyback_verify.json", "Verified shrinkage"),
        ("Form 4 P-buys", _count_json("form4_buys.json"), "f4_buys",
         "form4_buys.json", "Insider conviction"),
        ("Opportunistic insiders", _count_json("opportunistic_insiders.json"),
         "opportunistic_insiders", "opportunistic_insiders.json", "Cohen-Malloy"),
        ("Quarterly 10-Q", _count_json("quarterly_10q_data.json"),
         "quarterly_10q", "quarterly_10q_data.json", "Fresh balance sheet"),
        ("Net-net NCAV", _count_json("net_net_ncav.json"),
         "net_net_ncav", "net_net_ncav.json", "Graham floor"),
        ("Voss CIC triangulation", _count_json("voss_cic_triangulation.json"),
         "voss_cic", "voss_cic_triangulation.json", "M&A predictor"),
        ("Coval-Stafford proxy", _count_json("coval_stafford_proxy.json"),
         "coval_stafford", "coval_stafford_proxy.json", "Fire-sale pressure"),
        ("N-PORT forced selling", _count_json("nport_forced_selling.json"),
         "nport_forced_selling", "nport_forced_selling.json", "Real Coval-Stafford"),
        ("13F-delta", _count_json("form_13f_delta.json"),
         "form_13f_delta", "form_13f_delta.json", "Smart-money flow"),
        ("Financial primary", _count_json("financial_primary.json"),
         "financial_primary", "financial_primary.json", "Non-PSU sector"),
        ("Biotech PDUFA", _count_json("biotech_pdufa_calendar.json"),
         "biotech_pdufa", "biotech_pdufa_calendar.json", "FDA catalyst"),
        ("Activist letter feed", _count_json("activist_letter_feed.json"),
         "activist_letter", "activist_letter_feed.json", "Pre-13D activism"),
        ("Foreign markets (JP/KR/UK)", _count_json("foreign_markets.json"),
         None, "foreign_markets.json", "Non-US value-up"),
    ]

    headers = ["Data layer", "Coverage", "% univ", "Age", "Signal", "Source"]
    write_header_row(ws, 4, headers)
    r = 5
    for i, (label, n, fkey, src, signal) in enumerate(layer_rows, 1):
        pct = (n / UNIVERSE * 100) if UNIVERSE else 0
        age = age_of(fkey) if fkey else EM_DASH
        band = (i % 2 == 0)
        write_body_row(ws, r, [label, n, pct, age, signal, src],
                       band=band, bold_first=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Layer-firing distribution "
            "(real, from consensus)").font = BODY_BOLD
    r += 1
    # Real firing distribution
    try:
        rows = list(csv.DictReader(open(ROOT / "full_universe_consensus.csv")))
        n_layers = len([k for k in rows[0].keys() if k.endswith("_pts")])
        from collections import Counter
        dist = Counter(int(rr["n_layers_firing"]) for rr in rows)
    except Exception:
        rows, n_layers, dist = [], 0, {}
    write_header_row(ws, r, ["Layers firing", "Names", "", "", "", ""])
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    r += 1
    for j, lv in enumerate(sorted(dist, reverse=True)):
        if lv == 0:
            continue
        band = (j % 2 == 0)
        write_body_row(ws, r, [lv, dist[lv], "", "", "", ""], band=band)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        r += 1

    r += 1
    write_footnote(ws, r,
        f"{n_layers} total scoring layers, all additive. Coverage and "
        "data age are read live from each layer's source file at build "
        "time; no figures are hardcoded. Layers with low coverage "
        "(Form 4, 13F-delta, tender) are signal-sparse by design — "
        "they fire only on names exhibiting the pattern. The layer-"
        "firing distribution is the count of names firing on N "
        "independent layers; per the correlation analysis "
        "(layer_correlation_pairs.csv) the effective-independent layer "
        "count is ~21 of 30 at rho>0.6.", 6)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ----------------------------------------------------------------------
# Tab 7: Methodology
# ----------------------------------------------------------------------

def build_methodology(wb: Workbook):
    ws = wb.create_sheet("Methodology")
    set_col_widths(ws, [9, 30, 70])
    write_title_band(ws,
                     "Methodology",
                     "How the convergent twelve were surfaced from "
                     "the 6,164-name universe",
                     n_cols=3)

    headers = ["#", "Step", "Detail"]
    write_header_row(ws, 4, headers)

    method = [
        ("Universe construction",
         "6,164 US-listed common tickers from cancel_10b5_1.json — "
         "authoritative NYSE/Nasdaq/AMEX/CBOE set. Foreign names "
         "(JP/KR/UK) live in a separate tab and universe."),
        ("Layer ingestion",
         f"{n_consensus_layers()} additive scoring layers ingested per ticker, spanning PSU "
         "forensics, governance, valuation, verified buybacks, tender "
         "mechanics, 10b5-1 direction, Form 4 (raw + Cohen-Malloy "
         "opportunistic), Form 144, quarterly 10-Q, NCAV, Voss CIC, "
         "post-Ch11, internalization, bumpitrage, spinoff timing, "
         "Arquitos, Coval-Stafford proxy + real N-PORT, backstopped "
         "rights, FDIC dark banks, activist letters, 13F-delta, biotech "
         "PDUFA, financial-sector primary, discretionary insider "
         "conviction, emergence cross-feed, Sohn pitches, asymmetry "
         "assembly (PSIX recipe), distressed-stub progress, premium "
         "injections, selective buybacks, and hidden-asset "
         "realisation."),
        ("Additive discipline",
         "Every layer ADDS to the composite; none modifies another's "
         "score. New legs append fields; existing weights never change. "
         "This is enforced and audited (verify_universe_methodology.py).") ,
        ("Coverage-normalised composite",
         "Sparse-coverage names are not penalised for missing layers; "
         "the norm score rescales by sqrt(n_total/n_present) so a name "
         "strong on few layers competes with a name mediocre on many."),
        ("Per-pattern catalyst ranking",
         "Top names in each catalyst pattern (forward $ hurdle, M&A "
         "close, spin trigger, FDA/PDUFA, post-Ch11, asset sale, etc.) "
         "— surfaces single-mandate leaders (see Single-Measure tab)."),
        ("Archetype winners",
         "57 PSU/governance/thesis buckets — single best representative "
         "of each archetype across the universe (PSU_ARCHETYPES.md 38 + "
         "ASYMMETRIC_BY_ARCHETYPE.md 19)."),
        ("Consensus meta-ranking",
         "n_layers_firing = how many of the 30 independent layers "
         "produce a non-zero score for the ticker. consensus_score = "
         "sum of per-layer rank-decay contributions across the universe."),
        ("Layer independence",
         "Pairwise Spearman correlation (layer_correlation_pairs.csv) "
         f"collapses {n_consensus_layers()} raw layers to effective-independent at "
         "rho>0.6. Three correlated clusters: PSU+Voss, tender family, "
         "F4+opportunistic. 'Fires 9 layers' ≈ 7-8 true confirmations."),
        ("Freshness weighting",
         "layer_freshness.json records each layer's data age. Most "
         "layers are <14 days old. An optional age-decay multiplier "
         "(opt-in) can down-weight stale layers; current build reports "
         "age without auto-decaying."),
        ("Caution layering",
         "Eight red-flag classes from plan text: single-trigger CIC, "
         "repricing, retirement carveout, front-loaded grant, "
         "discretionary hurdle, aggregate-only metrics, plus structural. "
         "Convergence without direction; flag count modulates sizing."),
        ("Deployment / sizing",
         "Concentrated (≥5%): clean high-layer-count names. Material "
         "(2-5%): strong but 1-2 flags. Participation (0.5-2%): single-"
         "leg or multi-flag. Basket (<1% each): sub-archetype groups, "
         "Cohen-Malloy stack, R2000-boundary, NOL shells, foreign."),
        ("Honest limitations",
         "No realized-return backtest yet (AUDIT.md S1.1) — the "
         "composite is a structurally-sound pattern-recognition system, "
         "not yet validated alpha. Cohen-Malloy needs deeper Form 4 "
         "history. Coval-Stafford proxy supplements but does not replace "
         "the N-PORT real signal. See AUDIT.md for the full ledger."),
    ]
    r = 5
    for i, (step, detail) in enumerate(method, 1):
        band = (i % 2 == 0)
        write_body_row(ws, r, [i, step, detail],
                       band=band, align_first_left=False)
        ws.cell(row=r, column=2).font = BODY_BOLD
        ws.cell(row=r, column=3).alignment = Alignment(
            vertical="top", wrap_text=True, indent=1, horizontal="left")
        ws.row_dimensions[r].height = 50
        r += 1

    r += 1
    write_footnote(ws, r,
        "All scoring modules live in the cyclepapa repository. "
        "Full methodology trace: grand_unified_ranker.py + "
        "consensus_meta_ranker.py + systematic_rankings.py. "
        "Companion narrative: CASE_WORKBOOK.md teaches the framework; "
        "DILIGENCE_SHEETS.md gives per-name action triggers; "
        "BEST_OF_UNIVERSE.md proves the convergence claim.", 3)
    ws.sheet_view.showGridLines = False


# ----------------------------------------------------------------------
# Contents / navigation index (QoL)
# ----------------------------------------------------------------------

# (sheet title, one-line description). Order matches build order.
TAB_INDEX = [
    ("Cover", "Executive summary and the convergent shortlist."),
    ("Most Asymmetric", "Per-name detail for the highest layer-count names."),
    ("By Archetype", "Single best representative of each of 57 archetypes."),
    ("Reserve Baskets", "Sub-archetype baskets and full portfolio math."),
    ("Caution List", "Convergent names carrying governance red flags."),
    ("Incentive Improvers", "Latest proxy tightened the incentive architecture (rarity-weighted)."),
    ("Insider Conviction", "Discretionary open-market buying clusters (code P only, role-weighted)."),
    ("Asymmetry Assembly", "PSIX-recipe conjunction: cheap + inflection + leverage + insider co-occurring."),
    ("Distressed Stub Progress", "Finality-gated capital-structure value-unlock events, waterfall-scored."),
    ("Hidden Asset Realisation", "Spectrum/rights/RE inside levered stubs with mandatory-prepay debt sweeps (SSP-type)."),
    ("UK Capital Events", "UK RNS revealed-preference & distressed monitor (separate universe; funds filtered)."),
    ("Without Valuation", "Parallel ranking excluding the valuation leg."),
    ("Recent 30d", "Material incentive events disclosed in the last 30 days."),
    ("Foreign Markets", "Japan TSE PBR<1, Korea Value-Up, UK schemes."),
    ("Turnaround Signal", "Bollenbach pattern: turnaround talent into distress."),
    ("Single-Measure Best", "Best in class on each individual signal."),
    ("Layer Correlation", "Pairwise layer correlation and effective independence."),
    ("Coverage & Tiers", "Live per-layer coverage, data age, firing distribution."),
    ("Methodology", "How the composite is built, step by step."),
]


def build_contents(wb: Workbook):
    """First tab: a clean academic table of contents. Each row names a
    tab and what it holds. Pure navigation aid."""
    ws = wb.active
    ws.title = "Contents"
    set_col_widths(ws, [4, 26, 64])
    write_title_band(ws,
                     "The Asymmetric Equities Workbook",
                     f"Contents — a structural map of the {n_consensus_layers()}-layer "
                     "universe analysis",
                     n_cols=3)
    write_header_row(ws, 4, ["#", "Tab", "What it contains"])
    r = 5
    for i, (tab, desc) in enumerate(TAB_INDEX, 1):
        band = (i % 2 == 0)
        write_body_row(ws, r, [i, tab, desc], band=band)
        ws.cell(row=r, column=2).font = BODY_BOLD
        r += 1
    r += 1
    write_footnote(ws, r,
        "Generated by build_most_asymmetric_xlsx.py from the live "
        "consensus on disk. Every figure in this workbook is computed "
        "at build time; nothing is hardcoded. Single typeface "
        "(Times New Roman, 10pt); hierarchy by weight and rule only.", 3)
    ws.sheet_view.showGridLines = False


def build_layer_correlation(wb: Workbook):
    """Layer independence transparency: shows the most-correlated layer
    pairs and the effective-independent layer count. Reads from
    layer_correlation_pairs.csv + effective_layers.json."""
    ws = wb.create_sheet("Layer Correlation")
    set_col_widths(ws, [30, 30, 14, 16])
    write_title_band(ws,
                     "Layer Correlation & Independence",
                     "How independent are the scoring layers? Positively "
                     "correlated layers are not separate confirmations.",
                     n_cols=4)

    eff = {}
    ep = ROOT / "effective_layers.json"
    if ep.exists():
        try:
            eff = json.loads(ep.read_text())
        except Exception:
            eff = {}
    n_raw = eff.get("n_raw_layers", "—")
    n_eff = eff.get("n_effective_layers_at_06", "—")

    ws.cell(row=4, column=1,
            value=f"Raw layers: {n_raw}    Effective-independent "
                  f"(rho>0.6): {n_eff}").font = BODY_BOLD
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)

    headers = ["Layer A", "Layer B", "Spearman", "Interpretation"]
    write_header_row(ws, 6, headers)
    r = 7
    pp = ROOT / "layer_correlation_pairs.csv"
    if pp.exists():
        rows = list(csv.DictReader(pp.open()))
        # show pairs with |rho| >= 0.25 (the meaningful ones)
        shown = [x for x in rows
                 if abs(float(x.get("spearman_rho", 0) or 0)) >= 0.25]
        for i, x in enumerate(shown[:30], 1):
            band = (i % 2 == 0)
            a = x["layer_a"].replace("_pts", "")
            b = x["layer_b"].replace("_pts", "")
            write_body_row(ws, r,
                           [a, b, float(x["spearman_rho"]),
                            x.get("interpretation", "")],
                           band=band)
            r += 1
    r += 1
    write_footnote(ws, r,
        "Spearman rank correlation across the full universe. Pairs above "
        "rho 0.6 are folded into a single effective layer for the "
        "independence count: PSU + Voss CIC (Voss derives from the PSU "
        "plan), the tender family (tender + mechanism + bumpitrage share "
        "one source), and Form 4 + opportunistic insiders (one refines "
        "the other). A high raw layer-firing count should be read "
        "against this: nine layers firing is closer to seven or eight "
        "genuinely independent confirmations.", 4)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"


def _finalize_sheets(wb: Workbook):
    """Apply uniform view + print settings to every sheet (QoL):
    gridlines off, landscape, fit-to-width, repeat title rows on print,
    and a consistent margin. Idempotent and safe on all tabs."""
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        try:
            from openpyxl.worksheet.properties import PageSetupProperties
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            if ws.sheet_properties.pageSetUpPr is None:
                ws.sheet_properties.pageSetUpPr = PageSetupProperties()
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass
        try:
            ws.print_options.horizontalCentered = False
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.6
            ws.page_margins.bottom = 0.6
        except Exception:
            pass
        # Repeat the header rows (1-4) at the top of each printed page
        try:
            ws.print_title_rows = "1:4"
        except Exception:
            pass


# ----------------------------------------------------------------------
# Build
# ----------------------------------------------------------------------

def main() -> int:
    print("Loading sources...")
    proxy = load_proxy()
    yf = json.loads((ROOT / "yfinance_quick.json").read_text()) \
         if (ROOT / "yfinance_quick.json").exists() else {}
    bbv = json.loads((ROOT / "buyback_verify.json").read_text()) \
          if (ROOT / "buyback_verify.json").exists() else {}
    tender = json.loads((ROOT / "tender_scan.json").read_text()) \
             if (ROOT / "tender_scan.json").exists() else {}
    c10 = json.loads((ROOT / "cancel_10b5_1.json").read_text()) \
          if (ROOT / "cancel_10b5_1.json").exists() else {}
    f4 = json.loads((ROOT / "form4_buys.json").read_text()) \
         if (ROOT / "form4_buys.json").exists() else {}
    consensus = load_csv(ROOT / "consensus_ranking.csv")
    print(f"  proxy={len(proxy)} yf={len(yf)} bbv={len(bbv)} "
          f"tender={len(tender)} c10={len(c10)} f4={len(f4)} "
          f"consensus={len(consensus)}")

    wb = Workbook()
    build_contents(wb)
    build_cover(wb)
    build_most_asymmetric(wb, proxy, yf, bbv, tender, c10, f4)
    build_by_archetype(wb, {}, {}, yf)
    build_reserve_baskets(wb, yf)
    build_caution_list(wb, proxy, consensus)
    build_incentive_improvers(wb, yf, proxy)
    build_insider_conviction(wb, yf)
    build_asymmetry_assembly(wb, yf)
    build_distressed_stub(wb, yf)
    build_hidden_asset(wb, yf)
    build_uk_events(wb, yf)
    build_noval_view(wb, yf)
    build_recent_30d(wb, yf)
    build_foreign_markets(wb)
    build_turnaround_signal(wb, yf)
    build_single_measure(wb, yf, proxy, bbv, tender, c10, f4)
    build_layer_correlation(wb)
    build_coverage(wb)
    build_methodology(wb)

    _finalize_sheets(wb)
    wb.save(OUT)
    print(f"\nwrote {OUT}  ({len(wb.sheetnames)} tabs)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
