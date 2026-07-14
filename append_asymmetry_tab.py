"""Append an asymmetry-tier tab to screener_report.xlsx based on the
cross-region qualitative deep-research scoring.

Two scoring lenses:
  - asymmetry_stars: pure quality of setup (catalyst + insider + balance sheet)
  - entry_today_stars: re-rated for May 30 2026 entry — downgrades names
                       where the move has already happened (insider buys
                       are public, stock has lifted, narrative widely owned)
                       and upgrades names that are still un-rerated.

Tier ratings come from the 7-agent qualitative pass:
  ★★★★★ = exceptional (clear catalyst + downside-protected + insider buying)
  ★★★★  = strong (capital return + catalyst + clean governance)
  ★★★   = solid quality but limited asymmetry
  ★★    = mixed / structural overhang
  ★     = avoid / red flags
  0     = non-investable (delisted / shell / wrong ticker)
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# (ticker, region, asymmetry_stars, entry_today_stars, headline_thesis_short)
# entry_today_stars downgrades names where the move has been picked up by the market
ASYMMETRY = [
    # Original 5-star setups
    ('MGM',        'US', 5, 3, 'IAC accumulation widely known; stock has run; thesis no longer hidden'),
    ('ACRV',       'US', 5, 3, 'Cluster insider-buy Jan 14 2026 at $1.68 already absorbed; biotech runs post-disclosure'),
    ('G1A.DE',     'EU', 5, 3, 'Mission 30 narrative widely owned; premium quality-compounder multiple'),
    ('VPK.AS',     'EU', 5, 4, '€1.7bn distribution programme only 16% executed by April 2026 — multi-year runway'),
    ('GOOS.TO',    'CA', 5, 2, 'Trading near take-private bid; arb spread is only remaining return'),
    ('TOU.TO',     'CA', 5, 3, 'Mike Rose serial open-market buys public; LNG Canada thesis priced in'),
    # From-today ★★★★★ promotions
    ('TREE',       'US', 5, 5, 'LendingTree: Q1 sales +37%, EBITDA +71% organic; CEO Lebda died Oct 2025, Peyree (QuoteWizard founder) elevated; mgmt RAISED FY26 guide; stock at 52w low on non-operational overhang; <0.5x P/S, ~3x fwd EBITDA — single best un-priced operating leverage'),
    ('SHL.DE',     'EU', 4, 5, '-40% derate absorbed; Siemens spin-off doubles free float (forced buying); €230m buyback through Jan 2027'),
    ('INBK',       'US', 4, 5, 'Founder bought $18.60 Oct 2025; 13D filer surfaced; sharp discount to $41.41 book; credit overhang priced in'),
    ('BJRI',       'US', 4, 5, 'Shaich/Act III added at $34.92 March 2026; fresh CEO; +2.4% comps best in casual dining; $83M buyback runway'),
    ('SAX.DE',     'EU', 4, 5, 'Müller family insider-buy cluster Feb 6 2026; live PE sale process at €3-4bn; OOH +5.4% organic'),
    ('GRWG',       'US', 4, 5, '$46M cash = 38% of cap; CEO Lampert bought May 18 2026 at $1.54-1.55 (recent); cannabis at cyclical low'),
    ('FRE.DE',     'EU', 4, 5, 'Deleveraging validating; CFO bought late 2025; S&P upgraded outlook; Tyenne biosimilar share gains'),
    ('CFBK',       'US', 4, 5, 'Director Hoeweler bought $23.87 Dec 2025; 5% buyback through Aug 2026; tiny float'),
    # Tier 2 strong-but-partially-discovered
    ('UVSP',       'US', 4, 4, 'Buyback +50% expansion; div +4.5%; director open-market buys'),
    ('FRAF',       'US', 4, 4, 'EPS +69% YoY; director $106K open-market buy; 15% ROE; potential M&A target'),
    ('AMAL',       'US', 4, 4, 'Dividend +21%; active buyback; 21% ROE; isolated NPL pressure'),
    ('RLMD',       'US', 4, 4, 'Pre-funded into Phase 3; CFO bought; NDV-01 76% durable CR'),
    ('IRMD',       'US', 4, 3, '3870 product cycle; 31% op margin; CEO selling under 10b5-1'),
    ('RDVT',       'US', 4, 3, '85% gross / 41% EBITDA margin; active buyback; CEO sold near highs'),
    ('LXRX',       'US', 4, 4, '$199M cash; Novo partnership; HCM Phase 3 enroll mid-2026 still ahead'),
    ('NBN',        'US', 4, 4, '$525M+ loan purchases since 9/25; NIM 5.15%'),
    ('RBKB',       'US', 4, 4, 'Second-step thrift conversion closes Q3 2026 at $10'),
    ('LEU',        'US', 4, 3, '$1.87B cash; DOE HALEU re-award post-June 2026; priced for execution'),
    ('OFG',        'US', 4, 4, '$200M buyback + 17% div hike; PR discount; 16.4% ROATCE'),
    ('ENVX',       'US', 4, 4, 'Smartphone OEM qualification on new framework; $582M cash'),
    ('TWO-PA',     'US', 4, 4, 'Merger-arb pref; mandatory redemption $25 on UWMC deal close Q3'),
    ('HFBL',       'US', 4, 4, 'Quality LA thrift; EPS doubled YoY; 92% through buyback'),
    ('DAR',        'US', 4, 4, 'DGD turnaround; RVO tailwind; Point72 accumulating'),
    ('LAUR',       'US', 4, 3, 'LATAM ed +15%; raised guide; $181M buyback runway; some banked'),
    ('EIG',        'US', 4, 4, 'CFO bought Nov 2025 at $37-40; new $125M buyback'),
    ('ABOS',       'US', 4, 4, 'Cash through Phase 2 sabirnetug Alzheimer readout late 2026'),
    ('IFX.DE',     'EU', 4, 3, 'Guidance raised twice; FY26 segment margin ~20%; AI thesis priced'),
    ('MUV2.DE',    'EU', 4, 4, '€2.25bn buyback through Apr 2027; €24 div; Q1 net +57%'),
    ('NEM.DE',     'EU', 4, 3, '95% recurring rev; +17% cc; HCSS deal priced at premium'),
    ('SZG.DE',     'EU', 4, 4, 'Guidance raised twice; HKM closing; Aurubis windfall — partially banked'),
    ('HNR1.DE',    'EU', 4, 4, '€12.50 div (+39%); post-Henchoz reset; Talanx-controlled quality'),
    ('CWC.DE',     'EU', 4, 4, 'Foundation patient capital; 17-yr div streak; Cimpress monetisation H2 2026'),
    ('PAT.DE',     'EU', 4, 4, 'Trades near book; founder buying history; EBITDA +41% Q1; un-rerated RE manager'),
    ('ZURN.SW',    'EU', 4, 4, 'Beazley closing H2 2026; SST 265%; underowned post-raise'),
    ('SMHN.DE',    'EU', 4, 4, 'Q1 trough; record HBM/hybrid bonding order intake; H2 snapback'),
    ('ARIS.TO',    'CA', 4, 3, 'Marmato Q4 first-gold; Soto Norte stream eliminated; already ran on gold'),
    ('FTT.TO',     'CA', 4, 3, 'Record $3.8B backlog; 25-yr div streak; cyclical mining capex peak'),
    ('IAG.TO',     'CA', 4, 4, '$310M NCIB at trough; Côté ramp post-conveyor fix; net cash transition'),
    ('NA.TO',      'CA', 4, 4, 'CWB synergies ahead of guide ($176M Q1 > $116M FY25); rerating partial'),
    ('CM.TO',      'CA', 4, 4, 'Caribbean divestiture + new NCIB 30M sh; new CEO Culham reset pending'),
    ('QBR-B.TO',   'CA', 4, 3, 'Wireless +85% vs Big-3; dividend +14% — disruption thesis broadly recognized'),
    ('PXT.TO',     'CA', 4, 4, 'GeoPark activist proxy fight mid-2026; 40% share reduction; dividend cushion'),
    ('REI-UN.TO',  'CA', 4, 4, 'Chairman Sonshine net buyer LTM; 98.6% occ; 17.5% leasing spreads'),
    ('WSP.TO',     'CA', 4, 3, 'TRC/Power Engineers integration; backlog +18% — premium multiple, mostly banked'),
    # NEW: un-priced operating leverage confirmations
    ('QNST',       'US', 4, 4, 'Sales +28% Q3; adj EBITDA +53% on AI productivity + auto-insurance mix; $40M buyback active; net cash; auto-budget cycle is the risk (BUT HomeBuddy M&A drove half of Home growth + Q2 tax benefit inflated GAAP NI)'),
    ('NRDS',       'US', 4, 5, 'Sales +16%, margin +6.6pp, EBITDA +333%, P/S 0.6, stock -21% — cleanest un-priced operating leverage in our list'),
    ('EVER',       'US', 4, 4, 'Sales +24.5%, margin +7pp, EBITDA +403% — auto insurance carrier ad-budget cycle confirms operating leverage'),
    # NEW: downgraded — screener flagged but qualitative review caught accounting artifacts
    ('PRCH',       'US', 3, 2, 'Screener showed +30pp margin expansion BUT it is an accounting recharacterization (HOA carrier → PIRE Reciprocal MGA), not operating leverage; +15.6% real revenue growth not +42%; Texas 57% concentration; hurricane season tail risk'),
    ('AORT',       'US', 3, 2, '+18% sales partly cybersecurity-comp recovery (Q1 25 was depressed by Nov 2024 cyber incident); mgmt CUT guidance Q1; Endospan deal added $135M debt with no US revenue until 2027; zero insider buying through 57% drawdown'),
    # Avoid / value-trap flags (entry-today rating same as base)
    ('BYW.DE',     'EU', 1, 1, 'StaRUG restructuring; €2.7bn shortfall; criminal probe; audit delay'),
    ('GXI.DE',     'EU', 2, 2, 'BaFin probe; audit delay; technical default; SDAX expulsion'),
    ('OHB.DE',     'EU', 2, 1, 'Stock up ~8x in 12m; KKR placement = imminent overhang — asymmetry inverted'),
    ('VLN.TO',     'CA', 1, 1, 'Deal pre-priced at C$13.10 — no minority premium'),
    ('FRMM',       'US', 0, 0, 'Not Fremont — ETHZilla rebrand pivot shell'),
    ('ELTX',       'US', 0, 0, 'Ticker mismatch — ELEV was acquired by Concentra July 2025'),
    ('LUNA',       'US', 0, 0, 'Nasdaq delisted Jan 2025; OTC Expert Market only'),
    ('ASPU',       'US', 0, 0, 'Voluntarily delisted 2023; $0.4M cash; sub-scale wind-down'),
    ('HOTH',       'US', 1, 1, 'Pivoted to "nanomagnetic space-AI chips" — narrative shell'),
    ('ENVB',       'US', 1, 1, 'Reverse split + 5B share authority proposal + going concern'),
]

def main():
    df = pd.DataFrame(ASYMMETRY, columns=['ticker','region','asymmetry_stars','entry_today_stars','thesis'])
    # primary sort by entry-today stars (the more actionable lens), then base asymmetry
    df = df.sort_values(['entry_today_stars','asymmetry_stars','region','ticker'],
                        ascending=[False, False, True, True])

    xlsx = Path('screener_report.xlsx')
    with pd.ExcelWriter(xlsx, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
        df.to_excel(xw, sheet_name='asymmetry_tier', index=False)
        ws = xw.sheets['asymmetry_tier']
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 100
        ws.freeze_panes = 'A2'

    wb = load_workbook(xlsx)
    if 'asymmetry_tier' in wb.sheetnames:
        idx = wb.sheetnames.index('asymmetry_tier')
        wb.move_sheet('asymmetry_tier', offset=-idx)
        wb.save(xlsx)

    print(f'Appended asymmetry_tier tab with {len(df)} rows to {xlsx}')

if __name__ == '__main__':
    main()

