"""Add clean_topline tab to screener_report.xlsx."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

XLSX = Path('screener_report.xlsx')
SRC = Path('results_clean_topline/screener.csv')

if not SRC.exists():
    print(f'Missing {SRC} — run clean_topline_screener.py first')
    raise SystemExit(0)

df = pd.read_csv(SRC)

# Load universe metadata for name/sector/country
def load_meta():
    meta = {}
    for f in ['universe_us_wide.csv','universe_expanded.csv','universe_wider.csv',
              'universe_eu.csv','universe_eu_extra.csv','universe_canada.csv',
              'universe_us_large_mega.csv']:
        p = Path(f)
        if not p.exists(): continue
        try:
            d = pd.read_csv(p, usecols=lambda c: c in ['symbol','name','sector','industry','country'], low_memory=False)
            for _, r in d.iterrows():
                tk = r.get('symbol')
                if not isinstance(tk, str): continue
                k = tk.upper()
                if k in meta: continue
                meta[k] = {f: r.get(f) for f in ['name','sector','industry','country'] if f in d.columns}
        except Exception: pass
    return meta

META = load_meta()
for col in ('name','sector','industry','country'):
    df[col] = df['ticker'].astype(str).str.upper().map(lambda t: META.get(t, {}).get(col))

# Sanity: drop rows where gross margin > 100% (data error)
df = df[df['gross_margin_now_pct'].fillna(0) <= 100]

# Sensible column order
cols = ['ticker','name','sector','industry','country',
        'rev_ltm_now_M','rev_ltm_yoy_pct',
        'gross_ltm_now_M','gross_ltm_yoy_pct',
        'gross_margin_now_pct','gross_margin_chg_pp',
        'quality_score','perf_1y_pct',
        'market_cap','priceToSales','priceToBook','trailingPE',
        'enterpriseToEbitda','enterpriseToRevenue']
cols = [c for c in cols if c in df.columns]
df = df[cols].reset_index(drop=True)

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    df.to_excel(xw, sheet_name='clean_topline', index=False)
    ws = xw.sheets['clean_topline']
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 28
    for c in 'EFGHIJKLMNOPQRS': ws.column_dimensions[c].width = 14
    ws.freeze_panes = 'B2'

wb = load_workbook(XLSX)
if 'clean_topline' in wb.sheetnames:
    # Place right after forensic_audit (position 1)
    cur = wb.sheetnames.index('clean_topline')
    wb.move_sheet('clean_topline', offset=1 - cur)
    wb.save(XLSX)

print(f'Wrote clean_topline tab with {len(df)} rows')
