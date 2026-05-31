"""Add an archetype_top5 tab summarizing the top 5 names per setup archetype
with 1-year price performance and headline metrics. Helps surface
un-rerated picks within each setup style.
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

CACHE = Path('.cache/yf')
XLSX = Path('screener_report.xlsx')

def safe(t): return ''.join(c if c.isalnum() or c in '-_' else '_' for c in t)

def price_perf(ticker, days=252):
    p = CACHE / f'{safe(ticker)}__price.parquet'
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        if df.empty or 'Close' not in df.columns: return None
        s = pd.to_numeric(df['Close'], errors='coerce').dropna()
        if len(s) < days + 5: return None
        return float((s.iloc[-1] / s.iloc[-days] - 1) * 100)
    except Exception: return None

def load_info(ticker):
    p = CACHE / f'{safe(ticker)}__info_metrics.parquet'
    if not p.exists(): return {}
    try:
        d = pd.read_parquet(p)
        return d.iloc[0].to_dict() if not d.empty else {}
    except Exception: return {}

DEEP_RESEARCHED = {
    'UVSP','AXTI','GNE','FRAF','GRWG','AMAL','ENVB','RLMD','WFRD','LRN','MGM','NOV','HOTH','CRWS','ACRV','IRMD','RDVT',
    'TVRD','VVOS','LXRX','NBN','INDI','JBSS','CFBK','RBKB','TWO-PB','VELO','FIBK','COCO','LEU','OFG','OPEN','TITN','ENVX',
    'SNAP','FRMM','ELTX','TWO-PA','DJCO','HFBL','DMRC','LUNA','UHT','AMPX','DAR','LAUR','ASPU','EIG','HCWB','ABOS',
    'BJRI','ASTH','AMSC','ADUS','LIND','EVER','INTT','IIIN','RCMT','INBK',
    'IFX.DE','MUV2.DE','GXI.DE','BYW.DE','TTK.DE','D6H.DE','SAX.DE','NEM.DE','FRE.DE','HEN.DE','SZG.DE','G1A.DE','SRT3.DE',
    'GLJ.DE','HNR1.DE','EVD.DE','EVK.DE','DTE.DE','PAT.DE','VPK.AS','SHL.DE','ZURN.SW','SMHN.DE','OHB.DE','AIXA.DE','CWC.DE',
    'ARIS.TO','FTT.TO','IAG.TO','FNV.TO','GOOS.TO','X.TO','IGM.TO','DSG.TO','IFC.TO','CM.TO','NA.TO','BPF-UN.TO',
    'BNS.TO','QBR-B.TO','PXT.TO','SLF.TO','GWO.TO','REI-UN.TO','AEM.TO','TIH.TO','WSP.TO','VLN.TO','TOU.TO','ENB.TO',
}

ARCHETYPES = [
    ('multi_variant_US',      'results_us_wide/ranked.csv',          'avg_inflection_z', False, None),
    ('multi_variant_EU',      'results_eu_relaxed/ranked.csv',        'avg_inflection_z', False, None),
    ('multi_variant_CA',      'results_canada/ranked.csv',            'avg_inflection_z', False, None),
    ('fcf_signflip_strict',   'results_us_wide/fcf_inflections.csv',  'flip_magnitude_pct', False,
        lambda d: (d.get('view') == 'quarterly_strict') & (d.get('metric') == 'fcf_ps') & (d.get('is_flip') == True)),
    ('fcf_signflip_ttm',      'results_us_wide/fcf_inflections.csv',  'flip_magnitude_pct', False,
        lambda d: (d.get('view') == 'ttm_yoy') & (d.get('metric') == 'fcf_ps') & (d.get('is_flip') == True)),
    ('deep_value_inflection', 'results_us_wide/deep_value_screen.csv', 'n_variants_inflected', False,
        lambda d: d.get('is_value_plus_inflection') == True),
    ('cheap_inflecting',      'results_us_wide/valuation_screen.csv', 'n_variants_inflected', False,
        lambda d: d.get('is_cheap_inflecting') == True),
    ('52wh_cheap_on_growth',  'results_52wh/screener.csv',            'revenue_growth_ltm', False, None),
    ('multiple_compression',  'results_multiple_compression/clean.csv', 'multiple_compression_pct', True,
        lambda d: d.get('multiple_compression_pct') < -10),
    ('ev_ebitda_compression', 'results_ev_compression/screener.csv',  'compression_vs_sales_pct', True,
        lambda d: (d.get('ev_ebitda_now') > 0) & (d.get('ev_ebitda_now') < 30) & (d.get('sales_growth_pct') > 5)),
    ('operating_leverage',    'results_operating_leverage/screener.csv', 'leverage_score', False,
        lambda d: d.get('market_cap') > 100e6),
    ('ev_fcf_leverage',       'results_ev_fcf_leverage/screener.csv', 'fcf_growth_pct', False, None),
    ('fcf_yield_setup',       'results_fcf_yield/screener.csv',       'fcf_yield_now_pct', False,
        lambda d: (d.get('fcf_yield_now_pct') < 60) & (d.get('fcf_yield_now_pct') > 0)),
    ('segment_pre_rerate',    'pre_rerate_setups.csv',                'pre_rerate_score', False,
        lambda d: d.get('pre_rerate_score') > 5),
    ('volasym_bullish',       'results_volasym/volatility_asymmetry.csv', 'm_asym', False,
        lambda d: d.get('m_state').isin(['squeeze','hyper_squeeze']) & (d.get('m_asym_state') == 'upper')),
]

rows = []
for label, path, sort_col, asc, filt in ARCHETYPES:
    p = Path(path)
    if not p.exists(): continue
    try: df = pd.read_csv(p)
    except: continue
    if filt is not None:
        try: df = df[filt(df)].copy()
        except: pass
    if sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=asc)
    df = df.head(5).copy()
    tk_col = 'ticker' if 'ticker' in df.columns else df.columns[0]
    for rank, (_, row) in enumerate(df.iterrows(), 1):
        tk = str(row[tk_col])
        sig = row.get(sort_col, '')
        try: sig_val = float(sig) if pd.notna(sig) else None
        except: sig_val = None
        info = load_info(tk)
        rows.append({
            'archetype': label,
            'rank': rank,
            'ticker': tk,
            'signal_value': sig_val,
            'perf_1y_pct': price_perf(tk),
            'market_cap': info.get('marketCap'),
            'trailingPE': info.get('trailingPE'),
            'priceToBook': info.get('priceToBook'),
            'priceToSales': info.get('priceToSalesTrailing12Months'),
            'enterpriseToEbitda': info.get('enterpriseToEbitda'),
            'currentPrice': info.get('currentPrice'),
            'deep_researched': 'Y' if tk in DEEP_RESEARCHED else '',
        })

out = pd.DataFrame(rows)

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    out.to_excel(xw, sheet_name='archetype_top5', index=False)
    ws = xw.sheets['archetype_top5']
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    for col in 'FGHIJK':
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['L'].width = 12
    ws.freeze_panes = 'A2'

# Move tab to position 2 (right after asymmetry_tier)
wb = load_workbook(XLSX)
if 'archetype_top5' in wb.sheetnames:
    cur = wb.sheetnames.index('archetype_top5')
    desired = 1  # right after asymmetry_tier
    wb.move_sheet('archetype_top5', offset=desired - cur)
    wb.save(XLSX)

print(f'Wrote archetype_top5 tab with {len(out)} rows ({len(ARCHETYPES)} archetypes × 5 top picks)')
