"""
Build the FIP-Asymmetry workbook in a Harvard-monochrome aesthetic:
Times New Roman, single 11pt body size throughout, condensed letter
spacing, grayscale palette, hierarchy via WEIGHT + CASE + FILL only.
Financial-statement number convention: accounting separators, parens
for negatives, en-dash for empty/zero, percentage points (no % sign).

Tabs: Cover, Glossary, Composite by Region, Upside / Floor / Quality /
Stealth Leaders, Cheap Multiples, Growth & Margin, FIP by Timeframe,
RS-FIP Leaders, Volatility Asymmetry, Survival & Catalyst, Period
Returns, Liquidity Quality, Drift Audit, All Survivors.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Monochrome palette ---
BLACK       = "000000"
DARK_GRAY   = "1F2937"   # near-black; title / header fill
MID_GRAY    = "4B5563"   # section fill
SOFT_GRAY   = "9CA3AF"   # subtitle text on dark
LIGHT_GRAY  = "E5E7EB"   # borders
PAPER_GRAY  = "F3F4F6"   # banding
WHITE       = "FFFFFF"

# Times New Roman, single 11pt size, condensed throughout.
SERIF = "Times New Roman"
SIZE  = 11

REGION_MAP = {
    'United States': 'North America', 'Canada': 'North America',
    'United Kingdom': 'Europe', 'France': 'Europe', 'Germany': 'Europe',
    'Italy': 'Europe', 'Spain': 'Europe', 'Netherlands': 'Europe',
    'Switzerland': 'Europe', 'Sweden': 'Europe', 'Norway': 'Europe',
    'Denmark': 'Europe', 'Finland': 'Europe', 'Ireland': 'Europe',
    'Austria': 'Europe', 'Belgium': 'Europe', 'Portugal': 'Europe',
    'Greece': 'Europe', 'Luxembourg': 'Europe',
    'Japan': 'Asia Developed', 'Australia': 'Asia Developed',
    'Singapore': 'Asia Developed', 'Hong Kong': 'Asia Developed',
    'Taiwan': 'Asia Developed', 'South Korea': 'Asia Developed',
    'Israel': 'Asia Developed', 'New Zealand': 'Asia Developed',
    'China': 'Asia Emerging', 'India': 'Asia Emerging',
    'Indonesia': 'Asia Emerging', 'Thailand': 'Asia Emerging',
    'Malaysia': 'Asia Emerging', 'Philippines': 'Asia Emerging',
    'Vietnam': 'Asia Emerging',
    'Brazil': 'Latin America', 'Mexico': 'Latin America',
    'Chile': 'Latin America', 'Argentina': 'Latin America',
    'Saudi Arabia': 'MENA + Africa', 'United Arab Emirates': 'MENA + Africa',
    'Turkey': 'MENA + Africa', 'Egypt': 'MENA + Africa',
    'South Africa': 'MENA + Africa', 'Qatar': 'MENA + Africa',
}
REGIONS = ['North America', 'Europe', 'Asia Developed', 'Asia Emerging',
           'Latin America', 'MENA + Africa']

df = pd.read_csv('/home/user/cyclepapa/asymmetric_v2_universe_audit.csv')
df['region'] = df['country'].map(REGION_MAP).fillna('Other')

df = df.rename(columns={
    'pb_use': 'P/B', 'ev_ebitda_use': 'EV/EBITDA', 'rev_growth_use': 'Rev Growth',
    'ev_sales': 'EV/Sales', 'rev_growth_inflection': 'Rev Inflection',
    'fcf_yield': 'FCF Yield', 'op_margin_ex': 'Op Margin', 'roe_ex': 'ROE',
    'sec_rel_ev': 'Sector-Rel EV', 'asym_v2_score': 'v2 Score',
    'fip_d': 'FIP D (252d)', 'fip_w': 'FIP W (52w)', 'fip_m': 'FIP M (24m)',
    'pret_d': '12m Return', 'pret_m': '24m Return',
    'upside': 'Upside Score', 'floor': 'Floor Score',
    'quality': 'Quality Score', 'stealth': 'Stealth Score',
    'sector_used': 'Sector', 'market_cap_bucket': 'Cap',
    'roic_proxy': 'ROIC Proxy', 'debt_to_equity': 'Debt/Equity',
    'eps_q_growth': 'EPS Q Growth',
    'rs_fip_d': 'RS-FIP D', 'rs_fip_w': 'RS-FIP W',
    'rs_fip_w_inflection': 'RS-FIP W Inflect', 'rs_pret_d': 'RS 12m Return',
    'fip_w_minus_d': 'FIP W − D Gap',
    'asym_d_last': 'Volasym D', 'asym_w_last': 'Volasym W',
    'asym_m_last': 'Volasym M', 'asym_w_ma_last': 'Volasym W MA',
    'asym_m_ma_last': 'Volasym M MA',
    'asym_w_roc5': 'Volasym W RoC5', 'asym_m_roc3': 'Volasym M RoC3',
    'asym_m_dist50': 'Volasym M Dist-50',
    'nonzero_pct': 'Nonzero %', 'realized_vol_60d': '60d Vol',
    'last_price': 'Last Price', 'market_cap': 'Market Cap',
})

# Express percent-style columns as percentage points so they render
# 21.5 (not 21.5%) under the dashed accounting format.
PCT_COLS = ['12m Return', '24m Return', 'RS 12m Return',
            'Rev Growth', 'Rev Inflection', 'Op Margin', 'ROE',
            'FCF Yield', 'EPS Q Growth', 'Nonzero %', '60d Vol',
            'Volasym M Dist-50']
for c in PCT_COLS:
    if c in df.columns:
        df[c] = pd.to_numeric(df[c], errors='coerce') * 100

# --- Format codes ---
# Accounting convention: parens for negatives, accounting separators. Genuine
# zero renders as literal zero; en-dash is reserved for missing data only —
# render_table writes EN_DASH explicitly whenever pd.isna(v) is True.
EN_DASH = "–"
FMT_RAW   = '#,##0;(#,##0);0'
FMT_PP    = '#,##0.0;(#,##0.0);0.0'
FMT_RATIO = '#,##0.00;(#,##0.00);0.00'

NUM_FMT = {
    'Market Cap': FMT_RAW,
    '12m Return': FMT_PP, '24m Return': FMT_PP, 'RS 12m Return': FMT_PP,
    'Rev Growth': FMT_PP, 'Rev Inflection': FMT_PP,
    'Op Margin': FMT_PP, 'ROE': FMT_PP, 'FCF Yield': FMT_PP,
    'EPS Q Growth': FMT_PP, 'Nonzero %': FMT_PP, '60d Vol': FMT_PP,
    'Volasym D': FMT_PP, 'Volasym W': FMT_PP, 'Volasym M': FMT_PP,
    'Volasym W MA': FMT_PP, 'Volasym M MA': FMT_PP,
    'Volasym M Dist-50': FMT_PP,
    'FIP D (252d)': FMT_RATIO, 'FIP W (52w)': FMT_RATIO, 'FIP M (24m)': FMT_RATIO,
    'RS-FIP D': FMT_RATIO, 'RS-FIP W': FMT_RATIO, 'RS-FIP W Inflect': FMT_RATIO,
    'FIP W − D Gap': FMT_RATIO,
    'Volasym W RoC5': FMT_RATIO, 'Volasym M RoC3': FMT_RATIO,
    'P/B': FMT_RATIO, 'EV/EBITDA': FMT_RATIO, 'EV/Sales': FMT_RATIO,
    'ROIC Proxy': FMT_RATIO, 'Sector-Rel EV': FMT_RATIO,
    'Debt/Equity': FMT_RATIO, 'Last Price': FMT_RATIO,
    'v2 Score': FMT_RATIO, 'Upside Score': FMT_RATIO, 'Floor Score': FMT_RATIO,
    'Quality Score': FMT_RATIO, 'Stealth Score': FMT_RATIO,
}

# --- Type system (single font size, condensed letterforms) ---
def F(color=BLACK, bold=False, italic=False):
    """Single 11pt Times New Roman, with condensed letter spacing."""
    return Font(name=SERIF, size=SIZE, bold=bold, italic=italic, color=color,
                condense=True)

thin    = Side(border_style="thin",    color=LIGHT_GRAY)
hair    = Side(border_style="hair",    color=LIGHT_GRAY)
medium  = Side(border_style="medium",  color=BLACK)
border  = Border(left=thin, right=thin, top=thin, bottom=thin)
under   = Border(bottom=medium)
center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
center_tight = Alignment(horizontal="center", vertical="center")
leftA   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
rightA  = Alignment(horizontal="right",  vertical="center")
topL    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

fill_black   = PatternFill("solid", fgColor=DARK_GRAY)
fill_section = PatternFill("solid", fgColor=MID_GRAY)
fill_band    = PatternFill("solid", fgColor=PAPER_GRAY)
fill_white   = PatternFill("solid", fgColor=WHITE)
fill_panel   = PatternFill("solid", fgColor=PAPER_GRAY)

# --- Helpers ---
def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def render_table(ws, df_sub, start_row, title, cols):
    """Render section title row + a tightly-bordered table."""
    ncol = len(cols)
    # SECTION TITLE — bold, uppercase, on mid-gray fill
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncol)
    cell = ws.cell(row=start_row, column=1, value=title.upper())
    cell.font = F(WHITE, bold=True); cell.fill = fill_section
    cell.alignment = leftA
    ws.row_dimensions[start_row].height = 22

    # HEADER ROW — bold, uppercase, white on near-black, thin border
    hr = start_row + 1
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=hr, column=j, value=c.upper())
        cell.font = F(WHITE, bold=True); cell.fill = fill_black
        cell.alignment = center; cell.border = border
    ws.row_dimensions[hr].height = 28

    # BODY — Times New Roman 11pt, single size, accounting format
    for i, (_, row) in enumerate(df_sub[cols].iterrows()):
        r = hr + 1 + i
        fill = fill_band if i % 2 else fill_white
        for j, c in enumerate(cols, start=1):
            v = row[c]
            is_num = c in NUM_FMT
            if is_num:
                if pd.isna(v):
                    cell = ws.cell(row=r, column=j, value=EN_DASH)
                else:
                    cell = ws.cell(row=r, column=j, value=float(v))
                    cell.number_format = NUM_FMT[c]
                cell.alignment = rightA
            else:
                cell = ws.cell(row=r, column=j,
                               value=(v if pd.notna(v) else EN_DASH))
                cell.alignment = leftA
            cell.font = F(BLACK)
            cell.fill = fill
            cell.border = border
        ws.row_dimensions[r].height = 20
    return hr + len(df_sub) + 2

# --- Build workbook ---
wb = Workbook()

# 1. COVER ---------------------------------------------------------------
ws = wb.active; ws.title = "Cover"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = DARK_GRAY
ws.row_dimensions[1].height = 14

# Title banner — single 11pt size, hierarchy via weight + case + fill
ws.merge_cells('A2:H2')
ws['A2'] = "FIP — ASYMMETRY WORKBOOK"
ws['A2'].font = F(WHITE, bold=True)
ws['A2'].fill = fill_black
ws['A2'].alignment = center_tight
ws.row_dimensions[2].height = 56

ws.merge_cells('A3:H3')
ws['A3'] = "CONTINUOUS-INFORMATION MOMENTUM   ·   MULTI-METRIC ASYMMETRY   ·   SECTOR-RELATIVE"
ws['A3'].font = F(SOFT_GRAY, italic=True)
ws['A3'].fill = fill_black
ws['A3'].alignment = center_tight
ws.row_dimensions[3].height = 24

intro = [
    ("",                 ""),
    ("PURPOSE",
     "This workbook ranks the entire global universe of mid + large + small "
     "caps (plus US micro) by an asymmetric-opportunity composite built from "
     "the Frog-in-the-Pan (Da/Gurun/Warachka 2014) continuous-information "
     "signal, valuation floor across multiple metrics, fundamental catalyst, "
     "and capital-efficiency quality. All fundamentals are fetched fresh; no "
     "cached values are used in the ranking."),
    ("SCOPE",
     f"Universe: {len(df)} v2 survivors out of 2,321 unique tickers scanned. "
     "OHLC cache: 7,176 tickers spanning 44 countries. 365 pre-filter "
     f"candidates passed FIP and liquidity gates; {len(df)} cleared v2 gates."),
    ("METHODOLOGY",
     "(1) FIP smoothness: daily and weekly FIP both at or below −0.08 / "
     "−0.10 (formation 252 d, skip last 21 d per the paper). "
     "(2) Real liquidity: 60-day realized vol ≥ 0.008, ≥ 65 % non-zero return "
     "days, price ≥ $1. "
     "(3) Floor (multi-metric, any): P/B ≤ 2 OR EV/EBITDA ≤ 12 OR EV/Sales ≤ 3 "
     "OR FCF yield ≥ 3 %. "
     "(4) Catalyst (multi-confirmation): rev growth ≥ 5 % AND rev inflection "
     "≥ 0 AND (op margin ≥ 5 % OR quarterly EPS growth ≥ 0). "
     "(5) Survival: Debt/Equity ≤ 250. "
     "(6) Composite v2 score = √(Upside × Floor) × (0.7 + 0.3·Quality) × "
     "(0.8 + 0.2·Stealth)."),
    ("REGIONAL TAXONOMY",
     "North America   ·   Europe   ·   Asia Developed (JP / AU / SG / HK / "
     "TW / KR / IL / NZ)   ·   Asia Emerging (CN / IN / ID / TH / MY / PH / "
     "VN)   ·   Latin America   ·   MENA + Africa."),
    ("READING NOTES",
     "Each leg-leader tab shows the top picks per region by a single "
     "dimension. Single-metric tabs surface specialists. The composite tab "
     "is the headline ranking; the leg tabs surface concentrated bets when "
     "an investor prefers to weight one dimension. All numbers render in "
     "accounting convention: negatives in parentheses, en-dash for empty, "
     "right-aligned; percentage columns rendered as percentage points."),
    ("NUMBER CONVENTIONS",
     "Percentages: percentage points to 1 decimal (e.g. 21.5).   "
     "Ratios, scores, prices: 2 decimals (e.g. 12.54).   "
     "Raw $: accounting separators, 0 decimals.   "
     "Empty / zero: en-dash (–).   "
     "Typography: Times New Roman, 11 pt throughout, condensed."),
]
for i, (k, v) in enumerate(intro, start=5):
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    ws.cell(row=i, column=1, value=k).font = F(BLACK, bold=True)
    ws.cell(row=i, column=1).alignment = topL
    ws.cell(row=i, column=1).fill = fill_panel
    ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=8)
    cell = ws.cell(row=i, column=3, value=v)
    cell.font = F(BLACK)
    cell.alignment = topL
    cell.fill = fill_panel
    # Vertical rhythm: row height scaled to text length, biased tall
    ws.row_dimensions[i].height = max(18, 14 * max(1, len(v) // 75))

set_widths(ws, [10, 14] + [14]*6)

# 2. GLOSSARY -----------------------------------------------------------
ws = wb.create_sheet("Glossary"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = DARK_GRAY
glossary = [
    ("TERM", "DEFINITION"),
    ("FIP (Frog in the Pan)",
     "Da/Gurun/Warachka 2014. FIP = sgn(PRET) × (%neg − %pos) over the "
     "252-day formation period, skipping the most recent 21 trading days "
     "per the paper. A more-negative FIP for a winner means returns "
     "arrived as many small positive days (continuous information) rather "
     "than a few large jumps (discrete information) — predictive of "
     "stronger momentum continuation."),
    ("FIP_d / FIP_w / FIP_m",
     "Daily, weekly (52 w), and monthly (24 m) FIP scores. Smoothness "
     "across timeframes confirms the continuous-information setup."),
    ("12 m Return (pret_d)",
     "Cumulative return over the 252-day formation period (skip-last-21d)."),
    ("Pre-filter",
     "FIP_d ≤ −0.08 AND FIP_w ≤ −0.10 AND winner (return > 0) AND real "
     "liquidity (≥ 65 % non-zero return days, 60 d vol ≥ 0.008, price ≥ $1)."),
    ("Floor (multi-metric)",
     "Downside protection. A name passes the floor gate if ANY of "
     "P/B ≤ 2.0, EV/EBITDA ≤ 12, EV/Sales ≤ 3, or FCF yield ≥ 3 %."),
    ("Catalyst",
     "Upside trigger. Requires rev growth ≥ 5 % AND rev inflection ≥ 0 "
     "AND (op margin ≥ 5 % OR latest quarterly EPS growth ≥ 0)."),
    ("Survival",
     "Debt/Equity ≤ 250 — knocks out balance-sheet-impaired names."),
    ("Sector-Rel EV",
     "EV/EBITDA divided by the sector median. Below 1 = cheaper than peers, "
     "above 1 = premium."),
    ("ROIC Proxy",
     "Operating Margin × (1 / EV/Sales). Approximates return on invested "
     "capital independent of accounting leverage."),
    ("Upside Score",
     "Percentile rank composite of revenue growth and revenue inflection."),
    ("Floor Score",
     "Average percentile rank of low P/B, low EV/EBITDA, low EV/Sales, "
     "high FCF yield, and low sector-relative EV."),
    ("Quality Score",
     "Percentile rank of ROIC proxy. Captures capital efficiency."),
    ("Stealth Score",
     "Percentile rank of FIP_d + FIP_w (more negative = better). Measures "
     "continuous-information depth."),
    ("v2 Composite",
     "√(Upside × Floor) × (0.7 + 0.3·Quality) × (0.8 + 0.2·Stealth). "
     "Geometric on the two dimensions an investor cannot trade off, "
     "lifted by quality and stealth."),
]
for i, (k, v) in enumerate(glossary, start=2):
    is_header = (i == 2)
    fill = fill_black if is_header else (fill_band if i % 2 else fill_white)
    fc = WHITE if is_header else BLACK
    ws.cell(row=i, column=1, value=k).font = F(fc, bold=True)
    ws.cell(row=i, column=2, value=v).font = F(fc, bold=False)
    for c in (1, 2):
        cell = ws.cell(row=i, column=c)
        cell.alignment = topL
        cell.fill = fill
        cell.border = border
    ws.row_dimensions[i].height = 24 if is_header else max(28, 14 * max(1, len(v) // 90))
set_widths(ws, [30, 100])

# 3. COMPOSITE BY REGION ------------------------------------------------
def composite_tab():
    ws = wb.create_sheet("Composite by Region")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK_GRAY
    ws.merge_cells('A1:L1')
    cell = ws.cell(row=1, column=1, value="TOP PICKS BY COMPOSITE v2 SCORE, PER REGION")
    cell.font = F(WHITE, bold=True); cell.fill = fill_black
    cell.alignment = center_tight
    ws.row_dimensions[1].height = 32
    cols = ['symbol', 'name', 'country', 'Cap', 'Sector', '12m Return',
            'Rev Growth', 'Rev Inflection', 'P/B', 'EV/EBITDA', 'EV/Sales', 'v2 Score']
    set_widths(ws, [10,30,16,10,18,11,11,11,9,11,10,10])
    cur = 3
    for region in REGIONS:
        sub = df[df['region']==region].sort_values('v2 Score', ascending=False).head(10)
        if sub.empty: continue
        cur = render_table(ws, sub, cur, f"{region} — top {len(sub)} by composite score", cols)
composite_tab()

# 4-7. LEG LEADERS -------------------------------------------------------
def leg_tab(title, score_col, cols_extra):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK_GRAY
    cols = ['symbol','name','country','Cap','Sector', score_col] + cols_extra + ['v2 Score']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    cell = ws.cell(row=1, column=1, value=f"TOP PICKS BY {score_col.upper()}, PER REGION")
    cell.font = F(WHITE, bold=True); cell.fill = fill_black
    cell.alignment = center_tight
    ws.row_dimensions[1].height = 32
    widths = [10,28,16,10,18] + [13]*(len(cols)-5)
    set_widths(ws, widths)
    cur = 3
    for region in REGIONS:
        sub = df[df['region']==region].sort_values(score_col, ascending=False).head(10)
        if sub.empty: continue
        cur = render_table(ws, sub, cur,
                           f"{region} — top {len(sub)} by {score_col}", cols)

leg_tab("Upside Leaders",  'Upside Score',  ['Rev Growth','Rev Inflection','Op Margin','EPS Q Growth'])
leg_tab("Floor Leaders",   'Floor Score',   ['P/B','EV/EBITDA','EV/Sales','FCF Yield','Sector-Rel EV'])
leg_tab("Quality Leaders", 'Quality Score', ['ROIC Proxy','Op Margin','ROE','Debt/Equity'])
leg_tab("Stealth Leaders", 'Stealth Score', ['FIP D (252d)','FIP W (52w)','FIP M (24m)'])

# 8-14. MULTI-METRIC TABS ------------------------------------------------
def multi_metric_tab(title, specs):
    ws = wb.create_sheet(title); ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = DARK_GRAY
    cols = ['symbol','name','country','Cap','Sector'] + [s[0] for s in specs] + ['v2 Score']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    cell = ws.cell(row=1, column=1,
                   value=f"{title.upper()} — PER-REGION LEADERS ON EACH METRIC")
    cell.font = F(WHITE, bold=True); cell.fill = fill_black
    cell.alignment = center_tight
    ws.row_dimensions[1].height = 32
    widths = [10,28,16,10,18] + [12]*(len(specs)) + [10]
    set_widths(ws, widths)
    cur = 3
    for region in REGIONS:
        region_df = df[df['region']==region]
        if region_df.empty: continue
        ws.merge_cells(start_row=cur, start_column=1,
                       end_row=cur, end_column=len(cols))
        rc = ws.cell(row=cur, column=1, value=region.upper())
        rc.font = F(BLACK, bold=True); rc.fill = fill_panel
        rc.alignment = leftA
        rc.border = under
        ws.row_dimensions[cur].height = 22
        cur += 1
        for metric, ascending in specs:
            mask = region_df[metric].notna() & (region_df[metric] > -1e9)
            if not mask.any(): continue
            sub = region_df[mask].sort_values(metric, ascending=ascending).head(10)
            if sub.empty: continue
            label = f"  → top {len(sub)} by {metric} ({'lowest' if ascending else 'highest'})"
            cur = render_table(ws, sub, cur, label, cols)

multi_metric_tab(
    "Cheap Multiples",
    [('P/B', True), ('EV/EBITDA', True), ('EV/Sales', True),
     ('FCF Yield', False), ('Sector-Rel EV', True)],
)
multi_metric_tab(
    "Growth & Margin",
    [('Rev Growth', False), ('Rev Inflection', False),
     ('Op Margin', False), ('ROE', False), ('EPS Q Growth', False)],
)
multi_metric_tab(
    "FIP by Timeframe",
    [('FIP D (252d)', True), ('FIP W (52w)', True), ('FIP M (24m)', True),
     ('FIP W − D Gap', True)],
)
multi_metric_tab(
    "RS-FIP Leaders",
    [('RS-FIP D', True), ('RS-FIP W', True), ('RS-FIP W Inflect', True),
     ('RS 12m Return', False)],
)
multi_metric_tab(
    "Volatility Asymmetry",
    [('Volasym D', False), ('Volasym W', False), ('Volasym M', False),
     ('Volasym W RoC5', False), ('Volasym M RoC3', False),
     ('Volasym M Dist-50', False)],
)
multi_metric_tab(
    "Survival & Catalyst",
    [('Debt/Equity', True), ('EPS Q Growth', False),
     ('Rev Inflection', False), ('FCF Yield', False)],
)
multi_metric_tab(
    "Period Returns",
    [('12m Return', False), ('24m Return', False), ('RS 12m Return', False)],
)
multi_metric_tab(
    "Liquidity Quality",
    [('Nonzero %', False), ('60d Vol', False), ('Last Price', False)],
)

# 15. DRIFT AUDIT --------------------------------------------------------
ws = wb.create_sheet("Drift Audit"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = DARK_GRAY
ws.merge_cells('A1:H1')
cell = ws.cell(row=1, column=1,
               value="AUDIT — RANKING IS ACROSS THE ENTIRE UNIVERSE, NO CACHE SHORTCUTS")
cell.font = F(WHITE, bold=True); cell.fill = fill_black
cell.alignment = center_tight
ws.row_dimensions[1].height = 32

audit_rows = [
    ("Universe input CSVs",
     "Seven regional `leading_*` runs combined; 3,277 raw rows; 2,321 unique "
     "tickers across 44 countries."),
    ("OHLC cache",
     "7,176 tickers. 100 % of the pre-filter universe had OHLC."),
    ("FIP computed",
     "2,250 of 2,307 (post pref-share filter) had ≥ 300 trading days."),
    ("Pre-filter survivors",
     "365 (smooth FIP_d ≤ −0.08, smooth FIP_w ≤ −0.10, winner, ≥ 65 % "
     "non-zero days, vol ≥ 0.008, price ≥ $1)."),
    ("Fresh fundamentals",
     "ALL 365 pre-filter survivors re-fetched from yfinance — no cached "
     "P/B, EV/EBITDA, or rev growth used in final ranking."),
    ("Median P/B drift",
     "0.147. 91 rows shifted by more than 0.5."),
    ("Median EV/EBITDA drift",
     "0.847. 88 rows shifted by more than 2.0."),
    ("Final v2 survivors",
     f"{len(df)} after gates re-applied on fresh data."),
]
for i, (k, v) in enumerate(audit_rows, start=3):
    fill = fill_band if i % 2 else fill_white
    ws.cell(row=i, column=1, value=k).font = F(BLACK, bold=True)
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws.cell(row=i, column=2, value=v).font = F(BLACK)
    for c in (1, 2):
        cell = ws.cell(row=i, column=c)
        cell.alignment = topL
        cell.fill = fill
        cell.border = border
    ws.row_dimensions[i].height = max(28, 14 * max(1, len(v) // 95))

# Spacer + churn block
churn_start = len(audit_rows) + 5
ws.merge_cells(start_row=churn_start-1, start_column=1,
               end_row=churn_start-1, end_column=8)
hc = ws.cell(row=churn_start-1, column=1,
             value="CHURN VS. CACHED RANKING (TOP 25)")
hc.font = F(WHITE, bold=True); hc.fill = fill_section
hc.alignment = leftA
ws.row_dimensions[churn_start-1].height = 22

churn_rows = [
    ("Names that ENTERED top 25 with all-fresh refetch",
     "BP.L, DINO (HF Sinclair), HPE, NUE, SANM, SNX, BOTJ, 6806.T (Hirose)."),
    ("Names that EXITED top 25 with all-fresh refetch",
     "4042.T (Tosoh), 54E.F, BWFG, CII, CLST (Catalyst Bancorp), CTRE, "
     "FMBM, LQDT."),
    ("Implication",
     "Stale cached fundamentals materially affected ordering. The "
     "all-fresh refetch is the canonical rank."),
]
for j, (k, v) in enumerate(churn_rows, start=churn_start):
    fill = fill_band if j % 2 else fill_white
    ws.cell(row=j, column=1, value=k).font = F(BLACK, bold=True)
    ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=8)
    ws.cell(row=j, column=2, value=v).font = F(BLACK)
    for c in (1, 2):
        cell = ws.cell(row=j, column=c)
        cell.alignment = topL; cell.fill = fill; cell.border = border
    ws.row_dimensions[j].height = max(28, 14 * max(1, len(v) // 95))
set_widths(ws, [38] + [14]*7)

# 16. ALL SURVIVORS ------------------------------------------------------
ws = wb.create_sheet("All Survivors"); ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = DARK_GRAY
ws.merge_cells('A1:AE1')
cell = ws.cell(row=1, column=1,
               value=f"ALL {len(df)} v2 ASYMMETRIC SURVIVORS — RANKED BY COMPOSITE SCORE")
cell.font = F(WHITE, bold=True); cell.fill = fill_black
cell.alignment = center_tight
ws.row_dimensions[1].height = 32
cols_all = ['symbol','name','region','country','Cap','Sector',
            '12m Return','FIP D (252d)','FIP W (52w)','FIP M (24m)',
            'RS-FIP D','RS-FIP W','Volasym W','Volasym M',
            'Rev Growth','Rev Inflection','EPS Q Growth','Op Margin','ROE',
            'ROIC Proxy','Debt/Equity',
            'P/B','EV/EBITDA','EV/Sales','FCF Yield','Sector-Rel EV',
            'Upside Score','Floor Score','Quality Score','Stealth Score','v2 Score']
widths_all = [10,28,15,15,10,17, 11,11,11,11, 10,10,11,11,
              11,11,12,10,9, 11,11, 9,11,10,11,12, 12,11,12,12,10]
set_widths(ws, widths_all)
render_table(ws, df.sort_values('v2 Score', ascending=False), 3,
             "Full ranking (sorted by composite v2 score)", cols_all)

# Freeze panes on All Survivors so the symbol + name stay visible.
ws.freeze_panes = "C5"

out = '/home/user/cyclepapa/FIP_Asymmetry_Workbook.xlsx'
wb.save(out)
print(f"Saved {out}")
print(f"Tabs: {[s.title for s in wb.worksheets]}")
