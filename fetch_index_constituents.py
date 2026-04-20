"""Fetch current index constituents from official ETF/index provider files.

Each index is pulled from its tracking ETF's public holdings file. State
Street (SPDR) ships XLSX, BlackRock (iShares) ships CSV. We normalise to
data/constituents/<INDEX>.csv with columns: ticker, name, weight, sector.

Caveats:
- These are *current* constituents only — vendors don't publish history.
- Vendor URLs change occasionally; failures are logged but don't abort.
- iShares' CSVs have ~9 lines of metadata before the real table.

Run:
    pip install requests openpyxl
    python3 fetch_index_constituents.py
"""

from __future__ import annotations

import csv
import io
import sys
from pathlib import Path

import requests

OUT_DIR = Path(__file__).parent / "data" / "constituents"
OUT_DIR.mkdir(parents=True, exist_ok=True)

UA = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124 Safari/537.36"
)


def fetch(url: str) -> bytes:
    resp = requests.get(
        url,
        headers={"User-Agent": UA, "Accept": "*/*"},
        timeout=60,
        allow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def parse_ssga_xlsx(blob: bytes) -> list[dict]:
    """SSGA / SPDR holdings XLSX. Header row sits a few rows in; we scan
    down for a row containing 'Ticker' to anchor the table."""
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if any(cell.lower() == "ticker" for cell in cells):
            header_idx = i
            break
    if header_idx is None:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    out = []
    for r in rows[header_idx + 1:]:
        d = {h: r[i] for i, h in enumerate(header) if h}
        ticker = (d.get("Ticker") or "").strip() if isinstance(d.get("Ticker"), str) else ""
        if not ticker or ticker == "-":
            continue
        out.append(
            {
                "ticker": ticker,
                "name": str(d.get("Name", "")).strip(),
                "weight": str(d.get("Weight", "")).strip(),
                "sector": str(d.get("Sector", "")).strip(),
            }
        )
    return out


def parse_ishares_csv(blob: bytes) -> list[dict]:
    """iShares CSV. Skip leading metadata until we find the 'Ticker,Name,...'
    header row."""
    text = blob.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if line.lower().startswith("ticker,"):
            header_idx = i
            break
    if header_idx is None:
        return []
    reader = csv.DictReader(lines[header_idx:])
    out = []
    for row in reader:
        ticker = (row.get("Ticker") or "").strip()
        if not ticker or ticker == "-":
            continue
        out.append(
            {
                "ticker": ticker,
                "name": (row.get("Name") or "").strip(),
                "weight": (row.get("Weight (%)") or "").strip(),
                "sector": (row.get("Sector") or "").strip(),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Sources
# ---------------------------------------------------------------------------

SSGA = "https://www.ssga.com/us/en/intermediary/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{}.xlsx"


def ishares(pid: int, slug: str, ticker: str) -> str:
    """iShares holdings CSV needs the full product page path including slug
    plus the magic constant file id — not just the product ID."""
    return (
        f"https://www.ishares.com/us/products/{pid}/{slug}/"
        f"1467271812596.ajax?fileType=csv&fileName={ticker}_holdings"
        f"&dataType=fund"
    )

# Each entry: (index name, url, parser, source description)
SOURCES = [
    # SPDR XLSX
    ("SP500",         SSGA.format("spy"), parse_ssga_xlsx, "SPDR SPY"),
    ("Dow_Jones",     SSGA.format("dia"), parse_ssga_xlsx, "SPDR DIA"),
    ("SP_MidCap_400", SSGA.format("mdy"), parse_ssga_xlsx, "SPDR MDY"),
    # iShares CSVs (slug matters — without it the ajax endpoint 403/redirects)
    ("SP_SmallCap_600", ishares(239774, "ishares-core-sp-small-cap-etf", "IJR"), parse_ishares_csv, "iShares IJR"),
    ("NASDAQ_100",      ishares(239726, "ishares-nasdaq-100-etf", "QQQM"),      parse_ishares_csv, "iShares QQQM"),
    ("Russell_1000",  ishares(239707, "ishares-russell-1000-etf", "IWB"), parse_ishares_csv, "iShares IWB"),
    ("Russell_2000",  ishares(239710, "ishares-russell-2000-etf", "IWM"), parse_ishares_csv, "iShares IWM"),
    ("Russell_3000",  ishares(239714, "ishares-russell-3000-etf", "IWV"), parse_ishares_csv, "iShares IWV"),
    # Country / regional MSCI funds (proxy for DAX / CAC / FTSE / etc.)
    ("MSCI_United_Kingdom", ishares(239690, "ishares-msci-united-kingdom-etf", "EWU"), parse_ishares_csv, "iShares EWU"),
    ("MSCI_Germany",        ishares(239676, "ishares-msci-germany-etf",        "EWG"), parse_ishares_csv, "iShares EWG"),
    ("MSCI_France",         ishares(239674, "ishares-msci-france-etf",         "EWQ"), parse_ishares_csv, "iShares EWQ"),
    ("MSCI_Italy",          ishares(239680, "ishares-msci-italy-etf",          "EWI"), parse_ishares_csv, "iShares EWI"),
    ("MSCI_Spain",          ishares(239693, "ishares-msci-spain-etf",          "EWP"), parse_ishares_csv, "iShares EWP"),
    ("MSCI_Netherlands",    ishares(239686, "ishares-msci-netherlands-etf",    "EWN"), parse_ishares_csv, "iShares EWN"),
    ("MSCI_Belgium",        ishares(239669, "ishares-msci-belgium-etf",        "EWK"), parse_ishares_csv, "iShares EWK"),
    ("MSCI_Switzerland",    ishares(239695, "ishares-msci-switzerland-etf",    "EWL"), parse_ishares_csv, "iShares EWL"),
    ("MSCI_Sweden",         ishares(239692, "ishares-msci-sweden-etf",         "EWD"), parse_ishares_csv, "iShares EWD"),
    ("MSCI_Austria",        ishares(239667, "ishares-msci-austria-etf",        "EWO"), parse_ishares_csv, "iShares EWO"),
    ("MSCI_Japan",          ishares(239665, "ishares-msci-japan-etf",          "EWJ"), parse_ishares_csv, "iShares EWJ"),
    ("MSCI_Hong_Kong",      ishares(239678, "ishares-msci-hong-kong-etf",      "EWH"), parse_ishares_csv, "iShares EWH"),
    ("MSCI_Australia",      ishares(239668, "ishares-msci-australia-etf",      "EWA"), parse_ishares_csv, "iShares EWA"),
    ("MSCI_Canada",         ishares(239666, "ishares-msci-canada-etf",         "EWC"), parse_ishares_csv, "iShares EWC"),
    ("MSCI_Mexico",         ishares(239685, "ishares-msci-mexico-etf",         "EWW"), parse_ishares_csv, "iShares EWW"),
    ("MSCI_Brazil",         ishares(239612, "ishares-msci-brazil-etf",         "EWZ"), parse_ishares_csv, "iShares EWZ"),
    ("MSCI_South_Korea",    ishares(239681, "ishares-msci-south-korea-etf",    "EWY"), parse_ishares_csv, "iShares EWY"),
    ("MSCI_Taiwan",         ishares(239696, "ishares-msci-taiwan-etf",         "EWT"), parse_ishares_csv, "iShares EWT"),
    ("MSCI_Singapore",      ishares(239691, "ishares-msci-singapore-etf",      "EWS"), parse_ishares_csv, "iShares EWS"),
    ("MSCI_Emerging_Markets", ishares(239637, "ishares-msci-emerging-markets-etf", "EEM"),  parse_ishares_csv, "iShares EEM"),
    ("MSCI_EAFE",           ishares(239623, "ishares-msci-eafe-etf",           "EFA"), parse_ishares_csv, "iShares EFA"),
    ("MSCI_Europe",         ishares(239625, "ishares-core-msci-europe-etf",    "IEUR"), parse_ishares_csv, "iShares IEUR"),
    ("MSCI_Pacific",        ishares(239640, "ishares-core-msci-pacific-etf",   "IPAC"), parse_ishares_csv, "iShares IPAC"),
]


def write(name: str, rows: list[dict]) -> None:
    out = OUT_DIR / f"{name}.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["ticker", "name", "weight", "sector"])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print(f"  wrote {out.name}: {len(rows)} rows", flush=True)


def main() -> None:
    for name, url, parser, label in SOURCES:
        print(f"[{name}] <- {label}", flush=True)
        try:
            blob = fetch(url)
            rows = parser(blob)
            write(name, rows)
        except Exception as e:
            print(f"  FAILED ({type(e).__name__}): {e}", file=sys.stderr)
            write(name, [])


if __name__ == "__main__":
    main()
