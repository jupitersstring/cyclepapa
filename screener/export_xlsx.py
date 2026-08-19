"""Export the shortlist report to XLSX, mirroring the HTML artifact's
Harvard-minimalist style: Times New Roman, ONE font size (11pt) everywhere,
black on white, thin black rules only (under title, under table headers,
under last table row), no fills, no colors."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = 'screener/results/us_global_shortlist_2026-07-18.csv'
AUCT = 'screener/results/shortlist_auction_2026-07-18.csv'
OUT = 'screener/results/shortlist_report_2026-07-18.xlsx'

F = dict(name='Times New Roman', size=11)
FONT = Font(**F)
BOLD = Font(**F, bold=True)
THIN = Side(style='thin', color='000000')
B_BOT = Border(bottom=THIN)
LEFT = Alignment(horizontal='left', vertical='top')
RIGHT = Alignment(horizontal='right', vertical='top')
WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ticker -> (company, marker); markers per the report footnote:
# † mortgage REIT / BDC, ‡ deeply negative TTM FCF, § non-US listing
NAMES = {
    'RC': ('Ready Capital', '†'), 'BCIC': ('', '†'),
    'DXC': ('DXC Technology', ''), 'ABR': ('Arbor Realty Trust', '†'),
    'DAVA': ('Endava', ''), 'GPMT': ('Granite Point Mortgage', '†'),
    'EFOR': ('', ''), 'GTM': ('', ''),
    'NMFC': ('New Mountain Finance', '†'), 'BRSL': ('', '‡'),
    'SAMG': ('Silvercrest Asset Mgmt', ''), 'VRRM': ('Verra Mobility', ''),
    'AGNT': ('', ''), 'NUS': ('Nu Skin', ''),
    'LIGHT.AS': ('Signify', '§'), 'FVRR': ('Fiverr', ''),
    'JBI': ('Janus International', ''), 'CMCSA': ('Comcast', ''),
    'GSBD': ('Goldman Sachs BDC', '†'), 'SMPL': ('Simply Good Foods', ''),
    'HCKT': ('Hackett Group', ''), 'FISV': ('Fiserv', ''),
    'HGBL': ('Heritage Global', ''), 'OTEX': ('Open Text', ''),
    'GIB': ('CGI', ''), 'SGA': ('Saga Communications', ''),
    'IT': ('Gartner', ''), 'MAT': ('Mattel', ''), 'DOX': ('Amdocs', ''),
    'ACN': ('Accenture', ''), 'CTSH': ('Cognizant', ''),
    'PBH': ('Prestige Brands', ''), 'MMS': ('Maximus', ''),
    'G': ('Genpact', ''), 'CINT': ('CI&T', ''),
    'OTEX.TO': ('Open Text (Toronto)', '§'), 'SAM': ('Boston Beer', ''),
    'GIB-A.TO': ('CGI (Toronto, class A)', '§'), 'ACI': ('Albertsons', ''),
    'LULU': ('Lululemon Athletica', ''), 'BAH': ('Booz Allen Hamilton', ''),
    'PET.TO': ('Pet Valu', '§'), 'OPCH': ('Option Care Health', ''),
    'POOL': ('Pool Corp', ''), 'ACM': ('AECOM', ''),
    'MKTX': ('MarketAxess', ''), 'DLB': ('Dolby Laboratories', ''),
}
name_map = {t: v[0] for t, v in NAMES.items()}
mark_map = {t: v[1] for t, v in NAMES.items()}

CHECK, DASH = '✓', '—'


def put(ws, r, c, v, bold=False, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = BOLD if bold else FONT
    cell.alignment = align or (RIGHT if isinstance(v, (int, float)) else LEFT)
    if border:
        cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def table(ws, r0, headers, rows, fmts, widths):
    for j, h in enumerate(headers, 1):
        put(ws, r0, j, h, bold=True, border=B_BOT,
            align=LEFT if fmts[j - 1] is None else RIGHT)
    for i, row in enumerate(rows):
        last = i == len(rows) - 1
        for j, v in enumerate(row, 1):
            put(ws, r0 + 1 + i, j, v, border=B_BOT if last else None,
                fmt=fmts[j - 1])
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return r0 + len(rows) + 1


def num(v, nd=1):
    return None if pd.isna(v) else round(float(v), nd)


base = pd.read_csv(BASE).sort_values('norm_fcf_yield_5y_pct', ascending=False)
auct = pd.read_csv(AUCT).sort_values('alignment_score', ascending=False)

wb = Workbook()

# ---- Sheet 1: Shortlist --------------------------------------------------
ws = wb.active
ws.title = 'Shortlist'
put(ws, 1, 1, 'SHORTLIST: 200-WEEK LOWS × NORMALISED FCF × '
    'BUYBACKS — 2026-07-18', bold=True, border=B_BOT)
put(ws, 2, 1, 'Completed run over 8,109 listings across 24 indices; 568 '
    'stage-one survivors within 15% of their 200-week low; 47 passers. '
    'Yields are % of current market cap per year. Research triage only; '
    'not investment advice.', align=WRAP)
ws.row_dimensions[2].height = 30
rows = []
for _, r in base.iterrows():
    t = r['ticker']
    rows.append([t + mark_map.get(t, ''), name_map.get(t, '') or DASH,
                 num(r['pct_above_200w_low']), num(r['norm_fcf_yield_5y_pct']),
                 num(r['ttm_fcf_yield_pct']), num(r['buyback_yield_ttm_pct']),
                 num(r['net_buyback_yield_ttm_pct']), num(r['net_debt_5y_ebitda']),
                 num(r['market_cap_usd_m'], 0)])
table(ws, 4,
      ['Ticker', 'Company', 'Above 200w low %', 'Norm FCF yld 5y %',
       'TTM FCF yld %', 'Buyback yld %', 'Net buyback %', 'ND / EBITDA x',
       'Mkt cap M'],
      rows, [None, None, '0.0', '0.0', '0.0', '0.0', '0.0', '0.0', '#,##0'],
      [11, 24, 16, 17, 14, 13, 13, 13, 11])
ws.freeze_panes = 'A5'

# ---- Sheet 2: Auction Overlay --------------------------------------------
ws = wb.create_sheet('Auction Overlay')
put(ws, 1, 1, 'AUCTION & BREAKOUT OVERLAY — WEEKLY-LED (50 / 30 / 20)',
    bold=True, border=B_BOT)
put(ws, 2, 1, 'Dalton auction read from daily bars: value = volume-profile '
    'POC / 70% value area; states per the weekly 13-week bracket; RR is '
    'gross structural reward-to-risk before the estimated spread.',
    align=WRAP)
ws.row_dimensions[2].height = 30
mctx = {1: '+', 0: '0', -1: '−'}
rows = []
for _, r in auct.iterrows():
    rows.append([r['ticker'], str(r['auction_label']).replace('_', ' '),
                 num(r['alignment_score']), int(r['otf_up_weeks']),
                 CHECK if r['value_migration_up'] else DASH,
                 mctx.get(int(r['m_context']), '0'),
                 CHECK if r['d_excess_low'] else DASH,
                 num(r['rr_struct'], 2), num(r['cs_spread_pct'], 2),
                 num(r['dollar_vol_20d_m'])])
table(ws, 4,
      ['Ticker', 'Weekly auction state', 'Align', 'OTF wks', 'Val migr',
       'M', 'Excess', 'RR x', 'Est. spread %', '$ vol 20d M'],
      rows, [None, None, '0.0', '0', None, None, None, '0.00', '0.00',
             '#,##0.0'],
      [11, 20, 8, 9, 9, 5, 8, 8, 13, 12])
ws.freeze_panes = 'A5'

# ---- Sheet 3: Harmonics ---------------------------------------------------
ws = wb.create_sheet('Harmonics')
put(ws, 1, 1, 'HARMONIC COMPLETIONS — LAST 15 BARS', bold=True,
    border=B_BOT)
put(ws, 2, 1, 'Port of TradingView’s Harmonic Scanner '
    '(PatternsHunters) at base settings: ZigZag depth 10, error tolerance '
    '15%, X–D from up to the last 50 swing points. D at a swing low = '
    'buy, at a swing high = sell.', align=WRAP)
ws.row_dimensions[2].height = 30
rows = []
for _, r in auct.iterrows():
    for tf, sfx in (('D', '_d'), ('W', '_w')):
        sig = r.get('harm_signal' + sfx)
        if isinstance(sig, str) and sig:
            rows.append([r['ticker'], tf, sig,
                         str(r['harm_patterns' + sfx]).replace('+', ', '),
                         int(r['harm_bars_ago' + sfx]),
                         num(r['alignment_score']),
                         str(r['auction_label']).replace('_', ' ')])
rows.sort(key=lambda x: (x[2], -x[5]))
table(ws, 4,
      ['Ticker', 'TF', 'Signal', 'Patterns matched at D', 'Bars ago',
       'Align', 'Auction state'],
      rows, [None, None, None, None, '0', '0.0', None],
      [11, 5, 8, 52, 9, 8, 20])
ws.freeze_panes = 'A5'

# ---- Sheet 4: Notes --------------------------------------------------------
ws = wb.create_sheet('Notes')
ws.column_dimensions['A'].width = 110
notes = [
    ('DEFINITIONS & METHOD', True),
    ('Normalised FCF yield, 5y - mean(FCF, last five fiscal years) / '
     'current market cap x 100. Averaged across the cycle because a stock '
     'at multi-year lows usually has depressed trailing cash flow.', False),
    ('TTM FCF yield - sum(FCF, last four quarters) / market cap x 100. A '
     'large gap versus the normalised figure is the signature of a '
     'transient earnings shock.', False),
    ('Buyback yield, TTM - gross repurchase cash, last four quarters, / '
     'market cap x 100. The net variant subtracts share issuance and is '
     'the honest number.', False),
    ('Above 200-week low - last weekly close over the minimum weekly close '
     'of the trailing 200 weeks, minus one. The screen admits anything '
     'within 15%.', False),
    ('Thresholds: within 15% of 200-week low; normalised FCF yield >= 7%; '
     'gross buyback yield >= 3%; net debt <= 4.5x five-year EBITDA.', False),
    ('† Mortgage REITs / BDCs: reported operating cash flow is '
     'dominated by loan-book turnover, so an "FCF yield" is not an '
     'economic owner’s yield. ‡ TTM FCF deeply negative; the '
     'normalised figure is carried entirely by history. § Non-US '
     'listing; market cap in listing currency.', False),
    ('Auction overlay: monthly -> weekly -> daily hierarchy, weekly '
     'weighted heaviest. Failed auction low = probe under the trailing '
     '26-week low, bought back, never revisited; invalidation just under '
     'the probe. Acceptance = repeated closes outside the bracket plus '
     'value migration.', False),
    ('Data: Yahoo Finance. Research triage only; not investment advice.',
     False),
]
for i, (txt, bold) in enumerate(notes, 1):
    put(ws, i, 1, txt, bold=bold, align=WRAP,
        border=B_BOT if bold else None)
    ws.row_dimensions[i].height = 15 if bold else 45

wb.save(OUT)
print('wrote', OUT)
