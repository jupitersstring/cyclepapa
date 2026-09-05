"""Post-process an .xlsx into the Harvard house aesthetic.

Single typeface (Cambria) at a single size (10pt), black & white only,
no fills, whitespace-driven: a bold header row with a thin black rule
beneath it, muted grey for secondary text, gridlines off, header frozen.

This is applied to workbooks built via plain pandas.to_excel (which
default to Calibri 11 with no structure) so they match the hand-built
Harvard workbooks. For builders that already render through
build_harvard_workbook's helpers this pass is a harmless no-op-ish
normaliser (it only enforces font/size/fill, never touches values or
number formats).
"""
from __future__ import annotations
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


FONT_NAME = "Cambria"
FONT_SIZE = 10
INK = "FF000000"
MUTED = "FF707070"
RULE = "FFB0B0B0"

_NO_FILL = PatternFill(fill_type=None)
_HEADER_RULE = Border(bottom=Side(style="thin", color=INK))
_ROW_RULE = Border(bottom=Side(style="thin", color=RULE))

# Number formats: '–' shown for blank/zero so empty cells read as clean
# em-dashes rather than 0 (matches the build_harvard_workbook convention).
_FMT_MONEY = '#,##0;(#,##0);"–"'           # raw currency / large levels
_FMT_PCT = '0.0%;(0.0%);"–"'               # decimal-stored ratios shown as %
_FMT_RATIO = '0.00;(0.00);"–"'             # multiples / scores
_FMT_INT = '#,##0;(#,##0);"–"'             # counts


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# Header-name → number-format heuristic. Decimal-stored percentage fields
# (yields, margins, growth) get a % format; large levels get thousands
# separators; multiples / scores get two decimals.
_PCT_HINTS = ("yield", "margin", "_pct", "pct_", "growth", "momentum",
              "roe", "roa", "roce", "_yoy", "_cagr", "tax_rate",
              "return_yield", "off_52w", "accel")
_MONEY_HINTS = ("market_cap", "enterprise_value", "revenue", "ebitda_ttm",
                "fcf_ttm", "cfo_ttm", "cash", "debt", "equity", "assets",
                "goodwill", "intangibles", "ncav", "net_cash", "_ttm",
                "target_mean", "gross_profit", "shares_outstanding",
                "float_shares")
_RATIO_HINTS = ("ev_ebitda", "ev_ebit", "ev_sales", "p_e", "pb", "p_tb",
                "p_s", "p_ocf", "peg", "beta", "score", "asymmetry",
                "upside", "discount", "multiplier", "factor", "coverage",
                "net_debt_ebitda", "cash_pct_ev", "mcap_to_ncav",
                "debt_to_equity", "cash_conversion", "recommendation",
                "eta_", "hhi", "_n", "n_analysts")
_INT_HINTS = ("count", "_flag", "region_count", "segment_count",
              "product_line", "cluster_n", "years", "quarters",
              "_match", "_overrep", "_lt_2b", "_healthy", "af_")


# Literal junk strings that leak in when a NaN/None float is stringified
# (the `str(x or '')` gotcha: `NaN or ''` is NaN, so `str(NaN)` == "nan").
_BAD_TEXT = {"nan", "none", "inf", "-inf", "nat", "<na>",
             "#ref!", "#value!", "#div/0!", "#n/a", "nan%"}


def sanitize_nan_text(path: str) -> int:
    """Replace cells whose value is exactly a junk literal ('nan', 'None',
    'inf', error codes…) with an empty cell. Also nulls float inf/NaN.
    Surgical: only touches bad cells, never real data. Returns count fixed.
    """
    import math
    wb = load_workbook(path)
    fixed = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip().lower() in _BAD_TEXT:
                    cell.value = None
                    fixed += 1
                elif isinstance(v, float) and (math.isinf(v) or math.isnan(v)):
                    cell.value = None
                    fixed += 1
    if fixed:
        wb.save(path)
    return fixed


def _fmt_for_header(name) -> str | None:
    if not isinstance(name, str):
        return None
    h = name.strip().lower()
    # order matters: percent + int hints win over the broad money/ratio nets
    if any(k in h for k in _PCT_HINTS):
        return _FMT_PCT
    if any(k in h for k in _INT_HINTS):
        return _FMT_INT
    if any(k in h for k in _MONEY_HINTS):
        return _FMT_MONEY
    if any(k in h for k in _RATIO_HINTS):
        return _FMT_RATIO
    return None


def apply_harvard_style(path: str,
                        header_row: int = 1,
                        muted_text_cols: tuple = (),
                        freeze_header: bool = True,
                        zebra_rule: bool = False) -> None:
    """Rewrite every sheet of `path` into the Harvard aesthetic in place.

    Beyond fonts/fills, this now also applies sensible number formats by
    column-header heuristic (currency thousands-separators, %, ratios)
    and an AutoFilter on the header row so every data sheet is sortable
    and filterable in place — a big quality-of-life win for analysis.
    """
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row == 0 or max_col == 0:
            continue

        # Map each column to a number format from its header label
        col_fmt = {}
        for c in range(1, max_col + 1):
            hv = ws.cell(row=header_row, column=c).value
            fmt = _fmt_for_header(hv)
            if fmt:
                col_fmt[c] = fmt

        # Track content width per column for auto-sizing
        col_width = {c: 6 for c in range(1, max_col + 1)}

        for row in ws.iter_rows(min_row=1, max_row=max_row,
                                min_col=1, max_col=max_col):
            for cell in row:
                r, c = cell.row, cell.column
                # Strip any fill
                cell.fill = _NO_FILL
                is_header = (r == header_row)
                if is_header:
                    cell.font = Font(name=FONT_NAME, size=FONT_SIZE,
                                     bold=True, color=INK)
                    cell.alignment = Alignment(
                        horizontal="left" if isinstance(cell.value, str) else "right",
                        vertical="center", wrap_text=False)
                    cell.border = _HEADER_RULE
                else:
                    muted = c in muted_text_cols
                    cell.font = Font(name=FONT_NAME, size=FONT_SIZE,
                                     color=MUTED if muted else INK)
                    cell.alignment = Alignment(
                        horizontal="right" if _is_number(cell.value) else "left",
                        vertical="center", wrap_text=False)
                    if zebra_rule:
                        cell.border = _ROW_RULE
                    # Number format by column heuristic
                    if c in col_fmt and _is_number(cell.value):
                        cell.number_format = col_fmt[c]
                # Width tracking (use a formatted-length estimate for numbers)
                if cell.value is not None:
                    if _is_number(cell.value) and c in col_fmt and not is_header:
                        ln = len(f"{cell.value:,.0f}") + 3
                    else:
                        ln = len(str(cell.value))
                    if ln > col_width[c]:
                        col_width[c] = ln

        # Apply column widths (cap so prose columns don't explode the sheet)
        for c, w in col_width.items():
            ws.column_dimensions[get_column_letter(c)].width = min(max(w + 2, 8), 60)

        if freeze_header and max_row > header_row:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

        # AutoFilter across the header row → sortable/filterable in place
        if max_row > header_row:
            ws.auto_filter.ref = (
                f"{get_column_letter(1)}{header_row}:"
                f"{get_column_letter(max_col)}{max_row}")

    wb.save(path)
    # Clean any 'nan'/'inf'/error literals that leaked in upstream
    sanitize_nan_text(path)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--zebra", action="store_true",
                    help="add a thin grey rule under every data row")
    args = ap.parse_args()
    apply_harvard_style(args.path, zebra_rule=args.zebra)
    print(f"styled {args.path}", file=sys.stderr)


if __name__ == "__main__":
    main()
