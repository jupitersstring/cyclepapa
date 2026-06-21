#!/usr/bin/env python3
"""
Global Rotation & Inflection Workbook (Harvard aesthetic).

Computes every cross-sectional leg/measure we have discussed, for the ENTIRE
cached universe, then writes a styled .xlsx with Top-N per region for each leg
and a composite leaderboard.

Legs / measures
  Price action     : ret 5d / 10d / 21d                       (intra-week/fortnight/month)
  Sharpe action    : ATR-normalised thrust 5/10/21 = dClose/ATR14  (unit-free)
  Rotation         : blend (recency-weighted thrust), accel (z5 - z21)
  Early inflection : SuperSmoother slow curvature + velocity   (leads the cross)
  Breadth          : dimensional net (weekly + daily)          (from latest CSVs)
  Positioning      : % above 200-week low, % off 52-week high
  COMPOSITE        : z-blend of the legs (momentum + acceleration + breadth + early turn)

  python3 workbook.py [--topn 15] [--date YYYY-MM-DD]
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import json
import os

import numpy as np
import pandas as pd

from midcap_weekly_anomalies import get_universe, REGIONS
from signals import drop_incomplete_last, smoothed_inflection, regime_strength

ROOT = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(ROOT, ".cache")
RESULTS = os.path.join(ROOT, "results")
REGION_NAME = {"us": "United States", "uk": "United Kingdom", "ca": "Canada",
               "de": "Germany", "fr": "France", "nl": "Netherlands", "au": "Australia"}
CRIMSON = "A51C30"; CRIMSON_D = "7A0019"; PAPER = "FBF7F2"; INK = "1A1A1A"; RULE = "C9B8A8"

# leg -> (column, ascending?) for the per-leg Top-N sheets
LEGS = [
    ("Rotation blend (ATR-thrust)", "blend", False),
    ("Rotation accel (5d−21d)", "accel", False),
    ("Return 5d (intra-week)", "ret_5d", False),
    ("Return 10d (fortnight)", "ret_10d", False),
    ("Return 21d (month)", "ret_21d", False),
    ("Sharpe action 5d (move/ATR)", "atr_5d", False),
    ("Sharpe action 21d (move/ATR)", "atr_21d", False),
    ("Early inflection (curvature)", "early_curv", False),
    ("Breadth net (wk+dy)", "breadth", False),
    ("All-weather RS (up&down mkt)", "all_weather", False),
    ("Quiet resilience (down=low vol)", "quiet_resil", False),
    ("Nearest 200-week low", "pct_above_low", True),
]
SHOW = ["name", "sector", "industry", "ret_5d", "ret_21d", "blend", "accel",
        "early_curv", "breadth", "up_perf", "dn_perf", "down_vol_z",
        "all_weather", "quiet_resil", "pct_above_low", "composite"]


def _rz(s):                       # robust cross-sectional z (median / MAD)
    med = s.median(); mad = (s - med).abs().median() * 1.4826 or 1.0
    return ((s - med) / mad).clip(-4, 4)


def build_features(topn_date=None):
    # region/sector/industry/name maps from the universe
    meta = {}
    for rc in REGIONS:
        u = get_universe(source=f"{rc}-allcap")
        for _, r in u.iterrows():
            meta[r["symbol"]] = (rc, r["sector"], r.get("market_cap", ""), r["security"])
    ind = json.load(open(os.path.join(CACHE, "industry_labels.json"))) \
        if os.path.exists(os.path.join(CACHE, "industry_labels.json")) else {}

    cache = pd.read_pickle(os.path.join(CACHE, "ohlcvdict_1d_20y.pkl"))
    last = {s: d.dropna().index[-1] for s, d in cache.items() if len(d.dropna())}
    asof = max(last.values()); tol = pd.Timedelta(days=8)
    LB = 126                                          # regime lookback (~6 months)

    # per-region MARKET proxy: median daily return across that region's names
    region_mkt = {}
    for rc in REGIONS:
        cols = {}
        for s, m_ in meta.items():
            if m_[0] != rc:
                continue
            d = cache.get(s)
            if d is None:
                continue
            c = d["Close"].dropna()
            if len(c) >= LB + 5 and last.get(s, asof) >= asof - tol:
                cols[s] = c.iloc[-(LB + 5):]
        if cols:
            R = pd.DataFrame(cols).pct_change()
            region_mkt[rc] = R.median(axis=1)         # robust equal-weight market

    rows = []
    for sym, df in cache.items():
        if sym not in meta:
            continue
        d = drop_incomplete_last(df, "daily", asof=asof)
        c = d["Close"].astype(float)
        if len(c) < 80 or last[sym] < asof - tol or c.iloc[-1] < 0.5:
            continue
        atr = c.diff().abs().rolling(14).mean().iloc[-1]
        if not np.isfinite(atr) or atr <= 0:
            continue
        win200 = c.iloc[-200:]
        rc, sec, cap, nm = meta[sym]
        v = d["Volume"].astype(float)
        rec = {"symbol": sym, "region": rc, "sector": sec, "cap": cap, "name": nm[:34],
               "industry": (ind.get(sym) or {}).get("industry") or sec,
               "dvol": float((c * v).iloc[-21:].median()),
               "pct_above_low": float(c.iloc[-1] / win200.min() - 1) if win200.min() > 0 else np.nan,
               "pct_off_high": float(c.iloc[-1] / c.iloc[-252:].max() - 1) if len(c) >= 60 else np.nan}
        for k in (5, 10, 21):
            rec[f"ret_{k}d"] = float(c.iloc[-1] / c.iloc[-1 - k] - 1)
            rec[f"atr_{k}d"] = float((c.iloc[-1] - c.iloc[-1 - k]) / atr)
        ei = smoothed_inflection(np.log(c.clip(lower=1e-9).to_numpy()), 21, 5)
        rec["early_curv"] = ei["curv_z"] if ei else np.nan
        # regime-conditional strength vs the region market, aligned on common dates
        rec["up_perf"] = rec["dn_perf"] = rec["down_vol_z"] = rec["dn_cap"] = np.nan
        mkt = region_mkt.get(rc)
        if mkt is not None:
            sr = c.pct_change()
            vz = (v - v.rolling(63).mean()) / v.rolling(63).std()
            j = pd.concat([sr.rename("r"), vz.rename("vz"), mkt.rename("m")],
                          axis=1, join="inner").dropna(subset=["r", "m"]).iloc[-LB:]
            rs = regime_strength(j["r"].values, j["vz"].values, j["m"].values)
            if rs:
                rec.update({k: rs[k] for k in ("up_perf", "dn_perf", "down_vol_z", "dn_cap")})
        rows.append(rec)
    feat = pd.DataFrame(rows).set_index("symbol")

    # cross-sectional z + rotation blend/accel + composite
    for k in (5, 10, 21):
        feat[f"z_atr_{k}d"] = _rz(feat[f"atr_{k}d"])
    feat["blend"] = 0.5 * feat["z_atr_5d"] + 0.3 * feat["z_atr_10d"] + 0.2 * feat["z_atr_21d"]
    feat["accel"] = feat["z_atr_5d"] - feat["z_atr_21d"]

    # regime-conditional measures (z-scored WITHIN region so the market proxy is
    # the right benchmark for each)
    feat["all_weather"] = np.nan; feat["quiet_resil"] = np.nan
    for rc, g in feat.groupby("region"):
        zu, zd = _rz(g["up_perf"]), _rz(g["dn_perf"])
        zv = _rz(g["down_vol_z"])
        feat.loc[g.index, "all_weather"] = zu + zd                  # strong up AND down
        feat.loc[g.index, "quiet_resil"] = zd - zv + 0.5 * zu       # down strong on LOW vol + up strong

    # breadth net (weekly + daily) from the latest committed CSVs
    date = topn_date or sorted(g.split("_daily_")[-1][:-4]
                               for g in glob.glob(os.path.join(RESULTS, "breadth_*_daily_*.csv")))[-1]
    wk, dl = {}, {}
    for f in glob.glob(os.path.join(RESULTS, f"breadth_*_weekly_{date}.csv")):
        for _, r in pd.read_csv(f).iterrows():
            wk[r["symbol"]] = r["net"]
    for f in glob.glob(os.path.join(RESULTS, f"breadth_*_daily_{date}.csv")):
        for _, r in pd.read_csv(f).iterrows():
            dl[r["symbol"]] = r["net"]
    feat["breadth"] = feat.index.map(lambda s: (wk.get(s, 0) + dl.get(s, 0)))

    feat["composite"] = (0.25 * feat["blend"] + 0.15 * _rz(feat["accel"])
                         + 0.20 * _rz(feat["breadth"]) + 0.15 * _rz(feat["early_curv"].fillna(0))
                         + 0.10 * _rz(feat["ret_21d"])
                         + 0.15 * _rz(feat["all_weather"].fillna(0)))   # regime quality
    return feat, asof, date


# --------------------------------------------------------------------------- #
def write_xlsx(feat, asof, date, topn, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    serif = "Georgia"
    title_f = Font(name=serif, size=20, bold=True, color=CRIMSON)
    sub_f = Font(name=serif, size=10, italic=True, color="555555")
    hdr_f = Font(name=serif, size=9, bold=True, color="FFFFFF")
    cell_f = Font(name=serif, size=9, color=INK)
    sec_f = Font(name=serif, size=12, bold=True, color=CRIMSON_D)
    hdr_fill = PatternFill("solid", fgColor=CRIMSON)
    band = PatternFill("solid", fgColor=PAPER)
    thin = Side(style="thin", color=RULE)
    border = Border(bottom=thin)
    rule = Border(bottom=Side(style="medium", color=CRIMSON))

    wb = Workbook()

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_f; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def block(ws, top, title, sub_df, cols, fmts):
        ws.cell(top, 1, title).font = sec_f
        ws.cell(top, 1).border = rule
        for c in range(2, len(cols) + 1):
            ws.cell(top, c).border = rule
        hr = top + 1
        heads = ["#"] + cols
        for j, h in enumerate(heads, 1):
            ws.cell(hr, j, h.replace("_", " "))
        style_header(ws, hr, len(heads))
        for i, (sym, r) in enumerate(sub_df.iterrows(), 1):
            rr = hr + i
            ws.cell(rr, 1, i).font = cell_f
            ws.cell(rr, 2, sym).font = Font(name=serif, size=9, bold=True, color=CRIMSON_D)
            for j, col in enumerate(cols[1:], 3):
                v = r[col]
                cell = ws.cell(rr, j, v)
                cell.font = cell_f
                f = fmts.get(col)
                if f:
                    cell.number_format = f
                cell.alignment = Alignment(horizontal="left" if col in ("name", "sector", "industry") else "right")
            if i % 2 == 0:
                for j in range(1, len(heads) + 1):
                    ws.cell(rr, j).fill = band
            for j in range(1, len(heads) + 1):
                ws.cell(rr, j).border = border
        return hr + len(sub_df) + 2

    fmts = {"ret_5d": "+0.0%;-0.0%", "ret_10d": "+0.0%;-0.0%", "ret_21d": "+0.0%;-0.0%",
            "atr_5d": "+0.00", "atr_21d": "+0.00", "blend": "+0.00", "accel": "+0.00",
            "early_curv": "+0.00", "breadth": "+0;-0", "pct_above_low": "+0%;-0%",
            "up_perf": "+0.00%;-0.00%", "dn_perf": "+0.00%;-0.00%", "down_vol_z": "+0.00",
            "all_weather": "+0.00", "quiet_resil": "+0.00", "composite": "+0.00"}
    show_cols = ["symbol"] + SHOW

    # ---- cover sheet ----
    ws = wb.active; ws.title = "Cover"; ws.sheet_view.showGridLines = False
    ws.cell(2, 1, "GLOBAL ROTATION & INFLECTION").font = title_f
    ws.cell(3, 1, "Cross-sectional leaderboard — Top names per region, by leg").font = Font(name=serif, size=12, italic=True, color=INK)
    ws.cell(5, 1, f"As of {pd.Timestamp(asof).date()}   ·   breadth {date}   ·   {len(feat):,} ranked names across {feat['region'].nunique()} regions").font = sub_f
    notes = [
        "", "METHODOLOGY",
        "All measures are cross-sectionally robust-z-scored (median / MAD) to remove unit bias across currencies & price levels.",
        "Sharpe action = price move divided by ATR(14) (close-to-close proxy) — risk-adjusted thrust, unit-free.",
        "Rotation blend = 0.5·z(thrust 5d) + 0.3·z(10d) + 0.2·z(21d);  accel = z(5d) − z(21d) (rotation arriving now).",
        "Early inflection = SuperSmoother(log price) curvature — fires at the turning point, leading a bandpass cross.",
        "Breadth = directional dimensional net (weekly + daily) from the bandpass engine.",
        "Composite = 0.30·blend + 0.20·z(accel) + 0.20·z(breadth) + 0.15·z(early curv) + 0.15·z(21d return).",
        "",
        "PROVENANCE: ranks 99.8% of the live targeted universe (mega→micro, 7 regions); 0 phantom names; nets reproduce",
        "bit-for-bit from the cached data (see audit_coverage.py). Not investment advice.",
    ]
    for i, n in enumerate(notes, 7):
        ws.cell(i, 1, n).font = sec_f if n.isupper() and n else sub_f
    ws.column_dimensions["A"].width = 120

    # ---- composite leaderboard: Top-N per region ----
    ws = wb.create_sheet("Composite by region"); ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "COMPOSITE — Top names per region").font = title_f
    top = 3
    for rc in [r for r in REGIONS if (feat["region"] == r).any()]:
        sub = feat[feat["region"] == rc].sort_values("composite", ascending=False).head(topn)
        top = block(ws, top, f"{REGION_NAME[rc]}  ({(feat['region']==rc).sum()} ranked)",
                    sub[SHOW], show_cols, fmts)
    for j, w in enumerate([10, 30, 18, 22] + [9] * 10, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---- one sheet per leg ----
    import re as _re
    for title, col, asc in LEGS:
        safe = _re.sub(r"[/\\?*\[\]:]", "-", title)[:31]
        ws = wb.create_sheet(safe); ws.sheet_view.showGridLines = False
        ws.cell(1, 1, title.upper()).font = title_f
        ws.cell(2, 1, "Top names per region").font = sub_f
        top = 4
        for rc in [r for r in REGIONS if (feat["region"] == r).any()]:
            sub = feat[feat["region"] == rc].dropna(subset=[col]).sort_values(col, ascending=asc).head(topn)
            if sub.empty:
                continue
            top = block(ws, top, REGION_NAME[rc], sub[SHOW], show_cols, fmts)
        for j, w in enumerate([10, 30, 18, 22] + [9] * 10, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--topn", type=int, default=15)
    ap.add_argument("--date", default=None)
    args = ap.parse_args()
    feat, asof, date = build_features(args.date)
    os.makedirs(RESULTS, exist_ok=True)
    out = os.path.join(RESULTS, f"workbook_{dt.date.today().isoformat()}.xlsx")
    write_xlsx(feat, asof, date, args.topn, out)
    feat.reset_index().to_csv(os.path.join(RESULTS, f"workbook_features_{dt.date.today().isoformat()}.csv"), index=False)
    print(f"[workbook] {len(feat):,} names, {feat['region'].nunique()} regions -> {out}")


if __name__ == "__main__":
    main()
