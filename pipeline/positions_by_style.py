"""Positions-by-Substyle — the ORIGINAL view, restored with full detail.

For each fund substyle (Deep Value, US Activists Tier 1, Biotech Specialists,
Family Offices, etc.) list every fund's actual positions — conviction
holdings, material adds, new positions, 13D filings — in clean columns
parsed from the source raw_text. NOT collapsed, NOT scored. Just "what is
each kind of fund holding", organized properly.

Produces one workbook sheet per style FAMILY; within each sheet positions
are grouped by exact substyle -> fund -> section, so you can read e.g. all
Deep Value funds together. Output: positions_by_style.xlsx
"""
import os, re, sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "positions_by_style.xlsx")

# Family routing: exact-substyle -> family sheet. Order matters (first match).
FAMILY_RULES = [
    (r"biotech|healthcare", "Biotech & Healthcare"),
    (r"deep value|ben graham|bill miller|ugly duckling|crazy cheap", "Deep Value"),
    (r"vic-to-fund|skin-in-game|fat-pitch|value & multi-strat|value legends|institutional value|large-cap value|elite shorts", "Value Legends & VIC"),
    (r"quality|moat|monopoly|buffett|compounder|15% roe|long-duration", "Quality Compounders"),
    (r"p5 |microcap activist|microcap concentrat|microcap special|microcap discovery|microcap deep|nanocap|illiquid small", "Microcap Activists & P5"),
    (r"small-cap / multibagger|small-cap / deep|small concentrated activists|smid|small-cap div|small-cap", "Small-cap & Multibagger"),
    (r"us activist", "US Activists"),
    (r"uk activist|international activist|em governance|multi-jurisdiction", "Intl & UK Activists"),
    (r"activist|special sit|lichtenstein|conglomerate|long/short activist|catalyst-driven|wound down|advisory", "Other Activists & Special Sits"),
    (r"distressed|event-driven|post-bk|credit", "Distressed & Event-Driven"),
    (r"european|asia|spanish|canadian|japan|global quality \+ em|em/tech|quality global|sydney|london/milan|london/hk", "European / Asia / EM"),
    (r"tiger cub|l/s legends|concentrated growth|ultra-concentrated|growth -", "Growth & Tiger Cubs"),
    (r"family office|family-office|legendary family|allocator|macro / trend family", "Family Offices"),
    (r"mega multi-strat|quant|multi-strategy macro|soros|druckenmiller|appaloosa", "Mega Multi-Strat & Macro"),
    (r"cta|trend follow", "CTA / Trend"),
    (r"warrant", "Warrant Specialists"),
    (r"bank specialist|financial services|holdco|pe-style|concentrated microcap|concentrated quality \(sub|sma", "Concentrated / Specialist"),
]

SECTION_LABEL = {1: "CONVICTION", 2: "13D/13G", 3: "NEW POS", 4: "ADD"}
SECTION_ORDER = {4: 0, 3: 1, 2: 2, 1: 3, 0: 4}   # ADDs/NEWs first (actionable), then 13Ds, then conviction holds

def family_for(substyle):
    s = (substyle or "").lower()
    for pat, fam in FAMILY_RULES:
        if re.search(pat, s):
            return fam
    return "Other / Unclassified"

def parse_raw(raw, ticker, pct_value=None, dollar_m=None):
    """Parse 'TICKER | Company | X% portfolio | $YM | change | url' into fields.
    Falls back to the typed pct_value / dollar_m columns when raw doesn't parse."""
    if not raw:
        pct = f"{pct_value:.2f}%" if pct_value is not None else ""
        dol = f"${dollar_m:.1f}M" if dollar_m else ""
        return "", pct, dol, ""
    parts = [p.strip() for p in raw.split("|")]
    company = pct = dollar = ""
    notes = []
    # First pass: find structured fields
    for p in parts:
        if not p or p.startswith("http"):
            continue
        if p == ticker or p.startswith(ticker + " ") or p.startswith(ticker + "("):
            m = re.search(r"\(([^)]+)\)", p)
            if m and not company:
                company = m.group(1)[:50]
            continue
        # Standalone % with portfolio/book context = position size
        if "%" in p and not pct:
            pm = re.search(r"([\+\-]?\d+(?:\.\d+)?)\s*%", p)
            if pm:
                pct = pm.group(0)
                # Whole-cell tags get suppressed only if they were JUST the %
                if re.fullmatch(r"\s*[\+\-]?\d+(?:\.\d+)?\s*%\s*(portfolio|of\s+co\.?|of\s+company|of\s+book)?\s*", p, re.I):
                    continue
        # Dollar amount
        if "$" in p and not dollar:
            dm = re.search(r"\$[\d,.]+\s*[MBK]?", p)
            if dm:
                dollar = dm.group(0)
                if re.fullmatch(r"\s*\$[\d,.]+\s*[MBK]?\s*", p):
                    continue
        # Company name: short, non-numeric, not yet captured
        if not company and not re.search(r"\d", p) and 2 < len(p) < 50:
            company = p[:50]
            continue
        notes.append(p)
    # Fallbacks from typed columns
    if not pct and pct_value is not None:
        pct = f"{pct_value:.2f}%"
    if not dollar and dollar_m:
        dollar = f"${dollar_m:.1f}M"
    change = " / ".join(n for n in notes if n)[:160]
    return company, pct, dollar, change

def run():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    # gather all positions with substyle
    rows = list(conn.execute("""
        SELECT fm.fund_group AS substyle, fp.fund, fp.section, fp.ticker,
               fp.company, fp.pct_value, fp.pct_kind, fp.dollar_m, fp.raw_text
        FROM fund_positions fp JOIN fund_meta fm ON fm.fund = fp.fund
        WHERE fp.ticker IS NOT NULL AND fp.section IN (1,2,3,4)"""))

    # bucket by family
    by_family = {}
    for r in rows:
        fam = family_for(r["substyle"])
        by_family.setdefault(fam, []).append(r)

    # cross-tribe consensus: tickers held by N+ different macro families
    family_of_ticker = {}
    for fam, frs in by_family.items():
        for x in frs:
            family_of_ticker.setdefault(x["ticker"], set()).add(fam)
    cross_consensus = sorted(
        ((t, fams) for t, fams in family_of_ticker.items() if len(fams) >= 3),
        key=lambda x: -len(x[1]))

    HEAD = Font(bold=True, color="FFFFFF", size=11)
    HFILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    SUB = Font(bold=True, size=11, color="1F4E78")
    SUBFILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    FUNDF = Font(bold=True, italic=True, size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")

    wb = openpyxl.Workbook(); del wb["Sheet"]
    from openpyxl.utils import get_column_letter

    # index sheet
    idx = wb.create_sheet("INDEX")
    idx["A1"] = "POSITIONS BY FUND SUBSTYLE — what each kind of fund holds"
    idx["A1"].font = Font(bold=True, size=13)
    idx["A3"] = "Family sheet"; idx["B3"] = "# substyles"; idx["C3"] = "# positions"
    for c in ("A3","B3","C3"): idx[c].font = HEAD; idx[c].fill = HFILL

    # Cross-tribe consensus sheet — names held across multiple investing families
    xc = wb.create_sheet("CROSS_TRIBE_CONSENSUS")
    xc["A1"] = "TICKERS HELD ACROSS MULTIPLE INVESTING FAMILIES (>=3) — the universal picks"
    xc["A1"].font = Font(bold=True, size=12)
    xc_hdr = ["Ticker", "# families", "Families", "Total positions", "Funds holding"]
    for i, h in enumerate(xc_hdr, 1):
        cell = xc.cell(row=3, column=i, value=h); cell.font = HEAD; cell.fill = HFILL
    xc.freeze_panes = "A4"
    xr = 4
    for tkr, fams in cross_consensus[:80]:
        all_pos = [x for fr in by_family.values() for x in fr if x["ticker"] == tkr]
        all_funds = sorted({x["fund"] for x in all_pos})
        xc.cell(row=xr, column=1, value=tkr).font = Font(bold=True)
        xc.cell(row=xr, column=2, value=len(fams))
        xc.cell(row=xr, column=3, value=", ".join(sorted(fams)))
        xc.cell(row=xr, column=4, value=len(all_pos))
        xc.cell(row=xr, column=5, value="; ".join(f[:30] for f in all_funds[:10]))
        for col in range(1, 6):
            xc.cell(row=xr, column=col).alignment = WRAP
        xr += 1
    for i, w in enumerate([8, 12, 70, 14, 90], 1):
        xc.column_dimensions[get_column_letter(i)].width = w

    # Flat sheet for arbitrary filtering across ALL positions
    flat = wb.create_sheet("ALL_POSITIONS")
    flat_hdr = ["Family", "Substyle", "Fund", "Section", "Ticker", "Company",
                "% (book/co)", "$M", "Change / notes"]
    for i, h in enumerate(flat_hdr, 1):
        cell = flat.cell(row=1, column=i, value=h); cell.font = HEAD; cell.fill = HFILL
    flat.freeze_panes = "A2"
    flat.auto_filter.ref = "A1:I1"
    fr_idx = 2

    irow = 4
    used_names = set()
    for fam in sorted(by_family, key=lambda f: -len(by_family[f])):
        frows = by_family[fam]
        safe = fam.replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "")
        sheet_name = safe[:31]
        if sheet_name in used_names: sheet_name = (safe[:28] + "_2")[:31]
        used_names.add(sheet_name)
        substyles = sorted({r["substyle"] for r in frows})
        idx.cell(row=irow, column=1, value=fam)
        idx.cell(row=irow, column=2, value=len(substyles))
        idx.cell(row=irow, column=3, value=len(frows))
        irow += 1

        ws = wb.create_sheet(sheet_name)
        # Top of each family sheet: consensus block — most-held tickers within this tribe
        consensus = {}
        for x in frows:
            consensus.setdefault(x["ticker"], set()).add(x["fund"])
        top = sorted(consensus.items(), key=lambda c: -len(c[1]))[:15]
        ws["A1"] = f"▼ {fam} — top consensus picks within this tribe"
        ws["A1"].font = SUB; ws["A1"].fill = SUBFILL
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.cell(row=2, column=1, value="Ticker").font = HEAD
        ws.cell(row=2, column=1).fill = HFILL
        ws.cell(row=2, column=2, value="# funds in tribe").font = HEAD
        ws.cell(row=2, column=2).fill = HFILL
        ws.cell(row=2, column=3, value="Funds").font = HEAD
        ws.cell(row=2, column=3).fill = HFILL
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=8)
        cr = 3
        for tkr, funds in top:
            if len(funds) < 2: break
            ws.cell(row=cr, column=1, value=tkr).font = Font(bold=True)
            ws.cell(row=cr, column=2, value=len(funds))
            ws.cell(row=cr, column=3, value=", ".join(sorted(funds))[:200])
            ws.merge_cells(start_row=cr, start_column=3, end_row=cr, end_column=8)
            cr += 1
        cr += 1
        headers = ["Substyle","Fund","Section","Ticker","Company","% (book/co)","$M","Change / notes"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=cr, column=i, value=h); c.font = HEAD; c.fill = HFILL
        ws.freeze_panes = f"A{cr+1}"
        r = cr + 1
        # group: substyle -> fund -> section-order -> $ desc
        def sortkey(x):
            return (x["substyle"], x["fund"], SECTION_ORDER.get(x["section"], 9),
                    -(x["dollar_m"] or 0))
        cur_sub = cur_fund = None
        for x in sorted(frows, key=sortkey):
            if x["substyle"] != cur_sub:
                cur_sub = x["substyle"]; cur_fund = None
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
                cell = ws.cell(row=r, column=1, value=f"▼ {cur_sub}")
                cell.font = SUB; cell.fill = SUBFILL
                r += 1
            company, pct, dollar, change = parse_raw(x["raw_text"], x["ticker"],
                                                     x["pct_value"], x["dollar_m"])
            # also write to flat sheet
            for fi, fv in enumerate([fam, x["substyle"][:40], x["fund"][:34],
                                     SECTION_LABEL.get(x["section"], "?"),
                                     x["ticker"], company,
                                     pct, dollar, change], 1):
                flat.cell(row=fr_idx, column=fi, value=fv).alignment = WRAP
            flat.cell(row=fr_idx, column=5).font = Font(bold=True)
            fr_idx += 1
            ws.cell(row=r, column=1, value=x["substyle"][:40])
            ws.cell(row=r, column=2, value=x["fund"][:34])
            ws.cell(row=r, column=3, value=SECTION_LABEL.get(x["section"], "?"))
            ws.cell(row=r, column=4, value=x["ticker"]).font = Font(bold=True)
            ws.cell(row=r, column=5, value=company or (x["company"] or "")[:50])
            ws.cell(row=r, column=6, value=pct or (f"{x['pct_value']}%" if x["pct_value"] else ""))
            ws.cell(row=r, column=7, value=dollar or (f"${x['dollar_m']:.0f}M" if x["dollar_m"] else ""))
            ws.cell(row=r, column=8, value=change)
            for col in range(1, 9):
                ws.cell(row=r, column=col).alignment = WRAP
            r += 1
        widths = [34, 30, 12, 8, 30, 12, 9, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Flat sheet column widths
    for i, w in enumerate([22, 30, 30, 11, 8, 28, 12, 9, 55], 1):
        flat.column_dimensions[get_column_letter(i)].width = w

    for col, w in (("A", 32), ("B", 12), ("C", 12)):
        idx.column_dimensions[col].width = w
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(wb.sheetnames)-1} family sheets + index)")
    print(f"\nFamilies (by position count):")
    for fam in sorted(by_family, key=lambda f: -len(by_family[f])):
        subs = len({r['substyle'] for r in by_family[fam]})
        funds = len({r['fund'] for r in by_family[fam]})
        print(f"  {fam:<34} {funds:>3} funds, {subs:>2} substyles, {len(by_family[fam]):>4} positions")

if __name__ == "__main__":
    run()
