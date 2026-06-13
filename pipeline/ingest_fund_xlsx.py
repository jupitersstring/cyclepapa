"""Ingest the original fund_activity_last_6mo.xlsx (the foundational artifact) into typed tables.

For each fund tab, parse the 4-section structure into rows:
  Section (1) Highest conviction (recent adds)
  Section (2) >=5% / threshold disclosures
  Section (3) New positions sized large
  Section (4) Existing positions materially increased

Output:
  fund_positions  one row per (fund, ticker, section, snapshot)
  fund_meta       fund-level (group, sources)

Makes the workbook a primary data source, not a parallel one.
"""
import os, re, sqlite3
import openpyxl

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fund_activity_last_6mo.xlsx")

SKIP = {"Asymmetric Summary", "Consensus Buys", "Highest Conviction",
        "Conviction Adds", "Micro-Cap Conviction Adds", "Activist Catalysts",
        "Multi-Fund New Inits", "Cover", "Index", "All Activity"}

SECTION = {
    "highest conviction": 1,
    "threshold": 2, "5% / threshold": 2, ">=5%": 2,
    "new positions": 3,
    "materially increased": 4, "material adds": 4, "existing positions": 4,
}

TICKER_RE = re.compile(r'^([A-Z][A-Z0-9]{0,5}(?:\.[A-Z]{1,3})?)$')
# Junk tokens that pattern-match like tickers — extend as the data exposes more
NOT_TICKERS = {"NEW","ADD","CIK","AUM","RAUM","ADV","SEC","FDA","CEO","CFO","COO","CIO","CMO",
               "EU","US","UK","HK","TX","NY","CA","Q1","Q2","Q3","Q4","FY","YTD","TTM",
               "EBITDA","FCF","EPS","ROIC","ROE","ROA","NAV","ETF","SPAC","IPO","ESG","AI",
               "NYSE","TSX","LSE","ASX","NASDAQ","JV","PT","IV","II","III","DCF","DTC",
               "PDUFA","NDA","IND","MFN","JAN","FEB","MAR","APR","JUN","JUL","AUG","SEP","OCT","NOV","DEC",
               "QOQ","YOY","NOL","CAR","PE","VC","HF","BP","BPS","KPI","TAM","SAM","SOM"}
PCT_RE    = re.compile(r'([\+\-]?\d+(?:\.\d+)?)\s*%')
DOLLAR_RE = re.compile(r'\$([\d,]+(?:\.\d+)?)\s*([MBK]?)', re.I)
DATE_RE   = re.compile(r'(20\d{2})[-/](\d{1,2})[-/](\d{1,2})')

def detect_section(text):
    t = text.lower()
    for key, num in SECTION.items():
        if key in t: return num
    return None

def parse_dollar(s):
    m = DOLLAR_RE.search(s or "")
    if not m: return None
    v = float(m.group(1).replace(",", ""))
    mult = {"B": 1000, "M": 1, "K": 0.001, "": 1.0}.get(m.group(2).upper(), 1)
    return v * mult

def parse_pct(s):
    m = PCT_RE.search(s or "")
    return float(m.group(1)) if m else None

def parse_date(s):
    m = DATE_RE.search(s or "")
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else None

def run():
    conn = sqlite3.connect(DB)
    conn.executescript("""
    DROP TABLE IF EXISTS fund_positions;
    DROP TABLE IF EXISTS fund_meta;
    CREATE TABLE fund_meta (
      fund TEXT PRIMARY KEY, fund_group TEXT, source_block TEXT, total_rows INTEGER);
    CREATE TABLE fund_positions (
      id INTEGER PRIMARY KEY,
      fund TEXT, ticker TEXT, company TEXT, section INTEGER,
      pct_value REAL,     -- % of fund book or % of company
      pct_kind TEXT,      -- 'book' | 'company' | NULL
      dollar_m REAL, change_text TEXT, event_date TEXT, raw_text TEXT,
      asof TEXT NOT NULL);
    CREATE INDEX idx_fp_ticker ON fund_positions(ticker);
    CREATE INDEX idx_fp_fund ON fund_positions(fund);
    """)

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    funds_parsed = 0
    rows_parsed = 0
    snap = "2026-05-30"  # the workbook snapshot date

    for sheet in wb.sheetnames:
        if sheet in SKIP or sheet.startswith("Sheet"):
            continue
        ws = wb[sheet]
        fund = sheet
        section = 0
        group = ""
        srcs = []
        rows_in_tab = 0
        in_sources = False

        for row in ws.iter_rows(values_only=True):
            if not row: continue
            cells = [str(c).strip() if c is not None else "" for c in row]
            text = " | ".join(c for c in cells if c)
            if not text.strip(): continue
            rows_in_tab += 1
            tl = text.lower()

            # Header metadata
            if text.startswith("Group:"):
                group = text[6:].strip(); continue
            if text.startswith("Sources:"):
                in_sources = True; continue
            if in_sources and text.startswith(("http", "CIK", "AUM", "RAUM", "Key Person", "Strategy")):
                srcs.append(text); continue
            if in_sources and "(1)" in text and detect_section(tl):
                in_sources = False

            sec = detect_section(tl)
            if sec:
                section = sec
                continue

            # Try to extract a ticker from the row
            tkr = None; company = None
            for cell in cells:
                if not cell: continue
                first = cell.split()[0] if cell.split() else ""
                m = TICKER_RE.match(first)
                if m and m.group(1) not in NOT_TICKERS and len(m.group(1)) >= 2:
                    tkr = m.group(1)
                    if " " in cell: company = " ".join(cell.split()[1:])[:80]
                    break
            if not tkr:
                continue

            # Best-effort field extraction
            pct = parse_pct(text)
            usd = parse_dollar(text)
            dt  = parse_date(text)
            change = next((c for c in cells if any(kw in c for kw in ("ADD","NEW","+","trim","exit","%"))), "")
            # Heuristic: if % shows up adjacent to "company" or 13D context -> company%; else book%
            pct_kind = "company" if (section == 2 or any(k in tl for k in ("13d","13g","threshold"))) else "book"

            conn.execute("""INSERT INTO fund_positions
                (fund,ticker,company,section,pct_value,pct_kind,dollar_m,change_text,event_date,raw_text,asof)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (fund, tkr, company, section, pct, pct_kind, usd, change[:120], dt, text[:300], snap))
            rows_parsed += 1

        conn.execute("INSERT INTO fund_meta VALUES (?,?,?,?)",
                     (fund, group, "\n".join(srcs)[:2000], rows_in_tab))
        funds_parsed += 1

    conn.commit()

    print(f"Parsed {funds_parsed} fund tabs, {rows_parsed} position rows")
    print(f"\nUnique tickers tracked: {conn.execute('SELECT COUNT(DISTINCT ticker) FROM fund_positions').fetchone()[0]}")
    print(f"By section:")
    for s, n in conn.execute("SELECT section, COUNT(*) FROM fund_positions GROUP BY section ORDER BY section"):
        labels = {1:"Highest conviction", 2:"13D/13G threshold", 3:"New positions", 4:"Material adds"}
        print(f"  {s} {labels.get(s,'?'):<24} {n} rows")
    print(f"\nTop 15 multi-fund consensus tickers (count of funds holding):")
    for tkr, n in conn.execute("""SELECT ticker, COUNT(DISTINCT fund) c FROM fund_positions
                                  WHERE ticker IS NOT NULL
                                  GROUP BY ticker ORDER BY c DESC LIMIT 15"""):
        print(f"  {tkr:<8} {n} funds")

if __name__ == "__main__":
    run()
