"""Build MICRO/SMALL CAP INTENSITY analysis.

Large concentrated positions in micro/small caps are higher signal because:
  1. Fund's position = significant % of free float = cannot exit without impact
  2. Concentration = forced long-term holder = strong conviction signal
  3. Catalyst flow moves price violently in tight float
  4. Fund's % OWNERSHIP of company (not just % of book) is meaningful

Filter: market cap <$2B AND fund ownership >5% of company OR concentration >10% of fund book.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# MICRO/SMALL CAP POSITIONS WITH LARGE CONCENTRATION
# Each: (ticker, mcap, fund, % of company owned, % of fund book, $ size,
#        avg daily volume context, cost basis status, signal note)
# ================================================================

MICROSMALL = [
    # ============================================================
    # ULTRA-MICRO (<$200M cap) WITH MAJOR ANCHOR
    # ============================================================
    {'ticker': 'OPRX', 'mcap': '$93M', 'cap_tier': 'ULTRA-MICRO',
     'fund': 'Insider holders + $10M buyback + Presti (Doximity) board',
     'pct_company': '14.82% insider + buyback active',
     'pct_book': 'N/A (insider)', 'dollar': '$14M insider value',
     'liquidity': 'Sub-$100M cap = institutional minimum = ANY flow moves price violently',
     'cost_basis': 'FLAT — buyback at $4-5; Presti joined board April 2026 at $4.82',
     'signal': 'INSIDER 14.82% + corporate buyback + DOXIMITY EXEC (the digital health peer) joined board at depressed = strongest micro insider signal',
     'asym_rank': 1, 'tailwind': 'NOT (digital pharma EHR shift unpriced)'},

    {'ticker': 'FTLF', 'mcap': '$93M', 'cap_tier': 'ULTRA-MICRO',
     'fund': 'Smoak Capital (residual top-5 position)',
     'pct_company': 'Sub-5% institutional',
     'pct_book': 'N/A', 'dollar': 'Sub-$5M',
     'liquidity': 'Sub-$100M cap = sub-institutional minimum = pure microcap mean-reversion',
     'cost_basis': '-43.9% 6mo drawdown; cleanest Yartseva 7/7',
     'signal': 'Pure Yartseva 7/7 microcap; B/M 0.49 + ROA 7.2% + FCF yield 8.2%; sub-institutional minimum = any institutional flow violently re-rates',
     'asym_rank': 20, 'tailwind': 'NONE'},

    {'ticker': 'BNTC', 'mcap': '$200M', 'cap_tier': 'ULTRA-MICRO',
     'fund': 'Suvretta Capital (13D)',
     'pct_company': '44.1% of COMPANY (LARGEST stake by % in Suvretta portfolio)',
     'pct_book': '~44% of $250M Suvretta book', 'dollar': '~$110M',
     'liquidity': 'Suvretta owns ~half the float; cannot exit without crushing',
     'cost_basis': 'Built at sub-$10 levels',
     'signal': '44.1% of company = activist effectively running the show; LARGEST stake by % in Suvretta entire portfolio',
     'asym_rank': 28, 'tailwind': 'NONE'},

    {'ticker': 'QRHC', 'mcap': '$200M', 'cap_tier': 'ULTRA-MICRO',
     'fund': 'Wynnefield Capital',
     'pct_company': '13.3% (13D w/ board seat)',
     'pct_book': 'Significant book weighting', 'dollar': '~$27M',
     'liquidity': 'Sub-$200M; 13D + board = locked',
     'cost_basis': 'Built sub-$7; Cooperation Agreement w/ standstill until 30d before 2027 nom window',
     'signal': '13.3% + BOARD SEAT + Cooperation Agreement = formal multi-year lock; sub-$200M cap',
     'asym_rank': 30, 'tailwind': 'NONE'},

    {'ticker': 'POSTBPB', 'mcap': '$240M', 'cap_tier': 'ULTRA-MICRO',
     'fund': 'Nierenberg/D3 Family Fund',
     'pct_company': '7%+ (LARGEST single shareholder)',
     'pct_book': '~7% of D3 book ($16.4M)', 'dollar': '$16.4M',
     'liquidity': 'Sub-$300M; D3 largest = effectively activist anchor',
     'cost_basis': '13D/A active March 2025; Form 4 insider buys ongoing',
     'signal': 'Nierenberg LARGEST shareholder of PBPB + 7% of D3 book = micro-cap activist soft pressure',
     'asym_rank': 31, 'tailwind': 'NONE'},

    # ============================================================
    # MICRO/SMALL ($200M-$1B) WITH BIG FUND CONCENTRATION
    # ============================================================
    {'ticker': 'PRLD', 'mcap': '$340M', 'cap_tier': 'MICRO',
     'fund': 'Baker Bros + RA Capital (DUAL TOP-TIER)',
     'pct_company': '25.5% combined (Baker 15.5% + RA Cap 9.99%)',
     'pct_book': 'Baker small, RA small (both have large books) but 25% of COMPANY',
     'dollar': '$13M combined dollar but anchor at SAME $4.44',
     'liquidity': 'Sub-$340M; quarter of company owned by two top biotech crossover funds',
     'cost_basis': 'BOTH at $4.44 April 2026 $90M financing = current $4.26 BELOW entry',
     'signal': 'TWO top biotech crossover funds = 25.5% of company at SAME $4.44 = locked + below current = pure follow signal',
     'asym_rank': 2, 'tailwind': 'NOT (TPD platform + JAK2 MPN)'},

    {'ticker': 'CARS', 'mcap': '$560M', 'cap_tier': 'SMALL',
     'fund': 'Insider buying + Yartseva 7/7',
     'pct_company': 'Insider build at depressed',
     'pct_book': 'N/A', 'dollar': 'Insider',
     'liquidity': '$560M cap + $90M buyback (16% cap) = significant float compression',
     'cost_basis': 'Y7/7 still depressed; Q1 EPS $0.45 vs $0.13 est',
     'signal': '$90M buyback = 16% of cap = mechanical float compression + Yartseva 7/7 still depressed',
     'asym_rank': 12, 'tailwind': 'NONE'},

    {'ticker': 'PVLA', 'mcap': '$1.56B', 'cap_tier': 'SMALL',
     'fund': 'BVF + Suvretta (DUAL)',
     'pct_company': '8.5% (BVF) — large single-fund ownership',
     'pct_book': 'BVF 3.8% / Suvretta 3.67% (+29.3% Q1)', 'dollar': '$264M combined',
     'liquidity': '$1.56B small-cap pharma; BVF 8.5% of company',
     'cost_basis': 'Suvretta +29.3% Q1 ADD at $100-140; BVF long anchor',
     'signal': '$264M dual-anchor (BVF + Suvretta) at sub-$2B cap = strongest small-cap biotech signal',
     'asym_rank': 17, 'tailwind': 'NOT (rare disease + QTORIN platform)'},

    {'ticker': 'NRP', 'mcap': '$1.35B', 'cap_tier': 'SMALL',
     'fund': 'Saber + Right Tail + Greystone + Berkowitz + family insider',
     'pct_company': 'Family 31.75% insider; Saber 17.75% of Saber book; Right Tail NEW Q1',
     'pct_book': 'Saber 17.75% / Right Tail concentrated', 'dollar': '~$75M multi-fund',
     'liquidity': '$1.35B mid-small; 31.75% family + multi-fund anchor = limited float',
     'cost_basis': 'Right Tail NEW Q1 at ~$105 = current; family aligned forever',
     'signal': '5-fund concentration + 31.75% family insider on a $1.35B royalty trust = strongest anchor stack',
     'asym_rank': 10, 'tailwind': 'NOT (met coal + soda ash for solar/EV)'},

    {'ticker': 'CRTO', 'mcap': '$866M', 'cap_tier': 'MICRO',
     'fund': 'D3/Nierenberg + Petrus Advisers activist',
     'pct_company': 'D3 9.37% of book at small cap = meaningful ownership',
     'pct_book': 'D3 9.37% / +19% Q1', 'dollar': '$5M D3 + Petrus est $20M',
     'liquidity': '$866M micro; $200M buyback = 22% cap mechanical compression',
     'cost_basis': 'D3 +19% Q1 add at $11.50-14 = current FLAT',
     'signal': 'D3 activist book at micro + Petrus activist + 22% buyback = three forces at current entry',
     'asym_rank': 9, 'tailwind': 'NOT (retail media $200B TAM by 2028)'},

    {'ticker': 'INMD', 'mcap': '$869M', 'cap_tier': 'MICRO',
     'fund': 'Steel Partners (activist) + Mizrahy founder',
     'pct_company': 'Steel ~10% of co; founder large',
     'pct_book': 'Steel 25.5% of $338M book', 'dollar': '$86M + founder $10.7M Feb 2026',
     'liquidity': '$869M micro; Steel cannot easily exit 10%+ stake',
     'cost_basis': 'Steel doubled position Q1; founder bought $10.7M Feb 2026 at $13.71',
     'signal': 'Activist + founder both at current FLAT; cash $8.47/sh = 62% floor = bounded downside on a micro',
     'asym_rank': 3, 'tailwind': 'NONE (aesthetics cyclical)'},

    {'ticker': 'MGNI', 'mcap': '$1.89B', 'cap_tier': 'SMALL',
     'fund': 'Nine Ten Capital (Bares disciple)',
     'pct_company': '~2.3% (3.48M sh)', 'pct_book': '13.1% of $315M Nine Ten book',
     'dollar': '$41.4M', 'liquidity': '$1.89B; Bares-style concentration on independent CTV SSP',
     'cost_basis': 'Built at depressed CTV SSP prices',
     'signal': 'Nine Ten 13.1% concentration = Bares-style high-conviction independent CTV SSP winner',
     'asym_rank': 25, 'tailwind': 'NOT (CTV ad spend doubling 2024-2028)'},

    {'ticker': 'VITL', 'mcap': '$1.5B', 'cap_tier': 'SMALL',
     'fund': 'No major fund yet concentrated',
     'pct_company': '<5% institutional', 'pct_book': 'N/A',
     'dollar': 'Sub-$10M concentration', 'liquidity': '$1.5B with broad institutional float',
     'cost_basis': '-68.76% drawdown but no anchor has built',
     'signal': 'CONTRARIAN — no fund has piled in yet; entry preserved BUT no anchor signal',
     'asym_rank': 19, 'tailwind': 'PARTIAL (pasture-raised partial recognition)'},

    {'ticker': 'MTY.TO', 'mcap': 'C$887M (~$640M USD)', 'cap_tier': 'MICRO',
     'fund': 'CEO Lefebvre (insider)',
     'pct_company': 'CEO open-market buy at C$30.54',
     'pct_book': 'CEO personal', 'dollar': '~$1M+ CEO buy',
     'liquidity': 'C$887M; CEO buying = signal at trough multiple',
     'cost_basis': 'CEO at C$30.54 vs C$40.41 = in money but small position',
     'signal': 'CEO open-market buy at trough on micro Canadian QSR franchisor at BOOK valuation',
     'asym_rank': 21, 'tailwind': 'PARTIAL'},

    {'ticker': 'ACMR', 'mcap': '~$400M', 'cap_tier': 'MICRO',
     'fund': 'Kerrisdale + Steamboat (JOINT LETTER)',
     'pct_company': 'Joint position',
     'pct_book': 'Kerrisdale 3.36% of book', 'dollar': '$8.8M Kerrisdale + Steamboat',
     'liquidity': '$400M micro semiconductor equipment',
     'cost_basis': 'Joint letter Nov 2025; HK listing $100M offer pursuing',
     'signal': 'Two-fund joint letter for HK listing value unlock = forced corporate action on micro',
     'asym_rank': 'N/A', 'tailwind': 'PARTIAL'},
]

# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'MICRO+SMALL CAP INTENSITY' in wb.sheetnames:
    del wb['MICRO+SMALL CAP INTENSITY']

ws = wb.create_sheet('MICRO+SMALL CAP INTENSITY', 4)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ULTRA_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')  # blue best
MICRO_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')   # green
SMALL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')    # yellow
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:J1')
ws['A1'] = 'MICRO+SMALL CAP INTENSITY — Large Positions in Sub-$2B Names'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:J2')
ws['A2'] = ('Why this matters: Large concentrated positions in micro/small caps cannot easily exit '
            '= forced long-term holder = strong conviction signal. Fund % of COMPANY (vs % of book) '
            'is meaningful — at sub-$1B caps, even modest book weightings = high ownership share.')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

# Cap tier legend
ws.merge_cells('A4:C4')
ws['A4'] = 'CAP TIER LEGEND'
ws['A4'].font = HEADER; ws['A4'].fill = HEADER_FILL
ws['A5'] = 'ULTRA-MICRO'; ws['A5'].fill = ULTRA_FILL; ws['A5'].font = Font(bold=True)
ws['B5'] = '<$300M cap — sub-institutional minimum = any flow violently moves price'
ws['A6'] = 'MICRO'; ws['A6'].fill = MICRO_FILL; ws['A6'].font = Font(bold=True)
ws['B6'] = '$300M-$1B cap — institutional but tight float'
ws['A7'] = 'SMALL'; ws['A7'].fill = SMALL_FILL; ws['A7'].font = Font(bold=True)
ws['B7'] = '$1B-$2B cap — broader float but concentration still meaningful'

# Headers
headers = ['Cap Tier', 'Ticker', 'Mcap', 'Asym Rank', 'Fund', '% Company Owned', '% Fund Book', 'Dollar Size', 'Cost Basis Status', 'Why Signal Is Stronger Than Mid/Large Cap']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=9, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

# Sort: ultra-micro first, then by asym rank
def sort_key(x):
    tier_order = {'ULTRA-MICRO': 0, 'MICRO': 1, 'SMALL': 2}
    rank = x['asym_rank'] if isinstance(x['asym_rank'], int) else 99
    return (tier_order.get(x['cap_tier'], 3), rank)

MICROSMALL.sort(key=sort_key)

for i, n in enumerate(MICROSMALL, 10):
    tier = n['cap_tier']
    fill = {'ULTRA-MICRO': ULTRA_FILL, 'MICRO': MICRO_FILL, 'SMALL': SMALL_FILL}.get(tier)

    ws.cell(row=i, column=1, value=tier).fill = fill
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=i, column=3, value=n['mcap'])
    ws.cell(row=i, column=4, value=n['asym_rank'])
    ws.cell(row=i, column=5, value=n['fund'])
    ws.cell(row=i, column=6, value=n['pct_company'])
    ws.cell(row=i, column=7, value=n['pct_book'])
    ws.cell(row=i, column=8, value=n['dollar'])
    ws.cell(row=i, column=9, value=n['cost_basis'])
    ws.cell(row=i, column=10, value=n['signal'])

    for col in range(1, 11):
        ws.cell(row=i, column=col).alignment = LEFT_WRAP

    ws.row_dimensions[i].height = 65

widths = [12, 9, 14, 9, 35, 30, 22, 22, 35, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)
print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print("\nMICRO/SMALL CAP INTENSITY (sub-$2B with large fund concentration):")
print("=" * 80)
print(f"\n{'ULTRA-MICRO (<$300M cap):'}")
for n in MICROSMALL:
    if n['cap_tier'] == 'ULTRA-MICRO':
        print(f"  {n['ticker']:<10} {n['mcap']:<10} | {n['fund'][:35]:<35} | {n['pct_company'][:40]}")
print(f"\nMICRO ($300M-$1B cap):")
for n in MICROSMALL:
    if n['cap_tier'] == 'MICRO':
        print(f"  {n['ticker']:<10} {n['mcap']:<10} | {n['fund'][:35]:<35} | {n['pct_company'][:40]}")
print(f"\nSMALL ($1B-$2B cap):")
for n in MICROSMALL:
    if n['cap_tier'] == 'SMALL':
        print(f"  {n['ticker']:<10} {n['mcap']:<10} | {n['fund'][:35]:<35} | {n['pct_company'][:40]}")
