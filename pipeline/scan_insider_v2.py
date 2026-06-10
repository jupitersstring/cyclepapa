"""V2: Stricter scan for INSIDER BUYS + AGGRESSIVE FUND ADDS.

Improvements over V1:
  - Stricter ticker validation (must have stock-like context)
  - Filter "+X%" to only X >= 20% (aggressive adds, not minor adjustments)
  - Require specific insider phrases (Form 4, "open-market buy", "$X.XM buy")
  - Group by ticker, count tab mentions
"""
import openpyxl
import re
from collections import defaultdict

wb_path = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=True)

# Patterns
# Strong insider patterns (no false positives)
INSIDER_STRONG = [
    r'Form\s*4\s*(?:buys?|purchases?|filing)',
    r'open[\s-]market\s+(?:buy|purchase|buying)',
    r'(?:CEO|CFO|COO|founder|chairman|director|insider)\s+(?:bought|buying|purchase)',
    r'(?:CEO|CFO|COO|founder|chairman|director)\s+\+?\$\d',
    r'insider\s+(?:buys?|bought)\s+\$',
    r'\$\d+\.?\d*M\s+(?:CEO|CFO|founder|director|insider|chairman)\s+buy',
    r'(?:CEO|CFO|founder)\s+(?:open[\s-]market|bought|purchased|bought\s+\d)',
]

# Aggressive add patterns (only count meaningful adds 20%+)
AGGRESSIVE_ADD = [
    r'\+(\d+)%\s+(?:Q1|sh|ADD|add|shares)',
    r'\+(\d+)%\s+(?:Q[1-4])\s+2026',
    r'(doubled|tripled|quadrupled|quintupled)',
    r'\+(\d+)%\s+(?:position|stake)',
    r'(\d+)x\s+(?:add|increase)',
]

# Strict ticker validation — must look like a real ticker
TICKER_RE = re.compile(r'(?:^|\s|\(|\$)([A-Z][A-Z0-9]{1,5}(?:\.[A-Z]{1,3})?)\b')

# Block list for known false positives
NOT_TICKERS = {
    'CEO', 'CFO', 'COO', 'CTO', 'CIO', 'CMO', 'CCO',
    'AUM', 'RAUM', 'ADD', 'NEW', 'BUY', 'SELL', 'AND', 'THE',
    'FOR', 'WITH', 'TO', 'OF', 'AT', 'BY', 'IN', 'ON', 'OR',
    'IS', 'AS', 'IT', 'BE', 'HAS', 'WAS', 'ARE',
    'CIK', 'CRD', 'ADV', 'AGM', 'SEC', 'FDA', 'PDUFA', 'IRA',
    'CHIPS', 'AI', 'LLC', 'INC', 'CO', 'LTD', 'PLC', 'LP', 'GP',
    'Q1', 'Q2', 'Q3', 'Q4', 'FY', 'FYE', 'YTD', 'TTM',
    'IPO', 'SPAC', 'EBITDA', 'FCF', 'EPS', 'P', 'B', 'S',
    'NDA', 'IND', 'EU', 'US', 'UK', 'HK', 'JV', 'PT',
    'NYSE', 'NASDAQ', 'OTC', 'TSX', 'LSE', 'ASX', 'JPX',
    'PE', 'VC', 'HF', 'PB', 'MD', 'MM', 'BN', 'TN',
    'IRR', 'NPV', 'ROIC', 'ROE', 'ROA',
    'MFN', 'ETF', 'ESG', 'GP', 'LP',
    'HYPER', 'FILING', 'FUND', 'MULTI', 'INIT', 'CONVICTION', 'ACTIVIST',
    'CONSENSUS', 'HIGH', 'LOW', 'BUYS', 'POSITION', 'STAKE',
    'SH', 'SHARE', 'SHARES', 'NEW', 'SAME', 'CUT', 'MAX',
    'MULTIBAGGER', 'COMPANY', 'STOCK', 'CONVICTION', 'PRICE',
    'DOLLAR', 'MILLION', 'BILLION', 'AMOUNT', 'PERIOD',
    'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC',
    'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN',
    'AM', 'PM', 'BC', 'AD',
    'RM', 'RB', 'AR', 'BO',
    'YES', 'NO', 'OK', 'NA',
    'INC', 'LTD', 'GP', 'LP', 'PE', 'VC',
    'EV', 'MCAP', 'NAV', 'IRR',
    'PCT', 'PCS',
}

NUMERIC = re.compile(r'^\d+$')

SKIP_SHEETS = {'Cover', 'Index', 'All Activity', 'multibagger_systematic_synthesis',
               'asymmetric_ideas_2026Q2', 'phelps_mayer_lynch_screen',
               'yartseva_screen_corrected', 'positioning_methodology',
               'most_asymmetric_deep_dive', 'fund_coverage_gap',
               'qualitative_framework_supplement', 'final_most_asymmetric_revised',
               'not_yet_rerated_asymmetric', 'coverage_status'}


def extract_tickers(text):
    """Strict ticker extraction."""
    candidates = TICKER_RE.findall(text)
    return [t for t in candidates if t not in NOT_TICKERS and not NUMERIC.match(t) and len(t) >= 2 and len(t) <= 7]


# Track per-ticker signals
ticker_funds_insider = defaultdict(set)  # ticker -> set of fund tabs with insider buy
ticker_funds_add = defaultdict(set)      # ticker -> set of fund tabs with aggressive add
ticker_add_size = defaultdict(list)      # ticker -> list of (fund, %, snippet)
ticker_insider_detail = defaultdict(list)  # ticker -> list of (fund, snippet)
ticker_mentions = defaultdict(int)
ticker_max_add = defaultdict(int)


for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS or sheet_name.startswith('Sheet'):
        continue
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
        rows.append(' '.join(str(c) if c is not None else '' for c in row))

    for row_idx, row_text in enumerate(rows, 1):
        if not row_text.strip():
            continue

        tickers = extract_tickers(row_text)
        for t in tickers:
            ticker_mentions[t] += 1

        # Strong insider buy detection
        for pat in INSIDER_STRONG:
            if re.search(pat, row_text, re.IGNORECASE):
                for t in tickers:
                    ticker_funds_insider[t].add(sheet_name)
                    ticker_insider_detail[t].append((sheet_name, row_text[:200].strip()))
                break

        # Aggressive add detection (require 20%+)
        for pat in AGGRESSIVE_ADD:
            m = re.search(pat, row_text)
            if m:
                # Extract numeric % if present
                groups = m.groups()
                add_pct = 0
                add_text = m.group(0)
                if groups and groups[0] and groups[0].isdigit():
                    add_pct = int(groups[0])
                elif 'doubled' in add_text.lower():
                    add_pct = 100
                elif 'tripled' in add_text.lower():
                    add_pct = 200
                elif 'quadrupled' in add_text.lower():
                    add_pct = 300
                elif 'quintupled' in add_text.lower():
                    add_pct = 400

                if add_pct >= 20:
                    for t in tickers:
                        ticker_funds_add[t].add(sheet_name)
                        ticker_add_size[t].append((sheet_name, add_pct, row_text[:200].strip()))
                        if add_pct > ticker_max_add[t]:
                            ticker_max_add[t] = add_pct
                break


# Intersection: insider buys + fund adds
intersection_tickers = set(ticker_funds_insider.keys()) & set(ticker_funds_add.keys())

def score(t):
    insider_count = len(ticker_funds_insider[t])
    add_count = len(ticker_funds_add[t])
    max_add = ticker_max_add[t]
    return insider_count * 3 + add_count * 2 + min(max_add, 500) / 50

sorted_tickers = sorted(intersection_tickers, key=score, reverse=True)


print(f"NAMES WITH INSIDER BUYS + AGGRESSIVE FUND ADDS (>=20%)")
print("=" * 80)
print(f"{'Ticker':<10} {'Insider#':<10} {'Add#':<6} {'MaxAdd':<8} {'Score':<7} Total Mentions")
print('-' * 80)

# Filter to those with score >= 5 (real consensus, not noise)
filtered = [t for t in sorted_tickers if score(t) >= 5]
for t in filtered[:50]:
    print(f"{t:<10} {len(ticker_funds_insider[t]):<10} {len(ticker_funds_add[t]):<6} +{ticker_max_add[t]:<7}% {score(t):<7.1f} {ticker_mentions[t]}")

print(f"\nTOP 25 DETAIL (with snippets):")
print("=" * 80)
for t in filtered[:25]:
    print(f"\n{t}  (insider in {len(ticker_funds_insider[t])} tabs, add in {len(ticker_funds_add[t])} tabs, max add +{ticker_max_add[t]}%)")
    seen = set()
    for fund, snippet in ticker_insider_detail[t][:3]:
        if fund not in seen:
            seen.add(fund)
            print(f"  INSIDER [{fund}]: {snippet[:140]}")
    seen = set()
    sorted_adds = sorted(ticker_add_size[t], key=lambda x: -x[1])
    for fund, pct, snippet in sorted_adds[:3]:
        if fund not in seen:
            seen.add(fund)
            print(f"  ADD +{pct}% [{fund}]: {snippet[:140]}")
