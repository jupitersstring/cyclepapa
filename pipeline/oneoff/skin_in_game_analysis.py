"""Build MAX SKIN IN GAME + NOT DEEPLY IN MONEY analysis.

Best follow-the-smart-money signal:
  (a) Position concentration = max % of fund book (skin in game)
  (b) Recent add or anchor at current price band (entry economics intact)
  (c) Underwater or flat vs cost basis = fund has same upside as new entrant

Filter OUT positions where smart money is deeply in the money (already captured
multi-bagger move — new entrant gets diminished asymmetry).
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ================================================================
# SKIN-IN-GAME DATA
# Format: (ticker, fund, concentration_pct_or_size, cost_basis_or_entry,
#          current_price, in_money_status, signal_strength, source/note)
# ================================================================

SKIN_IN_GAME = [
    # ============================================================
    # TIER 1 — MAX SKIN + UNDERWATER (highest signal)
    # Smart money is at MAX conviction AND down/flat from cost = pure follow opportunity
    # ============================================================

    {'ticker': 'PRLD', 'fund': 'Baker Bros + RA Capital (DUAL ANCHOR)',
     'concentration': 'Baker 15.5% 13D/A + RA Capital 9.99% 13G',
     'cost_basis': '$4.44 (BOTH anchored same April 2026 $90M financing)',
     'current_price': '$4.26', 'pnl_status': '-4% UNDERWATER',
     'tier': 1, 'signal': 'EXTREME — TWO top biotech crossover funds anchored at SAME price 4% above current. Buy today = buy BELOW their entry. They cannot be selling here.',
     'asym_rank': 2},

    {'ticker': 'HHH', 'fund': 'Ackman/Pershing Square',
     'concentration': '47% of HHH via $900M common + $1B preferred',
     'cost_basis': '$100/sh (publicly disclosed acquisition price)',
     'current_price': '$63.77', 'pnl_status': '-36% UNDERWATER',
     'tier': 1, 'signal': 'EXTREME — Ackman publicly anchored at $100 with multi-billion commitment. $63 = sponsor PUT floor; he cannot mark down or sell. Any holdco progress = re-rate toward his cost.',
     'asym_rank': 4},

    {'ticker': 'KYMR', 'fund': 'BVF (MAX CONVICTION) + Baker Bros',
     'concentration': 'BVF #1 position 14.6% of $3.14B book ($458M) + Baker 4.1% / $721M',
     'cost_basis': '$86 (Dec 2025 follow-on, both participated)',
     'current_price': '$80', 'pnl_status': '-7% UNDERWATER',
     'tier': 1, 'signal': 'EXTREME — BVF largest holding with 14.6% concentration; followed on at $86. Current $80 means TWO biotech specialists below entry on highest-conviction names.',
     'asym_rank': 18},

    {'ticker': 'KBR', 'fund': 'D3 Family Fund (Nierenberg)',
     'concentration': '+425% Q1 2026 = LARGEST SINGLE ADD IN ENTIRE UNIVERSE',
     'cost_basis': '~$30 (Q1 2026 build)',
     'current_price': '$30', 'pnl_status': 'FLAT',
     'tier': 1, 'signal': 'EXTREME — D3 quintupled position in Q1 at current levels. Director Form 4 buys May 2026 at $30-31 confirm sponsor put. New entrant at $30 = same cost basis as the largest-add fund move of the quarter.',
     'asym_rank': 13},

    {'ticker': 'INMD', 'fund': 'Steel Partners + Mizrahy (founder)',
     'concentration': 'Steel 25.5% of $338M book / +155% Q1 2026 (DOUBLED). Founder bought 800k sh ($10.7M) Feb 2026',
     'cost_basis': '$13.50-14 (Steel Q1 build + founder Feb buy)',
     'current_price': '$13.71', 'pnl_status': 'FLAT',
     'tier': 1, 'signal': 'EXTREME — Founder bought $10.7M open-market AT current price. Steel doubled position to 25% of book at same level. Cash $8.47/sh = $11+ floor regardless. New entrant gets founder cost basis.',
     'asym_rank': 3},

    {'ticker': 'NRP', 'fund': 'Saber Capital (Huber) + Right Tail NEW Q1',
     'concentration': 'Saber 17.75% of book + Robertson family 31.75% insider + Right Tail NEW Q1 2026',
     'cost_basis': 'Right Tail NEW Q1 at ~$105 (just entered); Saber long-term',
     'current_price': '$105.81', 'pnl_status': 'FLAT (vs recent Right Tail entry)',
     'tier': 1, 'signal': 'STRONG — Right Tail concentrated NEW position Q1 2026 at current; Saber 17.75% concentration; insider family 31.75% — alignment intact at current. Nov 2026 distribution increase ahead.',
     'asym_rank': 10},

    {'ticker': 'OPRX', 'fund': 'Insider holders + Doximity exec board addition',
     'concentration': 'Insider 14.82% + $10M buyback through March 2027',
     'cost_basis': 'Buyback executing at $4-5 range; Presti (Doximity) joined board April 2026',
     'current_price': '$4.82', 'pnl_status': 'FLAT vs buyback execution',
     'tier': 1, 'signal': 'STRONG — Company buying back stock at current levels; high-quality industry exec (Doximity, the digital health peer) JOINED BOARD April 2026 at depressed prices = strong insider conviction.',
     'asym_rank': 1},

    # ============================================================
    # TIER 2 — MAX SKIN + MODEST IN MONEY
    # Smart money built at slightly lower; current price still close to entry
    # ============================================================

    {'ticker': 'AAP', 'fund': 'H Partners (Andrew Levander)',
     'concentration': '46.3% of $154M book / $71.2M / +50% sh Q1 2026 (1.35M new shares)',
     'cost_basis': '~$40-55 range (Q1 add at deeper discount)',
     'current_price': '$54.42', 'pnl_status': 'MODESTLY IN MONEY (~5-15%)',
     'tier': 2, 'signal': 'STRONG — H Partners HALF THEIR BOOK in one name + still ADDING Q1 (1.35M sh new). 46% concentration is extreme. Q2 margin proof Aug = continued add likely.',
     'asym_rank': 15},

    {'ticker': 'FUN', 'fund': 'H Partners + Jana Partners',
     'concentration': 'H Partners 53.7% / $82.5M / 5.7% of class. Jana 9% activist push',
     'cost_basis': 'H Partners built ~$22-30 range',
     'current_price': '$25', 'pnl_status': 'FLAT to modest',
     'tier': 2, 'signal': 'STRONG — H Partners 53.7% concentration = MAX conviction of any 13F filer in universe. Jana 9% pushing OUTRIGHT SALE. Two-front activist. Jaffer just joined board May 26.',
     'asym_rank': 14},

    {'ticker': 'CRTO', 'fund': "D3/Nierenberg + Petrus Advisers",
     'concentration': "D3 9.37% / +19% Q1 2026. Petrus activist position",
     'cost_basis': '$11.50-14 (D3 +19% Q1 at trough)',
     'current_price': '$13.50', 'pnl_status': 'FLAT (D3 added at current band)',
     'tier': 2, 'signal': 'STRONG — D3 added 19% to existing 9.37% in Q1 at this price band. $200M buyback (22% cap) executing. Petrus activist + Q3 anniversary lap.',
     'asym_rank': 9},

    {'ticker': 'PVLA', 'fund': 'BVF + Suvretta (DUAL)',
     'concentration': 'BVF 3.8% / $120M (8.5% of company) + Suvretta 3.67% / $144M (+29.3% Q1)',
     'cost_basis': 'Suvretta added 29% Q1 at $100-140 range',
     'current_price': '$130', 'pnl_status': 'FLAT to slightly in money',
     'tier': 2, 'signal': 'STRONG — Suvretta ADDED 29% Q1 post-Phase 3 positive. BVF + Suvretta both top-10 holders. Cash $25/sh = 19% floor. NDA/PDUFA staircase ahead.',
     'asym_rank': 17},

    {'ticker': 'MRP', 'fund': 'Lennar (parent spinoff sponsor)',
     'concentration': 'Lennar 80%+ owner (cost basis at spin = current)',
     'cost_basis': 'Spin cost basis = current price (paper position at spin date)',
     'current_price': '$25', 'pnl_status': 'FLAT (Lennar holds at cost)',
     'tier': 2, 'signal': 'STRONG — Lennar 80% ownership = patient capital lock-in. Lennar cannot exit easily (anti-stuffing). Deepest B/M 1.274 in cohort = land bank discount.',
     'asym_rank': 7},

    # ============================================================
    # TIER 3 — STRONG CONCENTRATION + RECENT ADD AT CURRENT
    # ============================================================

    {'ticker': 'FRPT', 'fund': 'Marlowe Partners',
     'concentration': '39.8% of book = HIGHEST single-fund concentration in universe',
     'cost_basis': 'Built at -40% drawdown levels',
     'current_price': '$60', 'pnl_status': 'FLAT vs cost basis',
     'tier': 3, 'signal': 'STRONG — Single-fund concentration 39.8% = bet-the-fund conviction. Manufacturing capacity online + margin inflection.',
     'asym_rank': 26},

    {'ticker': 'MGNI', 'fund': 'Nine Ten Capital (Bares)',
     'concentration': '13.1% of book / 3.48M sh / $41.4M',
     'cost_basis': 'Built at depressed CTV SSP prices',
     'current_price': '$13', 'pnl_status': 'FLAT to modest',
     'tier': 3, 'signal': 'STRONG — Bares-trained Nine Ten 13.1% concentration. Independent CTV SSP winner-take-most thesis.',
     'asym_rank': 25},

    {'ticker': 'BNTC', 'fund': 'Suvretta Capital',
     'concentration': '44.1% 13D = SUVRETTA LARGEST STAKE BY %',
     'cost_basis': 'Built at sub-$10 levels',
     'current_price': '$8', 'pnl_status': 'Likely close to cost or modest underwater',
     'tier': 3, 'signal': 'STRONG — Suvretta\'s LARGEST single position by % ownership. 13D filing = activist intent. Gene therapy platform.',
     'asym_rank': 28},

    {'ticker': 'ACHC', 'fund': 'Greenlight Capital (David Einhorn)',
     'concentration': 'Greenlight 4.1M shares + Sohn 5/12/2026 pitch',
     'cost_basis': 'Built at $30-40 range during DOJ overhang',
     'current_price': '$33', 'pnl_status': 'FLAT (Einhorn near cost)',
     'tier': 3, 'signal': 'STRONG — Einhorn surfaced thesis publicly at Sohn 5/12; -70% drawdown + 29% short float = squeeze fuel on DOJ resolution.',
     'asym_rank': 16},

    {'ticker': 'MTY.TO', 'fund': 'CEO Lefebvre (Eric Lefebvre)',
     'concentration': 'CEO open-market buy at C$30.54 (+33% increase to personal stake)',
     'cost_basis': 'C$30.54 (CEO buy)',
     'current_price': 'C$40.41', 'pnl_status': 'CEO +32% in money but at COST = current band signal',
     'tier': 3, 'signal': 'MODERATE — CEO buying $1M+ at depressed = sponsor put. At-book valuation + 16.7% FCF.',
     'asym_rank': 21},

    {'ticker': 'BZU.IM', 'fund': 'Kerrisdale (Sahm Adrangi) — public thesis',
     'concentration': 'Kerrisdale Oct 2025 published long thesis with €85 PT',
     'cost_basis': 'Thesis published at €49 (current)',
     'current_price': '€49', 'pnl_status': 'FLAT (published at current)',
     'tier': 3, 'signal': 'MODERATE — Public research firm anchored thesis at current. €85 PT (+73%). Italian cement + AI data-center demand.',
     'asym_rank': 22},

    # ============================================================
    # MEGA-CAP HIGH CONCENTRATION (in money — REFERENCE only)
    # ============================================================

    {'ticker': 'CVNA', 'fund': 'CAS Investment Partners',
     'concentration': '81.7% of CAS book = highest single-name concentration in universe',
     'cost_basis': '~$3-30 (legendary recovery position)',
     'current_price': '$71', 'pnl_status': '20-25x IN MONEY — REFERENCE ONLY',
     'tier': 'REF', 'signal': 'REFERENCE — CAS recovered CVNA from $3 to $71 = 40x+. No longer asymmetric.',
     'asym_rank': 'N/A'},

    {'ticker': 'GE', 'fund': 'TCI Fund Management (Chris Hohn)',
     'concentration': '29.85% of $45B book = largest single position',
     'cost_basis': 'Held since 2015',
     'current_price': '$209', 'pnl_status': 'DEEPLY IN MONEY — REFERENCE ONLY',
     'tier': 'REF', 'signal': 'REFERENCE — TCI deeply in money post-GE Aerospace breakup.',
     'asym_rank': 'N/A'},

    {'ticker': 'V', 'fund': 'TCI Fund Management',
     'concentration': '20.39% of $45B book',
     'cost_basis': 'Held long-term',
     'current_price': '$352', 'pnl_status': 'DEEPLY IN MONEY — REFERENCE ONLY',
     'tier': 'REF', 'signal': 'REFERENCE',
     'asym_rank': 'N/A'},

    {'ticker': 'CELC', 'fund': 'Baker Bros',
     'concentration': '19.99% ACTIVIST 13D = $904M / 5.48% of $17.4B book',
     'cost_basis': 'Built pre-data at $50-100',
     'current_price': '$165', 'pnl_status': 'DEEPLY IN MONEY post-May 1 — REFERENCE ONLY',
     'tier': 'REF', 'signal': 'REFERENCE — Baker captured the move; entry-today is post-data.',
     'asym_rank': 'N/A (re-rated)'},

    {'ticker': 'ASND', 'fund': 'RA Capital',
     'concentration': '24.9% of $9.44B book = $2.35B = LARGEST biotech crossover position in universe',
     'cost_basis': 'Held since 2015',
     'current_price': '$247', 'pnl_status': 'DEEPLY IN MONEY — REFERENCE ONLY',
     'tier': 'REF', 'signal': 'REFERENCE — RA Capital largest position but multiple captured.',
     'asym_rank': 'N/A (re-rated)'},
]

# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

# Remove existing if present
if 'MAX Skin + Recent Entry' in wb.sheetnames:
    del wb['MAX Skin + Recent Entry']

# Insert new sheet at position 2 (after Index)
ws = wb.create_sheet('MAX Skin + Recent Entry', 2)

# Styles
HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
TIER1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')  # blue best
TIER2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # green
TIER3_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # yellow
REF_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')    # grey
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:I1')
ws['A1'] = 'MAX SKIN IN GAME + NOT DEEPLY IN MONEY — Smart Money Conviction × Entry Economics'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = CENTER

ws.merge_cells('A2:I2')
ws['A2'] = 'Filter: highest position concentration (% of fund book) AND fund still at or near cost basis (= entry economics intact for new investor). Excludes positions deeply in money where fund already captured multi-bagger move.'
ws['A2'].font = Font(italic=True)
ws['A2'].alignment = LEFT_WRAP

ws.merge_cells('A4:C4')
ws['A4'] = 'TIER LEGEND'
ws['A4'].font = HEADER; ws['A4'].fill = HEADER_FILL
ws['A5'] = 'TIER 1'; ws['A5'].fill = TIER1_FILL; ws['A5'].font = Font(bold=True)
ws['B5'] = 'MAX skin + UNDERWATER (fund at MAX conviction + below cost = pure follow signal)'
ws['A6'] = 'TIER 2'; ws['A6'].fill = TIER2_FILL; ws['A6'].font = Font(bold=True)
ws['B6'] = 'MAX skin + flat to modest in money (Q1 ADDs at current band)'
ws['A7'] = 'TIER 3'; ws['A7'].fill = TIER3_FILL; ws['A7'].font = Font(bold=True)
ws['B7'] = 'Strong concentration + recent entry near current'
ws['A8'] = 'REF'; ws['A8'].fill = REF_FILL; ws['A8'].font = Font(bold=True)
ws['B8'] = 'High concentration but DEEPLY in money (reference only)'

# Headers
headers = ['Tier', 'Ticker', 'Asym Rank', 'Fund', 'Concentration / Skin', 'Cost Basis / Entry', 'Current', 'P&L Status', 'Signal Quality']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=10, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

# Data rows
for i, n in enumerate(SKIN_IN_GAME, 11):
    tier = n['tier']
    if tier == 1: fill = TIER1_FILL
    elif tier == 2: fill = TIER2_FILL
    elif tier == 3: fill = TIER3_FILL
    else: fill = REF_FILL

    ws.cell(row=i, column=1, value=str(tier)).fill = fill
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=i, column=3, value=n['asym_rank'])
    ws.cell(row=i, column=4, value=n['fund'])
    ws.cell(row=i, column=5, value=n['concentration'])
    ws.cell(row=i, column=6, value=n['cost_basis'])
    ws.cell(row=i, column=7, value=n['current_price'])
    ws.cell(row=i, column=8, value=n['pnl_status'])
    ws.cell(row=i, column=9, value=n['signal'])
    for col in range(1, 10):
        ws.cell(row=i, column=col).alignment = LEFT_WRAP

widths = [8, 10, 9, 30, 40, 30, 12, 28, 55]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)
print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
print(f"\nTIER 1 (MAX SKIN + UNDERWATER) — strongest follow-the-leader signals:")
for n in SKIN_IN_GAME:
    if n['tier'] == 1:
        print(f"  {n['ticker']:<8} | {n['fund'][:40]:<40} | {n['pnl_status']}")
print(f"\nTIER 2 (MAX SKIN + FLAT/MODEST IN MONEY):")
for n in SKIN_IN_GAME:
    if n['tier'] == 2:
        print(f"  {n['ticker']:<8} | {n['fund'][:40]:<40} | {n['pnl_status']}")
