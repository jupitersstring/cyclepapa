"""Re-rank archetype workbook from ENTRY-TODAY perspective.

Methodology:
  - Asymmetry score = (remaining upside vs current price) / (downside risk from current price)
  - Names that already moved on their catalyst have their UPSIDE captured = score drops
  - Names still in trough have FULL upside ahead = score rises
  - Adds 'rerated_status' field: NOT / PARTIAL / RERATED
  - New cross_rank reflects entry-today asymmetry, not original setup asymmetry
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Re-import the original ARCHETYPES from the build script
import sys, importlib.util
spec = importlib.util.spec_from_file_location("build_arch", "/tmp/build_archetype_workbook.py")
# Hack: read source, exec only the data section
with open("/tmp/build_archetype_workbook.py") as f:
    src = f.read()
# Extract just up to the ARCHETYPES dict close
ns = {}
exec(compile(src.split("# ================================================================\n# BUILD WORKBOOK")[0], "<arch>", "exec"), ns)
ARCHETYPES = ns['ARCHETYPES']

# ================================================================
# ENTRY-TODAY RE-RANKING TABLE
# Based on /home/user/cyclepapa/not_yet_rerated_asymmetric.txt
# Each ticker mapped to (new_cross_rank, rerated_status, entry_today_asym,
#                        remaining_upside, remaining_downside, rr, note)
# ================================================================

ENTRY_TODAY = {
    # ============================================================
    # TIER S — VIOLENT RE-RATE + DEEP INTRINSIC VALUE GAP TODAY
    # Setup combines: (a) cash/NAV floor bounding downside +
    # (b) binary or dated trigger + (c) tight float or smart-money
    # anchored at/near current price
    # ============================================================
    'OPRX':    (1,  'NOT', 9.5, '2-4x ($11.50 PT)',   '-30%', '12:1',
        'SUB-BOOK 0.70x + $19M FCF profitable + insider 14.82% + analyst $11.50 (138%) + sub-$100M float = ANY DSP Q4 print violently re-rates. Smart money still building (not exiting). True intrinsic value > 2x current.'),
    'PRLD':    (2,  'NOT', 9.5, '3-8x',               '-15%', '20:1',
        'NET CASH $2.20/sh = 51% of $4.26 price = downside FLOORED. Baker Bros 15.5% 13D + RA Capital 9.99% 13G BOTH anchored SAME April $90M financing at $4.44 — buying TODAY $4.26 is BELOW smart money entry. JAK2 binary upside 5-10x.'),
    'INMD':    (3,  'NOT', 9.0, '2-3x ($25-30)',      '-15%', '13:1',
        'CASH FLOOR: $8.47/sh = 62% of $13.71. Steel Partners 25.5% book + implied $18-20 takeout. Founder Mizrahy +800k sh ($10.7M Feb 2026) at current = sponsor put. Cash + M&A double-floor; upside on Steel re-bid.'),
    'HHH':     (4,  'NOT', 9.5, '1.5-2x to NAV',      '-15%', '13:1',
        'NAV $95-105 vs $63.77 = 33-39% DISCOUNT to real (not aspirational) NAV. Ackman 47% at $100 cost = sponsor PUT floor. Vantage close Q2 (30-45d) = dated trigger. Insurance float makes compounding violent post-close.'),
    'LQDA':    (5,  'NOT', 9.0, '2-3x or -50% (binary)', '-50%', '6:1 (9:1 risk-adj)',
        "'327 PATENT RULING PENDING = pure BINARY. Buckley 32% LONG + PUT structure ($44.7M notional) = sophisticated insider expectation of VIOLENT move. UTHR M&A target. Q1 $129.9M (+44% QoQ) momentum continues. Defined event, not slow re-rate."),
    'CTMX-WT': (6,  'NOT', 9.0, '2-4x or zero (5-wk timer)', '-100%', '5:1',
        'BVF Tranche 2 warrants EXPIRE 7/3/2026 = 5-WEEK binary timer. Most extreme dated catalyst in universe. BVF KYMR conviction signals biotech book quality. Pure binary structure.'),
    'MRP':     (7,  'NOT', 8.5, '1.8-2.5x to NAV',    '-15%', '13:1',
        'B/M 1.274 = DEEPEST in 6/7 cohort (deeper than NRP 0.47). Lennar 80% owner = guaranteed off-take + patient capital. Land bank at COST BASIS vs market. Pure NAV unwind; spinoff orphan technical pressure resolves into Q2 prints.'),

    # ============================================================
    # TIER A — DEEPLY UNDERVALUED + DEFINED CATALYST
    # ============================================================
    'GLOB':    (8,  'PARTIAL', 8.5, '3-5x if class action clears', '-35%', '10:1',
        'P/E 5.91 SECTOR LOW + P/FCF 5.56 + EV/EBITDA 4.95x + 17% FCF yield = DEEPLY undervalued by every metric. Class action JUN 23 IS the catalyst that creates the discount (clears = violent rerate). AI Pods $32.8M ARR from $0 in 12mo + $352M pipeline.'),
    'CRTO':    (9,  'NOT', 8.5, '2.5-4x',             '-20%', '13:1',
        'P/E 3.73 fwd + P/FCF 4.98 + EV/EBITDA 2.28x + 20.7% FCF yield = DEEP. Q3 2026 anniversary of scope losses REMOVES the headwind printed in Q1. Petrus Advisers activist + $200M buyback (22% cap) + Nierenberg D3 +19% Q1.'),
    'NRP':     (10, 'NOT', 8.5, '2.5-3x',             '-20%', '13:1',
        '14.5% FCF yield ALONE = pure mispricing. Royalty trust at 6.6x P/FCF vs TPL at 25x P/CF (SAME asset class) = comp gap re-rate. Distribution Nov 2026 + preferreds retire. Saber 17.75% + Right Tail NEW Q1 2026.'),
    'AAP CALL': (11, 'NOT', 8.5, '3-5x via calls',    '-100%', '6:1',
        'Cooper Creek $158M call notional +213% Q1 = leveraged play on H Partners 46% cash thesis. Q2 margin proof Aug = trigger. Two-fund double-conviction across styles. Options leverage on violent move.'),
    'CARS':    (12, 'NOT', 8.0, '2-3x',               '-20%', '11:1',
        'P/B 0.83 + P/FCF 3.5 + 28% TTM FCF yield + EV/EBITDA 2.7x = DEEP. Q1 EPS $0.45 vs $0.13 est (+246% BEAT). $90M buyback (16% cap raised from $60M) = buyback math compounds violently with cheap multiple.'),
    'KBR':     (13, 'NOT', 8.0, '2-2.5x ($55-65 SoTP)', '-25%', '10:1',
        'SoTP MTS $5.2B EV + STS $5.3B = $55-65 vs $30 today. MTS SPIN JAN 4 2027 = dated trigger 7 months out. D3 +425% Q1 (LARGEST single add in entire universe). Director Form 4 buys May 2026 = sponsor put.'),
    'FUN':     (14, 'NOT', 8.0, '2-3x (3x outright sale)', '-30%', '9:1',
        '42 parks IRREPLACEABLE RE @ $47M each = $2B floor. Jaffer joined board May 26 + Jana 9% pushing OUTRIGHT SALE = two-front activist pressure. H Partners 53.7% / 5.7% of class. Sale = violent 3x trigger.'),
    'AAP':     (15, 'NOT', 8.0, '2.5-3x ($130-150 AZO comp)', '-30%', '9:1',
        '-80% drawdown from 2021. H Partners 46.3% + Q1 +50% sh add ($71.2M). Bridge thesis: AZO-style 8-10% margins = $130-150. Q2 Aug 2026 = margin proof catalyst. Op margin +410bps already. Not cash-floored but H Partners concentration = sponsor put.'),
    'ACHC':    (16, 'NOT', 8.0, '2-3x + squeeze',     '-40%', '6:1',
        '-70% drawdown + 29% SHORT FLOAT = squeeze fuel. EV/EBITDA 6.5x vs peers 8-9x. DOJ resolution binary catalyst. Greenlight 4.1M sh + Sohn 5/12 pitch = thesis surfaced publicly.'),
    'PVLA':    (17, 'NOT', 7.5, '2-3x',               '-25%', '10:1',
        'Cash $25/sh = 19% of $130 price. Phase 3 SELVA POSITIVE Feb 24 2026 = DE-RISKED. Pre-NDA Q2 + NDA H2 + PDUFA H1 2027 = staircase of catalysts. BVF 8.5% of company + Suvretta 3.67% (+29.3% Q1).'),
    'KYMR':    (18, 'NOT', 7.5, '3-5x',               '-30%', '13:1',
        'Cash $1.55B + BVF 14.6% MAX conviction + Baker 4.1% = multi-fund convergence. BUT major catalyst BROADEN2 mid-2027 = 12+ months away (low velocity). Strong setup, slow trigger.'),
    'VITL':    (19, 'NOT', 7.5, '2-3x ($30-40)',      '-25%', '10:1',
        'Deepest drawdown -68.76% + 11% ROA sustained = exceptional. BUT B/M 0.74 = not deeply undervalued vs book; M&A precedent 4x sales = $50 = +260%. Brand-led re-rate is SLOW not violent. No defined binary catalyst.'),
    'FTLF':    (20, 'NOT', 7.5, '3-5x',               '-30%', '13:1',
        'Microcap $93M = sub-institutional minimum = any flow violently moves. Yartseva 7/7 cleanest setup. -43.9% 6mo. B/M 0.49. Pure micro-mean-reversion.'),
    'MTY.TO':  (21, 'NOT', 7.5, '2-2.5x',             '-15%', '13:1',
        'AT BOOK 1.04x + 16.7% FCF yield = pure value. M&A pool at 12-14x EBITDA vs trading 7.5x = M&A precedent arb. CEO Lefebvre open-market buy at C$30.54 = sponsor put. QSR consolidation trend.'),
    'BZU.IM':  (22, 'NOT', 7.5, '1.7-2x (Kerrisdale €85 PT)', '-20%', '8:1',
        '6.8x EV/EBITDA vs peers 9-11x. AI data-center concrete demand = real (not theoretical). 52% EBITDA from US. Kerrisdale Oct 2025 €85 PT (+73%). Italian listing = institutional underweight = re-rate slow but explicit.'),

    # ============================================================
    # TIER B — VALUE / SLOWER RE-RATE
    # ============================================================
    'COTY':    (23, 'NOT', 7.0, '2-4x',               '-30%', '9:1',
        'B/M 1.73 DEEPEST + FCF yield 17.4% but ROA -4.91% CAVEAT (Y3 fails). KKR overhang resolution = catalyst. Brand assets (Burberry, Tiffany, CK) worth multiples of EV. Operational risk balances deep value.'),
    'JOE':     (24, 'NOT', 7.0, '2-3x to LIFO-NAV',   '-20%', '10:1',
        '170k Florida Panhandle acres. GAAP B/M understates LIFO land basis vs market. True NAV-based B/M likely >1.0. Berkowitz anchor + Florida population growth.'),
    'MGNI':    (25, 'NOT', 7.0, '3-5x',               '-25%', '12:1',
        'Independent CTV SSP = winner of ad-tech wash-out. Netflix/Disney/Roku scaling. AI integration. Nine Ten 13.1% / 3.48M sh = highest concentration.'),
    'FRPT':    (26, 'NOT', 6.5, '2-3x',               '-30%', '8:1',
        'B/M 0.50 borderline + ROA 3.41% + -40% drawdown. Manufacturing capacity online = margin inflection. Marlowe 39.8% (highest concentration). M&A target. But valuation not as deep as Tier S/A.'),
    'CTSH':    (27, 'NOT', 6.5, '2-3x',               '-15%', '12:1',
        'Yartseva 6/7 + 9.88% FCF yield + ROA 10.44% + -27% 6mo. AI accelerator-not-disruptor thesis. Defensive value vs deep-value pure play.'),
    'BNTC':    (28, 'NOT', 6.5, '3-5x',               '-50%', '7:1',
        'Suvretta 44.1% 13D = LARGEST stake by %. Gene therapy platform. Multi-program optionality. Higher risk than other biotech-multi-fund names.'),
    'DLTR':    (29, 'PARTIAL', 6.5, '2-3x',           '-20%', '9:1',
        'B/M 0.20 Yartseva 6/7 + Family Dollar divestiture proceeds = $1B+ cash bolster. Post-divest clean Dollar Tree franchise. Multi-fund deep-value consensus already established = partial discount captured.'),
    'QRHC':    (30, 'NOT', 6.0, '2-3x',               '-30%', '8:1',
        'Wynnefield 13.3% 13D w/ board seat. Sub-$200M micro waste services. Recurring revenue.'),
    'POSTBPB': (31, 'NOT', 6.0, '2-3x',               '-30%', '8:1',
        'Nierenberg/D3 LARGEST shareholder (~$16.4M / 7% book). Restaurant turnaround + soft activist pressure.'),
    'ROCK':    (32, 'NOT', 5.5, '1.8-2.2x',           '-25%', '7:1',
        'OmniMax integration year. FY27 EPS $5+ at 15x = $75-90. Director Form 4 buy March 2026. Cyclical, not deep value.'),

    # ============================================================
    # SPECIAL STRUCTURE
    # ============================================================
    'XBI':     (33, 'NOT', 6.5, '1.5-2x via straddle', '-30%', '5:1',
        'Stonepine 40.9% in straddle = pure long-vol on biotech FDA cycle dispersion 2026-2027. Not directional value bet.'),

    # ============================================================
    # PARTIAL RE-RATE / WEAKER ASYMMETRY FROM TODAY
    # ============================================================
    'TPL':     (34, 'PARTIAL', 5.5, '1.3-1.5x',       '-25%', '5:1',
        'Already at 25x P/CF premium = ENDLESS DRILLING priced. NAV-vs-GAAP issue similar to JOE but multiple no longer cheap.'),
    'REGN':    (35, 'PARTIAL', 6.0, '1.5-2x',         '-15%', '8:1',
        'Defensive quality. Pipeline optionality. Partial trough recovery underway.'),
    'DV':      (36, 'PARTIAL', 5.5, '1.5-2x (was 3x)', '-25%', '6:1',
        'CTV verification rally underway = partial upside captured. Q1 +10% rev printed.'),

    # ============================================================
    # RE-RATED — CATALYST PLAYED, UPSIDE CAPTURED, AVOID NEW ENTRY
    # ============================================================
    'CELC':    (37, 'RERATED', 4.5, '1.3-2x post-data', '-30%', '5:1',
        'Phase 3 May 1 POSITIVE + +76% PT raises already in price. ASCO Jun 2 + PDUFA Jul 17 remain but pre-quantified. Quality+catalysts but entry economics weakened post-data.'),
    'HRMY':    (38, 'RERATED', 4.5, '1.5-2x',         '-30%', '6:1',
        'Q1 +17% WAKIX printed; pediatric FDA Feb priced; Lilly orexin validated = thesis worked. Buy-pre-catalyst opportunity passed.'),
    'TREE':    (39, 'RERATED', 4.0, '1.3-1.7x',       '-30%', '5:1',
        'Q1 +37% rev / +71% EBITDA already PRINTED. S&P upgraded. Insurance cycle now consensus. Inflection over.'),
    'ASND':    (40, 'RERATED', 4.0, '1.2-1.5x',       '-20%', '5:1',
        'YORVIPATH ramp priced. €247M Q1 (2x YoY) already at $247. Multiple normalized.'),
    'DHR':     (41, 'RERATED', 4.0, '1.3-1.5x',       '-15%', '5:1',
        'Bouncing from trough; bioprocessing cycle now consensus; multiple expansion captured. Defensive 2H 2026 fully priced.'),
}

# ================================================================
# APPLY RE-RANKING TO ARCHETYPES
# ================================================================

for arch_name, arch_data in ARCHETYPES.items():
    for n in arch_data['names']:
        ticker_key = n['ticker']
        if ticker_key in ENTRY_TODAY:
            rank, status, asym, upside, downside, rr, note = ENTRY_TODAY[ticker_key]
            n['cross_rank'] = rank
            n['rerated_status'] = status
            n['asym'] = asym
            n['upside'] = upside
            n['downside'] = downside
            n['rr'] = rr
            n['entry_today_note'] = note
        else:
            n['rerated_status'] = 'UNKNOWN'
            n['entry_today_note'] = ''

# ================================================================
# BUILD WORKBOOK
# ================================================================

wb = openpyxl.Workbook()

# Styles
BOLD = Font(bold=True, size=11)
HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
NOT_RERATED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # green
PARTIAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')      # yellow
RERATED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')      # red
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def status_fill(s):
    return {'NOT': NOT_RERATED_FILL, 'PARTIAL': PARTIAL_FILL, 'RERATED': RERATED_FILL}.get(s)


# ============ Index sheet ============
ws = wb.active
ws.title = "Index"
ws.merge_cells('A1:H1')
ws['A1'] = "INVESTMENT ARCHETYPES — ENTRY-TODAY ASYMMETRY RANKING"
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = CENTER

ws['A3'] = "Generated: 2026-05-30 (Entry-Today re-rank)"
ws['A4'] = "Methodology: Asymmetry scored from CURRENT price forward, not original setup"
ws['A5'] = "Re-rated names (catalyst already played, upside captured) downgraded vs Not-Re-rated names"

ws['A7'] = "STATUS LEGEND"
ws['A7'].font = HEADER
ws['A7'].fill = HEADER_FILL
ws.merge_cells('A7:H7')

ws['A8'] = "NOT"
ws['A8'].fill = NOT_RERATED_FILL
ws['B8'] = "Major catalyst still ahead; upside intact from current price"
ws['A9'] = "PARTIAL"
ws['A9'].fill = PARTIAL_FILL
ws['B9'] = "Some catalyst played; remaining upside reduced but meaningful"
ws['A10'] = "RERATED"
ws['A10'].fill = RERATED_FILL
ws['B10'] = "Catalyst played + price moved; entry-today economics weakened"

ws['A12'] = "ARCHETYPE INDEX"
ws['A12'].font = HEADER
ws['A12'].fill = HEADER_FILL
ws.merge_cells('A12:H12')

ws['A14'] = "#"
ws['B14'] = "Archetype"
ws['C14'] = "Names"
ws['D14'] = "Top Pick (Entry-Today Rank)"
for cell in ['A14', 'B14', 'C14', 'D14']:
    ws[cell].font = BOLD
    ws[cell].fill = SECTION_FILL

row = 15
for i, (arch_name, arch_data) in enumerate(ARCHETYPES.items(), 1):
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=arch_name)
    ws.cell(row=row, column=3, value=len(arch_data['names']))
    # Find top within archetype by ENTRY-TODAY asymmetry (use cross_rank ascending)
    top = sorted(arch_data['names'], key=lambda x: x['cross_rank'])[0]
    ws.cell(row=row, column=4, value=f"{top['ticker']} (#{top['cross_rank']}, {top['rerated_status']})")
    row += 1

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 30

# ============ MASTER Cross-Rank ============
ws_master = wb.create_sheet("MASTER Entry-Today")
ws_master.merge_cells('A1:N1')
ws_master['A1'] = "MASTER ENTRY-TODAY ASYMMETRIC RANKING"
ws_master['A1'].font = Font(bold=True, size=14)
ws_master['A1'].alignment = CENTER

# Collect all names, dedupe by ticker keeping lowest cross_rank
all_names = []
for arch_name, arch_data in ARCHETYPES.items():
    for name in arch_data['names']:
        nc = dict(name)
        nc['archetype'] = arch_name
        all_names.append(nc)
all_names.sort(key=lambda x: x['cross_rank'])
seen = set()
deduped = []
for n in all_names:
    if n['ticker'] not in seen:
        seen.add(n['ticker'])
        deduped.append(n)

headers = ['Rank', 'Status', 'Ticker', 'Name', 'Archetype', 'Mcap', 'Price', 'Asym',
          'Remaining Upside', 'Downside', 'R/R', 'Entry-Today Note', 'Smart Money', 'Catalyst']
for col, h in enumerate(headers, 1):
    cell = ws_master.cell(row=3, column=col, value=h)
    cell.font = HEADER
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

for row_idx, n in enumerate(deduped, 4):
    ws_master.cell(row=row_idx, column=1, value=n['cross_rank'])
    sc = ws_master.cell(row=row_idx, column=2, value=n['rerated_status'])
    if status_fill(n['rerated_status']):
        sc.fill = status_fill(n['rerated_status'])
    sc.font = Font(bold=True)
    ws_master.cell(row=row_idx, column=3, value=n['ticker']).font = Font(bold=True)
    ws_master.cell(row=row_idx, column=4, value=n['name'])
    ws_master.cell(row=row_idx, column=5, value=n['archetype'])
    ws_master.cell(row=row_idx, column=6, value=n['mcap'])
    ws_master.cell(row=row_idx, column=7, value=n['price'])
    ws_master.cell(row=row_idx, column=8, value=n['asym'])
    ws_master.cell(row=row_idx, column=9, value=n['upside'])
    ws_master.cell(row=row_idx, column=10, value=n['downside'])
    ws_master.cell(row=row_idx, column=11, value=n['rr'])
    ws_master.cell(row=row_idx, column=12, value=n.get('entry_today_note', ''))
    ws_master.cell(row=row_idx, column=13, value=n['smart_money'])
    ws_master.cell(row=row_idx, column=14, value=n['catalyst'])
    for col in range(1, 15):
        ws_master.cell(row=row_idx, column=col).alignment = LEFT_WRAP

widths = [6, 10, 9, 22, 32, 10, 10, 6, 18, 10, 7, 50, 35, 40]
for i, w in enumerate(widths, 1):
    ws_master.column_dimensions[get_column_letter(i)].width = w

# ============ Per-Archetype Sheets ============
for arch_name, arch_data in ARCHETYPES.items():
    sheet_name = arch_name.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '')[:31]
    ws_a = wb.create_sheet(sheet_name)

    ws_a.merge_cells('A1:L1')
    ws_a['A1'] = arch_name + ' — ENTRY-TODAY RANKING'
    ws_a['A1'].font = Font(bold=True, size=13)
    ws_a['A1'].alignment = CENTER

    ws_a.merge_cells('A2:L2')
    ws_a['A2'] = f"Definition: {arch_data['definition']}"
    ws_a['A2'].font = Font(italic=True, size=10)
    ws_a['A2'].alignment = LEFT_WRAP

    # Sort by entry-today cross_rank within archetype (lower rank = better)
    arch_sorted = sorted(arch_data['names'], key=lambda x: x['cross_rank'])

    headers = ['Within-Rank', 'Master-Rank', 'Status', 'Ticker', 'Name', 'Mcap', 'Price',
              'Asym', 'Rem Upside', 'Downside', 'R/R', 'Smart Money']
    for col, h in enumerate(headers, 1):
        cell = ws_a.cell(row=4, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, n in enumerate(arch_sorted):
        r = 5 + i*4
        ws_a.cell(row=r, column=1, value=i+1)
        ws_a.cell(row=r, column=2, value=n['cross_rank'])
        sc = ws_a.cell(row=r, column=3, value=n['rerated_status'])
        if status_fill(n['rerated_status']):
            sc.fill = status_fill(n['rerated_status'])
        sc.font = Font(bold=True)
        ws_a.cell(row=r, column=4, value=n['ticker']).font = Font(bold=True)
        ws_a.cell(row=r, column=5, value=n['name'])
        ws_a.cell(row=r, column=6, value=n['mcap'])
        ws_a.cell(row=r, column=7, value=n['price'])
        ws_a.cell(row=r, column=8, value=n['asym'])
        ws_a.cell(row=r, column=9, value=n['upside'])
        ws_a.cell(row=r, column=10, value=n['downside'])
        ws_a.cell(row=r, column=11, value=n['rr'])
        ws_a.cell(row=r, column=12, value=n['smart_money'])

        ws_a.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=12)
        ws_a.cell(row=r+1, column=1, value=f"ENTRY-TODAY: {n.get('entry_today_note', '')}").alignment = LEFT_WRAP
        ws_a.cell(row=r+1, column=1).fill = SECTION_FILL
        ws_a.cell(row=r+1, column=1).font = Font(italic=True)

        ws_a.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=12)
        ws_a.cell(row=r+2, column=1, value=f"THESIS: {n['thesis']} | VALUATION: {n['valuation']}").alignment = LEFT_WRAP

        ws_a.merge_cells(start_row=r+3, start_column=1, end_row=r+3, end_column=12)
        ws_a.cell(row=r+3, column=1, value=f"CATALYST: {n['catalyst']} | VARIANT: {n['variant']}").alignment = LEFT_WRAP

    widths = [10, 10, 9, 8, 22, 10, 10, 7, 15, 10, 8, 35]
    for i, w in enumerate(widths, 1):
        ws_a.column_dimensions[get_column_letter(i)].width = w
    ws_a.row_dimensions[4].height = 25

# Save
out_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
print(f"\nTotal asymmetric candidates: {len(deduped)} (deduplicated across archetypes)")
print(f"\nTop 15 by ENTRY-TODAY rank:")
print(f"{'#':<4} {'Status':<8} {'Ticker':<10} {'Asym':<6} {'Upside':<20} {'Note (truncated)'}")
for n in deduped[:15]:
    note = n.get('entry_today_note', '')[:60]
    print(f"  #{n['cross_rank']:<3} {n['rerated_status']:<8} {n['ticker']:<10} {n['asym']:<6} {n['upside']:<20} {note}")
print(f"\nRE-RATED names (downgraded for entry-today):")
for n in deduped:
    if n['rerated_status'] == 'RERATED':
        print(f"  #{n['cross_rank']:<3} {n['ticker']:<10} {n['asym']} — {n.get('entry_today_note', '')[:80]}")
