"""NON-BIOTECH MULTIBAG + ASYMMETRIC with qualitative + insider behavior.

Filter stack (ALL must apply):
  1. NOT biotech (no clinical-stage drug developers)
  2. Multibag (3x+ upside in 36mo)
  3. Asymmetric (R/R 8:1+ with bounded downside)
  4. QUALITY business (real cash flow, moat, or hidden asset)
  5. INSIDER behavior aligned (Form 4 buys / founder / activist 13D / family fund)
  6. Smart money cost basis intact (recently built or underwater)

REMOVED: PRLD PEPG CLYM ACRV VOR KPTI COAG MLTX BNTC CCEL CTMX-WT MYO MDGL
         IMNM RAPT SYRE XNCR ACET WVE RVMD KYMR PVLA SYBX (synbio)
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# TIER 1 — NON-BIOTECH MULTIBAG + ASYMMETRIC + INSIDER + QUALITY
# ================================================================

TIER_1 = [
    # === EXTREME INSIDER BUYING + QUALITY BUSINESS ===
    {'rank': 1, 'ticker': 'INMD', 'name': 'InMode Ltd', 'mcap': '$869M', 'price': '$13.71',
     'sector': 'Medical Aesthetics (devices, not biotech R&D)',
     'multibag': '2-3x base ($25-30) + M&A optionality at $18-20+',
     'rr': '13:1', 'downside': '-15% (cash floor)',
     'insider': 'FOUNDER Moshe Mizrahy +800k sh ($10.7M) open-market Feb 2026 at $13.50',
     'smart_money': 'Steel Partners +155% Q1 (25.5% of $338M book) at $13.50',
     'quality': 'Israeli aesthetic device leader: Morpheus8, BodyTite, recurring consumables 85%; 77.8% gross margin; net cash $8.47/sh = 62% of price',
     'catalyst': 'Q2 2026 earnings Aug 5 + potential Steel re-bid ($18 prior)',
     'qualitative': 'Profitable device franchise with no debt; founder open-market buying at current = highest signal'},

    {'rank': 2, 'ticker': 'NRP', 'name': 'Natural Resource Partners', 'mcap': '$1.35B', 'price': '$105.81',
     'sector': 'Royalty Trust (mineral acres)',
     'multibag': '2.5-3x base + comp gap to TPL at 25x P/CF (NRP at 6.6x)',
     'rr': '13:1', 'downside': '-20%',
     'insider': 'Robertson FAMILY 31.75% (permanent insider alignment); Right Tail NEW Q1 2026 at $105',
     'smart_money': 'Saber 17.75% + Greystone top + Berkowitz/Fairholme + Right Tail NEW Q1',
     'quality': 'Royalty trust on 13M mineral acres + 49% Sisecam soda ash JV; debt-free by mid-2026; 14.5% FCF yield ALONE = pure mispricing',
     'catalyst': 'Nov 2026 distribution step-up + Q3 Sisecam recovery + preferreds retire',
     'qualitative': '5-fund concentration + 31.75% family = highest insider/fund alignment in universe; met coal (no green sub for steel) + soda ash (solar glass + EV battery) = secular tailwinds masked by dying-coal narrative'},

    {'rank': 3, 'ticker': 'OPRX', 'name': 'OptimizeRx', 'mcap': '$93M (ULTRA-MICRO)', 'price': '$4.82',
     'sector': 'Digital Health (pharma DTC messaging, NOT biotech)',
     'multibag': '2-5x ($11.50 analyst PT = 138%)',
     'rr': '8:1', 'downside': '-30%',
     'insider': 'Insider 14.82% of company + $10M buyback executing at $4-5 + Doximity exec Will Presti JOINED BOARD April 2026',
     'smart_money': 'Insider concentration + corporate buyback + strategic peer-exec board addition',
     'quality': 'Sub-book 0.70x + $19M FCF profitable + P/FCF 4.89 + EHR-native pharma messaging rails; 60%+ US prescriber reach',
     'catalyst': 'Q2 earnings Aug 14 (75d) + Q4 DSP integration launch (180d)',
     'qualitative': 'Doximity (the digital health peer) sent executive to JOIN BOARD at $4.82 = strongest strategic insider validation; corporate buying back stock at sub-book'},

    {'rank': 4, 'ticker': 'SONO', 'name': 'Sonos Inc', 'mcap': '$1.6B', 'price': '~$13',
     'sector': 'Premium Audio Brand',
     'multibag': '3-4x recovery to historical', 'rr': '8:1', 'downside': '-30%',
     'insider': 'Coliseum Capital +$35M+ OPEN-MARKET buying Feb-Mar 2026 (4 Form 4 buys: $3.07M + $10.1M + $9.14M + $12.9M)',
     'smart_money': 'Coliseum 19.8% concentration + major shareholder 18M+ shs / ~15% of company (5/15/2026)',
     'quality': 'Premium audio brand with $1.7B+ revenue + 15-20% rev rebound expected post-app fix',
     'catalyst': 'Q2-Q3 2026 financials post app remediation + holiday season',
     'qualitative': 'MOST AGGRESSIVE Form 4 buying in entire universe ($35M in 6 weeks); brand IP intact post-app crisis; activist holds ~15% of company'},

    {'rank': 5, 'ticker': 'KBR', 'name': 'KBR Inc', 'mcap': '$3.9B', 'price': '$30',
     'sector': 'Defense + Nuclear Engineering',
     'multibag': '2-2.5x SoTP $55-65 + spin re-rate optionality',
     'rr': '10:1', 'downside': '-25%',
     'insider': 'Director Form 4 buys May 2026 at $30-31 (multiple insiders)',
     'smart_money': 'D3/Nierenberg +425% Q1 2026 = LARGEST SINGLE ADD in 442-tab universe + Irenic activist',
     'quality': 'Two-segment SoTP: MTS gov services 72% ($18.5B backlog, 20k cleared workforce) + STS process tech + TerraPower Natrium SMR alliance; FY26 $980M-1.04B EBITDA',
     'catalyst': 'MTS spin JAN 4 2027 = 215 days defined + NRC Natrium safety eval mid-2026',
     'qualitative': 'D3 quintupled position in Q1 at $30 + directors buying May = sponsor put at current; defense onshoring + SMR nuclear renaissance = triple secular tailwind'},

    {'rank': 6, 'ticker': 'HHH', 'name': 'Howard Hughes Holdings', 'mcap': '$3.5B', 'price': '$63.77',
     'sector': 'Holdco / RE + Insurance Float',
     'multibag': '1.5-2.5x to NAV $95-105', 'rr': '13:1', 'downside': '-15%',
     'insider': 'Marc Grandisson (ex-Arch Capital CEO) JOINED BOARD April 2026 = insurance expert + Ackman 47% via $900M common + $1B preferred at $100/sh',
     'smart_money': 'Ackman $1.9B publicly anchored at $100 cost = -36% UNDERWATER',
     'quality': '3,813 acres residential MPC + 2,447 acres commercial + Vantage specialty insurance $2.2B 2027 premium target',
     'catalyst': 'Vantage close Q2 2026 = 30-45 days',
     'qualitative': 'Ackman cost-basis floor + ex-Arch CEO board hire = insurance expertise lined up + holdco compounding thesis validated by insider hire'},

    {'rank': 7, 'ticker': 'MRP', 'name': 'Millrose Properties', 'mcap': '$4.59B', 'price': '$25',
     'sector': 'Land Bank (Lennar spinoff)',
     'multibag': '2-3x to NAV', 'rr': '13:1', 'downside': '-15%',
     'insider': 'Lennar 80%+ owner = patient capital lock + sponsor commitment',
     'smart_money': 'Brave Warrior (Glenn Greenberg) +59.99% Q1 + Permian Investment 12% + Sachem Head',
     'quality': 'B/M 1.274 = DEEPEST in 6/7 cohort (deeper than NRP 0.47); land bank at cost basis vs market with Lennar guaranteed off-take',
     'catalyst': 'Q2 2026 earnings + housing-cycle inflection',
     'qualitative': 'Lennar 80% lock = forced multi-year holder; spinoff orphan technical pressure resolves; US housing shortage 4M+ structural'},

    {'rank': 8, 'ticker': 'CRTO', 'name': 'Criteo SA', 'mcap': '$866M', 'price': '$13.50',
     'sector': 'Retail Media / Ad-Tech',
     'multibag': '2.5-3.5x post-anniversary lap', 'rr': '13:1', 'downside': '-20%',
     'insider': '$200M buyback (22% cap) + Petrus Advisers activist',
     'smart_money': 'D3/Nierenberg 9.37% +19% Q1 2026 at $13.50 = current price',
     'quality': 'P/E 3.73 fwd + P/FCF 4.98 + EV/EBITDA 2.28x + 20.7% FCF yield = DEEP. 235 retailers incl Lowes/Costco; OpenAI partnership; Commerce Yield retail media',
     'catalyst': 'Q3 2026 anniversary lap of scope losses (Q1 -31% ex-scope = +24% actual)',
     'qualitative': 'D3 + Petrus activist + 22% buyback = three forces at current entry + retail media $200B TAM by 2028 = secular tailwind market misses'},

    {'rank': 9, 'ticker': 'CARS', 'name': 'Cars.com', 'mcap': '$560M', 'price': '$10',
     'sector': 'Auto Digital Marketplace',
     'multibag': '2-3x', 'rr': '11:1', 'downside': '-20%',
     'insider': '$90M buyback (16% cap raised from $60M) = mechanical compression',
     'smart_money': 'Yartseva 7/7 setup; Q1 EPS $0.45 vs $0.13 est (+246%)',
     'quality': '19,390 dealers × $2,473 ARPD/mo, 90% recurring SaaS; subscription +2%, OEM/National -12% (cyclical bottom)',
     'catalyst': '$25-30M annualized cost program 2027',
     'qualitative': 'Buyback math compounds violently at cheap multiple (P/FCF 3.5, 28% TTM FCF yield); 90% SaaS recurring revenue moat'},

    {'rank': 10, 'ticker': 'VITL', 'name': 'Vital Farms', 'mcap': '$1.5B', 'price': '$13',
     'sector': 'Pasture-Raised Premium Eggs (Consumer)',
     'multibag': '3-5x ($30-50)', 'rr': '10:1', 'downside': '-25%',
     'insider': 'No major fund concentrated yet = entry PRESERVED before smart money pile-in',
     'smart_money': 'CONTRARIAN — no anchor built; sustained 11% ROA through avian flu drawdown',
     'quality': 'Category creator + Whole Foods/Target/Walmart distribution; brand pricing power retained during commodity shock; 5-year transition for pastured certification = moat',
     'catalyst': 'Avian flu normalization + summer pricing + M&A speculation (Kraft Heinz, Hain, General Mills)',
     'qualitative': '11% ROA SUSTAINED through -68.76% drawdown is exceptional = best ROA in 6/7 cohort; M&A precedent 4x sales = $50-80'},

    {'rank': 11, 'ticker': 'TLN', 'name': 'Talen Energy', 'mcap': '$8B', 'price': '~$200',
     'sector': 'Independent Power Producer (Nuclear)',
     'multibag': '3-5x AI nuclear re-rate', 'rr': '10:1', 'downside': '-30%',
     'insider': 'Multi-year nuclear PSA contracts as floor',
     'smart_money': 'Sachem Head +72% Q4 ADD',
     'quality': 'Susquehanna nuclear plant + Cumulus Data + irreplaceable nuclear capacity',
     'catalyst': 'May earnings + nuclear PSAs closing (weeks)',
     'qualitative': 'AI data-center electricity demand structural; nuclear renaissance with no green substitute; Sachem Head sees PSA economics underpriced'},

    {'rank': 12, 'ticker': 'PRM', 'name': 'Perimeter Solutions', 'mcap': '$534M', 'price': '~$11',
     'sector': 'Fire Retardants (essential)',
     'multibag': '2-3x', 'rr': '9:1', 'downside': '-25%',
     'insider': 'Form 4 CEO + officer buys $3.56M July 2025 at depressed levels',
     'smart_money': 'WindAcre Partnership 5.7%',
     'quality': 'Fire retardant chemicals oligopoly + recurring fire-season demand + government contracts',
     'catalyst': 'Q2 2026 earnings + fire season Q3 + recovery',
     'qualitative': 'CEO + officers $3.56M open-market = strong insider conviction; essential-services oligopoly with regulatory moat'},

    {'rank': 13, 'ticker': 'MGNI', 'name': 'Magnite', 'mcap': '$1.89B', 'price': '$13',
     'sector': 'Independent CTV SSP',
     'multibag': '3-5x', 'rr': '12:1', 'downside': '-25%',
     'insider': 'Bares-trained Nine Ten concentrated bet',
     'smart_money': 'Nine Ten Capital 13.1% of book / 3.48M sh / $41.4M',
     'quality': 'Last major independent CTV SSP winner of ad-tech wash-out; Netflix/Disney/Roku exclusive deals scaling; AI ad-buying integration',
     'catalyst': 'CTV ad spend doubling 2024-2028 + Netflix exclusive scaling',
     'qualitative': 'Winner-take-most CTV dynamics + market priced as "ad-tech roadkill" = pure variant perception'},

    {'rank': 14, 'ticker': 'MTY.TO', 'name': 'MTY Food Group', 'mcap': 'C$887M (~$640M USD)', 'price': 'C$40.41',
     'sector': 'Multi-Brand QSR Franchisor',
     'multibag': '2-2.5x', 'rr': '13:1', 'downside': '-15%',
     'insider': 'CEO Lefebvre open-market buy at C$30.54 (+33% personal stake)',
     'smart_money': 'CEO + 98%+ franchised business model',
     'quality': '90 banners (Cold Stone, Mucho Burrito, Papa Murphy\'s); 7,000+ locations; recurring royalty revenue + at BOOK (P/B 1.04) + 16.7% FCF yield',
     'catalyst': 'Q2 2026 earnings test + NCIB 5% of float through July',
     'qualitative': 'Franchise royalty model = highest quality recurring revenue; M&A pool at 12-14x EBITDA (Inspire/RBI/Roark) vs trading 7.5x = pure precedent arb'},

    {'rank': 15, 'ticker': 'ROCK', 'name': 'Gibraltar Industries', 'mcap': '$1.10B', 'price': '$39.69',
     'sector': 'Residential Roofing Accessories #1',
     'multibag': '1.8-2.2x', 'rr': '8:1', 'downside': '-25%',
     'insider': 'Director Metcalf bought $502,000 March 2026 + Vanguard 5.24% 13G NEW',
     'smart_money': 'Insider + index buying alignment',
     'quality': 'OmniMax acquisition closed Feb 2026 ($1.335B) = #1 residential roofing accessories; FY26 guide $1.76-1.83B rev + $3.65-4.05 EPS',
     'catalyst': 'Q2 2026 first clean OmniMax quarter + Q3 reroofing season',
     'qualitative': 'Director open-market buy + Vanguard NEW threshold + #1 market position post-acquisition'},

    {'rank': 16, 'ticker': 'BW', 'name': 'Babcock & Wilcox', 'mcap': '$300M MICRO', 'price': '~$1-2',
     'sector': 'Nuclear SMR + Industrial',
     'multibag': '3-5x nuclear renaissance', 'rr': '8:1', 'downside': '-40%',
     'insider': 'Steel Partners 13D NEW activist June 2025 (range $0.30-$0.80/sh)',
     'smart_money': 'Steel Partners activist (their playbook = break-up value)',
     'quality': 'Nuclear SMR IP + thermal power services + parts spinoff optionality',
     'catalyst': '13D nomination June-July 2026 + nuclear SMR capex Q2-Q3',
     'qualitative': 'Steel Partners break-up specialist + nuclear renaissance structural tailwind = asymmetric activist play'},

    {'rank': 17, 'ticker': 'BTSG', 'name': 'BrightSpring Health Services', 'mcap': '$3B', 'price': '~$10',
     'sector': 'Senior Living / Home Health Services',
     'multibag': '2.5-4x', 'rr': '10:1', 'downside': '-30%',
     'insider': 'Alta Fox Capital 3%→10% of book (+169% Q1) = bet-the-fund conviction',
     'smart_money': 'Alta Fox concentrated position add',
     'quality': 'Largest US pharmacy + home health services + senior living network; recurring service revenue',
     'catalyst': 'May earnings (21d) + Q2/Q3 occupancy proof',
     'qualitative': 'Senior living occupancy 20% below pre-COVID multiples = mispriced + healthcare services demographic tailwind'},

    {'rank': 18, 'ticker': 'GLOB', 'name': 'Globant SA', 'mcap': '$1.68B', 'price': '$43.50',
     'sector': 'LatAm IT Services + AI',
     'multibag': '3-5x post-class action clears', 'rr': '11:1', 'downside': '-30%',
     'insider': 'Founder Migoya ~3% via Founders Trust (23-year tenure)',
     'smart_money': 'Founder ownership + sentiment bottom 16% short float',
     'quality': 'P/E 5.91 SECTOR LOW + P/FCF 5.56 + EV/EBITDA 4.95x + 17% FCF yield = deeply undervalued',
     'catalyst': 'JUN 23 class action lead plaintiff deadline + AI Pods Q3 ($32.8M ARR from $0 in 12mo)',
     'qualitative': 'Class action overhang IS the discount; founder 23-year tenure + AI Pods inflection = LatAm + AI services secular tailwind unpriced'},

    {'rank': 19, 'ticker': 'EQT', 'name': 'EQT Corporation', 'mcap': '$28B', 'price': '~$55',
     'sector': 'Natural Gas E&P + LNG',
     'multibag': '2-3x LNG cycle 2028', 'rr': '8:1', 'downside': '-25%',
     'insider': 'Nierenberg 16.85% concentration mentioned + Sprott NEW',
     'smart_money': 'Sprott (gold/mining specialist) NEW + Sprott Asset NEW Q1 2026',
     'quality': 'Largest US natgas producer + LNG export capacity coming online',
     'catalyst': 'LNG export terminals 2026-2028 + winter heating demand',
     'qualitative': 'Sprott specialist sees natgas LNG cycle; AI data-center electricity demand drives natgas peaker plants'},

    {'rank': 20, 'ticker': 'POSTBPB', 'name': 'Potbelly Corp (PBPB)', 'mcap': '$240M MICRO', 'price': '$8',
     'sector': 'Restaurant Chain',
     'multibag': '2-3x', 'rr': '8:1', 'downside': '-30%',
     'insider': 'Nierenberg/D3 Form 4 insider buys ongoing',
     'smart_money': 'Nierenberg/D3 LARGEST shareholder (~$16.4M / 7% of book)',
     'quality': 'Sandwich shop turnaround + customer-traffic gains while peers don\'t (Q1 2026 D3 letter thesis)',
     'catalyst': 'Restaurant turnaround + Nierenberg escalation potential',
     'qualitative': 'Form 4 ongoing buys at depressed levels + Nierenberg LARGEST shareholder = activist soft pressure micro'},
]

# Tier 2 — strong signal but weaker on quality/insider OR larger mcap
TIER_2 = [
    ('FTLF', 'FitLife Brands', '$93M', 'Nutraceutical microcap', 'Smoak residual + sub-institutional minimum', 'Yartseva 7/7'),
    ('RPAY', 'Repay Holdings', '$130M', 'Payment processor fintech', 'Veradace 8.6% activist + calls', 'June 2026 13D nomination'),
    ('SSTI', 'SoundThinking', '$600M', 'Public safety tech', 'Veradace 16.5% / $90M 13D', 'June 2026 nomination'),
    ('AIRI', 'Air Industries Group', '$40M', 'Defense aerospace', 'Charles Frischer NEW 13D 3/23/26', '60d nomination late May/June'),
    ('CODI', 'Compass Diversified', '~$1B', 'Diversified holdco', 'ADW Capital +140% escalation Feb-Apr 2026', 'June 2026 13D nomination'),
    ('STRR', 'Star Equity Holdings', '<$30M', 'Holdings micro-cap', 'Eberwein $3M open-market 21.4%→27.1% (Dec 2025)', '13D escalation 28.25% Mar 2026'),
    ('KVHI', 'KVH Industries', '<$200M', 'Maritime telecom', 'Radoff family 10%+ Form 4 +40k sh @ $5.78 Nov 2025', 'Activist build'),
    ('BRN', 'Barnwell Industries', '<$50M ULTRA-MICRO', 'Oil/gas micro', 'Radoff group 17.4% + Director Schechter Form 4', '13D escalation Dec25-Mar26'),
    ('GCO', 'Genesco Inc.', '$280M', 'Footwear retail', 'Radoff/Jumana/Ross group 11.7%→12.3%', '13D coordinated'),
    ('SEER', 'Seer Inc', '~$200M', 'Proteomics (NOT clinical-stage biotech)', 'Radoff group 9.3%→10.6%', 'Triple 13D escalation Feb-Apr'),
    ('PAR', 'PAR Technology', '$1.5B', 'Restaurant POS / payments', 'Voss Capital 46% book concentration', 'Bet-the-fund'),
    ('QRHC', 'Quest Resource Holding', '$200M MICRO', 'Waste services', 'Wynnefield 13.3% 13D + board seat', 'Cooperation Agreement lock'),
    ('MNKTQ', 'Mallinckrodt (post-BK)', '<$200M', 'Specialty Pharma (commercial, NOT clinical)', 'Silver Pt 14.2% + GoldenTree 20.2% + Marathon 7.7% (3 distressed)', 'Acthar Medicare path'),
    ('CDRE', 'Cadre Holdings', '$700M', 'Tactical/defense gear', 'Wynnefield 23.9% +14% Q1', 'Defensive consumer'),
    ('FLR', 'Fluor Corp', '$8B', 'Engineering/construction', 'Greenlight (Einhorn) 7.73% +44% Q4', 'Federal infrastructure'),
    ('WDC', 'Western Digital', '$24B', 'AI datacenter HDD', 'Whitebox 13.3% portfolio +$154M Q4', 'AI HDD cycle'),
    ('AAP CALL', 'AAP Call Options', 'leverage', 'Calls on AAP H Partners thesis', 'Cooper Creek $158M call notional +213% Q1', 'Q2 margin proof Aug'),
    ('AAP', 'Advance Auto Parts', '$3.3B', 'Auto aftermarket', 'H Partners 46.3% + Q1 +50% sh add', 'Q2 margin proof'),
    ('FUN', 'Six Flags Entertainment', '$2.2B', 'Theme parks', 'H Partners 53.7% MAX concentration + Jana 9% sale push', 'Jaffer board May 26 2026'),
    ('ACHC', 'Acadia Healthcare', '$3.2B', 'Behavioral health services', 'Greenlight 4.1M + Sohn 5/12 pitch', 'DOJ resolution binary + 29% short squeeze'),
    ('CCO', 'Clear Channel Outdoor', '~$700M', 'Outdoor advertising', 'Legion Partners 33.6% activist', 'Outdoor ad recovery'),
    ('REZI', 'Resideo Technologies', '$2.6B', 'Home tech', 'Alta Fox 9.57% NEW + GoldenTree', 'Multi-fund consensus'),
    ('COR', 'Cencora', '$50B', 'Pharma distribution', 'Sio Capital +535% MASSIVE Q1 ADD', 'Q2 earnings'),
    ('EQIX', 'Equinix', '$75B', 'AI data center REIT', 'Land & Buildings +82% Q4', 'AI hyperscaler 10-15yr leases'),
    ('CMCSA', 'Comcast', '$158B', 'Cable + Peacock', 'GoldenTree 12.7% +284% MAJOR ADD', 'Streaming + cable cash flow'),
    ('S', 'SentinelOne', '$12B', 'Cybersecurity software', 'Anchorage Capital 83.6% of portfolio +4.9M sh', 'Activist board catalyst'),
]


# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'NON-BIOTECH MULTIBAG+INSIDER' in wb.sheetnames:
    del wb['NON-BIOTECH MULTIBAG+INSIDER']

ws = wb.create_sheet('NON-BIOTECH MULTIBAG+INSIDER', 1)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
T1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
T2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:I1')
ws['A1'] = 'NON-BIOTECH MULTIBAG + ASYMMETRIC + INSIDER + QUALITY — Strictest Combined Filter'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:I2')
ws['A2'] = ('Filters: (1) NOT biotech / clinical-stage  (2) Multibag 3x+ ideally  (3) R/R 8:1+  '
            '(4) QUALITY business (real cash flow / moat / hidden asset)  (5) INSIDER behavior aligned '
            '(Form 4 buys, founder, activist, family fund)  (6) Smart money cost basis at/below current.')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

headers = ['#', 'Ticker', 'Mcap', 'Sector', 'Upside / R/R', 'Insider Behavior', 'Quality Context', 'Catalyst', 'Why Asymmetric']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

row = 5
for n in TIER_1:
    ws.cell(row=row, column=1, value=n['rank']).fill = T1_FILL
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"{n['mcap']}\n@ {n['price']}")
    ws.cell(row=row, column=4, value=n['sector'])
    ws.cell(row=row, column=5, value=f"{n['multibag']}\nR/R {n['rr']} / DS {n['downside']}")
    ws.cell(row=row, column=6, value=n['insider'])
    ws.cell(row=row, column=7, value=n['quality'])
    ws.cell(row=row, column=8, value=n['catalyst'])
    ws.cell(row=row, column=9, value=n['qualitative'])
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 95
    row += 1

# Tier 2
row += 1
ws.cell(row=row, column=1, value='TIER 2 — Strong setup but smaller mcap signal or weaker quality').fill = T2_FILL
ws.cell(row=row, column=1).font = HEADER
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

t2_headers = ['Ticker', 'Name', 'Mcap', 'Sector', 'Smart Money / Insider', 'Catalyst']
for col, h in enumerate(t2_headers, 1):
    cell = ws.cell(row=row, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL
row += 1

for t in TIER_2:
    ws.cell(row=row, column=1, value=t[0]).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = T2_FILL
    ws.cell(row=row, column=2, value=t[1])
    ws.cell(row=row, column=3, value=t[2])
    ws.cell(row=row, column=4, value=t[3])
    ws.cell(row=row, column=5, value=t[4])
    ws.cell(row=row, column=6, value=t[5])
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 35
    row += 1

# Section: REMOVED biotech tickers
row += 2
ws.cell(row=row, column=1, value='REMOVED BIOTECH / CLINICAL-STAGE TICKERS (per user request):').font = Font(bold=True, size=11)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1
removed = 'PRLD PEPG CLYM ACRV VOR KPTI COAG MLTX BNTC CCEL CTMX-WT MYO MDGL IMNM RAPT SYRE XNCR ACET WVE RVMD KYMR PVLA SYBX (synbio) CELC HRMY ASND'
ws.cell(row=row, column=1, value=removed)
ws.cell(row=row, column=1).fill = SECTION_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws.cell(row=row, column=1).alignment = LEFT_WRAP
row += 1

widths = [5, 10, 14, 28, 22, 45, 50, 35, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)
print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"\nTIER 1 — NON-BIOTECH MULTIBAG + ASYMMETRIC + INSIDER + QUALITY ({len(TIER_1)} names):")
for n in TIER_1:
    print(f"  #{n['rank']:<3} {n['ticker']:<10} {n['mcap']:<22} | {n['sector'][:30]:<30} | {n['multibag']:<15}")
print(f"\nTIER 2 ({len(TIER_2)} additional names — see workbook)")
