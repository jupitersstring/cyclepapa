"""Build POSITION INTENSITY analysis weighted by % of fund book + dollar skin.

Metrics combined:
  - Sum of concentration % across all funds holding the name
  - Number of funds at >=5% / >=10% / >=20% / >=30% concentration
  - Total dollar skin across funds
  - Cross-reference vs in-money status

A 14.6% MAX position at BVF ($458M) is stronger signal than the same name
held at 1% by ten random funds. Multi-fund convergence WHERE EACH FUND has
significant book % = strongest conviction stack.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# POSITION INTENSITY DATA
# Each entry: ticker, list of (fund, pct_of_book, dollar_size, cost_basis_status)
# ================================================================

POSITIONS = {
    # ============================================================
    # MAX INTENSITY — UNDERWATER / FLAT (TIER 1)
    # ============================================================
    'HHH': {
        'funds': [
            ('Ackman/Pershing Square', '47% of HHH', '$1,900M ($900M common + $1B preferred)', 'cost $100 (-36% UNDERWATER)'),
            ('Northern Right Capital', 'NEW Q1 2026 +12%', '~$30M est', 'recent build'),
        ],
        'in_money': 'UNDERWATER',
        'sum_pct': 47, 'dollar_skin': 1900,
        'key': '$1.9B Ackman commitment publicly anchored at $100; literal billions of skin at -36% underwater = sponsor PUT floor',
    },
    'KYMR': {
        'funds': [
            ('BVF Partners (MAX CONVICTION)', '14.6% of $3.14B book', '$458M', 'cost $86 follow-on (-7%)'),
            ('Baker Bros', '4.1% of $17.4B book', '$721M', 'cost $86 same follow-on (-7%)'),
        ],
        'in_money': 'UNDERWATER',
        'sum_pct': 18.7, 'dollar_skin': 1179,
        'key': '$1.18B COMBINED dollar skin from TWO top biotech specialists; BVF #1 MAX conviction',
    },
    'PRLD': {
        'funds': [
            ('Baker Bros', '15.5% of company (13D/A)', '~$8M (small fund weighting)', 'cost $4.44 April financing (-4%)'),
            ('RA Capital', '9.99% of company (13G)', '~$5M', 'cost $4.44 SAME financing (-4%)'),
        ],
        'in_money': 'UNDERWATER',
        'sum_pct': 25.5, 'dollar_skin': 13,
        'key': '25.5% COMBINED ownership by TWO top biotech crossover funds anchored at SAME $4.44 = below current $4.26 = cannot be sellers',
    },
    'KBR': {
        'funds': [
            ('D3 Family Fund (Nierenberg)', 'Position +425% Q1 (LARGEST single add in universe)', '$105M+ est post-add', 'cost ~$30 (built Q1) FLAT'),
            ('Irenic Capital', '~1% activist position', '~$40M', 'activist anchor'),
            ('Director Form 4', 'Multiple May 2026 buys', '~$1M+', 'at $30-31 confirmed FLAT'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 'N/A', 'dollar_skin': 146,
        'key': 'D3 +425% = SINGLE LARGEST add of any fund in 442-sheet universe; new entrant gets D3 cost basis exactly',
    },
    'INMD': {
        'funds': [
            ('Steel Partners', '25.5% of $338M book / +155% Q1 (DOUBLED)', '$86M', 'cost ~$13.50 FLAT'),
            ('Founder Mizrahy', '800k sh Feb 2026 open-market buy', '$10.7M', 'cost $13.71 FLAT'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 25.5, 'dollar_skin': 97,
        'key': 'Activist + founder BOTH built positions AT CURRENT in 2026; cash $8.47/sh = 62% floor',
    },
    'NRP': {
        'funds': [
            ('Saber Capital (Huber)', '17.75% of book', '~$30M', 'long-term anchor'),
            ('Robertson Family (insider)', '31.75% of company', 'massive insider', 'aligned'),
            ('Right Tail Capital', 'NEW Q1 2026', '~$15M', 'cost ~$105 = current FLAT'),
            ('Greystone Capital', 'Top position', '~$10M', 'long-term'),
            ('Berkowitz/Fairholme', 'Anchor', '~$20M', 'long-term'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 'N/A', 'dollar_skin': 75,
        'key': 'FIVE-fund concentration + insider 31.75% family + Right Tail NEW Q1 at current = freshest entry signal',
    },
    'OPRX': {
        'funds': [
            ('Insider holders', '14.82% of company', '$14M @ $93M cap', 'aligned'),
            ('Buyback', '$10M authorized through March 2027', '$10M', 'executing at $4-5'),
            ('Presti (Doximity exec)', 'NEW board April 2026', 'signal', 'cost at depressed = current'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 14.82, 'dollar_skin': 24,
        'key': 'Insider 15% + buyback + Doximity peer exec to board April 2026 at $4-5 = strong corporate-insider signal',
    },

    # ============================================================
    # MAX INTENSITY — SLIGHTLY IN MONEY / FLAT (TIER 2)
    # ============================================================
    'FUN': {
        'funds': [
            ('H Partners', '53.7% of $154M book (MAX in 13F universe)', '$82.5M = 5.7% of FUN class', 'cost ~$22-30 modest in money'),
            ('Jana Partners', '9% activist push outright sale', '$130M est', 'recent build'),
            ('Travis Kelce co-invest', 'Sub-1% but signal', '$5M', 'aligned'),
        ],
        'in_money': 'FLAT/MODEST',
        'sum_pct': 62.7, 'dollar_skin': 218,
        'key': 'H Partners 53.7% = HIGHEST single-fund concentration of any actionable name; Jana 9% pushing OUTRIGHT SALE; Jaffer board seat May 26',
    },
    'AAP': {
        'funds': [
            ('H Partners', '46.3% of $154M book / +50% sh Q1 ADD (1.35M new)', '$71.2M', 'cost ~$45 modest in money'),
            ('Legion Partners (precedent)', '~3%', '$10M', 'activist precedent'),
        ],
        'in_money': 'MODEST IN MONEY (~5-15%)',
        'sum_pct': 49, 'dollar_skin': 81,
        'key': 'H Partners HALF THEIR BOOK + STILL ADDING Q1 (1.35M new sh); 46% concentration extreme',
    },
    'CRTO': {
        'funds': [
            ('D3/Nierenberg', '9.37% of book +19% Q1', '$5M (small fund)', 'cost ~$11.50-14 FLAT'),
            ('Petrus Advisers', 'Activist position', '~$20M est', 'activist anchor'),
            ('Buyback', '$200M = 22% cap', '$200M corporate', 'executing at $13-14'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 9.37, 'dollar_skin': 25,
        'key': 'D3 added 19% Q1 at current band + Petrus activist + 22% buyback = three-source pressure all at current price',
    },
    'PVLA': {
        'funds': [
            ('BVF Partners', '3.8% of $3.14B book / 8.5% of PVLA company', '$120M', 'cost depressed long'),
            ('Suvretta Capital', '3.67% of book / +29.3% Q1', '$144M', 'cost +29% Q1 added FLAT'),
        ],
        'in_money': 'FLAT to slight in money',
        'sum_pct': 7.5, 'dollar_skin': 264,
        'key': '$264M dual-anchor BVF + Suvretta; Suvretta +29% Q1 ADD AT current; cash 19% of price floor',
    },
    'MRP': {
        'funds': [
            ('Lennar Corporation', '80%+ owner (spinoff sponsor)', '~$3.7B at current MRP cap', 'cost = spin price = FLAT'),
        ],
        'in_money': 'FLAT (Lennar paper position)',
        'sum_pct': 80, 'dollar_skin': 3700,
        'key': 'Lennar 80% spinoff lock-in = patient capital cannot exit; deepest B/M 1.274 in cohort',
    },

    # ============================================================
    # TIER 3 — STRONG SINGLE-FUND CONCENTRATION
    # ============================================================
    'BNTC': {
        'funds': [
            ('Suvretta Capital', '44.1% of BNTC (13D = LARGEST stake by % ownership)', '~$110M', 'activist anchor'),
        ],
        'in_money': 'FLAT/UNDERWATER',
        'sum_pct': 44.1, 'dollar_skin': 110,
        'key': 'Suvretta\'s LARGEST single stake by % ownership in their entire portfolio',
    },
    'FRPT': {
        'funds': [
            ('Marlowe Partners', '39.8% of book (HIGHEST single-fund concentration in universe ex-CVNA)', '~$20M', 'cost depressed FLAT'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 39.8, 'dollar_skin': 20,
        'key': '39.8% concentration = bet-the-fund; manufacturing capacity coming online',
    },
    'ACHC': {
        'funds': [
            ('Greenlight Capital (Einhorn)', '4.1M shares', '$135M', 'cost $30-40 FLAT'),
            ('Sohn pitch 5/12/2026', 'public surfacing', 'signal', 'thesis broadcast at $33'),
        ],
        'in_money': 'FLAT',
        'sum_pct': 'small % book (~3%)', 'dollar_skin': 135,
        'key': 'Einhorn publicly anchored thesis at Sohn 5/12; -70% drawdown + 29% short float = squeeze',
    },
    'MGNI': {
        'funds': [
            ('Nine Ten Capital (Bares)', '13.1% of book / 3.48M sh', '$41.4M', 'cost depressed FLAT'),
        ],
        'in_money': 'FLAT to modest',
        'sum_pct': 13.1, 'dollar_skin': 41,
        'key': 'Bares-trained Nine Ten 13.1% concentration on independent CTV SSP',
    },
    'AAP CALL': {
        'funds': [
            ('Cooper Creek Partners', 'Largest single call exposure', '$158M notional', 'Q1 +213% = recent build at current'),
        ],
        'in_money': 'Calls in money on direction; leverage on H Partners thesis',
        'sum_pct': 'N/A (calls)', 'dollar_skin': 158,
        'key': '$158M Cooper Creek calls layered over H Partners $71M cash = $229M two-fund AAP thesis',
    },

    # ============================================================
    # REFERENCE — HIGH CONCENTRATION BUT DEEPLY IN MONEY
    # (DO NOT chase — fund already captured the move)
    # ============================================================
    'CVNA': {
        'funds': [
            ('CAS Investment Partners', '81.7% of book (HIGHEST single-name in 13F universe)', '$185M', '20-25x in money ($3->$71)'),
        ],
        'in_money': 'DEEPLY IN MONEY (REF)',
        'sum_pct': 81.7, 'dollar_skin': 185,
        'key': 'REFERENCE — CAS legendary 40x recovery; no longer asymmetric',
    },
    'GE': {
        'funds': [
            ('TCI Fund Management (Hohn)', '29.85% of $45B book', '$13B', 'held 2015 deeply in money'),
        ],
        'in_money': 'DEEPLY IN MONEY (REF)',
        'sum_pct': 29.85, 'dollar_skin': 13000,
        'key': 'REFERENCE — TCI captured GE Aerospace breakup; pure compounding',
    },
    'CELC': {
        'funds': [
            ('Baker Bros (ACTIVIST 13D)', '19.99% of CELC = 5.48% of $17.4B book', '$904M', 'pre-data entry deeply in money'),
        ],
        'in_money': 'DEEPLY IN MONEY post-May 1 (REF)',
        'sum_pct': 19.99, 'dollar_skin': 904,
        'key': 'REFERENCE — Baker captured May 1 Phase 3 positive; ASCO Jun 2 + PDUFA Jul 17 remaining but pre-quantified',
    },
    'ASND': {
        'funds': [
            ('RA Capital', '24.9% of $9.44B book = LARGEST biotech crossover position in universe', '$2,350M', 'held since 2015 deeply in money'),
        ],
        'in_money': 'DEEPLY IN MONEY (REF)',
        'sum_pct': 24.9, 'dollar_skin': 2350,
        'key': 'REFERENCE — RA Capital largest single position in entire biotech crossover universe; multiple normalized',
    },
}

# ================================================================
# COMPUTE INTENSITY SCORE
# Score = log10(dollar_skin) * 2 + sum_pct + bonus for multi-fund + bonus for underwater
# ================================================================

import math

def compute_intensity(p):
    score = 0
    # Dollar skin (log-scaled to avoid GE/HHH dominating)
    if p['dollar_skin']:
        score += math.log10(max(p['dollar_skin'], 1)) * 2
    # Sum concentration % (cap at 50)
    sp = p['sum_pct']
    if isinstance(sp, (int, float)):
        score += min(sp, 50) / 5
    # Multi-fund bonus
    score += len(p['funds']) * 1.5
    # Status multiplier
    if 'UNDERWATER' in p['in_money']:
        score *= 1.3
    elif 'FLAT' in p['in_money']:
        score *= 1.15
    elif 'MODEST' in p['in_money']:
        score *= 1.05
    elif 'DEEPLY IN MONEY' in p['in_money']:
        score *= 0.4
    return round(score, 1)


# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'POSITION INTENSITY' in wb.sheetnames:
    del wb['POSITION INTENSITY']

# Insert after MAX Skin sheet
ws = wb.create_sheet('POSITION INTENSITY', 3)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
TIER1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
TIER2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
TIER3_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
REF_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:I1')
ws['A1'] = 'POSITION INTENSITY — % of Fund Book × Dollar Skin × Multi-fund Convergence'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:I2')
ws['A2'] = ('Score = log10($dollar_skin) × 2 + sum_pct/5 + multi_fund_bonus × 1.5; '
            'multiplier 1.3x UNDERWATER, 1.15x FLAT, 1.05x MODEST IN MONEY, 0.4x DEEPLY IN MONEY')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

# Compute and sort
ranked = [(t, p, compute_intensity(p)) for t, p in POSITIONS.items()]
ranked.sort(key=lambda x: -x[2])

# Headers
headers = ['Rank', 'Ticker', 'Intensity', 'Sum % Book', 'Dollar Skin $M', '# Funds', 'P&L Status', 'Funds (concentration / size / cost basis)', 'Key Signal']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

row = 5
for rank, (ticker, p, score) in enumerate(ranked, 1):
    # Tier coloring based on P&L status
    if 'UNDERWATER' in p['in_money']:
        fill = TIER1_FILL
    elif 'FLAT' in p['in_money']:
        fill = TIER1_FILL if rank <= 7 else TIER2_FILL
    elif 'MODEST' in p['in_money']:
        fill = TIER2_FILL
    elif 'DEEPLY IN MONEY' in p['in_money']:
        fill = REF_FILL
    else:
        fill = TIER3_FILL

    ws.cell(row=row, column=1, value=rank).fill = fill
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=ticker).font = Font(bold=True)
    ws.cell(row=row, column=3, value=score)
    sp = p['sum_pct']
    ws.cell(row=row, column=4, value=str(sp) + ('%' if isinstance(sp, (int, float)) else ''))
    ws.cell(row=row, column=5, value=p['dollar_skin'])
    ws.cell(row=row, column=6, value=len(p['funds']))
    ws.cell(row=row, column=7, value=p['in_money'])

    fund_summary = '\n'.join(f"• {f[0]}: {f[1]} | {f[2]} | {f[3]}" for f in p['funds'])
    ws.cell(row=row, column=8, value=fund_summary)
    ws.cell(row=row, column=9, value=p['key'])

    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP

    # Set row height for multi-line fund summary
    ws.row_dimensions[row].height = max(60, len(p['funds']) * 30)
    row += 1

widths = [6, 10, 9, 10, 12, 7, 22, 60, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)

print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print("\nPOSITION INTENSITY RANKING (top 15):")
print(f"{'#':<4} {'Ticker':<10} {'Score':<7} {'Sum %':<8} {'$M Skin':<8} {'#Fund':<6} {'P&L'}")
for rank, (ticker, p, score) in enumerate(ranked[:15], 1):
    sp = p['sum_pct'] if isinstance(p['sum_pct'], (int, float)) else 'N/A'
    print(f"  {rank:<3} {ticker:<10} {score:<7} {str(sp):<8} {p['dollar_skin']:<8} {len(p['funds']):<6} {p['in_money'][:30]}")
