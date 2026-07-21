"""Comprehensive tier classification of ALL high-signal tickers across 455 fund tabs.

After comprehensive scan filtering out mega-caps + tokens, ~280 unique tickers
have meaningful signal (>=3 tab mentions). Of these, ~50 are NEW (not in our
41-name asymmetric universe).

Manually classify NEW candidates into Tier 1/2/3 based on:
  - Multi-fund consensus + adds + activist signal
  - Variant perception (market wrong on something specific)
  - Entry-today asymmetry from current price
  - Catalyst velocity
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# NEW CANDIDATES — surfaced from comprehensive 455-tab scan
# Classified by variant perception + entry-today asymmetry
# ================================================================

NEW_TIER_1 = [
    {'ticker': 'WBD', 'name': 'Warner Bros Discovery', 'mcap': '$25B',
     'fund_signal': '10 tabs / 8 material adds / +70% max ADD',
     'variant': 'Market: linear TV death + $36B debt overhang = terminal. Variant: HBO content moat + Zaslav cost cuts $5B + debt paydown $19B->$11B = cash flow inflection priced as bankruptcy candidate.',
     'catalyst': 'HBO Max international scaling Q3 2026 + studio M&A speculation + linear TV split planned',
     'asym': '2.5-4x', 'downside': '-30%', 'rr': '8:1',
     'sec_tailwind': 'PARTIAL (streaming consolidation; pricing power inflection)',
     'in_money_status': 'NOT (still 60%+ below 2021 peak)',
     'smart_money': '8 funds adding (Caxton + Brevan + others); zaslav insider aligned'},

    {'ticker': 'CRH', 'name': 'CRH plc (Irish building materials)', 'mcap': '$70B',
     'fund_signal': '7 tabs / 5 HC / 3 material adds',
     'variant': 'Market: legacy materials cyclical. Variant: US data-center concrete demand = AI capex tailwind (SAME thesis as BZU.IM but US-listed = lower friction for institutional adds).',
     'catalyst': 'AI data-center cement orders + US infrastructure cycle + buyback expansion',
     'asym': '1.7-2.2x', 'downside': '-20%', 'rr': '8:1',
     'sec_tailwind': 'NOT (AI data-center concrete unpriced — same as BZU.IM but US-friction)',
     'in_money_status': 'PARTIAL (up from trough but still below true potential)',
     'smart_money': '5 HC + multi-fund consensus on materials secular'},

    {'ticker': 'PCG', 'name': 'PG&E', 'mcap': '$40B',
     'fund_signal': '8 tabs / 8 HC / 2 material adds / +83% max ADD',
     'variant': 'Market: California wildfire risk + bankruptcy stigma. Variant: AB 1054 wildfire fund passed + Diablo Canyon extension + AI data-center load growth in NorCal = utility re-rate to peer multiple.',
     'catalyst': 'AI/EV electricity demand growth + rate case approvals + dividend restoration H2 2026',
     'asym': '1.8-2.5x', 'downside': '-25%', 'rr': '7:1',
     'sec_tailwind': 'NOT (AI data-center electricity demand unpriced for CA utility)',
     'in_money_status': 'PARTIAL (recovered from $3 trough, still below pre-bankruptcy)',
     'smart_money': '8 HC consensus = multi-fund deep value anchor on distressed utility recovery'},

    {'ticker': 'MDGL', 'name': 'Madrigal Pharmaceuticals', 'mcap': '$7.5B',
     'fund_signal': '5 HC / 1 new init / +149% max ADD',
     'variant': 'Market: single-asset post-approval (resmetirom for NASH/MASH). Variant: First-mover advantage on $5-10B MASH TAM with addressable 2-3M patients; current penetration <5% = early innings.',
     'catalyst': 'Rezdiffra commercial ramp Q2-Q3 2026 + Phase 3 cirrhotic readout 2027 + EU launch',
     'asym': '2-3x', 'downside': '-30%', 'rr': '8:1',
     'sec_tailwind': 'NOT (MASH treatment paradigm shift)',
     'in_money_status': 'NOT (still below pre-data run-up)',
     'smart_money': '5 HC + biotech crossover specialists; +149% max ADD on conviction'},

    {'ticker': 'SHC', 'name': 'Sotera Health', 'mcap': '$3.5B',
     'fund_signal': '4 HC / 3 material adds / +9% max ADD (also held by Kerrisdale +9% Q1)',
     'variant': 'Market: ethylene oxide litigation + COTY-style asbestos overhang. Variant: Critical sterilization MOAT for medical devices ($4B addressable, no green substitute) + lawsuits settled = pure compounder.',
     'catalyst': 'Litigation overhang resolution + medical device sterilization volume growth',
     'asym': '2-3x', 'downside': '-25%', 'rr': '9:1',
     'sec_tailwind': 'NOT (medical device sterilization moat unpriced post-litigation)',
     'in_money_status': 'NOT (still below pre-litigation peak)',
     'smart_money': 'Multi-fund + Kerrisdale top-10 long thesis emerging'},
]

NEW_TIER_2 = [
    {'ticker': 'MU', 'name': 'Micron Technology', 'mcap': '$110B',
     'fund_signal': '5 HC / 4 material adds / +200% max ADD',
     'variant': 'Market: HBM oversupply by 2027. Variant: AI memory demand structural; Micron HBM3E qualified for NVDA Blackwell = supply-constrained pricing power through 2027.',
     'catalyst': 'HBM3E ramp + DDR5 pricing + AI memory share gains', 'asym': '1.5-2x',
     'downside': '-25%', 'rr': '5:1', 'sec_tailwind': 'PARTIAL (AI memory partly priced)',
     'in_money_status': 'PARTIAL', 'smart_money': '5 HC consensus + cycle bull'},

    {'ticker': 'CLBT', 'name': 'Cellebrite DI', 'mcap': '$3.5B',
     'fund_signal': '4 HC / 3 material adds / +340% max ADD (massive)',
     'variant': 'Market: niche digital forensics. Variant: AI-enabled law enforcement evidence platform; mission-critical software with 90%+ gross margin = SaaS quality at unrecognized multiple.',
     'catalyst': 'Federal AI procurement budget + UFED Cloud growth', 'asym': '2-3x',
     'downside': '-30%', 'rr': '7:1', 'sec_tailwind': 'NOT (AI law enforcement spend secular)',
     'in_money_status': 'NOT', 'smart_money': '4 HC + 340% max ADD = singular conviction concentration'},

    {'ticker': 'KVUE', 'name': 'Kenvue (J&J consumer spinoff)', 'mcap': '$45B',
     'fund_signal': '3 HC / 3 new init / +169% max ADD',
     'variant': 'Market: spinoff orphan + Tylenol/autism litigation noise. Variant: Tylenol/Listerine/Band-Aid brand portfolio at sub-12x P/E with category leadership; lawsuits dismissed.',
     'catalyst': 'Q2 2026 brand growth proof + litigation closure + dividend hike',
     'asym': '1.6-2x', 'downside': '-20%', 'rr': '6:1',
     'sec_tailwind': 'PARTIAL (consumer staples re-rate)',
     'in_money_status': 'NOT (still below spin)', 'smart_money': '3 NEW positions Q1 2026 + 169% max = recent multi-fund accumulation'},

    {'ticker': 'CPNG', 'name': 'Coupang (Korean e-commerce)', 'mcap': '$50B',
     'fund_signal': '6 HC / 5 material adds / +66% max ADD',
     'variant': 'Market: Korean consumer weakness + Eats losses. Variant: Korean Prime monopoly + Eats break-even + AI logistics moat = post-MaaS profitability inflection.',
     'catalyst': 'Q2 2026 op margin proof + Taiwan expansion + Eats break-even',
     'asym': '1.8-2.2x', 'downside': '-25%', 'rr': '7:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL',
     'smart_money': 'Duquesne + Eminence + Tiger Global consensus'},

    {'ticker': 'BLDR', 'name': 'Builders FirstSource', 'mcap': '$14B',
     'fund_signal': '5 HC / 3 material adds / +597% MAX ADD',
     'variant': 'Market: housing slowdown cyclical. Variant: structural undersupply 4M+ homes + scale advantages over fragmented competitors + buyback math.',
     'catalyst': 'Mortgage rate cuts + housing starts inflection + buyback compounding',
     'asym': '2-3x', 'downside': '-30%', 'rr': '7:1',
     'sec_tailwind': 'NOT (US housing supply shortage unpriced — peer to MRP)',
     'in_money_status': 'NOT (below 2022 peak)',
     'smart_money': '5 HC + +597% max ADD by one fund = max conviction housing cycle bet'},

    {'ticker': 'SDRL', 'name': 'Seadrill', 'mcap': '$2.5B',
     'fund_signal': '1 13D / 3 material adds / +597% MAX ADD',
     'variant': 'Market: declining oil demand peak. Variant: offshore drilling rig shortage post-bankruptcies; day rates +60% off-cycle; 7th-gen drillships scarce.',
     'catalyst': 'Day rate spike + tender awards + Petrobras contracts H2 2026',
     'asym': '2-3x', 'downside': '-40%', 'rr': '6:1',
     'sec_tailwind': 'PARTIAL (deepwater supply gap)',
     'in_money_status': 'NOT (post-restructuring)', 'smart_money': '1 13D + +597% max = singular activist conviction'},

    {'ticker': 'MNKTQ', 'name': 'Mallinckrodt (post-bankruptcy)', 'mcap': '<$200M',
     'fund_signal': '3 13D / 3 material adds (post-BK emergence)',
     'variant': 'Market: opioid bankruptcy stigma + specialty pharma decline. Variant: post-BK clean balance sheet + Acthar gel + 3 distressed-debt funds anchored.',
     'catalyst': 'Post-BK equity rerate + specialty pharma growth + Acthar Medicare path',
     'asym': '2-4x', 'downside': '-40%', 'rr': '6:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT (post-BK trough)',
     'smart_money': 'Hudson Bay + Marathon + 1 more = distressed-debt fund convergence'},

    {'ticker': 'AERO', 'name': 'Grupo Aeromexico (post-BK)', 'mcap': '~$2B',
     'fund_signal': '1 13D / 2 material adds / 4 new positions',
     'variant': 'Market: post-BK Mexican carrier discount. Variant: Silver Point + SVPGlobal + Baupost all NEW = distressed value experts confirming thesis post-emergence.',
     'catalyst': 'Mexican aviation cycle + tourism rebound + JV with Delta',
     'asym': '1.7-2.5x', 'downside': '-30%', 'rr': '6:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT (post-BK)',
     'smart_money': 'Silver Point 9.0% 13G + SVPGlobal NEW + Baupost ~2% = three top distressed funds'},

    {'ticker': 'APG', 'name': 'API Group', 'mcap': '$11B',
     'fund_signal': '7 HC / 2 material adds / +48% max',
     'variant': 'Market: industrial services low-growth. Variant: Chubb Fire & Security acquisition synergy + recurring service revenue 70% + cross-sell.',
     'catalyst': 'Chubb integration + M&A platform growth',
     'asym': '1.8-2.2x', 'downside': '-20%', 'rr': '7:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL',
     'smart_money': '7 HC = institutional consensus on roll-up'},

    {'ticker': 'APP', 'name': 'AppLovin', 'mcap': '$110B (mega cap-ish)',
     'fund_signal': '5 HC / 4 material adds',
     'variant': 'Market: ad-tech roadkill peer (CRTO/DV). Variant: AXON AI ad targeting platform inflection + gaming installs = winner of post-ATT mobile ad recovery.',
     'catalyst': 'AXON expansion + AI ad targeting moat',
     'asym': '1.4-1.8x (high mcap dilutes asymmetry)', 'downside': '-25%', 'rr': '5:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL (large run-up)',
     'smart_money': '5 HC consensus'},
]

NEW_TIER_3 = [
    {'ticker': 'CRWV', 'name': 'CoreWeave', 'mcap': '$25B',
     'fund_signal': '4 HC / 2 material adds / 1 13D / +40% max',
     'variant': 'AI infra GPU-as-a-service; bubble valuation but moat real',
     'catalyst': 'Hyperscaler contracts + NVDA capacity allocation',
     'asym': '1.5-2x', 'downside': '-50%', 'rr': '3:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL'},

    {'ticker': 'SATS', 'name': 'EchoStar (Ergen)', 'mcap': '$3B',
     'fund_signal': '6 HC / 2 material adds / 2 new positions',
     'variant': 'Ergen-controlled spectrum + DISH merger optionality',
     'catalyst': 'DISH transaction + spectrum monetization',
     'asym': '2-4x', 'downside': '-50%', 'rr': '5:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT'},

    {'ticker': 'RVMD', 'name': 'Revolution Medicines', 'mcap': '$8B',
     'fund_signal': '5 HC / 1 new init',
     'variant': 'KRAS-mutant oncology multi-asset platform',
     'catalyst': 'RMC-6236 Phase 3 readouts 2027',
     'asym': '2-3x', 'downside': '-40%', 'rr': '5:1',
     'sec_tailwind': 'NOT (KRAS unpriced)', 'in_money_status': 'NOT'},

    {'ticker': 'MSTR', 'name': 'MicroStrategy', 'mcap': '$60B',
     'fund_signal': '5 HC / 3 material adds',
     'variant': 'Bitcoin treasury proxy; convertible structure',
     'catalyst': 'Bitcoin price + dilution mechanics', 'asym': '1.5-3x',
     'downside': '-50%', 'rr': '3:1', 'sec_tailwind': 'PARTIAL',
     'in_money_status': 'PARTIAL (BTC dependent)'},

    {'ticker': 'UNF', 'name': 'UniFirst', 'mcap': '$3B',
     'fund_signal': '1 HC / 1 13D / 2 material adds / +640% MAX ADD',
     'variant': 'Uniform rental cyclical; Cintas precedent at 25x; activist setup',
     'catalyst': 'PE bid or activist board pressure',
     'asym': '1.5-2.5x', 'downside': '-25%', 'rr': '6:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT'},

    {'ticker': 'DASH', 'name': 'DoorDash', 'mcap': '$70B',
     'fund_signal': '3 HC / 3 material adds / 2 new / +236% max ADD',
     'variant': 'Cycle through cohort math; profitable inflection',
     'catalyst': 'Operating leverage Q2-Q3 + grocery expansion',
     'asym': '1.5-2x', 'downside': '-25%', 'rr': '5:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL'},

    {'ticker': 'CNC', 'name': 'Centene', 'mcap': '$30B',
     'fund_signal': '3 HC / 3 material adds / 2 new / +69% max',
     'variant': 'Medicare Advantage rate cuts overhang; Medicaid redeterminations',
     'catalyst': '2026 bid cycle clarity + state Medicaid wins',
     'asym': '1.7-2.2x', 'downside': '-30%', 'rr': '6:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT'},

    {'ticker': 'COIN', 'name': 'Coinbase', 'mcap': '$60B',
     'fund_signal': '3 HC / 2 new init / +112% max',
     'variant': 'Crypto trading rebound + regulatory clarity',
     'catalyst': 'Spot ETF growth + Base L2 + staking yield',
     'asym': '1.5-2.5x', 'downside': '-40%', 'rr': '4:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL'},

    {'ticker': 'EXAS', 'name': 'Exact Sciences', 'mcap': '$15B',
     'fund_signal': 'Already mentioned earlier — 3 funds +157% max ADD',
     'variant': 'Cologuard moat + MCED expansion',
     'catalyst': 'CMS coverage + MCED launch',
     'asym': '1.7-2.2x', 'downside': '-30%', 'rr': '5:1',
     'sec_tailwind': 'NONE', 'in_money_status': 'NOT'},

    {'ticker': 'CHYM', 'name': 'Chime Financial', 'mcap': '~$15B (post-IPO)',
     'fund_signal': '3 HC / 4 new positions',
     'variant': 'Post-IPO neobank consolidation; growth thesis',
     'catalyst': 'Profitability inflection + customer add growth',
     'asym': '1.5-2.5x', 'downside': '-40%', 'rr': '4:1',
     'sec_tailwind': 'PARTIAL', 'in_money_status': 'PARTIAL'},
]


# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'FULL UNIVERSE NEW Tiers' in wb.sheetnames:
    del wb['FULL UNIVERSE NEW Tiers']

ws = wb.create_sheet('FULL UNIVERSE NEW Tiers', 6)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
T1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
T2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
T3_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:K1')
ws['A1'] = 'FULL UNIVERSE TIER CLASSIFICATION — NEW asymmetric candidates from 455-tab scan'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:K2')
ws['A2'] = ('Scanned all 445 fund tabs; found 288 tickers with >=3 tab mentions ex mega-caps. '
            'Of these, ~50 NOT in our current 41-name asymmetric universe. Top 25 classified into '
            'TIER 1/2/3 by variant perception + entry-today asymmetry + smart money signal.')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

ws.merge_cells('A4:C4')
ws['A4'] = 'TIER LEGEND'; ws['A4'].font = HEADER; ws['A4'].fill = HEADER_FILL
ws['A5'] = 'TIER 1'; ws['A5'].fill = T1_FILL; ws['A5'].font = Font(bold=True)
ws['B5'] = 'Multi-fund consensus + clear variant perception + entry-today asymmetry intact'
ws['A6'] = 'TIER 2'; ws['A6'].fill = T2_FILL; ws['A6'].font = Font(bold=True)
ws['B6'] = 'Strong setup with deep value OR binary catalyst OR fund concentration'
ws['A7'] = 'TIER 3'; ws['A7'].fill = T3_FILL; ws['A7'].font = Font(bold=True)
ws['B7'] = 'Speculative / partial signal / worth tracking'

headers = ['Tier', 'Ticker', 'Name', 'Mcap', 'Fund Signal', 'Variant Perception', 'Catalyst', 'Asym', 'Downside', 'R/R', 'Tailwind']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=9, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

row = 10
for tier_name, tier_data, fill in [('1', NEW_TIER_1, T1_FILL), ('2', NEW_TIER_2, T2_FILL), ('3', NEW_TIER_3, T3_FILL)]:
    for n in tier_data:
        ws.cell(row=row, column=1, value=f"TIER {tier_name}").fill = fill
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
        ws.cell(row=row, column=3, value=n['name'])
        ws.cell(row=row, column=4, value=n['mcap'])
        ws.cell(row=row, column=5, value=n['fund_signal'])
        ws.cell(row=row, column=6, value=n['variant'])
        ws.cell(row=row, column=7, value=n['catalyst'])
        ws.cell(row=row, column=8, value=n['asym'])
        ws.cell(row=row, column=9, value=n['downside'])
        ws.cell(row=row, column=10, value=n['rr'])
        ws.cell(row=row, column=11, value=n.get('sec_tailwind', ''))
        for col in range(1, 12):
            ws.cell(row=row, column=col).alignment = LEFT_WRAP
        ws.row_dimensions[row].height = 80
        row += 1

widths = [8, 10, 30, 9, 30, 70, 40, 12, 10, 7, 25]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)

print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"\nNEW TIER 1 (most asymmetric variant perception, multi-fund):")
for n in NEW_TIER_1:
    print(f"  {n['ticker']:<6} {n['mcap']:<8} {n['asym']:<10} | {n['variant'][:80]}")
print(f"\nNEW TIER 2 (strong setup with one signal dominant):")
for n in NEW_TIER_2:
    print(f"  {n['ticker']:<6} {n['mcap']:<8} {n['asym']:<10} | {n['variant'][:75]}")
print(f"\nNEW TIER 3 (speculative / partial):")
for n in NEW_TIER_3:
    print(f"  {n['ticker']:<6} {n['mcap']:<8} {n['asym']:<10}")
