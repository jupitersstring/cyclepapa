"""FORENSIC VERIFICATION + GAP SCAN

Don't trust memory. For each fund tab in source workbook, extract:
  - Section (1) Highest Conviction with % portfolio and Q-o-Q change
  - Section (2) 13D/13G threshold filings with dates
  - Section (3) New positions with % at init
  - Section (4) Material adds with prior/new %

Build comprehensive ticker → activity map.
Verify every Forensic Tier 1 claim against source rows.
Surface any HIGH-SIGNAL names I haven't elevated to Tier 1.
"""
import openpyxl
import re
from collections import defaultdict

wb = openpyxl.load_workbook('/home/user/cyclepapa/fund_activity_last_6mo.xlsx', data_only=True)

SKIP = {'Cover', 'Index', 'All Activity', 'Asymmetric Summary', 'Consensus Buys',
        'Highest Conviction', 'Conviction Adds', 'Micro-Cap Conviction Adds',
        'Activist Catalysts', 'Multi-Fund New Inits'}

# For each fund tab, extract structured data
fund_activity = {}  # fund_name -> {tickers_HC, tickers_13D, tickers_new, tickers_add, ...}

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP or sheet_name.startswith('Sheet'):
        continue
    ws = wb[sheet_name]

    activity = {
        'sheet': sheet_name,
        'HC': [],       # (row_text, tickers found)
        '13D': [],
        'new': [],
        'adds': [],
        'form4': [],
    }
    section = 0  # 1=HC, 2=13D, 3=new, 4=adds

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        text = ' '.join(str(c) for c in row if c is not None).strip()
        if not text:
            continue
        tl = text.lower()

        # Section detection
        if 'highest conviction' in tl and ('(1)' in text or '(1' in tl):
            section = 1
            continue
        if ('threshold' in tl or 'disclosure' in tl or '>=5%' in tl) and '(2' in tl:
            section = 2
            continue
        if 'new position' in tl and '(3' in tl:
            section = 3
            continue
        if ('material' in tl or 'existing positions' in tl) and '(4' in tl:
            section = 4
            continue

        # Extract tickers via patterns: capital letters 2-6 chars, optional .XX
        # Be strict: must be standalone or after $
        tickers = re.findall(r'(?:^|\s|\$|\()([A-Z]{2,6}(?:\.[A-Z]{1,3})?)\b', text)
        # Filter common non-tickers
        block = {'CEO','CFO','COO','CTO','LLC','INC','LP','GP','LTD','PLC','GMBH','LLP','LBP',
                 'Q1','Q2','Q3','Q4','FY','TTM','YTD','SEC','FDA','PDUFA','NDA','IND','HBO',
                 'AI','EU','US','UK','HK','JV','PT','MD','NX','EBITDA','FCF','EPS','EV',
                 'NYSE','NASDAQ','TSX','LSE','MAX','MIN','ADD','NEW','HIGH','LOW','NOT',
                 'III','II','IV','VI','VII','HC','TT','GS','MS','BS','PE','VC','BBB','AAA',
                 'TBD','TBA','TBA','BC','AD','OK','NA','SOTP','REIT','SPAC','RAUM','AUM',
                 'BUY','SELL','HOLD','LONG','SHORT','OWNS','HELD','OPEN','MARKET',
                 'FORM','TYPE','DATE','NOTE','REF','PER','BY','FOR','THE','WITH','AND',
                 'OR','AT','TO','OF','IN','ON','IS','AS','IT','BE','HAS','WAS','ARE','OFF',
                 'ALL','OUT','MAIN','BIG','TOP'}
        tickers = [t for t in tickers if t not in block and not t.isdigit() and len(t) >= 2]

        if not tickers:
            continue

        # Form 4 / insider detection
        if any(p in tl for p in ['form 4', 'open-market', 'open market buy', 'insider buy',
                                  'ceo buy', 'cfo buy', 'founder buy', 'director buy',
                                  'bought 800k', 'bought $']):
            activity['form4'].append({'text': text[:200], 'tickers': tickers})

        # Capture by section
        if section == 1:
            activity['HC'].append({'text': text[:200], 'tickers': tickers})
        elif section == 2:
            activity['13D'].append({'text': text[:200], 'tickers': tickers})
        elif section == 3:
            activity['new'].append({'text': text[:200], 'tickers': tickers})
        elif section == 4:
            activity['adds'].append({'text': text[:200], 'tickers': tickers})

    fund_activity[sheet_name] = activity


# Now build ticker → activity reverse map
ticker_to_signals = defaultdict(lambda: {
    'HC_funds': [],
    '13D_funds': [],
    'new_funds': [],
    'add_funds': [],
    'form4_funds': [],
})

for fund, act in fund_activity.items():
    for entry in act['HC']:
        for t in entry['tickers']:
            ticker_to_signals[t]['HC_funds'].append((fund, entry['text']))
    for entry in act['13D']:
        for t in entry['tickers']:
            ticker_to_signals[t]['13D_funds'].append((fund, entry['text']))
    for entry in act['new']:
        for t in entry['tickers']:
            ticker_to_signals[t]['new_funds'].append((fund, entry['text']))
    for entry in act['adds']:
        for t in entry['tickers']:
            ticker_to_signals[t]['add_funds'].append((fund, entry['text']))
    for entry in act['form4']:
        for t in entry['tickers']:
            ticker_to_signals[t]['form4_funds'].append((fund, entry['text']))


# Forensic verification of my Tier 1 list
FORENSIC_T1 = ['PRLD', 'HHH', 'INMD', 'OPRX', 'KBR', 'LQDA', 'CTMX', 'NRP', 'MRP', 'WBD',
               'SHC', 'MNKTQ', 'CPNG', 'PCG', 'AERO']

print("=" * 80)
print("FORENSIC VERIFICATION — Each TIER 1 name's source-workbook evidence")
print("=" * 80)
for t in FORENSIC_T1:
    s = ticker_to_signals[t]
    print(f"\n{t}:")
    print(f"  HC mentions: {len(s['HC_funds'])} | 13D: {len(s['13D_funds'])} | new: {len(s['new_funds'])} | adds: {len(s['add_funds'])} | form4: {len(s['form4_funds'])}")

    # Show most recent / interesting evidence
    if s['HC_funds']:
        print("  Sample HC:")
        for fund, text in s['HC_funds'][:2]:
            print(f"    [{fund}] {text[:150]}")
    if s['13D_funds']:
        print("  Sample 13D:")
        for fund, text in s['13D_funds'][:2]:
            print(f"    [{fund}] {text[:150]}")
    if s['add_funds']:
        print("  Sample ADDS:")
        for fund, text in s['add_funds'][:3]:
            print(f"    [{fund}] {text[:150]}")
    if s['form4_funds']:
        print("  Sample Form4:")
        for fund, text in s['form4_funds'][:2]:
            print(f"    [{fund}] {text[:150]}")
