"""Aggregate all 6 forensic agent findings into master TIER classification.

Total NEW candidates from 445-tab scan: 70 names
Tier 1 from agents: ~25 names
Tier 2: ~30 names
Tier 3: ~15 names

Build single sheet ranked by asymmetric entry-today framework with full
attribution to source agent + fund + signal.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# CONSOLIDATED FINDINGS FROM 6 PARALLEL FORENSIC AGENTS
# Each entry: ticker, name, mcap, fund (concentration), signal,
#             catalyst, variant, downside floor, tier, source_chunk
# ================================================================

FORENSIC_TIER_1 = [
    # ====== CHUNK 1 ======
    {'ticker': 'TLN', 'name': 'Talen Energy', 'mcap': '~$8B',
     'fund': 'Sachem Head Capital', 'signal': '+72% Q4 add',
     'catalyst': 'May earnings + nuclear PSAs closing (weeks)',
     'variant': 'AI nuclear energy demand unpriced vs 5yr PPA modeling',
     'floor': 'Hard assets + contractual PSA floors', 'chunk': 1},

    {'ticker': 'BTSG', 'name': 'BrightSpring Health', 'mcap': '~$3B',
     'fund': 'Alta Fox Capital', 'signal': '3%→10% of book (+169%)',
     'catalyst': 'May earnings (21 days)',
     'variant': 'Senior living occupancy 20% below pre-COVID multiples',
     'floor': 'Real estate NOI + 65% occupancy minimum', 'chunk': 1},

    {'ticker': 'EQIX', 'name': 'Equinix', 'mcap': '$75B',
     'fund': 'Land & Buildings Investment Mgmt', 'signal': '+82% Q4 (largest add)',
     'catalyst': 'May earnings + AI hyperscaler capex inflection',
     'variant': 'Infrastructure cycle modeled CYCLICAL vs STRUCTURAL secular',
     'floor': '10-15yr hyperscaler leases + SLAs', 'chunk': 1},

    {'ticker': 'CSR', 'name': 'Centerspace', 'mcap': '~$1B',
     'fund': 'Land & Buildings Investment Mgmt', 'signal': '+39% Q4 (+229k sh) 13D',
     'catalyst': 'May earnings (21 days)',
     'variant': 'Sunbelt apartment REIT valuation vs Sunbelt fundamentals',
     'floor': '70% contractual revenue + real estate', 'chunk': 1},

    {'ticker': 'COO', 'name': 'Cooper Companies', 'mcap': '$20B',
     'fund': 'JANA Partners', 'signal': '+45% Q3->Q4',
     'catalyst': 'May earnings + FDA approvals (21 days)',
     'variant': 'Medical device margin inflection misread',
     'floor': '28% FCF yield + device durability', 'chunk': 1},

    # ====== CHUNK 2 ======
    {'ticker': 'FLR', 'name': 'Fluor Corp', 'mcap': '$8B',
     'fund': 'Greenlight Capital (Einhorn)', 'signal': '7.73% of book +44% Q4',
     'catalyst': 'Federal/infrastructure deals execution',
     'variant': 'Einhorn sees transformation catalyst; market misprices execution risk',
     'floor': 'Engineering backlog + margin recovery', 'chunk': 2},

    {'ticker': 'WDC', 'name': 'Western Digital', 'mcap': '$24B',
     'fund': 'Whitebox Advisors', 'signal': '13.3% of portfolio +$154M Q4',
     'catalyst': 'Q1/Q2 2026 earnings + HDD AI datacenter stabilization',
     'variant': 'Activist accumulation vs market skepticism on AI HDD demand cycle',
     'floor': 'Enterprise SSD/HDD TAM recovery + NAND floor', 'chunk': 2},

    # ====== CHUNK 3 ======
    {'ticker': 'RPAY', 'name': 'Repay Holdings', 'mcap': '~$130M',
     'fund': 'Veradace Partners', 'signal': '8.6% (7.36M sh + 110k calls) 13D activist',
     'catalyst': '13D NOMINATION WINDOW opens early June 2026 (60 days from 4/15/26)',
     'variant': 'Small-cap payment processor undervalued vs GTECH/FISV',
     'floor': 'Cash 30-40% mcap + activist 8.6% + calls = voting power', 'chunk': 3},

    {'ticker': 'SSTI', 'name': 'SoundThinking', 'mcap': '~$600M',
     'fund': 'Veradace Partners', 'signal': '16.5% ($90M, 2.09M sh) 13D conversion',
     'catalyst': '13D NOMINATION WINDOW early June 2026 (post-3/10)',
     'variant': 'Gunshot detection monopoly + public safety trend underpriced',
     'floor': 'Cash ~50% mcap + 16.5% activist concentration', 'chunk': 3},

    {'ticker': 'SYBX', 'name': 'Synlogic', 'mcap': '~$50M',
     'fund': 'Funicular/Cable Car + Radoff', 'signal': '28.3% (3.3M sh) + 4.4% = 32.7% combined',
     'catalyst': 'Board response window 60d from 4/6/26 13D/A — into May/June 2026',
     'variant': 'Synthetic biology microbiome tech; activist proposed $0.64 buyout',
     'floor': '$0.64 NON-BINDING BUYOUT OFFER = HARD PRICE FLOOR', 'chunk': 3},

    # ====== CHUNK 5 — RA Capital biotech cluster ======
    {'ticker': 'CLYM', 'name': 'Climb Bio', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '33.0% portfolio / $20M PIPE anchor',
     'catalyst': 'Q2 2026 PIPE closing (weeks)',
     'variant': 'RA Cap anchored $20M PIPE; post-close runaway upside from $0 baseline',
     'floor': 'PIPE capital committed; post-close equity floor', 'chunk': 5},

    {'ticker': 'ACRV', 'name': 'Acrivon Therapeutics', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '28.8% portfolio; +3.89M sh @ $1.80 April 2026',
     'catalyst': 'Imminent trial readouts on cancer programs Q2-Q3 2026',
     'variant': 'RA Cap added $7M at $1.80 entry; market repricing delayed',
     'floor': 'Cash + trial value; post-BO-level $1 risk', 'chunk': 5},

    {'ticker': 'VOR', 'name': 'Vor Biopharma', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '19.9% portfolio; warrant cap raise Mar 2026',
     'catalyst': 'Q2 2026 trial data + donor availability inflection',
     'variant': 'Off-the-shelf engineered cells de-risked platform; market prices single-program',
     'floor': 'Cap raise completed + cash runway + post-data floor', 'chunk': 5},

    {'ticker': 'PEPG', 'name': 'PepGen', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '29.2% portfolio; +9.375M @ $3.20 Q4 = $30M PIPE',
     'catalyst': 'Q2-Q3 2026 trial data (genetically modified antibodies)',
     'variant': 'RA Cap first-mover PIPE @ $3.20; rare platform asset not yet de-risked',
     'floor': 'Post-PIPE cash floor + biotech recovery', 'chunk': 5},

    {'ticker': 'KPTI', 'name': 'Karyopharm Therapeutics', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '9.99% portfolio; $30M PIPE anchor Q1 2026',
     'catalyst': 'Q2-Q3 2026 pipeline readouts + partner activations',
     'variant': 'RA Cap stepped in as capital provider; market processing post-legacy-product transition',
     'floor': 'PIPE capital + pipeline optionality', 'chunk': 5},

    {'ticker': 'COAG', 'name': 'Hemab Therapeutics', 'mcap': 'micro',
     'fund': 'RA Capital Management', 'signal': '16.6% portfolio; 13D filed May 11, 2026',
     'catalyst': 'Q2-Q3 2026 IND-enabling work + capital close',
     'variant': 'Pre-clinical→IND; private biotech consolidation/JV signal via 13D',
     'floor': 'RA Cap 16.6% puts floor at deployment cost', 'chunk': 5},

    # ====== CHUNK 6 ======
    {'ticker': 'COR', 'name': 'Cencora (formerly AmerisourceBergen)', 'mcap': '$50B',
     'fund': 'Sio Capital', 'signal': '+535% MASSIVE Q1 2026 ADD (4.55% portfolio)',
     'catalyst': 'Q2 2026 earnings (weeks) + pharma pricing stability',
     'variant': 'Market OVERCOUNTS opioid settlement tail risk; core distribution oligopoly resilient',
     'floor': 'Pharma distribution oligopoly + healthcare baseline', 'chunk': 6},

    {'ticker': 'WAY', 'name': 'Waystar', 'mcap': '$7B',
     'fund': 'Sio Capital', 'signal': '+168% MATERIAL Q1 2026 ADD (5.68% portfolio)',
     'catalyst': 'Q2 earnings (weeks); cloud RCM adoption acceleration',
     'variant': 'SaaS margin expansion invisible to legacy earnings model',
     'floor': 'SaaS cohort multiples + sticky workflows', 'chunk': 6},

    {'ticker': 'ICLR', 'name': 'ICON plc', 'mcap': '$15B',
     'fund': 'Sio Capital', 'signal': '+141% MATERIAL Q1 2026 ADD (4.86% portfolio)',
     'catalyst': 'Q2 2026 earnings; biotech spend recovery post-correction',
     'variant': 'CRO valuations reset; outsourcing rebound overlooked',
     'floor': 'Biotech spend cycle + outsourcing secular trend', 'chunk': 6},

    {'ticker': 'EXPE', 'name': 'Expedia Group', 'mcap': '$25B',
     'fund': 'WindAcre Partnership', 'signal': '8.5% NEW entire position Q1 2026',
     'catalyst': 'Q2 earnings + summer travel demand inflection (weeks-June)',
     'variant': 'Travel structural recovery + OTA margin underweighted',
     'floor': 'Travel demand floor (pent-up leisure) + M&A strategic value', 'chunk': 6},

    {'ticker': 'BW', 'name': 'Babcock & Wilcox', 'mcap': '$300M',
     'fund': 'Steel Partners (13D activist)', 'signal': '13D NEW activist June 2025',
     'catalyst': '13D nomination June-July 2026 (weeks); nuclear SMR capex Q2-Q3',
     'variant': 'Nuclear renaissance structural; activist cost reduction opaque to market',
     'floor': 'Asset liquidation (SMR IP, parts spinoff) + activist break-up optionality', 'chunk': 6},

    {'ticker': 'ALC', 'name': 'Alcon', 'mcap': '$45B',
     'fund': 'AKO Capital', 'signal': '13.62% portfolio; +21.31% Q1 2026 ADD',
     'catalyst': 'Q2 2026 earnings (weeks) + FDA approval cycle',
     'variant': 'Eye care growth recovery underpriced vs FDA approval cycle',
     'floor': 'Healthcare device floor + institutional base', 'chunk': 6},

    {'ticker': 'ASML', 'name': 'ASML Holding', 'mcap': '$280B',
     'fund': 'Valley Forge Capital (Kantesaria)', 'signal': '6.69% portfolio; +35.5% MAJOR Q1 ADD',
     'catalyst': 'Q2 earnings + EUV fab ramp (weeks-months)',
     'variant': 'EUV monopoly massively underweighted vs AI capex inflection',
     'floor': 'Tech oligopoly pricing + capex commitment floor', 'chunk': 6},

    {'ticker': 'MYO', 'name': 'Myomo Inc', 'mcap': '<$100M',
     'fund': 'Horton Capital', 'signal': '6.0% portfolio; 13D NEW activist Dec 2025',
     'catalyst': '13D campaign through June 2026 (weeks); FDA robotics pending',
     'variant': 'Tiny float + adoption curve steep; market pricing conservative',
     'floor': 'Medical device floor + FDA approval optionality', 'chunk': 6},
]

FORENSIC_TIER_2 = [
    # CHUNK 1
    {'ticker': 'BL', 'name': 'BlackLine', 'fund': 'Engaged Capital', 'signal': '+138.79% Q4'},
    {'ticker': 'FISV', 'name': 'Fiserv', 'fund': 'JANA Partners', 'signal': 'NEW Q1'},
    {'ticker': 'GDS', 'name': 'GDS Holdings', 'fund': 'Sachem Head', 'signal': '+44% Q4'},
    {'ticker': 'CARG', 'name': 'CarGurus', 'fund': 'Alta Fox', 'signal': '+16% Q4'},
    {'ticker': 'DKNG', 'name': 'DraftKings', 'fund': 'Eminence Capital', 'signal': '+34% Q4'},
    {'ticker': 'REX', 'name': 'REX American Resources', 'fund': 'Mangrove Partners', 'signal': '+93% Q4'},
    {'ticker': 'FLYW', 'name': 'Flywire', 'fund': 'Mangrove + Voss', 'signal': 'Multi-fund'},
    {'ticker': 'ITGR', 'name': 'Integer Holdings', 'fund': 'Irenic Capital', 'signal': 'NEW activist'},
    # CHUNK 2
    {'ticker': 'BLK', 'name': 'BlackRock', 'fund': 'ValueAct Capital', 'signal': '10.59% NEW $744M Q4'},
    {'ticker': 'CMCSA', 'name': 'Comcast', 'fund': 'GoldenTree Asset Mgmt', 'signal': '12.7% +284% MAJOR ADD'},
    {'ticker': 'COLD', 'name': 'Americold Realty Trust', 'fund': 'Ancora Advisors', 'signal': '3.1% NEW Q4'},
    {'ticker': 'CPRI', 'name': 'Capri Holdings', 'fund': 'Greenlight (Einhorn)', 'signal': '4.08% +77% Q4'},
    {'ticker': 'CROX', 'name': 'Crocs Inc', 'fund': 'Himalaya Capital (Li Lu)', 'signal': '1.51% NEW Q4'},
    {'ticker': 'IMNM', 'name': 'ImmunoME', 'fund': 'Redmile Group', 'signal': '9.05% +13% recent'},
    {'ticker': 'MLTX', 'name': 'MoonLake Immunotherapeutics', 'fund': 'BVF Partners', 'signal': '8.8% +20M sh 13D/A'},
    {'ticker': 'S', 'name': 'SentinelOne', 'fund': 'Anchorage Capital', 'signal': '83.6% of portfolio +4.9M sh'},
    # CHUNK 3
    {'ticker': 'TIPT', 'name': 'Tiptree', 'fund': 'Veradace Partners', 'signal': '5.0% + 489k calls — blocked Fortegra-DB'},
    {'ticker': 'CCEL', 'name': 'Cryo-Cell Intl', 'fund': 'Camac Partners (Shahinian)', 'signal': '5.3% NEW 13D 3/25/26'},
    {'ticker': 'AIRI', 'name': 'Air Industries Group', 'fund': 'Charles L. Frischer', 'signal': '9.4% NEW 13D 3/23/26'},
    # CHUNK 4
    {'ticker': 'AMRZ', 'name': 'Amrize (Holcim spin)', 'fund': 'FPA Crescent (Romick)', 'signal': 'Spin received Q1'},
    {'ticker': 'AZE.BB', 'name': 'Azelis NV (European chems distrib)', 'fund': 'FPA Crescent (Romick)', 'signal': 'Added recent Q'},
    {'ticker': 'NOV', 'name': 'National Oilwell Varco', 'fund': 'FPA Crescent (Romick)', 'signal': 'NEW Q1 2026 — AI energy thesis'},
    {'ticker': 'DCH', 'name': 'Dauch Corp', 'fund': 'Miller Value (Bill Miller III)', 'signal': '3.92% NEW 13F 2/17/26'},
    {'ticker': 'ITRN', 'name': 'Ituran Location and Control (Israel)', 'fund': 'Miller Value', 'signal': '4.1% top-6'},
    # CHUNK 5
    {'ticker': 'WVE', 'name': 'Wave Life Sciences', 'fund': 'RA Capital Mgmt', 'signal': '17.1-17.5% 13D April 2026'},
    {'ticker': 'ACET', 'name': 'Adicet Bio', 'fund': 'RA Capital Mgmt', 'signal': '12.3% 13D April 2026'},
    {'ticker': 'Q', 'name': 'Qnity Electronics', 'fund': 'Durable Capital Partners', 'signal': '6.04% +82% MAJOR BUILD Q1'},
    {'ticker': 'ANDG', 'name': 'Andersen Group', 'fund': 'Durable Capital Partners', 'signal': '13.20% 13G NEW Dec 2025'},
    {'ticker': 'FERG', 'name': 'Ferguson Enterprises', 'fund': 'Durable Capital Partners', 'signal': '5.30% +40% Q1'},
    {'ticker': 'ARM', 'name': 'ARM Holdings', 'fund': 'Altimeter Capital (Gerstner)', 'signal': '4.55% NEW Q1 = $260M'},
    # CHUNK 6
    {'ticker': 'MCO', 'name': 'Moody\'s', 'fund': 'TCI Fund (Hohn)', 'signal': '13.84% +7.7% ADD Q1'},
    {'ticker': 'SPGI', 'name': 'S&P Global', 'fund': 'TCI Fund (Hohn)', 'signal': '13.22% +19% MATERIAL ADD Q1'},
    {'ticker': 'CNI', 'name': 'Canadian National Railway', 'fund': 'TCI Fund (Hohn)', 'signal': '2.24% — ACTIVIST 13D CEO REPLACEMENT CAMPAIGN'},
    {'ticker': 'CODI', 'name': 'Compass Diversified', 'fund': 'ADW Capital', 'signal': '13D escalation +71% to +140% Feb-Apr 2026'},
    {'ticker': 'PRM', 'name': 'Perimeter Solutions', 'fund': 'WindAcre Partnership', 'signal': '5.7% + Form 4 CEO/officer buys $3.56M Jul 2025'},
]

FORENSIC_TIER_3 = [
    # Multi-fund or partial signal
    {'ticker': 'RIOT', 'fund': 'Starboard +44.53%'},
    {'ticker': 'CRM', 'fund': 'Eminence +25%'},
    {'ticker': 'ABDP.L', 'fund': 'Sanford DeLand 4.73%'},
    {'ticker': 'BVXP.L', 'fund': 'Sanford DeLand 5.24%'},
    {'ticker': 'DPH.L', 'fund': 'Sanford DeLand 5.46%'},
    {'ticker': 'LSEG.L', 'fund': 'Sanford DeLand 4.93%'},
    {'ticker': 'GLD', 'fund': 'Bridgewater NEW Q4'},
    {'ticker': 'LYB', 'fund': 'Davis Selected new position'},
    {'ticker': 'JBS', 'fund': 'Davis Selected new position'},
    {'ticker': 'MDLZ', 'fund': 'Lindsell Train new'},
    {'ticker': 'UPS', 'fund': 'Miller Value 2.8%'},
    {'ticker': 'AXON', 'fund': 'Altimeter NEW Q1'},
    {'ticker': 'CLH', 'fund': 'Durable Capital +25% Q1'},
    {'ticker': 'CDRE', 'fund': 'Wynnefield 23.9% +14% Q1'},
    {'ticker': 'LYFT', 'fund': 'Nierenberg 8.55% +60% Q1'},
    {'ticker': 'HAYW', 'fund': 'Trigran 6.8% +14% Q1'},
    {'ticker': 'OI', 'fund': 'Cooper Creek +2.98M sh Q1'},
    {'ticker': 'HOOD', 'fund': 'Glynn Capital NEW Q1'},
    {'ticker': 'SCOR', 'fund': '180 Degree Capital (now MLCI)'},
]


# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'FORENSIC AGENT FINDINGS' in wb.sheetnames:
    del wb['FORENSIC AGENT FINDINGS']

ws = wb.create_sheet('FORENSIC AGENT FINDINGS', 2)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
T1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
T2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
T3_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.merge_cells('A1:I1')
ws['A1'] = 'FORENSIC AGENT FINDINGS — 6 Parallel Agents Scanned All 445 Fund Tabs'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:I2')
ws['A2'] = (f'TIER 1 NEW candidates: {len(FORENSIC_TIER_1)} | TIER 2: {len(FORENSIC_TIER_2)} | '
            f'TIER 3: {len(FORENSIC_TIER_3)} | TOTAL: {len(FORENSIC_TIER_1)+len(FORENSIC_TIER_2)+len(FORENSIC_TIER_3)} '
            f'NEW asymmetric candidates (NOT in our existing 41-name universe).')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

# === TIER 1 ===
ws['A4'] = 'TIER 1 — Passes all 5 forensic checks (smart money building + cost basis ≤ current + dated catalyst + variant intact + bounded downside)'
ws['A4'].font = HEADER; ws['A4'].fill = T1_FILL
ws.merge_cells('A4:I4')

headers = ['#', 'Ticker', 'Name', 'Mcap', 'Fund / Signal', 'Catalyst', 'Variant Perception', 'Downside Floor', 'Source Chunk']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

row = 6
for i, n in enumerate(FORENSIC_TIER_1, 1):
    ws.cell(row=row, column=1, value=i).fill = T1_FILL
    ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=n['name'])
    ws.cell(row=row, column=4, value=n.get('mcap', ''))
    ws.cell(row=row, column=5, value=f"{n['fund']}\n→ {n['signal']}")
    ws.cell(row=row, column=6, value=n['catalyst'])
    ws.cell(row=row, column=7, value=n['variant'])
    ws.cell(row=row, column=8, value=n['floor'])
    ws.cell(row=row, column=9, value=f"Chunk {n['chunk']}")
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 65
    row += 1

# === TIER 2 ===
row += 1
ws.cell(row=row, column=1, value='TIER 2 — Strong signal but partial criteria (3-4 of 5)').fill = T2_FILL
ws.cell(row=row, column=1).font = HEADER
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

t2_headers = ['#', 'Ticker', 'Name', 'Fund / Signal', '', '', '', '', '']
for col, h in enumerate(t2_headers, 1):
    if h:
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER; cell.fill = HEADER_FILL
row += 1

for i, n in enumerate(FORENSIC_TIER_2, 1):
    ws.cell(row=row, column=1, value=i).fill = T2_FILL
    ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=n.get('name', ''))
    ws.cell(row=row, column=4, value=f"{n['fund']} | {n['signal']}")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=9)
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 30
    row += 1

# === TIER 3 ===
row += 1
ws.cell(row=row, column=1, value='TIER 3 — Track but lower conviction or single-signal').fill = T3_FILL
ws.cell(row=row, column=1).font = HEADER
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

for n in FORENSIC_TIER_3:
    ws.cell(row=row, column=1, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = T3_FILL
    ws.cell(row=row, column=2, value=n['fund'])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    row += 1

# Cluster summary
row += 2
ws.cell(row=row, column=1, value='CROSS-FUND CONSENSUS CLUSTERS (multiple agents surfacing same fund pattern):').font = Font(bold=True, size=11)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

clusters = [
    'RA Capital biotech bench: CLYM, ACRV, VOR, PEPG, KPTI, COAG, WVE, ACET — 8 names with single fund max conviction = portfolio bet',
    'Durable Capital industrials: Q (+82%), ANDG (13.2% NEW), FERG (+40%), CLH (+25%) — manufacturing/services cycle play',
    'Sio Capital healthcare: COR (+535%!), WAY (+168%), ICLR (+141%) — most aggressive single-fund accumulation across healthcare',
    'Veradace activist micro: RPAY (8.6%), SSTI (16.5%), TIPT (5.0%) — coordinated activist 13D filings with June nomination windows',
    'Romick/FPA Crescent: AMRZ, AZE.BB, NOV — spin/European/AI energy theme',
    'Land & Buildings REIT: CSR (+39%), EQIX (+82%) — Sunbelt apartments + AI hyperscaler data centers',
    'TCI Fund Mgmt: MCO (+7.7%), SPGI (+19%), CNI (activist CEO) — large-cap quality + Canadian rail activist',
    'Steel Partners ranging: BW (Babcock & Wilcox nuclear 13D) + INMD (155% Q1) — activist breakup specialists',
]
for c in clusters:
    ws.cell(row=row, column=1, value=c).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 30
    row += 1

widths = [5, 12, 28, 9, 35, 35, 40, 30, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)

print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"\n=== AGGREGATE STATS ===")
print(f"TIER 1 NEW candidates: {len(FORENSIC_TIER_1)}")
print(f"TIER 2: {len(FORENSIC_TIER_2)}")
print(f"TIER 3: {len(FORENSIC_TIER_3)}")
print(f"TOTAL NEW: {len(FORENSIC_TIER_1)+len(FORENSIC_TIER_2)+len(FORENSIC_TIER_3)}")
print(f"\nTOP TIER 1 (most asymmetric entry-today + violent re-rate setups):")
for n in FORENSIC_TIER_1[:15]:
    print(f"  {n['ticker']:<8} {n['name'][:30]:<30} | {n['fund'][:30]:<30} | {n['catalyst'][:50]}")
