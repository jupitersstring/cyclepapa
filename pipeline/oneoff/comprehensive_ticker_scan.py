"""Comprehensive ticker scan across all 455 fund tabs.

Extract EVERY ticker mention with context:
  - # tabs mentioning the ticker
  - # tabs with the ticker in "Highest Conviction" section
  - # tabs with the ticker as 13D/13G threshold filing
  - # tabs with the ticker in "Material Adds" section
  - # tabs with the ticker as NEW position
  - Avg/max % portfolio weight when stated
  - Aggregate $ holdings if stated

Cross-reference against existing asymmetric universe (41 names) to find NEW
candidates worth elevating to Tier 1/2/3.
"""
import openpyxl
import re
from collections import defaultdict

wb = openpyxl.load_workbook('/home/user/cyclepapa/fund_activity_last_6mo.xlsx', data_only=True)

# Tickers to exclude (not real or already known false-positives from earlier scan)
NOT_TICKERS = set([
    'CEO', 'CFO', 'COO', 'CTO', 'CIO', 'CMO', 'CCO', 'CRO',
    'AUM', 'RAUM', 'ADD', 'NEW', 'BUY', 'SELL', 'AND', 'THE', 'FOR', 'WITH', 'TO', 'OF',
    'AT', 'BY', 'IN', 'ON', 'OR', 'IS', 'AS', 'IT', 'BE', 'HAS', 'WAS', 'ARE', 'HE',
    'CIK', 'CRD', 'ADV', 'AGM', 'SEC', 'FDA', 'PDUFA', 'IRA', 'CHIPS', 'AI',
    'LLC', 'INC', 'CO', 'LTD', 'PLC', 'LP', 'GP', 'GMBH', 'LLP', 'NV', 'AG', 'SA',
    'Q1', 'Q2', 'Q3', 'Q4', 'FY', 'FYE', 'YTD', 'TTM', 'QoQ', 'YoY',
    'IPO', 'SPAC', 'EBITDA', 'FCF', 'EPS',
    'NDA', 'IND', 'EU', 'US', 'UK', 'HK', 'JV', 'PT',
    'NYSE', 'NASDAQ', 'OTC', 'TSX', 'LSE', 'ASX', 'JPX',
    'PE', 'VC', 'HF', 'PB', 'MD', 'MM', 'BN', 'TN', 'NA',
    'GDP', 'CPI', 'FED', 'ECB', 'BOE', 'BOJ',
    'IRR', 'NPV', 'ROIC', 'ROE', 'ROA',
    'MFN', 'ETF', 'ESG', 'EV', 'MCAP', 'NAV',
    'PCT', 'PCS', 'MTM', 'EOY',
    'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC',
    'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN', 'AM', 'PM', 'BC', 'AD',
    'YES', 'NO', 'OK',
    'PRESS', 'COURT', 'FILING', 'FILES', 'JOINS', 'BOUGHT', 'SOLD', 'PRICE', 'SHARE',
    'SHARES', 'MILLION', 'BILLION', 'TRILLION', 'POSITION', 'STAKE', 'ACQUIRED',
    'HIRED', 'JOINED', 'EXITED', 'FORM', 'SCHEDULE',
    'HYPER', 'MULTI', 'INIT', 'CONVICTION', 'ACTIVIST', 'CONSENSUS',
    'HIGH', 'LOW', 'BUYS', 'MAJOR', 'SAME', 'CUT', 'MAX', 'MIN',
    'MULTIBAGGER', 'COMPANY', 'STOCK', 'DOLLAR', 'AMOUNT', 'PERIOD',
    'PUBLIC', 'PRIVATE', 'GROWTH', 'VALUE', 'QUALITY', 'CYCLICAL',
    'TIER', 'GROUP', 'NOTE', 'TOTAL', 'NET', 'GROSS', 'YIELD',
    'DEEP', 'FAST', 'SLOW', 'BIG', 'SMALL',
    'NORTH', 'SOUTH', 'EAST', 'WEST', 'GLOBAL', 'AMERICA', 'EUROPE', 'ASIA',
    'BUFFETT', 'MUNGER', 'KLARMAN', 'AKMAN', 'TEPPER',
    'TRUE', 'FALSE', 'BULL', 'BEAR', 'OFF',
    'IGNORE', 'INCLUDE', 'EXCEPT',
    'LARGE', 'MID',
    'HOLD', 'KEEP', 'TAKE', 'GIVE',
    'BACK', 'FRONT', 'SIDE',
])

# Skip summary/synthesis tabs (they double-count) — only scan fund tabs
SKIP_SHEETS = {
    'Cover', 'Index', 'All Activity', 'Asymmetric Summary',
    'Consensus Buys', 'Highest Conviction', 'Conviction Adds',
    'Micro-Cap Conviction Adds', 'Activist Catalysts',
    'Multi-Fund New Inits',
    'multibagger_systematic_synthesis', 'asymmetric_ideas_2026Q2',
    'phelps_mayer_lynch_screen', 'yartseva_screen_corrected',
    'positioning_methodology', 'most_asymmetric_deep_dive',
    'fund_coverage_gap', 'qualitative_framework_supplement',
    'final_most_asymmetric_revised', 'not_yet_rerated_asymmetric',
    'coverage_status',
}

# Ticker pattern — must be ALL-CAPS, 2-6 chars, optional .XX suffix, NOT in block list
TICKER_RE = re.compile(r'\b([A-Z]{2,6}(?:\.[A-Z]{1,3})?)\b')

ticker_stats = defaultdict(lambda: {
    'total_mentions': 0,
    'tabs_mentioning': set(),
    'tabs_highest_conviction': set(),
    'tabs_13d_threshold': set(),
    'tabs_material_add': set(),
    'tabs_new_position': set(),
    'add_pcts': [],
})


# Our existing 41 asymmetric universe
KNOWN_ASYMMETRIC = {
    'OPRX', 'PRLD', 'INMD', 'HHH', 'LQDA', 'CTMX-WT', 'MRP', 'GLOB', 'CRTO',
    'NRP', 'AAP', 'CARS', 'KBR', 'FUN', 'AAP CALL', 'AAP-CALL', 'ACHC', 'PVLA',
    'KYMR', 'VITL', 'FTLF', 'MTY.TO', 'BZU.IM', 'COTY', 'JOE', 'MGNI', 'FRPT',
    'CTSH', 'BNTC', 'DLTR', 'QRHC', 'POSTBPB', 'PBPB', 'ROCK', 'XBI', 'TPL',
    'REGN', 'DV', 'CELC', 'HRMY', 'TREE', 'ASND', 'DHR',
    'KVHI', 'BRN', 'SEER', 'GCO', 'STRR', 'SONO', 'BRZE',
    'RAPT', 'LPX', 'WK', 'XNCR', 'SYRE', 'PAR', 'CCO', 'REZI', 'NKTR', 'EXAS',
    'ALKT', 'CMA', 'GPK', 'EQT',
}

# Mega-cap "everybody owns these" exclusions — we want differentiated signals
MEGA_CAPS = {
    'AAPL', 'MSFT', 'GOOGL', 'GOOG', 'AMZN', 'META', 'NVDA', 'TSLA', 'BRK.B',
    'JPM', 'BAC', 'WMT', 'COST', 'V', 'MA', 'ICE', 'NFLX', 'HD', 'PG', 'KO',
    'XOM', 'CVX', 'JNJ', 'UNH', 'LLY', 'ABBV', 'PEP', 'CRM', 'AVGO', 'WFC',
    'C', 'CMCSA', 'GE', 'DIS', 'AAP', 'GS', 'MS', 'TMO', 'BLK', 'AXP',
    'INTC', 'AMD', 'QCOM', 'AMAT', 'TXN', 'ADI',
    'SPY', 'QQQ', 'IWM', 'DIA', 'IVV', 'IBIT', 'IEF', 'ASHR', 'XBI', 'XLF',
    'TSM', 'BABA', 'NIO', 'JD', 'PDD', 'BIDU',
}

# Scan all fund tabs
fund_tabs_scanned = 0
for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue
    fund_tabs_scanned += 1
    ws = wb[sheet_name]

    current_section = 0  # 0=header, 1=highest conviction, 2=13D, 3=new pos, 4=material adds
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        cells = [str(c) if c is not None else '' for c in row]
        text = ' '.join(cells)
        text_lower = text.lower()

        # Track section
        if '(1)' in text and 'highest conviction' in text_lower:
            current_section = 1
            continue
        elif '(2)' in text and ('threshold' in text_lower or 'disclos' in text_lower or '5%' in text):
            current_section = 2
            continue
        elif '(3)' in text and 'new position' in text_lower:
            current_section = 3
            continue
        elif '(4)' in text and ('material' in text_lower or 'increased' in text_lower):
            current_section = 4
            continue

        # Extract tickers from this row
        tickers = TICKER_RE.findall(text)
        for t in tickers:
            if t in NOT_TICKERS or len(t) < 2:
                continue
            if t.isdigit():
                continue
            ticker_stats[t]['total_mentions'] += 1
            ticker_stats[t]['tabs_mentioning'].add(sheet_name)
            if current_section == 1:
                ticker_stats[t]['tabs_highest_conviction'].add(sheet_name)
            elif current_section == 2:
                ticker_stats[t]['tabs_13d_threshold'].add(sheet_name)
            elif current_section == 3:
                ticker_stats[t]['tabs_new_position'].add(sheet_name)
            elif current_section == 4:
                ticker_stats[t]['tabs_material_add'].add(sheet_name)

            # Try to extract % add from text
            m = re.search(r'\+(\d+)%', text)
            if m and current_section == 4:
                ticker_stats[t]['add_pcts'].append(int(m.group(1)))

print(f"Scanned {fund_tabs_scanned} fund tabs")
print(f"Total unique tickers found: {len(ticker_stats)}")


# Compute signal score
def signal_score(t):
    s = ticker_stats[t]
    score = 0
    score += len(s['tabs_mentioning'])
    score += len(s['tabs_highest_conviction']) * 3
    score += len(s['tabs_13d_threshold']) * 5
    score += len(s['tabs_material_add']) * 2
    score += len(s['tabs_new_position']) * 2
    if s['add_pcts']:
        score += min(max(s['add_pcts']), 500) / 50
    return score


# Filter: meaningful signal (mentioned in >=3 tabs) AND not mega-cap AND not already in universe
top_tickers = []
for t, s in ticker_stats.items():
    if len(s['tabs_mentioning']) < 3:
        continue
    if t in MEGA_CAPS:
        continue
    score = signal_score(t)
    if score >= 5:
        top_tickers.append((t, score, s))

top_tickers.sort(key=lambda x: -x[1])

print(f"\nTotal tickers with >=3 tab mentions (ex mega-caps): {len(top_tickers)}")

print(f"\nTOP 80 TICKERS BY SIGNAL SCORE (excluding mega-caps):")
print(f"{'Ticker':<10} {'Score':<6} {'Tabs':<6} {'HC':<4} {'13D':<5} {'Add':<5} {'New':<5} {'MaxAdd':<7} {'In41?':<6}")
print("=" * 80)
for t, score, s in top_tickers[:80]:
    in_universe = '★' if t in KNOWN_ASYMMETRIC else ''
    max_add = max(s['add_pcts']) if s['add_pcts'] else 0
    print(f"{t:<10} {score:<6.1f} {len(s['tabs_mentioning']):<6} {len(s['tabs_highest_conviction']):<4} "
          f"{len(s['tabs_13d_threshold']):<5} {len(s['tabs_material_add']):<5} {len(s['tabs_new_position']):<5} "
          f"+{max_add:<6}% {in_universe}")

# Identify NEW asymmetric candidates (not in current 41)
print(f"\n\n=== TOP 50 NEW TICKERS NOT IN CURRENT 41-NAME UNIVERSE ===")
new_candidates = [(t, score, s) for t, score, s in top_tickers if t not in KNOWN_ASYMMETRIC][:50]
print(f"{'Ticker':<10} {'Score':<6} {'Tabs':<6} {'HC':<4} {'13D':<5} {'Add':<5} {'New':<5} {'MaxAdd':<7}")
print("=" * 70)
for t, score, s in new_candidates:
    max_add = max(s['add_pcts']) if s['add_pcts'] else 0
    print(f"{t:<10} {score:<6.1f} {len(s['tabs_mentioning']):<6} {len(s['tabs_highest_conviction']):<4} "
          f"{len(s['tabs_13d_threshold']):<5} {len(s['tabs_material_add']):<5} {len(s['tabs_new_position']):<5} +{max_add:<6}%")
