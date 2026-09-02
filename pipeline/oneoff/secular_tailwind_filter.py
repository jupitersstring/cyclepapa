"""Add SECULAR TAILWIND filter — re-rank for unpriced multi-year drivers.

Filter: Names with structural multi-year tailwinds NOT priced (vs cyclical bounces
or single binary catalysts). Tier S+ = passes asymmetric + secular tests.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import importlib.util

# Load ARCHETYPES via the original build script's data
ns = {}
with open("/tmp/build_archetype_workbook.py") as f:
    src = f.read()
exec(compile(src.split("# ================================================================\n# BUILD WORKBOOK")[0], "<arch>", "exec"), ns)
ARCHETYPES = ns['ARCHETYPES']

# Load ENTRY_TODAY from rerank script
ns2 = {}
with open("/tmp/rerank_entry_today.py") as f:
    src2 = f.read()
# Just extract the ENTRY_TODAY dict — find its start and end
import re
start = src2.find("ENTRY_TODAY = {")
# Find matching closing brace at column 0
end = src2.find("\n}\n", start) + 2
entry_today_block = src2[start:end]
exec(compile(entry_today_block, "<et>", "exec"), ns2)
ENTRY_TODAY = ns2['ENTRY_TODAY']

# ================================================================
# SECULAR TAILWIND CLASSIFICATION
# tailwind_status: NOT (unpriced) / PARTIAL / PRICED / NONE (cyclical or binary)
# ================================================================

SECULAR = {
    # NOT PRICED — structural multi-year drivers market hasn't recognized
    'BZU.IM':   ('NOT', 'AI data-center concrete demand (US data-center capex translates to cement orders); 52% EBITDA from US; institutional underweight on Italian listing'),
    'OPRX':     ('NOT', 'Digital pharma messaging shift to EHR-native channels (Doximity-style TAM); pharma DTC compliance with MFN actually accelerates digital share; market focused on near-term MFN overhang'),
    'KBR':      ('NOT', 'Triple secular: (1) US defense spend onshoring, (2) NRC small modular reactor / Natrium nuclear, (3) cleared workforce moat 20k; HomeSafe distraction masks underlying secular bookings'),
    'PRLD':     ('NOT', 'Targeted protein degradation platform structural pharma shift + JAK2V617F addresses 95% PV/60% ET/55% MF (aging demographic, no oral comp); SMARCA2 fail = thrown out with platform'),
    'NRP':      ('NOT', 'Met coal (no green substitute for steelmaking) + Sisecam soda ash (solar glass + EV battery) = TWO structural tailwinds masked by dying-thermal-coal narrative; 6.6x vs TPL 25x is a TRUE comp gap'),
    'MRP':      ('NOT', 'US housing supply shortage structural (4M+ undersupply) + Lennar 80% off-take = guaranteed demand; market treats as spinoff orphan, not housing-cycle play'),
    'CRTO':     ('NOT', 'Retail media TAM $200B by 2028 (Insider Intelligence); 235-retailer scale incl Lowes/Costco; OpenAI partnership live; market priced terminal decline, missed shift to Commerce Yield retail media'),
    'MGNI':     ('NOT', 'CTV ad spend doubling 2024-2028 (eMarketer); INDEPENDENT SSP that survived ad-tech wash-out = winner-take-most; Netflix/Disney/Roku exclusive deals scaling; market priced "ad-tech roadkill"'),
    'KYMR':     ('NOT', 'Oral biologic disruption (oral Dupixent = $14B+ franchise displacement); TPD structural pharma platform shift; Phase 1b execution risk masks platform value'),
    'PVLA':     ('NOT', 'Rare-disease pricing structural ($300k+ orphan); QTORIN platform with multiple programs (LMs, venous malformations, angiokeratomas); market priced single-asset story'),

    # PARTIAL — secular real but partially recognized
    'HHH':      ('PARTIAL', 'Sun Belt migration MPC (3,813 acres TX/NV/HI) + Ackman compounding holdco; NAV captures some but Vantage insurance float thesis not yet integrated'),
    'GLOB':     ('PARTIAL', 'LatAm IT services + AI services adoption (Globant AI Pods); class action overhang masks secular validation'),
    'ACHC':     ('PARTIAL', 'Behavioral health secular demand (mental health structural) but DOJ regulatory overhang dominates pricing'),
    'CARS':     ('PARTIAL', 'Auto digitization secular but dealer SaaS partially priced via subscription metrics'),
    'JOE':      ('PARTIAL', 'Florida population growth secular but slow re-rate; Berkowitz anchor recognized'),
    'DLTR':     ('PARTIAL', 'Consumer trade-down secular but already multi-fund consensus'),
    'FRPT':     ('PARTIAL', 'Pet humanization secular real but FRPT-specific manufacturing issues dominate pricing'),
    'CTSH':     ('PARTIAL', 'IT services AI accelerator (vs disruptor) debate; partial market acceptance'),
    'MTY.TO':   ('PARTIAL', 'QSR franchise consolidation secular (Inspire/RBI/Roark precedent) recognized but multiple at trough'),
    'VITL':     ('PARTIAL', 'Pasture-raised premium category creation real but partially recognized'),
    'REGN':     ('PARTIAL', 'Aging demographic + ophthalmology + immunology pipeline partially priced'),

    # PRICED — secular real but already captured in multiple
    'TPL':      ('PRICED', 'Permian premium already at 25x P/CF; secular real but no asymmetry from here'),
    'CELC':     ('PRICED', 'Targeted oncology secular but Phase 3 data + PT raises captured the move'),
    'HRMY':     ('PRICED', 'Orexin/sleep disorders secular validated by Lilly $6.3B Centessa; comp already priced'),
    'TREE':     ('PRICED', 'Insurance up-cycle secular now consensus post-Q1 inflection'),
    'ASND':     ('PRICED', 'Rare endocrine pricing secular but YORVIPATH ramp captured'),
    'DHR':      ('PRICED', 'Bioprocessing cycle now consensus 2H 2026'),

    # NONE — cyclical, single binary, or structural tailwind absent
    'INMD':     ('NONE', 'Aesthetic devices = cyclical consumer, not secular structural'),
    'LQDA':     ('NONE', 'Single product binary patent; PAH/PH-ILD secular but pricing dominated by patent ruling'),
    'CTMX-WT':  ('NONE', 'Pure 5-week binary warrant; no secular driver'),
    'AAP CALL': ('NONE', 'Auto aftermarket cyclical; calls leverage cyclical thesis'),
    'AAP':      ('NONE', 'Auto aftermarket mature; H Partners margin recovery thesis is operational not secular'),
    'FUN':      ('NONE', 'Theme parks cyclical consumer discretionary'),
    'COTY':     ('NONE', 'Prestige beauty cyclical; KKR overhang dominates pricing'),
    'FTLF':     ('NONE', 'Microcap nutraceutical reversion; no secular driver'),
    'QRHC':     ('NONE', 'Waste services minimal secular'),
    'POSTBPB':  ('NONE', 'Restaurant chain cyclical; no secular'),
    'ROCK':     ('NONE', 'Industrial roofing cyclical post-OmniMax integration'),
    'XBI':      ('NONE', 'Biotech basket straddle = vol bet, not directional secular'),
    'DV':       ('NONE', 'CTV verification rally already underway, partial capture'),
    'BNTC':     ('NONE', 'Gene therapy speculative single platform'),
}

# ================================================================
# BUILD UPDATED WORKBOOK
# ================================================================

# Apply rerank + secular to archetypes
for arch_name, arch_data in ARCHETYPES.items():
    for n in arch_data['names']:
        t = n['ticker']
        if t in ENTRY_TODAY:
            rank, status, asym, upside, downside, rr, note = ENTRY_TODAY[t]
            n['cross_rank'] = rank
            n['rerated_status'] = status
            n['asym'] = asym
            n['upside'] = upside
            n['downside'] = downside
            n['rr'] = rr
            n['entry_today_note'] = note
        else:
            n['rerated_status'] = 'UNKNOWN'
            n['entry_today_note'] = ''
        sec = SECULAR.get(t, ('NONE', ''))
        n['tailwind_status'] = sec[0]
        n['secular_tailwind'] = sec[1]

wb = openpyxl.Workbook()

# Styles
BOLD = Font(bold=True, size=11)
HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
NOT_RERATED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
PARTIAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RERATED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
TAILWIND_NOT = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')   # blue = best
TAILWIND_PARTIAL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # pale green
TAILWIND_PRICED = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')   # grey
TAILWIND_NONE = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')     # pale orange
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)


def status_fill(s):
    return {'NOT': NOT_RERATED_FILL, 'PARTIAL': PARTIAL_FILL, 'RERATED': RERATED_FILL}.get(s)


def tailwind_fill(s):
    return {'NOT': TAILWIND_NOT, 'PARTIAL': TAILWIND_PARTIAL, 'PRICED': TAILWIND_PRICED, 'NONE': TAILWIND_NONE}.get(s)


# ============ Index ============
ws = wb.active
ws.title = "Index"
ws.merge_cells('A1:H1')
ws['A1'] = "INVESTMENT ARCHETYPES — ENTRY-TODAY × SECULAR TAILWIND"
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = CENTER

ws['A3'] = "Generated: 2026-05-30 — VIOLENT RE-RATE + SECULAR TAILWIND lens"
ws['A4'] = "Filter stack: deep intrinsic value gap + bounded downside + dated/binary trigger + STRUCTURAL multi-year secular driver NOT priced"
ws['A5'] = "TIER S+ = passes all 3 filters: violent re-rate + deep value + unpriced secular tailwind"

ws['A7'] = "ENTRY-TODAY STATUS"
ws['A7'].font = HEADER; ws['A7'].fill = HEADER_FILL
ws.merge_cells('A7:D7')
ws['A8'] = "NOT"; ws['A8'].fill = NOT_RERATED_FILL
ws['B8'] = "Major catalyst still ahead; upside intact from current price"
ws['A9'] = "PARTIAL"; ws['A9'].fill = PARTIAL_FILL
ws['B9'] = "Some catalyst played; remaining upside reduced"
ws['A10'] = "RERATED"; ws['A10'].fill = RERATED_FILL
ws['B10'] = "Catalyst played; entry-today economics weakened"

ws['F7'] = "SECULAR TAILWIND"
ws['F7'].font = HEADER; ws['F7'].fill = HEADER_FILL
ws.merge_cells('F7:I7')
ws['F8'] = "NOT"; ws['F8'].fill = TAILWIND_NOT
ws['G8'] = "Structural multi-year driver NOT priced — best tier"
ws['F9'] = "PARTIAL"; ws['F9'].fill = TAILWIND_PARTIAL
ws['G9'] = "Secular real but partially recognized"
ws['F10'] = "PRICED"; ws['F10'].fill = TAILWIND_PRICED
ws['G10'] = "Secular real but multiple captures it"
ws['F11'] = "NONE"; ws['F11'].fill = TAILWIND_NONE
ws['G11'] = "Cyclical, binary, or no secular driver"

ws['A13'] = "ARCHETYPE INDEX"
ws['A13'].font = HEADER; ws['A13'].fill = HEADER_FILL
ws.merge_cells('A13:I13')

ws['A15'] = "#"; ws['B15'] = "Archetype"; ws['C15'] = "Names"; ws['D15'] = "Top Pick (Rank / Tailwind)"
for c in ['A15', 'B15', 'C15', 'D15']:
    ws[c].font = BOLD; ws[c].fill = SECTION_FILL

row = 16
for i, (arch_name, arch_data) in enumerate(ARCHETYPES.items(), 1):
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=arch_name)
    ws.cell(row=row, column=3, value=len(arch_data['names']))
    top = sorted(arch_data['names'], key=lambda x: x['cross_rank'])[0]
    ws.cell(row=row, column=4, value=f"{top['ticker']} (#{top['cross_rank']}, tw={top['tailwind_status']})")
    row += 1

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 32
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 50

# ============ TIER S+ — Best of All Filters ============
ws_tier = wb.create_sheet("TIER S+ Secular×Asymmetric")
ws_tier.merge_cells('A1:K1')
ws_tier['A1'] = "TIER S+ — VIOLENT RE-RATE + DEEP INTRINSIC GAP + UNPRICED SECULAR TAILWIND"
ws_tier['A1'].font = Font(bold=True, size=14)
ws_tier['A1'].alignment = CENTER

# Collect, dedupe, filter to tailwind_status == 'NOT'
all_names = []
seen = set()
for arch_name, arch_data in ARCHETYPES.items():
    for n in arch_data['names']:
        if n['ticker'] not in seen:
            seen.add(n['ticker'])
            nc = dict(n); nc['archetype'] = arch_name
            all_names.append(nc)
all_names.sort(key=lambda x: x['cross_rank'])

tier_s_plus = [n for n in all_names if n['tailwind_status'] == 'NOT']

ws_tier['A3'] = f"Universe: {len(tier_s_plus)} names with UNPRICED secular tailwind, sorted by violent-re-rate asymmetry"
ws_tier['A3'].font = Font(italic=True)

headers = ['Rerank', 'Ticker', 'Name', 'Mcap', 'Price', 'Asym', 'Upside', 'R/R', 'Secular Tailwind (NOT priced)', 'Entry-Today Setup', 'Catalyst']
for col, h in enumerate(headers, 1):
    cell = ws_tier.cell(row=5, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

for i, n in enumerate(tier_s_plus, 6):
    ws_tier.cell(row=i, column=1, value=tier_s_plus.index(n) + 1).font = Font(bold=True)
    ws_tier.cell(row=i, column=2, value=n['ticker']).font = Font(bold=True)
    ws_tier.cell(row=i, column=3, value=n['name'])
    ws_tier.cell(row=i, column=4, value=n['mcap'])
    ws_tier.cell(row=i, column=5, value=n['price'])
    ws_tier.cell(row=i, column=6, value=n['asym'])
    ws_tier.cell(row=i, column=7, value=n['upside'])
    ws_tier.cell(row=i, column=8, value=n['rr'])
    sc = ws_tier.cell(row=i, column=9, value=n['secular_tailwind'])
    sc.fill = TAILWIND_NOT
    ws_tier.cell(row=i, column=10, value=n.get('entry_today_note', ''))
    ws_tier.cell(row=i, column=11, value=n['catalyst'])
    for col in range(1, 12):
        ws_tier.cell(row=i, column=col).alignment = LEFT_WRAP

widths = [8, 10, 22, 10, 10, 7, 18, 8, 55, 50, 35]
for i, w in enumerate(widths, 1):
    ws_tier.column_dimensions[get_column_letter(i)].width = w

# ============ MASTER (re-include with tailwind column) ============
ws_master = wb.create_sheet("MASTER Entry-Today")
ws_master.merge_cells('A1:O1')
ws_master['A1'] = "MASTER ENTRY-TODAY × SECULAR TAILWIND RANKING"
ws_master['A1'].font = Font(bold=True, size=14); ws_master['A1'].alignment = CENTER

headers = ['Rank', 'Re-rate', 'Tailwind', 'Ticker', 'Name', 'Archetype', 'Mcap', 'Price', 'Asym', 'Upside', 'Downside', 'R/R', 'Secular Driver', 'Entry-Today Note', 'Smart Money']
for col, h in enumerate(headers, 1):
    cell = ws_master.cell(row=3, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

for row_idx, n in enumerate(all_names, 4):
    ws_master.cell(row=row_idx, column=1, value=n['cross_rank'])
    s1 = ws_master.cell(row=row_idx, column=2, value=n['rerated_status'])
    if status_fill(n['rerated_status']): s1.fill = status_fill(n['rerated_status'])
    s1.font = Font(bold=True)
    s2 = ws_master.cell(row=row_idx, column=3, value=n['tailwind_status'])
    if tailwind_fill(n['tailwind_status']): s2.fill = tailwind_fill(n['tailwind_status'])
    s2.font = Font(bold=True)
    ws_master.cell(row=row_idx, column=4, value=n['ticker']).font = Font(bold=True)
    ws_master.cell(row=row_idx, column=5, value=n['name'])
    ws_master.cell(row=row_idx, column=6, value=n['archetype'])
    ws_master.cell(row=row_idx, column=7, value=n['mcap'])
    ws_master.cell(row=row_idx, column=8, value=n['price'])
    ws_master.cell(row=row_idx, column=9, value=n['asym'])
    ws_master.cell(row=row_idx, column=10, value=n['upside'])
    ws_master.cell(row=row_idx, column=11, value=n['downside'])
    ws_master.cell(row=row_idx, column=12, value=n['rr'])
    ws_master.cell(row=row_idx, column=13, value=n['secular_tailwind'])
    ws_master.cell(row=row_idx, column=14, value=n.get('entry_today_note', ''))
    ws_master.cell(row=row_idx, column=15, value=n['smart_money'])
    for col in range(1, 16):
        ws_master.cell(row=row_idx, column=col).alignment = LEFT_WRAP

widths = [6, 9, 9, 9, 22, 32, 10, 10, 6, 18, 10, 7, 50, 45, 35]
for i, w in enumerate(widths, 1):
    ws_master.column_dimensions[get_column_letter(i)].width = w

# ============ Per-Archetype Sheets (add tailwind column) ============
for arch_name, arch_data in ARCHETYPES.items():
    sheet_name = arch_name.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '')[:31]
    ws_a = wb.create_sheet(sheet_name)

    ws_a.merge_cells('A1:M1')
    ws_a['A1'] = arch_name + ' — ENTRY-TODAY × TAILWIND'
    ws_a['A1'].font = Font(bold=True, size=13); ws_a['A1'].alignment = CENTER

    ws_a.merge_cells('A2:M2')
    ws_a['A2'] = f"Definition: {arch_data['definition']}"
    ws_a['A2'].font = Font(italic=True, size=10); ws_a['A2'].alignment = LEFT_WRAP

    arch_sorted = sorted(arch_data['names'], key=lambda x: x['cross_rank'])

    headers = ['Within-Rank', 'Master-Rank', 'Re-rate', 'Tailwind', 'Ticker', 'Name', 'Mcap', 'Price', 'Asym', 'Upside', 'R/R', 'Smart Money', 'Secular Driver']
    for col, h in enumerate(headers, 1):
        cell = ws_a.cell(row=4, column=col, value=h)
        cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

    for i, n in enumerate(arch_sorted):
        r = 5 + i*4
        ws_a.cell(row=r, column=1, value=i+1)
        ws_a.cell(row=r, column=2, value=n['cross_rank'])
        s1 = ws_a.cell(row=r, column=3, value=n['rerated_status'])
        if status_fill(n['rerated_status']): s1.fill = status_fill(n['rerated_status'])
        s1.font = Font(bold=True)
        s2 = ws_a.cell(row=r, column=4, value=n['tailwind_status'])
        if tailwind_fill(n['tailwind_status']): s2.fill = tailwind_fill(n['tailwind_status'])
        s2.font = Font(bold=True)
        ws_a.cell(row=r, column=5, value=n['ticker']).font = Font(bold=True)
        ws_a.cell(row=r, column=6, value=n['name'])
        ws_a.cell(row=r, column=7, value=n['mcap'])
        ws_a.cell(row=r, column=8, value=n['price'])
        ws_a.cell(row=r, column=9, value=n['asym'])
        ws_a.cell(row=r, column=10, value=n['upside'])
        ws_a.cell(row=r, column=11, value=n['rr'])
        ws_a.cell(row=r, column=12, value=n['smart_money'])
        ws_a.cell(row=r, column=13, value=n['secular_tailwind'])

        ws_a.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=13)
        ws_a.cell(row=r+1, column=1, value=f"ENTRY-TODAY: {n.get('entry_today_note', '')}").alignment = LEFT_WRAP
        ws_a.cell(row=r+1, column=1).fill = SECTION_FILL
        ws_a.cell(row=r+1, column=1).font = Font(italic=True)

        ws_a.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=13)
        ws_a.cell(row=r+2, column=1, value=f"THESIS: {n['thesis']} | VALUATION: {n['valuation']}").alignment = LEFT_WRAP

        ws_a.merge_cells(start_row=r+3, start_column=1, end_row=r+3, end_column=13)
        ws_a.cell(row=r+3, column=1, value=f"CATALYST: {n['catalyst']} | VARIANT: {n['variant']}").alignment = LEFT_WRAP

    widths = [10, 10, 9, 9, 8, 22, 10, 10, 7, 15, 8, 30, 50]
    for i, w in enumerate(widths, 1):
        ws_a.column_dimensions[get_column_letter(i)].width = w
    ws_a.row_dimensions[4].height = 25

# Save
out_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
print(f"\nTIER S+ (Violent re-rate + deep value + UNPRICED secular tailwind):")
for i, n in enumerate(tier_s_plus, 1):
    note = n['secular_tailwind'][:80]
    print(f"  TS+#{i:<2} (Master #{n['cross_rank']:<2}) {n['ticker']:<10} asym {n['asym']} — {note}")
print(f"\nBy tailwind status counts:")
from collections import Counter
c = Counter(n['tailwind_status'] for n in all_names)
for k in ['NOT', 'PARTIAL', 'PRICED', 'NONE']:
    print(f"  {k:<8} = {c[k]}")
