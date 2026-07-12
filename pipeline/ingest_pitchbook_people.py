"""Ingest PitchBook people-search exports into pb_people / pb_affiliation.

These are watchlists of influential individuals — hedge-fund titans, concentrated
value managers, sovereign-wealth heads, dynastic family offices, tech founders —
keyed off a Biography search, with each person's PUBLIC-company board seats and
primary affiliation. The alpha "tea leaves": which public tickers these people
are attached to (active board seat = live signal), and which of their funds/FOs
are missing from our roster.

Four of the five uploaded files load; the fifth (0a7e48ee, 2025-09-19) is
physically truncated (only 64KB of the zip survived — data streams gone), so its
rows are unrecoverable. Its metadata shows it was a same-author 2025-09-20 search.

Tables:
  pb_people       one row per (person, position) — a person can hold several
  pb_affiliation  person -> company edge, typed, active/former, ticker if mapped
"""
import os, re, sqlite3, glob
import openpyxl

# Words that mark an entity (not a person) in the quoted search terms — used to
# keep only person-like principals when parsing Search Criteria.
_ORG_RE = re.compile(
    r"(Office|Group|Capital|Holding|Fund|Enterpris|Trading|Element|Bank|Foundation|"
    r"Council|Purse|Duchy|House| Ltd| SA |Investment|Suisse|Privy|Keeper|Royal|"
    r"Partners|Ventures|Advisor|Management|Company|Qatar|Zabeel|Shamal|Seed)", re.I)

def _name_key(s):
    return re.sub(r"[^a-z]", "", (s or "").lower())

def parse_principals(paths):
    """The quoted names in each file's Search Criteria ARE the tracked people —
    the search returns anyone whose biography mentions them, so flagging rows
    where the person IS a searched name separates principals from associates."""
    principals = set()
    for path in paths:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["Data"]
        crit = ""
        for r in ws.iter_rows(max_row=8, values_only=True):
            if r and r[0] == "Search Criteria:":
                crit = str(r[1] or ""); break
        wb.close()
        for n in re.findall("[“”\"]([^“”\"]+)[“”\"]", crit):
            n = n.strip()
            if 1 < len(n) < 40 and not _ORG_RE.search(n):
                principals.add(n)
    return principals

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(BASE, "data", "cyclepapa.db")
PB = os.path.join(BASE, "data", "pitchbook")

# Map each export (by filename prefix) to the theme it represents. The theme is
# how the monitor groups signals and how gap-fill routes funds to a style.
THEMES = {
    "524d5a6b": "Sequoia Network",
    "a50dae15": "Gulf / Royal Family Offices",
    "f2aec892": "Billionaire / Oligarch Vehicles",
    "fbc86b51": "Titans & Concentrated Managers",
}

def _cell(r, ci, key):
    i = ci.get(key)
    if i is None or i >= len(r) or r[i] is None:
        return None
    v = str(r[i]).strip()
    return v or None

def init_schema(conn):
    conn.executescript("""
    DROP TABLE IF EXISTS pb_people;
    DROP TABLE IF EXISTS pb_affiliation;
    CREATE TABLE pb_people (
      person_id TEXT, first_name TEXT, last_name TEXT, full_name TEXT,
      primary_company TEXT, primary_company_type TEXT, primary_position TEXT,
      is_former INTEGER, board_seats TEXT, roles TEXT,
      location TEXT, country TEXT, biography TEXT,
      theme TEXT, company_website TEXT, is_principal INTEGER DEFAULT 0);
    CREATE INDEX idx_pbp_name ON pb_people(full_name);
    CREATE INDEX idx_pbp_company ON pb_people(primary_company);
    CREATE INDEX idx_pbp_theme ON pb_people(theme);
    CREATE TABLE pb_affiliation (
      full_name TEXT, company TEXT, company_type TEXT, position TEXT,
      is_former INTEGER, theme TEXT, ticker TEXT, is_principal INTEGER DEFAULT 0);
    CREATE INDEX idx_pba_company ON pb_affiliation(company);
    CREATE INDEX idx_pba_ticker ON pb_affiliation(ticker);
    """)

_FORMER_RE = re.compile(r"\(former\)", re.I)

def load_file(path, conn, principal_keys):
    prefix = os.path.basename(path).split("-")[0]
    theme = THEMES.get(prefix, "Other")
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = next((i for i, r in enumerate(rows) if r and r[0] == "Person ID"), None)
    if hdr_i is None:
        return 0
    hdr = rows[hdr_i]
    ci = {h: i for i, h in enumerate(hdr) if h}
    n = 0
    for r in rows[hdr_i + 1:]:
        if not r or not r[0]:
            continue
        first = _cell(r, ci, "First Name") or ""
        last = _cell(r, ci, "Last Name") or ""
        full = (first + " " + last).strip() or _cell(r, ci, "People") or ""
        company = _cell(r, ci, "Primary Company")
        ctype = _cell(r, ci, "Primary Company Type")
        pos = _cell(r, ci, "Primary Position")
        # PitchBook marks stale affiliations "(Former)" in the company field
        raw_company = company or ""
        is_former = 1 if _FORMER_RE.search(raw_company) else 0
        company = _FORMER_RE.sub("", raw_company).strip(" ()") or None
        # some company cells carry a trailing "(Industry ...)" — strip parenthetical tail
        if company:
            company = re.sub(r"\s*\([^)]*\)\s*$", "", company).strip() or company
        loc = _cell(r, ci, "Location")
        country = _cell(r, ci, "Country/Territory/Region")
        is_prin = 1 if _name_key(full) in principal_keys else 0
        conn.execute("""INSERT INTO pb_people VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
            _cell(r, ci, "Person ID"), first, last, full,
            company, ctype, pos, is_former,
            _cell(r, ci, "Board Seats"), _cell(r, ci, "Roles"),
            loc, country, _cell(r, ci, "Biography"),
            theme, _cell(r, ci, "Primary Company Website"), is_prin))
        if company:
            conn.execute("""INSERT INTO pb_affiliation
                (full_name, company, company_type, position, is_former, theme, ticker, is_principal)
                VALUES (?,?,?,?,?,?,NULL,?)""",
                (full, company, ctype, pos, is_former, theme, is_prin))
        n += 1
    return n

def run():
    conn = sqlite3.connect(DB)
    init_schema(conn)
    paths = [p for p in sorted(glob.glob(os.path.join(PB, "*.xlsx")))
             if not os.path.basename(p).startswith(("0a7e48ee", "repaired"))]
    principals = parse_principals(paths)
    pkeys = {_name_key(p) for p in principals}
    conn.execute("DROP TABLE IF EXISTS pb_principal")
    conn.execute("CREATE TABLE pb_principal (name TEXT PRIMARY KEY)")
    for p in sorted(principals):
        conn.execute("INSERT OR IGNORE INTO pb_principal VALUES (?)", (p,))
    total = 0
    for path in paths:
        base = os.path.basename(path)
        k = load_file(path, conn, pkeys)
        print(f"  {base.split('-')[0]} [{THEMES.get(base.split('-')[0],'?')}]: {k} people")
        total += k
    conn.commit()
    n_prin = conn.execute("SELECT COUNT(DISTINCT full_name) FROM pb_people WHERE is_principal=1").fetchone()[0]
    print(f"loaded {total} people rows; "
          f"{conn.execute('SELECT COUNT(DISTINCT full_name) FROM pb_people').fetchone()[0]} distinct individuals; "
          f"{len(principals)} principals parsed, {n_prin} present in data")
    conn.close()

if __name__ == "__main__":
    run()
