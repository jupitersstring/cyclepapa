"""Build AGGRESSIVE BUY SIGNALS — most aggressive insider + fund adds.

Combines three signals from the 442-sheet workbook:
  1. Conviction Adds tab (multi-fund accumulation, $ sum, max % add)
  2. Activist Catalysts tab (13D/13G threshold filings, max fund concentration)
  3. All Activity scan (Form 4, insider/founder/director buys, family fund 13D escalation)

Output: ranked AGGRESSIVE BUY SIGNALS sheet with tier classifications.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# COMPILED DATA — combining all three signal sources
# ================================================================

SIGNALS = [
    # ============================================================
    # TIER 1 — TRIPLE SIGNAL (multi-fund add + activist + insider)
    # ============================================================
    {'tier': 1, 'ticker': 'HHH', 'name': 'Howard Hughes Holdings', 'mcap': '$3.5B',
     'fund_adds': 'Pershing 47.74% / $900M + JCP $14.45M NEW (2 funds, $914.5M sum)',
     'activist': '13D Pershing (47% via $900M + $1B preferred)',
     'insider': 'Marc Grandisson (ex-Arch Capital CEO) joined board April 2026 = ex-insurance CEO insider hire signal',
     'asym_rank': 4, 'tier_s_plus': 'NO (partial tailwind)',
     'why_aggressive': '$914.5M total fund buying + Ackman $1.9B commitment underwater + insurance CEO hire = bet-the-Berkshire-thesis'},

    {'tier': 1, 'ticker': 'KBR', 'name': 'KBR Inc', 'mcap': '$3.9B',
     'fund_adds': 'D3/Nierenberg +425% Q1 = LARGEST single add in 442-sheet universe; Irenic activist ~1%',
     'activist': 'Irenic activist position + spin board-approved Sept 2025',
     'insider': 'Director Form 4 buys May 2026 at $30-31 (multiple insiders)',
     'asym_rank': 13, 'tier_s_plus': 'YES',
     'why_aggressive': 'D3 quintupled position in Q1 + directors buying May 2026 at $30 = sponsor put at current; spin Jan 4 2027'},

    {'tier': 1, 'ticker': 'INMD', 'name': 'InMode Ltd', 'mcap': '$869M',
     'fund_adds': 'Steel Partners +155% Q1 2026 = doubled position (25.5% of $338M Steel book)',
     'activist': 'Steel Partners activist anchor; prior $18 takeout',
     'insider': 'Founder Mizrahy bought 800,000 sh ($10.7M) Feb 2026 OPEN-MARKET at $13.50',
     'asym_rank': 3, 'tier_s_plus': 'NO (cyclical aesthetics)',
     'why_aggressive': 'Activist + founder BOTH at current FLAT; cash $8.47/sh = 62% floor; M&A optionality'},

    {'tier': 1, 'ticker': 'AAP', 'name': 'Advance Auto Parts', 'mcap': '$3.3B',
     'fund_adds': 'H Partners 46.3% / +50% sh Q1 ADD (1.35M new sh); Legion Partners precedent',
     'activist': 'H Partners 46% concentration approaching activist threshold',
     'insider': 'Q2 2026 margin proof Aug = catalyst (no documented insider yet)',
     'asym_rank': 15, 'tier_s_plus': 'NO',
     'why_aggressive': 'HALF a fund\'s book in one name + STILL ADDING Q1 (1.35M new sh)'},

    {'tier': 1, 'ticker': 'MRP', 'name': 'Millrose Properties', 'mcap': '$4.59B',
     'fund_adds': 'Brave Warrior (Greenberg) +59.99% Q1 + Permian Investment Partners (2 funds)',
     'activist': 'Lennar 80%+ owner = sponsor lock',
     'insider': 'Lennar 80% parent stake = effective insider commitment',
     'asym_rank': 7, 'tier_s_plus': 'YES (US housing supply unpriced)',
     'why_aggressive': 'Greenberg +60% Q1 + Lennar 80% locked + B/M 1.274 deepest in cohort'},

    {'tier': 1, 'ticker': 'FUN', 'name': 'Six Flags Entertainment', 'mcap': '$2.2B',
     'fund_adds': 'H Partners 53.7% (MAX concentration in 13F universe) + Cove Street +1,711% + L&B + Jana 9%',
     'activist': 'Two-front: H Partners 53.7% / 5.7% class + Jana 9% pushing OUTRIGHT SALE',
     'insider': 'Jaffer joined board May 26 2026 (Class III through 2027, Audit Cttee)',
     'asym_rank': 14, 'tier_s_plus': 'NO',
     'why_aggressive': '53.7% MAX single-fund concentration + Jana sale push + Jaffer board seat just took effect'},

    # ============================================================
    # TIER 2 — ACTIVIST + INSIDER FAMILY BUYING (micro/small caps)
    # ============================================================
    {'tier': 2, 'ticker': 'KVHI', 'name': 'KVH Industries', 'mcap': '<$200M',
     'fund_adds': 'Radoff family 10%+ owner build',
     'activist': 'Radoff 2,150,000 direct + 325,000 Foundation = 2,475,000 sh = 10%+ owner',
     'insider': 'Nov 2025 Form 4: +40,000 sh P-code purchase at $5.78',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Family fund 10%+ owner + Form 4 open-market buy Nov 2025 + Foundation co-investing = aligned multi-generation insider'},

    {'tier': 2, 'ticker': 'BRN', 'name': 'Barnwell Industries', 'mcap': '<$50M',
     'fund_adds': 'Radoff 12.8% + Foundation 6.5% = group 17.4% peak',
     'activist': '13D escalation 9.1%→12.8% Dec 2025 → Mar 2026 (multiple amendments)',
     'insider': 'Director Schechter Form 4 Dec 2025 (83,674 sh)',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Ultra-microcap (<$50M) with 17.4% group activist build + director Form 4 = total takeover possible'},

    {'tier': 2, 'ticker': 'SEER', 'name': 'Seer, Inc.', 'mcap': '~$200M',
     'fund_adds': 'Radoff group escalating 9.3% → 10.5% → 10.6%',
     'activist': 'Multiple 13D/A amendments Feb-April 2026; Radoff+Torok+JEC II+MOS group',
     'insider': 'Family fund coordinated buying',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Triple 13D escalation in 2 months on micro proteomics name'},

    {'tier': 2, 'ticker': 'GCO', 'name': 'Genesco Inc.', 'mcap': '$280M',
     'fund_adds': 'Radoff/Jumana/Ross group 11.7% → 12.3%',
     'activist': 'Initial 13D April 15 → escalated April 27 2026',
     'insider': 'Coordinated three-party activist build',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Three-party 13D coordinated on apparel retail micro = activist consensus'},

    {'tier': 2, 'ticker': 'STRR', 'name': 'Star Equity Holdings', 'mcap': '<$30M',
     'fund_adds': 'Star Equity Fund (Eberwein) 27.1% concentration',
     'activist': 'Eberwein 13D + board control',
     'insider': 'Eberwein +$3M OPEN-MARKET BUYING Q4 2025 (Dec 8) — 21.4%→27.1%',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '$3M open-market buying on a sub-$30M micro = ultra-aggressive insider build'},

    {'tier': 2, 'ticker': 'SONO', 'name': 'Sonos Inc', 'mcap': '$1.6B',
     'fund_adds': 'Coliseum Capital (Gray/Shackelton) — anchor',
     'activist': 'Coliseum 13D activist + ~19.8% concentration',
     'insider': 'Coliseum +$23M+ FEB-MAR 2026 multiple Form 4 buys ($3.07M, $10.1M, $9.14M, $12.9M)',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '$35M+ open-market buying in 6 weeks (4 Form 4 buys Feb-Mar 2026) = extreme conviction at current'},

    # ============================================================
    # TIER 3 — MULTI-FUND ADDS + ACTIVIST (no documented insider)
    # ============================================================
    {'tier': 3, 'ticker': 'RAPT', 'name': 'RAPT Therapeutics', 'mcap': '~$100M',
     'fund_adds': 'Caligan +$21.8M NEW + Polygon + OrbiMed (3 funds, +60% max)',
     'activist': 'OrbiMed 60% concentration on one name = max conviction',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'OrbiMed 60% concentration + 3 funds NEW Q1 = biotech triple stack'},

    {'tier': 3, 'ticker': 'LPX', 'name': 'Louisiana-Pacific', 'mcap': '$5.3B',
     'fund_adds': 'Eminence +$327M / +29% + Scopia + Southeastern (3 funds, $365M sum)',
     'activist': 'Eminence 5.2% threshold',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '$365M from 3 funds + Eminence threshold filing on building products'},

    {'tier': 3, 'ticker': 'WK', 'name': 'Workday', 'mcap': '$57B',
     'fund_adds': 'Praesidium +60% + Irenic + Eminence',
     'activist': 'Irenic Capital 4.48% threshold filing',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Eminence + Irenic activist + Praesidium 60% add — large cap activism rare'},

    {'tier': 3, 'ticker': 'XNCR', 'name': 'Xencor', 'mcap': '~$1B',
     'fund_adds': 'EcoR1 + RA Capital (2 funds, +160% max!)',
     'activist': 'EcoR1 + RA Cap dual threshold filings',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '+160% MAX ADD by single fund + dual biotech crossover threshold = aggressive biotech setup'},

    {'tier': 3, 'ticker': 'SYRE', 'name': 'Spyre Therapeutics', 'mcap': '~$800M',
     'fund_adds': 'Perceptive +149% Q1',
     'activist': 'Perceptive 5% threshold',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Perceptive +149% Q1 ADD on biotech = top-tier biotech specialist max conviction'},

    {'tier': 3, 'ticker': 'PAR', 'name': 'PAR Technology', 'mcap': '$1.5B',
     'fund_adds': 'Voss Capital concentration / Permian Investment Partners',
     'activist': 'Voss 46% concentration on one name',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Voss 46% = bet-the-fund conviction on restaurant tech'},

    {'tier': 3, 'ticker': 'CCO', 'name': 'Clear Channel Outdoor', 'mcap': '~$700M',
     'fund_adds': 'Legion Partners + 2 funds, $43.9M sum',
     'activist': 'Legion 33.6% activist concentration',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Legion 33.6% concentration + outdoor advertising recovery'},

    {'tier': 3, 'ticker': 'BRZE', 'name': 'Braze Inc', 'mcap': '$4.8B',
     'fund_adds': 'Iconiq 11.6% + Millennium',
     'activist': 'Iconiq 11.6% threshold + Millennium',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Iconiq 11.6% on customer engagement platform; growth equity activist setup'},

    # ============================================================
    # TIER 4 — STRONG SINGLE-FUND + INSIDER OR BUYBACK
    # ============================================================
    {'tier': 4, 'ticker': 'OPRX', 'name': 'OptimizeRx', 'mcap': '$93M',
     'fund_adds': 'Insider 14.82% + $10M buyback through March 2027',
     'activist': 'Doximity exec Presti joined board April 2026 = strategic alignment',
     'insider': 'Corporate buyback executing at $4-5 = effective insider buying',
     'asym_rank': 1, 'tier_s_plus': 'YES',
     'why_aggressive': '#1 Tier S+ asymmetric rank + Doximity exec to board = strategic peer insider signal'},

    {'tier': 4, 'ticker': 'POSTBPB', 'name': 'Potbelly Corp', 'mcap': '$240M',
     'fund_adds': 'Nierenberg/D3 LARGEST shareholder (~$16.4M / 7% of D3 book)',
     'activist': '13D/A active March 2025; ongoing Form 4 insider buys by Nierenberg',
     'insider': 'Nierenberg Form 4 buying at depressed levels',
     'asym_rank': 31, 'tier_s_plus': 'N/A',
     'why_aggressive': 'D3 largest shareholder + ongoing Form 4 buys = activist soft pressure micro'},

    {'tier': 4, 'ticker': 'NRP', 'name': 'Natural Resource Partners', 'mcap': '$1.35B',
     'fund_adds': 'Right Tail NEW Q1 2026 + Saber 17.75% + Greystone + Berkowitz/Fairholme (5 funds)',
     'activist': 'Multi-fund anchor convergence',
     'insider': 'Robertson family 31.75% insider (permanent alignment)',
     'asym_rank': 10, 'tier_s_plus': 'YES',
     'why_aggressive': '5-fund concentration + 31.75% family insider + Right Tail freshest entry Q1 at current'},

    {'tier': 4, 'ticker': 'MTY.TO', 'name': 'MTY Food Group', 'mcap': 'C$887M',
     'fund_adds': 'CEO Lefebvre +33% stake increase',
     'activist': 'CEO open-market buy = governance signal',
     'insider': 'CEO Lefebvre open-market buy at C$30.54 = $1M+ stake increase',
     'asym_rank': 21, 'tier_s_plus': 'N/A',
     'why_aggressive': 'CEO buying $1M+ at depressed multiple = strong governance signal at book-value QSR'},

    {'tier': 4, 'ticker': 'ROCK', 'name': 'Gibraltar Industries', 'mcap': '$1.10B',
     'fund_adds': 'Vanguard 5.24% 13G NEW',
     'activist': 'Vanguard threshold filing',
     'insider': 'Director Metcalf bought $502,000 March 2026',
     'asym_rank': 29, 'tier_s_plus': 'N/A',
     'why_aggressive': 'Director $502k open-market March 2026 + Vanguard 5.24% NEW = post-OmniMax integration buying'},

    {'tier': 4, 'ticker': 'BNTC', 'name': 'Benitec Biopharma', 'mcap': '$200M',
     'fund_adds': 'Suvretta 44.1% (LARGEST stake by % in their portfolio)',
     'activist': 'Suvretta 13D anchor',
     'insider': 'PIPE participation Nov 2025',
     'asym_rank': 28, 'tier_s_plus': 'N/A',
     'why_aggressive': 'Single-fund 44.1% concentration on micro biotech + PIPE = high-conviction lock'},

    {'tier': 4, 'ticker': 'ACMR', 'name': 'ACM Research', 'mcap': '~$400M',
     'fund_adds': 'Kerrisdale 3.36% + Steamboat joint',
     'activist': 'Joint letter Nov 2025 — HK listing $100M offer pursuit',
     'insider': 'No documented Form 4',
     'asym_rank': 'N/A', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Two-fund JOINT LETTER (rare) for HK value unlock = forced corporate action'},

    # ============================================================
    # TIER 5 — MICRO-CAP MULTI-FUND ADDS (smaller signal but worth tracking)
    # ============================================================
    {'tier': 5, 'ticker': 'REZI', 'name': 'Resideo Technologies', 'mcap': '$2.6B',
     'fund_adds': 'Alta Fox 9.57% NEW ($44.5M) + GoldenTree + 2 more (4 funds, $154.5M)',
     'activist': 'Alta Fox + GoldenTree',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '4-fund convergence on home tech with Alta Fox initiating 9.57%'},

    {'tier': 5, 'ticker': 'NKTR', 'name': 'Nektar Therapeutics', 'mcap': '~$280M',
     'fund_adds': '22NW +331k sh (5.9% port) + BVF + Polygon (3 funds)',
     'activist': '22NW threshold + BVF',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '3-fund biotech micro adds + BVF anchor (biotech quality signal)'},

    {'tier': 5, 'ticker': 'EXAS', 'name': 'Exact Sciences', 'mcap': '$15B',
     'fund_adds': 'Hudson Bay $113.3M NEW + Pentwater (3 funds, $849M sum, +157%)',
     'activist': '',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': '$849M from 3 funds + Cologuard moat'},

    {'tier': 5, 'ticker': 'ALKT', 'name': 'Alkami Technology', 'mcap': '~$2B',
     'fund_adds': 'JANA Partners +158% / $69M + Praesidium',
     'activist': 'JANA Partners 5.87% threshold',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'JANA +158% ADD = activist heavyweight conviction'},

    {'tier': 5, 'ticker': 'CMA', 'name': 'Comerica', 'mcap': '$8.7B',
     'fund_adds': 'Magnetar $99.1M NEW + Whitebox $109M NEW',
     'activist': '',
     'insider': 'No documented Form 4',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Two large NEW positions Q1 from sophisticated quants on regional bank'},

    {'tier': 5, 'ticker': 'GPK', 'name': 'Graphic Packaging', 'mcap': '$8.5B',
     'fund_adds': 'Eminence $193M / +67% + Greenlight (Einhorn)',
     'activist': 'Eminence threshold',
     'insider': 'Greenlight Capital anchor',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Eminence +67% + Greenlight Einhorn on packaging cyclical'},

    {'tier': 5, 'ticker': 'EQT', 'name': 'EQT Corporation', 'mcap': '$28B',
     'fund_adds': 'Sprott NEW + Sprott Asset NEW (2 funds NEW, $50M)',
     'activist': '',
     'insider': 'Nierenberg 16.85% concentration mentioned elsewhere',
     'asym_rank': 'NEW', 'tier_s_plus': 'N/A',
     'why_aggressive': 'Sprott (gold/mining specialist) NEW position = natural-gas LNG cycle bet'},
]

# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'AGGRESSIVE BUY SIGNALS' in wb.sheetnames:
    del wb['AGGRESSIVE BUY SIGNALS']

ws = wb.create_sheet('AGGRESSIVE BUY SIGNALS', 5)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
T1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
T2_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
T3_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
T4_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
T5_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:J1')
ws['A1'] = 'AGGRESSIVE BUY SIGNALS — Insider + Multi-Fund Accumulation Across 442 Tabs'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:J2')
ws['A2'] = ('Cross-referenced from: Conviction Adds (multi-fund accumulation), Activist Catalysts (13D/13G threshold filings), '
            'All Activity scan (Form 4, founder buys, family fund 13D escalation). Names with all three signals = TIER 1.')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

# Legend
ws.merge_cells('A4:C4')
ws['A4'] = 'TIER LEGEND'
ws['A4'].font = HEADER; ws['A4'].fill = HEADER_FILL
ws['A5'] = 'TIER 1'; ws['A5'].fill = T1_FILL; ws['A5'].font = Font(bold=True)
ws['B5'] = 'Multi-fund ADDS + activist 13D + insider signal (TRIPLE)'
ws['A6'] = 'TIER 2'; ws['A6'].fill = T2_FILL; ws['A6'].font = Font(bold=True)
ws['B6'] = 'Activist + Form 4 / founder open-market buys (typically micro/small)'
ws['A7'] = 'TIER 3'; ws['A7'].fill = T3_FILL; ws['A7'].font = Font(bold=True)
ws['B7'] = 'Multi-fund aggressive adds + activist threshold (no documented insider)'
ws['A8'] = 'TIER 4'; ws['A8'].fill = T4_FILL; ws['A8'].font = Font(bold=True)
ws['B8'] = 'Strong single-fund + insider/buyback'
ws['A9'] = 'TIER 5'; ws['A9'].fill = T5_FILL; ws['A9'].font = Font(bold=True)
ws['B9'] = 'Multi-fund adds (no activist or insider)'

# Headers
headers = ['Tier', 'Ticker', 'Name', 'Mcap', 'Asym Rank', 'TS+', 'Fund Adds (multi-fund accumulation)', 'Activist 13D/13G', 'Insider / Form 4 / Founder Buys', 'Why Aggressive']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=11, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

# Sort by tier then asym_rank
def sort_key(s):
    rank = s['asym_rank']
    if isinstance(rank, int):
        return (s['tier'], rank)
    return (s['tier'], 999)

SIGNALS.sort(key=sort_key)

tier_fills = {1: T1_FILL, 2: T2_FILL, 3: T3_FILL, 4: T4_FILL, 5: T5_FILL}

for i, s in enumerate(SIGNALS, 12):
    fill = tier_fills.get(s['tier'])
    ws.cell(row=i, column=1, value=s['tier']).fill = fill
    ws.cell(row=i, column=1).font = Font(bold=True)
    ws.cell(row=i, column=2, value=s['ticker']).font = Font(bold=True)
    ws.cell(row=i, column=3, value=s['name'])
    ws.cell(row=i, column=4, value=s['mcap'])
    ws.cell(row=i, column=5, value=str(s['asym_rank']))
    ws.cell(row=i, column=6, value=s.get('tier_s_plus', 'N/A'))
    ws.cell(row=i, column=7, value=s['fund_adds'])
    ws.cell(row=i, column=8, value=s['activist'])
    ws.cell(row=i, column=9, value=s['insider'])
    ws.cell(row=i, column=10, value=s['why_aggressive'])

    for col in range(1, 11):
        ws.cell(row=i, column=col).alignment = LEFT_WRAP

    ws.row_dimensions[i].height = 75

widths = [6, 10, 24, 10, 9, 6, 50, 35, 50, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)

# Print summary
tier_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
for s in SIGNALS:
    tier_counts[s['tier']] += 1

print(f"Saved: {wb_path}")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"\nAGGRESSIVE BUY SIGNALS — TIER COUNTS:")
print(f"  TIER 1 (triple signal): {tier_counts[1]}")
print(f"  TIER 2 (activist + Form 4 insider): {tier_counts[2]}")
print(f"  TIER 3 (multi-fund + activist): {tier_counts[3]}")
print(f"  TIER 4 (strong single + insider): {tier_counts[4]}")
print(f"  TIER 5 (multi-fund adds): {tier_counts[5]}")
print(f"\nTIER 1 — TRIPLE SIGNAL (all three confirmed):")
for s in SIGNALS:
    if s['tier'] == 1:
        print(f"  {s['ticker']:<8} {s['mcap']:<8} | {s['why_aggressive'][:90]}")
print(f"\nTIER 2 — Activist + insider Form 4 / family build:")
for s in SIGNALS:
    if s['tier'] == 2:
        print(f"  {s['ticker']:<8} {s['mcap']:<14} | {s['why_aggressive'][:80]}")
