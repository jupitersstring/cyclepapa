"""
Add new fund tabs to fund_activity_last_6mo.xlsx using same structure as existing tabs.

Each new tab follows the format:
  Row 0: Fund Name
  Row 1: "Group: <group>"
  Row 2: "Sources:"
  Rows 3-9: source URLs
  Row 10: "(1) Highest conviction positions (recent adds)"
  Row 11: Ticker | Company | % Portfolio | $ Value | Change | Source
  Rows 12-N: data rows
  Blank row
  Row X: "(2) >=5% / threshold disclosures"
  Row X+1: Ticker | Company | % of Co. | Filing Date | Type | Source
  Rows: data
  Blank row
  "(3) New positions sized large"  | Ticker | % of Portfolio at Init | Quarter | Source
  Blank row
  "(4) Existing positions materially increased" | Ticker | Prior % | New % | Quarter | Source
"""
import openpyxl
from openpyxl.styles import Font, Alignment


def add_fund_tab(wb, fund_name, group, sources, conviction, disclosures, new_inits, mat_adds):
    """
    Add a new tab to the workbook with standard fund-tab structure.

    fund_name: str e.g. "TCI Fund Management"
    group: str e.g. "UK Activists - Tier 1"
    sources: list of source URLs
    conviction: list of dicts: {ticker, company, pct, value, change, source}
    disclosures: list of dicts: {ticker, company, pct_co, filing_date, type, source}
    new_inits: list of dicts: {ticker, pct_at_init, quarter, source}
    mat_adds: list of dicts: {ticker, prior_pct, new_pct, quarter, source}
    """
    # Sheet name truncated to 31 chars (Excel limit)
    sheet_name = fund_name[:31]
    if sheet_name in wb.sheetnames:
        print(f"  Tab '{sheet_name}' already exists, skipping")
        return
    ws = wb.create_sheet(sheet_name)

    # Header
    bold = Font(bold=True)
    ws.cell(row=1, column=1, value=fund_name).font = bold
    ws.cell(row=2, column=1, value=f"Group: {group}")
    ws.cell(row=3, column=1, value="Sources:")
    for i, src in enumerate(sources):
        ws.cell(row=4 + i, column=1, value=f"- {src}")

    # Determine first data section start (allow for source list)
    row = 4 + max(len(sources), 6) + 2  # 2-row buffer

    # Section 1: Highest conviction positions (recent adds)
    ws.cell(row=row, column=1, value="(1) Highest conviction positions (recent adds)").font = bold
    row += 1
    headers = ['Ticker', 'Company', '% of Portfolio', '$ Value', 'Change', 'Source']
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h).font = bold
    row += 1
    for entry in conviction:
        ws.cell(row=row, column=1, value=entry.get('ticker', ''))
        ws.cell(row=row, column=2, value=entry.get('company', ''))
        ws.cell(row=row, column=3, value=entry.get('pct', ''))
        ws.cell(row=row, column=4, value=entry.get('value', ''))
        ws.cell(row=row, column=5, value=entry.get('change', ''))
        ws.cell(row=row, column=6, value=entry.get('source', ''))
        row += 1
    row += 1  # blank

    # Section 2: >=5% / threshold disclosures
    ws.cell(row=row, column=1, value="(2) >=5% / threshold disclosures").font = bold
    row += 1
    headers = ['Ticker', 'Company', '% of Co.', 'Filing Date', 'Type', 'Source']
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h).font = bold
    row += 1
    for entry in disclosures:
        ws.cell(row=row, column=1, value=entry.get('ticker', ''))
        ws.cell(row=row, column=2, value=entry.get('company', ''))
        ws.cell(row=row, column=3, value=entry.get('pct_co', ''))
        ws.cell(row=row, column=4, value=entry.get('filing_date', ''))
        ws.cell(row=row, column=5, value=entry.get('type', ''))
        ws.cell(row=row, column=6, value=entry.get('source', ''))
        row += 1
    row += 1

    # Section 3: New positions sized large
    ws.cell(row=row, column=1, value="(3) New positions sized large").font = bold
    row += 1
    headers = ['Ticker', '% of Portfolio at Init', 'Quarter', 'Source']
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h).font = bold
    row += 1
    for entry in new_inits:
        ws.cell(row=row, column=1, value=entry.get('ticker', ''))
        ws.cell(row=row, column=2, value=entry.get('pct_at_init', ''))
        ws.cell(row=row, column=3, value=entry.get('quarter', ''))
        ws.cell(row=row, column=4, value=entry.get('source', ''))
        row += 1
    row += 1

    # Section 4: Existing positions materially increased
    ws.cell(row=row, column=1, value="(4) Existing positions materially increased").font = bold
    row += 1
    headers = ['Ticker', 'Prior %', 'New %', 'Quarter', 'Source']
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h).font = bold
    row += 1
    for entry in mat_adds:
        ws.cell(row=row, column=1, value=entry.get('ticker', ''))
        ws.cell(row=row, column=2, value=entry.get('prior_pct', ''))
        ws.cell(row=row, column=3, value=entry.get('new_pct', ''))
        ws.cell(row=row, column=4, value=entry.get('quarter', ''))
        ws.cell(row=row, column=5, value=entry.get('source', ''))
        row += 1

    # Column widths
    for col, w in enumerate([12, 28, 12, 14, 40, 60], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w


def main():
    wb_path = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'
    wb = openpyxl.load_workbook(wb_path)
    print(f"Loaded workbook with {len(wb.sheetnames)} sheets")

    # ============================================================
    # Add fund data here as agent outputs arrive
    # Each call to add_fund_tab() = one new fund tab
    # ============================================================

    # Example template - replace with actual agent output data:
    # add_fund_tab(wb,
    #     fund_name="TCI Fund Management",
    #     group="UK Activists - Tier 1",
    #     sources=["https://13f.info/manager/...", "https://whalewisdom.com/filer/..."],
    #     conviction=[
    #         {'ticker': 'V', 'company': 'Visa', 'pct': '12.5%', 'value': '$5.2B',
    #          'change': '+8% Q3->Q4', 'source': '...'},
    #     ],
    #     disclosures=[
    #         {'ticker': 'XYZ', 'company': 'XYZ Corp', 'pct_co': '5.2%',
    #          'filing_date': '2026-03-15', 'type': '13D', 'source': '...'},
    #     ],
    #     new_inits=[],
    #     mat_adds=[],
    # )

    wb.save(wb_path)
    print(f"Saved. New sheet count: {len(wb.sheetnames)}")


if __name__ == '__main__':
    main()
