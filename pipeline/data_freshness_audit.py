"""DATA FRESHNESS AUDIT across all 455 fund tabs.

Check each fund tab for:
  - Latest quarter mentioned (Q1 2026 = freshest; Q4 2025 = recent; older = stale)
  - Source links present (13f.info, secform4, stockzoa, SEC EDGAR, etc.)
  - Q1 2026 13F filing data (filed by May 15, 2026)
  - Form 4 dates within last 6 months
  - 13D/13G amendment dates within last 12 months

Flag stale tabs that need refresh.
"""
import openpyxl
import re
from collections import defaultdict
from datetime import datetime

wb = openpyxl.load_workbook('/home/user/cyclepapa/fund_activity_last_6mo.xlsx', data_only=True)

SKIP = {'Cover', 'Index', 'All Activity', 'Asymmetric Summary', 'Consensus Buys',
        'Highest Conviction', 'Conviction Adds', 'Micro-Cap Conviction Adds',
        'Activist Catalysts', 'Multi-Fund New Inits'}

freshness_data = {}

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP:
        continue
    ws = wb[sheet_name]

    record = {
        'sheet': sheet_name,
        'has_q1_2026': False,
        'has_q4_2025': False,
        'has_q3_2025': False,
        'has_older': False,
        'has_form4': False,
        'has_13d': False,
        'has_13g': False,
        'source_links': 0,
        'total_rows': 0,
        'recent_dates': [],
        'oldest_year': None,
        'newest_year': None,
        'data_richness': 0,  # rows with content / total rows
    }

    all_text = []
    rows_with_data = 0
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        record['total_rows'] += 1
        text = ' '.join(str(c) for c in row if c is not None).strip()
        if text:
            rows_with_data += 1
            all_text.append(text)

    full_text = ' '.join(all_text)
    tl = full_text.lower()

    # Quarter detection
    if 'q1 2026' in tl or 'q1/2026' in tl or '2026-q1' in tl or 'march 2026' in tl or 'mar 2026' in tl or 'apr 2026' in tl or 'may 2026' in tl:
        record['has_q1_2026'] = True
    if 'q4 2025' in tl or 'q4/2025' in tl or '2025-q4' in tl or 'dec 2025' in tl or 'november 2025' in tl or 'nov 2025' in tl:
        record['has_q4_2025'] = True
    if 'q3 2025' in tl or 'q3/2025' in tl or '2025-q3' in tl:
        record['has_q3_2025'] = True
    if any(yr in full_text for yr in ['2024', '2023', '2022', '2021', '2020']) and not any(yr in full_text for yr in ['2025', '2026']):
        record['has_older'] = True

    # Filing types
    if 'form 4' in tl:
        record['has_form4'] = True
    if '13d' in tl or '13D' in full_text:
        record['has_13d'] = True
    if '13g' in tl:
        record['has_13g'] = True

    # Source links
    record['source_links'] = full_text.count('http')

    # Date extraction
    dates = re.findall(r'(20\d{2})[-/](\d{1,2})[-/](\d{1,2})', full_text)
    if dates:
        years = [int(d[0]) for d in dates]
        record['newest_year'] = max(years)
        record['oldest_year'] = min(years)
        # Latest date in 2026
        dates_2026 = [d for d in dates if d[0] == '2026']
        if dates_2026:
            record['recent_dates'] = sorted([f"{d[0]}-{d[1]}-{d[2]}" for d in dates_2026])[-3:]

    # Data richness score
    if record['total_rows']:
        record['data_richness'] = round(rows_with_data / record['total_rows'], 2)
    else:
        record['data_richness'] = 0

    freshness_data[sheet_name] = record


# Categorize tabs
fresh_tabs = []      # Q1 2026 data + sources
recent_tabs = []     # Q4 2025 data + sources
stale_tabs = []      # Older or missing Q1/Q4 2026
sparse_tabs = []     # Few rows / minimal data

for sheet, rec in freshness_data.items():
    if rec['total_rows'] < 12:
        sparse_tabs.append(sheet)
    elif rec['has_q1_2026']:
        fresh_tabs.append(sheet)
    elif rec['has_q4_2025']:
        recent_tabs.append(sheet)
    else:
        stale_tabs.append(sheet)


print(f"=== DATA FRESHNESS AUDIT — 445 fund tabs ===")
print(f"FRESH (Q1 2026 data):     {len(fresh_tabs)} tabs ({100*len(fresh_tabs)//len(freshness_data)}%)")
print(f"RECENT (Q4 2025 data):    {len(recent_tabs)} tabs ({100*len(recent_tabs)//len(freshness_data)}%)")
print(f"STALE (older data):       {len(stale_tabs)} tabs ({100*len(stale_tabs)//len(freshness_data)}%)")
print(f"SPARSE (<12 rows):        {len(sparse_tabs)} tabs ({100*len(sparse_tabs)//len(freshness_data)}%)")

print(f"\n=== STALE TABS (no Q1 or Q4 2025 reference) ===")
for s in stale_tabs[:30]:
    rec = freshness_data[s]
    print(f"  {s:<40} rows={rec['total_rows']:<4} links={rec['source_links']:<3} newest_yr={rec['newest_year']}")

print(f"\n=== SPARSE TABS (< 12 rows = likely skeleton or incomplete) ===")
for s in sparse_tabs[:30]:
    rec = freshness_data[s]
    print(f"  {s:<40} rows={rec['total_rows']:<4} richness={rec['data_richness']}")

print(f"\n=== HEALTHY FRESH TABS sample (Q1 2026 + multiple source links) ===")
healthy = [s for s in fresh_tabs if freshness_data[s]['source_links'] >= 3]
print(f"Total healthy: {len(healthy)} tabs")
for s in healthy[:15]:
    rec = freshness_data[s]
    dates = ' '.join(rec['recent_dates'][-2:]) if rec['recent_dates'] else ''
    print(f"  {s:<40} rows={rec['total_rows']:<4} links={rec['source_links']:<3} recent={dates}")

# Save categorization
import json
with open('/tmp/freshness_audit.json', 'w') as f:
    json.dump({
        'fresh': fresh_tabs,
        'recent': recent_tabs,
        'stale': stale_tabs,
        'sparse': sparse_tabs,
    }, f)
print(f"\nCategorization saved to /tmp/freshness_audit.json")
