"""Build investment archetype workbook with valuation, thesis, and ranking."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
BOLD = Font(bold=True, size=11)
HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ================================================================
# ARCHETYPE DATA STRUCTURE
# Each idea: ticker, mcap, price, val_metrics, thesis, catalyst,
#            variant, asym_within, upside, downside, rr, cross_rank
# ================================================================

ARCHETYPES = {
    "1. Yartseva Pure Multibagger": {
        "definition": "Sub-$2B mcap; B/M >0.40; FCF yield >5%; ROA >0; near 52w low; 6mo negative. Pure quantitative multibagger setup per CAFE WP 33 (Feb 2025).",
        "names": [
            {"ticker": "NRP", "name": "Natural Resource Partners", "mcap": "$1.35B", "price": "$105.81",
             "valuation": "P/B 2.26 (B/M 0.47), 14.5% FCF yield, ROA 10.5%, 6.6x P/FCF",
             "thesis": "Royalty trust on 13M mineral acres + 49% Sisecam soda ash JV. Near-zero debt by mid-2026 after $109M paydown + preferred retirement. Distribution increase guided for November 2026 (delayed from prior guide due to Sisecam capital call).",
             "catalyst": "Nov 2026 distribution step-up + preferreds retire; Q3 2026 Sisecam recovery sig",
             "variant": "Market: dying coal MLP. Smart money: debt-free royalty with 12% FCF yield, comparable to TPL at 25x P/CF (NRP at 6x).",
             "asym": 9.0, "upside": "2.5-3.0x", "downside": "-20%", "rr": "13:1", "cross_rank": 4,
             "smart_money": "Saber/Huber 17.75%, Greystone top, Right Tail NEW Q1 2026, Berkowitz/Fairholme"},

            {"ticker": "GLOB", "name": "Globant SA", "mcap": "$1.68B", "price": "$43.50",
             "valuation": "P/E 5.91 (sector low), P/B 0.78, P/FCF 5.56, 17% FCF yield, EV/EBITDA 4.95",
             "thesis": "LatAm IT services with AI Pods inflection ($32.8M ARR from $0 in 12 months; $352M pipeline; 40% top-20 client penetration). Founder Migoya 23-year tenure + ~3% via Founders Trust.",
             "catalyst": "Q3 2026 AI Pods print + Argentina FX tailwind + ENRGI platform launch H2",
             "variant": "Market: melting India-IT comp. Smart money: AI mix shift to higher-margin work + LatAm cost arb. CAVEAT: securities class action (lead plaintiff June 23, 2026); COO resigned July 2025 not replaced.",
             "asym": 9.0, "upside": "4-5x", "downside": "-25%", "rr": "16:1", "cross_rank": 11,
             "smart_money": "Cartica historical 6.5% Q4 2022; sentiment bottom 16% short float"},

            {"ticker": "CRTO", "name": "Criteo SA", "mcap": "$866M", "price": "$13.50",
             "valuation": "P/E 3.73 (fwd), P/B 0.79, P/FCF 4.98, EV/EBITDA 2.28, 20.7% FCF yield",
             "thesis": "Performance ad-tech pivoting to Commerce Yield retail media (235 retailers incl Lowes/Costco). $1B+ media spend Q1 2026 (+8% cc). OpenAI partnership live with 1,000+ brands. Q1 retail media -31% ex-scope -- actual underlying +24%.",
             "catalyst": "Q3 2026 anniversary of scope losses + Luxembourg redomiciliation + $200M buyback (22% cap)",
             "variant": "Market prices terminal decline; Petrus Advisers activist + 33% FCF yield + retail media TAM $200B by 2028.",
             "asym": 8.0, "upside": "2.5-3.5x", "downside": "-25%", "rr": "12:1", "cross_rank": 15,
             "smart_money": "Nierenberg D3 9.37% / +19% Q1 2026; Petrus Advisers activist"},

            {"ticker": "TREE", "name": "LendingTree", "mcap": "$525M", "price": "$37.50",
             "valuation": "P/E 2.92, P/FCF 7.2, ROE 88.6%, B/M 0.60",
             "thesis": "Three-segment marketplace (Insurance 68%, Consumer 20%, Home 12%) at Q1 2026 inflection: rev +37%, VMM +28%, EBITDA +71%, Insurance +51%. Net leverage falling 3.4x→2.1x; S&P upgraded.",
             "catalyst": "Q2 2026 print + $575M July 2025 converts retired + Fed rate cuts unlock Home/Consumer",
             "variant": "Street disbelieves Insurance durability. Variant: insurance carriers in year-2 of 5-7yr up-cycle = $200M+ EBITDA into 2027.",
             "asym": 9.0, "upside": "3-4x", "downside": "-30%", "rr": "10:1", "cross_rank": 7,
             "smart_money": "Yartseva 7/7 + post-Q1 inflection demonstrated"},

            {"ticker": "OPRX", "name": "OptimizeRx", "mcap": "$93M", "price": "$4.82",
             "valuation": "P/B 0.70 (sub-book), P/FCF 4.89, 15.7% FCF yield, EV/Sales 0.88, Adj EBITDA 4.4x EV",
             "thesis": "Digital health POC pharma messaging to 60%+ US prescribers. May 21 2026 DeepIntent integration = first DSP accessing OPRX EHR network (Q3 2026 launch) opens >80% of digital pharma budgets.",
             "catalyst": "Q4 2026 DSP integration revenue start + Q2 2026 earnings Aug 14",
             "variant": "Market: MFN victim; Smart money: profitable ($19M FCF) at 0.7x book, target $11.50 (138% upside).",
             "asym": 9.0, "upside": "3.5-6x", "downside": "-45%", "rr": "8:1", "cross_rank": 13,
             "smart_money": "Insider 14.82%; $10M buyback through March 2027; Doximity exec Presti to board April 2026"},

            {"ticker": "VITL", "name": "Vital Farms", "mcap": "$1.5B", "price": "$13",
             "valuation": "B/M 0.74, ROA 11% (BEST in 6/7 tier), -68.76% from peak",
             "thesis": "Pasture-raised egg category creator with brand moat. Walmart/Target/Whole Foods distribution. 11% ROA sustained THROUGH the avian flu crash = exceptional earnings power.",
             "catalyst": "Avian flu egg-supply normalization + summer pricing + M&A speculation",
             "variant": "Market lumped premium VITL with commodity egg crash. Reality: brand pricing power retained. Strategic buyers: Kraft Heinz, Hain, General Mills.",
             "asym": 8.0, "upside": "3-5x", "downside": "-30%", "rr": "13:1", "cross_rank": 6,
             "smart_money": "No major fund concentrated yet = entry preserved; Yartseva 6/7"},

            {"ticker": "CARS", "name": "Cars.com", "mcap": "$560M", "price": "$10",
             "valuation": "P/B 0.83, P/FCF 3.5, 28% TTM FCF yield, EV/EBITDA 2.7x",
             "thesis": "Digital dealer SaaS (19,390 dealers x $2,473 ARPD/mo, 90% recurring). Q1 EPS $0.45 vs $0.13 est (+246%). Subscription +2%; OEM/National -12% (cyclical bottom).",
             "catalyst": "$25-30M annualized cost program 2027 + $90M buyback (16% cap raised from $60M)",
             "variant": "Street: declining auto Yellow Pages at 4x P/FCF. Variant: 90% SaaS dealer recurring + buyback math.",
             "asym": 7.0, "upside": "2-2.5x", "downside": "-25%", "rr": "8:1", "cross_rank": 24,
             "smart_money": "Yartseva 7/7 - smart money still building"},

            {"ticker": "DV", "name": "DoubleVerify", "mcap": "$1.50B", "price": "$9",
             "valuation": "P/FCF 11x (trough), EV/EBITDA 9.6x, EV/Sales 1.9x, 9% FCF yield",
             "thesis": "Pure-play ad verification (68% market share in fraud detection vs IAS 2.7%). 4,300+ customers vs IAS 169. Q1 2026 rev +10%, EBITDA 31% margin. CTV +28%, social activation +92%.",
             "catalyst": "ABS streaming Do Not Air list rollout Q2/Q3 + Meta activation $12M ARR to $50M+ path",
             "variant": "Market: Google MRC self-measurement threat. Variant: CTV verification underpenetrated (10%) vs display (70%).",
             "asym": 8.0, "upside": "3-3.5x", "downside": "-25%", "rr": "12:1", "cross_rank": 17,
             "smart_money": "Insider 12.69% (high for ad-tech); Yartseva 7/7"},

            {"ticker": "ROCK", "name": "Gibraltar Industries", "mcap": "$1.10B", "price": "$39.69",
             "valuation": "P/B 1.34, FCF yield 6.5%, ROA 2.95%, EV/Sales 1.0x vs peer 1.5x",
             "thesis": "3-segment industrial (Residential 80%, Renewables, Agtech). OmniMax acquisition closed Feb 2026 ($1.335B) = #1 in residential roofing accessories. FY26 guide $1.76-1.83B rev, $3.65-4.05 EPS.",
             "catalyst": "Q2 2026 earnings = first clean OmniMax quarter; aluminum normalization; Q3 reroofing season",
             "variant": "Market: busted SPAC with 1.56 D/E. Variant: one-time integration year; FY27 EPS $5+ at 15x = $75-90.",
             "asym": 7.0, "upside": "1.8-2.2x", "downside": "-25%", "rr": "8:1", "cross_rank": 26,
             "smart_money": "Director Metcalf bought $502k March 2026; Vanguard 5.24% 13G"},

            {"ticker": "HRMY", "name": "Harmony Biosciences", "mcap": "$1.77B", "price": "$31",
             "valuation": "P/B 1.91, P/Sales 2.05, FCF yield 19.36%, 26% net cash position",
             "thesis": "$1B+ WAKIX franchise (pitolisant for narcolepsy). Q1 2026 rev +17% to $215M; pediatric FDA approval Feb 13 2026 extends IP to 2030 (pediatric exclusivity to 2040s). 5 Phase 3 trials in pipeline.",
             "catalyst": "Pitolisant GR NDA Q2 2026 (PDUFA Q1 2027); BP-205 Phase 1 PK mid-2026",
             "variant": "Market: single-product cliff. Smart money: Lilly's $6.3B Centessa (Q3 2026 close) validated orexin space + IP estate.",
             "asym": 9.0, "upside": "2.5-3.5x", "downside": "-30%", "rr": "12:1", "cross_rank": 5,
             "smart_money": "M&A buyers: Jazz Pharma, Lundbeck, Otsuka; Catalyst Pharma precedent"},

            {"ticker": "INMD", "name": "InMode Ltd", "mcap": "$869M", "price": "$13.71",
             "valuation": "P/B 1.32, FCF $537M cash = $8.47/sh (62% of price), 77.8% gross margin",
             "thesis": "Israeli aesthetic device leader (Morpheus8, BodyTite). 85% recurring consumables. Steel Partners +155% Q1 2026 = 25.5% of their book; founder Mizrahy bought 800k sh ($10.7M) Feb 2026.",
             "catalyst": "Q2 2026 earnings ~Aug 5 + potential Steel renewed bid ($18 prior)",
             "variant": "Market: Israeli geopolitical + aesthetic cycle. Smart money: $11-12 floor on cash alone; Steel implied $20+ takeout.",
             "asym": 8.0, "upside": "2.2x", "downside": "-15%", "rr": "10:1", "cross_rank": 16,
             "smart_money": "Steel Partners 25.5% of book; Mizrahy +800k sh ($10.7M)"},

            {"ticker": "FTLF", "name": "FitLife Brands", "mcap": "$93M", "price": "$10",
             "valuation": "P/B 2.05, B/M 0.49, FCF yield 8.2%, ROA 7.2%, near 52w low, -43.9% 6mo",
             "thesis": "Microcap nutraceutical at $93M cap. Sub-institutional minimum. Classic Yartseva discarded-microcap profile with positive FCF + ROA + extreme drawdown.",
             "catalyst": "Q2 2026 earnings; M&A speculation; recovery in supplement category",
             "variant": "Market: forgotten micro. Smart money: pure Yartseva mean-reversion target.",
             "asym": 8.0, "upside": "3-5x", "downside": "-30%", "rr": "13:1", "cross_rank": 14,
             "smart_money": "Smoak Capital residual position; previously a Smoak top-5"},

            {"ticker": "MTY.TO", "name": "MTY Food Group", "mcap": "C$887M", "price": "C$40.41",
             "valuation": "P/B 1.04 (trading at book!), EV/EBITDA 7.5x (vs QSR peer 14-18x), 16.7% FCF yield",
             "thesis": "Multi-brand Canadian QSR franchisor (90 banners incl Cold Stone, Mucho Burrito, Papa Murphy's). 7,000+ locations, 98%+ franchised. Q1 EPS C$0.98 BEAT vs C$0.80 cons.",
             "catalyst": "Q2 2026 earnings test SSS stabilization + NCIB 5% of float through July",
             "variant": "Market: melting ice cube. Variant: SSS troughing + debt paying down + M&A pool (RBI, Inspire at 12-14x EBITDA).",
             "asym": 8.0, "upside": "2-2.5x", "downside": "-15%", "rr": "10:1", "cross_rank": 19,
             "smart_money": "CEO Lefebvre open-market buy at C$30.54 (+33% stake)"},
        ]
    },

    "2. Activist Board Catalyst": {
        "definition": "Schedule 13D filing with board seat secured or imminent + activist demands for value unlock. Dated trigger creates near-term volatility around catalyst.",
        "names": [
            {"ticker": "FUN", "name": "Six Flags Entertainment", "mcap": "$2.2B", "price": "$25",
             "valuation": "$5.27B net debt = 7x leverage; EV/EBITDA 9x on $1.05B 2027 EBITDA",
             "thesis": "Post-Cedar Fair/Six Flags merger consolidation. 42 parks irreplaceable real estate. Q1 2026 rev +12%, attendance +4%, per-cap +6%. $200M synergy target ($120M cost + $80M growth).",
             "catalyst": "May 26 2026 AGM: Jaffer joins board (Class III through 2027, Audit&Finance Cttee). Jana 9% pushing OUTRIGHT SALE.",
             "variant": "Street: broken merger story. Two-front activist pressure (governance + sale) + per-park RE floor $47M each.",
             "asym": 8.5, "upside": "2-3x (3x sale)", "downside": "-30%", "rr": "9:1", "cross_rank": 8,
             "smart_money": "H Partners 53.7% / $82.5M / 5.7% of class; Jana 9% + Travis Kelce co-invest"},

            {"ticker": "AAP", "name": "Advance Auto Parts", "mcap": "$3.3B", "price": "$54.42",
             "valuation": "3.8% op margin (rising +410bps), FY26 guide $2.40-3.10 EPS; sub-1x sales; M&A floor 0.8x sales = $80/sh",
             "thesis": "#3 aftermarket retailer post-WorldPac sale. Q1 +3.5% comps (Pro mid-single, DIY low-single). 700+ stores closed; ~1,300 strategic locations remain. Bridge to AZO-style 8-10% margins = $130-150/sh.",
             "catalyst": "Q2 2026 earnings Aug = needs comps +2%+ and op margin >5%; H Partners 13D escalation possible",
             "variant": "Consensus PT $57 (dead money). H Partners 46% concentration implies AAP returns to AZO margins.",
             "asym": 8.0, "upside": "2.5-3.0x", "downside": "-30%", "rr": "9:1", "cross_rank": 10,
             "smart_money": "H Partners 46.3% / $71.2M / +50% sh Q1 2026 (1.35M sh); Legion Partners precedent"},

            {"ticker": "MRP", "name": "Millrose Properties", "mcap": "$4.59B", "price": "$25",
             "valuation": "B/M 1.274 (DEEPEST in 6/7 tier; deeper than NRP's 0.47), pure NAV play",
             "thesis": "Lennar 80%+ owner spinoff with 100k+ optioned lots at cost basis. Land bank trading below NAV with Lennar as guaranteed off-take partner. Housing-cycle recovery levered.",
             "catalyst": "Q2 2026 print + Lennar volume scaling + housing cycle inflection",
             "variant": "Market priced spinoff orphan technical pressure. Smart money: pure land cash-cow at discount to NAV.",
             "asym": 7.5, "upside": "2-3x", "downside": "-20%", "rr": "12:1", "cross_rank": 22,
             "smart_money": "Lennar 80%+ owner = patient capital lock-in; SITG multi-fund flag"},

            {"ticker": "QRHC", "name": "Quest Resource Holding", "mcap": "$200M", "price": "$7",
             "valuation": "Sub-$200M micro; recurring waste services revenue model",
             "thesis": "Wynnefield filed 13D 13.3% May 2025; Robert Lipstein joined board. Cooperation Agreement w/ standstill until 30 days before 2027 nomination window. Asset-light recurring waste mgmt.",
             "catalyst": "Operational improvements + Wynnefield additional buying if dips",
             "variant": "Market: small-cap value trap. Activist: board seat + potential value-unlock initiatives.",
             "asym": 7.0, "upside": "2-3x", "downside": "-30%", "rr": "8:1", "cross_rank": 28,
             "smart_money": "Wynnefield 13.3% / 2.73M sh / 13D w/ board seat"},

            {"ticker": "POSTBPB", "name": "Potbelly Corp (PBPB)", "mcap": "$240M", "price": "$8",
             "valuation": "Sub-$300M micro restaurant chain; turnaround stage",
             "thesis": "Nierenberg/D3 is LARGEST shareholder (~$16.4M / 7% of D3 book). Customer-traffic gains while peers don't (Q1 2026 letter thesis). Form 4 insider buying by Nierenberg at depressed levels.",
             "catalyst": "Restaurant turnaround + Nierenberg potential escalation",
             "variant": "Market: tier-2 restaurant chain. Activist soft-pressure thesis.",
             "asym": 7.0, "upside": "2-3x", "downside": "-30%", "rr": "8:1", "cross_rank": 29,
             "smart_money": "Nierenberg/D3 largest shareholder; 13D/A active March 2025"},

            {"ticker": "ACHC", "name": "Acadia Healthcare", "mcap": "$3.2B", "price": "$33",
             "valuation": "EV/EBITDA 6.5x (vs HCA 9x, UHS 8x); 29% short float; -70% drawdown",
             "thesis": "Largest US behavioral health network (260+ facilities). DOJ Medicaid investigation overhang. Q1 2026 admissions stabilizing. Greenlight 4.1M activist; Sohn 5/12 pitch.",
             "catalyst": "DOJ resolution + Q2 2026 earnings + 29% short squeeze potential",
             "variant": "Market: DOJ/Medicaid existential. Greenlight: provider necessity + facility footprint + de-rate captured = floor near current.",
             "asym": 8.0, "upside": "2-2.5x", "downside": "-40%", "rr": "6:1", "cross_rank": 11,
             "smart_money": "Greenlight Capital 4.1M sh (David Einhorn); Sohn presentation 5/12/2026"},
        ]
    },

    "3. Biotech Multi-Fund Convergence": {
        "definition": "Two or more top biotech crossover specialists (Baker Bros, RA Capital, BVF, Stonepine, Suvretta) anchored at high conviction in same name. Convergence is the high-signal pattern.",
        "names": [
            {"ticker": "CELC", "name": "Celcuity", "mcap": "$8B (post-data)", "price": "$165",
             "valuation": "$387M cash + accepted NDA + Priority Review = de-risked; 4-5x peak rev precedent",
             "thesis": "Phase 3 VIKTORIA-1 PIK3CA-mut POSITIVE May 1 2026 (76% PFS hazard reduction). Gedatolisib pan-PI3K/mTOR oncology. Triplet vs alpelisib+fulvestrant. Multiple HR+/HER2- breast cancer cohorts.",
             "catalyst": "JUNE 2 ASCO LBA1008 oral + JULY 17 PDUFA goal date (2 catalysts in 8 weeks)",
             "variant": "Market: alpelisib-class peak ($1-2B). Baker Bros 19.99% 13D activist: $5-10B peak (triplet displaces alpelisib + captures WT segment).",
             "asym": 9.5, "upside": "1.3-2x base / 2-3x M&A", "downside": "-20%", "rr": "8:1", "cross_rank": 1,
             "smart_money": "Baker Bros 19.99% ACTIVIST 13D = $904M / 5.48% of $17.4B book"},

            {"ticker": "KYMR", "name": "Kymera Therapeutics", "mcap": "$6.5B", "price": "$80",
             "valuation": "$1.55B cash (runway 2029); Dec 2025 follow-on $602M at $86",
             "thesis": "Targeted protein degradation platform. Lead KT-621 oral STAT6 degrader competes with injectable Dupixent ($14B+ franchise). IRAK4 partnered with Sanofi.",
             "catalyst": "BROADEN2 AD mid-2027 (12+ months); BREADTH asthma late 2027; FDA Fast Track granted both",
             "variant": "Market: Phase 1b execution risk. Two top biotech specialists: oral Dupixent-disruptor with peak $5-10B.",
             "asym": 9.0, "upside": "3-5x", "downside": "-30%", "rr": "13:1", "cross_rank": 4,
             "smart_money": "BVF #1 position 14.6% of $3.14B book ($458M) + Baker Bros 4.1% / $721M; BOTH anchored $86 follow-on"},

            {"ticker": "PRLD", "name": "Prelude Therapeutics", "mcap": "$340M", "price": "$4.26",
             "valuation": "$175M cash vs $340M mcap = $2.20/sh net cash (51% of price); buyout floor near current",
             "thesis": "Precision oncology / molecular glue / degrader platform. SMARCA2 discontinued 2025; pivot to JAK2V617F (PRT12396) + KAT6A (PRT13722). AbCellera partnership.",
             "catalyst": "PRT12396 Phase 1 H2 2026; PRT13722 IND mid-2026; Incyte option agreement Nov 2025",
             "variant": "Market: broken story post-SMARCA2 fail. Smart money: April 2026 refunding = platform-validation thesis. JAK2 mutant-selective addresses 95% PV/60% ET/55% MF.",
             "asym": 8.5, "upside": "3-5x", "downside": "-30%", "rr": "13:1", "cross_rank": 9,
             "smart_money": "Baker Bros 15.5% 13D/A + RA Capital 9.99% 13G; BOTH anchored SAME $90M April financing at $4.44"},

            {"ticker": "PVLA", "name": "Palvella Therapeutics", "mcap": "$1.56B", "price": "$130",
             "valuation": "$261.9M cash = $25/sh (19% of price). Phase 3 derisked. M&A floor 4-6x peak rev",
             "thesis": "Rare-disease topical platform. QTORIN rapamycin gel for microcystic LMs (Breakthrough + Orphan + Fast Track). Phase 3 SELVA POSITIVE Feb 24 2026 (mLM-IGA +2.13, p<0.001).",
             "catalyst": "Pre-NDA FDA meeting Q2 2026; NDA H2 2026; PDUFA H1 2027",
             "variant": "Market: single-asset rare-disease (~$1.5B). Smart money: platform with >$1B peak QTORIN + venous malformations/angiokeratomas/pitavastatin programs.",
             "asym": 8.0, "upside": "2-3x", "downside": "-30%", "rr": "9:1", "cross_rank": 18,
             "smart_money": "BVF 3.8% / $120M (8.5% of company) + Suvretta 3.67% / $144M (+29.3% Q1)"},

            {"ticker": "ASND", "name": "Ascendis Pharma", "mcap": "$15.3B", "price": "$247",
             "valuation": "Profitable (22% non-IFRS op margin); €247M Q1 (2x YoY); peak rev YORVIPATH $4-5B",
             "thesis": "TransCon prodrug platform: SKYTROFA (GHD), YORVIPATH (hypopara), YUVIWEL (achondroplasia approved Feb 2026). >1,000 new US YORVIPATH enrollments Q1; 6,300+ patients across 2,700 prescribers.",
             "catalyst": "YUVIWEL ramp + Q2/Q3 YORVIPATH growth + TransCon CNP follow-on indications",
             "variant": "Sell-side: $15B endocrine specialty maturing. RA Capital: $30B+ at 70-90K addressable hypopara population. Analyst high $345 (Barclays).",
             "asym": 8.0, "upside": "1.4-1.8x", "downside": "-20%", "rr": "8:1", "cross_rank": 20,
             "smart_money": "RA Capital 24.9% of $9.44B = $2.35B = LARGEST single biotech crossover position; held since 2015"},

            {"ticker": "BNTC", "name": "Benitec Biopharma", "mcap": "$200M", "price": "$8",
             "valuation": "Sub-$300M biotech; gene therapy platform optionality",
             "thesis": "Gene therapy platform anchor position. Suvretta 44.1% 13D anchor stake (largest activist ownership in their portfolio). PIPE participation Nov 2025.",
             "catalyst": "Multiple gene-therapy programs in development; clinical readouts upcoming",
             "variant": "44.1% ownership creates board influence + asymmetric upside on clinical reads.",
             "asym": 7.0, "upside": "3-5x", "downside": "-50%", "rr": "8:1", "cross_rank": 30,
             "smart_money": "Suvretta 44.1% 13D (Suvretta's LARGEST stake by % ownership)"},
        ]
    },

    "4. Sponsor-Anchored Holdco Transformation": {
        "definition": "Activist or family-office anchor with significant ownership AND cost-basis floor that creates sponsor put. Holdco transformation thesis (insurance float, multi-asset compounding).",
        "names": [
            {"ticker": "HHH", "name": "Howard Hughes Holdings", "mcap": "$3.5B", "price": "$63.77",
             "valuation": "NAV $95-105/sh = 33-39% discount; $1.84B cash + $515M revolver liquidity",
             "thesis": "Ackman-led 'Modern Berkshire'. 3,813 acres residential MPC ($4.8B NAV) + 2,447 acres commercial ($2.0B) + Vantage specialty insurance ($2.2B 2027 premium). Marc Grandisson (ex-Arch Capital CEO) on board April 2026.",
             "catalyst": "Vantage close Q2 2026 (30-45 days); Park Ward Village condo closings Q2-Q3 (~$1B rev); first post-Vantage capital allocation",
             "variant": "Street: sub-scale RE dev with leverage. Ackman: insurance float + permanent capital. NAV $95-105 vs $63 = 50-67% margin of safety.",
             "asym": 9.0, "upside": "2-3x", "downside": "-20%", "rr": "12:1", "cross_rank": 3,
             "smart_money": "Ackman 47% via $900M @ $100/sh + $1B preferred; Northern Right NEW Q1 +12%"},
        ]
    },

    "5. Spinoff / SoTP Arbitrage": {
        "definition": "Sum-of-parts valuation discount with defined corporate action (spin-off, divestiture, breakup) as catalyst.",
        "names": [
            {"ticker": "KBR", "name": "KBR Inc", "mcap": "$3.9B", "price": "$30",
             "valuation": "SoTP: MTS at 11x EBITDA ($5.2B EV) + STS at 14x ($5.3B) = $55-65/sh combined",
             "thesis": "Two segments: MTS (gov services 72%, $18.5B backlog, cleared workforce 20k) + STS (process tech + TerraPower Natrium nuclear EPC alliance, 22% EBITDA margin). FY26 $980M-1.04B EBITDA, $3.87-4.22 EPS.",
             "catalyst": "MTS spin January 4 2027 (board-approved Sept 2025); NRC Natrium safety eval mid-2026",
             "variant": "Market: broken post-HomeSafe termination + class action overhang. Variant: HomeSafe disposed; spin unlocks SoTP value.",
             "asym": 8.5, "upside": "2-2.5x", "downside": "-25%", "rr": "10:1", "cross_rank": 12,
             "smart_money": "D3/Nierenberg +425% Q1 (largest single add in entire universe); Irenic activist ~1%; Director Form 4 buys May 2026"},

            {"ticker": "DLTR", "name": "Dollar Tree", "mcap": "$25B", "price": "$80",
             "valuation": "B/M 0.20 (Yartseva 6/7); Family Dollar divestiture proceeds = $1B+ cash bolster",
             "thesis": "Post-divestiture clean Dollar Tree franchise (16k+ stores). Multi-fund deep-value consensus Q1 2026. Mayer M6 turnaround pattern.",
             "catalyst": "Family Dollar sale completion + comparable sales recovery + multiple re-rate",
             "variant": "Market priced Family Dollar mistake. Smart money: post-divestiture clean franchise.",
             "asym": 7.5, "upside": "2-3x", "downside": "-20%", "rr": "11:1", "cross_rank": 21,
             "smart_money": "Multi-fund deep-value funds Q1 2026 adds"},
        ]
    },

    "6. Deep Drawdown Mean Reversion": {
        "definition": "Yartseva mechanics: extreme drawdown (>40% from peak) + retained profitability/cash flow + brand or category strength = mean reversion setup.",
        "names": [
            {"ticker": "VITL", "name": "Vital Farms", "mcap": "$1.5B", "price": "$13",
             "valuation": "B/M 0.74, ROA 11%, -68.76% from peak (DEEPEST drawdown in 6/7)",
             "thesis": "[See Archetype 1 for full thesis]",
             "catalyst": "Avian flu normalization + M&A + summer pricing",
             "variant": "[See Archetype 1]",
             "asym": 9.0, "upside": "3-5x", "downside": "-30%", "rr": "13:1", "cross_rank": 6,
             "smart_money": "No major fund concentrated yet"},

            {"ticker": "COTY", "name": "Coty Inc", "mcap": "$1.79B", "price": "$12",
             "valuation": "B/M 1.73 (DEEPEST in 6/7 tier), FCF yield 17.4%, but ROA -4.91% (Y3 fails)",
             "thesis": "Prestige/fragrance beauty (Burberry, Tiffany, CK, Calvin Klein brands). KKR overhang anchor. -37.5% drawdown extreme contrarian setup. Brand assets worth multiples of EV.",
             "catalyst": "Debt restructuring + KKR exit dynamics + cost program execution",
             "variant": "Market: earnings deterioration. Smart: 6-year FCF payback floor with brand asset coverage.",
             "asym": 7.0, "upside": "2-4x", "downside": "-30%", "rr": "9:1", "cross_rank": 25,
             "smart_money": "KKR legacy anchor; deep-value funds incrementally adding"},

            {"ticker": "FRPT", "name": "Freshpet", "mcap": "$2.53B", "price": "$60",
             "valuation": "B/M 0.50 (borderline Yartseva 6/7), ROA 3.41%, -40% drawdown",
             "thesis": "Premium fresh dog food brand leader. Manufacturing capacity coming online (Ennis plant). Margin inflection at 70%+ utilization. M&A target for Nestle Purina / Mars / General Mills.",
             "catalyst": "Capacity utilization + margin inflection + M&A speculation",
             "variant": "Market: consumer weakness + capex burden. Smart: brand pricing power + post-capex FCF inflection.",
             "asym": 7.5, "upside": "3-4x", "downside": "-30%", "rr": "10:1", "cross_rank": 18,
             "smart_money": "Marlowe Partners 39.8% in last 13F (highest single-fund signal)"},

            {"ticker": "OPRX", "name": "OptimizeRx", "mcap": "$93M", "price": "$4.82",
             "valuation": "P/B 0.70 sub-book, -65.4% 6mo drawdown",
             "thesis": "[See Archetype 1]",
             "catalyst": "DSP integration Q4 2026 + MFN clarity",
             "variant": "[See Archetype 1]",
             "asym": 9.0, "upside": "3.5-6x", "downside": "-45%", "rr": "8:1", "cross_rank": 13,
             "smart_money": "Insider 14.82%; $10M buyback"},
        ]
    },

    "7. Hard Asset / Royalty": {
        "definition": "Irreplaceable asset ownership (royalty, land, minerals) with no recurring capex requirement and cash-flow distribution structure.",
        "names": [
            {"ticker": "NRP", "name": "Natural Resource Partners", "mcap": "$1.35B", "price": "$105.81",
             "valuation": "[See Archetype 1]",
             "thesis": "[See Archetype 1]",
             "catalyst": "Nov 2026 distribution + preferreds retire",
             "variant": "[See Archetype 1]",
             "asym": 9.0, "upside": "2.5-3x", "downside": "-20%", "rr": "13:1", "cross_rank": 4,
             "smart_money": "Saber 17.75% + Greystone top + Right Tail NEW + Berkowitz"},

            {"ticker": "JOE", "name": "St. Joe Company", "mcap": "$2.8B", "price": "$48",
             "valuation": "Land bank: 170k acres NW Florida; GAAP B/M understates LIFO land basis",
             "thesis": "Florida Panhandle landbank with 170k+ acres. GAAP B/M understates LIFO land basis vs market value. True NAV-based B/M likely >1.0 even at current price.",
             "catalyst": "Florida population growth + master plan community development + Berkowitz anchor",
             "variant": "Yartseva fails on size + low GAAP B/M; correct framework is Lynch asset play.",
             "asym": 7.0, "upside": "2-3x", "downside": "-20%", "rr": "10:1", "cross_rank": 27,
             "smart_money": "Berkowitz/Fairholme longstanding anchor"},

            {"ticker": "TPL", "name": "Texas Pacific Land Trust", "mcap": "$30B", "price": "$1300",
             "valuation": "Permian royalty trust at 25x P/CF (vs NRP at 6x)",
             "thesis": "Permian water + royalty optionality. Same NAV-vs-GAAP-book issue as JOE. Pricing 'endless drilling' premium.",
             "catalyst": "Permian production growth + water revenue optionality",
             "variant": "Yartseva captures FCF yield + ROA but understates B/M for land assets.",
             "asym": 6.0, "upside": "1.3-1.5x", "downside": "-25%", "rr": "5:1", "cross_rank": 31,
             "smart_money": "Multi-fund quality holders; HRC anchor"},
        ]
    },

    "8. Foreign / Underfollowed Value": {
        "definition": "Listed outside primary US market or in segment poorly covered by US sell-side. Currency arbitrage + institutional underweight = entry edge.",
        "names": [
            {"ticker": "MTY.TO", "name": "MTY Food Group", "mcap": "C$887M", "price": "C$40.41",
             "valuation": "[See Archetype 1]",
             "thesis": "[See Archetype 1]",
             "catalyst": "Q2 2026 SSS test + NCIB",
             "variant": "[See Archetype 1]",
             "asym": 8.0, "upside": "2-2.5x", "downside": "-15%", "rr": "10:1", "cross_rank": 19,
             "smart_money": "CEO open-market buy"},

            {"ticker": "BZU.IM", "name": "Buzzi SpA", "mcap": "€8.9B", "price": "€49",
             "valuation": "6.8x EV/EBITDA (vs peer 9-11x); €85 PT = +73% upside",
             "thesis": "Italian cement (€8.9B mcap). 52% of EBITDA from US. AI data-center concrete demand thesis. Trading at sector-discount multiple to LafargeHolcim/CRH peers.",
             "catalyst": "AI data-center capex translating to concrete orders; multiple re-rate to peer levels",
             "variant": "Italian listing + cement = institutional underweight. Kerrisdale Oct 2025 long thesis 73% upside.",
             "asym": 8.0, "upside": "1.7-2x", "downside": "-20%", "rr": "8:1", "cross_rank": 20,
             "smart_money": "Kerrisdale Capital published long thesis Oct 2025 (€85 PT)"},

            {"ticker": "MGNI", "name": "Magnite", "mcap": "$1.89B", "price": "$13",
             "valuation": "Yartseva 6/7 (sub-$2B), B/M 0.49, FCF yield positive, ROA positive",
             "thesis": "Independent CTV SSP. Netflix/Disney/Roku exclusive deals scaling. AI ad-buying integration. Last major non-Google scaled SSP.",
             "catalyst": "CTV ad spend doubling 2024-2028 + Netflix exclusive scaling + AI integration",
             "variant": "Market: ad-tech roadkill. Smart: winning independent CTV SSP that survived ad-tech wash-out.",
             "asym": 7.0, "upside": "3-5x", "downside": "-25%", "rr": "12:1", "cross_rank": 23,
             "smart_money": "Nine Ten Capital (Bares) 13.1% of book / 3.48M sh / $41.4M"},
        ]
    },

    "9. Quality Compounder at Trough": {
        "definition": "Defensive quality business at multi-year low multiple due to transitory headwind. Re-rate when consensus catches up to recovery.",
        "names": [
            {"ticker": "DHR", "name": "Danaher", "mcap": "$122B", "price": "$171",
             "valuation": "19.7x forward P/E (trough vs 5yr avg 30x); ROA pos; B/M 0.43",
             "thesis": "Pure-play life sciences (Cytiva bioprocessing, Beckman Dx, Cepheid). Q1 EPS $2.06 beat by 6.2%; bioprocessing inflecting positive after destock; Cepheid respiratory comp resets Q4 2026.",
             "catalyst": "Bioprocessing book-to-bill durably >1.0 + China VBP stabilization H2 2026 + Cepheid Q4 reset",
             "variant": "Consensus: show-me on growth post-destock. Sundheim/D1: same stock 2019-2021 (30%/yr) at trough multiple while H2 2026 comps inflect.",
             "asym": 7.5, "upside": "1.8-2.0x", "downside": "-15%", "rr": "9:1", "cross_rank": 22,
             "smart_money": "D1 Capital +312% Q1 ($107M→$437M, 4th-largest position)"},

            {"ticker": "CTSH", "name": "Cognizant", "mcap": "$25B", "price": "$52.75",
             "valuation": "P/B 1.66 (B/M 0.60), 9.88% FCF yield, ROA 10.44%, -27% 6mo",
             "thesis": "Yartseva 6/7 IT services franchise. Multi-fund consensus Q1 2026. 300k+ delivery workforce moat. AI/IT services capex super-cycle 2026-2028.",
             "catalyst": "AI implementation winner via enterprise partners + Q2 2026 earnings",
             "variant": "Market: AI disrupts IT services. Smart: AI accelerator not disruptor.",
             "asym": 7.0, "upside": "2-3x", "downside": "-15%", "rr": "12:1", "cross_rank": 26,
             "smart_money": "Multi-fund deep-value funds Q1 2026"},

            {"ticker": "REGN", "name": "Regeneron", "mcap": "$65B", "price": "$638",
             "valuation": "B/M 0.48, FCF yield 6.33%, P/B 2.07; Yartseva 6/7 (size only fail)",
             "thesis": "Founder/CEO Schleifer + Yancopoulos R&D-led. $31.4B book equity + EYLEA franchise + Linvoseltamab launching. Pipeline broad. Itepekimab Phase 3 readout potential catalyst.",
             "catalyst": "Itepekimab 2026-2027 data + Linvoseltamab ramp",
             "variant": "Market: fianlimab miss + Eylea biosimilar. Smart: broad pipeline + buyback + founder skin.",
             "asym": 7.5, "upside": "1.5-2x", "downside": "-15%", "rr": "10:1", "cross_rank": 23,
             "smart_money": "Founder-led; multi-fund pharma specialist holds"},
        ]
    },

    "10. Hedged / Special Structure": {
        "definition": "Position structured with options/warrants to cap downside while preserving upside optionality. Sophisticated risk-managed bet.",
        "names": [
            {"ticker": "LQDA", "name": "Liquidia Corp", "mcap": "$5.3B", "price": "$60",
             "valuation": "P/Sales 12x on $130M Q1 run-rate; cash $200M+ floor; M&A premium 50-100%",
             "thesis": "Yutrepia (treprostinil dry-powder inhaler) for PAH and PH-ILD. Q1 2026 product sales $129.9M (+44% QoQ); 4,500 unique patients; mgmt $1B+ 2027 target.",
             "catalyst": "'327 patent ruling pending (binary); Q2 print Aug 2026; UTHR Tresmi competitive launch",
             "variant": "Consensus PT $63.75 assumes flat share. Buckley's PUT-hedged structure suggests binary patent risk; UTHR M&A target.",
             "asym": 7.0, "upside": "2-2.5x", "downside": "-50% (-60% on adverse patent)", "rr": "5:1 (9:1 risk-adj)", "cross_rank": 28,
             "smart_money": "Buckley Capital 32% LONG + PUT structure ($44.7M notional)"},

            {"ticker": "XBI", "name": "SPDR S&P Biotech ETF (Stonepine straddle)", "mcap": "n/a", "price": "$80",
             "valuation": "Biotech basket at multi-year low; straddle structure captures vol",
             "thesis": "Stonepine $51M long + $51M short = $102M notional. Pure long-vol bet on biotech 2026-2027 FDA cycle dispersion.",
             "catalyst": "Biotech FDA cycle 2026-2027 (multiple catalysts driving dispersion)",
             "variant": "Long-vol thesis: biotech sector cycle creates dispersion that benefits straddle structure regardless of direction.",
             "asym": 7.0, "upside": "1.5-2x via straddle", "downside": "-30%", "rr": "5:1", "cross_rank": 31,
             "smart_money": "Stonepine 40.9% of $250M book in XBI straddle"},

            {"ticker": "CTMX-WT", "name": "CytomX Tranche 2 Warrants", "mcap": "n/a", "price": "varies",
             "valuation": "5.77M shares warrants; group ownership equivalent ~3.5%",
             "thesis": "BVF Partners holds CTMX Tranche 2 warrants expiring 7/3/2026 (next 5 weeks). Embedded biotech IPO/crossover leverage near expiration creates urgency catalyst.",
             "catalyst": "Warrant expiration 7/3/2026 = IMMINENT URGENCY",
             "variant": "BVF's KYMR top conviction (14.6%) signals their biotech book has high conviction quality.",
             "asym": 7.5, "upside": "2-4x via warrant", "downside": "-100% (warrant expires worthless)", "rr": "5:1", "cross_rank": 32,
             "smart_money": "BVF Partners 13G/A #5 (warrants exp 7/3/2026)"},

            {"ticker": "AAP CALL", "name": "Advance Auto Parts calls (Cooper Creek)", "mcap": "n/a", "price": "varies",
             "valuation": "$158M call notional; +213% Q1 2026 (Cooper Creek)",
             "thesis": "Cooper Creek's largest single call exposure. Overlaps with H Partners' AAP cash 46% concentration = double-conviction across two fund styles.",
             "catalyst": "Q2 2026 margin recovery proof + H Partners 13D escalation",
             "variant": "Catalyst-driven leverage on H Partners' AAP thesis.",
             "asym": 7.5, "upside": "3-5x via calls", "downside": "-100% on time decay", "rr": "6:1", "cross_rank": 33,
             "smart_money": "Cooper Creek call stack ($518M total notional incl SIG/TTWO/DOCN/GXO)"},
        ]
    },
}

# ================================================================
# BUILD WORKBOOK
# ================================================================

# Sheet 1: Cover / Index
ws = wb.active
ws.title = "Index"
ws.merge_cells('A1:H1')
ws['A1'] = "INVESTMENT ARCHETYPES — Asymmetric Multibagger Universe"
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = CENTER

ws['A3'] = "Generated: 2026-05-27"
ws['A4'] = "Source: 442-sheet fund_activity_last_6mo.xlsx + 11 synthesis docs"
ws['A5'] = "Universe: 33+ asymmetric candidates organized into 10 investment archetypes"

ws['A7'] = "ARCHETYPE INDEX"
ws['A7'].font = HEADER
ws['A7'].fill = HEADER_FILL
ws.merge_cells('A7:H7')

ws['A9'] = "#"
ws['B9'] = "Archetype"
ws['C9'] = "Names"
ws['D9'] = "Top Pick (Cross-Rank)"
for cell in ['A9', 'B9', 'C9', 'D9']:
    ws[cell].font = BOLD
    ws[cell].fill = SECTION_FILL

row = 10
for i, (arch_name, arch_data) in enumerate(ARCHETYPES.items(), 1):
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=arch_name)
    ws.cell(row=row, column=3, value=len(arch_data['names']))
    # Find top within archetype
    top = sorted(arch_data['names'], key=lambda x: -x['asym'])[0]
    ws.cell(row=row, column=4, value=f"{top['ticker']} (#{top['cross_rank']})")
    row += 1

# Index column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 25

# Sheet 2: MASTER CROSS-ARCHETYPE RANKING
ws_master = wb.create_sheet("MASTER Cross-Rank")
ws_master.merge_cells('A1:M1')
ws_master['A1'] = "MASTER ASYMMETRIC RANKING — Cross-Archetype"
ws_master['A1'].font = Font(bold=True, size=14)
ws_master['A1'].alignment = CENTER

# Collect all names
all_names = []
for arch_name, arch_data in ARCHETYPES.items():
    for name in arch_data['names']:
        # Skip cross-references to archetype 1 names appearing in other archetypes
        name_copy = dict(name)
        name_copy['archetype'] = arch_name
        all_names.append(name_copy)

# Sort by cross_rank
all_names.sort(key=lambda x: x['cross_rank'])

# Dedup by ticker (keep first occurrence by cross_rank)
seen_tickers = set()
deduped = []
for n in all_names:
    if n['ticker'] not in seen_tickers:
        seen_tickers.add(n['ticker'])
        deduped.append(n)

# Master headers
headers = ['Cross-Rank', 'Ticker', 'Name', 'Archetype', 'Mcap', 'Price', 'Asym', '36mo Upside', 'Downside', 'R/R', 'Smart Money', 'Catalyst', 'Variant Perception']
for col, h in enumerate(headers, 1):
    cell = ws_master.cell(row=3, column=col, value=h)
    cell.font = HEADER
    cell.fill = HEADER_FILL
    cell.alignment = CENTER

# Data rows
for row_idx, n in enumerate(deduped, 4):
    ws_master.cell(row=row_idx, column=1, value=n['cross_rank'])
    ws_master.cell(row=row_idx, column=2, value=n['ticker'])
    ws_master.cell(row=row_idx, column=3, value=n['name'])
    ws_master.cell(row=row_idx, column=4, value=n['archetype'])
    ws_master.cell(row=row_idx, column=5, value=n['mcap'])
    ws_master.cell(row=row_idx, column=6, value=n['price'])
    ws_master.cell(row=row_idx, column=7, value=n['asym'])
    ws_master.cell(row=row_idx, column=8, value=n['upside'])
    ws_master.cell(row=row_idx, column=9, value=n['downside'])
    ws_master.cell(row=row_idx, column=10, value=n['rr'])
    ws_master.cell(row=row_idx, column=11, value=n['smart_money'])
    ws_master.cell(row=row_idx, column=12, value=n['catalyst'])
    ws_master.cell(row=row_idx, column=13, value=n['variant'])
    for col in range(1, 14):
        ws_master.cell(row=row_idx, column=col).alignment = LEFT_WRAP

# Column widths
widths = [10, 10, 25, 35, 10, 10, 7, 12, 10, 8, 40, 40, 50]
for i, w in enumerate(widths, 1):
    ws_master.column_dimensions[get_column_letter(i)].width = w

# Per-archetype sheets
for arch_name, arch_data in ARCHETYPES.items():
    # Sheet name truncated to 31 chars; replace invalid chars
    sheet_name = arch_name.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '')[:31]
    ws_a = wb.create_sheet(sheet_name)

    ws_a.merge_cells('A1:K1')
    ws_a['A1'] = arch_name
    ws_a['A1'].font = Font(bold=True, size=13)
    ws_a['A1'].alignment = CENTER

    ws_a.merge_cells('A2:K2')
    ws_a['A2'] = f"Definition: {arch_data['definition']}"
    ws_a['A2'].font = Font(italic=True, size=10)
    ws_a['A2'].alignment = LEFT_WRAP

    # Sort by within-archetype asymmetry
    arch_sorted = sorted(arch_data['names'], key=lambda x: -x['asym'])

    headers = ['Within-Rank', 'Cross-Rank', 'Ticker', 'Name', 'Mcap', 'Price', 'Asym', 'Upside', 'Downside', 'R/R', 'Smart Money']
    for col, h in enumerate(headers, 1):
        cell = ws_a.cell(row=4, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for i, n in enumerate(arch_sorted):
        r = 5 + i*4  # leave space for thesis below
        ws_a.cell(row=r, column=1, value=i+1)
        ws_a.cell(row=r, column=2, value=n['cross_rank'])
        ws_a.cell(row=r, column=3, value=n['ticker']).font = Font(bold=True)
        ws_a.cell(row=r, column=4, value=n['name'])
        ws_a.cell(row=r, column=5, value=n['mcap'])
        ws_a.cell(row=r, column=6, value=n['price'])
        ws_a.cell(row=r, column=7, value=n['asym'])
        ws_a.cell(row=r, column=8, value=n['upside'])
        ws_a.cell(row=r, column=9, value=n['downside'])
        ws_a.cell(row=r, column=10, value=n['rr'])
        ws_a.cell(row=r, column=11, value=n['smart_money'])

        # Thesis row
        ws_a.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=11)
        ws_a.cell(row=r+1, column=1, value=f"THESIS: {n['thesis']}").alignment = LEFT_WRAP
        ws_a.cell(row=r+1, column=1).fill = SECTION_FILL

        # Valuation row
        ws_a.merge_cells(start_row=r+2, start_column=1, end_row=r+2, end_column=11)
        ws_a.cell(row=r+2, column=1, value=f"VALUATION: {n['valuation']}").alignment = LEFT_WRAP

        # Catalyst + Variant row
        ws_a.merge_cells(start_row=r+3, start_column=1, end_row=r+3, end_column=11)
        ws_a.cell(row=r+3, column=1, value=f"CATALYST: {n['catalyst']} | VARIANT: {n['variant']}").alignment = LEFT_WRAP

    # Column widths
    widths = [10, 10, 8, 25, 10, 10, 7, 12, 10, 8, 35]
    for i, w in enumerate(widths, 1):
        ws_a.column_dimensions[get_column_letter(i)].width = w
    ws_a.row_dimensions[4].height = 25

# Save
out_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"  - {s}")
print(f"\nTotal asymmetric candidates: {len(deduped)} (deduplicated across archetypes)")
print(f"\nTop 10 by cross-rank:")
for n in deduped[:10]:
    print(f"  #{n['cross_rank']}: {n['ticker']:6s} ({n['asym']}/10, {n['upside']}) - {n['archetype'][:30]}")
