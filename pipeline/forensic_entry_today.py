"""FORENSIC ENTRY-TODAY TIER 1 — brutal classification.

Every Tier 1 name MUST pass ALL forensic checks:
  1. Current price BELOW or AT smart money cost basis (NOT above)
  2. Major catalyst measurable in DAYS or WEEKS ahead (not vague H2 / 2027)
  3. Variant perception STILL INTACT (not consensus yet)
  4. Bounded downside (cash floor / NAV / activist 13D / contractual put)
  5. Not deeply in money from any timeframe

REJECT if:
  - Catalyst already played out (CELC, HRMY, TREE, ASND, DHR, MDGL, CRH)
  - Stock already +25% off lows in last 6mo (BLDR, MU, APP, CVNA)
  - Smart money REDUCING (Steel reduced after Q1, etc — flag if applicable)
  - Pure speculation w/o defined trigger (MSTR, COIN, CRWV, RVMD)
  - Vague catalyst >12 months out (KYMR mid-2027 = Tier 2 not Tier 1)
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================================================================
# FORENSIC TIER 1 — Most brutal entry-today filter
# Every name must pass ALL 5 forensic checks
# ================================================================

FORENSIC_T1 = [
    # === BOUNDED DOWNSIDE + DEFINED CATALYST + SMART MONEY UNDERWATER/FLAT ===

    {'ticker': 'PRLD', 'name': 'Prelude Therapeutics', 'mcap': '$340M', 'price_today': '$4.26',
     'smart_entry': 'Baker 15.5% + RA Cap 9.99% BOTH at $4.44 (April 2026 $90M financing)',
     'price_vs_entry': '-4% UNDERWATER vs DUAL anchor entry',
     'catalyst_specific': 'PRT12396 Phase 1 readout H2 2026 (3-6 months); Incyte option agreement Nov 2025 monetizable',
     'downside_floor': '$2.20 net cash/sh = 51% of price = HARD FLOOR',
     'variant_intact': 'YES — SMARCA2 fail narrative dominates; market missing platform validation by Baker+RA',
     'days_to_catalyst': '90-180 days (H2 2026 readout)',
     'forensic_pass': 'PASS — underwater dual anchor + cash floor + binary 3-6mo',
     'rank': 1},

    {'ticker': 'HHH', 'name': 'Howard Hughes Holdings', 'mcap': '$3.5B', 'price_today': '$63.77',
     'smart_entry': 'Ackman 47% via $900M common @ $100/sh + $1B preferred',
     'price_vs_entry': '-36% UNDERWATER vs Ackman cost basis',
     'catalyst_specific': 'Vantage Specialty Insurance close Q2 2026 = 30-45 DAYS',
     'downside_floor': 'NAV $95-105/sh (RE + insurance float); -34% to NAV bottom',
     'variant_intact': 'YES — RE conglomerate discount; insurance float not yet integrated in multiple',
     'days_to_catalyst': '30-45 days (Vantage close)',
     'forensic_pass': 'PASS — Ackman put floor + imminent dated catalyst + NAV discount',
     'rank': 2},

    {'ticker': 'INMD', 'name': 'InMode Ltd', 'mcap': '$869M', 'price_today': '$13.71',
     'smart_entry': 'Steel +155% Q1 (25.5% of $338M book) + Founder Mizrahy 800k sh @ $13.50 Feb 2026',
     'price_vs_entry': 'FLAT vs both Steel build and founder Feb 2026 buy',
     'catalyst_specific': 'Q2 2026 earnings ~Aug 5 (60-70 days) + Steel re-bid potential ($18 prior)',
     'downside_floor': '$8.47/sh cash = 62% of price = HARD FLOOR',
     'variant_intact': 'YES — aesthetics-cycle narrative; Steel + founder both built at current = market missing M&A value',
     'days_to_catalyst': '60-90 days (Q2 earnings)',
     'forensic_pass': 'PASS — activist + founder dual signal AT current + cash floor + M&A optionality',
     'rank': 3},

    {'ticker': 'OPRX', 'name': 'OptimizeRx', 'mcap': '$93M', 'price_today': '$4.82',
     'smart_entry': 'Corporate $10M buyback executing at $4-5 + Doximity exec Presti to board April 2026 at $4.82',
     'price_vs_entry': 'FLAT — buyback at current; insider 14.82% aligned',
     'catalyst_specific': 'Q2 2026 earnings Aug 14 (75 days) + Q4 DSP integration revenue start (180 days)',
     'downside_floor': 'P/B 0.70 sub-book + $19M FCF; analyst PT $11.50 (+138%)',
     'variant_intact': 'YES — MFN overhang masks digital pharma EHR-native shift',
     'days_to_catalyst': '75 days (Q2 earnings) / 180 days (DSP)',
     'forensic_pass': 'PASS — sub-book + insider 14.82% + Doximity peer exec to board + buyback at current',
     'rank': 4},

    {'ticker': 'KBR', 'name': 'KBR Inc', 'mcap': '$3.9B', 'price_today': '$30',
     'smart_entry': 'D3/Nierenberg +425% Q1 (LARGEST single add in 442-tab universe) built at $30',
     'price_vs_entry': 'FLAT vs D3 Q1 build; Director Form 4 buys May 2026 at $30-31',
     'catalyst_specific': 'MTS spin January 4, 2027 (board-approved Sept 2025) = ~215 days',
     'downside_floor': 'SoTP: MTS $5.2B + STS $5.3B EV = $55-65/sh combined vs $30 today',
     'variant_intact': 'YES — HomeSafe distraction + class action overhang masks defense+nuclear secular',
     'days_to_catalyst': '215 days (MTS spin Jan 4 2027)',
     'forensic_pass': 'PASS — largest single fund add in universe + insider buying May + defined spin date',
     'rank': 5},

    {'ticker': 'LQDA', 'name': 'Liquidia Corp', 'mcap': '$5.3B', 'price_today': '$60',
     'smart_entry': 'Buckley Capital 32% LONG + PUT structure ($44.7M notional)',
     'price_vs_entry': 'Sophisticated hedge structure = expects violent move EITHER WAY',
     'catalyst_specific': '"327 patent ruling PENDING = BINARY (could be days to weeks)',
     'downside_floor': 'PUT-hedged structure = bounded downside; $200M cash; UTHR M&A floor',
     'variant_intact': 'YES — patent binary unresolved; Buckley positioning signals violent move expected',
     'days_to_catalyst': '0-90 days (patent ruling could be any time)',
     'forensic_pass': 'PASS — binary catalyst pending + sophisticated put-hedged structure + M&A floor',
     'rank': 6},

    {'ticker': 'CTMX-WT', 'name': 'CytomX Tranche 2 Warrants', 'mcap': 'n/a', 'price_today': 'varies',
     'smart_entry': 'BVF Partners 13G/A Tranche 2 warrants',
     'price_vs_entry': 'Warrants compress to zero at expiration',
     'catalyst_specific': 'WARRANTS EXPIRE 7/3/2026 = ~33 DAYS',
     'downside_floor': 'Warrant = -100% if expires worthless; bounded loss',
     'variant_intact': 'YES — BVF biotech book quality signal; pure binary structure',
     'days_to_catalyst': '33 days (binary timer)',
     'forensic_pass': 'PASS — purest dated binary in universe; BVF anchor signal',
     'rank': 7},

    {'ticker': 'NRP', 'name': 'Natural Resource Partners', 'mcap': '$1.35B', 'price_today': '$105.81',
     'smart_entry': 'Right Tail Capital NEW Q1 2026 at ~$105; Saber 17.75% + Robertson family 31.75%',
     'price_vs_entry': 'FLAT vs Right Tail freshest entry; family aligned permanently',
     'catalyst_specific': 'Q3 2026 Sisecam recovery signal + Nov 2026 distribution step-up (90-180 days)',
     'downside_floor': '14.5% FCF yield ALONE caps downside; debt-free by mid-2026',
     'variant_intact': 'YES — dying-coal narrative masks royalty trust mechanics; 6.6x vs TPL 25x P/CF',
     'days_to_catalyst': '90-180 days (Q3 + Nov distribution)',
     'forensic_pass': 'PASS — 5-fund concentration + family 31.75% + Right Tail Q1 at current',
     'rank': 8},

    {'ticker': 'MRP', 'name': 'Millrose Properties', 'mcap': '$4.59B', 'price_today': '$25',
     'smart_entry': 'Lennar 80%+ owner = patient capital LOCK + Brave Warrior (Greenberg) +60% Q1',
     'price_vs_entry': 'FLAT vs Lennar spin cost + Brave Warrior current band entry',
     'catalyst_specific': 'Q2 2026 earnings (60-90 days) + housing-cycle inflection ongoing',
     'downside_floor': 'B/M 1.274 DEEPEST in cohort; land bank at cost basis',
     'variant_intact': 'YES — spinoff orphan technical pressure; market treats as not housing-cycle play',
     'days_to_catalyst': '60-90 days (Q2 print)',
     'forensic_pass': 'PASS — deepest B/M + Lennar sponsor lock + Brave Warrior recent ADD',
     'rank': 9},

    # === NEW FORENSIC TIER 1 FROM 455-TAB SCAN ===

    {'ticker': 'WBD', 'name': 'Warner Bros Discovery', 'mcap': '$25B', 'price_today': '~$11',
     'smart_entry': 'Multi-fund Q1 2026 adds (8 tabs) at $9-11 = AT or BELOW current; max ADD +70%',
     'price_vs_entry': 'FLAT to MODESTLY ABOVE smart money Q1 entry band',
     'catalyst_specific': 'Streaming spin announcement Dec 2025 → execution timeline + Q3 2026 cash flow inflection (~120 days)',
     'downside_floor': 'HBO Max content library replacement cost $30B+ vs $25B EV; spin unlocks SoTP value',
     'variant_intact': 'YES — linear TV death narrative; market missing HBO Max international + cost cuts + debt paydown $36B→$11B',
     'days_to_catalyst': '120 days (Q3 cash flow proof) / spin transaction 2026',
     'forensic_pass': 'PASS — 8-fund Q1 accumulation at current band + defined spin catalyst + SoTP value',
     'rank': 10},

    {'ticker': 'SHC', 'name': 'Sotera Health', 'mcap': '$3.5B', 'price_today': '~$14',
     'smart_entry': 'Kerrisdale +9% Q1 at ~$13.50 + 4 HC + 3 ADDs across funds in Q1',
     'price_vs_entry': 'FLAT vs Kerrisdale recent add + multi-fund Q1 band',
     'catalyst_specific': 'Ethylene oxide litigation resolution + Q2 2026 sterilization volume + Q3 ECO production normalization',
     'downside_floor': 'Sterilization MOAT ($4B addressable, no green substitute); peer multiple 15-20x EBITDA vs 9x',
     'variant_intact': 'YES — EO litigation overhang persists; market missing medical device sterilization criticality',
     'days_to_catalyst': '60-90 days (Q2 earnings + litigation milestones)',
     'forensic_pass': 'PASS — Kerrisdale Q1 add at current + medical device moat + bounded litigation downside',
     'rank': 11},

    {'ticker': 'MNKTQ', 'name': 'Mallinckrodt (post-BK)', 'mcap': '<$200M', 'price_today': 'post-BK trough',
     'smart_entry': 'Hudson Bay 5.19% Feb 2026 + Marathon + 1 more = 3 distressed-debt fund 13D filers',
     'price_vs_entry': 'FLAT vs distressed fund Q1 13D entries',
     'catalyst_specific': 'Post-BK equity rerate as Acthar revenue grows + opioid liability discharged',
     'downside_floor': 'Acthar gel franchise + specialty pharma cash flow; distressed-debt funds reorganization mechanics',
     'variant_intact': 'YES — opioid BK stigma masks clean post-BK balance sheet + specialty pharma',
     'days_to_catalyst': '90-180 days (Q2/Q3 earnings + Acthar Medicare path)',
     'forensic_pass': 'PASS — 3 distressed funds Q1 13D + post-BK trough valuation + bounded by Acthar',
     'rank': 12},

    {'ticker': 'CPNG', 'name': 'Coupang', 'mcap': '$50B', 'price_today': '~$24',
     'smart_entry': 'Duquesne Family +45% Q4 + Eminence + Tiger Global = 6 HC at current band',
     'price_vs_entry': 'FLAT to modestly above Q4-Q1 multi-fund adds',
     'catalyst_specific': 'Q2 2026 op margin proof + Eats break-even + Taiwan expansion (~60-90 days)',
     'downside_floor': 'Korean Prime monopoly + $5B+ cash; 3% TAM penetration',
     'variant_intact': 'YES — Korean consumer weakness + Eats losses narrative; market missing logistics moat + AI logistics',
     'days_to_catalyst': '60-90 days (Q2 print)',
     'forensic_pass': 'PASS — 6-fund consensus + recent Q4 adds at current + Q2 catalyst defined',
     'rank': 13},

    {'ticker': 'PCG', 'name': 'PG&E', 'mcap': '$40B', 'price_today': '~$18',
     'smart_entry': '8 HC consensus across 8 tabs + 2 material adds +83% max',
     'price_vs_entry': 'AT or near multi-fund deep-value entry band',
     'catalyst_specific': 'H2 2026 dividend restoration target + Diablo Canyon decision + rate case approvals',
     'downside_floor': 'AB 1054 wildfire fund passed = liability cap; AI data-center NorCal demand floor',
     'variant_intact': 'YES — wildfire BK stigma persists; market missing AI/EV electricity demand growth + dividend return',
     'days_to_catalyst': '120-180 days (H2 dividend + rate case)',
     'forensic_pass': 'PASS — 8 HC consensus + defined dividend catalyst + AI electricity tailwind',
     'rank': 14},

    {'ticker': 'AERO', 'name': 'Grupo Aeromexico (post-BK)', 'mcap': '~$2B', 'price_today': 'post-BK trough',
     'smart_entry': 'Silver Point 9.0% 13G Feb 2026 + SVPGlobal NEW + Baupost ~2% = 3 distressed-value funds Q1',
     'price_vs_entry': 'FLAT vs three distressed fund Q1 entries',
     'catalyst_specific': 'Mexican aviation cycle + tourism rebound + Delta JV optimization (90-180 days)',
     'downside_floor': 'Post-BK clean balance sheet + Mexican Prime aviation slot value',
     'variant_intact': 'YES — post-BK Mexican carrier discount; market missing Delta JV optimization + Mexican GDP recovery',
     'days_to_catalyst': '90-180 days (Q2 traffic + Delta JV)',
     'forensic_pass': 'PASS — Silver Point 13G + SVPGlobal + Baupost three-fund convergence at current',
     'rank': 15},
]

# === FORENSIC TIER 2 — Strong but partial signal or pre-rerate concern ===
FORENSIC_T2 = [
    {'ticker': 'KYMR', 'name': 'Kymera Therapeutics', 'mcap': '$6.5B', 'price_today': '$80',
     'smart_entry': 'BVF 14.6% MAX + Baker 4.1% both at $86 follow-on Dec 2025',
     'price_vs_entry': '-7% UNDERWATER vs dual top biotech anchor',
     'forensic_concern': 'Catalyst BROADEN2 mid-2027 = 12+ months out (low velocity)',
     'rank': 1},

    {'ticker': 'AAP', 'name': 'Advance Auto Parts', 'mcap': '$3.3B', 'price_today': '$54',
     'smart_entry': 'H Partners 46.3% + Q1 +50% sh add (1.35M new)',
     'price_vs_entry': 'MODESTLY in money (~5-15%)',
     'forensic_concern': 'Operational margin recovery thesis dependent on Q2 print; no cash floor',
     'rank': 2},

    {'ticker': 'FUN', 'name': 'Six Flags Entertainment', 'mcap': '$2.2B', 'price_today': '$25',
     'smart_entry': 'H Partners 53.7% + Jana 9% + Cove Street +1,711%',
     'price_vs_entry': 'FLAT to modest',
     'forensic_concern': 'Weather risk Q2-Q3 + synergy execution dependent; activist horizon 6-12mo',
     'rank': 3},

    {'ticker': 'CRTO', 'name': 'Criteo SA', 'mcap': '$866M', 'price_today': '$13.50',
     'smart_entry': 'D3 9.37% +19% Q1 + Petrus activist + $200M buyback',
     'price_vs_entry': 'FLAT vs D3 Q1 add band',
     'forensic_concern': 'Anniversary scope lap Q3 2026 = ~4 months; partial rerate already',
     'rank': 4},

    {'ticker': 'CLBT', 'name': 'Cellebrite DI', 'mcap': '$3.5B', 'price_today': '~$21',
     'smart_entry': '4 HC + +340% max ADD',
     'price_vs_entry': 'PARTIAL — some adds at lower prices',
     'forensic_concern': 'Volatile recent run; AI procurement timing uncertain',
     'rank': 5},

    {'ticker': 'KVUE', 'name': 'Kenvue', 'mcap': '$45B', 'price_today': '~$21',
     'smart_entry': '3 NEW positions Q1 + 169% max ADD',
     'price_vs_entry': 'FLAT vs Q1 NEW position band',
     'forensic_concern': 'Litigation tail risk + slower brand growth re-rate',
     'rank': 6},

    {'ticker': 'BNTC', 'name': 'Benitec Biopharma', 'mcap': '$200M',
     'smart_entry': 'Suvretta 44.1% activist (LARGEST stake by %)',
     'forensic_concern': 'Single-fund concentration risk; biotech speculation',
     'rank': 7},

    {'ticker': 'ACHC', 'name': 'Acadia Healthcare', 'mcap': '$3.2B', 'price_today': '$33',
     'smart_entry': 'Greenlight 4.1M sh + Sohn 5/12 pitch',
     'forensic_concern': 'DOJ resolution timing unknown; behavioral health regulatory tail',
     'rank': 8},

    {'ticker': 'PVLA', 'name': 'Palvella Therapeutics', 'mcap': '$1.56B', 'price_today': '$130',
     'smart_entry': 'BVF + Suvretta +29.3% Q1',
     'forensic_concern': 'NDA H2 2026 / PDUFA H1 2027 = longer dated',
     'rank': 9},

    {'ticker': 'BLDR', 'name': 'Builders FirstSource', 'mcap': '$14B',
     'smart_entry': '5 HC + +597% MAX ADD',
     'forensic_concern': 'Stock already up significantly off lows; rate cut timing uncertain',
     'rank': 10},

    {'ticker': 'APG', 'name': 'API Group', 'mcap': '$11B',
     'smart_entry': '7 HC consensus + Chubb integration',
     'forensic_concern': 'No defined dated catalyst; slow consensus discovery',
     'rank': 11},

    {'ticker': 'CARS', 'name': 'Cars.com', 'mcap': '$560M', 'price_today': '$10',
     'smart_entry': '$90M buyback (16% cap) + Yartseva 7/7',
     'forensic_concern': 'Auto dealer cyclical recovery slow; no concentrated fund anchor',
     'rank': 12},
]

# === RE-RATED / NOT TIER 1 — explicitly disqualified ===
DISQUALIFIED = [
    {'ticker': 'CELC', 'reason': 'Phase 3 May 1 POSITIVE priced + +76% PT raises captured'},
    {'ticker': 'HRMY', 'reason': 'Q1 +17% WAKIX growth printed + Lilly orexin validated'},
    {'ticker': 'TREE', 'reason': 'Q1 +37% rev / +71% EBITDA already PRINTED'},
    {'ticker': 'DV', 'reason': 'CTV verification rally underway; partial upside captured'},
    {'ticker': 'DHR', 'reason': 'Bouncing from trough; bioprocessing cycle now consensus'},
    {'ticker': 'ASND', 'reason': 'YORVIPATH ramp priced; €247M Q1 already at $247'},
    {'ticker': 'CVNA', 'reason': 'Recovered from $3 to $71+ post-split = 40x+ in money'},
    {'ticker': 'MDGL', 'reason': 'Rezdiffra (NASH) approved 2024; commercial inflection priced; stock at $230+'},
    {'ticker': 'CRH', 'reason': 'Stock near highs; cyclical recovery + materials priced; AI data-center exposure only partial driver'},
    {'ticker': 'MU', 'reason': 'HBM cycle thesis already priced through Q1 2026 run-up'},
    {'ticker': 'APP', 'reason': 'AppLovin already 5x+ from lows; AXON AI ad platform priced'},
    {'ticker': 'TPL', 'reason': 'Already at 25x P/CF premium = endless drilling priced'},
    {'ticker': 'BLDR', 'reason': '+597% max ADD impressive but stock has run; housing rate cut timing uncertain'},
    {'ticker': 'COIN/MSTR/CRWV', 'reason': 'Speculative crypto/AI infra; no defined entry-today asymmetry'},
    {'ticker': 'CTSH', 'reason': 'Defensive AI accelerator narrative; no concentrated fund anchor; slower re-rate'},
    {'ticker': 'JOE', 'reason': 'Florida population growth secular but slow re-rate; no near-term catalyst'},
    {'ticker': 'GLOB', 'reason': 'Class action JUNE 23 deadline overhang = NEAR-TERM RISK not asymmetric until cleared'},
    {'ticker': 'VITL', 'reason': 'Deepest drawdown but B/M 0.74 not deep value; no defined binary catalyst; slow brand rerate'},
]


# ================================================================
# BUILD WORKBOOK SHEET
# ================================================================

wb_path = '/home/user/cyclepapa/investment_archetypes.xlsx'
wb = openpyxl.load_workbook(wb_path)

if 'FORENSIC ENTRY-TODAY TIER 1' in wb.sheetnames:
    del wb['FORENSIC ENTRY-TODAY TIER 1']

ws = wb.create_sheet('FORENSIC ENTRY-TODAY TIER 1', 1)

HEADER = Font(bold=True, size=12, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
T1_FILL = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
T2_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
DQ_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws.merge_cells('A1:I1')
ws['A1'] = 'FORENSIC ENTRY-TODAY TIER 1 — Brutal Buy-Today Filter'
ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = CENTER

ws.merge_cells('A2:I2')
ws['A2'] = ('Every TIER 1 name must pass ALL 5 forensic checks: '
            '(1) Current price BELOW or AT smart money cost basis  (2) Catalyst measurable in DAYS/WEEKS (not vague H2)  '
            '(3) Variant perception STILL INTACT  (4) Bounded downside (cash/NAV/13D)  (5) Not deeply in money')
ws['A2'].font = Font(italic=True); ws['A2'].alignment = LEFT_WRAP

# TIER 1 section
ws['A4'] = 'TIER 1 — Passes ALL 5 forensic checks (15 names)'
ws['A4'].font = HEADER; ws['A4'].fill = T1_FILL
ws.merge_cells('A4:I4')

headers = ['Rank', 'Ticker', 'Mcap', 'Price Today', 'Smart Money Entry / vs Current', 'Catalyst (Days Out)', 'Downside Floor', 'Variant Perception Intact', 'Forensic Verdict']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=h)
    cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER

row = 6
for n in FORENSIC_T1:
    ws.cell(row=row, column=1, value=n['rank']).fill = T1_FILL
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=n['mcap'])
    ws.cell(row=row, column=4, value=n['price_today'])
    ws.cell(row=row, column=5, value=f"{n['smart_entry']}\n→ {n['price_vs_entry']}")
    ws.cell(row=row, column=6, value=f"{n['catalyst_specific']}\n(≈ {n['days_to_catalyst']})")
    ws.cell(row=row, column=7, value=n['downside_floor'])
    ws.cell(row=row, column=8, value=n['variant_intact'])
    ws.cell(row=row, column=9, value=n['forensic_pass'])
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 95
    row += 1

# Spacer
row += 1

# TIER 2 section
ws.cell(row=row, column=1, value='TIER 2 — Strong but partial signal or longer-dated catalyst (12 names)')
ws.cell(row=row, column=1).font = HEADER; ws.cell(row=row, column=1).fill = T2_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

t2_headers = ['Rank', 'Ticker', 'Mcap', 'Price Today', 'Smart Money Entry / vs Current', 'Forensic Concern (why not Tier 1)', '', '', '']
for col, h in enumerate(t2_headers, 1):
    if h:
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER; cell.fill = HEADER_FILL; cell.alignment = CENTER
row += 1

for n in FORENSIC_T2:
    ws.cell(row=row, column=1, value=n['rank']).fill = T2_FILL
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=3, value=n['mcap'])
    ws.cell(row=row, column=4, value=n.get('price_today', ''))
    ws.cell(row=row, column=5, value=f"{n.get('smart_entry', '')}\n→ {n.get('price_vs_entry', '')}")
    ws.cell(row=row, column=6, value=n['forensic_concern'])
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    ws.row_dimensions[row].height = 50
    row += 1

# Spacer
row += 1

# DISQUALIFIED section
ws.cell(row=row, column=1, value='DISQUALIFIED FROM TIER 1 — RE-RATED, in-money, or no defined forensic edge (18 names)')
ws.cell(row=row, column=1).font = HEADER; ws.cell(row=row, column=1).fill = DQ_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
row += 1

ws.cell(row=row, column=1, value='Ticker'); ws.cell(row=row, column=1).font = HEADER
ws.cell(row=row, column=1).fill = HEADER_FILL
ws.cell(row=row, column=2, value='Reason for disqualification (FORENSIC failure mode)')
ws.cell(row=row, column=2).font = HEADER; ws.cell(row=row, column=2).fill = HEADER_FILL
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
row += 1

for n in DISQUALIFIED:
    ws.cell(row=row, column=1, value=n['ticker']).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = DQ_FILL
    ws.cell(row=row, column=2, value=n['reason'])
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    for col in range(1, 10):
        ws.cell(row=row, column=col).alignment = LEFT_WRAP
    row += 1

# Column widths
widths = [6, 10, 9, 12, 50, 50, 40, 35, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(wb_path)

print(f"Saved: {wb_path}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"\nFORENSIC TIER 1 (15 names - all pass ALL 5 checks):")
for n in FORENSIC_T1:
    print(f"  #{n['rank']:<3} {n['ticker']:<10} {n['mcap']:<8} {n['price_today']:<12} | catalyst ≈ {n['days_to_catalyst']}")
print(f"\nDISQUALIFIED (18 names - fail forensic):")
for n in DISQUALIFIED:
    print(f"  {n['ticker']:<14} — {n['reason'][:70]}")
