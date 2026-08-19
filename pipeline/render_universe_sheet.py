"""Render the universe analysis workbook — monochrome academic aesthetic.

A companion to style_analysis.xlsx. This file ranks the FULL 5,862-ticker
universe (no style cut) with the unified_signal score, sorted into sheets
by size bucket and signal type.

Design language is shared with render_style_workbook.py via _style_bw.py.
"""
import os, re, sqlite3
import openpyxl
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _style_bw import (
    write_title, write_section_heading, write_table_header, write_table_rows,
    autosize, write_legend_sheet, add_contents_index, set_print_layout,
    NUMFMT_USD, NUMFMT_PCT, NUMFMT_NUM, NUMFMT_INT, NUMFMT_USD2,
    NUMFMT_MCAP, NUMFMT_M_TO_B,
    TNR, SIZE_BODY, BODY_FONT, BODY_ITALIC, SECTION_FONT, TICKER_FONT, MONO_FONT,
    BLACK, ROW_BORDER, LAPIS, CRIMSON, color_directional, color_fixed,
)
from openpyxl.styles import Font, Alignment, Border, Side

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "universe_analysis.xlsx")

ETFs = {"SPY","QQQ","VOO","IWM","IEF","IEFA","EFA","EEM","BIL","IVV","XBI","HYG",
        "GLD","SLV","TLT","XLE","XLF","XLK","XLY","XLP","XLU","XLI","XLV","XLB","XLRE",
        "ARKK","JNK","LQD","TIP","AGG","BND","VEA","VWO","SHY","TIPS"}
MEGA = {"AMZN","MSFT","NVDA","META","GOOGL","GOOG","AAPL","TSLA","BRK-A","BRK-B"}

BIOTECH_PATTERNS = ("pharmaceutic","biological","therapeutic")
def is_biotech(desc):
    if not desc: return False
    d = desc.lower()
    return any(p in d for p in BIOTECH_PATTERNS)

SIG_HDR = ["Ticker","Score","Why","Mcap","ADV $M","Bucket","13F Wt","S1","S3","S4","Act %",
           "pB Max","pB ≥5%","13D","Clu $M",
           "F4 Buy 180d","F4 Buy ≤30d","F4 Sell 180d","F4 Sell ≤30d",
           "EV/EBITDA","P/B",
           "Entry","vs Entry %","3mo %","Off Hi %","Anchor $",
           "Name","Sector","Px","Industry","Business"]

# components string -> readable "why" (top-3 contributing terms). Turns an opaque
# 90+ score into its drivers, e.g. "s1 24 · pb5 18 · s4 14".
_WHY_LABEL = {"sm": "sm", "s3*": "s3", "s4*": "s4", "s1*": "s1", "act": "act",
              "pb_max": "pb", "pb_n5": "pb5", "clust": "clu", "clust$": "clu$",
              "f4buy": "f4buy", "f4rec+": "f4rec", "f4sell": "f4sell",
              "f4recsell": "f4sell30", "mic": "micro", "er": "er", "entry": "entry",
              "cat8k": "8k"}
def _why(components):
    if not components:
        return ""
    terms = []
    for tok in components.split():
        if "=" not in tok:
            continue
        k, v = tok.split("=", 1)
        try:
            val = float(v)
        except ValueError:
            continue
        if abs(val) >= 1 and k in _WHY_LABEL:
            terms.append((_WHY_LABEL[k], val))
    terms.sort(key=lambda t: -abs(t[1]))
    return " · ".join(f"{lbl} {val:.0f}" for lbl, val in terms[:3])

def get_signal_rows(conn, where_extra="", limit=None, params=()):
    sql = """SELECT us.ticker, us.score, us.mcap_m, us.mcap_bucket, us.smart_money_n,
        us.s1_top, us.s3_new, us.s4_add, us.activist_max_pct, us.max_pct_book,
        us.n_funds_5pct_book, us.activist_filings,
        us.insider_cluster_dollars_m,
        us.form4_buy_usd_m, us.form4_buy_30d_m, us.form4_sell_usd_m, us.form4_sell_30d_m,
        us.ev_ebitda, us.pb_ratio,
        us.entry_bucket, us.vs_entry_pct, us.anchor_px,
        us.expected_return_pct, tm.name, tm.sic_description, tm.price,
        COALESCE(yf.industry, tm.industry, tm.sic_description), yf.business_summary,
        tm.adv_3m_usd_m, us.components, ps.mom_3mo, ps.off_high
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf  yf ON yf.ticker = us.ticker
        LEFT JOIN price_stats ps ON ps.ticker = us.ticker
        WHERE us.sec_type='common' """ + where_extra + " ORDER BY us.score DESC"
    if limit: sql += f" LIMIT {limit}"
    return list(conn.execute(sql, params))

def signal_row_to_cells(r):
    """Map a raw signal row to display values.
    Row order: ticker,score,mcap,bucket,13F,s1,s3,s4,act,pbmax,pb5,13d,clu,
               f4_180,f4_30,f4sell_180,f4sell_30,ev_ebitda,pb,entry,vsentry,anchor,
               er,name,sector,price
    """
    eb = r[19] or ""
    if eb == "BELOW_ENTRY":         eb_label = "below"
    elif eb == "NEAR_ENTRY":        eb_label = "near"
    elif eb == "MODERATELY_ABOVE":  eb_label = "mod above"
    elif eb == "WELL_ABOVE":        eb_label = "well above"
    else:                           eb_label = ""
    return [
        r[0], round(r[1] or 0, 1),
        _why(r[29]),                             # Why (top-3 score drivers)
        r[2] or "",                              # Mcap
        round(r[28], 1) if r[28] else "",        # ADV $M (adv_3m_usd_m)
        r[3] or "",                              # Bucket
        round(r[4] or 0, 1), r[5] or 0, r[6] or 0, r[7] or 0,
        round(r[8] or 0, 1),
        round(r[9] or 0, 1),
        r[10] or 0,
        r[11] or 0,
        round(r[12] or 0, 1) if r[12] else "",
        round(r[13] or 0, 1) if r[13] else "",   # F4 buy 180d
        round(r[14] or 0, 1) if r[14] else "",   # F4 buy ≤30d
        round(r[15] or 0, 1) if r[15] else "",   # F4 sell 180d
        round(r[16] or 0, 1) if r[16] else "",   # F4 sell ≤30d
        round(r[17], 1) if r[17] is not None else "",   # EV/EBITDA
        round(r[18], 2) if r[18] is not None else "",   # P/B
        eb_label,
        round(r[20] or 0, 1) if r[20] else "",   # vs entry
        round(r[30], 0) if r[30] is not None else "",   # 3mo % momentum (price_stats)
        round(r[31], 0) if r[31] is not None else "",   # off 3mo high (drawdown)
        round(r[21] or 0, 2) if r[21] else "",   # anchor px
        (r[23] or "")[:38],                      # Name
        (r[24] or "")[:32],                      # Sector
        round(r[25] or 0, 2) if r[25] else "",   # Px
        (r[26] or "")[:26],            # Industry
        _one_liner(r[27], 90),         # Business (one-line summary)
    ]

def format_signal_row(ws, ridx):
    """Apply number formats to a signal row. Columns (1-idx):
    1 Ticker 2 Score 3 Why 4 Mcap 5 ADV$M 6 Bucket 7 13F 8 S1 9 S3 10 S4 11 Act%
    12 pBMax 13 pB≥5% 14 13D 15 Clu$M 16 F4B180 17 F4B30 18 F4S180 19 F4S30
    20 EV/EBITDA 21 P/B 22 Entry 23 vsEntry% 24 Anchor$ 25 Name 26 Sector 27 Px"""
    ws.cell(row=ridx, column=2).number_format = '0.0'          # Score (align decimals)
    ws.cell(row=ridx, column=7).number_format = '0.0'          # 13F Wt (decimal = weighted)
    ws.cell(row=ridx, column=4).number_format = NUMFMT_MCAP    # mcap
    ws.cell(row=ridx, column=5).number_format = NUMFMT_M_TO_B  # ADV $M
    ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT    # act
    ws.cell(row=ridx, column=12).number_format = NUMFMT_PCT    # pB max
    ws.cell(row=ridx, column=15).number_format = NUMFMT_M_TO_B # Clu $M
    ws.cell(row=ridx, column=16).number_format = NUMFMT_M_TO_B # F4 buy 180d
    ws.cell(row=ridx, column=17).number_format = NUMFMT_M_TO_B # F4 buy 30d
    ws.cell(row=ridx, column=18).number_format = NUMFMT_M_TO_B # F4 sell 180d
    ws.cell(row=ridx, column=19).number_format = NUMFMT_M_TO_B # F4 sell 30d
    ws.cell(row=ridx, column=20).number_format = '0.0"x"'      # EV/EBITDA
    ws.cell(row=ridx, column=21).number_format = '0.00"x"'     # P/B
    ws.cell(row=ridx, column=23).number_format = NUMFMT_PCT    # vs entry
    ws.cell(row=ridx, column=24).number_format = NUMFMT_PCT    # 3mo %
    ws.cell(row=ridx, column=25).number_format = NUMFMT_PCT    # off hi %
    ws.cell(row=ridx, column=26).number_format = NUMFMT_USD2   # anchor px
    ws.cell(row=ridx, column=29).number_format = NUMFMT_USD2   # px

def add_signal_heatmap(ws, first_row, last_row):
    """Times-Lattice 'colour is data': faint lapis (good) / crimson (bad) washes on
    the decision columns — a dark ink reduced to a whisper, never a bright pastel.
    Score higher = deeper lapis; valuation cheaper = deeper lapis; vs-entry diverges
    lapis(below, opportunity)↔crimson(well above)."""
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    if last_row < first_row:
        return
    rng = lambda col: f"{col}{first_row}:{col}{last_row}"
    WHITE = "FFFFFF"
    LAPIS_WASH, CRIMSON_WASH = "DDE2EA", "EDDCE0"   # the two inks at ~7% strength
    # Score (col B): higher = deeper lapis (more attractive)
    ws.conditional_formatting.add(rng("B"),
        ColorScaleRule(start_type="min", start_color=WHITE, end_type="max", end_color=LAPIS_WASH))
    ws.conditional_formatting.add(rng("B"), DataBarRule(start_type="min", end_type="max", color="C7CEDA"))
    # EV/EBITDA (T) and P/B (U): cheaper (lower) = deeper lapis
    for col in ("T", "U"):
        ws.conditional_formatting.add(rng(col),
            ColorScaleRule(start_type="min", start_color=LAPIS_WASH, end_type="max", end_color=WHITE))
    # vs Entry % (W): below entry (negative) lapis opportunity ↔ well above crimson
    ws.conditional_formatting.add(rng("W"),
        ColorScaleRule(start_type="min", start_color=LAPIS_WASH, mid_type="num", mid_value=0, mid_color=WHITE,
                       end_type="max", end_color=CRIMSON_WASH))

# ---- sheets -----------------------------------------------------------------
def sheet_readme(wb, conn):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    n_tk = conn.execute("SELECT COUNT(*) FROM unified_signal").fetchone()[0]
    n_fd = conn.execute("SELECT COUNT(DISTINCT fund) FROM fund_meta").fetchone()[0]
    n_hold = conn.execute("SELECT COUNT(*) FROM fund_13f_holdings").fetchone()[0]
    n_13f_funds = conn.execute("SELECT COUNT(DISTINCT fund) FROM fund_13f_holdings").fetchone()[0]
    write_title(ws,
        "Cyclepapa — Universe Analysis",
        f"A data-driven ranking of the smart-money universe ({n_tk:,} tickers, {n_fd} funds, primary EDGAR sources).",
        1)
    ws.column_dimensions["A"].width = 92

    # --- data-as-of dates (be honest about the 13F lag) ---
    def _maxdate(sql):
        r = conn.execute(sql).fetchone()
        return str(r[0])[:10] if r and r[0] else "—"
    f13_filed = _maxdate("SELECT MAX(filed) FROM fund_13f_holdings")
    # 13F holdings are as-of the quarter-end BEFORE the filing (statutory +45d lag)
    import datetime as _dt, calendar as _cal
    f13_asof = "—"
    try:
        fd = _dt.date.fromisoformat(f13_filed)
        qe_month = ((fd.month - 1) // 3) * 3          # 0,3,6,9 -> prior quarter-end month
        yr = fd.year if qe_month else fd.year - 1
        qe_month = qe_month or 12
        f13_asof = f"{yr}-{qe_month:02d}-{_cal.monthrange(yr, qe_month)[1]}"
    except Exception:
        pass
    f4_date = _maxdate("SELECT MAX(trans_date) FROM form4_transactions")
    d13_date = _maxdate("SELECT MAX(filed) FROM holder_13d")
    c8_date = _maxdate("SELECT MAX(filed) FROM catalysts_8k")
    yf_date = _maxdate("SELECT MAX(asof) FROM ticker_yf")

    rows = [
        ("",),
        ("DATA AS-OF (read before trusting any number)",),
        (f"13F holdings   position as-of ~{f13_asof}  (latest filing {f13_filed}; SEC allows +45d, so 'smart money'",),
        ("    reflects quarter-END positions and can be up to ~3-4 months old — a fund may have since exited).",),
        (f"Form 4 insider {f4_date}     ·   13D/G activist {d13_date}     ·   8-K catalysts {c8_date}   (near-current)",),
        (f"Valuations     {yf_date}     (Yahoo; price/mcap current to within days)",),
        ("    → The 13F-derived columns (smart_money, section counts, %book) are the LAGGED layer; the Form 4 /",),
        ("      13D / 8-K / valuation columns are current. Don't read a 13F consensus as a live position.",),
        ("",),
        ("Universe",),
        (f"{n_tk:,} tickers — the union of fund_13f_holdings, fund_positions, and holder_13d.subject_ticker.",),
        (f"{n_fd} funds in fund_meta; ETFs/preferreds/warrants are classified (sec_type) and excluded from pick tables.",),
        ("",),
        ("Score formula",),
        ("score = log(smart_money) × 2          smart_money = CONVICTION-WEIGHTED 13F holders",),
        ("    The '13F' column is NOT a raw holder count. Each fund's vote is weighted",),
        ("    min(1, 75 / n_positions): a focused book (<=75 names) counts fully, while",),
        ("    pod-shops / quants are heavily downweighted — Citadel (6,687 names) ×0.011,",),
        ("    Millennium ×0.019, AQR ×0.020 — so statistical-arb breadth can't fake consensus.",),
        ("      + 3.0 × n_funds_section3        new major positions",),
        ("      + 1.5 × n_funds_section4        existing material adds",),
        ("      + 2.0 × n_funds_section1        top picks",),
        ("      + 0.5 × activist_max_pct        13D/G concentration",),
        ("      + 0.6 × max_pct_book            single-fund concentration",),
        ("      + 1.5 × n_funds_5pct_book       concentration cluster",),
        ("      + cluster_step(n_insiders)      live insider buy cluster",),
        ("      + log(form4_buys + 1) × 2       cumulative open-market buying",),
        ("      − log(form4_sells + 1) × 1.5    insider sells (counter-signal)",),
        ("      + micro_bonus                   +5 if <$300M, +3 if <$2B",),
        ("      + 0.5 × expected_return_pct     base-rate weighted excess",),
        ("",),
        ("Data sources",),
        (f"fund_13f_holdings     {n_hold:,} rows from SEC 13F-HR XML across {n_13f_funds} funds",),
        ("fund_positions        6,748 rows from XLSX research-team classifications",),
        ("holder_13d            current SC 13D/G filings via efts.sec.gov full-text search",),
        ("form4_transactions    P-code open-market buys + S-code sells, ≤180d",),
        ("insider_clusters      live ≤180d clusters",),
        ("catalysts_8k          8-K filings with parsed Item codes",),
        ("ticker_meta           Yahoo chart price + SEC XBRL shares-out (76% mcap coverage)",),
        ("",),
        ("Filters",),
        ("ex-ETF — SPY, QQQ, IWM, sector ETFs removed for noise reduction.",),
        ("ex-Mega — top-10 mega-caps removed where noted.",),
        ("ex-Biotech — SIC matching pharmaceutic / biological / therapeutic excluded where noted.",),
        ("",),
        ("Methodology note",),
        ("This is a pure SQL aggregation. No curated ticker lists, no editorial picks, no memory.",),
        ("The score formula is shared between universe and style workbooks. Re-rank by editing pipeline/unified_score.py.",),
    ]
    for i, r in enumerate(rows, 4):
        c = ws.cell(row=i, column=1, value=r[0])
        ws.row_dimensions[i].height = 18
        if not r[0].strip(): continue
        if r[0][:6] not in ("score ", "      "):
            # section heading test: short capitalized phrases
            if r[0] in ("Universe","Score formula","Data sources","Filters","Methodology note"):
                c.font = SECTION_FONT
            else:
                c.font = BODY_FONT
        else:
            c.font = MONO_FONT

def write_signal_sheet(wb, conn, name, where_extra="", limit=200, subtitle="", exclude_biotech=False):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    write_table_header(ws, 4, SIG_HDR)
    # Fetch a generous superset, then apply the ETF/mega/biotech exclusions in
    # Python and truncate to `limit` — so the exclusions actually reduce the list
    # (the old Non-Biotech sheet passed limit=400 and never filtered biotech).
    fetch = (limit * 4) if limit else None
    rows = get_signal_rows(conn, where_extra=where_extra, limit=fetch)
    rows = [r for r in rows if r[0] not in ETFs and r[0] not in MEGA]
    if exclude_biotech:
        rows = [r for r in rows if not is_biotech(r[24])]   # r[24] = sic_description
    matched = len(rows)
    if limit and matched > limit:
        rows = rows[:limit]
        subtitle = (subtitle + "  " if subtitle else "") + f"[showing top {limit} of {matched} matching names]"
    # Transparency: if the LIMIT truncates, say so in the subtitle rather than
    # silently showing a subset.
    subtitle = (subtitle + "  " if subtitle else "") + "· signal-detail columns (13F/S1-4/pB/13D/insider) are collapsed — click the ＋ above column T to expand."
    write_title(ws, name, subtitle, len(SIG_HDR))
    out = [signal_row_to_cells(r) for r in rows]
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        format_signal_row(ws, ridx)
    add_signal_heatmap(ws, 5, 4 + len(out))
    last = 4 + len(out)
    # colour is data, pervasively: momentum by sign; buys lapis, sells crimson;
    # insider cluster $ lapis; activist stake lapis (its presence is the signal).
    color_directional(ws, 5, last, 24, higher_is_better=True)     # 3mo % momentum
    color_directional(ws, 5, last, 23, higher_is_better=False)    # vs Entry % (below entry = good)
    color_fixed(ws, 5, last, [15, 16, 17], LAPIS)                 # Clu $, F4 buy 180/30
    color_fixed(ws, 5, last, [18, 19], CRIMSON)                   # F4 sell 180/30
    color_fixed(ws, 5, last, [11, 14], LAPIS)                     # Act %, 13D count
    ws.freeze_panes = "B5"
    if out:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(SIG_HDR))}{4 + len(out)}"
    autosize(ws)
    # ticker col narrower; Business (last col) wide for the one-line summary
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["C"].width = 22  # Why
    ws.column_dimensions[get_column_letter(len(SIG_HDR))].width = 80
    ws.column_dimensions[get_column_letter(len(SIG_HDR) - 1)].width = 24  # Industry
    # Collapse the granular signal-detail block (cols H–S: section counts, activist,
    # pB, 13D, cluster, Form-4 detail) into an outline group — the "Why" column
    # already summarizes it, so the default view is identity → score → why → size →
    # liquidity → valuation → entry. Click the [+] to expand the full breakdown.
    ws.sheet_properties.outlinePr.summaryRight = False
    for _c in range(8, 20):   # cols H(8)..S(19)
        cd = ws.column_dimensions[get_column_letter(_c)]
        cd.outline_level = 1
        cd.hidden = True

def _signal_flags(r):
    """Independent signal types firing for a unified_signal row (as a dict).
    r keys: smart_money_n, activist_max_pct, form4_buy_30d_m, insider_n,
    insider_cluster_dollars_m, entry_bucket, ev_ebitda, pb_ratio, cat8k_ma,
    cat8k_ctrl, s3_new, s4_add, form4_sell_30d_m."""
    cheap = ((r["ev_ebitda"] is not None and 2 <= r["ev_ebitda"] <= 12)
             or (r["pb_ratio"] is not None and 0 < r["pb_ratio"] <= 1.2))
    return {
        "Smart$≥3":  (r["smart_money_n"] or 0) >= 3,
        "Activist":  (r["activist_max_pct"] or 0) >= 10,
        "Insider30d":(r["form4_buy_30d_m"] or 0) > 0,
        "Cluster":   (r["insider_n"] or 0) >= 2,
        "New/Add":   (r["s3_new"] or 0) > 0 or (r["s4_add"] or 0) > 0,
        "BelowEntry":r["entry_bucket"] == "BELOW_ENTRY",
        "Cheap":     cheap,
        "Catalyst":  bool(r["cat8k_ma"] or r["cat8k_ctrl"]),
    }

def sheet_convergence(wb, conn):
    """Names where several INDEPENDENT signal types fire at once — the single most
    decision-relevant pattern, and one with no home until now (a name recurs
    across five sheets today with no consolidated view). Check-glyph matrix."""
    ws = wb.create_sheet("Convergence")
    ws.sheet_view.showGridLines = False
    flag_names = ["Smart$≥3","Activist","Insider30d","Cluster","New/Add","BelowEntry","Cheap","Catalyst"]
    write_title(ws, "Convergence — where independent signals stack up",
                "Names firing ≥3 independent signal types. Convergence of unrelated signals is the strongest read. Sorted by signal count, then score.", 6 + len(flag_names))
    hdr = ["Ticker","# Sig","Score","Mcap","Off Hi %"] + flag_names + ["Name"]
    write_table_header(ws, 4, hdr)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT us.*, ps.off_high FROM unified_signal us
        LEFT JOIN price_stats ps ON ps.ticker = us.ticker
        WHERE us.sec_type='common'""").fetchall()
    conn.row_factory = None
    scored = []
    for r in rows:
        if r["ticker"] in ETFs or r["ticker"] in MEGA:
            continue
        flags = _signal_flags(r)
        n = sum(flags.values())
        if n >= 3:
            scored.append((n, r, flags))
    scored.sort(key=lambda t: (-t[0], -(t[1]["score"] or 0)))
    out = []
    for n, r, flags in scored[:120]:
        out.append([r["ticker"], n, round(r["score"] or 0, 1), r["mcap_m"] or "",
                    round(r["off_high"], 0) if r["off_high"] is not None else ""]
                   + ["●" if flags[f] else "" for f in flag_names]
                   + [(r["name"] or "")[:32]])
    write_table_rows(ws, out, 5, ticker_col=1)
    from openpyxl.formatting.rule import DataBarRule
    if out:
        ws.conditional_formatting.add(f"B5:B{4+len(out)}",
            DataBarRule(start_type="num", start_value=3, end_type="num", end_value=8, color="808080"))
        ws.auto_filter.ref = f"A4:{get_column_letter(len(hdr))}{4+len(out)}"
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=5).number_format = NUMFMT_PCT
    ws.freeze_panes = "C5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_action_dashboard(wb, conn):
    """Front-page scannable summary: the top actionable setups across ALL signal
    types on one screen, so a reader gets the 'what should I look at today' answer
    without opening 30 tabs. Ranks by convergence (signal count) then score,
    tradeable names first."""
    ws = wb.create_sheet("Action Dashboard")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Action Dashboard — top setups across every signal",
                "The most actionable names right now: strongest convergence of independent signals, with the driver, valuation, momentum and a one-line read. Start here.", 9)
    hdr = ["Ticker","Setup","Score","# Sig","Mcap","ADV $M","EV/EBITDA","Off Hi %","Read"]
    write_table_header(ws, 4, hdr)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT us.*, tm.adv_3m_usd_m AS adv, ps.off_high, ps.mom_3mo
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN price_stats ps ON ps.ticker = us.ticker
        WHERE us.sec_type='common'""").fetchall()
    conn.row_factory = None
    cand = []
    for r in rows:
        if r["ticker"] in ETFs or r["ticker"] in MEGA:
            continue
        flags = _signal_flags(r)
        n = sum(flags.values())
        if n >= 2 and (r["score"] or 0) >= 15:
            cand.append((n, r, flags))
    cand.sort(key=lambda t: (-t[0], -(t[1]["score"] or 0)))
    out = []
    for n, r, flags in cand[:25]:
        fired = [f for f in flags if flags[f]]
        setup = " + ".join(fired[:3])
        # one-line read synthesizing the strongest angle
        bits = []
        if flags["Activist"]:  bits.append(f"activist {r['activist_max_pct']:.0f}%")
        if flags["Cluster"]:   bits.append(f"{r['insider_n']} insiders buying")
        elif flags["Insider30d"]: bits.append("insiders buying")
        if flags["New/Add"]:
            _nb = (r['s3_new'] or 0) + (r['s4_add'] or 0)
            bits.append(f"{_nb} fund{'s' if _nb != 1 else ''} building")
        if flags["Cheap"] and r["ev_ebitda"]: bits.append(f"{r['ev_ebitda']:.0f}x EV/EBITDA")
        if flags["BelowEntry"] and r["vs_entry_pct"]: bits.append(f"{r['vs_entry_pct']:.0f}% vs entry")
        if r["off_high"] is not None and r["off_high"] <= -15: bits.append(f"{r['off_high']:.0f}% off high")
        read = "; ".join(bits[:4]) or setup
        out.append([r["ticker"], setup, round(r["score"] or 0, 1), n, r["mcap_m"] or "",
                    round(r["adv"], 1) if r["adv"] else "",
                    round(r["ev_ebitda"], 1) if r["ev_ebitda"] is not None else "",
                    round(r["off_high"], 0) if r["off_high"] is not None else "",
                    read])
    write_table_rows(ws, out, 5, ticker_col=1)
    from openpyxl.formatting.rule import ColorScaleRule
    if out:
        ws.conditional_formatting.add(f"C5:C{4+len(out)}",
            ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="595959"))
        ws.auto_filter.ref = f"A4:I{4+len(out)}"
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=6).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=7).number_format = '0.0"x"'
        ws.cell(row=ridx, column=8).number_format = NUMFMT_PCT
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["I"].width = 62

def _has_prior(conn):
    try:
        return conn.execute("SELECT COUNT(*) FROM fund_13f_prior").fetchone()[0] > 0
    except sqlite3.OperationalError:
        return False

def sheet_qoq_change(wb, conn):
    """Quarter-over-quarter 13F change: current holdings diffed against each fund's
    PRIOR filing. A single-quarter snapshot cannot tell accumulation from quiet
    distribution — this can. Net funds building (new+add) minus trimming
    (reduced+exited), and net $ flow, per ticker."""
    ws = wb.create_sheet("QoQ Change")
    ws.sheet_view.showGridLines = False
    write_title(ws, "QoQ Position Change — building vs quietly trimming",
                "Current 13F vs each fund's PRIOR filing, matched on CUSIP + share count (mapping/value-unit safe). Net Funds = (new+added) − (trimmed+exited). All-new names are often IPOs/spins; mixed-churn rows (both adds and trims) are the cleanest accumulation/distribution reads.", 10)
    hdr = ["Ticker","Net Funds","New","Added","Trimmed","Exited","Δ Shares %","Score","Mcap","Name"]
    write_table_header(ws, 4, hdr)
    # Match on CUSIP (stable across quarters), not ticker — the two quarters were
    # mapped by different logic, so a ticker-level diff is dominated by mapping
    # noise (Comcast CMCSA vs CCZ). Share counts are unit-independent, so we
    # ignore the value_k unit stragglers entirely and diff shares.
    rows = list(conn.execute("""
        WITH ok_funds AS (
             -- guard against PARTIAL prior filings: Berkshire's prior accession
             -- covered only $67B of a $263B book, manufacturing fake adds/exits.
             -- A fund's prior book must be within [40%, 250%] of current to diff.
             SELECT c.fund FROM
               (SELECT fund, SUM(value_k) v FROM fund_13f_holdings GROUP BY fund) c
               JOIN (SELECT fund, SUM(value_k) v FROM fund_13f_prior GROUP BY fund) p
               ON p.fund=c.fund
             WHERE c.v > 0 AND p.v BETWEEN c.v*0.4 AND c.v*2.5),
             -- match at the TICKER level (via cusip_map), not raw CUSIP: an
             -- ADR->ordinary CUSIP change between quarters (AZN 046353108 ->
             -- G0593M107) otherwise fabricates "19 funds new + 19 exited".
             cur AS (SELECT h.fund, COALESCE(cm.ticker, h.cusip) tk, SUM(h.shares) sh
                     FROM fund_13f_holdings h LEFT JOIN cusip_map cm ON cm.cusip=h.cusip
                     WHERE h.cusip IS NOT NULL AND h.sh_type IN ('SH','')
                       AND substr(h.cusip,7,1) BETWEEN '0' AND '9'
                       AND substr(h.cusip,8,1) BETWEEN '0' AND '9'
                       AND h.fund IN (SELECT fund FROM ok_funds)
                     GROUP BY h.fund, tk),
             pri AS (SELECT h.fund, COALESCE(cm.ticker, h.cusip) tk, SUM(h.shares) sh
                     FROM fund_13f_prior h LEFT JOIN cusip_map cm ON cm.cusip=h.cusip
                     WHERE h.cusip IS NOT NULL AND h.sh_type IN ('SH','')
                       AND substr(h.cusip,7,1) BETWEEN '0' AND '9'
                       AND substr(h.cusip,8,1) BETWEEN '0' AND '9'
                       AND h.fund IN (SELECT fund FROM ok_funds)
                     GROUP BY h.fund, tk),
             chg AS (
               SELECT cur.fund, cur.tk, cur.sh cur_sh, pri.sh pri_sh
               FROM cur LEFT JOIN pri ON pri.fund=cur.fund AND pri.tk=cur.tk
               UNION ALL
               SELECT pri.fund, pri.tk, NULL, pri.sh
               FROM pri LEFT JOIN cur ON cur.fund=pri.fund AND cur.tk=pri.tk
               WHERE cur.fund IS NULL)
        SELECT tk,
          SUM(CASE WHEN pri_sh IS NULL AND cur_sh>0 THEN 1 ELSE 0 END) n_new,
          SUM(CASE WHEN pri_sh IS NOT NULL AND cur_sh>pri_sh*1.05 THEN 1 ELSE 0 END) n_add,
          SUM(CASE WHEN cur_sh IS NOT NULL AND pri_sh IS NOT NULL AND cur_sh<pri_sh*0.95 THEN 1 ELSE 0 END) n_trim,
          SUM(CASE WHEN cur_sh IS NULL AND pri_sh>0 THEN 1 ELSE 0 END) n_exit,
          SUM(COALESCE(cur_sh,0)) - SUM(COALESCE(pri_sh,0)) d_sh,
          SUM(COALESCE(pri_sh,0)) p_sh
        FROM chg GROUP BY tk""").fetchall())
    scored = []
    for r in rows:
        tk, n_new, n_add, n_trim, n_exit, d_sh, p_sh = r
        net_funds = (n_new + n_add) - (n_trim + n_exit)
        if abs(net_funds) < 2:
            continue
        us = conn.execute("SELECT score, mcap_m, name, sec_type FROM unified_signal WHERE ticker=?", (tk,)).fetchone()
        if not us or us[3] != 'common' or tk in ETFs or tk in MEGA:
            continue
        d_pct = (d_sh / p_sh * 100) if p_sh else (100 if d_sh > 0 else 0)
        scored.append((net_funds, tk, n_new, n_add, n_trim, n_exit, d_pct, us[0], us[1], us[2]))
    # Mixed-churn rows (some adds/trims/exits, i.e. funds with an existing view
    # changing it) are the cleanest accumulation/distribution read; all-new rows
    # (IPO/SPAC allocations) sort below them rather than walling off the top.
    scored.sort(key=lambda x: (-(1 if (x[3] + x[4] + x[5]) > 0 else 0), -x[0]))
    churn = [t for t in scored if (t[3] + t[4] + t[5]) > 0]
    allnew = [t for t in scored if (t[3] + t[4] + t[5]) == 0]
    top = churn[:70] + churn[-30:] + allnew[:20]
    out = []
    for nf, tk, n_new, n_add, n_trim, n_exit, d_pct, score, mcap, name in top:
        pure_new = (n_add + n_trim + n_exit) == 0
        out.append([tk, nf, n_new, n_add, n_trim, n_exit,
                    "new" if pure_new else round(max(-99, min(999, d_pct)), 0),
                    round(score or 0, 1), mcap or "", (name or "")[:38]])
    write_table_rows(ws, out, 5, ticker_col=1)
    # colour is data: Net Funds & Δ Shares — lapis building, crimson trimming
    color_directional(ws, 5, 4 + len(out), [2, 7], higher_is_better=True)
    if out:
        ws.auto_filter.ref = f"A4:J{4+len(out)}"
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=7).number_format = '0"%"'      # Δ Shares % (was $M — wrong unit)
        ws.cell(row=ridx, column=8).number_format = '0.0'       # Score
        ws.cell(row=ridx, column=9).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_dossier(wb, conn, top_n=45):
    """One consolidated block per ticker — score + drivers, holders, insiders,
    activist, catalysts, valuation, momentum — so vetting an idea doesn't mean
    hand-cross-referencing six ticker-keyed sheets. Covers the top-N by score."""
    from _canon import canon
    ws = wb.create_sheet("Ticker Dossier")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Ticker Dossier — every signal per name, one block each",
                f"Top {top_n} by score. Each block: drivers · top holders (13F %book) · insiders · activist · recent 8-Ks · valuation & momentum. The single-idea vetting view.", 8)
    conn.row_factory = sqlite3.Row
    names = conn.execute("""SELECT us.*, tm.adv_3m_usd_m adv, ps.off_high, ps.mom_3mo,
               yf.rev_growth, yf.profit_margin, yf.fwd_pe
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker=us.ticker
        LEFT JOIN price_stats ps ON ps.ticker=us.ticker
        LEFT JOIN ticker_yf yf ON yf.ticker=us.ticker
        WHERE us.sec_type='common' AND us.ticker NOT IN ({})
        ORDER BY us.score DESC LIMIT ?""".format(",".join("?"*len(MEGA))),
        list(MEGA) + [top_n]).fetchall()
    conn.row_factory = None
    row = 4
    from openpyxl.styles import Font as _F
    for r in names:
        tk = r["ticker"]
        # header line: ticker — name | score | mcap | bucket
        h = ws.cell(row=row, column=1, value=f"{tk} — {(r['name'] or '')[:40]}")
        h.font = _F(name="Times New Roman", size=11, bold=True)
        ws.cell(row=row, column=6, value=f"Score {r['score']:.0f}")
        ws.cell(row=row, column=7, value=f"{(r['mcap_m'] or 0)/1000:.1f}B" if r['mcap_m'] else "")
        ws.cell(row=row, column=8, value=r["mcap_bucket"] or "")
        row += 1
        # drivers
        ws.cell(row=row, column=1, value="Drivers"); ws.cell(row=row, column=2, value=_why(r["components"])); row += 1
        # top holders (13F, canonical, %book)
        holders = []
        seen = set()
        for hr in conn.execute("""SELECT fund, pct_book FROM fund_13f_holdings
                WHERE ticker=? AND sh_type IN ('SH','') ORDER BY value_k DESC LIMIT 8""", (tk,)):
            c = canon(hr[0])
            if c in seen: continue
            seen.add(c)
            nm = re.sub(r"\s*\(.*$", "", hr[0]).strip()[:22]
            holders.append(f"{nm}{f' {hr[1]:.0f}%' if hr[1] else ''}")
        ws.cell(row=row, column=1, value="Held by"); ws.cell(row=row, column=2, value=", ".join(holders[:6])[:120]); row += 1
        # insiders + activist
        ins = conn.execute("""SELECT COUNT(DISTINCT owner), SUM(shares*price)/1e6,
                MAX(CASE WHEN role LIKE '%CEO%' OR role LIKE '%CFO%' OR role LIKE '%Chief%' OR role LIKE '%President%' THEN 1 ELSE 0 END)
                FROM form4_transactions WHERE ticker=? AND code='P' AND trans_date>=date('now','-180 days')
                  AND price<200000""", (tk,)).fetchone()
        act = conn.execute("""SELECT holder, pct_class, form FROM holder_13d WHERE subject_ticker=?
                ORDER BY pct_class DESC LIMIT 1""", (tk,)).fetchone()
        parts = []
        if ins and ins[0]: parts.append(f"{ins[0]} insiders bought ${ins[1]:.1f}M{' (C-suite)' if ins[2] else ''}")
        if r["insider_n"] and r["insider_n"] >= 2: parts.append(f"{r['insider_n']}-insider cluster")
        if act: parts.append(f"{re.sub(r'( |).*$','',act[0])[:18]} {act[1]:.0f}% ({'13D' if '13D' in (act[2] or '') else '13G'})" if act[1] else "")
        ws.cell(row=row, column=1, value="Insiders"); ws.cell(row=row, column=2, value=" · ".join(p for p in parts if p)[:120]); row += 1
        # recent catalysts
        cats = conn.execute("""SELECT filed, has_ma, has_control, has_director, has_pipe FROM catalysts_8k
                WHERE ticker=? ORDER BY filed DESC LIMIT 3""", (tk,)).fetchall()
        clabels = []
        for c in cats:
            tags = [t for t, on in [("M&A", c[1]), ("control", c[2]), ("director", c[3]), ("PIPE", c[4])] if on]
            if tags: clabels.append(f"{c[0][:10]} {'/'.join(tags)}")
        ws.cell(row=row, column=1, value="Catalysts"); ws.cell(row=row, column=2, value="; ".join(clabels)[:120] or "—"); row += 1
        # valuation + momentum
        val = []
        if r["ev_ebitda"] is not None: val.append(f"EV/EBITDA {r['ev_ebitda']:.1f}x")
        if r["pb_ratio"] is not None: val.append(f"P/B {r['pb_ratio']:.2f}")
        if r["fwd_pe"] is not None and r["fwd_pe"] > 0: val.append(f"Fwd P/E {r['fwd_pe']:.0f}")
        if r["rev_growth"] is not None: val.append(f"Rev {r['rev_growth']*100:+.0f}%")
        if r["profit_margin"] is not None: val.append(f"Margin {r['profit_margin']*100:.0f}%")
        if r["mom_3mo"] is not None: val.append(f"3mo {r['mom_3mo']:+.0f}%")
        if r["off_high"] is not None: val.append(f"{r['off_high']:+.0f}% off high")
        ws.cell(row=row, column=1, value="Valuation"); ws.cell(row=row, column=2, value=" · ".join(val)[:120]); row += 1
        # thin rule between blocks
        row += 1
    for rr in range(4, row):
        c = ws.cell(row=rr, column=1)
        if c.value in ("Drivers", "Held by", "Insiders", "Catalysts", "Valuation"):
            c.font = _F(name="Times New Roman", size=9, italic=True, color="7F7F7F")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 96
    ws.freeze_panes = "A4"

def sheet_whos_buying(wb, conn):
    """The NAMES behind the s3/s4 counts — which specific funds are initiating new
    positions and materially adding. Counts tell you 'how many'; this tells you
    'who', which is the part that actually matters (a Baupost new position reads
    very differently from an anonymous count of 3)."""
    ws = wb.create_sheet("Who's Buying")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Who's Buying — the funds behind the New / Add counts",
                "Per name: which funds are INITIATING (S3 new major position) and ADDING (S4 material add), from fund_positions. De-duplicated by canonical manager.", 7)
    hdr = ["Ticker", "Company", "Score", "# New", "New Initiators (funds)", "# Add", "Material Adders (funds)"]
    write_table_header(ws, 4, hdr)
    # canonical-manager de-dupe so a fund's name variants don't list twice
    from _canon import canon
    rows = conn.execute("""
        SELECT us.ticker, us.name, us.score, us.s3_new, us.s4_add
        FROM unified_signal us
        WHERE us.sec_type='common' AND (us.s3_new > 0 OR us.s4_add > 0)
        ORDER BY (us.s3_new*2 + us.s4_add) DESC, us.score DESC LIMIT 200""").fetchall()
    def funds_for(tk, sec):
        seen, out = set(), []
        for (f,) in conn.execute("""SELECT DISTINCT fund FROM fund_positions
                WHERE ticker=? AND section=? AND ticker IS NOT NULL""", (tk, sec)):
            c = canon(f)
            if c in seen:
                continue
            seen.add(c)
            # display the raw fund but trimmed of the trailing manager parenthetical
            out.append(re.sub(r"\s*\(.*$", "", f).strip()[:26])
        return out
    out = []
    for tk, name, score, s3, s4 in rows:
        new_f = funds_for(tk, 3)
        # the same manager can't be simultaneously INITIATING and ADDING — when
        # research notes place it in both sections, the initiation wins.
        from _canon import canon as _cn2
        new_keys = {_cn2(f) for f in new_f}
        add_f = [f for f in funds_for(tk, 4) if _cn2(f) not in new_keys]
        out.append([tk, (name or "")[:24], round(score or 0, 1),
                    len(new_f), ", ".join(new_f)[:70],
                    len(add_f), ", ".join(add_f)[:70]])
    write_table_rows(ws, out, 5, ticker_col=1)
    ws.freeze_panes = "B5"
    if out:
        ws.auto_filter.ref = f"A4:G{4 + len(out)}"
    autosize(ws)
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["G"].width = 60

def sheet_best_in_bucket(wb, conn, per_bucket=20):
    """Top names WITHIN each size bucket. The flat Top-100 is ~44% large-cap
    because raw scores favor names held by many funds; ranking within-bucket
    surfaces the best small/micro/mid ideas that the mega-caps otherwise bury.
    """
    ws = wb.create_sheet("Best in Bucket")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Best in Bucket — top ideas within each size class",
                "Score is cross-sectionally biased toward large caps; this ranks the best names within nano/micro/small/mid/large separately.",
                len(SIG_HDR))
    row = 4
    for bucket, label in [("nano", "Nano (<$50M)"), ("micro", "Micro ($50–300M)"),
                          ("small", "Small ($300M–2B)"), ("mid", "Mid ($2–10B)"),
                          ("large", "Large (>$10B)")]:
        write_section_heading(ws, row, label, len(SIG_HDR)); row += 1
        write_table_header(ws, row, SIG_HDR); row += 1
        rows = get_signal_rows(conn, where_extra=f"AND us.mcap_bucket='{bucket}'", limit=per_bucket)
        rows = [r for r in rows if r[0] not in ETFs and r[0] not in MEGA]
        out = [signal_row_to_cells(r) for r in rows]
        write_table_rows(ws, out, row)
        for ridx in range(row, row + len(out)):
            format_signal_row(ws, ridx)
        row += len(out) + 2
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions[get_column_letter(len(SIG_HDR))].width = 80
    ws.column_dimensions[get_column_letter(len(SIG_HDR) - 1)].width = 24

def sheet_activist(wb, conn):
    ws = wb.create_sheet("Activist 10+")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Activist Concentration",
                "SC 13D/G filings disclosing ≥10% stake. Type: 13D = ACTIVIST intent; 13G = PASSIVE holder. Filer shown. Ex-biotech, ex-ETF.", 13)
    hdr = ["Ticker","Type","Top Filer","Act %","Mcap","Bucket","13D #","13F #","pB Max","EV/EBITDA","P/B","Name","Sector"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.mcap_m, us.mcap_bucket,
               us.activist_max_pct, us.activist_filings, us.smart_money_n,
               us.max_pct_book, us.ev_ebitda, us.pb_ratio, tm.name, tm.sic_description,
               (SELECT h.holder FROM holder_13d h WHERE h.subject_ticker=us.ticker
                  ORDER BY h.pct_class DESC LIMIT 1) AS top_filer,
               (SELECT h.form FROM holder_13d h WHERE h.subject_ticker=us.ticker
                  ORDER BY h.pct_class DESC LIMIT 1) AS top_form,
               (SELECT MAX(CASE WHEN h.form LIKE '%13D%' THEN 1 ELSE 0 END)
                  FROM holder_13d h WHERE h.subject_ticker=us.ticker) AS any_13d
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.activist_max_pct >= 10 AND us.sec_type='common'
        ORDER BY any_13d DESC, us.activist_max_pct DESC"""))
    out = []
    for r in rows:
        if r[0] in ETFs or r[0] in MEGA: continue
        if is_biotech(r[10]): continue
        # 13D anywhere = activist; else the top filer's form (13G = passive)
        typ = "13D activist" if r[13] else ("13G passive" if r[12] and "13G" in (r[12] or "") else "—")
        filer = re.sub(r"\s*\(.*$", "", (r[11] or "")).strip()[:26]
        out.append([r[0], typ, filer, round(r[3] or 0, 1),
                    r[1] or "", r[2] or "",
                    r[4] or 0, r[5] or 0, round(r[6] or 0, 1),
                    round(r[7], 1) if r[7] is not None else "",
                    round(r[8], 2) if r[8] is not None else "",
                    (r[9] or "")[:38], (r[10] or "")[:32]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=10).number_format = '0.0"x"'
        ws.cell(row=ridx, column=11).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_broker_radar(wb, conn):
    """Single-desk share-count jumps = candidate swap-hedge footprints for
    stakes nobody has disclosed yet (activist TRS exposure sits on the
    counterparty's 13F, not the activist's)."""
    ws = wb.create_sheet("Broker Swap Radar")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Broker Swap Radar — single-desk hedge footprints",
                "QoQ share-count change per swap-desk broker 13F (UBS, GS, MS, JPM...). One desk absorbing a block ≥0.35% of shares out is the classic total-return-swap hedge print — activist economic exposure with no 13D yet. Idio % = this desk's share of all desks' movement (high = NOT index flow). Context: 13D/G filers ≤12mo + our activist-style funds holding. Caveat: ETF baskets, index adds and plain custody flows also move these books — treat as leads, not proof.", 12)
    hdr = ["Ticker","Broker","Δ Sh (M)","Δ % Out","Δ $M","Desk $M","Idio %","Score","Mcap","13D/G ≤12mo","Activist holders","Name"]
    write_table_header(ws, 4, hdr)
    try:
        rows = list(conn.execute("""SELECT ticker, broker, delta_sh_m, pct_out, delta_m,
                cur_m, idio_pct, score, mcap_m, recent_13d, activist_holders, name
            FROM broker_swap_radar
            ORDER BY pct_out * (idio_pct/100.0) DESC LIMIT 120"""))
    except Exception:
        rows = []
    out = []
    for r in rows:
        out.append([r[0], (r[1] or "")[:20], round(r[2] or 0, 1), r[3] or 0,
                    round(r[4] or 0, 0), round(r[5] or 0, 0), r[6] or 0,
                    round(r[7], 1) if r[7] is not None else "",
                    r[8] or "", (r[9] or "")[:36], (r[10] or "")[:36], (r[11] or "")[:34]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = '0.00"%"'
        ws.cell(row=ridx, column=5).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=6).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=7).number_format = '0"%"'
        ws.cell(row=ridx, column=9).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_broker_radar(wb, conn):
    """Swap-desk share-count jumps — the pre-13D shadow-accumulation radar."""
    ws = wb.create_sheet("Broker Swap Radar")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Broker Swap Radar — possible swap-hedge accumulation",
                "QoQ share-count jumps inside ONE swap-desk broker's 13F. An activist building via cash-settled swaps files nothing — the desk hedging the swap buys the physical and prints HERE first. "
                "Idio % = this desk's move vs all desks (high = idiosyncratic, low = index flow). Leads, not proof: ETF baskets and custody flows also move desks. Shadow = accumulation × idiosyncrasy × activist-context × cheapness.", 13)
    hdr = ["Ticker","Shadow","Broker","Δ Sh (M)","Δ % Out","Δ $M","Desk $M","Idio %","Mcap","Score","Disclosed Swap","Live Action","13D Momentum","144 Sale (contra)","13D (12mo)","Activist holders","Name"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""SELECT ticker, shadow_score, broker, delta_sh_m,
               pct_out, delta_m, cur_m, idio_pct, mcap_m, score,
               live_action, d13_momo, recent_13d, activist_holders, name, disclosed_swap, f144_sale
        FROM broker_swap_radar ORDER BY shadow_score DESC"""))
    out = []
    for r in rows[:200]:
        out.append([r[0], round(r[1] or 0, 1),
                    r[2], r[3], r[4], r[5],
                    r[6], r[7], r[8] or "",
                    round(r[9], 0) if r[9] is not None else "",
                    (r[15] or "")[:26], (r[10] or "")[:22], (r[11] or "")[:28],
                    (r[16] or "")[:18], (r[12] or "")[:24], (r[13] or "")[:24],
                    (r[14] or "")[:32]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = '0.0'
        ws.cell(row=ridx, column=5).number_format = '0.00"%"'
        ws.cell(row=ridx, column=6).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=7).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=8).number_format = '0"%"'
        ws.cell(row=ridx, column=9).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_latent_ownership(wb, conn):
    """Hidden economic control: warrants/converts/blockers/board rights/swaps
    named in 13D text — a 4.9% header can mask a far larger latent stake."""
    try:
        rows = list(conn.execute("""SELECT lo.ticker, lo.holder, lo.filed, lo.form,
                   lo.flags, lo.blocker_pct, lo.swap_counterparties, lo.n_features,
                   u.score, u.mcap_m
            FROM latent_ownership lo LEFT JOIN unified_signal u ON u.ticker = lo.ticker
            WHERE lo.n_features >= 1 ORDER BY
              (lo.swap_counterparties IS NOT NULL) DESC, lo.n_features DESC, lo.filed DESC"""))
    except sqlite3.OperationalError:
        return
    ws = wb.create_sheet("Latent Ownership")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Latent Ownership — hidden economic control in 13D text",
                "Warrants, convertibles, ownership blockers (+ the % ceiling), board-designation rights, ROFRs and DISCLOSED swaps parsed from each holder's latest 13D/G. "
                "A small header % with a rich structure = large latent optionality or overhang. Swap rows (top) name a counterparty desk — cross-check the Broker Swap Radar.", 10)
    hdr = ["Ticker","Holder","Filed","Form","# Feat","Blocker %","Swap Cpty","Hidden Features","Score","Mcap","Name"]
    write_table_header(ws, 4, hdr)
    nm = {r[0]: r[1] for r in conn.execute("SELECT ticker, name FROM unified_signal")}
    out = []
    for r in rows[:160]:
        out.append([r[0] or "", (r[1] or "")[:26], r[2], (r[3] or "")[:9],
                    r[7], r[5] if r[5] else "", (r[6] or "")[:16],
                    (r[4] or "")[:52], round(r[8],0) if r[8] is not None else "",
                    r[9] or "", (nm.get(r[0]) or "")[:30]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=6).number_format = '0.00"%"'
        ws.cell(row=ridx, column=10).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"; autosize(ws); ws.column_dimensions["A"].width = 8

def sheet_nport(wb, conn):
    """Monthly N-PORT holdings of marquee single-manager funds — fresher than
    13F and includes FOREIGN names 13F omits (Sequoia's Rolls-Royce, etc.)."""
    try:
        series = list(conn.execute("""SELECT series, MAX(filed) FROM nport_holdings
            GROUP BY series ORDER BY MAX(filed) DESC"""))
    except sqlite3.OperationalError:
        return
    if not series:
        return
    ws = wb.create_sheet("N-PORT Monthly")
    ws.sheet_view.showGridLines = False
    write_title(ws, "N-PORT Monthly — marquee fund holdings (fresher than 13F)",
                "Top holdings from registered funds' monthly N-PORT-P filings. Monthly cadence beats quarterly 13F, and N-PORT reports FOREIGN listings a 13F never shows. Supplementary (RIC data), not counted as 13F smart money.", 8)
    hdr = ["Series (fund)","Filed","Ticker","Issuer","$M","% Fund"]
    write_table_header(ws, 4, hdr)
    out = []
    for ser, filed in series:
        for r in conn.execute("""SELECT ticker, issuer, ROUND(val_usd/1e6,1), pct
            FROM nport_holdings WHERE series=? AND filed=? ORDER BY val_usd DESC LIMIT 15""", (ser, filed)):
            out.append([ser[:34], filed, r[0] or "", (r[1] or "")[:36], r[2], r[3]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=5).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=6).number_format = '0.0"%"'
    ws.freeze_panes = "C5"; autosize(ws)

def sheet_insider_f4(wb, conn):
    """Insider buying ranked by RECENCY-weighted total. ≤30d buys shown separately."""
    ws = wb.create_sheet("Insider F4 Buys")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Form 4 Insider Buying — recency weighted",
                "Open-market P-code buys. ≤30d weight 1.0; 31–60 d 0.6; 61–120 d 0.3; 121–180 d 0.1. Sorted by recency-weighted dollars.", 13)
    hdr = ["Ticker","Weighted $M","≤30d $M","31-60 $M","61-180 $M","# Buyers","Avg Px","Mcap","Bucket","13F","EV/EBITDA","P/B","Name"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT f.ticker,
            SUM(CASE WHEN julianday('now')-julianday(f.trans_date) <= 30  THEN f.shares*f.price ELSE 0 END)/1e6 AS d_30,
            SUM(CASE WHEN julianday('now')-julianday(f.trans_date) BETWEEN 31 AND 60  THEN f.shares*f.price ELSE 0 END)/1e6 AS d_60,
            SUM(CASE WHEN julianday('now')-julianday(f.trans_date) > 60   THEN f.shares*f.price ELSE 0 END)/1e6 AS d_180,
            COUNT(DISTINCT f.owner), AVG(f.price),
            us.mcap_m, us.mcap_bucket, us.smart_money_n,
            us.ev_ebitda, us.pb_ratio,
            tm.name
        FROM form4_transactions f
        LEFT JOIN ticker_meta tm ON tm.ticker = f.ticker
        LEFT JOIN unified_signal us ON us.ticker = f.ticker
        WHERE f.code='P' AND f.acquired=1 AND f.price IS NOT NULL
          AND f.trans_date >= date('now','-180 days')
                  AND COALESCE(us.sec_type,'common')='common'
          AND NOT EXISTS (SELECT 1 FROM ticker_yf y WHERE y.ticker = f.ticker
              AND ((y.mcap_m > 0 AND f.shares*f.price/1e6 > y.mcap_m)
                OR (y.price > 0 AND (f.price > y.price*5 OR f.price < y.price*0.10))))
          AND NOT (f.shares*f.price/1e6 > 250
                   AND NOT EXISTS (SELECT 1 FROM ticker_yf y2 WHERE y2.ticker = f.ticker))
        GROUP BY f.ticker
        HAVING (d_30*1.0 + d_60*0.6 + d_180*0.3) >= 0.1
        ORDER BY (d_30*1.0 + d_60*0.6 + d_180*0.3) DESC"""))
    out = []
    for r in rows:
        d30, d60, d180 = r[1] or 0, r[2] or 0, r[3] or 0
        weighted = d30 * 1.0 + d60 * 0.6 + d180 * 0.3
        out.append([r[0], round(weighted, 1),
                    round(d30, 1) if d30 else "",
                    round(d60, 1) if d60 else "",
                    round(d180, 1) if d180 else "",
                    r[4], round(r[5] or 0, 2),
                    r[6] or "", r[7] or "unknown",
                    r[8] or 0,
                    round(r[9], 1) if r[9] is not None else "",
                    round(r[10], 2) if r[10] is not None else "",
                    (r[11] or "")[:38]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        for col in (2, 3, 4, 5):
            ws.cell(row=ridx, column=col).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=7).number_format = NUMFMT_USD2
        ws.cell(row=ridx, column=8).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=11).number_format = '0.0"x"'
        ws.cell(row=ridx, column=12).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_insider_recent(wb, conn):
    """Pure recent (≤30d) insider buying — most actionable."""
    ws = wb.create_sheet("Insider Buys ≤30d")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Recent Insider Buying — last 30 days only",
                "Buys reported in the last 30 days. C-Suite = a CEO/CFO/COO/President/Chair bought (personal-cash conviction beats a passive 10%-owner). Net nets out any sells.", 16)
    hdr = ["Ticker","Buy $M","Sell $M","Net $M","# Buyers","C-Suite","Days Ago","Avg Px","Mcap","Bucket","13F","S3","S4","Act %","EV/EBITDA","Name"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT f.ticker, SUM(f.shares*f.price)/1e6 AS dollars_m,
               COUNT(DISTINCT CASE WHEN COALESCE(f.owner,'')='' THEN f.accession ELSE f.owner END), MAX(f.trans_date),
               AVG(f.price),
               us.mcap_m, us.mcap_bucket,
               us.smart_money_n, us.s3_new, us.s4_add, us.activist_max_pct,
               us.ev_ebitda, us.pb_ratio,
               tm.name,
               MAX(CASE WHEN f.role LIKE '%CEO%' OR f.role LIKE '%Chief Exec%'
                        OR f.role LIKE '%CFO%' OR f.role LIKE '%Chief Fin%'
                        OR f.role LIKE '%COO%' OR f.role LIKE '%President%'
                        OR f.role LIKE '%Chair%' THEN 1 ELSE 0 END) AS csuite,
               CAST(julianday('now') - julianday(MAX(f.trans_date)) AS INT) AS days_ago,
               (SELECT COALESCE(SUM(s.shares*s.price),0)/1e6 FROM form4_transactions s
                  WHERE s.ticker=f.ticker AND s.code='S' AND s.trans_date >= date('now','-30 days')
                    AND s.price IS NOT NULL AND s.price < 200000) AS sell_m
        FROM form4_transactions f
        LEFT JOIN ticker_meta tm ON tm.ticker = f.ticker
        LEFT JOIN unified_signal us ON us.ticker = f.ticker
        WHERE f.code='P' AND f.acquired=1 AND f.price IS NOT NULL
          AND f.trans_date >= date('now','-30 days')
                  AND COALESCE(us.sec_type,'common')='common'
          AND NOT EXISTS (SELECT 1 FROM ticker_yf y WHERE y.ticker = f.ticker
              AND ((y.mcap_m > 0 AND f.shares*f.price/1e6 > y.mcap_m)
                OR (y.price > 0 AND (f.price > y.price*5 OR f.price < y.price*0.10))))
          AND NOT (f.shares*f.price/1e6 > 250
                   AND NOT EXISTS (SELECT 1 FROM ticker_yf y2 WHERE y2.ticker = f.ticker))
        GROUP BY f.ticker
        HAVING dollars_m >= 0.05
        ORDER BY csuite DESC, dollars_m DESC"""))
    out = []
    for r in rows:
        buy, sell = r[1] or 0, r[16] or 0
        out.append([r[0], round(buy, 2), round(sell, 2) if sell else "",
                    round(buy - sell, 2), r[2],
                    "CEO/CFO" if r[14] else "", r[15],
                    round(r[4] or 0, 2),
                    r[5] or "", r[6] or "unknown",
                    r[7] or 0, r[8] or 0, r[9] or 0,
                    round(r[10] or 0, 1),
                    round(r[11], 1) if r[11] is not None else "",
                    (r[13] or "")[:38]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=2).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=3).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=4).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=8).number_format = NUMFMT_USD2    # Avg Px
        ws.cell(row=ridx, column=9).number_format = NUMFMT_MCAP    # Mcap
        ws.cell(row=ridx, column=14).number_format = NUMFMT_PCT    # Act %
        ws.cell(row=ridx, column=15).number_format = '0.0"x"'      # EV/EBITDA
    # colour is data: Net $ (buy − sell) — lapis net buying, crimson net selling
    color_directional(ws, 5, 4 + len(out), 4, higher_is_better=True)
    ws.freeze_panes = "B5"
    if out:
        ws.auto_filter.ref = f"A4:P{4 + len(out)}"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_clusters(wb, conn):
    ws = wb.create_sheet("Insider Clusters")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Live Insider Clusters",
                "Insider buy clusters (≤180-day window) — multiple insiders, same ticker. Days-Ago from the window end: a 3-day-old cluster is far stronger than a 29-day-old one.", 12)
    hdr = ["Ticker","Trigger","Days Ago","Window End","# Insiders","Cluster $M","Avg Px","Top Buyer","Mcap","Bucket","EV/EBITDA","P/B"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT ic.ticker, ic.trigger, ic.window_end, ic.n_insiders, ic.total_usd_m,
               ic.avg_price, ic.top_buyer, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio,
               CAST(julianday('now') - julianday(ic.window_end) AS INT) AS days_ago
        FROM insider_clusters ic
        LEFT JOIN ticker_meta tm ON tm.ticker = ic.ticker
        LEFT JOIN unified_signal us ON us.ticker = ic.ticker
        WHERE DATE(ic.window_end) >= DATE('now', '-180 days')
          AND COALESCE(us.sec_type,'common')='common'
          AND ic.n_insiders >= 2
        ORDER BY ic.total_usd_m DESC"""))
    out = [[r[0], r[1], r[11], r[2], r[3], round(r[4] or 0, 2), round(r[5] or 0, 2),
            r[6][:30] if r[6] else "", r[7] or "", r[8] or "unknown",
            round(r[9], 1) if r[9] is not None else "",
            round(r[10], 2) if r[10] is not None else ""] for r in rows]
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=6).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=7).number_format = NUMFMT_USD2
        ws.cell(row=ridx, column=9).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=11).number_format = '0.0"x"'
        ws.cell(row=ridx, column=12).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_unknown(wb, conn):
    ws = wb.create_sheet("Unknown Mcap")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Unknown Market Cap",
                "Tickers where Yahoo + SEC XBRL share-out resolution failed. Foreign listings, SPACs, warrants, defunct.", 13)
    hdr = ["Ticker","Score","Bucket","13F","S1","S3","S4","Act %","pB Max","F4 $M","Name","Sector","Exch"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.score, us.mcap_bucket, us.smart_money_n,
               us.s1_top, us.s3_new, us.s4_add, us.activist_max_pct, us.max_pct_book,
               us.form4_buy_usd_m, tm.name, tm.sic_description, tm.exchange
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.mcap_bucket = 'unknown' AND us.sec_type='common'
          AND COALESCE(tm.name, us.name) IS NOT NULL
        ORDER BY us.score DESC LIMIT 200"""))
    out = []
    for r in rows:
        if r[0] in ETFs: continue
        out.append([r[0], round(r[1] or 0, 1), r[2] or "",
                    r[3] or 0, r[4] or 0, r[5] or 0, r[6] or 0,
                    round(r[7] or 0, 1), round(r[8] or 0, 1),
                    round(r[9] or 0, 1) if r[9] else "",
                    (r[10] or "")[:38], (r[11] or "")[:30], (r[12] or "")[:12]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=8).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_fund_coverage(wb, conn):
    ws = wb.create_sheet("Fund Coverage")
    ws.sheet_view.showGridLines = False
    n_fund = conn.execute("SELECT COUNT(*) FROM fund_meta").fetchone()[0]
    n_data = conn.execute("""SELECT COUNT(DISTINCT fm.fund) FROM fund_meta fm
        WHERE fm.fund IN (SELECT fund FROM fund_13f_state WHERE n_holdings > 0)
           OR fm.fund IN (SELECT fund FROM fund_positions)
           OR fm.fund IN (SELECT holder FROM holder_13d)""").fetchone()[0]
    write_title(ws, "Fund Coverage",
                f"{n_data} of {n_fund} funds ({n_data*100//n_fund}%) have at least one primary-source data row. "
                f"Remaining are documented categorical gaps (below-threshold, CTA, private office, ...).", 3)
    hdr = ["Category","Funds","Pct"]
    write_table_header(ws, 4, hdr)
    # COUNT(DISTINCT fm.fund): the LEFT JOINs to fund_positions/holder_13d fan out
    # (a fund with 50 positions x 5 filings = 250 rows), so a plain COUNT(*)
    # inflated every category ~200x. We also collapse the joins to EXISTS-style
    # per-fund flags first so each fund lands in exactly one category.
    cats = list(conn.execute("""
        WITH f AS (
          SELECT fm.fund,
                 (SELECT n_holdings FROM fund_13f_state s WHERE s.fund=fm.fund) AS nh,
                 EXISTS(SELECT 1 FROM fund_positions p WHERE p.fund=fm.fund) AS has_fp,
                 EXISTS(SELECT 1 FROM holder_13d h WHERE h.holder=fm.fund) AS has_13d,
                 (SELECT status FROM fund_resolution_state r WHERE r.fund=fm.fund) AS status
          FROM fund_meta fm)
        SELECT
          CASE
            WHEN nh > 0 THEN '1. 13F-HR holdings ingested'
            WHEN has_fp AND has_13d THEN '2. fund_positions + 13D/G'
            WHEN has_fp THEN '3. fund_positions only'
            WHEN has_13d THEN '4. 13D/G only (foreign activists)'
            WHEN status LIKE '%non_filer%' THEN '5. Foreign non-filer (explicit)'
            WHEN status = 'below_13f_threshold' THEN '6. Below AUM threshold'
            WHEN status = 'non_equity_strategy' THEN '7. CTA / options (no equity to track)'
            WHEN status = 'historical_13f_only' THEN '8. Historical only (pre-2013 format)'
            WHEN status = 'individual' THEN '9. Individual (not a fund)'
            WHEN status = 'meta_rollup' THEN '10. Meta tab (not a real fund)'
            WHEN status = 'private_office' THEN '11. Private office (no disclosure)'
            ELSE '12. Other / unresolved'
          END as cat, COUNT(DISTINCT fund)
        FROM f GROUP BY cat ORDER BY 1"""))
    total = sum(r[1] for r in cats)
    out = [[r[0], r[1], round(r[1]*100/total, 1)] for r in cats]
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
    autosize(ws)
    ws.column_dimensions["A"].width = 40

def sheet_all_holdings_consolidated(wb, conn):
    """Union view across 13F, fund_positions, and holder_13d — EVERY disclosed
    position across the whole fund universe. Source column indicates origin.
    No per-fund cap and all fund_positions sections (1-4) are included, so no
    fund and no position is silently dropped (fits well within Excel's limit).
    """
    ws = wb.create_sheet("All Positions")
    ws.sheet_view.showGridLines = False
    hdr = ["Fund","Ticker","Source","Value $M","%Book","Section","Activist %","Mcap","Bucket","EV/EBITDA","P/B"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT h.fund, h.ticker, '13F-HR' AS source,
               h.value_k/1000.0 AS value_m,
               h.pct_book, NULL AS section,
               NULL AS act_pct,
               us.mcap_m, us.mcap_bucket, us.ev_ebitda, us.pb_ratio
        FROM fund_13f_holdings h
        LEFT JOIN unified_signal us ON us.ticker = h.ticker
        WHERE h.ticker IS NOT NULL
        UNION ALL
        SELECT fp.fund, fp.ticker, 'XLSX' AS source,
               fp.dollar_m AS value_m,
               fp.pct_value AS pct_book,
               fp.section,
               us.activist_max_pct,
               us.mcap_m, us.mcap_bucket, us.ev_ebitda, us.pb_ratio
        FROM fund_positions fp
        LEFT JOIN unified_signal us ON us.ticker = fp.ticker
        WHERE fp.ticker IS NOT NULL
          AND fp.section IN (1,2,3,4)
        UNION ALL
        SELECT h.holder AS fund, h.subject_ticker AS ticker, 'SC 13D/G' AS source,
               NULL AS value_m,
               h.pct_class AS pct_book,
               NULL AS section,
               us.activist_max_pct,
               us.mcap_m, us.mcap_bucket, us.ev_ebitda, us.pb_ratio
        FROM holder_13d h
        LEFT JOIN unified_signal us ON us.ticker = h.subject_ticker
        WHERE h.subject_ticker IS NOT NULL AND h.pct_class >= 5
        ORDER BY fund, value_m DESC
    """))
    # Show ALL rows (they fit well within Excel's ~1.05M limit). The prior version
    # capped at 6000 ordered by FUND NAME (dropping every fund after ~"P") AND
    # top-30 per fund AND excluded section-2 positions — silently hiding 45 funds
    # and ~65k positions. Now genuinely complete: every fund, every position.
    n_funds = len(set(r[0] for r in rows))
    CAP = 500000  # extreme-safety backstop only
    truncated = len(rows) > CAP
    if truncated:
        rows = rows[:CAP]
    write_title(ws, "All Fund Positions — consolidated view",
                f"Every disclosed position: 13F-HR + fund_positions (all sections) + 13D/G (≥5%). "
                f"{len(rows):,} rows across {n_funds} funds — complete, no per-fund cap.", 11)
    out = []
    for r in rows:
        out.append([(r[0] or "")[:45], r[1] or "", r[2],
                    round(r[3] or 0, 1) if r[3] else "",
                    round(r[4] or 0, 2),
                    r[5] or "",
                    round(r[6] or 0, 1) if r[6] else "",
                    r[7] or "", r[8] or "",
                    round(r[9], 1) if r[9] is not None else "",
                    round(r[10], 2) if r[10] is not None else ""])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=5).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=7).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=8).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=10).number_format = '0.0"x"'
        ws.cell(row=ridx, column=11).number_format = '0.00"x"'
    ws.freeze_panes = "A5"
    autosize(ws)

def sheet_all_funds(wb, conn):
    ws = wb.create_sheet("All Funds")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Per-Fund Inventory",
                "Every fund in fund_meta with data-availability status. Sorted by 13F holdings count.", 7)
    hdr = ["Fund","Status","CIK","13F #","13F $M","13D #","Pos Count"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT fm.fund, fr.status, fr.best_cik,
               COALESCE(st.n_holdings, 0),
               COALESCE(st.total_value_k, 0) / 1e3,
               (SELECT COUNT(*) FROM holder_13d h WHERE h.holder = fm.fund),
               (SELECT COUNT(*) FROM fund_positions fp WHERE fp.fund = fm.fund)
        FROM fund_meta fm
        LEFT JOIN fund_resolution_state fr ON fr.fund = fm.fund
        LEFT JOIN fund_13f_state st ON st.fund = fm.fund
        ORDER BY st.n_holdings DESC NULLS LAST"""))
    out = []
    for r in rows:
        out.append([r[0][:55], r[1] or "", r[2] or "",
                    r[3], round(r[4] or 0), r[5], r[6]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=5).number_format = NUMFMT_M_TO_B   # 13F total $M
    ws.freeze_panes = "B5"
    autosize(ws)

def sheet_asymmetry(wb, conn):
    """MOST ASYMMETRIC setups: cheap valuation + smart money in below entry +
    catalyst + room to multiply. Downside protection × upside potential."""
    ws = wb.create_sheet("Asymmetry")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Asymmetry — best risk/reward setups",
                "asymmetry = margin-of-safety (cheap EV/EBITDA + low P/B + below smart-money entry) × upside (conviction + activist/insider catalyst + small-cap room). Ranked desc.", 16)
    hdr = ["Ticker","Asym","Score","Mcap","Bucket","EV/EBITDA","P/B","P/E","pB Max","Act %","Entry","vs Entry %","Catalyst","Name","Industry","Business"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.asymmetry_score, us.score, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio, us.pe_ttm, us.max_pct_book, us.activist_max_pct,
               us.entry_bucket, us.vs_entry_pct, us.cat8k_ma, us.cat8k_ctrl,
               us.insider_cluster_dollars_m, tm.name
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.asymmetry_score IS NOT NULL AND us.sec_type='common'
        ORDER BY us.asymmetry_score DESC LIMIT 150"""))
    out = []
    for r in rows:
        if r[0] in ETFs or r[0] in MEGA: continue
        eb = r[10] or ""
        eb_label = ("below" if eb=="BELOW_ENTRY" else "near" if eb=="NEAR_ENTRY"
                    else "above" if "ABOVE" in eb else "")
        cat = []
        if r[12]: cat.append("M&A")
        if r[13]: cat.append("CTRL")
        if r[14] and r[14] > 0: cat.append("clstr")
        d = desc_for(conn, r[0])
        out.append([r[0], round(r[1] or 0, 1), round(r[2] or 0, 1),
                    r[3] or "", r[4] or "",
                    round(r[5], 1) if r[5] is not None else "",
                    round(r[6], 2) if r[6] is not None else "",
                    round(r[7], 1) if r[7] is not None else "",
                    round(r[8] or 0, 1), round(r[9] or 0, 1),
                    eb_label, round(r[11] or 0, 1) if r[11] else "",
                    " ".join(cat), (r[15] or "")[:34], d[0], d[1]])
        if len(out) >= 100: break
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5+len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=6).number_format = '0.0"x"'
        ws.cell(row=ridx, column=7).number_format = '0.00"x"'
        ws.cell(row=ridx, column=8).number_format = '0.0"x"'
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=10).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=12).number_format = NUMFMT_PCT
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions[get_column_letter(15)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(16)].width = 80   # Business

def sheet_revealed_pref(wb, conn):
    """Revealed preference — what smart money is ACTIVELY buying (new + adds),
    not just holding. Ranked by revealed_pref = 2*S3 + S4 + 0.5*S1.
    Cross-cut by size bucket so micro/small revealed conviction is visible."""
    ws = wb.create_sheet("Revealed Preference")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Revealed Preference — what they're actively buying",
                "Ranked by revealed_pref = 2×(new major positions) + 1×(material adds) + 0.5×(top-conviction holds). Reveals active accumulation, not static holdings.", 15)
    hdr = ["Ticker","Rev Pref","S3 New","S4 Add","S1 Top","13F","Mcap","Bucket","EV/EBITDA","P/B","Act %","Entry","Name","Industry","Business"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.revealed_pref, us.s3_new, us.s4_add, us.s1_top,
               us.smart_money_n, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio, us.activist_max_pct, us.entry_bucket, tm.name
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.revealed_pref > 0 AND us.sec_type='common'
        ORDER BY us.revealed_pref DESC, us.smart_money_n DESC LIMIT 120"""))
    out = []
    for r in rows:
        if r[0] in ETFs or r[0] in MEGA: continue
        eb = r[11] or ""
        eb_label = ("below" if eb == "BELOW_ENTRY" else "near" if eb == "NEAR_ENTRY"
                    else "above" if "ABOVE" in eb else "")
        d = desc_for(conn, r[0])
        out.append([r[0], round(r[1] or 0, 1), r[2] or 0, r[3] or 0, r[4] or 0,
                    r[5] or 0, r[6] or "", r[7] or "",
                    round(r[8], 1) if r[8] is not None else "",
                    round(r[9], 2) if r[9] is not None else "",
                    round(r[10] or 0, 1), eb_label, (r[12] or "")[:38], d[0], d[1]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=7).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=9).number_format = '0.0"x"'
        ws.cell(row=ridx, column=10).number_format = '0.00"x"'
        ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT
    ws.column_dimensions[get_column_letter(14)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(15)].width = 80   # Business
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_valuation(wb, conn):
    """Cheapest names by valuation among smart-money holdings."""
    ws = wb.create_sheet("Valuation")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Valuation — cheap on trailing multiples, checked for quality",
                "Names held by ≥3 funds, cheapest EV/EBITDA first. Growth/margin columns separate a cheap compounder from a value trap (neg growth/margin flagged). EV/Rev covers names with no EBITDA.", 19)
    hdr = ["Ticker","EV/EBITDA","P/B","P/E","Fwd P/E","EV/Rev","Rev Gr %","Margin %","Score","Mcap","Bucket","13F","Act %","Entry","vs Entry %","Name","Sector","Industry","Business"]
    write_table_header(ws, 4, hdr)
    # Floor at 2x — below that is almost always a data artifact (warrant,
    # near-zero EBITDA, ADR currency mismatch). Exclude warrant/preferred tickers.
    rows = list(conn.execute("""
        SELECT us.ticker, us.ev_ebitda, us.pb_ratio, us.pe_ttm, us.score, us.mcap_m, us.mcap_bucket,
               us.smart_money_n, us.activist_max_pct, us.entry_bucket, us.vs_entry_pct,
               tm.name, tm.sic_description,
               yf.fwd_pe, yf.ev_revenue, yf.rev_growth, yf.profit_margin
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf yf ON yf.ticker = us.ticker
        WHERE us.ev_ebitda IS NOT NULL AND us.ev_ebitda >= 2 AND us.ev_ebitda < 40 AND us.sec_type='common'
          AND us.smart_money_n >= 3
          AND us.ticker NOT LIKE '%-P%'   -- preferreds
          AND us.ticker NOT LIKE '%-W%'   -- warrants/when-issued
          AND us.ticker NOT LIKE '%W'     -- warrant suffix
          AND us.ticker NOT LIKE '%.W%'
        ORDER BY us.ev_ebitda ASC LIMIT 120"""))
    out = []
    for r in rows:
        if r[0] in ETFs or r[0] in MEGA: continue
        eb = r[9] or ""
        eb_label = ("below" if eb == "BELOW_ENTRY" else "near" if eb == "NEAR_ENTRY"
                    else "above" if "ABOVE" in eb else "")
        d = desc_for(conn, r[0])
        out.append([r[0], round(r[1], 1), round(r[2], 2) if r[2] is not None else "",
                    round(r[3], 1) if r[3] is not None else "",
                    round(r[13], 1) if (r[13] is not None and r[13] > 0) else "",   # Fwd P/E (neg = meaningless)
                    round(r[14], 1) if r[14] is not None else "",          # EV/Rev
                    round(r[15]*100, 0) if r[15] is not None else "",      # Rev Gr %
                    round(r[16]*100, 0) if r[16] is not None else "",      # Margin %
                    round(r[4] or 0, 1), r[5] or "", r[6] or "", r[7] or 0,
                    round(r[8] or 0, 1), eb_label,
                    round(r[10] or 0, 1) if r[10] else "",
                    (r[11] or "")[:38], (r[12] or "")[:30], d[0], d[1]])
    write_table_rows(ws, out, 5)
    # colour is data: Rev Gr % and Margin % — lapis growing/profitable, crimson
    # shrinking/loss-making (the value-trap tell).
    color_directional(ws, 5, 4 + len(out), [7, 8], higher_is_better=True)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=2).number_format = '0.0"x"'    # EV/EBITDA
        ws.cell(row=ridx, column=3).number_format = '0.00"x"'   # P/B
        ws.cell(row=ridx, column=4).number_format = '0.0"x"'    # P/E
        ws.cell(row=ridx, column=5).number_format = '0.0"x"'    # Fwd P/E
        ws.cell(row=ridx, column=6).number_format = '0.0"x"'    # EV/Rev
        ws.cell(row=ridx, column=7).number_format = '0"%"'      # Rev Gr %
        ws.cell(row=ridx, column=8).number_format = '0"%"'      # Margin %
        ws.cell(row=ridx, column=10).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=13).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=15).number_format = NUMFMT_PCT
    ws.freeze_panes = "B5"
    if out:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(hdr))}{4 + len(out)}"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions[get_column_letter(18)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(19)].width = 80   # Business

def sheet_catalysts(wb, conn):
    """8-K material-event tickers: M&A, control change, director shuffle, PIPE, bankruptcy."""
    ws = wb.create_sheet("Catalysts 8-K")
    ws.sheet_view.showGridLines = False
    write_title(ws, "8-K Material-Event Catalysts (≤180d)",
                "M&A (1.01 / 2.01), Control change (5.01), Director change (5.02), PIPE/dilution (3.02), Bankruptcy (1.03). Sorted by unified score, then M&A/control weight.", 15)
    hdr = ["Ticker","Score","Mcap","M&A","Ctrl","Director","PIPE","Bnk","Total Events","13F","Activist %","EV/EBITDA","P/B","Name","Sector"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.score, us.mcap_m,
               us.cat8k_ma, us.cat8k_ctrl, us.cat8k_dir, us.cat8k_pipe, us.cat8k_bnk, us.cat8k_n,
               us.smart_money_n, us.activist_max_pct, us.ev_ebitda, us.pb_ratio, tm.name, tm.sic_description
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.cat8k_n > 0 AND us.sec_type='common'
        ORDER BY us.score DESC, (us.cat8k_ma*5 + us.cat8k_ctrl*4) DESC LIMIT 200"""))
    out = []
    for r in rows:
        if r[0] in ETFs or r[0] in MEGA: continue
        out.append([r[0], round(r[1] or 0, 1), r[2] or "",
                    "✓" if r[3] else "", "✓" if r[4] else "",
                    "✓" if r[5] else "", "✓" if r[6] else "",
                    "✓" if r[7] else "", r[8] or 0,
                    r[9] or 0, round(r[10] or 0, 1),
                    round(r[11], 1) if r[11] is not None else "",
                    round(r[12], 2) if r[12] is not None else "",
                    (r[13] or "")[:38], (r[14] or "")[:32]])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=12).number_format = '0.0"x"'
        ws.cell(row=ridx, column=13).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

# Approximate FX -> USD (mid-2026). Foreign mcap in ticker_yf is in the LISTING
# currency; printing it with a "$" made 4676.T (¥586B ≈ $3.9B) read as "$586B".
# These are approximate and drift; the sheet labels them so and shows the
# currency. Minor units (GBp pence, ZAc cents) convert via their major /100.
_FX_USD = {
    "USD": 1.0, "CAD": 0.73, "EUR": 1.08, "GBP": 1.28, "GBp": 0.0128, "JPY": 0.0064,
    "HKD": 0.128, "AUD": 0.66, "CHF": 1.12, "SGD": 0.74, "INR": 0.012, "KRW": 0.00073,
    "TWD": 0.031, "ZAR": 0.055, "ZAc": 0.00055, "NOK": 0.093, "DKK": 0.145, "SEK": 0.095,
    "PLN": 0.25, "IDR": 0.0000615, "TRY": 0.030, "HUF": 0.0028, "MYR": 0.21, "CNY": 0.138,
    "BRL": 0.18, "MXN": 0.055, "THB": 0.028, "PHP": 0.017, "NZD": 0.60, "ILS": 0.27,
}

def _mcap_usd(mcap_m, currency):
    if not mcap_m:
        return None
    return mcap_m * _FX_USD.get(currency or "USD", None) if (currency or "USD") in _FX_USD else None

def sheet_global_picks(wb, conn):
    """Foreign-exchange tickers — scored on a GLOBAL-FAIR formula.

    The standard unified_score includes US-only signals (Form 4 buys,
    insider clusters) which SEC doesn't provide for foreign listings.
    This sheet uses global_score (smart_money + sections + activist +
    pct_book + micro_bonus + entry_bonus only) so foreign tickers rank
    on the same footing as US tickers stripped of the same signals.
    """
    ws = wb.create_sheet("Global Picks")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Global Picks — non-US listings, fair-score",
                "Foreign-exchange tickers (.L London, .T Tokyo, .TO Toronto, .HK Hong Kong, .AX Sydney, .MI Milan, .DE Frankfurt, .PA Paris, .AS Amsterdam, .MC Madrid). Ranked by global_score which excludes US-only signals.", 17)
    ws.cell(row=2, column=1).value = (ws.cell(row=2, column=1).value or "") + \
        "  Mcap converted to USD at approximate mid-2026 FX (see Currency col)."
    hdr = ["Ticker","Global Score","Exchange","Mcap $ (USD)","Ccy","13F","S1","S3","S4","pB Max","Act %","Entry","vs Entry %","EV/EBITDA","P/B","Name","Industry","Business"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.global_score, tm.exchange, us.mcap_m,
               us.smart_money_n, us.s1_top, us.s3_new, us.s4_add,
               us.max_pct_book, us.activist_max_pct,
               us.entry_bucket, us.vs_entry_pct, us.ev_ebitda, us.pb_ratio, tm.name,
               yf.currency
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf yf ON yf.ticker = us.ticker
        WHERE us.is_us = 0 AND us.sec_type='common'
          AND (us.s1_top + us.s3_new + us.s4_add + us.smart_money_n) >= 1
          AND COALESCE(tm.name, yf.long_name) IS NOT NULL
        ORDER BY us.global_score DESC LIMIT 150"""))
    out = []
    for r in rows:
        eb = r[10] or ""
        eb_label = ("below" if eb == "BELOW_ENTRY" else
                    "near"  if eb == "NEAR_ENTRY" else
                    "above" if "ABOVE" in eb else "")
        ccy = r[15] or "USD"
        mcap_usd = r[3]   # unified_signal.mcap_m is already FX-converted to USD
        out.append([r[0], round(r[1] or 0, 1),
                    (r[2] or "")[:14],
                    round(mcap_usd) if mcap_usd is not None else "", ccy, r[4] or 0,
                    r[5] or 0, r[6] or 0, r[7] or 0,
                    round(r[8] or 0, 1),
                    round(r[9] or 0, 1),
                    eb_label,
                    round(r[11] or 0, 1) if r[11] else "",
                    round(r[12], 1) if r[12] is not None else "",
                    round(r[13], 2) if r[13] is not None else "",
                    (r[14] or "")[:38], *desc_for(conn, r[0])])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_MCAP    # Mcap $ (USD)
        ws.cell(row=ridx, column=10).number_format = NUMFMT_PCT    # pB Max
        ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT    # Act %
        ws.cell(row=ridx, column=13).number_format = NUMFMT_PCT    # vs Entry %
        ws.cell(row=ridx, column=14).number_format = '0.0"x"'      # EV/EBITDA
        ws.cell(row=ridx, column=15).number_format = '0.00"x"'     # P/B
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions[get_column_letter(17)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(18)].width = 80   # Business


_ANCHOR_LABEL = {"candidates": "analyst anchor", "raw_text": "filing text",
                 "form4_p_buy": "insider avg cost", "p80_close": "80th-pctl price"}
def sheet_in_the_money(wb, conn):
    """Below-entry / in-the-money picks — buy below where smart money entered."""
    ws = wb.create_sheet("In The Money")
    ws.sheet_view.showGridLines = False
    write_title(ws, "In The Money — buy below smart-money entry",
                "Current price below the smart-money cost anchor (cost_basis / raw_text / Form-4 P-buy avg / 80th-pctl). Asymmetric setup.", 19)
    hdr = ["Ticker","Score","Mcap","Bucket","Now $","Anchor $","vs Entry %",
           "13F","S1","S3","S4","Act %","pB Max","Anchor Src","EV/EBITDA","P/B","Name","Industry","Business"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.score, us.mcap_m, us.mcap_bucket,
               COALESCE(yf.price, us.price) AS now_px,
               us.anchor_px, us.vs_entry_pct,
               us.smart_money_n, us.s1_top, us.s3_new, us.s4_add,
               us.activist_max_pct, us.max_pct_book,
               us.anchor_source, us.ev_ebitda, us.pb_ratio, tm.name
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf yf ON yf.ticker = us.ticker
        WHERE us.entry_bucket = 'BELOW_ENTRY' AND us.sec_type='common'
        ORDER BY us.score DESC LIMIT 100"""))
    out = []
    for r in rows:
        if r[0] in ETFs: continue
        # ONE price vintage: recompute vs-entry from the freshest price shown in
        # the NOW column, and drop rows that are no longer below the anchor —
        # the stored vs_entry_pct could disagree with its own row (FEIM showed
        # -18% while NOW > ANCHOR).
        now_px, anchor = r[4], r[5]
        vs = ((now_px / anchor) - 1) * 100 if (now_px and anchor) else None
        if vs is None or vs > 2:      # not meaningfully below entry any more
            continue
        out.append([r[0], round(r[1] or 0, 1), r[2] or "", r[3] or "",
                    round(now_px or 0, 2) if now_px else "",
                    round(anchor or 0, 2) if anchor else "",
                    round(vs, 1),
                    r[7] or 0, r[8] or 0, r[9] or 0, r[10] or 0,
                    round(r[11] or 0, 1), round(r[12] or 0, 1),
                    _ANCHOR_LABEL.get(r[13], (r[13] or ""))[:18],
                    round(r[14], 1) if r[14] is not None else "",
                    round(r[15], 2) if r[15] is not None else "",
                    (r[16] or "")[:38], *desc_for(conn, r[0])])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=5).number_format = NUMFMT_USD2
        ws.cell(row=ridx, column=6).number_format = NUMFMT_USD2
        ws.cell(row=ridx, column=7).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=12).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=13).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=15).number_format = '0.0"x"'
        ws.cell(row=ridx, column=16).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions[get_column_letter(18)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(19)].width = 80   # Business

def sheet_bill_miller(wb, conn):
    ws = wb.create_sheet("Bill Miller")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Bill Miller — both Funds",
                "Miller Value Partners (Bill IV, Sarasota) + Patient Capital Management (Bill III). Side-by-side with overlap.", 12)
    funds = [
        ("Bill IV — Miller Value Partners", "Miller Value Partners%"),
        ("Bill III — Patient Capital",       "Patient Capital%"),
    ]
    write_section_heading(ws, 4, "Top 20 holdings — per fund", 12)
    hdr = ["Ticker","Issuer","Value $M","%Book","Mcap","Bucket","13F","Act %","Cluster?","EV/EBITDA","P/B","Name"]
    write_table_header(ws, 5, hdr)
    row = 6
    for label, like in funds:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        c = ws.cell(row=row, column=1, value=label)
        c.font = BODY_ITALIC
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1
        rows = list(conn.execute("""
            SELECT h.ticker, h.issuer, h.value_k, h.pct_book,
                   us.mcap_m, us.mcap_bucket, us.smart_money_n, us.activist_max_pct,
                   us.insider_cluster_dollars_m, us.ev_ebitda, us.pb_ratio, tm.name
            FROM fund_13f_holdings h
            LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
            LEFT JOIN unified_signal us ON us.ticker = h.ticker
            WHERE h.fund LIKE ?
            ORDER BY h.value_k DESC LIMIT 20""", (like,)))
        out = []
        for r in rows:
            cluster_mark = "yes" if (r[8] and r[8] > 0) else ""
            out.append([r[0] or "-", (r[1] or "")[:30],
                        round((r[2] or 0)/1000, 1) if r[2] else "",
                        round(r[3] or 0, 2),
                        r[4] or "", r[5] or "",
                        r[6] or 0, round(r[7] or 0, 1),
                        cluster_mark,
                        round(r[9], 1) if r[9] is not None else "",
                        round(r[10], 2) if r[10] is not None else "",
                        (r[11] or "")[:30]])
        write_table_rows(ws, out, row)
        for ridx in range(row, row + len(out)):
            ws.cell(row=ridx, column=3).number_format = NUMFMT_NUM
            ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
            ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
            ws.cell(row=ridx, column=8).number_format = NUMFMT_PCT
            ws.cell(row=ridx, column=10).number_format = '0.0"x"'
            ws.cell(row=ridx, column=11).number_format = '0.00"x"'
        row += len(out) + 1

    row += 1
    write_section_heading(ws, row, "Shared overlap — held by both funds", 12)
    row += 1
    hdr2 = ["Ticker","Issuer","Bill IV %","Bill III %","Combined %","Mcap","Bucket","13F","Act %","EV/EBITDA","P/B","Name"]
    write_table_header(ws, row, hdr2)
    row += 1
    overlap = list(conn.execute("""
        SELECT h4.ticker, h4.issuer, h4.pct_book pct4, h3.pct_book pct3,
               (h4.pct_book + h3.pct_book) AS combined,
               us.mcap_m, us.mcap_bucket, us.smart_money_n, us.activist_max_pct,
               us.ev_ebitda, us.pb_ratio, tm.name
        FROM fund_13f_holdings h4
        JOIN fund_13f_holdings h3 ON h3.ticker = h4.ticker
        LEFT JOIN ticker_meta tm ON tm.ticker = h4.ticker
        LEFT JOIN unified_signal us ON us.ticker = h4.ticker
        WHERE h4.fund LIKE 'Miller Value%' AND h3.fund LIKE 'Patient Capital%'
          AND h4.ticker IS NOT NULL
        ORDER BY combined DESC"""))
    out = []
    for r in overlap:
        out.append([r[0], (r[1] or "")[:30],
                    round(r[2] or 0, 2),
                    round(r[3] or 0, 2),
                    round(r[4] or 0, 2),
                    r[5] or "", r[6] or "",
                    r[7] or 0, round(r[8] or 0, 1),
                    round(r[9], 1) if r[9] is not None else "",
                    round(r[10], 2) if r[10] is not None else "",
                    (r[11] or "")[:30]])
    write_table_rows(ws, out, row)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=6).number_format = NUMFMT_MCAP    # mcap
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=10).number_format = '0.0"x"'
        ws.cell(row=ridx, column=11).number_format = '0.00"x"'
    ws.freeze_panes = "B6"
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_best_ideas(wb, conn):
    """Composite shortlist — names that fire on MULTIPLE independent signals at
    once: cheap valuation, below smart-money entry, recent insider buying,
    activist / concentration, and a live catalyst. Each row carries a plain-text
    rationale. Ex-biotech, ex-mega; small enough to multiply (< $10B)."""
    ws = wb.create_sheet("Best Ideas")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Best Ideas — multi-signal shortlist",
                "Names firing on several independent signals at once (cheap + below entry + insider buying + activist/concentration + catalyst). Ranked by a blended idea score; rationale in the Why column. Ex-biotech, ex-mega, < $10B.", 18)
    hdr = ["Ticker", "Idea Score", "# Signals", "Base Score", "Asym", "Mcap", "Bucket",
           "EV/EBITDA", "P/B", "Entry", "vs Entry %", "F4 ≤30d $M", "Act %",
           "Catalyst", "Name", "Why", "Industry", "Business"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker, us.score, us.asymmetry_score, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio, us.entry_bucket, us.vs_entry_pct,
               us.form4_buy_30d_m, us.insider_cluster_dollars_m,
               us.activist_max_pct, us.max_pct_book, us.n_funds_5pct_book,
               us.smart_money_n, us.s3_new, us.s4_add, us.activist_filings,
               us.cat8k_ma, us.cat8k_ctrl, tm.name, tm.sic_description
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        WHERE us.mcap_bucket IN ('nano','micro','small','mid') AND us.sec_type='common'"""))
    scored = []
    for r in rows:
        (tk, score, asym, mcap, bucket, ev, pb, eb, vse, f4_30, clu,
         actpct, maxpb, n5, smn, s3, s4, actf, c_ma, c_ctrl, name, sic) = r
        if tk in ETFs or tk in MEGA or is_biotech(sic): continue
        cheap   = (ev is not None and 0 < ev <= 12) or (pb is not None and 0 < pb < 2)
        below   = eb == "BELOW_ENTRY"
        insider = (f4_30 or 0) > 0 or (clu or 0) > 0
        activ   = (actpct or 0) >= 5 or (maxpb or 0) >= 5 or (n5 or 0) >= 1
        catal   = bool(c_ma) or bool(c_ctrl)
        backing = (smn or 0) >= 2 or (actf or 0) > 0 or ((s3 or 0) + (s4 or 0)) >= 1
        n_flags = sum([cheap, below, insider, activ, catal])
        if n_flags < 2 or not backing: continue
        idea = round((asym or 0) + 0.4 * (score or 0) + 4 * n_flags, 1)
        # plain-text rationale
        why = []
        if smn: why.append(f"{smn}×13F")
        if s3:  why.append(f"{s3} new")
        if s4:  why.append(f"{s4} add")
        if ev is not None and 0 < ev <= 12: why.append(f"{ev:.1f}x EV/EBITDA")
        if pb is not None and 0 < pb < 2:   why.append(f"{pb:.1f} P/B")
        if below and vse: why.append(f"{vse:.0f}% vs entry")
        if (f4_30 or 0) > 0: why.append(f"insider ${f4_30:.1f}M ≤30d")
        if (clu or 0) > 0:   why.append(f"cluster ${clu:.1f}M")
        if (actpct or 0) >= 5: why.append(f"activist {actpct:.0f}%")
        if c_ma:   why.append("M&A")
        if c_ctrl: why.append("control")
        cat = []
        if c_ma: cat.append("M&A")
        if c_ctrl: cat.append("CTRL")
        if (clu or 0) > 0: cat.append("clstr")
        scored.append([tk, idea, n_flags, round(score or 0, 1), round(asym or 0, 1),
                       mcap or "", bucket or "",
                       round(ev, 1) if ev is not None else "",
                       round(pb, 2) if pb is not None else "",
                       ("below" if below else "near" if eb == "NEAR_ENTRY"
                        else "above" if eb and "ABOVE" in eb else ""),
                       round(vse, 1) if vse else "",
                       round(f4_30, 2) if (f4_30 or 0) > 0 else "",
                       round(actpct or 0, 1),
                       " ".join(cat), (name or "")[:34], " · ".join(why),
                       *desc_for(conn, tk)])
    scored.sort(key=lambda x: -x[1])
    out = scored[:90]
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=6).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=8).number_format = '0.0"x"'
        ws.cell(row=ridx, column=9).number_format = '0.00"x"'
        ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=12).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=13).number_format = NUMFMT_PCT
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["P"].width = 76                     # Why
    ws.column_dimensions[get_column_letter(17)].width = 24   # Industry
    ws.column_dimensions[get_column_letter(18)].width = 80   # Business

def sheet_adversarial_review(wb, conn):
    """Surface the 52-agent adversarial review of the top picks — a red-team that
    stress-tested each headline name across data-quality, thesis-soundness and
    recency lenses, separating confirmed setups from data-inflated artifacts."""
    import json
    path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "data", "adversarial_review.json")
    if not os.path.exists(path):
        return
    data = json.load(open(path))
    ws = wb.create_sheet("Adversarial Review")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Adversarial Review — red-team of the top picks",
                f"A {data.get('agent_count','multi')}-agent stress test of the highest-ranked names across three lenses (data quality · thesis soundness · recency/news), separating confirmed setups from data-inflated artifacts. Scores as of {data.get('asof')}.", 4)
    row = 4
    write_section_heading(ws, row, "Synthesis — confirmed vs data-inflated", 4)
    row += 1
    for raw in data.get("synthesis", "").split("\n"):
        line = raw.strip()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        if not line:
            ws.row_dimensions[row].height = 6
            row += 1
            continue
        is_head = line.startswith("#")
        clean = (line.replace("**", "").replace("### ", "").replace("## ", "")
                     .replace("# ", "").lstrip("- ").strip())
        if line.startswith("- "):
            clean = "•  " + clean
        c = ws.cell(row=row, column=1, value=clean)
        c.font = SECTION_FONT if is_head else BODY_FONT
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(16, 15 * (1 + len(clean) // 105))
        row += 1
    row += 1
    write_section_heading(ws, row, "Per-ticker lens detail", 4)
    row += 1
    LENS = {"data-quality": "Data Quality", "thesis-soundness": "Thesis",
            "recency-and-news": "Recency/News"}
    write_table_header(ws, row, ["Ticker", "Score*", "Lens", "Assessment"])
    row += 1
    for rev in data.get("reviews", []):
        first = True
        for L in rev.get("lenses", []):
            txt = L.get("text", "").strip()
            ws.cell(row=row, column=1, value=rev["ticker"] if first else "").font = TICKER_FONT
            sc = ws.cell(row=row, column=2, value=rev.get("score_asof") if first else "")
            sc.font = BODY_FONT; sc.alignment = Alignment(horizontal="right", vertical="top")
            lc = ws.cell(row=row, column=3, value=LENS.get(L.get("lens"), L.get("lens")))
            lc.font = BODY_FONT; lc.alignment = Alignment(horizontal="left", vertical="top")
            ac = ws.cell(row=row, column=4, value=txt)
            ac.font = BODY_FONT
            ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
            ws.row_dimensions[row].height = max(28, 14 * (1 + len(txt) // 95))
            first = False
            row += 1
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 120
    ws.freeze_panes = "A4"

def _one_liner(s, limit=200):
    """Condense a stored business summary to a short one-liner for display."""
    if not s:
        return ""
    s = str(s).strip().replace("\n", " ")
    dot = s.find(". ")
    if 0 < dot <= limit:
        return s[:dot + 1]
    return s if len(s) <= limit else s[:limit].rstrip() + "…"

_DESC_CACHE = None
def desc_for(conn, ticker):
    """(industry, one-line business summary) for a ticker — memoized so every
    sheet can append a self-explanatory description per row."""
    global _DESC_CACHE
    if _DESC_CACHE is None:
        _DESC_CACHE = {}
        for r in conn.execute("""
            SELECT us.ticker,
                   COALESCE(yf.industry, tm.industry, tm.sic_description),
                   yf.business_summary, us.name
            FROM unified_signal us
            LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
            LEFT JOIN ticker_yf  yf ON yf.ticker = us.ticker"""):
            summ = r[2] or ""
            # Yahoo summaries open with "<Company Name>, together with its
            # subsidiaries," — a 90-char cut then just repeats the Name column.
            # Strip that lead-in so the Business cell carries new information.
            nm = (r[3] or "").rstrip(".")
            if nm and summ.upper().startswith(nm.upper()):
                summ = summ[len(nm):]
            summ = re.sub(r"^[,\s]*(together with its subsidiaries|and its subsidiaries"
                          r"|through its subsidiaries)?[,\s]*", "", summ, flags=re.I)
            _DESC_CACHE[r[0]] = ((r[1] or "")[:26], _one_liner(summ, 90))
    return _DESC_CACHE.get(ticker, ("", ""))

def sheet_ticker_reference(wb, conn):
    """Glossary: every ticker with name, sector, industry, and a short business
    summary — so any symbol in the workbook can be looked up. Sorted A–Z."""
    ws = wb.create_sheet("Ticker Reference")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Ticker Reference — name, industry, business",
                "Every symbol in the universe with its company name, sector, industry, market cap, and a one-line description of what it does. Sourced from Yahoo Finance + SEC. Sorted A–Z.", 6)
    hdr = ["Ticker", "Name", "Sector", "Industry", "Mcap", "Business Summary"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker,
               COALESCE(tm.name, yf.long_name)                       AS name,
               COALESCE(yf.sector, tm.sector)                        AS sector,
               COALESCE(yf.industry, tm.industry, tm.sic_description) AS industry,
               us.mcap_m,
               yf.business_summary
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf  yf ON yf.ticker = us.ticker
        WHERE COALESCE(tm.name, yf.long_name, yf.business_summary, yf.sector,
                       tm.sic_description) IS NOT NULL
        ORDER BY us.ticker"""))
    out = []
    for r in rows:
        if r[0] in ETFs: continue
        out.append([r[0], (r[1] or "")[:46], (r[2] or "")[:22],
                    (r[3] or "")[:30], r[4] or "", _one_liner(r[5])])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 100

TAB_COLORS = {
    # Navigation / meta — lightest
    "README":          "F2F2F2",
    "Legend":          "F2F2F2",
    "Fund Coverage":   "F2F2F2",
    "All Funds":       "F2F2F2",
    # Universe ranking — darkest
    "Action Dashboard": "1A1A1A",
    "Convergence":     "1A1A1A",
    "Best Ideas":      "262626",
    "Adversarial Review": "262626",
    "Top 100":         "262626",
    "Non-Biotech Top 100": "262626",
    "Asymmetry":       "262626",
    # Size buckets — mid-dark gradient
    "Nano (<$50M)":            "404040",
    "Micro ($50M–$300M)":      "595959",
    "Small ($300M–$2B)":       "595959",
    "Mid ($2B–$10B)":          "595959",
    # Signal sheets — mid
    "Material + New":          "808080",
    "Who's Buying":            "808080",
    "Revealed Preference":     "808080",
    "Activist 10+":            "808080",
    "Insider Buys ≤30d":       "808080",
    "Insider F4 Buys":         "808080",
    "Insider Clusters":        "808080",
    "Catalysts 8-K":           "808080",
    # Setup sheets — mid-light
    "In The Money":            "A6A6A6",
    "Valuation":               "A6A6A6",
    "Global Picks":            "A6A6A6",
    "Bill Miller":             "A6A6A6",
    # Reference / support — lighter
    "Unknown Mcap":            "BFBFBF",
    "All Positions":           "BFBFBF",
    "Ticker Reference":        "BFBFBF",
}

def main():
    conn = sqlite3.connect(DB)
    try:                                  # ensure momentum/drawdown are fresh
        import build_price_stats; build_price_stats.run()
    except Exception as e:
        print(f"  (price_stats skipped: {e})")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheet_readme(wb, conn)
    write_legend_sheet(wb, 1)
    sheet_action_dashboard(wb, conn)      # front-page scannable summary
    sheet_convergence(wb, conn)           # multi-signal convergence matrix
    sheet_dossier(wb, conn)               # per-ticker consolidated view
    if _has_prior(conn):
        sheet_qoq_change(wb, conn)        # quarter-over-quarter accumulation/distribution
    sheet_best_ideas(wb, conn)
    sheet_adversarial_review(wb, conn)
    write_signal_sheet(wb, conn, "Top 100",
        where_extra="AND us.mcap_bucket != 'unknown'", limit=140,
        subtitle="Top 100 by unified_score across the full 5,862-ticker universe (ex-ETF, ex-mega-cap, mcap known).")
    for bucket, title in [("nano","Nano (<$50M)"),
                          ("micro","Micro ($50M–$300M)"),
                          ("small","Small ($300M–$2B)"),
                          ("mid","Mid ($2B–$10B)")]:
        write_signal_sheet(wb, conn, title,
            where_extra=f"AND us.mcap_bucket = '{bucket}'", limit=60,
            subtitle=f"Top {bucket} cap by unified_score. Ex-ETF, ex-mega.")
    sheet_best_in_bucket(wb, conn)
    write_signal_sheet(wb, conn, "Material + New",
        where_extra="AND (us.s3_new + us.s4_add) >= 2 AND us.mcap_bucket != 'unknown'",
        limit=80,
        subtitle="≥2 funds adding to existing (S4) OR initiating major new (S3) — smart money is BUILDING.")
    sheet_whos_buying(wb, conn)
    sheet_activist(wb, conn)
    sheet_broker_radar(wb, conn)
    sheet_latent_ownership(wb, conn)
    sheet_nport(wb, conn)
    sheet_insider_recent(wb, conn)
    sheet_insider_f4(wb, conn)
    sheet_clusters(wb, conn)
    write_signal_sheet(wb, conn, "Non-Biotech Top 100",
        where_extra="AND us.mcap_bucket != 'unknown'", limit=140,
        subtitle="Top ex-biotech, ex-ETF, ex-mega.", exclude_biotech=True)
    sheet_in_the_money(wb, conn)
    sheet_asymmetry(wb, conn)
    sheet_revealed_pref(wb, conn)
    sheet_valuation(wb, conn)
    sheet_catalysts(wb, conn)
    sheet_global_picks(wb, conn)
    sheet_bill_miller(wb, conn)
    sheet_unknown(wb, conn)
    sheet_all_holdings_consolidated(wb, conn)
    sheet_fund_coverage(wb, conn)
    sheet_all_funds(wb, conn)
    sheet_ticker_reference(wb, conn)

    # AutoFilter on single-table sheets (header at row 4) — lets the reader
    # sort / filter by any column in Excel. Multi-table sheets are excluded.
    AF_SHEETS = {
        "Best Ideas",
        "Top 100", "Nano (<$50M)", "Micro ($50M–$300M)", "Small ($300M–$2B)",
        "Mid ($2B–$10B)", "Material + New", "Activist 10+", "Insider Buys ≤30d",
        "Insider F4 Buys", "Insider Clusters", "Non-Biotech Top 100", "In The Money",
        "Asymmetry", "Revealed Preference", "Valuation", "Catalysts 8-K",
        "Global Picks", "Unknown Mcap", "All Positions", "All Funds",
        "Fund Coverage", "Ticker Reference",
    }
    for sn in AF_SHEETS:
        if sn in wb.sheetnames:
            ws = wb[sn]
            if ws.max_row > 4:
                ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Tab colour-coding — grayscale tones by theme
    for sname, color in TAB_COLORS.items():
        if sname in wb.sheetnames:
            wb[sname].sheet_properties.tabColor = color

    add_contents_index(wb["README"], wb.sheetnames)
    set_print_layout(wb)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
