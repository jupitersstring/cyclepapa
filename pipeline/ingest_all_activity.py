"""All Activity ingest — the master flow log we should have been using.

The original workbook's `All Activity` tab has 3,849 cross-fund rows with
clean columns: Group | Fund | Category | Ticker | Detail1 | Detail2 | Detail3 | Source.
This is far cleaner than re-parsing per-fund tabs, so we ingest it FIRST
and merge with the per-fund-tab data.

Also ingests the 8 pre-aggregated summary tabs:
  Consensus Buys, Highest Conviction, Conviction Adds, Micro-Cap Conviction Adds,
  Activist Catalysts, Multi-Fund New Inits, Asymmetric Summary
"""
import os, re, sqlite3
import openpyxl

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fund_activity_last_6mo.xlsx")

CATEGORY_TO_SECTION = {
    "highest conviction": 1, "(1)": 1,
    "threshold": 2, ">=5%": 2, "5%": 2, "(2)": 2,
    "new position": 3, "(3)": 3,
    "material": 4, "existing": 4, "(4)": 4,
}
TICKER_RE = re.compile(r"^([A-Z][A-Z0-9.\-]{0,7})(?:\s|\(|$)")
# Japanese 4-digit ticker pattern: "4464 Soft99" or "7984 JT" or "1234 / Company"
JP_TICKER_RE = re.compile(r"^(\d{4})(?:\s*(?:JT|JP|TYO|/)?\s+|[\s/]\s*)([A-Z][\w &/.,'-]{1,60})")
# European/Asian exchange-suffixed tickers: "BZU.IM", "VOD.L", "ASML.AS", "STAN LN"
INTL_RE = re.compile(r"^([A-Z][A-Z0-9]{0,5}(?:\.[A-Z]{1,3}|\s+[A-Z]{2})?)(?:\s|$)")
PCT_RE = re.compile(r"([\+\-]?\d+(?:\.\d+)?)\s*%")
DOLLAR_RE = re.compile(r"\$([\d,]+(?:\.\d+)?)\s*([MBK]?)", re.I)
DATE_RE = re.compile(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})")

NOT_TICKERS = {"NEW","ADD","CIK","AUM","RAUM","ADV","SEC","FDA","CEO","CFO","COO","CIO","CMO",
               "EU","US","UK","HK","TX","NY","CA","Q1","Q2","Q3","Q4","FY","YTD","TTM",
               "EBITDA","FCF","EPS","ROIC","ROE","ROA","NAV","ETF","SPAC","IPO","ESG","AI",
               "JV","PT","IV","II","III","DCF","DTC","PDUFA","NDA","IND","MFN",
               "JAN","FEB","MAR","APR","JUN","JUL","AUG","SEP","OCT","NOV","DEC",
               "TIER","GROUP","FUND","NEXT","SAME","CUT","MAX","TOTAL","NET","NAV","SUB",
               "OVER","UNDER","ABOVE","BELOW","NONE","BUYS","NOTES","LAST","BASE","LARGE",
               "HOLD","KEEP","TOP","HOT","WAR","ARE","WAS","HAD","HAS","HIS","ITS"}

def detect_section(text):
    if not text: return 0
    t = str(text).lower()
    for k, v in CATEGORY_TO_SECTION.items():
        if k in t: return v
    return 0

def parse_pct(s):
    if not s: return None
    m = PCT_RE.search(str(s))
    return float(m.group(1)) if m else None

def parse_dollar(s):
    if not s: return None
    m = DOLLAR_RE.search(str(s))
    if not m: return None
    v = float(m.group(1).replace(",", ""))
    mult = {"B": 1000.0, "M": 1.0, "K": 0.001, "": 1.0}[m.group(2).upper()]
    return v * mult

def extract_ticker(s):
    if not s: return None, None
    s = str(s).strip()
    # Japanese 4-digit code first (Effissimo, Oasis etc.)
    jm = JP_TICKER_RE.match(s)
    if jm:
        return f"{jm.group(1)}.T", jm.group(2).strip()[:60]
    # Generic US/intl alphanumeric
    first = s.split()[0] if s else ""
    m = INTL_RE.match(s) or TICKER_RE.match(first)
    if not m: return None, None
    t = m.group(1).strip(".- ").replace(" ", ".")
    if t in NOT_TICKERS or len(t) < 2: return None, None
    if t.isdigit() and len(t) != 4: return None, None  # allow 4-digit JP, reject others
    rest = s[len(first):].strip(" |/-:")
    co = None
    cm = re.match(r"\(([^)]{2,60})\)", rest) or re.match(r"([A-Z][A-Za-z &.,'-]{2,60})", rest)
    if cm: co = cm.group(1).strip()
    return t, co

def ingest_all_activity(conn, snap="2026-05-30"):
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    if "All Activity" not in wb.sheetnames:
        print("All Activity tab not found"); return 0
    ws = wb["All Activity"]
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row): continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Schema: Group | Fund | Category | Ticker col | Detail1 | Detail2 | Detail3 | Source
        if len(cells) < 4: continue
        group, fund, category, ticker_cell = cells[0], cells[1], cells[2], cells[3]
        if not fund or not ticker_cell: continue
        section = detect_section(category)
        ticker, company = extract_ticker(ticker_cell)
        if not ticker: continue
        # Pull % and $ from detail cells
        details = cells[4:7] if len(cells) >= 7 else cells[4:]
        all_detail = " | ".join(d for d in details if d)
        pct = parse_pct(all_detail)
        usd = parse_dollar(all_detail)
        # company fallback from detail-1 if it's a name
        if not company and details and not re.search(r"\d", details[0]):
            company = details[0][:60]
        pct_kind = "company" if (section == 2 or any(k in all_detail.lower() for k in ("13d","13g","threshold"))) else "book"
        dt = None
        dm = DATE_RE.search(all_detail)
        if dm: dt = f"{dm.group(1)}-{int(dm.group(2)):02d}-{int(dm.group(3)):02d}"
        raw = " | ".join(c for c in cells if c)
        conn.execute("""INSERT INTO fund_positions
            (fund,ticker,company,section,pct_value,pct_kind,dollar_m,change_text,event_date,raw_text,asof)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (fund, ticker, (company or "")[:80], section, pct, pct_kind, usd,
             all_detail[:160], dt, raw[:300], snap))
        n += 1
    print(f"  All Activity: {n} rows ingested")
    return n

def ingest_summary_tabs(conn):
    """Ingest pre-aggregated cross-fund summary tabs into separate tables."""
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    conn.executescript("""
    DROP TABLE IF EXISTS conviction_adds;
    DROP TABLE IF EXISTS highest_conviction;
    DROP TABLE IF EXISTS consensus_buys;
    DROP TABLE IF EXISTS activist_catalysts;
    DROP TABLE IF EXISTS microcap_conviction_adds;
    DROP TABLE IF EXISTS multi_fund_new_inits;
    CREATE TABLE conviction_adds (
      ticker TEXT PRIMARY KEY, n_funds_adding INTEGER, n_new_inits INTEGER,
      sum_dollar_m REAL, max_dollar_m REAL, max_pct_add REAL,
      funds TEXT, narratives TEXT);
    CREATE TABLE microcap_conviction_adds (
      ticker TEXT PRIMARY KEY, n_funds_adding INTEGER, n_new_inits INTEGER,
      sum_dollar_m REAL, max_dollar_m REAL, max_pct_add REAL,
      funds TEXT, narratives TEXT);
    CREATE TABLE highest_conviction (
      ticker TEXT, n_funds INTEGER, max_pct_book REAL, n_threshold INTEGER,
      n_hyper INTEGER, flags TEXT, top_funds TEXT,
      PRIMARY KEY (ticker, max_pct_book));
    CREATE TABLE consensus_buys (
      ticker TEXT PRIMARY KEY, n_funds INTEGER, max_pct REAL,
      flags TEXT, funds TEXT, narratives TEXT);
    CREATE TABLE activist_catalysts (
      ticker TEXT PRIMARY KEY, n_threshold_filings INTEGER, funds_5pct TEXT,
      n_funds_positive INTEGER, max_pct_any_fund REAL, flags TEXT, sample_notes TEXT);
    CREATE TABLE multi_fund_new_inits (
      ticker TEXT, n_funds_initiating INTEGER, sum_dollar_m REAL, funds TEXT,
      PRIMARY KEY (ticker, n_funds_initiating));
    """)

    def grab(tab, table, cols):
        if tab not in wb.sheetnames: return 0
        ws = wb[tab]
        n = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            vals = list(row[:len(cols)]) + [None]*(len(cols)-len(row))
            tkr = str(vals[0]).strip()
            if not tkr or len(tkr) > 8: continue
            try:
                ph = ",".join("?"*len(cols))
                conn.execute(f"INSERT OR REPLACE INTO {table} VALUES ({ph})", vals)
                n += 1
            except sqlite3.IntegrityError:
                pass
        print(f"  {tab}: {n} rows ingested")
        return n

    grab("Conviction Adds", "conviction_adds",
         ["ticker","n_funds_adding","n_new_inits","sum_dollar_m","max_dollar_m","max_pct_add","funds","narratives"])
    grab("Micro-Cap Conviction Adds", "microcap_conviction_adds",
         ["ticker","n_funds_adding","n_new_inits","sum_dollar_m","max_dollar_m","max_pct_add","funds","narratives"])
    grab("Highest Conviction", "highest_conviction",
         ["ticker","n_funds","max_pct_book","n_threshold","n_hyper","flags","top_funds"])
    grab("Consensus Buys", "consensus_buys",
         ["ticker","n_funds","max_pct","flags","funds","narratives"])
    grab("Activist Catalysts", "activist_catalysts",
         ["ticker","n_threshold_filings","funds_5pct","n_funds_positive","max_pct_any_fund","flags","sample_notes"])
    grab("Multi-Fund New Inits", "multi_fund_new_inits",
         ["ticker","n_funds_initiating","sum_dollar_m","funds"])

def run():
    conn = sqlite3.connect(DB)
    # Clear all-activity-sourced rows so we can re-ingest cleanly
    print("ingesting summary tabs:")
    ingest_summary_tabs(conn)
    print("\ningesting All Activity master log:")
    ingest_all_activity(conn)
    conn.commit()

    # Coverage check
    print(f"\nfund_positions row total: {conn.execute('SELECT COUNT(*) FROM fund_positions').fetchone()[0]}")
    print(f"Unique tickers: {conn.execute('SELECT COUNT(DISTINCT ticker) FROM fund_positions').fetchone()[0]}")
    print(f"conviction_adds: {conn.execute('SELECT COUNT(*) FROM conviction_adds').fetchone()[0]}")
    print(f"highest_conviction: {conn.execute('SELECT COUNT(*) FROM highest_conviction').fetchone()[0]}")
    print(f"activist_catalysts: {conn.execute('SELECT COUNT(*) FROM activist_catalysts').fetchone()[0]}")

if __name__ == "__main__":
    run()
