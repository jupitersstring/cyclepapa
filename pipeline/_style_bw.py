"""Shared Harvard-academic monochrome aesthetic for cyclepapa workbooks.

Design grammar:
  - All black on white. No fills, no shading, no colors.
  - Single body font size (10pt Times New Roman) — like an HBR article.
  - Headers in SMALL CAPS, bold, same size as body.
  - Thin black bottom rule under headers (FT/HBR style).
  - Title 14pt bold; subtitle 10pt italic.
  - Section headers: 11pt bold, small caps.
  - Generous row heights (18 pt) for breathability.
  - Right-align numbers, left-align text.
  - Hairline gray rules between rows (D9D9D9), only where helpful.
  - No gridlines visible.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

TNR = "Times New Roman"

# Single body size — restraint
SIZE_BODY = 10
SIZE_HDR = 10
SIZE_SECTION = 11
SIZE_SUBTITLE = 10
SIZE_TITLE = 15

BLACK = "000000"
LIGHT_GREY = "BFBFBF"
HAIRLINE = "D9D9D9"
MUTED = "3F3F3F"
# Times-Lattice style guide: the ONLY two accent inks. Colour is data — lapis =
# good/improving, crimson = bad/deteriorating. Used ONLY as directional font
# colour on value cells (never fills, headings, or borders).
LAPIS = "061933"
CRIMSON = "7A0019"

def ink_for(value, higher_is_better=True):
    """Return LAPIS (good) / CRIMSON (bad) / None for a signed value."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v == 0:
        return None
    good = (v > 0) if higher_is_better else (v < 0)
    return LAPIS if good else CRIMSON

def color_directional(ws, first_row, last_row, cols, higher_is_better=True):
    """Colour the given 1-indexed columns lapis/crimson by the cell's sign, over
    [first_row, last_row]. cols may be one int or a list. Keeps Times + size."""
    if isinstance(cols, int):
        cols = [cols]
    for r in range(first_row, last_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            ink = ink_for(cell.value, higher_is_better)
            if ink:
                f = cell.font
                cell.font = Font(name=f.name or TNR, size=f.size or SIZE_BODY,
                                 bold=f.bold, italic=f.italic, color=ink)

def color_fixed(ws, first_row, last_row, cols, ink):
    """Colour non-blank/non-zero cells in `cols` a FIXED ink — for columns whose
    mere presence carries direction (a buy $ is always good=lapis, a sell $ always
    bad=crimson, regardless of sign)."""
    if isinstance(cols, int):
        cols = [cols]
    for r in range(first_row, last_row + 1):
        for c in cols:
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v in (None, "", "—", 0, 0.0):
                continue
            f = cell.font
            cell.font = Font(name=f.name or TNR, size=f.size or SIZE_BODY,
                             bold=f.bold, italic=f.italic, color=ink)

# Fonts
TITLE_FONT     = Font(name=TNR, bold=True, size=SIZE_TITLE, color=BLACK)
SUBTITLE_FONT  = Font(name=TNR, italic=True, size=SIZE_SUBTITLE, color=BLACK)
SECTION_FONT   = Font(name=TNR, bold=True, size=SIZE_SECTION, color=BLACK)
HDR_FONT       = Font(name=TNR, bold=True, size=SIZE_HDR, color=BLACK)
BODY_FONT      = Font(name=TNR, size=SIZE_BODY, color=BLACK)
BODY_ITALIC    = Font(name=TNR, italic=True, size=SIZE_BODY, color=BLACK)
TICKER_FONT    = Font(name=TNR, bold=True, size=SIZE_BODY, color=BLACK)
MONO_FONT      = Font(name="Consolas", size=SIZE_BODY, color=BLACK)

# Borders
THIN_BLK       = Side(border_style="thin", color=BLACK)
HAIRLINE_SIDE  = Side(border_style="hair", color=HAIRLINE)
NO_SIDE        = Side(border_style=None)

# header row: black 0.5pt under, nothing above/sides
HDR_BORDER     = Border(top=NO_SIDE, bottom=THIN_BLK, left=NO_SIDE, right=NO_SIDE)
# body row: hairline gray bottom only
ROW_BORDER     = Border(top=NO_SIDE, bottom=HAIRLINE_SIDE, left=NO_SIDE, right=NO_SIDE)
TITLE_BORDER   = Border(bottom=THIN_BLK, top=NO_SIDE, left=NO_SIDE, right=NO_SIDE)

NO_FILL = PatternFill(fill_type=None)

NUMFMT_USD  = '"$"#,##0'
NUMFMT_PCT  = '0.0"%"'
NUMFMT_NUM  = '#,##0.0'
NUMFMT_INT  = '#,##0'
NUMFMT_USD2 = '"$"#,##0.00'

# Smart scale-aware money format — input is in millions.
#   < 1,000          → "$NNN M"     (i.e. < $1B)
#   1,000 – 999,999  → "$N.NN B"   (i.e. $1B to $1T)
#   >= 1,000,000     → "$N.N T"    (>= $1T)
NUMFMT_MCAP = '[>=1000000]"$"#,##0.0,,"T";[>=1000]"$"#,##0.0,"B";"$"#,##0" M"'
# Same idea for dollar quantities in $M (Form 4 $M, cluster $M, position $M):
NUMFMT_M_TO_B = '[>=1000]"$"#,##0.0,"B";"$"#,##0.0" M"'

# Times-Lattice density spec — rows are separated by a hairline, not padding, so
# heights are tight (the ledger-like feel). Was 22/18/30 (airy); now dense.
HDR_HEIGHT = 17
BODY_HEIGHT = 14
TITLE_HEIGHT = 26

# ---------- helpers ----------
MEDIUM_BLK = Side(border_style="medium", color=BLACK)

def write_title(ws, title, subtitle, ncols):
    """Broadsheet masthead — an engraved nameplate: the title UPPERCASE bold with
    a fleur-de-lis, over an italic-muted subline, closed by a MEDIUM black rule
    (the 19th-century financial-broadsheet look of the Times-Lattice style guide)."""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = TITLE_HEIGHT
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws.cell(row=1, column=1, value="⚜  " + str(title).upper())   # ⚜ fleur + nameplate
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center")
    # heavy rule directly under the nameplate
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).border = Border(bottom=MEDIUM_BLK, top=NO_SIDE, left=NO_SIDE, right=NO_SIDE)
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = SUBTITLE_FONT
    s.alignment = Alignment(horizontal="left", vertical="center")
    # thin black rule closing the masthead under the subline
    for col in range(1, ncols + 1):
        ws.cell(row=2, column=col).border = Border(bottom=THIN_BLK, top=NO_SIDE, left=NO_SIDE, right=NO_SIDE)

def write_section_heading(ws, row, text, ncols):
    """Section heading: small caps, 11pt bold, hairline rule beneath."""
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text.upper())
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal="left", vertical="bottom")

def write_table_header(ws, row, cols):
    ws.row_dimensions[row].height = HDR_HEIGHT
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h.upper())
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="right" if i > 1 else "left",
                                 vertical="bottom", wrap_text=False)
        c.border = HDR_BORDER

def write_table_rows(ws, rows, start_row, ticker_col=1, hairline=True, blank="—"):
    """Write rows. First column treated as ticker (bold, left). Rest right-aligned if numeric.

    Empty cells (None or "") are rendered as a placeholder em-dash so every cell
    reads as intentional — "—" is the standard "not applicable / not available"
    mark. Genuine zeros are preserved as 0 (callers pass `x or 0` for counts).
    """
    for i, row in enumerate(rows):
        ws.row_dimensions[start_row + i].height = BODY_HEIGHT
        for j, v in enumerate(row, 1):
            is_blank = v is None or (isinstance(v, str) and v.strip() == "")
            display = blank if is_blank else v
            c = ws.cell(row=start_row + i, column=j, value=display)
            is_num = isinstance(v, (int, float))
            if j == ticker_col and isinstance(v, str) and not is_blank:
                c.font = TICKER_FONT
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.font = BODY_FONT
                # numbers and placeholders right-align (placeholder usually
                # stands in for a numeric column); real text left-aligns.
                c.alignment = Alignment(
                    horizontal="right" if (is_num or is_blank) else "left",
                    vertical="center")
            if hairline:
                c.border = ROW_BORDER

def autosize(ws):
    """Column widths chosen for visual rhythm — content + 2."""
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            try:
                v = ws.cell(row=row_idx, column=col_idx).value
            except Exception:
                continue
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 50))
        ws.column_dimensions[letter].width = max(8, min(max_len + 2.5, 42))

def set_default_font(wb):
    """Apply Times New Roman as the workbook default styles where possible."""
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = None

# ---------- legend / glossary (shared across both workbooks) ----------
LEGEND = [
    ("Position classification", [
        ("Section 1 / S1", "Top-conviction holding — the manager's highest-conviction, largest-weight positions."),
        ("Section 3 / S3", "New major position — a newly initiated significant holding this period (an initiation)."),
        ("Section 4 / S4", "Material add — a meaningful increase to an existing position."),
        ("Section 5", "Researcher-flagged position (letters, interviews, primary research) without a 13F section."),
        ("ST  (prefix)", "Within-style: counted only across funds in THIS macro-style. e.g. ST S3 = funds in this style initiating a new major position; ST Holders = holders within the style."),
        ("Sub (prefix)", "Within sub-group: counted only across funds in this sub-group tier (e.g. Sub Holders)."),
        ("Uni (prefix)", "Universe-wide: counted across every tracked fund (e.g. Uni 13F)."),
    ]),
    ("Smart money & conviction", [
        ("13F", "Number of distinct 13F filers (funds) holding the name."),
        ("Holders", "Count of funds holding the name (overall, or within style / sub-group when prefixed)."),
        ("pB Max", "Largest single-fund position weight — the maximum % of any one fund's book in the name."),
        ("pB ≥5%", "Number of funds with at least 5% of their book in the name (a concentration cluster)."),
        ("%Book", "A position's weight as a percent of the fund's reported equity book."),
        ("Score", "Unified score: log(13F)×2 + section weights + concentration + activist + insider + catalyst − sells (full formula on README)."),
        ("Global Score", "Score excluding US-only signals (Form 4, clusters) so foreign listings rank on equal footing."),
        ("Rev Pref", "Revealed preference = 2×S3 + 1×S4 + 0.5×S1 — measures active accumulation, not static holding."),
        ("Asym", "Asymmetry score — margin-of-safety (cheap valuation + below smart-money entry) × upside (conviction + catalyst + small-cap room)."),
        ("Why", "The top-3 terms driving the Score, as compact codes: sm=smart-money holders, s1=top-pick funds, s3=new-position funds, s4=add funds, pb=low price/book, pb5=funds ≥5% book, clu$=insider-cluster $, f4buy=insider buys, f4rec=very-recent buys, f4sell=insider sells, act=activist %, 8k=catalyst, micro=small-cap, entry=below smart-money entry. e.g. 's1 24 · pb5 18 · pb 14'."),
        ("Lift  (Signature picks)", "How much a fund STYLE over-indexes on a name vs the whole universe: (style holder-share) ÷ (universe holder-share). >1 = the style's distinctive bet."),
    ]),
    ("Activist & insider (SEC)", [
        ("13D", "Number of SC 13D / 13G beneficial-ownership filings (a ≥5% stake)."),
        ("Act %", "Largest activist stake disclosed via 13D/G (max percent of share class)."),
        ("Clu $M", "Insider cluster size — total dollars of a live (≤180-day) multi-insider open-market buy cluster."),
        ("F4 Buy / F4 $M", "Form 4 open-market insider purchases (transaction code P), in millions of dollars."),
        ("F4 Buy ≤30d / 180d", "Insider open-market buys reported within the last 30 / 180 days."),
        ("F4 Sell", "Form 4 open-market insider sales (code S) — a counter-signal."),
        ("Weighted $M", "Recency-weighted insider buys: ≤30d ×1.0, 31–60d ×0.6, 61–120d ×0.3, 121–180d ×0.1."),
        ("# Insiders / # Buyers", "Distinct insiders buying in the window."),
        ("Cluster? / clstr", "A live insider buy cluster is present for the name."),
    ]),
    ("Valuation", [
        ("Mcap", "Market capitalization in USD ($M, auto-scaled to $B / $T). Foreign caps are FX-converted to USD."),
        ("ADV $M", "Average daily dollar volume traded (3-month, $M) — the tradeability check: a high score on a $0.3M/day nano is hard to act on."),
        ("EV/EBITDA", "Enterprise value ÷ trailing EBITDA (shown ×). A negative figure means negative EBITDA."),
        ("P/B", "Price ÷ book value per share (shown ×). A negative figure means negative book equity."),
        ("P/E  /  Fwd P/E", "Price ÷ trailing (or forward) EPS (shown ×)."),
        ("Rev Gr %", "Year-over-year revenue growth. Negative (crimson) flags a possible value trap on an otherwise-cheap multiple."),
        ("Margin %", "Net profit margin. Negative (crimson) = loss-making."),
        ("EV/Rev  /  PEG", "Enterprise value ÷ revenue; PEG = P/E ÷ growth. Cover names with no meaningful EV/EBITDA."),
    ]),
    ("Price & momentum (from daily closes)", [
        ("3mo %  /  20d %", "Price change over the last ~3 months / ~20 trading days. Positive = lapis, negative = crimson."),
        ("Off Hi %", "Percent below the 3-month high (drawdown). A name deep off its high with insiders buying is a different setup from one at highs."),
    ]),
    ("Quarter-over-quarter (QoQ Change sheet)", [
        ("Net Funds", "(New + Added) − (Trimmed + Exited) funds this quarter vs each fund's prior 13F. Lapis = accumulating, crimson = distributing."),
        ("New / Added / Trimmed / Exited", "Count of funds that started / grew (>5%) / cut (>5%) / closed the position, matched on CUSIP + share count."),
        ("Δ Shares %", "Aggregate share-count change across all funds vs the prior quarter."),
        ("Form", "Security form being accumulated, from each 13F line's titleOfClass. \"common\" = ordinary common/ordinary shares (a clean directional bet). \"+preferred / +warrant / +unit / +right / +note\" flags that non-common equity forms are held under this ticker — optionality or financing, which should NOT be read as the same conviction as buying common."),
    ]),
    ("Broker Swap Radar", [
        ("Δ Sh (M) / Δ % Out", "Quarter-over-quarter share-count change in ONE swap-desk broker's 13F (UBS, GS, MS, JPM...), absolute and as % of shares outstanding."),
        ("Idio %", "This desk's move as a share of ALL tracked desks' movement in the name. High = idiosyncratic (swap-hedge-like); low = every desk moved (index/ETF flow)."),
        ("Why it matters", "An activist building via total-return swaps appears on NO 13F/13D of their own — the counterparty desk hedges with physical shares, which print HERE. Leads, not proof: baskets and custody flows also move desks."),
    ]),
    ("Latent Ownership (13D text)", [
        ("# Feat / Hidden Features", "Count and list of economic-control features parsed from the holder's 13D: prefunded/ordinary warrants, convertibles, ownership blocker, board-designation rights, registration rights, ROFR, anti-dilution, standstill, disclosed swap."),
        ("Blocker %", "The ownership-limitation ceiling (4.99 / 9.99 / 19.99%) — the holder's economic exposure can sit just under it while the header % looks small; the blocker is often contractually raisable."),
        ("Swap Cpty", "A total-return / cash-settled swap named in the 13D text, with counterparty desk if disclosed — the clearest hidden-economic-exposure tell; cross-check the Broker Swap Radar."),
    ]),
    ("N-PORT Monthly", [
        ("Series (fund)", "A registered fund's monthly N-PORT-P holdings — fresher than quarterly 13F and inclusive of FOREIGN listings 13F never reports. Supplementary RIC data, not counted as 13F smart money."),
    ]),
    ("Entry / setup", [
        ("Entry / Bucket", "Where the current price sits versus the smart-money cost anchor: below / near / above."),
        ("Anchor $", "Estimated smart-money cost basis (cost_basis / filing text / Form-4 buy average / 80th-percentile)."),
        ("vs Entry %", "Current price relative to the anchor (negative = trading below where smart money bought)."),
        ("Now $", "Current share price (USD)."),
        ("ER %", "Expected return — base-rate-weighted historical 12-month excess for the name's factor tags."),
    ]),
    ("Catalysts (8-K, ≤180 days)", [
        ("M&A", "Item 1.01 / 2.01 — merger, acquisition, or material definitive agreement."),
        ("Ctrl / CTRL", "Item 5.01 — change of control of the registrant."),
        ("Director", "Item 5.02 — departure / appointment of directors or officers."),
        ("PIPE", "Item 3.02 — unregistered sale of equity (potential dilution)."),
        ("Bnk", "Item 1.03 — bankruptcy or receivership."),
        ("Total Events", "Count of distinct 8-K material events in the window."),
    ]),
    ("Size buckets", [
        ("nano", "Under $50M market cap."),
        ("micro", "$50M – $300M."),
        ("small", "$300M – $2B."),
        ("mid", "$2B – $10B."),
        ("large", "Over $10B."),
        ("unknown", "Market cap unresolved (foreign / SPAC / warrant / defunct)."),
    ]),
    ("Sources & symbols", [
        ("13F-HR", "SEC quarterly institutional holdings filing (the standard smart-money source). NOTE: 13F holdings are quarter-END positions filed up to 45 days later — the smart-money columns can be up to ~3–4 months old (see each sheet's as-of date). Form 4 / 13D / 8-K / valuation columns are near-current."),
        ("XLSX", "Research-team position classification (sections 1 / 3 / 4 / 5)."),
        ("SC 13D/G", "SEC beneficial-ownership filing (a ≥5% stake)."),
        ("Value $M", "Position market value in $M (13F holdings); blank for 13D/G rows."),
        ("—", "An em-dash means not applicable / not available for that cell."),
    ]),
    ("Colour (Times-Lattice — 'colour is data')", [
        ("Lapis blue", "Good / improving: positive momentum & growth, insider buying, net funds accumulating, cheap valuation."),
        ("Crimson red", "Bad / deteriorating: negative momentum & growth, insider selling, net funds distributing."),
        ("Faint wash", "Score / valuation / vs-entry heatmaps — a lapis or crimson tint at ~7% strength, darker = more attractive."),
        ("Black only", "Everything structural. Colour appears ONLY where it carries a good/bad meaning, never for decoration."),
    ]),
]

def write_legend_sheet(wb, index=1):
    """Insert a 'Legend' sheet defining every column header / abbreviation used
    across the workbook. Shared so both books document the same vocabulary."""
    ws = wb.create_sheet("Legend", index)
    ws.sheet_view.showGridLines = False
    write_title(ws, "Legend — column definitions & abbreviations",
                "What every column header and code means. Prefixes: ST = within macro-style · Sub = within sub-group · Uni = universe-wide.", 2)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 116
    row = 4
    for group, items in LEGEND:
        write_section_heading(ws, row, group, 2)
        row += 1
        for term, definition in items:
            a = ws.cell(row=row, column=1, value=term)
            a.font = TICKER_FONT
            a.alignment = Alignment(horizontal="left", vertical="top")
            b = ws.cell(row=row, column=2, value=definition)
            b.font = BODY_FONT
            b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 30 if len(definition) <= 110 else 44
            a.border = ROW_BORDER
            b.border = ROW_BORDER
            row += 1
        row += 1  # gap between groups
    ws.freeze_panes = "A4"
    return ws

def add_contents_index(ws, sheetnames, exclude=("README",)):
    """Append a clickable 'Contents' index to the README sheet — one internal
    hyperlink per sheet so the reader can jump straight to any tab. Monochrome:
    black text, underlined to signal it is clickable."""
    row = (ws.max_row or 1) + 2
    write_section_heading(ws, row, "Contents — click to open a sheet", 2)
    row += 1
    link_font = Font(name=TNR, size=SIZE_BODY, color=BLACK, underline="single")
    for name in sheetnames:
        if name in exclude:
            continue
        c = ws.cell(row=row, column=1, value=name)
        c.hyperlink = f"#'{name}'!A1"
        c.font = link_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 16
        row += 1

def set_print_layout(wb, header_rows=4):
    """Sensible print defaults on every sheet: landscape, fit-to-width, and the
    title/header rows repeated at the top of each printed page."""
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_options.horizontalCentered = False
        try:
            ws.print_title_rows = f"1:{header_rows}"
        except Exception:
            pass
