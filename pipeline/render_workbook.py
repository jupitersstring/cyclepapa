"""Render investment_archetypes.xlsx FROM the SQLite DB.

Same content as the prior 22-sheet workbook, but every cell is a query result
not a hardcoded dict. xlsx becomes a render target, not a source of truth.

Sheets produced:
  - 1_VERIFIED_TIER1  (live, joined with prices + liquidity + base rates)
  - 2_BASE_RATES       (backtest results by bucket → numbers users can challenge)
  - 3_BACKTEST_DETAIL  (every event + 6/12/18mo returns)
  - 4_MASTER_CSV       (full master_candidates table, typed columns)
  - 5_CATALYSTS_LIVE   (catalysts with computed days_remaining)
  - 6_FORM4_BUYS       (verified open-market buys, code=P)
  - 7_LIQUIDITY        (ADV + days-to-exit sizing budget)
  - 8_ARCHETYPES       (original archetype theses, preserved)
"""
import os, sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "investment_archetypes.xlsx")

HEAD = Font(bold=True, color="FFFFFF", size=11)
HFILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YEL   = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
WRAP  = Alignment(wrap_text=True, vertical="top")

def write_sheet(wb, name, headers, rows, highlight_col=None, highlight_neg=False, widths=None):
    ws = wb.create_sheet(name[:31])
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font = HEAD; c.fill = HFILL
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = WRAP
            if highlight_col == c and highlight_neg and isinstance(v, (int, float)):
                cell.fill = GREEN if v > 0 else RED
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def run():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    wb = openpyxl.Workbook(); del wb["Sheet"]

    # 1. Empirical Tier 1 — ranked by base-rate-weighted ER (the new methodology)
    rows = list(conn.execute("""
        SELECT c.ticker, c.name, c.sector, c.price, c.mcap_m,
               e.weighted_excess_12m AS er_12m, e.best_tag, e.best_tag_excess,
               e.worst_tag, e.worst_tag_excess, e.cluster_live,
               l.adv_usd_m, l.days_to_exit_1pct_adv10,
               c.kill_criteria, c.factor_tags
        FROM candidates c
        JOIN expected_return e ON e.ticker=c.ticker
        LEFT JOIN liquidity l ON l.ticker=c.ticker
        ORDER BY e.weighted_excess_12m DESC"""))
    fmt = lambda v: f"{v*100:+.0f}%" if v is not None else ""
    write_sheet(wb, "1_EMPIRICAL_TIER1",
        ["Ticker","Name","Sector","Price","Mcap $M","ER 12m","Best tag","Best %",
         "Worst tag","Worst %","Live cluster?","ADV $M","Days to exit","Kill criteria","All tags"],
        [(r["ticker"], r["name"], r["sector"], r["price"], r["mcap_m"], fmt(r["er_12m"]),
          r["best_tag"], fmt(r["best_tag_excess"]), r["worst_tag"], fmt(r["worst_tag_excess"]),
          "YES" if r["cluster_live"] else "", r["adv_usd_m"], r["days_to_exit_1pct_adv10"],
          r["kill_criteria"], r["factor_tags"]) for r in rows],
        widths=[8,28,30,8,10,9,22,9,22,9,11,9,11,48,32])

    # 9_INTACT: high conviction × stock at or below fund entry price (the answer
    # to "find large conviction where price hasn't moved from entry")
    rows = list(conn.execute("""SELECT ticker, current_px, anchor_px, anchor_source,
        vs_entry_pct, bucket, conviction_score, n_funds, n_hyper, has_insider_cobuy,
        sum_dollar_m, anchors_seen
        FROM ticker_entry_intact
        WHERE bucket IN ('NEAR_ENTRY','BELOW_ENTRY')
        ORDER BY conviction_score DESC, vs_entry_pct ASC"""))
    write_sheet(wb, "9_INTACT_ENTRY",
        ["Ticker","Current","Entry anchor","Anchor source","vs entry %","Bucket",
         "Conv score","# funds","# hyper","Insider co-buy","Sum $M","Anchors seen"],
        [(r["ticker"], r["current_px"], r["anchor_px"], r["anchor_source"],
          f"{r['vs_entry_pct']:+.1f}%", r["bucket"], r["conviction_score"],
          r["n_funds"], r["n_hyper"], "YES" if r["has_insider_cobuy"] else "",
          r["sum_dollar_m"], r["anchors_seen"]) for r in rows],
        widths=[8,9,12,14,11,16,11,9,9,13,9,60])

    # S0. Mega Style Overview — 445 funds collapsed to 13 macro investing styles
    rows = list(conn.execute("""SELECT macro_style, n_funds, total_rows, n_conviction,
        n_threshold, n_new, n_adds, top_funds, top_consensus
        FROM style_summary ORDER BY n_funds DESC"""))
    write_sheet(wb, "S0_STYLES_OVERVIEW",
        ["Macro style","# funds","Total rows","Conviction","13D/G","New init","Material adds",
         "Top funds in style","Style top-consensus tickers"],
        [tuple(r) for r in rows], widths=[36,9,11,11,7,9,13,60,75])

    # S1. Style x ticker convergence — which styles agree on which names
    rows = list(conn.execute("""SELECT tsc.ticker, tsc.macro_style, tsc.score,
        tsc.n_funds, tsc.n_hyper, tsc.dollar_m, tc.score AS total_score, tc.n_funds AS total_funds
        FROM ticker_style_conviction tsc JOIN ticker_conviction tc ON tc.ticker = tsc.ticker
        WHERE tc.ticker NOT IN ('AMZN','MSFT','GOOGL','GOOG','NVDA','META','AAPL','TSLA','SPY','QQQ','IWM','IVV','IEF','BABA','TSM','BAC','BRK.B','BRK.A','NFLX','JPM','CRM','JNJ','WMT','H2','SEC','BN','AVGO')
        ORDER BY tc.score DESC, tsc.score DESC LIMIT 200"""))
    write_sheet(wb, "S1_STYLE_CONVERGENCE",
        ["Ticker","Macro style","Style score","# funds in style","# hyper-conv","$M",
         "Total ticker score","Total # funds (all styles)"],
        [tuple(r) for r in rows], widths=[8,36,11,16,12,8,16,18])

    # 9e. Multi-factor conviction score — combines hyper-conviction, 13D vs 13G,
    # NEW init, material add, public letter, follow-on, holding persistence,
    # insider co-buy, and multi-fund peer signals into one ranked score
    mega = ("AMZN","MSFT","GOOGL","GOOG","NVDA","META","AAPL","TSLA","SPY","QQQ","IWM","IVV","IEF","BABA","TSM","BAC","BRK.B","BRK.A","NFLX","JPM","CRM","JNJ","WMT","H2")
    placeholders = ",".join("?"*len(mega))
    rows = list(conn.execute(f"""SELECT * FROM ticker_conviction
        WHERE ticker NOT IN ({placeholders}) ORDER BY score DESC LIMIT 80""", mega))
    write_sheet(wb, "9e_CONVICTION_SCORE",
        ["Ticker","Score (style-wtd)","Raw","# funds","Hyper (>=10%)","Activist 13D","Passive 13G",
         "NEW init","Material add","Public letter","Follow-on","Persist","Insider co-buy","Sum $M",
         "Max % book","Max % co","Styles converging","Top fund signals"],
        [(r["ticker"], r["score"], r["raw_score"], r["n_funds"], r["n_hyper"], r["n_activist_13d"],
          r["n_passive_13g"], r["n_new_init"], r["n_material_add"], r["n_public_letter"],
          r["n_follow_on"], r["n_persist"], "YES" if r["has_insider_cobuy"] else "",
          r["sum_dollar_m"], r["max_pct_book"], r["max_pct_company"],
          r["styles_summary"], r["fund_signals_summary"])
         for r in rows],
        widths=[8,14,6,9,13,13,12,9,13,13,11,9,13,9,11,10,55,55])

    rows = list(conn.execute("""SELECT fund, ticker, signals, score, pct_book, pct_company, dollar_m
        FROM fund_conviction WHERE score >= 6 ORDER BY score DESC LIMIT 200"""))
    write_sheet(wb, "9f_FUND_CONVICTION_DETAIL",
        ["Fund","Ticker","Signals (combined)","Score","% book","% company","$M"],
        [(r["fund"], r["ticker"], r["signals"], r["score"], r["pct_book"],
          r["pct_company"], r["dollar_m"]) for r in rows],
        widths=[36,8,42,7,9,11,9])

    # 9. FUND POSITIONING MONITOR — the original conviction/adds/13D view, typed
    rows = list(conn.execute("""SELECT ticker, funds, dollar_m, max_pct, funds_list
        FROM v_top_material_adds WHERE funds>=2 ORDER BY funds DESC, dollar_m DESC LIMIT 60"""))
    write_sheet(wb, "9_MATERIAL_ADDS",
        ["Ticker","# funds adding","Sum $M","Max % book","Funds (multi-fund consensus)"],
        [(r["ticker"], r["funds"], round(r["dollar_m"] or 0, 1),
          f"{r['max_pct']:.1f}%" if r["max_pct"] else "",
          r["funds_list"]) for r in rows], widths=[8,15,9,11,90])

    rows = list(conn.execute("""SELECT ticker, funds, dollar_m, funds_list
        FROM v_top_new_positions WHERE funds>=2 ORDER BY funds DESC, dollar_m DESC LIMIT 40"""))
    write_sheet(wb, "9a_NEW_POSITIONS",
        ["Ticker","# funds initiating","Sum $M","Funds"],
        [(r["ticker"], r["funds"], round(r["dollar_m"] or 0, 1), r["funds_list"]) for r in rows],
        widths=[8,18,9,90])

    rows = list(conn.execute("""SELECT ticker, funds, max_pct_company, funds_list
        FROM v_top_13d_filings WHERE funds>=2 ORDER BY funds DESC, max_pct_company DESC LIMIT 50"""))
    write_sheet(wb, "9b_13D_THRESHOLD",
        ["Ticker","# filers","Max % of company","Filers"],
        [(r["ticker"], r["funds"],
          f"{r['max_pct_company']:.1f}%" if r["max_pct_company"] else "",
          r["funds_list"]) for r in rows], widths=[8,9,16,90])

    rows = list(conn.execute("""SELECT ticker, funds, dollar_m, max_pct
        FROM v_top_conviction WHERE funds>=3 ORDER BY funds DESC, dollar_m DESC LIMIT 60"""))
    write_sheet(wb, "9c_HIGHEST_CONVICTION",
        ["Ticker","# funds holding","Sum $M","Max % book in any fund"],
        [(r["ticker"], r["funds"], round(r["dollar_m"] or 0, 1),
          f"{r['max_pct']:.1f}%" if r["max_pct"] else "") for r in rows],
        widths=[8,16,12,22])

    rows = list(conn.execute("""SELECT fund, conviction_n, threshold_n, new_pos_n, adds_n, total
        FROM v_fund_activity ORDER BY total DESC LIMIT 80"""))
    write_sheet(wb, "9d_FUND_ACTIVITY",
        ["Fund","Conviction rows","13D/G rows","New positions","Material adds","Total"],
        [tuple(r) for r in rows], widths=[40,16,11,15,15,9])

    # 8a. Archetype status — original 10 archetypes reconciled to empirical re-rank
    rows = list(conn.execute("""SELECT archetype, mapped_factor, base_rate_excess,
        members_total, members_live_t1, members_live_t2, members_demoted,
        members_graduated, members_dead, members_untracked, members_excluded,
        best_member, best_member_er, verdict FROM archetype_status ORDER BY archetype"""))
    pc = lambda v: f"{v*100:+.0f}%" if v is not None else ""
    write_sheet(wb, "8a_ARCHETYPE_STATUS",
        ["Archetype","Mapped factor","Base rate 12m","Total","Live T1","Live T2","Demoted",
         "Graduated (rerated)","Dead","Untracked","Excluded","Best surviving","Best ER","Verdict"],
        [(r["archetype"], r["mapped_factor"], pc(r["base_rate_excess"]),
          r["members_total"], r["members_live_t1"], r["members_live_t2"], r["members_demoted"],
          r["members_graduated"], r["members_dead"], r["members_untracked"], r["members_excluded"],
          r["best_member"], pc(r["best_member_er"]), r["verdict"]) for r in rows],
        widths=[42,22,13,7,8,8,8,18,6,9,9,14,8,22])

    # 8b. Each original-archetype member with its current status + thesis
    rows = list(conn.execute("""SELECT archetype, ticker, status, er, factor_tags, thesis
        FROM archetype_member_status ORDER BY archetype,
        CASE status WHEN 'LIVE-T1' THEN 1 WHEN 'LIVE-T2' THEN 2 WHEN 'LIVE-T3' THEN 3
                    WHEN 'DEMOTED' THEN 4 WHEN 'GRADUATED' THEN 5 WHEN 'DEAD' THEN 6
                    ELSE 7 END, ticker"""))
    write_sheet(wb, "8b_ARCHETYPE_MEMBERS",
        ["Archetype","Ticker","Status","ER 12m","Factor tags","Original thesis (preserved)"],
        [(r["archetype"], r["ticker"], r["status"], pc(r["er"]),
          r["factor_tags"], r["thesis"]) for r in rows],
        widths=[42,8,18,9,32,80])

    # 0. Market-wide discovery (EDGAR daily indices, last 9 business days)
    rows = list(conn.execute("""SELECT ticker, issuer, n_buyers, total_usd_m, avg_px,
                                       last_close, off_52w_high, top_buyer, top_role, asof
                                FROM discovery ORDER BY (n_buyers*2 + total_usd_m) DESC"""))
    pc = lambda v: f"{v*100:+.0f}%" if v is not None else ""
    write_sheet(wb, "0_DISCOVERY_CLUSTERS",
        ["Ticker","Issuer","# buyers","Total $M","Avg px","Last","Off 52w high","Top buyer","Role","Asof"],
        [(r["ticker"], r["issuer"], r["n_buyers"], r["total_usd_m"], r["avg_px"],
          r["last_close"], pc(r["off_52w_high"]), r["top_buyer"], r["top_role"], r["asof"]) for r in rows],
        widths=[8,38,9,9,8,8,12,28,22,11])
    rows = list(conn.execute("""SELECT ticker, subject, filer_hint, filed, last_close, off_52w_high
                                FROM discovery_13d_subjects ORDER BY off_52w_high"""))
    write_sheet(wb, "0b_NEW_13D_FILINGS",
        ["Ticker","Subject company","Filer","Filed","Last","Off 52w high"],
        [(r["ticker"], r["subject"], r["filer_hint"], r["filed"], r["last_close"],
          pc(r["off_52w_high"])) for r in rows],
        widths=[8,42,34,10,8,12])

    # 2. Base rates
    rows = list(conn.execute("SELECT factor, hit_rate, avg_excess_12m, sample_n FROM base_rates ORDER BY avg_excess_12m DESC"))
    write_sheet(wb, "2_BASE_RATES",
        ["Factor (signal bucket)","12mo hit rate","Avg excess vs SPY","Sample n"],
        [(r["factor"], f"{r['hit_rate']*100:.0f}%", f"{r['avg_excess_12m']*100:+.1f}%", r["sample_n"]) for r in rows],
        widths=[28,16,20,10])

    # 3. Backtest detail
    rows = list(conn.execute("""SELECT e.ticker, e.bucket, e.event_date, e.description,
                                       r.entry_date, r.entry_px, r.ret_6m, r.ret_12m, r.ret_18m,
                                       r.spy_12m, r.excess_12m, r.excess_18m
                                FROM backtest_events e JOIN backtest_results r ON r.event_id=e.id
                                ORDER BY e.bucket, e.event_date"""))
    pct = lambda v: f"{v*100:+.1f}%" if v is not None else ""
    write_sheet(wb, "3_BACKTEST_DETAIL",
        ["Ticker","Bucket","Event date","Description","Entry","Px","6m","12m","18m","SPY12m","Excess12m","Excess18m"],
        [(r["ticker"], r["bucket"], r["event_date"], r["description"], r["entry_date"], r["entry_px"],
          pct(r["ret_6m"]), pct(r["ret_12m"]), pct(r["ret_18m"]), pct(r["spy_12m"]),
          pct(r["excess_12m"]), pct(r["excess_18m"])) for r in rows],
        widths=[8,18,11,40,11,9,8,8,8,8,11,11])

    # 4. Master CSV mirror
    rows = list(conn.execute("""SELECT ticker, name, sector, currency, price, price_asof,
                                       mcap_m, shares_out_m, tier, verification_status,
                                       known_issues, source_url FROM candidates ORDER BY tier, ticker"""))
    write_sheet(wb, "4_MASTER_CSV",
        ["Ticker","Name","Sector","Ccy","Price","Asof","Mcap $M","Shares M","Tier","Verification","Issues","Source"],
        [tuple(r) for r in rows], widths=[8,28,30,5,9,12,10,10,14,28,50,40])

    # 5. Catalysts live (days_remaining computed at query)
    rows = list(conn.execute("""SELECT ticker, description, expected_date, effective_status, days_remaining, source_url
                                FROM v_catalysts_live ORDER BY days_remaining IS NULL, days_remaining"""))
    write_sheet(wb, "5_CATALYSTS_LIVE",
        ["Ticker","Catalyst","Expected","Status (live)","Days remaining","Source"],
        [tuple(r) for r in rows], widths=[8,60,14,16,14,40])

    # 6. Form 4 verified buys (code=P, acquired=1)
    rows = list(conn.execute("""SELECT ticker, owner, role, trans_date, code, shares, price,
                                       ROUND(shares*price/1e6, 2) AS usd_m, source_url
                                FROM form4_transactions
                                WHERE code='P' AND acquired=1
                                ORDER BY trans_date DESC"""))
    write_sheet(wb, "6_FORM4_BUYS",
        ["Ticker","Owner","Role","Date","Code","Shares","Price","$M","Source"],
        [tuple(r) for r in rows], widths=[8,28,28,11,5,10,9,7,55])

    # 6b. Live insider clusters (the +109% bucket signal, live)
    rows = list(conn.execute("""SELECT ticker, trigger, window_start, window_end,
                                       n_insiders, total_usd_m, avg_price,
                                       top_buyer, top_buyer_usd_m
                                FROM insider_clusters
                                ORDER BY (n_insiders*2 + total_usd_m) DESC"""))
    write_sheet(wb, "6b_LIVE_CLUSTERS",
        ["Ticker","Trigger","Window start","Window end","# insiders","Total $M",
         "Avg px","Top buyer","Top $M"],
        [tuple(r) for r in rows], widths=[8,12,12,12,11,9,8,30,9])

    # 7. Liquidity / sizing budget
    rows = list(conn.execute("""SELECT c.ticker, c.mcap_m, l.adv_shares, l.adv_usd_m,
                                       l.days_to_exit_1pct_adv10, l.asof
                                FROM candidates c JOIN liquidity l ON l.ticker=c.ticker
                                ORDER BY l.days_to_exit_1pct_adv10"""))
    write_sheet(wb, "7_LIQUIDITY",
        ["Ticker","Mcap $M","ADV shares","ADV $M","Days to exit 1% pos","Asof"],
        [tuple(r) for r in rows], widths=[8,10,12,9,18,12])

    # 8. Original archetype theses preserved
    rows = list(conn.execute("""SELECT archetype, ticker, thesis, valuation, catalyst, variant, smart_money
                                FROM archetype_members ORDER BY archetype, ticker"""))
    write_sheet(wb, "8_ARCHETYPES",
        ["Archetype","Ticker","Thesis","Valuation","Catalyst","Variant","Smart money"],
        [tuple(r) for r in rows], widths=[40,8,60,40,40,40,40])

    wb.save(OUT)
    print(f"wrote {OUT}")
    for s in wb.sheetnames: print(f"  - {s}")

if __name__ == "__main__":
    run()
