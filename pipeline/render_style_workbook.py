"""Style-subcategory workbook — monochrome academic aesthetic.

Companion to universe_analysis.xlsx. Where that book ranks the FULL universe
flat, this slices the picks BY style of fund originating the signal.
Visual grammar shared via _style_bw.py.
"""
import os, sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _style_bw import (
    write_title, write_section_heading, write_table_header, write_table_rows,
    autosize, write_legend_sheet, add_contents_index, set_print_layout,
    NUMFMT_USD, NUMFMT_PCT, NUMFMT_NUM, NUMFMT_INT, NUMFMT_USD2,
    NUMFMT_MCAP, NUMFMT_M_TO_B,
    BODY_FONT, BODY_ITALIC, SECTION_FONT, MONO_FONT, TICKER_FONT,
    TNR, SIZE_BODY,
)

DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "cyclepapa.db")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "style_analysis.xlsx")

ETFs = {"SPY","QQQ","VOO","IWM","IEF","IEFA","EFA","EEM","BIL","IVV","XBI","HYG",
        "GLD","SLV","TLT","XLE","XLF","XLK","XLY","XLP","XLU","XLI","XLV","XLB","XLRE",
        "ARKK","JNK","LQD","TIP","AGG","BND","VEA","VWO","SHY"}

# Excel: ≤31 chars, no /\?*[]
def safe_sheet_name(name):
    s = name.replace("/", "-").replace("\\", "-").replace("?","").replace("*","")
    s = s.replace("[","(").replace("]",")").replace(":", "-")
    return s[:31]

def style_macro_list(conn):
    return list(conn.execute("""SELECT macro_style, COUNT(*) FROM fund_style
        WHERE macro_style IS NOT NULL
        GROUP BY macro_style ORDER BY 2 DESC"""))

def sheet_readme(wb, conn):
    ws = wb.create_sheet("README", 0)
    write_title(ws,
        "Cyclepapa — Style Analysis",
        "Top picks within each fund style and size bucket. 445 funds in 15 macro styles.",
        1)
    ws.column_dimensions["A"].width = 92
    styles = style_macro_list(conn)
    rows = [
        ("",),
        ("Organisation",),
        ("Overview              — universe-wide style summary, 10 names per style.",),
        ("Sub-Group Tiers       — top picks per sub_group (e.g. US Activist Tier 1).",),
        ("Per-style sheets ×15  — one per macro_style:",),
        ("                        top picks within style, S3 new initiations,",),
        ("                        S4 material adds, concentration leaders,",),
        ("                        by-size cross-cut (nano → large).",),
        ("",),
        ("Macro styles",),
    ]
    for s, c_count in styles:
        rows.append((f"{s:<40} {c_count:>4} funds",))
    rows += [
        ("",),
        ("Methodology",),
        ("All measures computed from primary EDGAR data. No memory-based picks.",),
        ("",),
        ("score = log(n_funds) × 2",),
        ("      + 3.0 × s3_new + 1.5 × s4_add + 2.0 × s1_top",),
        ("      + 0.5 × activist_max_pct",),
        ("      + 0.6 × max(pct_book)              position concentration",),
        ("      + 1.5 × n_funds_5pct_book          concentration cluster",),
        ("      + cluster_step(n_insiders)",),
        ("      + log(form4_buys + 1) × 2",),
        ("      - log(form4_sells + 1) × 1.5      counter-signal",),
        ("      + micro_bonus",),
    ]
    for i, r in enumerate(rows, 4):
        c = ws.cell(row=i, column=1, value=r[0])
        ws.row_dimensions[i].height = 18
        if not r[0].strip(): continue
        if r[0] in ("Organisation","Macro styles","Methodology"):
            c.font = SECTION_FONT
        elif r[0].startswith("score") or r[0].startswith("      "):
            c.font = MONO_FONT
        else:
            c.font = BODY_FONT

def write_style_sheet(wb, conn, macro_style, sheet_name):
    ws = wb.create_sheet(sheet_name)
    style_funds = [r[0] for r in conn.execute(
        "SELECT fund FROM fund_style WHERE macro_style = ?", (macro_style,))]
    write_title(ws, macro_style,
                f"Top picks held by funds in this style ({len(style_funds)} funds).",
                17)
    row = 4
    write_section_heading(ws, row, "Top picks — most held within this style", 17)
    row += 1
    hdr = ["Ticker","St Holders","pB Max","pB ≥5%","S3","S4","S1","Mcap","Bucket","EV/EBITDA","P/B","Act %","Clu $M","F4 Buy","Name","Industry","Business"]
    write_table_header(ws, row, hdr)
    row += 1
    ph = ",".join("?" * len(style_funds))
    sql = f"""
        SELECT h.ticker,
               COUNT(DISTINCT h.fund) AS holders,
               MAX(h.pct_book) AS max_pb,
               SUM(CASE WHEN h.pct_book >= 5 THEN 1 ELSE 0 END) AS n5,
               (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                 WHERE fp.ticker = h.ticker AND fp.section = 3
                   AND fp.fund IN ({ph})) AS s3_st,
               (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                 WHERE fp.ticker = h.ticker AND fp.section = 4
                   AND fp.fund IN ({ph})) AS s4_st,
               (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                 WHERE fp.ticker = h.ticker AND fp.section = 1
                   AND fp.fund IN ({ph})) AS s1_st,
               us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio,
               us.activist_max_pct, us.insider_cluster_dollars_m, us.form4_buy_usd_m,
               tm.name
        FROM fund_13f_holdings h
        LEFT JOIN unified_signal us ON us.ticker = h.ticker
        LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
        WHERE h.fund IN ({ph}) AND h.ticker IS NOT NULL
        GROUP BY h.ticker
        ORDER BY holders DESC, max_pb DESC
        LIMIT 60
    """
    rows = list(conn.execute(sql, style_funds * 4))
    out = []
    for r in rows:
        if r[0] in ETFs: continue
        out.append([r[0], r[1], round(r[2] or 0, 1), r[3] or 0,
                    r[4] or 0, r[5] or 0, r[6] or 0,
                    r[7] or "", r[8] or "",
                    round(r[9], 1) if r[9] is not None else "",
                    round(r[10], 2) if r[10] is not None else "",
                    round(r[11] or 0, 1),
                    round(r[12] or 0, 1) if r[12] else "",
                    round(r[13] or 0, 1) if r[13] else "",
                    (r[14] or "")[:38], *desc_for(conn, r[0])])
        if len(out) >= 30: break
    write_table_rows(ws, out, row)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=8).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=10).number_format = '0.0"x"'
        ws.cell(row=ridx, column=11).number_format = '0.00"x"'
        ws.cell(row=ridx, column=12).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=13).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=14).number_format = NUMFMT_M_TO_B
    row += len(out) + 2

    # New initiations
    write_section_heading(ws, row, "New major positions — section 3", 13)
    row += 1
    hdr2 = ["Ticker","St S3 #","Uni 13F","pB Max","Mcap","Bucket","EV/EBITDA","P/B","Act %","Name"]
    write_table_header(ws, row, hdr2)
    row += 1
    rows = list(conn.execute(f"""
        SELECT fp.ticker, COUNT(DISTINCT fp.fund) AS n,
               us.smart_money_n, us.max_pct_book, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio,
               us.activist_max_pct, tm.name
        FROM fund_positions fp
        LEFT JOIN unified_signal us ON us.ticker = fp.ticker
        LEFT JOIN ticker_meta tm ON tm.ticker = fp.ticker
        WHERE fp.fund IN ({ph}) AND fp.section = 3 AND fp.ticker IS NOT NULL
        GROUP BY fp.ticker ORDER BY n DESC LIMIT 20""", style_funds))
    out = [[r[0], r[1], r[2] or 0, round(r[3] or 0, 1),
            r[4] or "", r[5] or "",
            round(r[6], 1) if r[6] is not None else "",
            round(r[7], 2) if r[7] is not None else "",
            round(r[8] or 0, 1),
            (r[9] or "")[:38]] for r in rows if r[0] not in ETFs]
    write_table_rows(ws, out, row)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=7).number_format = '0.0"x"'
        ws.cell(row=ridx, column=8).number_format = '0.00"x"'
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
    row += len(out) + 2

    # Material adds
    write_section_heading(ws, row, "Material adds — section 4", 13)
    row += 1
    write_table_header(ws, row, hdr2)
    row += 1
    rows = list(conn.execute(f"""
        SELECT fp.ticker, COUNT(DISTINCT fp.fund) AS n,
               us.smart_money_n, us.max_pct_book, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio,
               us.activist_max_pct, tm.name
        FROM fund_positions fp
        LEFT JOIN unified_signal us ON us.ticker = fp.ticker
        LEFT JOIN ticker_meta tm ON tm.ticker = fp.ticker
        WHERE fp.fund IN ({ph}) AND fp.section = 4 AND fp.ticker IS NOT NULL
        GROUP BY fp.ticker ORDER BY n DESC LIMIT 20""", style_funds))
    out = [[r[0], r[1], r[2] or 0, round(r[3] or 0, 1),
            r[4] or "", r[5] or "",
            round(r[6], 1) if r[6] is not None else "",
            round(r[7], 2) if r[7] is not None else "",
            round(r[8] or 0, 1),
            (r[9] or "")[:38]] for r in rows if r[0] not in ETFs]
    write_table_rows(ws, out, row)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=7).number_format = '0.0"x"'
        ws.cell(row=ridx, column=8).number_format = '0.00"x"'
        ws.cell(row=ridx, column=9).number_format = NUMFMT_PCT
    row += len(out) + 2

    # Concentration leaders
    write_section_heading(ws, row, "Concentration leaders — single-fund pct_book ≥5%", 13)
    row += 1
    hdr3 = ["Ticker","Fund","%Book","Uni 13F","Mcap","Bucket","EV/EBITDA","P/B","Name"]
    write_table_header(ws, row, hdr3)
    row += 1
    rows = list(conn.execute(f"""
        SELECT h.ticker, h.fund, h.pct_book, us.smart_money_n, us.mcap_m, us.mcap_bucket,
               us.ev_ebitda, us.pb_ratio, tm.name
        FROM fund_13f_holdings h
        LEFT JOIN unified_signal us ON us.ticker = h.ticker
        LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
        WHERE h.fund IN ({ph}) AND h.pct_book >= 5 AND h.pct_book <= 100
        ORDER BY h.pct_book DESC LIMIT 20""", style_funds))
    out = [[r[0], r[1][:35], round(r[2] or 0, 1), r[3] or 0,
            r[4] or "", r[5] or "",
            round(r[6], 1) if r[6] is not None else "",
            round(r[7], 2) if r[7] is not None else "",
            (r[8] or "")[:32]]
           for r in rows if r[0] not in ETFs]
    write_table_rows(ws, out, row)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=7).number_format = '0.0"x"'
        ws.cell(row=ridx, column=8).number_format = '0.00"x"'
    row += len(out) + 2

    # Size cross-cut
    write_section_heading(ws, row, "By size bucket — top within style by mcap class", 13)
    row += 1
    hdr4 = ["Bucket","Ticker","Holders","pB Max","Mcap","S3","S4","Act %","EV/EBITDA","P/B","Name"]
    write_table_header(ws, row, hdr4)
    row += 1
    out = []
    for bucket in ["nano","micro","small","mid","large"]:
        rows = list(conn.execute(f"""
            SELECT h.ticker, COUNT(DISTINCT h.fund) AS holders,
                   MAX(h.pct_book) AS max_pb,
                   us.mcap_m,
                   (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                    WHERE fp.ticker=h.ticker AND fp.section=3 AND fp.fund IN ({ph})),
                   (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                    WHERE fp.ticker=h.ticker AND fp.section=4 AND fp.fund IN ({ph})),
                   us.activist_max_pct, us.ev_ebitda, us.pb_ratio, tm.name
            FROM fund_13f_holdings h
            JOIN unified_signal us ON us.ticker = h.ticker
            LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
            WHERE h.fund IN ({ph}) AND us.mcap_bucket = ?
              AND h.ticker NOT IN ({','.join('?'*len(ETFs))})
            GROUP BY h.ticker
            ORDER BY (holders * 2 + COALESCE(max_pb,0) * 0.5) DESC LIMIT 5""",
            style_funds + style_funds + style_funds + [bucket] + list(ETFs)))
        for r in rows:
            out.append([bucket, r[0], r[1], round(r[2] or 0, 1),
                        r[3] or "", r[4] or 0, r[5] or 0,
                        round(r[6] or 0, 1),
                        round(r[7], 1) if r[7] is not None else "",
                        round(r[8], 2) if r[8] is not None else "",
                        (r[9] or "")[:32]])
    write_table_rows(ws, out, row, ticker_col=2)
    for ridx in range(row, row + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
        ws.cell(row=ridx, column=8).number_format = NUMFMT_PCT
        ws.cell(row=ridx, column=9).number_format = '0.0"x"'
        ws.cell(row=ridx, column=10).number_format = '0.00"x"'
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions[get_column_letter(16)].width = 24   # Industry (Top picks)
    ws.column_dimensions[get_column_letter(17)].width = 80   # Business (Top picks)

def sheet_overview(wb, conn):
    ws = wb.create_sheet("Overview")
    write_title(ws, "Universe Overview by Style",
                "For each macro_style, the top 10 most-held names within that style.", 15)
    row = 4
    for ms, count in style_macro_list(conn):
        write_section_heading(ws, row, f"{ms} — {count} funds", 15)
        row += 1
        hdr = ["Ticker","St Holders","pB Max","pB ≥5%","S3","S4","Mcap","Bucket","EV/EBITDA","P/B","Act %","Clu $M","F4 Buy","Sell","Name"]
        write_table_header(ws, row, hdr)
        row += 1
        style_funds = [r[0] for r in conn.execute(
            "SELECT fund FROM fund_style WHERE macro_style = ?", (ms,))]
        if not style_funds: continue
        ph = ",".join("?" * len(style_funds))
        rows = list(conn.execute(f"""
            SELECT h.ticker, COUNT(DISTINCT h.fund) AS holders,
                   MAX(h.pct_book), SUM(CASE WHEN h.pct_book>=5 THEN 1 ELSE 0 END),
                   (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                    WHERE fp.ticker = h.ticker AND fp.section = 3 AND fp.fund IN ({ph})),
                   (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                    WHERE fp.ticker = h.ticker AND fp.section = 4 AND fp.fund IN ({ph})),
                   us.mcap_m, us.mcap_bucket,
                   us.ev_ebitda, us.pb_ratio, us.activist_max_pct,
                   us.insider_cluster_dollars_m, us.form4_buy_usd_m, us.form4_sell_usd_m,
                   tm.name
            FROM fund_13f_holdings h
            LEFT JOIN unified_signal us ON us.ticker = h.ticker
            LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
            WHERE h.fund IN ({ph}) AND h.ticker IS NOT NULL
            GROUP BY h.ticker
            ORDER BY holders DESC LIMIT 12""", style_funds * 3))
        out = []
        for r in rows:
            if r[0] in ETFs: continue
            out.append([r[0], r[1], round(r[2] or 0, 1), r[3] or 0, r[4] or 0, r[5] or 0,
                        r[6] or "", r[7] or "",
                        round(r[8], 1) if r[8] is not None else "",
                        round(r[9], 2) if r[9] is not None else "",
                        round(r[10] or 0, 1),
                        round(r[11] or 0, 1) if r[11] else "",
                        round(r[12] or 0, 1) if r[12] else "",
                        round(r[13] or 0, 1) if r[13] else "",
                        (r[14] or "")[:30]])
            if len(out) >= 10: break
        write_table_rows(ws, out, row)
        for ridx in range(row, row + len(out)):
            ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
            ws.cell(row=ridx, column=7).number_format = NUMFMT_MCAP
            ws.cell(row=ridx, column=9).number_format = '0.0"x"'
            ws.cell(row=ridx, column=10).number_format = '0.00"x"'
            ws.cell(row=ridx, column=11).number_format = NUMFMT_PCT
            ws.cell(row=ridx, column=12).number_format = NUMFMT_M_TO_B
            ws.cell(row=ridx, column=13).number_format = NUMFMT_M_TO_B
            ws.cell(row=ridx, column=14).number_format = NUMFMT_M_TO_B
        row += len(out) + 2
    autosize(ws)
    ws.column_dimensions["A"].width = 8

def sheet_subgroup_focus(wb, conn):
    """Sub_group tier picks. Multi-fund sub_groups shown explicitly;
    single-fund specialists consolidated into a 'Specialist Funds' bucket
    per macro_style so every fund is visible without 90 separate sections."""
    multi = list(conn.execute("""SELECT macro_style, sub_group, COUNT(*) c
        FROM fund_style WHERE sub_group IS NOT NULL AND macro_style IS NOT NULL
        GROUP BY macro_style, sub_group
        HAVING c >= 2
        ORDER BY macro_style, c DESC, sub_group"""))
    # All singletons grouped per macro_style
    singletons = {}
    for r in conn.execute("""SELECT macro_style, sub_group, fund FROM fund_style
        WHERE sub_group IS NOT NULL AND macro_style IS NOT NULL
        AND sub_group IN (SELECT sub_group FROM fund_style GROUP BY sub_group HAVING COUNT(*) = 1)
        ORDER BY macro_style, sub_group"""):
        singletons.setdefault(r[0], []).append((r[1], r[2]))

    ws = wb.create_sheet("Sub-Group Tiers")
    write_title(ws, "Sub-Group Tier Picks",
                "Multi-fund tiers shown explicitly. Single-fund specialists consolidated per macro_style — every fund is visible.", 13)
    # Group multi-fund subgroups by macro_style so we can render the
    # singletons-roll-up right after each macro's multi-fund tiers.
    multi_by_macro = {}
    for ms, sg, n in multi:
        multi_by_macro.setdefault(ms, []).append((sg, n))
    all_macros = sorted(set(list(multi_by_macro) + list(singletons)))

    row = 4
    for ms in all_macros:
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        c = ws.cell(row=row, column=1, value=ms.upper())
        c.font = Font(name=TNR, bold=True, size=SIZE_BODY + 1, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="bottom")
        row += 1
        # Multi-fund sub_groups for this macro_style — EACH gets its own
        # heading + tier table (previously only the last one rendered).
        hdr = ["Ticker","Sub Holders","pB Max","Mcap","Bucket","S3","S4","Act %","Clu $M","F4 $M","EV/EBITDA","P/B","Name"]
        for sg, n in multi_by_macro.get(ms, []):
            write_section_heading(ws, row, f"  {sg}  ({n} funds)", 13)
            row += 1
            sg_funds = [r[0] for r in conn.execute(
                "SELECT fund FROM fund_style WHERE sub_group = ? ORDER BY fund", (sg,))]
            # name every member fund so the tier is self-documenting
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            fc = ws.cell(row=row, column=1, value="     " + "  ·  ".join(sg_funds))
            fc.font = BODY_ITALIC
            fc.alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            write_table_header(ws, row, hdr)
            row += 1
            if not sg_funds:
                row += 1
                continue
            ph = ",".join("?" * len(sg_funds))
            rows = list(conn.execute(f"""
                SELECT h.ticker, COUNT(DISTINCT h.fund) holders, MAX(h.pct_book) max_pb,
                       us.mcap_m, us.mcap_bucket,
                       (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                        WHERE fp.ticker=h.ticker AND fp.section=3 AND fp.fund IN ({ph})),
                       (SELECT COUNT(DISTINCT fp.fund) FROM fund_positions fp
                        WHERE fp.ticker=h.ticker AND fp.section=4 AND fp.fund IN ({ph})),
                       us.activist_max_pct, us.insider_cluster_dollars_m, us.form4_buy_usd_m,
                       us.ev_ebitda, us.pb_ratio,
                       tm.name
                FROM fund_13f_holdings h
                LEFT JOIN unified_signal us ON us.ticker = h.ticker
                LEFT JOIN ticker_meta tm ON tm.ticker = h.ticker
                WHERE h.fund IN ({ph}) AND h.ticker IS NOT NULL
                GROUP BY h.ticker
                ORDER BY (holders * 2 + (COALESCE(max_pb,0)) * 0.5) DESC LIMIT 10
            """, sg_funds * 3))
            out = []
            for r in rows:
                if r[0] in ETFs: continue
                out.append([r[0], r[1], round(r[2] or 0, 1),
                            r[3] or "", r[4] or "", r[5] or 0, r[6] or 0,
                            round(r[7] or 0, 1),
                            round(r[8] or 0, 1) if r[8] else "",
                            round(r[9] or 0, 1) if r[9] else "",
                            round(r[10], 1) if r[10] is not None else "",
                            round(r[11], 2) if r[11] is not None else "",
                            (r[12] or "")[:32]])
                if len(out) >= 8: break
            write_table_rows(ws, out, row)
            for ridx in range(row, row + len(out)):
                ws.cell(row=ridx, column=3).number_format = NUMFMT_PCT
                ws.cell(row=ridx, column=4).number_format = NUMFMT_MCAP
                ws.cell(row=ridx, column=8).number_format = NUMFMT_PCT
                ws.cell(row=ridx, column=9).number_format = NUMFMT_M_TO_B
                ws.cell(row=ridx, column=10).number_format = NUMFMT_M_TO_B
                ws.cell(row=ridx, column=11).number_format = '0.0"x"'
                ws.cell(row=ridx, column=12).number_format = '0.00"x"'
            row += len(out) + 2
        # ── end inner sub_group loop ──

        # SPECIALIST FUNDS — consolidate all singletons in this macro_style
        macro_singletons = singletons.get(ms, [])
        if macro_singletons:
            write_section_heading(ws, row,
                f"  Specialist Funds — {len(macro_singletons)} single-fund sub_groups consolidated",
                11)
            row += 1
            spec_hdr = ["Fund","Sub-Group","Top Holding","%Book","13F #","Mcap","Bucket"]
            write_table_header(ws, row, spec_hdr)
            row += 1
            spec_out = []
            for sg, fund in macro_singletons:
                # Top 1 holding for this fund (by value)
                top = conn.execute("""SELECT h.ticker, h.pct_book,
                       us.smart_money_n, us.mcap_m, us.mcap_bucket
                    FROM fund_13f_holdings h
                    LEFT JOIN unified_signal us ON us.ticker = h.ticker
                    WHERE h.fund = ? AND h.ticker IS NOT NULL
                    ORDER BY h.value_k DESC LIMIT 1""", (fund,)).fetchone()
                if top:
                    spec_out.append([fund[:45], sg[:38], top[0],
                                     round(top[1] or 0, 2),
                                     top[2] or 0,
                                     top[3] or "",
                                     top[4] or ""])
                else:
                    spec_out.append([fund[:45], sg[:38], "—", 0, 0, "", ""])
            write_table_rows(ws, spec_out, row)
            for ridx in range(row, row + len(spec_out)):
                ws.cell(row=ridx, column=4).number_format = NUMFMT_PCT
                ws.cell(row=ridx, column=6).number_format = NUMFMT_MCAP
            row += len(spec_out) + 2
    autosize(ws)
    ws.column_dimensions["A"].width = 30

# Grayscale tab palette — content stays B&W; tab tint just aids navigation.
# 15 macro_styles mapped to 8 distinct greyscale tones, grouped by character.
TAB_COLORS = {
    # Activists / Special Sits / Distressed — darkest (high-action)
    "Activists / Special Situations":      "262626",
    "Distressed / Event-Driven":           "404040",
    # Value / Quality — mid-dark
    "Value / Concentrated Quality":        "595959",
    "Foreign / EM Value":                  "595959",
    "Small-cap / Multibagger Specialists": "595959",
    "Microcap-Tactical":                   "595959",
    # Tiger Cubs / Family Offices — mid
    "Tiger Cubs / L/S Legends":            "808080",
    "Family Offices / Individual Filers":  "808080",
    # Mega multi-strat / Macro — mid-light
    "Mega Multi-Strats / Quants":          "A6A6A6",
    "Macro / Trend":                       "A6A6A6",
    "CTA / Trend Followers":               "A6A6A6",
    # Specialist — light
    "Biotech Specialists":                 "BFBFBF",
    "Warrant Specialists":                 "BFBFBF",
    "PE / SPAC / Gold / Mining":           "BFBFBF",
    "Other / Unclassified":                "D9D9D9",
}

def sheet_fund_roster(wb, conn):
    """Definitive roster — EVERY fund with its macro_style + sub_group + data
    counts, grouped by style then sub-group. Guarantees no fund is missed."""
    ws = wb.create_sheet("Fund Roster")
    write_title(ws, "Fund Roster — all funds by style & sub-group",
                "Every fund in the universe with its macro_style, sub_group, and data coverage. Grouped style → sub-group.", 8)
    row = 4
    macros = [r[0] for r in conn.execute(
        "SELECT DISTINCT macro_style FROM fund_style WHERE macro_style IS NOT NULL ORDER BY macro_style")]
    total_funds = 0
    for ms in macros:
        n_in_macro = conn.execute("SELECT COUNT(*) FROM fund_style WHERE macro_style=?", (ms,)).fetchone()[0]
        ws.row_dimensions[row].height = 22
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f"{ms.upper()}  ({n_in_macro} funds)")
        c.font = Font(name=TNR, bold=True, size=SIZE_BODY + 1, color="000000")
        c.alignment = Alignment(horizontal="left", vertical="bottom")
        row += 1
        hdr = ["Fund","Sub-Group","13F #","13F $M","13D #","Foreign Pos","Total Pos","Status"]
        write_table_header(ws, row, hdr)
        row += 1
        # all funds in this macro_style, ordered by sub_group then fund
        funds = list(conn.execute("""
            SELECT fm.fund, fs.sub_group,
                   COALESCE(st.n_holdings,0),
                   COALESCE(st.total_value_k,0)/1e3,
                   (SELECT COUNT(*) FROM holder_13d h WHERE h.holder=fm.fund),
                   (SELECT COUNT(*) FROM fund_positions fp WHERE fp.fund=fm.fund AND fp.ticker LIKE '%.%'),
                   (SELECT COUNT(*) FROM fund_positions fp WHERE fp.fund=fm.fund),
                   fr.status
            FROM fund_meta fm
            JOIN fund_style fs ON fs.fund=fm.fund
            LEFT JOIN fund_13f_state st ON st.fund=fm.fund
            LEFT JOIN fund_resolution_state fr ON fr.fund=fm.fund
            WHERE fs.macro_style=?
            ORDER BY fs.sub_group, fm.fund""", (ms,)))
        out = []
        for f in funds:
            out.append([f[0][:46], (f[1] or "")[:40], f[2], round(f[3] or 0),
                        f[4], f[5], f[6], (f[7] or "")[:22]])
        write_table_rows(ws, out, row)
        for ridx in range(row, row+len(out)):
            ws.cell(row=ridx, column=4).number_format = NUMFMT_M_TO_B
        total_funds += len(out)
        row += len(out) + 2
    # footer tally
    ws.cell(row=row, column=1, value=f"TOTAL: {total_funds} funds across {len(macros)} macro-styles").font = SECTION_FONT
    ws.freeze_panes = "A4"
    autosize(ws)

def _one_liner(s, limit=200):
    if not s:
        return ""
    s = str(s).strip().replace("\n", " ")
    dot = s.find(". ")
    if 0 < dot <= limit:
        return s[:dot + 1]
    return s if len(s) <= limit else s[:limit].rstrip() + "…"

_DESC_CACHE = None
def desc_for(conn, ticker):
    """(industry, one-line business summary) for a ticker — memoized."""
    global _DESC_CACHE
    if _DESC_CACHE is None:
        _DESC_CACHE = {}
        for r in conn.execute("""
            SELECT us.ticker,
                   COALESCE(yf.industry, tm.industry, tm.sic_description),
                   yf.business_summary
            FROM unified_signal us
            LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
            LEFT JOIN ticker_yf  yf ON yf.ticker = us.ticker"""):
            _DESC_CACHE[r[0]] = ((r[1] or "")[:26], _one_liner(r[2], 90))
    return _DESC_CACHE.get(ticker, ("", ""))

def sheet_ticker_reference(wb, conn):
    """Glossary: every ticker with name, sector, industry, and a short business
    summary — so any symbol in the style book can be looked up. Sorted A–Z."""
    ws = wb.create_sheet("Ticker Reference")
    write_title(ws, "Ticker Reference — name, industry, business",
                "Every symbol in the universe with its company name, sector, industry, market cap, and a one-line description of what it does. Sourced from Yahoo Finance + SEC. Sorted A–Z.", 6)
    hdr = ["Ticker", "Name", "Sector", "Industry", "Mcap", "Business Summary"]
    write_table_header(ws, 4, hdr)
    rows = list(conn.execute("""
        SELECT us.ticker,
               COALESCE(tm.name, yf.long_name)                       AS name,
               COALESCE(yf.sector, tm.sector)                        AS sector,
               COALESCE(yf.industry, tm.industry, tm.sic_description) AS industry,
               us.mcap_m,
               yf.business_summary
        FROM unified_signal us
        LEFT JOIN ticker_meta tm ON tm.ticker = us.ticker
        LEFT JOIN ticker_yf  yf ON yf.ticker = us.ticker
        WHERE COALESCE(tm.name, yf.long_name, yf.business_summary, yf.sector,
                       tm.sic_description) IS NOT NULL
        ORDER BY us.ticker"""))
    out = []
    for r in rows:
        if r[0] in ETFs: continue
        out.append([r[0], (r[1] or "")[:46], (r[2] or "")[:22],
                    (r[3] or "")[:30], r[4] or "", _one_liner(r[5])])
    write_table_rows(ws, out, 5)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=5).number_format = NUMFMT_MCAP
    ws.freeze_panes = "B5"
    autosize(ws)
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 100

def main():
    conn = sqlite3.connect(DB)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sheet_readme(wb, conn)
    write_legend_sheet(wb, 1)
    sheet_fund_roster(wb, conn)
    sheet_overview(wb, conn)
    sheet_subgroup_focus(wb, conn)
    for ms, _ in style_macro_list(conn):
        sn = safe_sheet_name(ms)
        write_style_sheet(wb, conn, ms, sn)
        ws = wb[sn]
        # Tab color — grayscale per macro_style
        if ms in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ms]

    sheet_ticker_reference(wb, conn)

    # AutoFilter on the single-table Ticker Reference (header at row 4)
    if "Ticker Reference" in wb.sheetnames:
        ws = wb["Ticker Reference"]
        if ws.max_row > 4:
            ws.auto_filter.ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"

    # README + meta tabs in lightest grey for distinction
    for nav in ("README", "Legend", "Fund Roster", "Overview", "Sub-Group Tiers", "Ticker Reference"):
        if nav in wb.sheetnames:
            wb[nav].sheet_properties.tabColor = "F2F2F2"

    add_contents_index(wb["README"], wb.sheetnames)
    set_print_layout(wb)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
