"""People & Board Tea-Leaves monitor — a SEPARATE workbook from the universe/style
books. The signal: where network-connected influential people are TAKING BOARD
SEATS. A director/advisor being placed onto a company's board is a forward tea
leaf — it reveals where a smart-money network (a titan, a Sequoia operator, a
sovereign/family-office) is quietly positioning, often before it shows in price
or 13F filings.

Sourced from five PitchBook people-searches, each filtered to board/advisory
positions by design. We classify each seat as an OUTSIDE-DIRECTOR placement (an
investor/operator joining a board they don't run — the signal) vs an OPERATOR
(founder/CEO on their own board — not a signal), map the company to a ticker, and
overlay our own smart-money score. Concentration (several connected people on one
board) and cross-network convergence are the strongest reads.
"""
import os, sqlite3
import openpyxl
from _style_bw import (
    write_title, write_section_heading, write_table_header, write_table_rows,
    autosize, set_default_font, add_contents_index, set_print_layout,
    NUMFMT_MCAP, NUMFMT_M_TO_B,
)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(BASE, "data", "cyclepapa.db")
OUT = os.path.join(BASE, "people_monitor.xlsx")

NEW_FUNDS = ["Cat Rock Capital Management", "Theleme Partners", "Clarkston Capital Partners",
             "Hosking Partners", "Tybourne Capital Management", "Man Group",
             "Mubadala Investment Company"]

# investor-type primary affiliations — a person from one of these taking a board
# seat elsewhere is the classic activism / involvement tea leaf.
INVESTOR_TYPES = ("Asset Manager", "Hedge Fund", "PE/Buyout", "Venture Capital",
                  "Family Office (Multi)", "Family Office (Single)", "Investor",
                  "Growth/Expansion")


def sheet_readme(wb, conn):
    ws = wb.create_sheet("README")
    ws.sheet_view.showGridLines = False
    n_people = conn.execute("SELECT COUNT(DISTINCT full_name) FROM pb_people").fetchone()[0]
    n_seats = conn.execute("SELECT COUNT(*) FROM pb_affiliation WHERE role_class='outside_director' AND is_former=0").fetchone()[0]
    n_uni = conn.execute("""SELECT COUNT(DISTINCT a.ticker) FROM pb_affiliation a
        JOIN unified_signal u ON u.ticker=a.ticker
        WHERE a.role_class='outside_director' AND a.is_former=0 AND u.sec_type='common'""").fetchone()[0]
    write_title(ws, "People & Board Tea-Leaves — Alpha Monitor",
                "Where network-connected influential people are taking board seats — the forward signal, cross-referenced with our smart money.", 4)
    ws.column_dimensions["A"].width = 102
    rows = [
        ("",),
        ("The thesis — reading the tea leaves",),
        ("A board seat is a leading indicator. When a network-connected investor or operator is placed onto a",),
        ("company's board — an activist taking a non-exec seat, a Sequoia operator joining a portfolio company, a",),
        ("sovereign/family-office nominee appearing — it reveals where the smart money is positioning, often before",),
        ("it shows up in price or 13F filings. Concentration (several connected directors on one board) and",),
        ("cross-network convergence are the strongest reads.",),
        ("",),
        (f"Sourced from five PitchBook people-searches ({n_people:,} individuals), each filtered to board/advisory roles.",),
        (f"{n_seats} active OUTSIDE-director seats identified (investor/operator joining a board they don't run);",),
        (f"{n_uni} land on companies inside our tradeable universe and carry a live smart-money score.",),
        ("",),
        ("How to read it",),
        ("• 'Board Signal — In Universe' — the actionable sheet: tradeable companies where connected people hold",),
        ("   active board seats, ranked by our score, with the director names, network theme, and smart-money overlay.",),
        ("• 'Investor Board Seats' — highest signal: people from an investment firm (AM / hedge fund / PE / VC /",),
        ("   family office) taking an OUTSIDE board seat — the classic activism / involvement tell.",),
        ("• 'Board Convergence' — companies where MULTIPLE connected directors, or multiple distinct networks,",),
        ("   converge on one board. Coordinated positioning = the strongest signal.",),
        ("• 'Watchlist — Not in Universe' — the same board concentration on companies we don't yet cover (mostly",),
        ("   foreign — Harbour Energy, Playtech, NOVATEK, En+ ...). Candidates to add.",),
        ("• 'Family Office / SWF Map' — where Gulf, royal, and billionaire-vehicle nominees sit.",),
        ("• 'Network Roster Adds' — investment firms found in the data and now tracked for holdings (context).",),
        ("",),
        ("Signal classification & caveats",),
        ("• OUTSIDE-director = a Board Member / Non-Exec / Advisor / Chairman who is NOT the company's CEO/founder.",),
        ("   Founder/CEO-on-own-board rows are labelled OPERATOR and excluded from the signal sheets.",),
        ("• '(Former)' seats are flagged and excluded — only live placements count.",),
        ("• A person is 'network-connected' because their PitchBook biography ties them to a tracked titan / Sequoia /",),
        ("   family-office; 'principal' rows are the tracked individual themselves.",),
        ("• One of the five source files (2025-09-19) arrived truncated and unrecoverable; a same-author 2025-09-20",),
        ("   search is present. Board data is a point-in-time snapshot, not a live feed.",),
    ]
    for i, r in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=r[0])
    return ws


def sheet_board_signal(wb, conn):
    ws = wb.create_sheet("Board Signal — In Universe")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Board Signal — connected directors on tradeable companies",
                "Companies where network-connected people hold an ACTIVE outside board seat. Ranked by our score. The forward tea leaf + our smart money on one line.", 8)
    hdr = ["Ticker", "Company", "Our Score", "Smart$ n", "# Dir", "Connected Directors", "Network", "Sector"]
    write_table_header(ws, 4, hdr)
    rows = conn.execute("""
        SELECT a.ticker, u.name, u.score, u.smart_money_n,
               COUNT(DISTINCT a.full_name) nd,
               GROUP_CONCAT(DISTINCT a.full_name),
               GROUP_CONCAT(DISTINCT a.theme), u.sector
        FROM pb_affiliation a
        JOIN unified_signal u ON u.ticker=a.ticker
        WHERE a.role_class='outside_director' AND a.is_former=0 AND u.sec_type='common'
        GROUP BY a.ticker
        ORDER BY u.score DESC, nd DESC LIMIT 140""").fetchall()
    out = []
    for r in rows:
        out.append([r[0], (r[1] or "")[:26], round(r[2], 1) if r[2] is not None else "",
                    round(r[3], 1) if r[3] is not None else "", r[4],
                    (r[5] or "")[:44], (r[6] or "").replace(" / ", "/")[:24], (r[7] or "")[:16]])
    write_table_rows(ws, out, 5, ticker_col=1)
    return ws


def sheet_investor_seats(wb, conn):
    ws = wb.create_sheet("Investor Board Seats")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Investor Board Seats — the activism / involvement tell",
                "People whose day job is investing (AM / hedge fund / PE / VC / family office) taking an OUTSIDE board seat. The highest-signal placements.", 7)
    hdr = ["Director", "Investor Identity", "Board Seat At", "Ticker", "Our Score", "Smart$ n", "Role"]
    write_table_header(ws, 4, hdr)
    ph = ",".join("?" * len(INVESTOR_TYPES))
    # High-signal = an OUTSIDE board seat held by either (a) a tracked principal
    # (a titan/activist personally), or (b) someone whose PitchBook identity is an
    # investment firm. The investor firm, when known, is shown for context.
    rows = conn.execute(f"""
        WITH investor_id AS (
            SELECT full_name, MAX(primary_company) FILTER (WHERE primary_company_type IN ({ph})) AS firm
            FROM pb_people GROUP BY full_name)
        SELECT a.full_name,
               COALESCE(inv.firm, '(principal)') AS firm,
               a.company, a.ticker, u.score, u.smart_money_n, a.position, a.is_principal
        FROM pb_affiliation a
        LEFT JOIN investor_id inv ON inv.full_name = a.full_name
        LEFT JOIN unified_signal u ON u.ticker = a.ticker
        WHERE a.role_class='outside_director' AND a.is_former=0
          AND a.company_type='Public Company'
          AND (a.is_principal=1 OR inv.firm IS NOT NULL)
          AND (inv.firm IS NULL OR a.company <> inv.firm)
        GROUP BY a.full_name, a.company
        ORDER BY (u.score IS NULL), u.score DESC, a.is_principal DESC, a.full_name
        LIMIT 160""", INVESTOR_TYPES).fetchall()
    out = []
    for r in rows:
        firm = r[1] if r[1] != "(principal)" else ("tracked principal" if r[7] else "")
        out.append([r[0], (firm or "")[:24], (r[2] or "")[:24], r[3] or "",
                    round(r[4], 1) if r[4] is not None else "",
                    round(r[5], 1) if r[5] is not None else "", (r[6] or "")[:22]])
    write_table_rows(ws, out, 5, ticker_col=4)
    return ws


def sheet_convergence(wb, conn):
    ws = wb.create_sheet("Board Convergence")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Board Convergence — where networks stack up",
                "Companies with MULTIPLE connected outside directors, or directors from DISTINCT networks. Coordinated positioning is the strongest signal.", 7)
    hdr = ["Company", "Ticker", "# Directors", "# Networks", "Networks", "Directors", "Our Score"]
    write_table_header(ws, 4, hdr)
    rows = conn.execute("""
        SELECT a.company, a.ticker,
               COUNT(DISTINCT a.full_name) nd,
               COUNT(DISTINCT a.theme) nt,
               GROUP_CONCAT(DISTINCT a.theme),
               GROUP_CONCAT(DISTINCT a.full_name), u.score
        FROM pb_affiliation a
        LEFT JOIN unified_signal u ON u.ticker=a.ticker
        WHERE a.role_class='outside_director' AND a.is_former=0
          AND a.company_type='Public Company'
        GROUP BY a.company
        HAVING COUNT(DISTINCT a.full_name) >= 2
        ORDER BY nd DESC, nt DESC LIMIT 80""").fetchall()
    out = []
    for r in rows:
        out.append([(r[0] or "")[:30], r[1] or "—", r[2], r[3],
                    (r[4] or "").replace(" / ", "/")[:30], (r[5] or "")[:40],
                    round(r[6], 1) if r[6] is not None else ""])
    write_table_rows(ws, out, 5, ticker_col=2)
    return ws


def sheet_watchlist(wb, conn):
    ws = wb.create_sheet("Watchlist — Not in Universe")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Board Signal — Not Yet in Our Universe",
                "Companies with connected outside directors that our US-13F universe doesn't cover (mostly foreign). Board concentration as a watch/add list.", 4)
    hdr = ["Company", "# Directors", "Connected Directors", "Networks"]
    write_table_header(ws, 4, hdr)
    rows = conn.execute("""
        SELECT a.company, COUNT(DISTINCT a.full_name) nd,
               GROUP_CONCAT(DISTINCT a.full_name), GROUP_CONCAT(DISTINCT a.theme)
        FROM pb_affiliation a
        WHERE a.role_class='outside_director' AND a.is_former=0
          AND a.company_type='Public Company' AND a.ticker IS NULL
        GROUP BY a.company
        HAVING COUNT(DISTINCT a.full_name) >= 2
        ORDER BY nd DESC LIMIT 120""").fetchall()
    out = [[(r[0] or "")[:40], r[1], (r[2] or "")[:50], (r[3] or "").replace(" / ", "/")[:28]] for r in rows]
    write_table_rows(ws, out, 5, ticker_col=1)
    return ws


def sheet_fo_swf(wb, conn):
    ws = wb.create_sheet("Family Office & SWF Map")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Family Office / Sovereign-Wealth Board Map",
                "Where Gulf, royal, and billionaire-vehicle nominees hold board seats. Active roles; in-universe tickers first.", 6)
    hdr = ["Individual", "Board Seat At", "Type", "Network", "Ticker", "Our Score"]
    write_table_header(ws, 4, hdr)
    rows = conn.execute("""
        SELECT a.full_name, a.company, a.company_type, a.theme, a.ticker, u.score
        FROM pb_affiliation a
        LEFT JOIN unified_signal u ON u.ticker=a.ticker
        WHERE a.theme IN ('Gulf / Royal Family Offices','Billionaire / Oligarch Vehicles')
          AND a.is_former=0 AND a.role_class IN ('outside_director','operator')
          AND a.company_type IN ('Public Company','Asset Manager','PE/Buyout',
               'Family Office (Multi)','Family Office (Single)','Investor','Real Estate')
        ORDER BY (u.score IS NULL), u.score DESC, a.full_name LIMIT 200""").fetchall()
    out = []
    for r in rows:
        out.append([r[0], (r[1] or "")[:28], (r[2] or "")[:20], (r[3] or "").replace(" / ", "/")[:22],
                    r[4] or "", round(r[5], 1) if r[5] is not None else ""])
    write_table_rows(ws, out, 5, ticker_col=5)
    return ws


def sheet_new_funds(wb, conn):
    ws = wb.create_sheet("Network Roster Adds")
    ws.sheet_view.showGridLines = False
    write_title(ws, "Network Roster Adds — firms found & now tracked",
                "Investment firms surfaced in the PitchBook data that were missing from our tracker, verified as active 13F filers, now ingested for holdings.", 6)
    hdr = ["Fund", "Style", "Holdings", "13F Book", "Top Holding", "Top $M"]
    write_table_header(ws, 4, hdr)
    out = []
    for f in NEW_FUNDS:
        n = conn.execute("SELECT COUNT(*) FROM fund_13f_holdings WHERE fund=?", (f,)).fetchone()[0]
        book = conn.execute("SELECT SUM(value_k)/1e3 FROM fund_13f_holdings WHERE fund=?", (f,)).fetchone()[0]
        style = conn.execute("SELECT macro_style FROM fund_style WHERE fund=?", (f,)).fetchone()
        top = conn.execute("""SELECT ticker, issuer, value_k/1e3 FROM fund_13f_holdings
            WHERE fund=? ORDER BY value_k DESC LIMIT 1""", (f,)).fetchone()
        top_name = (top[0] or top[1]) if top else ""
        out.append([f, style[0] if style else "", n, round(book or 0, 1),
                    (top_name or "")[:22], round(top[2], 1) if top else ""])
    write_table_rows(ws, out, 5, ticker_col=1)
    for ridx in range(5, 5 + len(out)):
        ws.cell(row=ridx, column=4).number_format = NUMFMT_M_TO_B
        ws.cell(row=ridx, column=6).number_format = NUMFMT_M_TO_B
    return ws


def run():
    conn = sqlite3.connect(DB)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_readme(wb, conn)
    sheet_board_signal(wb, conn)
    sheet_investor_seats(wb, conn)
    sheet_convergence(wb, conn)
    sheet_watchlist(wb, conn)
    sheet_fo_swf(wb, conn)
    sheet_new_funds(wb, conn)
    for ws in wb.worksheets:
        autosize(ws)
        # parity with the other two workbooks: freeze the header + enable filtering
        # on every data sheet (was missing on all six people sheets).
        if ws.title != "README" and ws.max_row > 4:
            ws.freeze_panes = "B5"
            from openpyxl.utils import get_column_letter as _gcl
            ws.auto_filter.ref = f"A4:{_gcl(ws.max_column)}{ws.max_row}"
    set_default_font(wb)
    set_print_layout(wb, header_rows=4)
    add_contents_index(wb["README"], [s.title for s in wb.worksheets])
    wb.save(OUT)
    print(f"wrote {OUT}")
    conn.close()


if __name__ == "__main__":
    run()
