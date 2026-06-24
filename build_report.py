"""Harvard-style report builder — single-font, single-size, monochrome
Excel workbook from the latest screen output.

Aesthetic rules (strict):
  * Single typeface throughout: Garamond 11pt
  * No fills, no colours, no shading
  * Borders: thin rule below header row only
  * Bold reserved for sheet titles and column headers
  * Numbers right-aligned, text left-aligned, headers centred
  * Generous column widths; row heights slightly taller than default
  * Sheet names: lowercase, no abbreviations beyond standard ones

Sheets:
  1. about            — methodology and how to read the workbook
  2. summary          — universe + sleeve counts, top-line metrics
  3. activist watch   — resolution score > 0.20 ordered by it
  4. insider buys     — PDMR direction-resolved buys
  5. wind-downs       — committed + likely
  6. open-end / DCM   — OPEN_END_CONVERSION_PROPOSED + DCM_ACTIVE
  7. setup sleeve     — chart confirmation present
  8. fundamentals     — event catalyst, IRR-only ranking
  9. saba UKIT        — 8 named trusts + 3 equity stakes
 10. discount league  — sorted by current discount
 11. universe         — full ranked listing
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HERE = Path(os.path.dirname(os.path.abspath(__file__)))

# -- Aesthetic constants --
FONT_NAME = "Garamond"
FONT_SIZE = 11
THIN_RULE = Side(style="thin", color="000000")
HEADER_BORDER = Border(bottom=THIN_RULE)
TOP_RULE_BORDER = Border(top=THIN_RULE)


def _font(bold: bool = False, italic: bool = False) -> Font:
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold, italic=italic,
                color="000000")


def _apply_style(ws, title: str, header_row: int, n_rows: int,
                 n_cols: int, col_widths: list[float] | None = None) -> None:
    """Apply single-font/single-size aesthetic to a sheet."""
    # Default font on every cell
    for row in ws.iter_rows(min_row=1, max_row=header_row + n_rows + 4,
                            min_col=1, max_col=max(n_cols, 1) + 1):
        for c in row:
            if c.font is None or c.font.name != FONT_NAME:
                c.font = _font()
    # Title row (row 1)
    ws.cell(1, 1, title).font = _font(bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal="left")
    # Generated stamp (row 2)
    ws.cell(2, 1, f"As of {datetime.utcnow().strftime('%Y-%m-%d')}").font = _font(
        italic=True)
    # Header row
    for c in range(1, n_cols + 1):
        cell = ws.cell(header_row, c)
        cell.font = _font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                   wrap_text=True)
        cell.border = HEADER_BORDER
    # Body
    for r in range(header_row + 1, header_row + n_rows + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.font = _font()
            v = cell.value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left",
                                           vertical="top", wrap_text=False)
    # Closing rule under last data row
    for c in range(1, n_cols + 1):
        ws.cell(header_row + n_rows, c).border = Border(bottom=THIN_RULE)
    # Column widths
    if col_widths is None:
        col_widths = [16] * n_cols
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Row heights
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[header_row].height = 26
    # Freeze title + header
    ws.freeze_panes = ws.cell(header_row + 1, 1)


def _write_table(ws, headers: list[str], rows: list[list],
                 col_widths: list[float], title: str,
                 subtitle: str | None = None) -> None:
    """Write a table starting at row 4 (after title + date + blank)."""
    header_row = 4
    if subtitle:
        ws.cell(3, 1, subtitle).font = _font(italic=True)
        ws.row_dimensions[3].height = 14
    for j, h in enumerate(headers, 1):
        ws.cell(header_row, j, h)
    for i, r in enumerate(rows, 1):
        for j, v in enumerate(r, 1):
            ws.cell(header_row + i, j, v)
    _apply_style(ws, title=title, header_row=header_row, n_rows=len(rows),
                 n_cols=len(headers), col_widths=col_widths)


def _fmt_pct(v) -> str:
    if v is None or pd.isna(v):
        return ""
    try:
        return f"{float(v) * 100:.1f}%"
    except (TypeError, ValueError):
        return ""


def _fmt_pp(v) -> str:
    if v is None or pd.isna(v):
        return ""
    try:
        return f"{float(v):+.1f}pp"
    except (TypeError, ValueError):
        return ""


def _fmt_money(v, unit: str = "k") -> str:
    if v is None or pd.isna(v):
        return ""
    try:
        f = float(v)
        if unit == "k":
            return f"£{f / 1000:,.0f}k"
        if unit == "m":
            return f"£{f:,.1f}m"
    except (TypeError, ValueError):
        return ""
    return ""


def _fmt_n(v) -> int | str:
    if v is None or pd.isna(v):
        return ""
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return ""


def _s(v, n: int = 50) -> str:
    """Safe string truncation — pd.NaN is a float and would explode on
    subscript; the ternary handles every flavour of empty."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v)[:n]


# ---------------------------------------------------------------------

def build(results_csv: Path, out_path: Path) -> None:
    df = pd.read_csv(results_csv)
    inv = df[(df["error"].isna()) & (df["investable"] == True)].copy()
    wb = Workbook()

    # ---------- 1. about ----------
    ws = wb.active
    ws.title = "about"
    ws.cell(1, 1, "Wind-down / NAV-discount screener — model report")
    ws.cell(1, 1).font = _font(bold=True)
    ws.cell(2, 1, f"As of {datetime.utcnow().strftime('%Y-%m-%d')}").font = _font(italic=True)
    methodology = [
        "",
        "I. METHODOLOGY",
        "",
        "The screener ranks UK and US closed-end funds across two compounding",
        "axes. The first is whether a corporate-action catalyst is present or",
        "imminent — managed wind-down, strategic review, open-end conversion,",
        "tender, return-of-capital, or discount-control intervention. The",
        "second is whether the chart already shows the market absorbing the",
        "expected re-rating: a tight base near the volume-profile point of",
        "control with rising directional volume.",
        "",
        "Expected internal rate of return is computed as:",
        "",
        "  expected IRR = annualised(recovery × discount × NAV-trajectory",
        "                            × (1 − path risk) × P(catalyst))",
        "                  + dividend carry",
        "",
        "where recovery rate is asset-class-specific (0.97 for listed equity",
        "books down to 0.40 for distressed), and P(catalyst) is the prior",
        "probability adjusted upward by qualitative signal density (insider",
        "buying, advisor appointment, institutional accumulation, strategic-",
        "review filings) and discount stretch versus the three-year average.",
        "",
        "II. SLEEVE STRUCTURE",
        "",
        "Setup sleeve — chart confirmation present (BASE_ABSORBING, BASE_",
        "BREAKOUT, CAPITULATION, BASE_QUIET). Ranked by setup × IRR.",
        "",
        "Fundamentals sleeve — event catalyst tagged or auto-promoted from",
        "RNS, ranked on IRR alone. Captures stub wind-downs that fail to",
        "print a chart footprint as their liquidity dries up.",
        "",
        "Activist watch — resolution score > 0.20. Composite of advisor",
        "appointment, strategic-review filing, buyback frequency, insider",
        "buys, institutional accumulation, and known-activist TR-1 buys,",
        "decayed with a fifteen-day half-life.",
        "",
        "Micro sleeve — investability-gated names with non-trivial IRR.",
        "Position size capped at one per cent per name.",
        "",
        "III. DATA",
        "",
        "Discount and NAV data sourced from the AIC live Drupal feed for UK",
        "names; Yahoo book-value proxy for US CEFs. Regulatory news (TR-1",
        "stake disclosures, PDMR insider deals, buybacks, wind-down and",
        "review announcements) scraped from Investegate. Activist holder",
        "identification matches the holder string against a curated list",
        "of UK and US closed-end-fund activist managers including Saba",
        "Capital, Asset Value Investors, City of London Investment, Lazard,",
        "Elliott, 1607 Capital, Metage Capital, and Almitas Capital.",
        "",
        "IV. CAVEATS",
        "",
        "Catalyst probabilities and durations remain hand-calibrated; an",
        "event-study backtest exists but has not yet been integrated as a",
        "calibration loop. Wind-down realised returns measured on price",
        "alone understate true total return by the value of cash",
        "distributions and remain a follow-up. Per-name exclusions and",
        "overrides live in universe.csv and signal_exclusions.",
    ]
    for i, line in enumerate(methodology, start=3):
        ws.cell(i, 1, line).font = _font()
    ws.column_dimensions["A"].width = 88
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    for i in range(3, 3 + len(methodology)):
        ws.row_dimensions[i].height = 15
    ws.cell(5, 1).font = _font(bold=True)
    # Bold the roman-numeral section heads
    for r in range(3, 3 + len(methodology)):
        v = ws.cell(r, 1).value or ""
        if v.startswith(("I.", "II.", "III.", "IV.")):
            ws.cell(r, 1).font = _font(bold=True)

    # ---------- 2. summary ----------
    ws = wb.create_sheet("summary")
    setup = inv[inv["in_setup_sleeve"] == True]
    fund = inv[inv["in_fundamentals_sleeve"] == True]
    micro = df[df["in_micro_sleeve"] == True]
    activists = inv[inv["resolution_score"].fillna(0) > 0.20]
    saba = inv[inv["saba_ukit_member"] == True]
    cat_counts = inv["catalyst"].value_counts().reset_index()
    cat_counts.columns = ["catalyst", "count"]
    rows = [
        ["Universe size", _fmt_n(len(df))],
        ["Investable", _fmt_n(len(inv))],
        ["Setup sleeve", _fmt_n(len(setup))],
        ["Fundamentals sleeve", _fmt_n(len(fund))],
        ["Micro sleeve", _fmt_n(len(micro))],
        ["Activist watch (resolution > 0.20)", _fmt_n(len(activists))],
        ["Saba UKIT members", _fmt_n(len(saba))],
        ["", ""],
        ["Median discount (investable)",
         _fmt_pct(inv["nav_discount_est"].median())],
        ["Median expected IRR", _fmt_pct(inv["expected_irr"].median())],
        ["Top IRR", _fmt_pct(inv["expected_irr"].max())],
        ["", ""],
    ]
    for _, r in cat_counts.iterrows():
        rows.append([f"  {r['catalyst']}", _fmt_n(r["count"])])
    _write_table(ws, headers=["metric", "value"], rows=rows,
                 col_widths=[40, 16], title="Summary statistics",
                 subtitle="Universe composition and catalyst distribution")

    # ---------- 3. activist watch ----------
    ws = wb.create_sheet("activist watch")
    aw = activists.sort_values("resolution_score", ascending=False).head(40)
    headers = ["ticker", "name", "phase", "catalyst", "resolution",
               "PDMR buys", "TR-1 material", "activist buys",
               "activist holders", "discount", "IRR", "saba UKIT"]
    rows = []
    for _, r in aw.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 42),
            r.get("phase", ""), r.get("catalyst", ""),
            float(r["resolution_score"] or 0),
            _fmt_n(r.get("rns_pdmr_buys")),
            _fmt_n(r.get("rns_tr1_material_adds")),
            _fmt_n(r.get("rns_tr1_activist_buys")),
            _s(r.get("activist_holders"), 50),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("expected_irr")),
            "yes" if r.get("saba_ukit_member") else "",
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 36, 14, 22, 11, 9, 11, 11, 36, 9, 9, 10],
                 title="Activist watch",
                 subtitle=("Names with resolution score above 0.20 — fresh "
                           "advisor / review / buyback / insider / "
                           "institutional accumulation cluster"))

    # ---------- 4. insider buys ----------
    ws = wb.create_sheet("insider buys")
    ib = inv[inv["rns_pdmr_buys"].fillna(0) > 0].sort_values(
        "rns_pdmr_buys", ascending=False).head(40)
    headers = ["ticker", "name", "PDMR buys", "£ value", "PDMR sells",
               "catalyst", "phase", "resolution", "IRR"]
    rows = []
    for _, r in ib.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 42),
            _fmt_n(r["rns_pdmr_buys"]),
            _fmt_money(r.get("pdmr_buy_gbp")),
            _fmt_n(r.get("rns_pdmr_sells")),
            r.get("catalyst", ""), r.get("phase", ""),
            float(r.get("resolution_score") or 0),
            _fmt_pct(r.get("expected_irr")),
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 38, 9, 11, 9, 22, 14, 11, 8],
                 title="Insider conviction — PDMR buys",
                 subtitle=("Director / PDMR purchases over the lookback "
                           "window; conviction-buys only, scrip and "
                           "vesting filtered"))

    # ---------- 5. wind-downs ----------
    ws = wb.create_sheet("wind-downs")
    wd = inv[inv["catalyst"].isin(
        ["WIND_DOWN_COMMITTED", "WIND_DOWN_LIKELY",
         "RETURN_OF_CAPITAL_LIVE"])].sort_values(
        "expected_irr", ascending=False).head(40)
    headers = ["ticker", "name", "catalyst", "NAV quality", "phase",
               "discount", "age months", "duration months",
               "P (event)", "IRR", "composite"]
    rows = []
    for _, r in wd.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 38),
            r.get("catalyst", ""), r.get("nav_quality", ""),
            r.get("phase", ""),
            _fmt_pct(r.get("nav_discount_est")),
            f"{float(r['catalyst_age_months']):.0f}" if pd.notna(
                r.get("catalyst_age_months")) else "",
            f"{float(r['expected_duration_months']):.0f}" if pd.notna(
                r.get("expected_duration_months")) else "",
            _fmt_pct(r.get("catalyst_prob_signal_adj")),
            _fmt_pct(r.get("expected_irr")),
            f"{float(r['composite_score']):.3f}" if pd.notna(
                r.get("composite_score")) else "",
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 34, 22, 16, 14, 9, 10, 12, 10, 8, 11],
                 title="Wind-downs and return-of-capital",
                 subtitle=("Committed, likely and live capital-return "
                           "names, ranked on annualised IRR"))

    # ---------- 6. open-end / DCM ----------
    ws = wb.create_sheet("DCM and open-end")
    dcm = inv[inv["catalyst"].isin(
        ["DCM_ACTIVE", "OPEN_END_CONVERSION_PROPOSED",
         "STRATEGIC_REVIEW"])].sort_values(
        "expected_irr", ascending=False).head(50)
    headers = ["ticker", "name", "catalyst", "phase", "discount",
               "vs sector", "resolution", "P (event)", "IRR", "composite"]
    rows = []
    for _, r in dcm.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 38),
            r.get("catalyst", ""), r.get("phase", ""),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pp(r.get("discount_vs_sector_pp")),
            float(r["resolution_score"] or 0),
            _fmt_pct(r.get("catalyst_prob_signal_adj")),
            _fmt_pct(r.get("expected_irr")),
            f"{float(r['composite_score']):.3f}" if pd.notna(
                r.get("composite_score")) else "",
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 36, 26, 14, 9, 10, 11, 10, 8, 11],
                 title="Discount control and open-end conversion",
                 subtitle=("Active DCM, proposed open-end conversion and "
                           "strategic review candidates"))

    # ---------- 7. setup sleeve ----------
    ws = wb.create_sheet("setup sleeve")
    su = setup.sort_values("composite_score", ascending=False).head(40)
    headers = ["ticker", "name", "phase", "catalyst", "discount",
               "POC dist", "base wks", "IRR", "composite"]
    rows = []
    for _, r in su.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 40),
            r.get("phase", ""), r.get("catalyst", ""),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("poc_distance_pct")),
            _fmt_n(r.get("base_length_weeks")),
            _fmt_pct(r.get("expected_irr")),
            f"{float(r['composite_score']):.3f}" if pd.notna(
                r.get("composite_score")) else "",
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 38, 16, 22, 9, 10, 9, 8, 10],
                 title="Setup sleeve — chart confirmation present",
                 subtitle=("Active base or breakout near the volume "
                           "profile point of control, ranked on "
                           "composite (setup × IRR)"))

    # ---------- 8. fundamentals sleeve ----------
    ws = wb.create_sheet("fundamentals sleeve")
    fu = fund.sort_values("expected_irr", ascending=False).head(40)
    headers = ["ticker", "name", "catalyst", "NAV quality", "discount",
               "P (event)", "duration months", "IRR", "drivers"]
    rows = []
    for _, r in fu.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 36),
            r.get("catalyst", ""), r.get("nav_quality", ""),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("catalyst_prob_signal_adj")),
            f"{float(r['expected_duration_months']):.0f}" if pd.notna(
                r.get("expected_duration_months")) else "",
            _fmt_pct(r.get("expected_irr")),
            _s(r.get("top_drivers"), 50),
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 32, 22, 16, 9, 10, 12, 8, 38],
                 title="Fundamentals sleeve — IRR ranked",
                 subtitle=("Event catalyst names without a chart "
                           "footprint requirement; useful when the "
                           "stub has dried up"))

    # ---------- 9. saba UKIT ----------
    ws = wb.create_sheet("saba UKIT")
    sb = saba.sort_values("resolution_score", ascending=False)
    headers = ["ticker", "name", "phase", "catalyst", "resolution",
               "discount", "IRR"]
    rows = []
    for _, r in sb.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 40),
            r.get("phase", ""), r.get("catalyst", ""),
            float(r["resolution_score"] or 0),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("expected_irr")),
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 38, 16, 22, 11, 9, 8],
                 title="Saba UKIT holdings",
                 subtitle=("The eight named investment trusts in Saba "
                           "Capital's UK Investment Trusts ETF — the "
                           "purest forward-looking activist signal"))

    # ---------- 10. discount league ----------
    ws = wb.create_sheet("discount league")
    league = inv.sort_values("nav_discount_est", ascending=False).head(60)
    headers = ["ticker", "name", "catalyst", "discount",
               "3yr avg", "52w high", "vs sector", "phase"]
    rows = []
    for _, r in league.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 38),
            r.get("catalyst", ""),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("discount_3y_avg")),
            _fmt_pct(r.get("discount_52w_high")),
            _fmt_pp(r.get("discount_vs_sector_pp")),
            r.get("phase", ""),
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 38, 26, 9, 9, 10, 10, 14],
                 title="Discount league",
                 subtitle="All investable names sorted by current discount")

    # ---------- 11. universe ----------
    ws = wb.create_sheet("universe")
    full = inv.sort_values("composite_score", ascending=False)
    headers = ["ticker", "name", "phase", "catalyst", "nav_quality",
               "discount", "IRR", "resolution", "composite",
               "PDMR buys", "TR-1 buys", "saba"]
    rows = []
    for _, r in full.iterrows():
        rows.append([
            r["ticker"], _s(r.get("name"), 32),
            r.get("phase", ""), r.get("catalyst", ""),
            r.get("nav_quality", ""),
            _fmt_pct(r.get("nav_discount_est")),
            _fmt_pct(r.get("expected_irr")),
            float(r["resolution_score"] or 0)
            if pd.notna(r.get("resolution_score")) else "",
            f"{float(r['composite_score']):.3f}" if pd.notna(
                r.get("composite_score")) else "",
            _fmt_n(r.get("rns_pdmr_buys")),
            _fmt_n(r.get("rns_tr1_buys")),
            "yes" if r.get("saba_ukit_member") else "",
        ])
    _write_table(ws, headers=headers, rows=rows,
                 col_widths=[8, 32, 14, 22, 16, 9, 8, 11, 11, 9, 9, 7],
                 title="Full ranked universe",
                 subtitle="All investable names")

    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("results_csv", nargs="?",
                   help="Path to results CSV (defaults to latest)")
    p.add_argument("--out", default=None)
    args = p.parse_args()
    if args.results_csv is None:
        cands = [c for c in HERE.glob("results_2026*.csv")
                 if "_top30" not in c.name and "_sleeves" not in c.name]
        if not cands:
            print("No results CSV found", file=sys.stderr)
            return 1
        args.results_csv = str(sorted(cands)[-1])
    stamp = datetime.utcnow().strftime("%Y%m%d")
    out = Path(args.out) if args.out else HERE / f"report_{stamp}.xlsx"
    build(Path(args.results_csv), out)
    print(f"Wrote {out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
