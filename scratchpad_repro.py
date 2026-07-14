"""
Reproduce the FIP asymmetric-v2 ranking to verify no silent drops
between the audit CSV and current code.

Steps:
 1) Re-apply v2 gates on the columns already in the CSV.
 2) Recompute sector median EV/EBITDA and diff vs sec_ev_med.
 3) Refetch fundamentals for 3 deterministic tickers via yfinance.
 4) Recompute asym_v2_score from (upside, floor, quality, stealth).
 5) Run build_workbook.py and verify structure.
"""
import subprocess
import sys
import math
import pandas as pd
import numpy as np

CSV = "/home/user/cyclepapa/asymmetric_v2_universe_audit.csv"
WB  = "/home/user/cyclepapa/FIP_Asymmetry_Workbook.xlsx"

df = pd.read_csv(CSV)
print(f"[load] rows in audit CSV: {len(df)}")
assert len(df) == 62, f"expected 62 rows, got {len(df)}"


# -----------------------------------------------------------------
# STEP 1 — Load-and-gate reproduction
# -----------------------------------------------------------------
print("\n=== STEP 1: v2 gates on audit values ===")
def check_row(r):
    fails = []
    if not (r["fip_d"] <= -0.08):                fails.append(f"fip_d={r['fip_d']:.4f} not <= -0.08")
    if not (r["fip_w"] <= -0.10):                fails.append(f"fip_w={r['fip_w']:.4f} not <= -0.10")
    if not (r["pret_d"] > 0):                    fails.append(f"pret_d={r['pret_d']:.4f} not > 0")
    if not (r["nonzero_pct"] >= 0.65):           fails.append(f"nonzero_pct={r['nonzero_pct']:.4f} not >= 0.65")
    if not (r["realized_vol_60d"] >= 0.008):     fails.append(f"realized_vol_60d={r['realized_vol_60d']:.4f} not >= 0.008")
    if not (r["last_price"] >= 1.0):             fails.append(f"last_price={r['last_price']:.4f} not >= 1.0")

    # floor gate ANY
    pb  = r["pb_use"]
    eve = r["ev_ebitda_use"]
    evs = r["ev_sales"]
    fcy = r["fcf_yield"]
    floor_ok = ( (pd.notna(pb)  and pb  <= 2.0) or
                 (pd.notna(eve) and eve <= 12.0) or
                 (pd.notna(evs) and evs <= 3.0) or
                 (pd.notna(fcy) and fcy >= 0.03) )
    if not floor_ok:
        fails.append(f"floor gate: pb={pb} evE={eve} evS={evs} fcy={fcy}")

    # catalyst gate
    rg   = r["rev_growth_use"]
    rgi  = r["rev_growth_inflection"]
    opm  = r["op_margin_ex"]
    eqg  = r["eps_q_growth"]
    catalyst_ok = (pd.notna(rg) and rg >= 0.05 and
                   pd.notna(rgi) and rgi >= 0 and
                   ((pd.notna(opm) and opm >= 0.05) or (pd.notna(eqg) and eqg >= 0)))
    if not catalyst_ok:
        fails.append(f"catalyst: rg={rg} rgi={rgi} opm={opm} eqg={eqg}")

    # survival gate
    de = r["debt_to_equity"]
    survival_ok = pd.isna(de) or (de <= 250)
    if not survival_ok:
        fails.append(f"survival: debt/eq={de}")

    return fails

pass_count = 0
fail_report = []
for _, r in df.iterrows():
    f = check_row(r)
    if not f:
        pass_count += 1
    else:
        fail_report.append((r["symbol"], f))

print(f"passing all gates on audit values: {pass_count} / {len(df)}")
if fail_report:
    print("FAILERS:")
    for sym, fails in fail_report:
        print(f"  {sym}: {'; '.join(fails)}")
else:
    print("no failers")


# -----------------------------------------------------------------
# STEP 2 — Sector-median reproduction
# -----------------------------------------------------------------
print("\n=== STEP 2: sector median EV/EBITDA ===")
sec_med = df.groupby("sector_used")["ev_ebitda_use"].median()
df["_sec_med_recomp"] = df["sector_used"].map(sec_med)
diffs = (df["_sec_med_recomp"] - df["sec_ev_med"]).abs()
print(f"max |sec_ev_med_recomp - sec_ev_med| = {diffs.max():.6f}")
# also spot-print per sector
per_sector = df.groupby("sector_used").agg(
    n=("symbol", "size"),
    csv_med=("sec_ev_med", "first"),
    recomp_med=("_sec_med_recomp", "first"),
)
print(per_sector.to_string())


# -----------------------------------------------------------------
# STEP 3 — Fresh fundamentals spot check
# -----------------------------------------------------------------
print("\n=== STEP 3: fresh fundamentals spot check ===")
survivors_sorted = df.sort_values("symbol").reset_index(drop=True)
picks = [survivors_sorted.iloc[0], survivors_sorted.iloc[30], survivors_sorted.iloc[61]]

import yfinance as yf
def _num(v):
    try:
        v = float(v)
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except Exception:
        return None

def fetch(sym):
    try:
        t = yf.Ticker(sym)
        info = t.info or {}
    except Exception as e:
        return {"err": str(e)}
    pb = _num(info.get("priceToBook"))
    ev = _num(info.get("enterpriseValue"))
    ebitda = _num(info.get("ebitda"))
    rev = _num(info.get("totalRevenue"))
    fcf = _num(info.get("freeCashflow"))
    mcap = _num(info.get("marketCap"))
    ev_ebitda = (ev / ebitda) if (ev is not None and ebitda not in (None, 0)) else None
    ev_sales  = (ev / rev)   if (ev is not None and rev   not in (None, 0)) else None
    fcf_yield = (fcf / mcap) if (fcf is not None and mcap not in (None, 0)) else None
    return {"pb": pb, "ev_ebitda": ev_ebitda, "ev_sales": ev_sales, "fcf_yield": fcf_yield}

for r in picks:
    sym = r["symbol"]
    fresh = fetch(sym)
    print(f"\n  {sym} ({r['name']})")
    if "err" in fresh:
        print(f"    yfinance error: {fresh['err']}")
        continue
    for k, csv_col in [("pb", "pb_use"),
                       ("ev_ebitda", "ev_ebitda_use"),
                       ("ev_sales", "ev_sales"),
                       ("fcf_yield", "fcf_yield")]:
        csv_v = r[csv_col]
        new_v = fresh[k]
        if new_v is None:
            drift = "n/a"
        else:
            try:
                drift = f"{(new_v - csv_v):+.4f} ({(new_v - csv_v) / csv_v * 100:+.1f}%)"
            except Exception:
                drift = "n/a"
        print(f"    {k:12s}: csv={csv_v}  fresh={new_v}  drift={drift}")


# -----------------------------------------------------------------
# STEP 4 — Rank-order reproduction
# -----------------------------------------------------------------
print("\n=== STEP 4: asym_v2_score recompute ===")
u = df["upside"].astype(float)
f = df["floor"].astype(float)
q = df["quality"].astype(float)
s = df["stealth"].astype(float)
recomp = np.sqrt(u * f) * (0.7 + 0.3 * q) * (0.8 + 0.2 * s)
csv_score = df["asym_v2_score"].astype(float)
drift = (recomp - csv_score).abs()
print(f"max |recomp - csv asym_v2_score| = {drift.max():.8f}")
print(f"# rows with drift > 1e-4 : {(drift > 1e-4).sum()}")
worst = drift.nlargest(5)
if not worst.empty:
    print("top 5 drifts:")
    for idx, d in worst.items():
        print(f"  {df.at[idx, 'symbol']:12s} csv={csv_score[idx]:.6f} recomp={recomp[idx]:.6f} drift={d:.2e}")


# -----------------------------------------------------------------
# STEP 5 — Workbook build reproduction
# -----------------------------------------------------------------
print("\n=== STEP 5: workbook build ===")
proc = subprocess.run(
    ["/usr/local/bin/python", "/home/user/cyclepapa/build_workbook.py"],
    capture_output=True, text=True,
)
print(f"exit: {proc.returncode}")
if proc.stdout.strip():
    print("stdout:", proc.stdout.strip()[-500:])
if proc.stderr.strip():
    print("stderr:", proc.stderr.strip()[-500:])

if proc.returncode == 0:
    from openpyxl import load_workbook
    wb = load_workbook(WB, read_only=False)
    print(f"tabs ({len(wb.sheetnames)}):", wb.sheetnames)
    if "All Survivors" in wb.sheetnames:
        ws = wb["All Survivors"]
        # count non-empty rows
        rows = list(ws.iter_rows(values_only=True))
        nonempty = sum(1 for row in rows if any(c not in (None, "") for c in row))
        # Header structure varies; report both non-empty rows and rows minus 1
        print(f"All Survivors: total non-empty rows = {nonempty}")
    if "Composite by Region" in wb.sheetnames:
        ws = wb["Composite by Region"]
        rows = list(ws.iter_rows(values_only=True))
        # print first 30 rows to identify region body rows
        print("Composite by Region — first 40 non-empty rows:")
        n = 0
        for row in rows:
            if any(c not in (None, "") for c in row):
                n += 1
                if n <= 40:
                    print(" ", row)
