"""Add forensic_audit tab to screener_report.xlsx."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

XLSX = Path('screener_report.xlsx')
SRC = Path('results_forensic/audit.csv')

df = pd.read_csv(SRC)
# Promote ticker to first column
if 'ticker' not in df.columns and df.columns[0] == 'Unnamed: 0':
    df = df.rename(columns={'Unnamed: 0': 'ticker'})
# Order columns
preferred = ['ticker','verdict','market_cap_B',
             'edgar_sales_ltm_yoy_pct','yf_sales_qyoy_pct',
             'gross_margin_chg_pp','yf_gross_qyoy_pct',
             'ps_now','ev_ebitda','notes']
cols = [c for c in preferred if c in df.columns]
df = df[cols]

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    df.to_excel(xw, sheet_name='forensic_audit', index=False)
    ws = xw.sheets['forensic_audit']
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    for c in 'DEFGHIJ': ws.column_dimensions[c].width = 18
    ws.column_dimensions['J'].width = 120
    ws.freeze_panes = 'B2'

# Move to front
wb = load_workbook(XLSX)
if 'forensic_audit' in wb.sheetnames:
    cur = wb.sheetnames.index('forensic_audit')
    wb.move_sheet('forensic_audit', offset=-cur)  # to position 0
    wb.save(XLSX)
print(f'Wrote forensic_audit tab with {len(df)} rows')
