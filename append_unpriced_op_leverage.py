"""Add unpriced_op_leverage tab to screener_report.xlsx.

Cross-filter on operating_leverage screener:
  - sales_growth_pct >= 15
  - margin_expansion_pp > 0 (operating leverage actually materializing)
  - ps_now < 2.5 AND ev_sales_now < 3 (cheap on sales)
  - market_cap >= $100M
  - 1y price perf in [-30%, +20%] (un-priced)
  - exclude India
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

CACHE = Path('.cache/yf')
XLSX = Path('screener_report.xlsx')

def safe(t): return ''.join(c if c.isalnum() or c in '-_' else '_' for c in t)

def price_perf(tk, days=252):
    p = CACHE / f'{safe(tk)}__price.parquet'
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        s = pd.to_numeric(df['Close'], errors='coerce').dropna()
        if len(s) < days+5: return None
        return float((s.iloc[-1]/s.iloc[-days] - 1) * 100)
    except Exception: return None


def _col(df, candidates):
    """Get a sorted Series for a line item, handling both orientations
    (dates-as-index for US/EU, items-as-index for some Asian markets)."""
    if df is None or df.empty: return None
    items_in_index = pd.api.types.is_datetime64_any_dtype(df.columns) or \
                     any(isinstance(c, pd.Timestamp) for c in df.columns[:3])
    for c in candidates:
        if items_in_index:
            matches = [ix for ix in df.index if str(ix) == c or str(ix).startswith(c[:10])]
            if matches:
                s = pd.to_numeric(df.loc[matches[0]], errors='coerce').dropna()
                if not s.empty: return s.sort_index()
        else:
            if c in df.columns:
                s = pd.to_numeric(df[c], errors='coerce').dropna()
                if not s.empty: return s.sort_index()
    return None


def gross_margin_change_pp(tk):
    """Latest gross margin minus prior-period gross margin, in percentage points."""
    p = CACHE / f'{safe(tk)}__income.parquet'
    if not p.exists(): return None
    try:
        inc = pd.read_parquet(p)
        if inc.empty: return None
    except Exception: return None
    rev = _col(inc, ['Total Revenue','Revenue','Operating Revenue'])
    gp  = _col(inc, ['Gross Profit'])
    if rev is None or gp is None: return None
    if len(rev) < 2 or len(gp) < 2: return None
    # Align indexes
    common = rev.index.intersection(gp.index)
    if len(common) < 2: return None
    rev = rev.reindex(common); gp = gp.reindex(common)
    # latest vs same period prior year if 5+ datapoints else first vs last
    if len(rev) >= 5:
        cur_m = float(gp.iloc[-1]) / float(rev.iloc[-1]) if rev.iloc[-1] > 0 else None
        prv_m = float(gp.iloc[-5]) / float(rev.iloc[-5]) if rev.iloc[-5] > 0 else None
    else:
        cur_m = float(gp.iloc[-1]) / float(rev.iloc[-1]) if rev.iloc[-1] > 0 else None
        prv_m = float(gp.iloc[0])  / float(rev.iloc[0])  if rev.iloc[0]  > 0 else None
    if cur_m is None or prv_m is None: return None
    return (cur_m - prv_m) * 100

def load_meta():
    meta = {}
    for f in ['universe_us_wide.csv','universe_expanded.csv','universe_wider.csv',
              'universe_eu.csv','universe_eu_extra.csv','universe_canada.csv',
              'universe_us_large_mega.csv']:
        p = Path(f)
        if not p.exists(): continue
        try:
            df = pd.read_csv(p, usecols=lambda c: c in ['symbol','name','sector','industry','country'], low_memory=False)
            for _, r in df.iterrows():
                tk = r.get('symbol')
                if not isinstance(tk, str): continue
                k = tk.upper()
                if k in meta: continue
                meta[k] = {f: r.get(f) for f in ['name','sector','industry','country'] if f in df.columns}
        except Exception: pass
    return meta

META = load_meta()

df = pd.read_csv('results_operating_leverage/screener.csv')
df = df[~df['ticker'].astype(str).str.upper().str.endswith(('.NS','.BO','_NS','_BO'))]
df['perf_1y_pct'] = df['ticker'].apply(price_perf)
df['gross_margin_chg_pp'] = df['ticker'].apply(gross_margin_change_pp)

mask = (
    (df['sales_growth_pct'] >= 15) &
    (df['margin_expansion_pp'] > 0) &
    (df['ps_now'] < 2.5) &
    (df.get('ev_sales_now', 999) < 3) &
    (df['market_cap'] > 100e6) &
    (df['perf_1y_pct'].between(-30, 20))
)
out = df[mask].sort_values(['margin_expansion_pp','sales_growth_pct'], ascending=[False, False]).copy()

# Flag rows where the operating leverage is STRUCTURAL (gross margin also expanding,
# not just OpEx leverage) — a stronger signal of pricing power / mix shift
out['structural_op_lev'] = out['gross_margin_chg_pp'].apply(
    lambda v: 'Y' if pd.notna(v) and v > 0 else ('N' if pd.notna(v) else '')
)

# add name/sector/industry/country from universe lookup
for col in ('name','sector','industry','country'):
    out[col] = out['ticker'].astype(str).str.upper().map(lambda t: META.get(t, {}).get(col))

cols = ['ticker','name','sector','industry','country',
        'rev_now_M','sales_growth_pct',
        'gross_margin_chg_pp','margin_expansion_pp','structural_op_lev',
        'ebitda_margin_now_pct','ebitda_growth_pct',
        'ps_now','ev_sales_now','ev_ebitda_now','psg','leverage_score',
        'market_cap','pb_now','perf_1y_pct']
cols = [c for c in cols if c in out.columns]
out = out[cols].reset_index(drop=True)

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    out.to_excel(xw, sheet_name='unpriced_op_leverage', index=False)
    ws = xw.sheets['unpriced_op_leverage']
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 28
    for c in 'EFGHIJKLMNOPQR':
        ws.column_dimensions[c].width = 14
    ws.freeze_panes = 'B2'

wb = load_workbook(XLSX)
if 'unpriced_op_leverage' in wb.sheetnames:
    cur = wb.sheetnames.index('unpriced_op_leverage')
    # Place right after asymmetry_tier (position 1)
    wb.move_sheet('unpriced_op_leverage', offset=1 - cur)
    wb.save(XLSX)

print(f'Wrote unpriced_op_leverage tab with {len(out)} rows')
