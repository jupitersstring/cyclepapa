"""Append akre_inflection and flat_inflection tabs to screener_report.xlsx."""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

CACHE = Path('.cache/yf')
XLSX = Path('screener_report.xlsx')

def _safe(t): return ''.join(c if c.isalnum() or c in '-_' else '_' for c in t)

def load_meta_lookup():
    """sector/industry/name from universe CSVs."""
    meta = {}
    for f in ['universe_us_wide.csv','universe_expanded.csv','universe_wider.csv',
              'universe_eu.csv','universe_eu_extra.csv','universe_canada.csv']:
        p = Path(f)
        if not p.exists(): continue
        try:
            df = pd.read_csv(p, usecols=lambda c: c in ['symbol','name','sector','industry','country'], low_memory=False)
            for _, r in df.iterrows():
                tk = r.get('symbol')
                if not isinstance(tk, str): continue
                key = tk.upper()
                if key in meta: continue
                meta[key] = {f: r.get(f) for f in ['name','sector','industry','country'] if f in df.columns}
        except Exception: pass
    return meta

META = load_meta_lookup()

def enrich(df, tk_col='ticker'):
    if df.index.name == 'ticker':
        df = df.reset_index()
    rows = []
    for tk in df[tk_col].astype(str):
        m = META.get(tk.upper(), {})
        rows.append({'name': m.get('name'), 'sector': m.get('sector'),
                     'industry': m.get('industry'), 'country': m.get('country')})
    extras = pd.DataFrame(rows, index=df.index)
    return pd.concat([df, extras], axis=1)

TABS = [
    ('akre_inflection',  'results_akre/screener.csv',           50),
    ('flat_inflection',  'results_flat_inflection/screener.csv', 50),
]

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    for name, path, top_n in TABS:
        p = Path(path)
        if not p.exists():
            print(f'  [skip] {name}: {path} missing')
            continue
        df = pd.read_csv(p)
        if 'ticker' in df.columns: df = df.head(top_n)
        else: df = df.head(top_n)
        df = enrich(df)
        df.to_excel(xw, sheet_name=name, index=False)
        ws = xw.sheets[name]
        ws.column_dimensions['A'].width = 12
        ws.freeze_panes = 'B2'
        print(f'  wrote {name}: {len(df)} rows')

# Move new tabs to position 1-2 (right after asymmetry_tier which is at 0)
wb = load_workbook(XLSX)
target_order = ['asymmetry_tier','akre_inflection','flat_inflection','stealth_breakout',
                'archetype_top5_unrerated','archetype_top5','composite_score']
present = [s for s in target_order if s in wb.sheetnames]
# remaining = everything else in current order
for s in present:
    cur = wb.sheetnames.index(s)
    desired = present.index(s)
    if cur != desired:
        wb.move_sheet(s, offset=desired - cur)
wb.save(XLSX)

print(f'\nWorkbook tabs in order:')
wb2 = load_workbook(XLSX)
for i, s in enumerate(wb2.sheetnames):
    print(f'  {i}: {s}')
