"""Add stealth_breakout tab — names that:
  - hit a screener archetype's top of list
  - 1y perf is modest (-25% to +30%, hasn't violently rerated)
  - within 8% of 5y high (close to ATH)
  - trade above 200d moving average (in uptrend)
  - market cap >= $100M
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

CACHE = Path('.cache/yf')
XLSX = Path('screener_report.xlsx')

PERF_LO, PERF_HI = -25.0, 30.0
ATH_TOL = 8.0
MIN_MCAP = 100e6

def safe(t): return ''.join(c if c.isalnum() or c in '-_' else '_' for c in t)

def technicals(ticker):
    p = CACHE / f'{safe(ticker)}__price.parquet'
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        if df.empty or 'Close' not in df.columns: return None
        s = pd.to_numeric(df['Close'], errors='coerce').dropna()
        if len(s) < 252: return None
        last = float(s.iloc[-1])
        five_y_high = float(s.max())
        out = {
            'last': last,
            'pct_off_5y_high': round((last/five_y_high - 1) * 100, 1),
            'perf_1y': round((last/float(s.iloc[-252]) - 1) * 100, 1),
            'perf_13w': round((last/float(s.iloc[-65]) - 1) * 100, 1) if len(s) >= 65 else None,
        }
        if len(s) >= 200:
            ma200 = float(s.iloc[-200:].mean())
            out['above_200d_ma'] = bool(last > ma200)
            out['pct_above_200d_ma'] = round((last/ma200 - 1) * 100, 1)
        s_52w = s.iloc[-252:]
        out['high_52w'] = float(s_52w.max())
        out['pct_off_52w_high'] = round((last/out['high_52w'] - 1) * 100, 1)
        return out
    except Exception: return None

def load_info(ticker):
    p = CACHE / f'{safe(ticker)}__info_metrics.parquet'
    if not p.exists(): return {}
    try:
        d = pd.read_parquet(p)
        return d.iloc[0].to_dict() if not d.empty else {}
    except Exception: return {}

DEEP_RESEARCHED = {
    'UVSP','AXTI','GNE','FRAF','GRWG','AMAL','ENVB','RLMD','WFRD','LRN','MGM','NOV','HOTH','CRWS','ACRV','IRMD','RDVT',
    'TVRD','VVOS','LXRX','NBN','INDI','JBSS','CFBK','RBKB','TWO-PB','VELO','FIBK','COCO','LEU','OFG','OPEN','TITN','ENVX',
    'SNAP','FRMM','ELTX','TWO-PA','DJCO','HFBL','DMRC','LUNA','UHT','AMPX','DAR','LAUR','ASPU','EIG','HCWB','ABOS',
    'BJRI','ASTH','AMSC','ADUS','LIND','EVER','INTT','IIIN','RCMT','INBK',
    'IFX.DE','MUV2.DE','GXI.DE','BYW.DE','TTK.DE','D6H.DE','SAX.DE','NEM.DE','FRE.DE','HEN.DE','SZG.DE','G1A.DE','SRT3.DE',
    'GLJ.DE','HNR1.DE','EVD.DE','EVK.DE','DTE.DE','PAT.DE','VPK.AS','SHL.DE','ZURN.SW','SMHN.DE','OHB.DE','AIXA.DE','CWC.DE',
    'ARIS.TO','FTT.TO','IAG.TO','FNV.TO','GOOS.TO','X.TO','IGM.TO','DSG.TO','IFC.TO','CM.TO','NA.TO','BPF-UN.TO',
    'BNS.TO','QBR-B.TO','PXT.TO','SLF.TO','GWO.TO','REI-UN.TO','AEM.TO','TIH.TO','WSP.TO','VLN.TO','TOU.TO','ENB.TO',
}

ARCHETYPES = [
    ('multi_variant_US',      'results_us_wide/ranked.csv',          'avg_inflection_z', False, None),
    ('multi_variant_EU',      'results_eu_relaxed/ranked.csv',        'avg_inflection_z', False, None),
    ('multi_variant_CA',      'results_canada/ranked.csv',            'avg_inflection_z', False, None),
    ('fcf_signflip_strict',   'results_us_wide/fcf_inflections.csv',  'flip_magnitude_pct', False,
        lambda d: (d.get('view') == 'quarterly_strict') & (d.get('metric') == 'fcf_ps') & (d.get('is_flip') == True)),
    ('fcf_signflip_ttm',      'results_us_wide/fcf_inflections.csv',  'flip_magnitude_pct', False,
        lambda d: (d.get('view') == 'ttm_yoy') & (d.get('metric') == 'fcf_ps') & (d.get('is_flip') == True)),
    ('deep_value_inflection', 'results_us_wide/deep_value_screen.csv', 'n_variants_inflected', False,
        lambda d: d.get('is_value_plus_inflection') == True),
    ('cheap_inflecting',      'results_us_wide/valuation_screen.csv', 'n_variants_inflected', False,
        lambda d: d.get('is_cheap_inflecting') == True),
    ('52wh_cheap_on_growth',  'results_52wh/screener.csv',            'revenue_growth_ltm', False, None),
    ('multiple_compression',  'results_multiple_compression/clean.csv', 'multiple_compression_pct', True,
        lambda d: d.get('multiple_compression_pct') < -10),
    ('ev_ebitda_compression', 'results_ev_compression/screener.csv',  'compression_vs_sales_pct', True,
        lambda d: (d.get('ev_ebitda_now') > 0) & (d.get('ev_ebitda_now') < 30) & (d.get('sales_growth_pct') > 5)),
    ('operating_leverage',    'results_operating_leverage/screener.csv', 'leverage_score', False,
        lambda d: d.get('market_cap') > 100e6),
    ('ev_fcf_leverage',       'results_ev_fcf_leverage/screener.csv', 'fcf_growth_pct', False, None),
    ('fcf_yield_setup',       'results_fcf_yield/screener.csv',       'fcf_yield_now_pct', False,
        lambda d: (d.get('fcf_yield_now_pct') < 60) & (d.get('fcf_yield_now_pct') > 0)),
    ('segment_pre_rerate',    'pre_rerate_setups.csv',                'pre_rerate_score', False,
        lambda d: d.get('pre_rerate_score') > 5),
    ('volasym_bullish',       'results_volasym/volatility_asymmetry.csv', 'm_asym', False,
        lambda d: d.get('m_state').isin(['squeeze','hyper_squeeze']) & (d.get('m_asym_state') == 'upper')),
]

rows = []
for label, path, sort_col, asc, filt in ARCHETYPES:
    p = Path(path)
    if not p.exists(): continue
    try: df = pd.read_csv(p)
    except: continue
    if filt is not None:
        try: df = df[filt(df)].copy()
        except: pass
    if sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=asc)
    tk_col = 'ticker' if 'ticker' in df.columns else df.columns[0]
    rank, seen = 0, set()
    for _, row in df.iterrows():
        if rank >= 5: break
        tk = str(row[tk_col])
        if tk in seen: continue
        if tk.endswith('.NS') or tk.endswith('.BO') or tk.endswith('_NS') or tk.endswith('_BO'): continue
        seen.add(tk)
        info = load_info(tk)
        mcap = info.get('marketCap')
        if mcap is None or mcap < MIN_MCAP: continue
        t = technicals(tk)
        if t is None or t.get('perf_1y') is None: continue
        if not (PERF_LO <= t['perf_1y'] <= PERF_HI): continue
        if t['pct_off_5y_high'] < -ATH_TOL: continue
        if not t.get('above_200d_ma'): continue
        rank += 1
        try: sig_val = float(row.get(sort_col)) if pd.notna(row.get(sort_col)) else None
        except: sig_val = None
        rows.append({
            'archetype': label,
            'rank': rank,
            'ticker': tk,
            'signal': sig_val,
            'perf_1y_pct': t['perf_1y'],
            'perf_13w_pct': t.get('perf_13w'),
            'pct_off_5y_high': t['pct_off_5y_high'],
            'pct_off_52w_high': t['pct_off_52w_high'],
            'pct_above_200d_ma': t.get('pct_above_200d_ma'),
            'market_cap': mcap,
            'trailingPE': info.get('trailingPE'),
            'priceToBook': info.get('priceToBook'),
            'priceToSales': info.get('priceToSalesTrailing12Months'),
            'currentPrice': t['last'],
            'deep_researched': 'Y' if tk in DEEP_RESEARCHED else '',
        })

out = pd.DataFrame(rows)
# convergence flag — count how many archetypes each ticker appears in
counts = out.groupby('ticker').size().rename('n_archetypes').reset_index()
out = out.merge(counts, on='ticker', how='left')
out = out.sort_values(['n_archetypes','archetype','rank'], ascending=[False, True, True])

with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
    out.to_excel(xw, sheet_name='stealth_breakout', index=False)
    ws = xw.sheets['stealth_breakout']
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 12
    for c in 'DEFGHIJK': ws.column_dimensions[c].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 13
    ws.column_dimensions['N'].width = 13
    ws.column_dimensions['O'].width = 12
    ws.column_dimensions['P'].width = 14
    ws.freeze_panes = 'A2'

wb = load_workbook(XLSX)
if 'stealth_breakout' in wb.sheetnames:
    cur = wb.sheetnames.index('stealth_breakout')
    wb.move_sheet('stealth_breakout', offset=1 - cur)  # right after asymmetry_tier
    wb.save(XLSX)

print(f'Wrote stealth_breakout with {len(out)} rows, {out["ticker"].nunique()} unique tickers')
print(f'Filter: 1y perf [{PERF_LO}%, {PERF_HI}%], within {ATH_TOL}% of 5y high, above 200d MA, mkt cap >= ${MIN_MCAP/1e6:.0f}M')
