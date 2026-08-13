#!/usr/bin/env python3
"""
psu_crossfeed.py — cross-feed from the PSU/governance 31-layer engine.

The repository carries two independent research systems:

  1. THIS subsystem (emergence/pollers): event-driven special-situations
     screening ranked by bottom-up reward/risk waterfalls.
  2. The PSU engine (branch claude/discretionary-insider-conviction):
     31 additive scoring layers over 6,166 US names — PSU/governance
     forensics, insider behavior (incl. the discretionary open-market
     conviction leg), tender mechanics, valuation, forced selling.

They source differently and rank differently, so agreement between them
is a genuine independent confirmation. This module reads a committed
read-only snapshot of the engine's outputs (data/psu_engine/) and adds
one sheet to the risk-reward workbook:

  - Engine top names (n_layers_firing, consensus score, key layer pts)
  - The CONVERGENCE set: names ranked by this subsystem's RR that also
    fire multiple engine layers
  - Top discretionary insider-conviction clusters

Degrades gracefully: if the snapshot is absent the sheet is skipped and
the workbook builds exactly as before.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook

REPO = Path(__file__).resolve().parent.parent
PSU_DIR = REPO / "data" / "psu_engine"
CONSENSUS_CSV = PSU_DIR / "full_universe_consensus.csv"
CONVICTION_JSON = PSU_DIR / "discretionary_insider_conviction.json"


def _stem(ticker: str) -> str:
    """'NYSE:TGS' -> 'TGS'; bare tickers pass through."""
    return re.sub(r"[^A-Za-z0-9.\-]", "", str(ticker).split(":")[-1]).upper()


def load_psu_engine() -> tuple[dict, dict]:
    """Return (consensus_by_ticker, conviction_by_ticker); empty dicts
    when the snapshot is not present."""
    consensus = {}
    if CONSENSUS_CSV.exists():
        with CONSENSUS_CSV.open() as f:
            for r in csv.DictReader(f):
                consensus[r["ticker"].upper()] = r
    conviction = {}
    if CONVICTION_JSON.exists():
        try:
            data = json.loads(CONVICTION_JSON.read_text())
            conviction = {k.upper(): v for k, v in data.items()
                          if isinstance(v, dict) and (v.get("score") or 0) > 0}
        except Exception:
            conviction = {}
    return consensus, conviction


def build_psu_engine_sheet(wb: Workbook, universe_rr: list[dict],
                           style_header_row, style_body_band,
                           autosize_cols, fonts: dict) -> None:
    """Add the 'PSU Engine x-feed' sheet. fonts carries the workbook's
    shared Font objects: {subhead, footnote, body_b}."""
    consensus, conviction = load_psu_engine()
    if not consensus:
        print("  (psu_engine snapshot absent — skipping x-feed sheet)")
        return

    ws = wb.create_sheet("PSU Engine x-feed")
    ws.sheet_view.showGridLines = False
    ws["A1"] = ("PSU/governance engine cross-feed — independent "
                "confirmation from the 31-layer consensus")
    ws["A1"].font = fonts["subhead"]
    ws["A2"] = (f"Snapshot of {len(consensus):,} US names scored by the "
                "PSU engine (31 additive layers: PSU forensics, "
                "governance, insider conviction, tender mechanics, "
                "valuation, forced selling). Names appearing in BOTH "
                "systems' rankings carry genuinely independent "
                "confirmation: the two systems share no sourcing.")
    ws["A2"].font = fonts["footnote"]
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 30

    # ---- Section 1: convergence set ---------------------------------
    rr_by_stem = {}
    for r in universe_rr:
        s = _stem(r.get("ticker", ""))
        if s and s not in rr_by_stem:
            rr_by_stem[s] = r

    convergent = []
    for tk, c in consensus.items():
        try:
            n_layers = int(float(c.get("n_layers_firing") or 0))
        except Exception:
            n_layers = 0
        if n_layers < 2:
            continue
        rr = rr_by_stem.get(tk)
        if rr is None:
            continue
        convergent.append((tk, n_layers,
                           float(c.get("consensus_score") or 0), rr))
    convergent.sort(key=lambda x: (-x[1], -x[2]))

    row = 4
    ws.cell(row=row, column=1,
            value="Convergence set — in this universe AND firing 2+ "
                  "engine layers").font = fonts["body_b"]
    row += 1
    headers = ["Ticker", "Name", "Engine layers", "Engine score",
               "RR here", "Source", "Bucket", "Archetype",
               "Insider conviction", "Conviction flags"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(headers))
    row += 1
    for i, (tk, n_layers, cons, rr) in enumerate(convergent[:40]):
        conv = conviction.get(tk) or {}
        cells = [
            tk, rr.get("name", "")[:40], n_layers, round(cons, 2),
            f"{float(rr.get('rr') or 0):.1f}×", rr.get("source", ""),
            rr.get("bucket", ""), rr.get("archetype", "")[:10],
            (conv.get("score") or "—"),
            "; ".join(conv.get("flags") or [])[:60] or "—",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))
        row += 1

    # ---- Section 2: engine top names not in this universe -----------
    row += 2
    ws.cell(row=row, column=1,
            value="Engine top names NOT in this subsystem's universe — "
                  "candidate sourcing gap").font = fonts["body_b"]
    row += 1
    headers2 = ["Ticker", "Engine layers", "Engine score", "PSU pts",
                "Valuation pts", "Insider conviction", "Conviction flags"]
    for j, h in enumerate(headers2, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(headers2))
    row += 1
    missing = []
    for tk, c in consensus.items():
        try:
            n_layers = int(float(c.get("n_layers_firing") or 0))
        except Exception:
            n_layers = 0
        if n_layers >= 5 and tk not in rr_by_stem:
            missing.append((tk, n_layers,
                            float(c.get("consensus_score") or 0), c))
    missing.sort(key=lambda x: (-x[1], -x[2]))
    for i, (tk, n_layers, cons, c) in enumerate(missing[:30]):
        conv = conviction.get(tk) or {}
        cells = [
            tk, n_layers, round(cons, 2),
            round(float(c.get("psu_pts") or 0), 0),
            round(float(c.get("valuation_pts") or 0), 0),
            (conv.get("score") or "—"),
            "; ".join(conv.get("flags") or [])[:60] or "—",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers2), banded=(i % 2 == 1))
        row += 1

    # ---- Section 3: top discretionary insider-conviction clusters ---
    row += 2
    ws.cell(row=row, column=1,
            value="Top discretionary insider-conviction clusters "
                  "(open-market Form 4 code P only)").font = fonts["body_b"]
    row += 1
    headers3 = ["Ticker", "Conviction", "Insiders", "Same-day", "C-suite",
                "Top buyer $M", "Configuration"]
    for j, h in enumerate(headers3, 1):
        ws.cell(row=row, column=j, value=h)
    style_header_row(ws, row, len(headers3))
    row += 1
    top_conv = sorted(conviction.items(),
                      key=lambda x: -(x[1].get("score") or 0))[:20]
    for i, (tk, v) in enumerate(top_conv):
        cells = [
            tk, v.get("score"), v.get("n_insiders"),
            v.get("same_day_cluster"), v.get("csuite_buyers"),
            round((v.get("top_person_dollar") or 0) / 1e6, 2),
            "; ".join(v.get("flags") or [])[:70],
        ]
        for j, val in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=val)
        style_body_band(ws, row, len(headers3), banded=(i % 2 == 1))
        row += 1

    row += 2
    ws.cell(row=row, column=1, value=(
        "Source: data/psu_engine/ — read-only snapshot of "
        "full_universe_consensus.csv + discretionary_insider_conviction"
        ".json from branch claude/discretionary-insider-conviction "
        "(see data/psu_engine/SNAPSHOT.md for refresh instructions). "
        "The engine is US-common-stock only, so non-US names in this "
        "workbook legitimately have no engine row. 'Engine layers' is "
        "n_layers_firing: how many of the 31 independent additive "
        "layers produce a non-zero score."
    )).font = fonts["footnote"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row,
                   end_column=10)
    ws.row_dimensions[row].height = 44
    autosize_cols(ws)

    n_conv = len(convergent)
    print(f"  psu_crossfeed: {n_conv} convergent names, "
          f"{len(missing)} engine-only 5+ layer names, "
          f"{len(conviction)} conviction clusters")
