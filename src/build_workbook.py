#!/usr/bin/env python3
"""
build_workbook.py — generate a Harvard-style special-situations workbook.

Pulls the canonical pieces of the framework into one .xlsx:
- Cover + investment recommendation
- Executive summary: top picks ranked by risk-reward ratio
- Universe table (all Tier 1+2 names)
- Waterfall matrix (bear/base/bull P × R)
- Catalyst timeline
- One detail sheet per top pick (deal, scorecard, catalysts, waterfall,
  triangulation, red flags, kill criteria, pre-mortem)
- Factor exposure / correlation map
- Portfolio sizing (Kelly → correlation haircut → cluster cap)
- Methodology

Output: output/cyclepapa_risk_reward_workbook.xlsx

Usage:
    python3 -m src.build_workbook
"""

from __future__ import annotations

import re
import sys
from datetime import date
from pathlib import Path

try:
    import yaml
except ImportError:
    print("Install PyYAML", file=sys.stderr); sys.exit(1)
try:
    from openpyxl import Workbook
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print("Install openpyxl", file=sys.stderr); sys.exit(1)

REPO = Path(__file__).resolve().parent.parent
CANDIDATES = REPO / "data" / "candidates"
PORTFOLIO_MD = REPO / "output" / "portfolio.md"
UNIVERSE_RR_CSV = REPO / "output" / "universe_risk_reward.csv"
LISTED_EQUITY_MD = REPO / "output" / "listed_equity_watchlist.md"
POSTREORG_MD = REPO / "output" / "postreorg_watchlist.md"
OUT_DIR = REPO / "output"
OUT_PATH = OUT_DIR / "cyclepapa_risk_reward_workbook.xlsx"

# Harvard-academic palette: dark navy + ledger grey
NAVY = "0E2A47"
NAVY_LIGHT = "5B6B85"
RULE = "8B939F"
BAND = "F2F4F6"
GOLD = "B58A2A"
RED = "9E2A2B"
WHITE = "FFFFFF"
GREEN = "2C5F2D"

THIN_GREY = Side(style="thin", color=RULE)
MED_NAVY = Side(style="medium", color=NAVY)

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_BODY = Font(name="Calibri", size=10, color="111111")
FONT_BODY_B = Font(name="Calibri", size=10, color="111111", bold=True)
FONT_SUBHEAD = Font(name="Calibri", size=11, bold=True, color=NAVY)
FONT_TITLE = Font(name="Calibri", size=22, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Calibri", size=13, color=NAVY_LIGHT, italic=True)
FONT_FOOTNOTE = Font(name="Calibri", size=9, color="555555", italic=True)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_BAND = PatternFill("solid", fgColor=BAND)
FILL_HIGHLIGHT = PatternFill("solid", fgColor="FBEAB2")

BORDER_BOX = Border(left=THIN_GREY, right=THIN_GREY,
                    top=THIN_GREY, bottom=THIN_GREY)
BORDER_HEAD = Border(left=MED_NAVY, right=MED_NAVY,
                     top=MED_NAVY, bottom=MED_NAVY)

ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ----- Load candidates --------------------------------------------------------

def load_candidates() -> list[dict]:
    """Active Tier 1+2 candidates with derived risk-reward metrics."""
    out = []
    for p in sorted(CANDIDATES.glob("*.yaml")):
        with p.open() as f:
            d = yaml.safe_load(f) or {}
        if not isinstance(d, dict):
            continue
        if d.get("state") == "pass":
            continue
        if d.get("tier") not in (1, 2):
            continue
        # Expected value across waterfall scenarios
        w = d.get("waterfall", {}) or {}
        ev = 0.0
        for k in ("bear", "base", "bull"):
            s = w.get(k, {}) or {}
            p_ = float(s.get("p", 0) or 0)
            r_ = float(s.get("return_multiple", 0) or 0)
            ev += p_ * r_
        bear_r = float((w.get("bear") or {}).get("return_multiple", 1.0) or 1.0)
        bull_r = float((w.get("bull") or {}).get("return_multiple", 1.0) or 1.0)
        downside = max(0.01, 1.0 - bear_r)  # % loss in bear case
        upside = max(0.01, bull_r - 1.0)
        skew = upside / downside
        rr_ratio = (ev - 1.0) / downside if downside > 0 else 0.0
        d["_ev"] = round(ev, 2)
        d["_downside"] = round(downside, 2)
        d["_upside"] = round(upside, 2)
        d["_skew"] = round(skew, 1)
        d["_rr_ratio"] = round(rr_ratio, 1)
        d["_bear_r"] = bear_r
        d["_bull_r"] = bull_r
        d["_base_r"] = float((w.get("base") or {}).get("return_multiple", 1.0) or 1.0)
        out.append(d)
    return sorted(out, key=lambda x: -x["_rr_ratio"])


# ----- Parse portfolio weights from output/portfolio.md ----------------------

def load_portfolio_weights() -> dict[str, dict]:
    """Pull per-name weight + cluster + factor from portfolio.md."""
    out: dict[str, dict] = {}
    if not PORTFOLIO_MD.exists():
        return out
    text = PORTFOLIO_MD.read_text()
    # 'Risk-budgeted basket weights' table parses the row pattern
    pat = re.compile(
        r"\|\s*\*\*([^|*]+)\*\*\s*\|\s*([^|]+)\|\s*([\d.]+%)\s*\|"
        r"\s*([\d.]+%)\s*\|\s*([\d.]+%)\s*\|\s*([\d.]+)% × ([\d.]+)\s*\|"
        r"\s*([\d.]+) bps\s*\|"
    )
    for m in pat.finditer(text):
        ticker = m.group(1).strip()
        cluster = m.group(2).strip()
        out[ticker] = {
            "ticker":    ticker,
            "cluster":   cluster,
            "raw_kelly": m.group(3),
            "corr_haircut": m.group(4),
            "capped":    m.group(5),
            "weight":    float(m.group(6)),
            "ev_mult":   float(m.group(7)),
            "contrib_bps": float(m.group(8)),
        }
    return out


# ----- Sheet builders --------------------------------------------------------

def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = BORDER_HEAD


def style_body_band(ws, row: int, n_cols: int, banded: bool = False) -> None:
    fill = FILL_BAND if banded else None
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_BODY
        if fill: cell.fill = fill
        cell.border = BORDER_BOX
        if c == 1:
            cell.alignment = ALIGN_L
        else:
            cell.alignment = ALIGN_C


def autosize_cols(ws, max_w: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        best = 8
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            best = max(best, min(max_w, len(v) + 2))
        ws.column_dimensions[col_letter].width = best


def build_cover(wb: Workbook, candidates: list[dict],
                weights: dict[str, dict]) -> None:
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws["A2"] = "cyclepapa"
    ws["A2"].font = FONT_TITLE
    ws["A3"] = "Special Situations · Risk-Reward Workbook"
    ws["A3"].font = FONT_SUBTITLE
    ws["A5"] = f"Generated {date.today().isoformat()}"
    ws["A5"].font = FONT_FOOTNOTE

    # Recommendation block
    ws["A8"] = "Investment recommendation"
    ws["A8"].font = FONT_SUBHEAD
    ws["A8"].border = Border(bottom=MED_NAVY)
    n_active = len(candidates)
    invested = sum(w["weight"] for w in weights.values())
    port_ev = sum(w["weight"] * w["ev_mult"]
                  for w in weights.values()) / max(invested, 0.01)

    bullets = [
        ("Universe coverage: 697 named candidates screened quantitatively; "
         "Executive Summary ranks the investable subset by reward/risk."),
        (f"Hand-built YAMLs with bottom-up waterfalls: {n_active} names "
         "(Tier 1 + 2). Remaining ranking uses a transparent "
         "score+archetype proxy; coverage gap sheet lists the highest-RR "
         "names that still need hand-deepening."),
        (f"Risk-budgeted invested weight on the built basket: "
         f"{invested*100:.1f}% of NAV "
         f"({(1-invested)*100:.1f}% cash sleeve — Kelly haircut binding). "
         f"Expected multiple on invested capital: {port_ev:.2f}×."),
        ("Reading order: Cover → Executive Summary (universe-wide ranking) "
         "→ Coverage gap (what to build next) → per-pick detail sheets → "
         "Methodology."),
    ]
    for i, t in enumerate(bullets):
        ws.cell(row=10 + i, column=1, value=("•  " + t)).font = FONT_BODY
        ws.cell(row=10 + i, column=1).alignment = ALIGN_WRAP

    # Methodology footer
    ws["A18"] = "How to read this workbook"
    ws["A18"].font = FONT_SUBHEAD
    ws["A18"].border = Border(bottom=MED_NAVY)
    legend = [
        "Reward/Downside ratio = (Expected EV× − 1) ÷ Bear-case loss.",
        "Skew = (Bull return − 1) ÷ Bear-case loss.",
        "Expected EV× = Σ pᵢ × Rᵢ across waterfall scenarios.",
        "Weight = post-Kelly haircut + cluster cap, from portfolio.py.",
        ("Confidence column on tabular sheets: ● verified, ◐ partial, "
         "○ skeleton/option."),
    ]
    for i, t in enumerate(legend):
        ws.cell(row=20 + i, column=1, value=("•  " + t)).font = FONT_BODY
        ws.cell(row=20 + i, column=1).alignment = ALIGN_WRAP

    # Set column widths
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[2].height = 32


def load_universe_rr() -> list[dict]:
    """Load the universe-wide reward/risk ranking from CSV (produced by
    src/universe_risk_reward.py)."""
    if not UNIVERSE_RR_CSV.exists():
        return []
    import csv
    out = []
    with UNIVERSE_RR_CSV.open() as f:
        for r in csv.DictReader(f):
            out.append(r)
    return out


def build_executive_summary(wb: Workbook, candidates: list[dict],
                            weights: dict[str, dict],
                            universe_rr: list[dict]) -> None:
    """Executive summary ranks the FULL universe by quantitative
    reward/risk — REAL waterfalls (hand-built YAMLs) where they exist,
    PROXY (universe-screener score × archetype tilt) where they don't."""
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Universe-wide top 100 · ranked by quantitative reward/risk (asymmetry)"
    ws["A1"].font = FONT_SUBHEAD
    ws["A2"] = (f"Universe coverage: {len(universe_rr)} investable names "
                f"({sum(1 for r in universe_rr if r['source']=='REAL')} "
                "REAL waterfalls · "
                f"{sum(1 for r in universe_rr if r['source']=='PROXY')} "
                "PROXY).  REAL uses the hand-built YAML's bottom-up "
                "bear/base/bull. PROXY uses a transparent formula on the "
                "universe-screener score + archetype tilt.")
    ws["A2"].font = FONT_FOOTNOTE
    ws["A2"].alignment = ALIGN_WRAP
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 38

    headers = ["Rank", "Src", "Ticker", "Name", "Region", "Score",
               "Bucket", "Archetype", "Bear loss", "Base", "Bull",
               "EV×", "Reward/Risk"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=4, column=j, value=h)
    style_header_row(ws, 4, len(headers))

    top = universe_rr[:100]
    for i, r in enumerate(top):
        row = 5 + i
        cells = [
            i + 1,
            r["source"],
            r["ticker"],
            r["name"][:50],
            r["region"][:22],
            f"{float(r['score']):.2f}",
            r["bucket"],
            r["archetype"][:10],
            f"{float(r['bear_loss'])*100:.0f}%",
            f"{float(r['base_r']):.2f}×",
            f"{float(r['bull_r']):.2f}×",
            f"{float(r['ev']):.2f}×",
            f"{float(r['rr']):.1f}×",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))
        # Bold the rank cell to make Src visible at a glance
        if r["source"] == "REAL":
            ws.cell(row=row, column=2).font = FONT_BODY_B
        else:
            ws.cell(row=row, column=2).font = Font(name="Calibri", size=10,
                                                    color=NAVY_LIGHT, italic=True)

    last_row = 4 + len(top)
    ws.conditional_formatting.add(
        f"M5:M{last_row}",
        ColorScaleRule(start_type="min", start_color="F2F4F6",
                       mid_type="percentile", mid_value=50,
                       mid_color="C7E8C0",
                       end_type="max", end_color=GREEN),
    )

    # Footnote
    ws.cell(row=last_row + 2, column=1, value=(
        "Reward/Risk = (EV× − 1) ÷ Bear loss = the asymmetry metric.  "
        "The broad universe-screener scores the full candidate set on "
        "archetype, vintage, status, size, region, and cross-source "
        "corroboration; this asymmetric scan ranks the investable subset "
        f"({len(universe_rr)} names) by reward/risk. Investable filter drops "
        "PASS_FALSE_FRIEND / ACQUIRED / ARC_DONE / REPEAT_RX and "
        "non-equity placeholder tickers ((state), (private), (delisted)). "
        "PROXY waterfall: bear_loss = max(0.10, 0.65 − 0.30·score), "
        "tilted ×0.75 for A1/A2 (sovereign-anchored floor). "
        "bull = 1.50 + 1.50·score, tilted ×1.10 for H, ×1.15 for F. "
        "Hand-built YAMLs override the proxy with bottom-up numbers."
    )).font = FONT_FOOTNOTE
    ws.cell(row=last_row + 2, column=1).alignment = ALIGN_WRAP
    ws.merge_cells(start_row=last_row + 2, start_column=1,
                   end_row=last_row + 2, end_column=len(headers))
    ws.row_dimensions[last_row + 2].height = 60

    autosize_cols(ws)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 32


def build_coverage_gap(wb: Workbook, universe_rr: list[dict]) -> None:
    """Show the highest-RR names by PROXY that still need a YAML.
    Anti-selection-bias check: which universe names would enter the
    basket if hand-deepened?"""
    ws = wb.create_sheet("Coverage gap (need YAML)")
    ws.sheet_view.showGridLines = False
    ws["A1"] = ("Coverage gap — highest-RR universe names with no "
                "hand-built YAML")
    ws["A1"].font = FONT_SUBHEAD
    ws["A2"] = ("Building YAMLs for these names extends the framework's "
                "comprehensive coverage. Their RR is currently from the "
                "proxy formula; real bottom-up numbers may rank them "
                "higher or lower.")
    ws["A2"].font = FONT_FOOTNOTE
    ws["A2"].alignment = ALIGN_WRAP
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 28

    headers = ["Universe rank", "Ticker", "Name", "Region", "Score",
               "Archetype", "Proxy RR", "Suggested priority"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=4, column=j, value=h)
    style_header_row(ws, 4, len(headers))

    proxy_only = [(i, r) for i, r in enumerate(universe_rr)
                  if r["source"] == "PROXY"][:20]
    for idx, (uni_rank, r) in enumerate(proxy_only):
        row = 5 + idx
        score = float(r["score"])
        priority = ("Tier 1 — hand-deepen ASAP"
                    if score >= 0.80 else
                    "Tier 2 — verify primary docs first"
                    if score >= 0.55 else
                    "Tier 3 — skeleton acceptable")
        cells = [
            uni_rank + 1, r["ticker"], r["name"][:50], r["region"][:22],
            f"{score:.2f}", r["archetype"][:10], f"{float(r['rr']):.1f}×",
            priority,
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(idx % 2 == 1))

    autosize_cols(ws)
    ws.row_dimensions[4].height = 32


def build_old_yaml_only(wb: Workbook, candidates: list[dict],
                       weights: dict[str, dict]) -> None:
    """Original 21-YAML view, retained for reference."""
    ws = wb.create_sheet("YAML-only (hand-built)")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Hand-built Tier 1+2 YAMLs only (legacy view)"
    ws["A1"].font = FONT_SUBHEAD
    ws["A2"] = ("This is the 21 hand-built names ranked by their bottom-up "
                "waterfall numbers. Compare against Executive Summary to "
                "see selection bias.")
    ws["A2"].font = FONT_FOOTNOTE

    headers = ["Rank", "Ticker", "Name", "Region", "Bucket", "Archetype",
               "EV×", "Skew", "Bear loss", "Reward/Risk", "Weight",
               "Dominant factor"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header_row(ws, 4, len(headers))

    for i, c in enumerate(candidates[:15]):
        row = 5 + i
        ticker = c.get("ticker", "?")
        wt = weights.get(str(ticker), {})
        cells = [
            i + 1, ticker, (c.get("name") or "")[:50],
            c.get("jurisdiction") or "",
            c.get("bucket") or "",
            ", ".join(c.get("archetype") or [])
            if isinstance(c.get("archetype"), list) else str(c.get("archetype") or ""),
            f"{c['_ev']:.2f}×", f"{c['_skew']:.1f}",
            f"{c['_downside']*100:.0f}%", f"{c['_rr_ratio']:.1f}×",
            f"{wt.get('weight', 0):.2f}%" if wt else "—",
            (c.get("factors") or {}).get("primary", "") or "",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))
    autosize_cols(ws)
    ws.row_dimensions[4].height = 32


def build_universe(wb: Workbook, candidates: list[dict],
                   weights: dict[str, dict]) -> None:
    ws = wb.create_sheet("Universe (Tier 1+2)")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "All active Tier 1 + 2 candidates"
    ws["A1"].font = FONT_SUBHEAD

    headers = ["Ticker", "Name", "Jurisdiction", "Sector", "Bucket",
               "Archetype", "Tier", "Bear", "Base", "Bull", "EV×",
               "Skew", "Bear loss", "Reward/Risk", "Weight"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    for i, c in enumerate(candidates):
        row = 4 + i
        w = weights.get(str(c.get("ticker")), {})
        archetype = c.get("archetype") or []
        if isinstance(archetype, list):
            archetype = ", ".join(archetype)
        cells = [
            c.get("ticker"),
            (c.get("name") or "")[:60],
            c.get("jurisdiction") or "",
            (c.get("sector") or "")[:35],
            c.get("bucket") or "",
            archetype,
            c.get("tier") or "",
            f"{c['_bear_r']:.2f}×",
            f"{c['_base_r']:.2f}×",
            f"{c['_bull_r']:.2f}×",
            f"{c['_ev']:.2f}×",
            f"{c['_skew']:.1f}",
            f"{c['_downside']*100:.0f}%",
            f"{c['_rr_ratio']:.1f}×",
            f"{w.get('weight', 0):.2f}%" if w else "—",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))

    last_row = 3 + len(candidates)
    ws.conditional_formatting.add(
        f"N4:N{last_row}",
        ColorScaleRule(start_type="min", start_color="F2F4F6",
                       mid_type="percentile", mid_value=50,
                       mid_color="C7E8C0",
                       end_type="max", end_color=GREEN),
    )
    autosize_cols(ws)
    ws.row_dimensions[3].height = 32


def build_waterfall_matrix(wb: Workbook, candidates: list[dict]) -> None:
    ws = wb.create_sheet("Waterfall matrix")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Bear / Base / Bull probability × return multiples"
    ws["A1"].font = FONT_SUBHEAD

    headers = ["Ticker", "Name",
               "Bear P", "Bear R", "Bear P×R",
               "Base P", "Base R", "Base P×R",
               "Bull P", "Bull R", "Bull P×R",
               "EV×", "Bear rationale", "Bull rationale"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    for i, c in enumerate(candidates):
        w = c.get("waterfall", {}) or {}
        bear = w.get("bear", {}) or {}
        base = w.get("base", {}) or {}
        bull = w.get("bull", {}) or {}
        row = 4 + i
        bp, br = float(bear.get("p", 0) or 0), float(bear.get("return_multiple", 0) or 0)
        np_, nr = float(base.get("p", 0) or 0), float(base.get("return_multiple", 0) or 0)
        up, ur = float(bull.get("p", 0) or 0), float(bull.get("return_multiple", 0) or 0)
        cells = [
            c.get("ticker"),
            (c.get("name") or "")[:50],
            f"{bp:.2f}", f"{br:.2f}×", f"{bp*br:.2f}",
            f"{np_:.2f}", f"{nr:.2f}×", f"{np_*nr:.2f}",
            f"{up:.2f}", f"{ur:.2f}×", f"{up*ur:.2f}",
            f"{bp*br + np_*nr + up*ur:.2f}",
            (bear.get("rationale") or "")[:120],
            (bull.get("rationale") or "")[:120],
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))

    autosize_cols(ws)
    # Wider rationale cols
    ws.column_dimensions["M"].width = 55
    ws.column_dimensions["N"].width = 55
    ws.row_dimensions[3].height = 32
    for r in range(4, 4 + len(candidates)):
        ws.row_dimensions[r].height = 45


def build_catalyst_timeline(wb: Workbook, candidates: list[dict]) -> None:
    ws = wb.create_sheet("Catalyst timeline")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Catalysts across the basket, sorted by start of window"
    ws["A1"].font = FONT_SUBHEAD

    headers = ["Window start", "Window end", "Ticker", "Name",
               "Event", "P(favourable)", "Re-rate if yes",
               "Hit if no", "Source"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    rows = []
    for c in candidates:
        for cat in (c.get("catalysts") or []):
            w_ = cat.get("window") or []
            start = str(w_[0]) if len(w_) >= 1 else ""
            end = str(w_[1]) if len(w_) >= 2 else ""
            rerate = cat.get("rerate_if_yes") or []
            hit = cat.get("hit_if_no") or []
            rows.append({
                "start":   start,
                "end":     end,
                "ticker":  c.get("ticker"),
                "name":    (c.get("name") or "")[:40],
                "event":   (cat.get("event") or "")[:90],
                "p":       cat.get("p_favorable", ""),
                "rerate":  f"{rerate[0]}×-{rerate[1]}×" if len(rerate) == 2 else "",
                "hit":     f"{hit[0]} to {hit[1]}" if len(hit) == 2 else "",
                "source":  (cat.get("source") or "")[:50],
            })

    rows.sort(key=lambda r: str(r["start"]))
    for i, r in enumerate(rows):
        row = 4 + i
        ws.cell(row=row, column=1, value=r["start"])
        ws.cell(row=row, column=2, value=r["end"])
        ws.cell(row=row, column=3, value=r["ticker"])
        ws.cell(row=row, column=4, value=r["name"])
        ws.cell(row=row, column=5, value=r["event"])
        ws.cell(row=row, column=6, value=r["p"])
        ws.cell(row=row, column=7, value=r["rerate"])
        ws.cell(row=row, column=8, value=r["hit"])
        ws.cell(row=row, column=9, value=r["source"])
        style_body_band(ws, row, 9, banded=(i % 2 == 1))

    autosize_cols(ws)
    ws.row_dimensions[3].height = 32


def write_kv_block(ws, start_row: int, title: str,
                   pairs: list[tuple[str, str]]) -> int:
    """Two-column label/value block. Returns next free row."""
    ws.cell(row=start_row, column=1, value=title).font = FONT_SUBHEAD
    ws.cell(row=start_row, column=1).border = Border(bottom=MED_NAVY)
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=4)
    r = start_row + 1
    for k, v in pairs:
        ws.cell(row=r, column=1, value=k).font = FONT_BODY_B
        ws.cell(row=r, column=1).alignment = ALIGN_L
        ws.cell(row=r, column=2, value=v).font = FONT_BODY
        ws.cell(row=r, column=2).alignment = ALIGN_WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    return r + 1


def build_pick_detail(wb: Workbook, c: dict,
                      weights: dict[str, dict]) -> None:
    ticker = str(c.get("ticker", "?"))
    sheet_name = f"{ticker}"[:31].replace("/", "_")
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 60

    # Title
    ws["A1"] = c.get("name") or ticker
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 30
    ws["A2"] = (f"{ticker}  ·  {c.get('jurisdiction', '')}  ·  "
                f"Bucket {c.get('bucket', '')}  ·  "
                f"Archetype "
                f"{', '.join(c.get('archetype', []) or []) or c.get('archetype')}  ·  "
                f"Tier {c.get('tier', '')}")
    ws["A2"].font = FONT_SUBTITLE

    # Risk-reward snapshot
    w = weights.get(ticker, {})
    pairs = [
        ("Expected EV×",      f"{c['_ev']:.2f}×"),
        ("Bear / Base / Bull return",
         f"{c['_bear_r']:.2f}× / {c['_base_r']:.2f}× / {c['_bull_r']:.2f}×"),
        ("Bear loss",         f"{c['_downside']*100:.0f}%"),
        ("Skew (upside/downside)", f"{c['_skew']:.1f}×"),
        ("Reward/Risk ratio", f"{c['_rr_ratio']:.1f}×"),
        ("Portfolio weight",  f"{w.get('weight', 0):.2f}%" if w else "—"),
        ("Cluster",           w.get("cluster", "—") if w else "—"),
        ("Primary factor",
         (c.get("factors") or {}).get("primary", "—")),
    ]
    next_row = write_kv_block(ws, 4, "Risk-reward snapshot", pairs)

    # Deal mechanic
    deal = c.get("deal") or {}
    next_row = write_kv_block(ws, next_row, "Deal", [
        ("Date",      str(deal.get("date", ""))),
        ("Mechanic",  (deal.get("mechanic") or "")[:1200]),
    ])

    # Scorecard
    sc = c.get("scorecard") or {}
    sc_pairs = [(k.replace("_", " ").title(),
                 "—" if v is None else str(v)) for k, v in sc.items()]
    if sc_pairs:
        next_row = write_kv_block(ws, next_row, "Scorecard", sc_pairs)

    # Catalysts
    ws.cell(row=next_row, column=1, value="Catalysts").font = FONT_SUBHEAD
    ws.cell(row=next_row, column=1).border = Border(bottom=MED_NAVY)
    ws.merge_cells(start_row=next_row, start_column=1,
                   end_row=next_row, end_column=4)
    next_row += 1
    cat_hdr = ["Event", "Window", "P(yes)", "Re-rate if yes / hit if no"]
    for j, h in enumerate(cat_hdr, 1):
        ws.cell(row=next_row, column=j, value=h)
    style_header_row(ws, next_row, len(cat_hdr))
    next_row += 1
    cats = c.get("catalysts") or []
    for idx, cat in enumerate(cats):
        w_ = cat.get("window") or []
        window_s = (f"{w_[0]} → {w_[1]}" if len(w_) >= 2 else
                    str(w_[0]) if len(w_) == 1 else "—")
        rerate = cat.get("rerate_if_yes") or []
        hit = cat.get("hit_if_no") or []
        ws.cell(row=next_row, column=1, value=(cat.get("event") or "")[:120])
        ws.cell(row=next_row, column=2, value=window_s)
        ws.cell(row=next_row, column=3, value=str(cat.get("p_favorable", "")))
        ws.cell(row=next_row, column=4, value=(
            f"Up {rerate[0]}-{rerate[1]}× / Down {hit[0]}-{hit[1]}"
            if rerate and hit else "—"))
        style_body_band(ws, next_row, 4, banded=(idx % 2 == 1))
        ws.cell(row=next_row, column=1).alignment = ALIGN_WRAP
        ws.row_dimensions[next_row].height = 32
        next_row += 1
    next_row += 1

    # Triangulation
    tri = c.get("triangulation") or {}
    next_row = write_kv_block(ws, next_row, "Triangulation (3-leg test)", [
        ("Leg 1 — Valuation",      str(tri.get("leg1_valuation", "—"))),
        ("Leg 2 — Game theory",    str(tri.get("leg2_game_theory", "—"))),
        ("Leg 3 — Revealed preference", str(tri.get("leg3_revealed_pref", "—"))),
        ("Notes",                   (tri.get("notes") or "—")[:1200]),
    ])

    # Anchor / red flags / kill / pre-mortem / rationale
    anchor = c.get("anchor") or {}
    parties = anchor.get("parties") or []
    next_row = write_kv_block(ws, next_row, "Anchor", [
        ("Parties",       ", ".join(parties) if parties else "—"),
        ("Stake (% est.)", str(anchor.get("stake_pct", "—"))),
    ])

    rf = c.get("red_flags") or {}
    active_rf = [k for k, v in rf.items() if v is True]
    watch = rf.get("_watch", "")
    next_row = write_kv_block(ws, next_row, "Red flags", [
        ("Triggered",  ", ".join(active_rf) if active_rf else "None of 11 active"),
        ("Watch list", watch or "—"),
    ])

    kc = c.get("kill_criteria") or []
    next_row = write_kv_block(ws, next_row, "Kill criteria", [
        (f"{i+1}.", v) for i, v in enumerate(kc)
    ])

    pm = c.get("pre_mortem") or "—"
    next_row = write_kv_block(ws, next_row, "Pre-mortem", [
        ("", pm[:2500]),
    ])

    tr = c.get("tier_rationale") or ""
    if tr:
        next_row = write_kv_block(ws, next_row, "Tier rationale", [
            ("", tr[:2500]),
        ])


def build_portfolio_sizing(wb: Workbook, weights: dict[str, dict]) -> None:
    ws = wb.create_sheet("Portfolio sizing")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Risk-budgeted portfolio weights (Kelly → corr. haircut → cluster cap)"
    ws["A1"].font = FONT_SUBHEAD

    headers = ["Ticker", "Cluster", "Raw Kelly",
               "After corr. haircut", "After cluster cap",
               "EV× on weight", "Weight × EV (bps)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))

    items = sorted(weights.values(), key=lambda x: -x["weight"])
    for i, w in enumerate(items):
        row = 4 + i
        cells = [
            w["ticker"], w["cluster"], w["raw_kelly"],
            w["corr_haircut"], w["capped"],
            f"{w['ev_mult']:.2f}×", f"{w['contrib_bps']:.1f}",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))

    last = 3 + len(items)
    total_w = sum(w["weight"] for w in items)
    total_b = sum(w["contrib_bps"] for w in items)
    ws.cell(row=last + 1, column=1, value="Total invested")
    ws.cell(row=last + 1, column=5, value=f"{total_w:.2f}%")
    ws.cell(row=last + 1, column=7, value=f"{total_b:.1f}")
    for c in (1, 5, 7):
        ws.cell(row=last + 1, column=c).font = FONT_BODY_B
        ws.cell(row=last + 1, column=c).border = Border(top=MED_NAVY)
    ws.cell(row=last + 2, column=1, value="Cash sleeve")
    ws.cell(row=last + 2, column=5, value=f"{100 - total_w:.2f}%")
    ws.cell(row=last + 2, column=1).font = FONT_BODY_B
    ws.cell(row=last + 2, column=5).font = FONT_BODY_B

    autosize_cols(ws)
    ws.row_dimensions[3].height = 32


def build_all_names(wb: Workbook, universe_rr: list[dict]) -> None:
    """Every ranked name (not just the top 100), with all asymmetry
    lenses, so nothing is hidden below a cutoff."""
    ws = wb.create_sheet("All names")
    ws.sheet_view.showGridLines = False
    ws["A1"] = (f"All {len(universe_rr)} ranked names · reward/risk + "
                "alternative asymmetry lenses")
    ws["A1"].font = FONT_SUBHEAD
    headers = ["Rank", "Src", "Ticker", "Name", "Region", "Arch",
               "Bear", "EV×", "Rew/Risk", "Skew", "Downside",
               "Conv.", "Hard", "Val", "Composite"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j, value=h)
    style_header_row(ws, 3, len(headers))
    for i, r in enumerate(universe_rr):
        row = 4 + i

        def fnum(k, d=0.0):
            try:
                return float(r.get(k) or d)
            except (TypeError, ValueError):
                return d
        cells = [
            i + 1, r["source"], r["ticker"], r["name"][:44],
            r["region"][:20], r["archetype"][:8],
            f"{fnum('bear_loss')*100:.0f}%", f"{fnum('ev'):.2f}×",
            f"{fnum('rr'):.1f}×", f"{fnum('skew'):.1f}",
            f"{fnum('downside_prot'):.2f}", f"{fnum('conviction'):.2f}",
            f"{fnum('hardness'):.2f}",
            (f"{fnum('val_score'):.2f}" if r.get("val_score") else "—"),
            f"{fnum('composite'):.3f}",
        ]
        for j, v in enumerate(cells, 1):
            ws.cell(row=row, column=j, value=v)
        style_body_band(ws, row, len(headers), banded=(i % 2 == 1))
    ws.freeze_panes = "A4"
    last = 3 + len(universe_rr)
    for col in ("I", "O"):   # heat Reward/Risk + Composite
        ws.conditional_formatting.add(
            f"{col}4:{col}{last}",
            ColorScaleRule(start_type="min", start_color="F2F4F6",
                           mid_type="percentile", mid_value=50,
                           mid_color="C7E8C0", end_type="max",
                           end_color=GREEN))
    autosize_cols(ws)
    ws.row_dimensions[3].height = 30


def build_lenses(wb: Workbook, universe_rr: list[dict]) -> None:
    """Side-by-side leaders under each asymmetry lens, so the user can see
    WHICH lens each name wins on — different lenses surface different
    opportunities (a floored microcap vs a convex sovereign bet vs a
    cheap-on-valuation restructuring vs a heavily-corroborated event)."""
    ws = wb.create_sheet("Asymmetry lenses")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Top 20 under each asymmetry lens"
    ws["A1"].font = FONT_SUBHEAD
    ws["A2"] = ("Different lenses find different asymmetry. Reward/Risk = "
                "(EV−1)/bear. Skew = pure upside/downside. Downside = "
                "capital floor (1−bear). Conviction = weighted cross-source "
                "signal. Valuation = cheapness (net-cash/EV-EBITDA/P-B/"
                "discount, where data exists). Composite blends them.")
    ws["A2"].font = FONT_FOOTNOTE
    ws["A2"].alignment = ALIGN_WRAP
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 42

    def fnum(r, k):
        try:
            return float(r.get(k) or 0)
        except (TypeError, ValueError):
            return 0.0

    lenses = [
        ("Reward/Risk", "rr", "{:.1f}×"),
        ("Skew (up/down)", "skew", "{:.1f}"),
        ("Downside floor", "downside_prot", "{:.2f}"),
        ("Signal conviction", "conviction", "{:.2f}"),
        ("Valuation cheapness", "val_score", "{:.2f}"),
        ("Composite (blended)", "composite", "{:.3f}"),
    ]
    col = 1
    for label, key, fmt in lenses:
        pool = [r for r in universe_rr if fnum(r, key) > 0]
        ranked = sorted(pool, key=lambda r: -fnum(r, key))[:20]
        ws.cell(row=4, column=col, value=label).font = FONT_HEADER
        ws.cell(row=4, column=col).fill = FILL_HEADER
        ws.cell(row=4, column=col + 1, value="val").font = FONT_HEADER
        ws.cell(row=4, column=col + 1).fill = FILL_HEADER
        for i, r in enumerate(ranked):
            rr = 5 + i
            tk = r["ticker"].split(":")[-1]
            ws.cell(row=rr, column=col,
                    value=f"{tk} · {r['name'][:20]}").font = FONT_BODY
            ws.cell(row=rr, column=col + 1,
                    value=fmt.format(fnum(r, key))).font = FONT_BODY
            ws.cell(row=rr, column=col + 1).alignment = ALIGN_R
        col += 3
    autosize_cols(ws, max_w=34)
    ws.row_dimensions[4].height = 26


def _parse_md_tables(path: Path) -> list[tuple[list[str], list[list[str]]]]:
    """Return every GitHub-markdown table in a file as (headers, rows).
    Lets the workbook fold in the post-reorg watchlists without re-running
    their network-bound screens — the .md files are the durable artifact."""
    tables: list[tuple[list[str], list[list[str]]]] = []
    if not path.exists():
        return tables
    header: list[str] | None = None
    rows: list[list[str]] = []
    for line in path.read_text().splitlines():
        s = line.strip()
        is_row = s.startswith("|") and s.endswith("|")
        if is_row and set(s) <= set("|-: "):   # separator row
            continue
        if is_row:
            cells = [c.strip() for c in s.strip("|").split("|")]
            if header is None:
                header = cells
            else:
                rows.append(cells)
        else:
            if header is not None:
                tables.append((header, rows))
            header, rows = None, []
    if header is not None:
        tables.append((header, rows))
    return tables


def _render_table_sheet(wb: Workbook, sheet_name: str, title: str,
                        subtitle: str, headers: list[str],
                        rows: list[list[str]], heat_col: int | None = None,
                        subtitle2: str = "") -> None:
    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = FONT_SUBHEAD
    r0 = 2
    for sub in (subtitle, subtitle2):
        if sub:
            ws.cell(row=r0, column=1, value=sub).font = FONT_FOOTNOTE
            ws.cell(row=r0, column=1).alignment = ALIGN_WRAP
            ws.merge_cells(start_row=r0, start_column=1, end_row=r0,
                           end_column=max(4, len(headers)))
            ws.row_dimensions[r0].height = 30
            r0 += 1
    hdr_row = r0 + 1
    for j, h in enumerate(headers, 1):
        ws.cell(row=hdr_row, column=j, value=h)
    style_header_row(ws, hdr_row, len(headers))
    for i, row in enumerate(rows):
        rr = hdr_row + 1 + i
        for j, v in enumerate(row, 1):
            ws.cell(row=rr, column=j, value=v)
        style_body_band(ws, rr, len(headers), banded=(i % 2 == 1))
    ws.freeze_panes = f"A{hdr_row + 1}"
    autosize_cols(ws)
    ws.row_dimensions[hdr_row].height = 30


def build_postreorg_listed(wb: Workbook) -> None:
    """Listed-equity reorganization watchlist — the tradable post-reorg
    slice on the six-question screen. Folded in from its durable .md."""
    tables = _parse_md_tables(LISTED_EQUITY_MD)
    if not tables:
        return
    # main watchlist = the table whose header starts with Name|Ticker|Conf
    main = next((t for t in tables if len(t[0]) >= 6
                 and t[0][0].lower() == "name" and "conf" in
                 " ".join(t[0]).lower()), tables[0])
    headers, rows = main
    _render_table_sheet(
        wb, "Post-reorg (listed common)",
        "Listed-equity reorganization watchlist — six-question sweet-spot screen",
        "Exchange-listed common only (no claims, DIP paper, rights, "
        "backstops, creditor-only securities). Q's: Listed · Unnatural "
        "owners (live forced-seller overhang) · Repaired B/S · Overstated "
        "count/debt · Catalyst · Quality (EBIT-yield). Conf ✓ = filer's own "
        "emergence verified from the filing (or PACER); ~ = unverified/kept.",
        headers, rows,
        subtitle2="Source: output/listed_equity_watchlist.md (make "
        "listed-equity-screen). Emergence dates and filer-emergence "
        "verification are read from each SEC filing.")
    # set-aside table (unconfirmed filer emergence), if present
    aside = next((t for t in tables if t is not main and t[0]
                  and t[0][0].lower() == "name"), None)
    if aside and aside[1]:
        h2, r2 = aside
        ws = wb["Post-reorg (listed common)"]
        start = ws.max_row + 3
        ws.cell(row=start, column=1,
                value="Set aside — filer's own emergence unconfirmed "
                "(verify; not scored, not dropped)").font = FONT_SUBHEAD
        for j, h in enumerate(h2, 1):
            ws.cell(row=start + 1, column=j, value=h)
        style_header_row(ws, start + 1, len(h2))
        for i, row in enumerate(r2):
            for j, v in enumerate(row, 1):
                ws.cell(row=start + 2 + i, column=j, value=v)
            style_body_band(ws, start + 2 + i, len(h2), banded=(i % 2 == 1))
        autosize_cols(ws)


def build_postreorg_assembly(wb: Workbook) -> None:
    """Post-reorg assembly scorecard — Verdad EBIT-yield + Chapter-22 veto
    across the whole fresh-start cohort. Folded in from its durable .md."""
    tables = _parse_md_tables(POSTREORG_MD)
    main = next((t for t in tables if t[0] and t[0][0].lower() == "name"),
                None)
    if not main:
        return
    headers, rows = main
    _render_table_sheet(
        wb, "Post-reorg assembly",
        "Post-reorg assembly scorecard — Verdad EBIT-yield + Chapter-22 veto",
        "The fresh-start cohort graded on the two highest-signal filters: "
        "EBIT/EV > 20% → +61% avg 2yr (PRIORITY), < 0% → −21% (AVOID); and "
        "a Chapter-22 auto-veto (re-filed after emergence = fixed the "
        "balance sheet, not the business).",
        headers, rows,
        subtitle2="Source: output/postreorg_watchlist.md (make "
        "postreorg-score).")


def build_methodology(wb: Workbook) -> None:
    ws = wb.create_sheet("Methodology")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Methodology · cyclepapa framework"
    ws["A1"].font = FONT_SUBHEAD
    ws.column_dimensions["A"].width = 130

    paragraphs = [
        ("Buckets",
         "A = listed common as primary instrument. B = anchor instrument "
         "(rights, convertibles, MCB cascades). C = legacy cancelled / new "
         "common issued post-restructuring."),
        ("Archetypes",
         "A1 sovereign-strategic. A2 sovereign industrial policy. B "
         "convertible / MCB. C LME (liability management exchange). D "
         "customer-strategic. E court-supervised. F MCB cascade. "
         "G regulator-forced. H governance reset."),
        ("Waterfall scenarios",
         "Each YAML carries three named scenarios with probability and "
         "return multiple. The framework's expected EV× is the dot "
         "product of probability and return multiple vectors."),
        ("Reward/Risk ratio",
         "Defined as (EV× − 1) ÷ Bear-case loss. Captures the asymmetry "
         "of the expected outcome relative to the downside the position "
         "is exposed to. Names with EV× < 1 receive negative ratios; the "
         "workbook ranks by this ratio."),
        ("Triangulation (3 legs)",
         "Leg 1: independent valuation (PF EV/EBITDA, SOTP, or DCF). "
         "Leg 2: game-theoretic check on the anchor/insider stack. "
         "Leg 3: revealed preference — what did the cap stack do?"),
        ("Portfolio sizing",
         "¼-Kelly per name, then correlation haircut (cosine over factor "
         "vectors), then cluster cap at 50% of the cluster's raw Kelly "
         "sum, then a global 60% invested cap forces a 40% cash sleeve. "
         "Kelly haircut is the binding constraint."),
        ("Discovery pipeline",
         "Pollers — EDGAR (US), NSM (UK), SEDAR+ (CA), TDnet (JP) — "
         "write hits to data/inbox/. inbox_promote.py promotes them "
         "into universe.md by tier+sub-query, dedup against existing "
         "rows. universe_screen.py scores. yaml_skeleton.py builds "
         "Tier-3 skeletons. Hand-deepening promotes to Tier 1+2."),
        ("Discipline",
         "Every file representing work is tracked, committed, pushed. "
         "src/audit.py blocks runs on durability violations. Each "
         "scorecard input carries a source tag (reported/verified/"
         "estimated) so the verification gap is visible."),
    ]
    r = 3
    for h, body in paragraphs:
        ws.cell(row=r, column=1, value=h).font = FONT_BODY_B
        r += 1
        ws.cell(row=r, column=1, value=body).font = FONT_BODY
        ws.cell(row=r, column=1).alignment = ALIGN_WRAP
        ws.row_dimensions[r].height = max(28, 14 * (1 + len(body) // 90))
        r += 2


def main() -> int:
    candidates = load_candidates()
    if not candidates:
        print("No candidates loaded — check data/candidates/", file=sys.stderr)
        return 1
    weights = load_portfolio_weights()
    universe_rr = load_universe_rr()
    print(f"Loaded {len(candidates)} hand-built Tier 1+2 YAMLs, "
          f"{len(weights)} portfolio rows, "
          f"{len(universe_rr)} universe-wide rows.")
    if not universe_rr:
        print("  ! universe_risk_reward.csv missing — run "
              "`python3 -m src.universe_risk_reward` first.",
              file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    build_cover(wb, candidates, weights)
    build_executive_summary(wb, candidates, weights, universe_rr)
    build_lenses(wb, universe_rr)
    build_all_names(wb, universe_rr)
    build_coverage_gap(wb, universe_rr)
    build_old_yaml_only(wb, candidates, weights)
    build_universe(wb, candidates, weights)
    build_waterfall_matrix(wb, candidates)
    build_catalyst_timeline(wb, candidates)
    build_postreorg_listed(wb)
    build_postreorg_assembly(wb)
    build_portfolio_sizing(wb, weights)

    # Detail sheets: the top-10 hand-built YAMLs by their own bottom-up
    # reward/risk ratio. Anchoring on the candidates list (not the
    # universe rank) keeps detail sheets stable as new PROXY candidates
    # flood the universe-wide ranking and push REAL names below rank 30.
    yaml_by_ticker = {}
    for c in candidates:
        t = str(c.get("ticker", ""))
        yaml_by_ticker[t.upper()] = c
        stem = re.sub(r"[^A-Za-z0-9-]", "", t.split(":")[-1]).upper()
        yaml_by_ticker[stem] = c
    detail_added = set()
    # candidates is already sorted by _rr_ratio descending in
    # load_candidates(); take the top 10 distinct.
    for c in candidates:
        stem = re.sub(r"[^A-Za-z0-9-]", "",
                      str(c.get("ticker", "")).split(":")[-1]).upper()
        if not stem or stem in detail_added:
            continue
        build_pick_detail(wb, c, weights)
        detail_added.add(stem)
        if len(detail_added) >= 10:
            break

    build_methodology(wb)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
