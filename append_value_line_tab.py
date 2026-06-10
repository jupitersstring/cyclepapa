"""Add value_line tab to screener_report.xlsx.

Computes classic Value-Line "Bargain Basement" style deep-value metrics
from cached info_metrics + derived fields, with a sector-relative composite
that replaces the original binary 7-point score (Issue 5).

  - Net cash per share = (totalCash - totalDebt) / shares
  - Net cash as % of market cap (Graham net-net proxy)
  - P/E (trailing), P/B, P/S, P/FCF
  - FCF yield (= 1 / P/FCF × 100)
  - EV/EBITDA, EV/Sales, EV/GrossProfit
  - Current ratio, Quick ratio, Debt-to-Equity
  - ROE, ROA, gross/operating/profit margins
  - 5y / 1y price performance

Scoring:
  * bargain_score (LEGACY): binary 0-7 sum of fixed thresholds (P/B<1, P/E<12,
    CR>2, NetCash>0, FCFyld>8%, D/E<50, GM>30%). Same as before, kept for
    backward compatibility but it conflates sectors (e.g. a 28% gross margin
    is great for utilities, mediocre for software).
  * bargain_pct_score (NEW): mean of sector-percentile ranks across the same
    seven metrics, on a 0-100 scale. A score of 80 means: in this ticker's
    sector, it sits in the top fifth across the basket of deep-value metrics.
    Direction: lower-is-better for P/B, P/E, D/E; higher-is-better for the
    other four. Missing metrics are skipped (composite averages only the
    available percentiles).

Output tab is sortable by bargain_pct_score (descending).
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook

CACHE = Path('.cache/yf')
XLSX = Path('screener_report.xlsx')

def _safe(t): return ''.join(c if c.isalnum() or c in '-_' else '_' for c in t)


def price_perf_5y(tk):
    p = CACHE / f'{_safe(tk)}__price.parquet'
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        s = pd.to_numeric(df['Close'], errors='coerce').dropna()
        if len(s) < 252: return None
        return float((s.iloc[-1]/s.iloc[0] - 1) * 100)
    except Exception: return None


def price_perf_1y(tk):
    p = CACHE / f'{_safe(tk)}__price.parquet'
    if not p.exists(): return None
    try:
        df = pd.read_parquet(p)
        s = pd.to_numeric(df['Close'], errors='coerce').dropna()
        if len(s) < 252: return None
        return float((s.iloc[-1]/s.iloc[-252] - 1) * 100)
    except Exception: return None


def load_info(tk):
    p = CACHE / f'{_safe(tk)}__info_metrics.parquet'
    if not p.exists(): return None
    try:
        d = pd.read_parquet(p)
        return d.iloc[0].to_dict() if not d.empty else None
    except Exception: return None


def load_meta():
    """Load name/sector/industry/country from universe CSVs."""
    meta = {}
    for f in ['universe_us_wide.csv','universe_expanded.csv','universe_wider.csv',
              'universe_eu.csv','universe_eu_extra.csv','universe_canada.csv',
              'universe_us_large_mega.csv','universe_japan.csv','universe_korea.csv',
              'universe_hongkong.csv','universe_australia.csv']:
        p = Path(f)
        if not p.exists(): continue
        try:
            d = pd.read_csv(p, usecols=lambda c: c in ['symbol','name','sector','industry','country'], low_memory=False)
            for _, r in d.iterrows():
                tk = r.get('symbol')
                if not isinstance(tk, str): continue
                k = tk.upper()
                if k in meta: continue
                meta[k] = {f: r.get(f) for f in ['name','sector','industry','country'] if f in d.columns}
                # also safe form
                safe_k = k.replace('.','_')
                if safe_k != k and safe_k not in meta: meta[safe_k] = meta[k]
        except Exception: pass
    return meta


META = load_meta()


def compute_value_line(tk):
    i = load_info(tk)
    if i is None: return None

    mcap   = i.get('marketCap')
    if mcap is None or mcap < 50e6: return None
    if tk.upper().endswith(('_NS','_BO','.NS','.BO')): return None  # exclude India

    price   = i.get('currentPrice')
    shares  = i.get('sharesOutstanding')
    cash    = i.get('totalCash')
    debt    = i.get('totalDebt')
    book    = i.get('bookValue')
    pe      = i.get('trailingPE')
    pb      = i.get('priceToBook')
    ps      = i.get('priceToSalesTrailing12Months')
    ev_ebd  = i.get('enterpriseToEbitda')
    ev_sale = i.get('enterpriseToRevenue')
    eps     = i.get('trailingEps')
    fcf     = i.get('freeCashflow')
    revenue = i.get('totalRevenue')
    de      = i.get('debtToEquity')
    curr_r  = i.get('currentRatio')
    quick_r = i.get('quickRatio')
    roe     = i.get('returnOnEquity')
    roa     = i.get('returnOnAssets')
    gross_m = i.get('grossMargins')
    op_m    = i.get('operatingMargins')
    prof_m  = i.get('profitMargins')

    # Derived
    net_cash = None
    if cash is not None and debt is not None:
        net_cash = float(cash) - float(debt)
    net_cash_per_share = None
    if net_cash is not None and shares and shares > 0:
        net_cash_per_share = net_cash / float(shares)
    net_cash_pct_mcap = None
    if net_cash is not None and mcap > 0:
        net_cash_pct_mcap = net_cash / mcap * 100

    p_fcf = None
    fcf_yield = None
    if fcf and fcf != 0:
        p_fcf = mcap / float(fcf) if float(fcf) > 0 else None
        fcf_yield = float(fcf) / mcap * 100 if mcap > 0 else None

    # Bargain score (0-7)
    score = 0
    score_breakdown = []
    def to_num(x):
        try:
            v = float(x)
            return v if not pd.isna(v) else None
        except (TypeError, ValueError): return None
    pb_n = to_num(pb); pe_n = to_num(pe); curr_n = to_num(curr_r); de_n = to_num(de)
    gm_n = to_num(gross_m); fy_n = to_num(fcf_yield); nc_n = to_num(net_cash)
    def add(cond, label):
        nonlocal score
        if cond:
            score += 1
            score_breakdown.append(label)
    add(pb_n is not None and pb_n < 1.0, 'P/B<1')
    add(pe_n is not None and 0 < pe_n < 12, 'P/E<12')
    add(curr_n is not None and curr_n > 2, 'CR>2')
    add(nc_n is not None and nc_n > 0, 'NetCash>0')
    add(fy_n is not None and fy_n > 8, 'FCFyld>8%')
    add(de_n is not None and de_n < 50, 'D/E<50')
    add(gm_n is not None and gm_n > 0.30, 'GM>30%')

    meta = META.get(tk.upper(), {})

    return {
        'ticker': tk,
        'name': meta.get('name'),
        'sector': meta.get('sector'),
        'industry': meta.get('industry'),
        'country': meta.get('country'),
        'currentPrice': price,
        'market_cap': mcap,
        'shares_out': shares,
        # Value Line core: working capital / net cash measures
        'net_cash':              net_cash,
        'net_cash_per_share':    net_cash_per_share,
        'net_cash_pct_mcap':     net_cash_pct_mcap,
        'bookValue':             book,
        # Liquidity / leverage
        'currentRatio':          curr_r,
        'quickRatio':            quick_r,
        'debtToEquity':          de,
        # Earnings / FCF
        'trailingEps':           eps,
        'totalRevenue':          revenue,
        'freeCashflow':          fcf,
        # Valuation multiples
        'trailingPE':            pe,
        'priceToBook':           pb,
        'priceToSales':          ps,
        'priceToFCF':            p_fcf,
        'fcfYield_pct':          fcf_yield,
        'enterpriseToEbitda':    ev_ebd,
        'enterpriseToRevenue':   ev_sale,
        # Margins / returns
        'returnOnEquity':        roe,
        'returnOnAssets':        roa,
        'grossMargin':           gross_m,
        'operatingMargin':       op_m,
        'profitMargin':          prof_m,
        # Performance
        'perf_1y_pct':           price_perf_1y(tk),
        'perf_5y_pct':           price_perf_5y(tk),
        # Bargain composite
        'bargain_score':         score,
        'bargain_passed':        ';'.join(score_breakdown),
    }


def _sector_pct(series, sector, direction):
    """Per-row percentile rank within sector (0-100). NaN where invalid.

    direction='hi' means higher value scores higher percentile; 'lo' inverts.
    Tickers with NaN metric are excluded from the rank and get NaN percentile.
    """
    s = pd.to_numeric(series, errors='coerce')
    # rank(pct=True) gives the cumulative fraction (higher value -> closer to 1)
    pct = s.groupby(sector).rank(pct=True, method='average') * 100
    if direction == 'lo':
        pct = 100 - pct
    return pct


def main():
    info_files = sorted(CACHE.glob('*__info_metrics.parquet'))
    print(f'Scanning {len(info_files)} tickers...')
    rows = []
    for i, p in enumerate(info_files):
        tk = p.name.split('__')[0]
        rec = compute_value_line(tk)
        if rec is None: continue
        rows.append(rec)
        if (i+1) % 2000 == 0: print(f'  {i+1}/{len(info_files)} kept={len(rows)}')

    df = pd.DataFrame(rows)
    print(f'\nTotal value-line scored: {len(df)}')

    # ---- SECTOR-PERCENTILE BARGAIN SCORE (Issue 5) ----
    # Replaces the binary 0-7 thresholds with sector-relative percentile ranks.
    # 'sector' is filled with '__unknown__' so those tickers form their own
    # comparison group instead of being NaN.
    df['_sector_rank'] = df['sector'].fillna('__unknown__')

    # Only positive-value multiples are economically meaningful for "cheap":
    #   P/B < 0  => negative book value (impaired)
    #   P/E < 0  => loss-making (not cheap)
    #   D/E < 0  => negative shareholder equity (bankrupt)
    # Strip these to NaN so they don't poison the "lower is better" rank.
    pb_valid = pd.to_numeric(df['priceToBook'],    errors='coerce').where(lambda s: s > 0)
    pe_valid = pd.to_numeric(df['trailingPE'],     errors='coerce').where(lambda s: s > 0)
    de_valid = pd.to_numeric(df['debtToEquity'],   errors='coerce').where(lambda s: s >= 0)
    cr_valid = pd.to_numeric(df['currentRatio'],   errors='coerce').where(lambda s: s > 0)

    # The seven sector-percentile components — same metrics as legacy score
    df['_pct_pb']     = _sector_pct(pb_valid,                  df['_sector_rank'], 'lo')
    df['_pct_pe']     = _sector_pct(pe_valid,                  df['_sector_rank'], 'lo')
    df['_pct_cr']     = _sector_pct(cr_valid,                  df['_sector_rank'], 'hi')
    df['_pct_ncap']   = _sector_pct(df['net_cash_pct_mcap'],   df['_sector_rank'], 'hi')
    df['_pct_fcfy']   = _sector_pct(df['fcfYield_pct'],        df['_sector_rank'], 'hi')
    df['_pct_de']     = _sector_pct(de_valid,                  df['_sector_rank'], 'lo')
    df['_pct_gm']     = _sector_pct(df['grossMargin'],         df['_sector_rank'], 'hi')

    pct_cols = ['_pct_pb','_pct_pe','_pct_cr','_pct_ncap','_pct_fcfy','_pct_de','_pct_gm']
    # Require at least 4 of 7 valid components — one or two non-null metrics
    # is too noisy to call a sector-relative composite.
    n_valid = df[pct_cols].notna().sum(axis=1)
    df['bargain_pct_score'] = df[pct_cols].mean(axis=1, skipna=True).round(1)
    df.loc[n_valid < 4, 'bargain_pct_score'] = np.nan
    df['bargain_pct_n'] = n_valid.astype(int)
    # Also surface per-component percentiles (rounded) for inspection
    rename = {'_pct_pb':'sec_pct_PB','_pct_pe':'sec_pct_PE','_pct_cr':'sec_pct_CR',
              '_pct_ncap':'sec_pct_NetCash','_pct_fcfy':'sec_pct_FCFyld',
              '_pct_de':'sec_pct_DE','_pct_gm':'sec_pct_GM'}
    for old, new in rename.items():
        df[new] = df[old].round(1)
    df = df.drop(columns=pct_cols + ['_sector_rank'])

    # Sort by NEW sector-percentile composite, then by legacy bargain_score, then P/B
    df = df.sort_values(['bargain_pct_score','bargain_score','priceToBook'],
                        ascending=[False, False, True], na_position='last')

    with pd.ExcelWriter(XLSX, engine='openpyxl', mode='a', if_sheet_exists='replace') as xw:
        df.to_excel(xw, sheet_name='value_line', index=False)
        ws = xw.sheets['value_line']
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 28
        for c in 'EFGHIJKLMNOPQRSTUVWXYZ': ws.column_dimensions[c].width = 14
        ws.freeze_panes = 'C2'

    # Move tab to a sensible position (after forensic_audit / cheap_on_growth area)
    wb = load_workbook(XLSX)
    if 'value_line' in wb.sheetnames:
        target_idx = 4  # after forensic_audit/cheap_on_growth/clean_topline area
        cur = wb.sheetnames.index('value_line')
        wb.move_sheet('value_line', offset=target_idx - cur)
        wb.save(XLSX)

    print(f'\nWrote value_line tab with {len(df)} rows')
    print('Top 20 by bargain_pct_score (sector-relative Value-Line composite):')
    cols_show = ['ticker','name','sector','bargain_pct_score','bargain_score',
                 'priceToBook','trailingPE','currentRatio','net_cash_pct_mcap',
                 'fcfYield_pct','perf_1y_pct']
    cols_show = [c for c in cols_show if c in df.columns]
    head = df.head(20)[cols_show].copy()
    for c in cols_show:
        if c not in ('ticker','name','sector','bargain_score'):
            head[c] = pd.to_numeric(head[c], errors='coerce').round(1)
    print(head.to_string(index=False))


if __name__ == '__main__':
    main()
