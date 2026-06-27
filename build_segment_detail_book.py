"""Per-name segment-detail workbook — what the binary archetypes hide.

Shows for every name with EDGAR segment coverage:
  - Top 3 segments by % of revenue (with names + share)
  - Largest segment $ revenue + share
  - Fastest-growing segment + YoY %
  - Geographic mix (top regions + shares)
  - Number of business segments, geographic regions, product lines
  - HHI of segment revenue concentration

Output: segment_detail_book.xlsx
  Cover           headline tiles + archetype index
  All Segments    every name with segment data, ranked by ETA
  Plus 4 per-archetype drill-downs:
  Diversified     names with 4+ segments + HHI <= 0.40
  Concentrated    HHI >= 0.70 or top segment >= 70%
  Global          4+ geographies
  Fastest         single segment growing > 25% YoY
"""
from __future__ import annotations
import argparse
import os
import sys

import pandas as pd

import build_harvard_workbook as bhw
from build_harvard_workbook import (
    Workbook,
    INK, DARK_GREY, MUTED, RULE, LIGHT_GREY, PALE_GREY, WHITE,
    FONT_NAME, FONT_SIZE,
    _font, _border, _align,
    _section_rule, _verdict_badge,
    _write_money, _write_pct, _write_ratio, _write_score, _write_int,
    _NUM_ALIGN_RIGHT, _NUM_ALIGN_CENTER, _TXT_ALIGN_LEFT,
)
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_data():
    """Merge asymmetry_global + segment_signals + valuation columns."""
    import glob
    df = pd.read_csv('asymmetry_global.csv').drop_duplicates('symbol')
    sig = pd.read_csv('edgar_segment_signals.csv')
    df = df[df['symbol'].isin(sig['symbol'])].copy()
    df = df.drop(columns=[c for c in df.columns if c.endswith('_arch')])
    df = df.merge(sig, on='symbol', how='left')

    # Pull valuation ratios from per-country yartseva CSVs (master is sparse)
    val_cols = ['symbol', 'ev_ebitda', 'p_e', 'pb', 'fcf_yield', 'roce',
                'net_debt_ebitda', 'ebitda_margin', 'momentum_12m']
    val_frames = []
    for f in sorted(glob.glob('*_yartseva.csv')):
        try:
            d = pd.read_csv(f, usecols=lambda c: c in val_cols)
        except Exception:
            continue
        if 'symbol' in d.columns:
            val_frames.append(d)
    if val_frames:
        val = pd.concat(val_frames, ignore_index=True).drop_duplicates('symbol', keep='first')
        # Only merge columns we don't already have
        merge_cols = ['symbol'] + [c for c in val.columns if c != 'symbol' and c not in df.columns]
        df = df.merge(val[merge_cols], on='symbol', how='left')

    if 'entry_today_asymmetry' not in df.columns:
        df['entry_today_asymmetry'] = df.get('asymmetry_score', 0)
    df['entry_today_asymmetry'] = df['entry_today_asymmetry'].fillna(0)
    if 'market_cap' in df.columns:
        df = df[df['market_cap'].fillna(0) >= 10_000_000]
    if 'verdict' in df.columns:
        df = df[df['verdict'] != 'RED']
    return df


def _sheet_safe(s: str) -> str:
    out = ''.join(ch if ch.isalnum() or ch in '_+- ' else '_' for ch in str(s))
    return out[:31]


def _write_segment_table(ws, df_subset, label, n_total, sort_col='entry_today_asymmetry'):
    f_bold = _font(bold=True, color=INK)
    f_bold_muted = _font(bold=True, color=MUTED)
    f_text = _font(color=INK)
    f_text_muted = _font(color=MUTED)
    f_italic_muted = _font(italic=True, color=MUTED)

    NCOLS = 21
    # Title
    t = ws.cell(row=2, column=1, value=label)
    t.font = f_bold
    t.alignment = _TXT_ALIGN_LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    ws.row_dimensions[2].height = 22
    for c in range(1, NCOLS + 1):
        ws.cell(row=3, column=c).border = Border(bottom=Side(style='thin', color=INK))
    ws.row_dimensions[3].height = 4

    # Headline tiles
    n_match = len(df_subset)
    pct_universe = (n_match / n_total * 100) if n_total else 0
    n_diverse = int((df_subset['segment_count'].fillna(0) >= 4).sum())
    n_concentrated = int((df_subset['segment_revenue_hhi'].fillna(0) >= 0.7).sum())
    n_global = int((df_subset['geographic_region_count'].fillna(0) >= 4).sum())
    fastest_row = df_subset.dropna(subset=['fastest_segment_yoy']).nlargest(1, 'fastest_segment_yoy')
    fastest_label = (f"{fastest_row.iloc[0]['symbol']} {fastest_row.iloc[0]['fastest_segment_name']} "
                     f"+{fastest_row.iloc[0]['fastest_segment_yoy']*100:.0f}%"
                     if len(fastest_row) else "—")

    headline = [
        ("MATCHES", f"{n_match:,}", f"{pct_universe:.1f}% of segment-covered"),
        ("DIVERSIFIED", f"{n_diverse:,}", "4+ segments"),
        ("CONCENTRATED", f"{n_concentrated:,}", "HHI >= 0.70"),
        ("GLOBAL", f"{n_global:,}", "4+ regions"),
        ("FASTEST", fastest_label[:24], "yoy among matches"),
    ]
    for i, (lbl, val, sub) in enumerate(headline):
        col = 1 + i * 4
        ws.cell(row=5, column=col, value=lbl).font = f_italic_muted
        ws.cell(row=5, column=col).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=6, column=col, value=val).font = f_bold
        ws.cell(row=6, column=col).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=7, column=col, value=sub).font = f_italic_muted
        ws.cell(row=7, column=col).alignment = _TXT_ALIGN_LEFT
    ws.row_dimensions[6].height = 22
    for c in range(1, NCOLS + 1):
        ws.cell(row=8, column=c).border = Border(top=Side(style='thin', color=INK))
    ws.row_dimensions[8].height = 4

    _section_rule(ws, 10, "Headline valuation + segment detail — top by ETA", span_cols=NCOLS)

    headers = ['#', 'Ticker', 'Name', 'Country', 'Sector', 'Mcap (USD)',
               'Verdict',
               'EV/EBITDA', 'P/E', 'FCF yld %', 'ROIC %', 'EBITDA m %',
               'ND/EBITDA', 'Mom 12m %',
               'Segs', 'HHI', 'Largest segment (share)',
               'Top 3 segments', 'Regs', 'Top regions', 'Fastest segment YoY']
    text_cols = {2, 3, 4, 5, 17, 18, 20, 21}  # ticker/name/country/sector/segment text
    center_cols = {7, 15}  # verdict + segs count
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=11, column=i, value=h)
        c.font = f_bold_muted
        c.alignment = (_TXT_ALIGN_LEFT if i in text_cols else
                       _NUM_ALIGN_CENTER if i in center_cols else
                       _NUM_ALIGN_RIGHT)
    for c in range(1, NCOLS + 1):
        ws.cell(row=12, column=c).border = Border(top=Side(style='thin', color=INK))

    for r_idx, (_, r) in enumerate(df_subset.iterrows(), start=13):
        _write_int(ws, r_idx, 1, r_idx - 12, font=f_text_muted)
        ws.cell(row=r_idx, column=2, value=r['symbol']).font = f_bold
        ws.cell(row=r_idx, column=2).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=3, value=str(r.get('name') or '')[:30]).font = f_text
        ws.cell(row=r_idx, column=3).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=4, value=str(r.get('src') or '')).font = f_text_muted
        ws.cell(row=r_idx, column=4).alignment = _NUM_ALIGN_CENTER
        ws.cell(row=r_idx, column=5, value=str(r.get('sector') or '')[:14]).font = f_text_muted
        ws.cell(row=r_idx, column=5).alignment = _TXT_ALIGN_LEFT
        _write_money(ws, r_idx, 6, r.get('market_cap'), font=f_text)
        _verdict_badge(ws, r_idx, 7, r.get('verdict') or 'UNRESEARCHED')
        # Valuation columns
        _write_score(ws, r_idx, 8, r.get('ev_ebitda'), font=f_text)
        _write_score(ws, r_idx, 9, r.get('p_e'), font=f_text)
        _write_pct(ws, r_idx, 10, r.get('fcf_yield'), font=f_text)
        _write_pct(ws, r_idx, 11, r.get('roce'), font=f_text)
        _write_pct(ws, r_idx, 12, r.get('ebitda_margin'), font=f_text)
        _write_score(ws, r_idx, 13, r.get('net_debt_ebitda'), font=f_text)
        _write_pct(ws, r_idx, 14, r.get('momentum_12m'), font=f_text)
        # Segment columns
        _write_int(ws, r_idx, 15, int(r.get('segment_count') or 0), font=f_text)
        _write_ratio(ws, r_idx, 16, r.get('segment_revenue_hhi'), font=f_text)
        ls_name = str(r.get('largest_segment_name') or '')[:22]
        ls_share = r.get('largest_segment_share')
        if pd.notna(ls_share) and ls_share is not None:
            largest_str = f"{ls_name} ({ls_share*100:.0f}%)"
        else:
            largest_str = ls_name
        ws.cell(row=r_idx, column=17, value=largest_str).font = f_text
        ws.cell(row=r_idx, column=17).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=18, value=str(r.get('top_segments') or '')[:80]).font = f_text_muted
        ws.cell(row=r_idx, column=18).alignment = _TXT_ALIGN_LEFT
        _write_int(ws, r_idx, 19, int(r.get('geographic_region_count') or 0), font=f_text)
        ws.cell(row=r_idx, column=20, value=str(r.get('top_regions') or '')[:60]).font = f_text_muted
        ws.cell(row=r_idx, column=20).alignment = _TXT_ALIGN_LEFT
        fs_name = str(r.get('fastest_segment_name') or '')[:20]
        fs_yoy = r.get('fastest_segment_yoy')
        if pd.notna(fs_yoy):
            fast_str = f"{fs_name} {fs_yoy*100:+.0f}%"
        else:
            fast_str = fs_name
        ws.cell(row=r_idx, column=21, value=fast_str).font = f_text
        ws.cell(row=r_idx, column=21).alignment = _TXT_ALIGN_LEFT
        for c in range(1, NCOLS + 1):
            ws.cell(row=r_idx, column=c).border = Border(
                bottom=Side(style='thin', color=RULE))
        ws.row_dimensions[r_idx].height = 16

    # Column widths tuned for the new wider layout
    widths = {1: 4, 2: 10, 3: 22, 4: 6, 5: 14, 6: 14, 7: 12,
              8: 9, 9: 8, 10: 9, 11: 8, 12: 10, 13: 9, 14: 10,
              15: 5, 16: 7, 17: 28, 18: 52, 19: 5, 20: 32, 21: 22}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A13'
    # QoL: sortable/filterable table (header row 11 → last data row)
    if ws.max_row >= 13:
        ws.auto_filter.ref = f"A11:{get_column_letter(ws.max_column)}{ws.max_row}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--n', type=int, default=60, help='top N per archetype tab')
    ap.add_argument('--detail-n', type=int, default=300,
                    help='top N in the All Segments master tab')
    ap.add_argument('--out', default='segment_detail_book.xlsx')
    args = ap.parse_args()

    print('loading data...', file=sys.stderr)
    df = load_data()
    print(f'  {len(df):,} segment-covered rows', file=sys.stderr)
    if 'entry_today_asymmetry' not in df.columns:
        df['entry_today_asymmetry'] = 0
    sort_col = 'entry_today_asymmetry'

    wb = Workbook()
    cover = wb.active
    cover.title = 'Cover'
    cover.column_dimensions['A'].width = 6
    for col_letter in 'BCDEFG':
        cover.column_dimensions[col_letter].width = 28
    cover.column_dimensions['H'].width = 6

    f_bold = _font(bold=True, color=INK)
    f_italic = _font(italic=True, color=INK)
    f_italic_muted = _font(italic=True, color=MUTED)
    f_text = _font(color=INK)

    t = cover.cell(row=3, column=2, value="Segment-Level Detail")
    t.font = f_bold
    cover.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    cover.row_dimensions[3].height = 22
    for c in range(2, 8):
        cover.cell(row=4, column=c).border = Border(bottom=Side(style='thin', color=INK))
    cover.row_dimensions[4].height = 4

    sub = cover.cell(row=5, column=2,
                     value="What the binary archetypes hide — actual segment names, % of sales, YoY growth")
    sub.font = f_italic
    cover.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    src_note = cover.cell(row=7, column=2,
                          value="Source: SEC EDGAR dimensional XBRL (us-gaap StatementBusinessSegmentsAxis, "
                                "srt StatementGeographicalAxis). Coverage: US filers with multi-segment 10-K disclosures.")
    src_note.font = f_italic_muted
    cover.merge_cells(start_row=7, start_column=2, end_row=7, end_column=7)

    _section_rule(cover, 9, "Universe stats", span_cols=7)
    n_total = len(df)
    n_4plus = int((df['segment_count'].fillna(0) >= 4).sum())
    n_hhi_diverse = int((df['segment_revenue_hhi'].fillna(1.0) <= 0.40).sum())
    n_hhi_concentr = int((df['segment_revenue_hhi'].fillna(0) >= 0.70).sum())
    n_global = int((df['geographic_region_count'].fillna(0) >= 4).sum())
    n_fast = int((df['fastest_segment_yoy'].fillna(0) >= 0.25).sum())

    tiles = [
        ("FILERS", f"{n_total:,}", "with segment data"),
        ("DIVERSIFIED", f"{n_4plus:,}", "4+ segments"),
        ("BALANCED", f"{n_hhi_diverse:,}", "HHI <= 0.40"),
        ("CONCENTRATED", f"{n_hhi_concentr:,}", "HHI >= 0.70"),
        ("GLOBAL", f"{n_global:,}", "4+ regions"),
        ("FAST SEG", f"{n_fast:,}", "+25% YoY"),
    ]
    for i, (lbl, val, sub_lbl) in enumerate(tiles):
        col = 2 + i
        cover.cell(row=10, column=col, value=lbl).font = f_italic_muted
        cover.cell(row=11, column=col, value=val).font = f_bold
        cover.cell(row=12, column=col, value=sub_lbl).font = f_italic_muted
        for r in (10, 11, 12):
            cover.cell(row=r, column=col).alignment = _TXT_ALIGN_LEFT
    cover.row_dimensions[11].height = 24
    for c in range(2, 8):
        cover.cell(row=13, column=c).border = Border(top=Side(style='thin', color=INK))
    cover.row_dimensions[13].height = 4

    # Per-archetype top scorers index
    _section_rule(cover, 15, "Tabs in this workbook", span_cols=7)
    rows_meta = [
        ("All Segments", f"top {args.detail_n:,} by ETA", "every name we have segment data on"),
        ("Diversified", f"top {args.n:,}", "4+ segments AND HHI <= 0.40 (real diversification)"),
        ("Concentrated", f"top {args.n:,}", "HHI >= 0.70 or top segment >= 70% (single-segment risk)"),
        ("Global", f"top {args.n:,}", "4+ reporting geographies"),
        ("Fastest", f"top {args.n:,}", "single segment growing > 25% YoY (hidden engine)"),
    ]
    for i, (tab, n, desc) in enumerate(rows_meta, start=16):
        cover.cell(row=i, column=2, value=tab).font = f_bold
        cover.cell(row=i, column=2).hyperlink = f"#'{_sheet_safe(tab)}'!A1"
        cover.cell(row=i, column=2).alignment = _TXT_ALIGN_LEFT
        cover.cell(row=i, column=3, value=n).font = f_text
        cover.cell(row=i, column=3).alignment = _TXT_ALIGN_LEFT
        cover.cell(row=i, column=4, value=desc).font = f_text
        cover.cell(row=i, column=4).alignment = _TXT_ALIGN_LEFT
        cover.merge_cells(start_row=i, start_column=4, end_row=i, end_column=7)

    cover.sheet_view.showGridLines = False

    # === All Segments tab ===
    all_seg = df.sort_values(sort_col, ascending=False).head(args.detail_n).reset_index(drop=True)
    ws = wb.create_sheet('All Segments')
    _write_segment_table(ws, all_seg, 'All Segment-Covered Names (top by ETA)', n_total, sort_col)

    # === Per-archetype tabs ===
    archetypes = [
        ('Diversified',
         df[(df['segment_count'].fillna(0) >= 4) & (df['segment_revenue_hhi'].fillna(1.0) <= 0.40)],
         '4+ segments AND HHI <= 0.40'),
        ('Concentrated',
         df[((df['segment_revenue_hhi'].fillna(0) >= 0.70) |
             (df['largest_segment_share'].fillna(0) >= 0.70)) &
            (df['segment_count'].fillna(0) >= 2)],
         'HHI >= 0.70 or top segment >= 70%'),
        ('Global',
         df[df['geographic_region_count'].fillna(0) >= 4],
         '4+ reporting geographies'),
        ('Fastest',
         df[(df['fastest_segment_yoy'].fillna(0) >= 0.25) & (df['segment_count'].fillna(0) >= 2)],
         'Single segment growing > 25% YoY'),
    ]

    for label, sub_df, desc in archetypes:
        sub_df = sub_df.nlargest(args.n, sort_col).reset_index(drop=True)
        if sub_df.empty:
            continue
        ws = wb.create_sheet(_sheet_safe(label))
        _write_segment_table(ws, sub_df, f"{label} — {desc}", n_total, sort_col)

    wb.save(args.out)
    print(f'wrote {args.out}: {len(wb.worksheets)} sheets', file=sys.stderr)


if __name__ == '__main__':
    main()
