"""Per-archetype top-N book.

For each of the 30 archetypes, one tab listing the top-N names that
match it, ranked by entry_today_asymmetry. Plus a Cover with archetype
counts and a Density tab surfacing the names that match the MOST
archetypes (cross-archetype winners).

Reuses the elegant Harvard helpers from build_harvard_workbook.

Output: top_by_archetype_book.xlsx
"""
from __future__ import annotations
import argparse
import os
import sys

import pandas as pd

import build_harvard_workbook as bhw
from build_harvard_workbook import (
    Workbook, Color,
    INK, DARK_GREY, MUTED, RULE, LIGHT_GREY, PALE_GREY, WHITE,
    FONT_NAME, FONT_SIZE,
    _font, _fill, _border, _align,
    _set_col_widths, _crimson_banner, _section_rule, _verdict_badge,
    _write_money, _write_pct, _write_ratio, _write_score, _write_int,
    _NUM_ALIGN_RIGHT, _NUM_ALIGN_CENTER, _TXT_ALIGN_LEFT,
)
from openpyxl.styles import Alignment, Border, Side


# Friendly archetype label mapping (matches archetype_tags.py pretty dict)
ARCHETYPE_LABELS = {
    'arch_narrative_lag': 'Narrative Lag',
    'arch_fixed_cost_demand_shock': 'Fixed-Cost + Demand Shock',
    'arch_discounted_vehicle': 'Discounted Vehicle',
    'arch_capital_discipline': 'Capital Discipline',
    'arch_regime_cyclical': 'Regime-Change Cyclical',
    'arch_dead_option': 'Dead Option',
    'arch_kpi_threshold': 'KPI Threshold',
    'arch_blindspot': 'Blind-Spot Geography',
    'arch_micro_activist_inflect': 'Microcap Activist Inflect',
    'arch_durable_reinvestment': 'Durable Reinvestment',
    'arch_cash_reinvest': 'Cash Reinvestment',
    'arch_roic_inflect': 'ROIC Inflection',
    'arch_cheap_per_roiic': 'Cheap per ROIIC',
    'arch_tangible_value': 'Tangible Value',
    'arch_lindy_margin': 'Lindy Margin',
    'arch_lindy_fcf': 'Lindy FCF',
    'arch_no_dilution': 'No Dilution',
    'arch_lindy_growth': 'Lindy Growth',
    'arch_quiet_compounder': 'Quiet Compounder',
    'arch_buyback_compounder': 'Buyback Compounder',
    'arch_owner_operator': 'Owner-Operator',
    'arch_qarp': 'QARP',
    'arch_reinvest_inflect': 'Reinvestment Inflection',
    'arch_double_inflect': 'Double Inflection',
    'arch_cash_quality': 'Cash Quality',
    'arch_capital_light_pivot': 'Capital-Light Pivot',
    'arch_capital_returner': 'Capital Returner',
    'arch_low_sbc_quality': 'Low-SBC Quality',
    'arch_tax_efficient': 'Tax Efficient',
    'arch_strong_coverage': 'Strong Coverage',
    'arch_diversified_segments': 'Diversified Segments',
    'arch_concentrated_segments': 'Concentrated Segments',
    'arch_geographic_global': 'Global Geographic Footprint',
    'arch_fastest_segment': 'Fastest Segment Inflection',
}


def _sheet_safe(s: str) -> str:
    """Excel sheet names: max 31 chars, no /\\?*[]:"""
    out = ''.join(ch if ch.isalnum() or ch in '_+- ' else '_' for ch in str(s))
    return out[:31]


def load_data():
    """Merge asymmetry_global + archetype_tags + verdicts + valuation."""
    df = pd.read_csv('asymmetry_global.csv').drop_duplicates('symbol')
    # Strip any stale suffixed columns that prior merges left behind
    df = df.drop(columns=[c for c in df.columns if c.endswith('_arch')])
    arch = pd.read_csv('archetype_tags.csv')
    arch_cols = [c for c in arch.columns if c.startswith('arch_')]
    # Drop overlapping columns from arch before merge to avoid suffix collision
    overlap = [c for c in arch.columns if c != 'symbol' and c in df.columns]
    df = df.merge(arch.drop(columns=overlap), on='symbol', how='left')

    # Verdicts
    frames = []
    for path, default in [
        ('qualitative_aligned_green.csv', 'GREEN'),
        ('qualitative_red_avoid.csv', 'RED'),
        ('qualitative_extended_verdicts.csv', None),
    ]:
        if not os.path.exists(path):
            continue
        try:
            d = pd.read_csv(path)
        except pd.errors.ParserError:
            d = pd.read_csv(path, engine='python', on_bad_lines='skip', quoting=3)
        if 'verdict' not in d.columns and default is not None:
            d['verdict'] = default
        frames.append(d[[c for c in ['symbol', 'verdict'] if c in d.columns]])
    if frames:
        v = pd.concat(frames, ignore_index=True).drop_duplicates('symbol', keep='last')
        df = df.drop(columns=[c for c in ('verdict',) if c in df.columns])
        df = df.merge(v, on='symbol', how='left')
    df['verdict'] = df['verdict'].fillna('UNRESEARCHED')

    # Valuation columns from per-country yartseva CSVs
    import glob
    val_cols = ['symbol', 'ev_ebitda', 'p_e', 'pb', 'fcf_yield', 'roce',
                'net_debt_ebitda', 'ebitda_margin', 'momentum_12m']
    val_frames = []
    for f in sorted(glob.glob('*_yartseva.csv')):
        try:
            d = pd.read_csv(f, usecols=lambda c: c in val_cols)
        except Exception:
            continue
        if 'symbol' in d.columns:
            val_frames.append(d)
    if val_frames:
        val = pd.concat(val_frames, ignore_index=True).drop_duplicates('symbol', keep='first')
        mc = ['symbol'] + [c for c in val.columns if c != 'symbol' and c not in df.columns]
        df = df.merge(val[mc], on='symbol', how='left')

    # Apply min mcap + exclude RED
    df = df[df['market_cap'].fillna(0) >= 10_000_000]
    df = df[df['verdict'] != 'RED']
    return df, arch_cols


def _write_archetype_table(ws, df_subset, archetype_label, total_universe, sort_col):
    """Render one archetype sheet with headline figures + top-N table."""
    f_bold = _font(bold=True, color=INK)
    f_bold_muted = _font(bold=True, color=MUTED)
    f_text = _font(color=INK)
    f_text_muted = _font(color=MUTED)
    f_italic_muted = _font(italic=True, color=MUTED)

    # Title banner
    t = ws.cell(row=2, column=1, value=archetype_label)
    t.font = f_bold
    t.alignment = _TXT_ALIGN_LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    ws.row_dimensions[2].height = 22
    for c in range(1, 20):
        ws.cell(row=3, column=c).border = Border(bottom=Side(style='thin', color=INK))
    ws.row_dimensions[3].height = 4

    # Headline figures (5 tiles)
    n_match = len(df_subset)
    n_green = int((df_subset['verdict'] == 'GREEN').sum())
    n_yellow = int((df_subset['verdict'] == 'YELLOW').sum())
    n_unr = int((df_subset['verdict'] == 'UNRESEARCHED').sum())
    pct_universe = (n_match / total_universe * 100) if total_universe else 0

    headline = [
        ("MATCHES", f"{n_match:,}", f"{pct_universe:.1f}% of universe"),
        ("GREEN", f"{n_green}", "high conviction"),
        ("YELLOW", f"{n_yellow}", "risk-flagged"),
        ("UNRESEARCHED", f"{n_unr:,}", "no thesis"),
        ("TOP SCORER",
         (df_subset.iloc[0]['symbol'] if len(df_subset) else '—'),
         f"by {sort_col.replace('_', ' ')}"),
    ]
    for i, (lbl, val, sub) in enumerate(headline):
        col = 1 + i * 3
        ws.cell(row=5, column=col, value=lbl).font = f_italic_muted
        ws.cell(row=5, column=col).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=6, column=col, value=val).font = f_bold
        ws.cell(row=6, column=col).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=7, column=col, value=sub).font = f_italic_muted
        ws.cell(row=7, column=col).alignment = _TXT_ALIGN_LEFT
    ws.row_dimensions[6].height = 22
    for c in range(1, 20):
        ws.cell(row=8, column=c).border = Border(top=Side(style='thin', color=INK))
    ws.row_dimensions[8].height = 4

    # Top-N table heading
    _section_rule(ws, 10, f"Top names matching {archetype_label}", span_cols=19)

    headers = ['#', 'Ticker', 'Name', 'Country', 'Sector', 'Bucket',
               'Mcap (USD)', 'Verdict', 'ETA', 'Asym',
               'EV/EBITDA', 'P/E', 'P/B',
               'FCF yld %', 'ROIC %', 'ND/EBITDA', 'EBITDA m %',
               'Mom 12m %', 'Arch #']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=11, column=i, value=h)
        c.font = f_bold_muted
        c.alignment = (_TXT_ALIGN_LEFT if i in (2, 3, 4, 5) else
                       _NUM_ALIGN_CENTER if i in (6, 8) else
                       _NUM_ALIGN_RIGHT)
    for c in range(1, 20):
        ws.cell(row=12, column=c).border = Border(top=Side(style='thin', color=INK))

    # Top-N data rows
    for r_idx, (_, r) in enumerate(df_subset.iterrows(), start=13):
        _write_int(ws, r_idx, 1, r_idx - 12, font=f_text_muted)
        ws.cell(row=r_idx, column=2, value=r['symbol']).font = f_bold
        ws.cell(row=r_idx, column=2).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=3, value=str(r.get('name') or '')[:50]).font = f_text
        ws.cell(row=r_idx, column=3).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=4, value=str(r.get('src') or '')).font = f_text_muted
        ws.cell(row=r_idx, column=4).alignment = _NUM_ALIGN_CENTER
        ws.cell(row=r_idx, column=5, value=str(r.get('sector') or '')).font = f_text_muted
        ws.cell(row=r_idx, column=5).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=r_idx, column=6, value=str(r.get('market_cap_bucket') or '')).font = f_text_muted
        ws.cell(row=r_idx, column=6).alignment = _NUM_ALIGN_CENTER
        _write_money(ws, r_idx, 7, r.get('market_cap'), font=f_text)
        _verdict_badge(ws, r_idx, 8, r['verdict'])
        _write_score(ws, r_idx, 9, r.get('entry_today_asymmetry'), font=f_bold)
        _write_score(ws, r_idx, 10, r.get('asymmetry_score'), font=f_text)
        _write_score(ws, r_idx, 11, r.get('ev_ebitda'), font=f_text)
        _write_score(ws, r_idx, 12, r.get('p_e'), font=f_text)
        _write_score(ws, r_idx, 13, r.get('pb'), font=f_text)
        _write_pct(ws, r_idx, 14, r.get('fcf_yield'), font=f_text)
        _write_pct(ws, r_idx, 15, r.get('roce'), font=f_text)
        _write_score(ws, r_idx, 16, r.get('net_debt_ebitda'), font=f_text)
        _write_pct(ws, r_idx, 17, r.get('ebitda_margin'), font=f_text)
        _write_pct(ws, r_idx, 18, r.get('momentum_12m'), font=f_text)
        _write_int(ws, r_idx, 19, int(r.get('archetype_count') or 0), font=f_text_muted)
        for c in range(1, 20):
            ws.cell(row=r_idx, column=c).border = Border(
                bottom=Side(style='thin', color=RULE))
        ws.row_dimensions[r_idx].height = 16

    # Column widths
    widths = {1: 4, 2: 11, 3: 32, 4: 6, 5: 16, 6: 11, 7: 17, 8: 12,
              9: 8, 10: 8, 11: 10, 12: 8, 13: 8, 14: 10, 15: 9,
              16: 10, 17: 10, 18: 11, 19: 7}
    from openpyxl.utils import get_column_letter
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A13'
    # QoL: sortable/filterable table (header row 11 → last data row)
    if ws.max_row >= 13:
        ws.auto_filter.ref = f"A11:{get_column_letter(ws.max_column)}{ws.max_row}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--n', type=int, default=30, help='top N per archetype')
    ap.add_argument('--out', default='top_by_archetype_book.xlsx')
    args = ap.parse_args()

    print('loading data...', file=sys.stderr)
    df, arch_cols = load_data()
    print(f'  {len(df):,} eligible rows, {len(arch_cols)} archetypes', file=sys.stderr)

    # ETA — verdict-aware ranking
    mult = {'GREEN': 1.10, 'YELLOW': 0.85, 'RED': 0.40}
    qm = df['verdict'].map(mult).fillna(1.0)
    if 'entry_today_asymmetry' not in df.columns:
        df['entry_today_asymmetry'] = df['asymmetry_score'].fillna(0) * qm
    sort_col = 'entry_today_asymmetry'

    wb = Workbook()

    # === Cover ===
    cover = wb.active
    cover.title = 'Cover'
    cover.column_dimensions['A'].width = 6
    for col_letter in 'BCDEFG':
        cover.column_dimensions[col_letter].width = 24
    cover.column_dimensions['H'].width = 6

    f_bold = _font(bold=True, color=INK)
    f_italic = _font(italic=True, color=INK)
    f_italic_muted = _font(italic=True, color=MUTED)
    f_text = _font(color=INK)

    t = cover.cell(row=3, column=2, value=f"Top {args.n} per Archetype")
    t.font = f_bold
    cover.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    cover.row_dimensions[3].height = 22
    for c in range(2, 8):
        cover.cell(row=4, column=c).border = Border(bottom=Side(style='thin', color=INK))
    cover.row_dimensions[4].height = 4

    sub = cover.cell(row=5, column=2,
                     value="One tab per multibagger archetype — the top names that match each pattern")
    sub.font = f_italic
    cover.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    note = cover.cell(row=7, column=2,
                     value="Yartseva-aligned upside  ·  Graham downside floor  ·  30 archetypes  ·  EDGAR XBRL coverage on US filers  ·  As of 24 June 2026")
    note.font = f_italic_muted
    cover.merge_cells(start_row=7, start_column=2, end_row=7, end_column=7)

    # Universe headline tiles
    _section_rule(cover, 9, "Headline figures", span_cols=7)
    n_total = len(df)
    n_green = int((df['verdict'] == 'GREEN').sum())
    n_yellow = int((df['verdict'] == 'YELLOW').sum())
    n_unr = int((df['verdict'] == 'UNRESEARCHED').sum())
    multi_count = int((df[arch_cols].fillna(0).astype(int).sum(axis=1) >= 2).sum())
    high_density = int((df[arch_cols].fillna(0).astype(int).sum(axis=1) >= 5).sum())

    tiles = [
        ("UNIVERSE", f"{n_total:,}", "names ranked"),
        ("ARCHETYPES", f"{len(arch_cols)}", "patterns"),
        ("MULTI-MATCH", f"{multi_count:,}", "2+ archetypes"),
        ("DENSE", f"{high_density:,}", "5+ archetypes"),
        ("GREEN", f"{n_green}", "high conviction"),
        ("UNRESEARCHED", f"{n_unr:,}", "no thesis"),
    ]
    for i, (lbl, val, sub_lbl) in enumerate(tiles):
        col = 2 + i
        cover.cell(row=10, column=col, value=lbl).font = f_italic_muted
        cover.cell(row=10, column=col).alignment = _TXT_ALIGN_LEFT
        cover.cell(row=11, column=col, value=val).font = f_bold
        cover.cell(row=11, column=col).alignment = _TXT_ALIGN_LEFT
        cover.cell(row=12, column=col, value=sub_lbl).font = f_italic_muted
        cover.cell(row=12, column=col).alignment = _TXT_ALIGN_LEFT
    cover.row_dimensions[11].height = 24
    for c in range(2, 8):
        cover.cell(row=13, column=c).border = Border(top=Side(style='thin', color=INK))
    cover.row_dimensions[13].height = 4

    # Archetype index — count + top scorer per archetype, sorted by count
    _section_rule(cover, 15, "Archetype index", span_cols=7)
    idx_headers = ['Archetype', 'Matches', 'GREEN', 'YELLOW', 'UNR', 'Top scorer', 'Top ETA']
    for i, h in enumerate(idx_headers, start=2):
        c = cover.cell(row=16, column=i, value=h)
        c.font = _font(bold=True, color=MUTED)
        c.alignment = _TXT_ALIGN_LEFT if i in (2, 7) else _NUM_ALIGN_RIGHT
    for c in range(2, 9):
        cover.cell(row=17, column=c).border = Border(top=Side(style='thin', color=INK))

    arch_summary = []
    for col in arch_cols:
        sub_df = df[df[col].fillna(0).astype(int) == 1]
        if sub_df.empty:
            continue
        top = sub_df.nlargest(1, sort_col)
        arch_summary.append({
            'arch_col': col,
            'label': ARCHETYPE_LABELS.get(col, col),
            'matches': len(sub_df),
            'green': int((sub_df['verdict'] == 'GREEN').sum()),
            'yellow': int((sub_df['verdict'] == 'YELLOW').sum()),
            'unr': int((sub_df['verdict'] == 'UNRESEARCHED').sum()),
            'top_sym': top.iloc[0]['symbol'] if len(top) else '',
            'top_eta': top.iloc[0][sort_col] if len(top) else None,
        })
    arch_summary.sort(key=lambda r: -r['matches'])

    for row_i, s in enumerate(arch_summary, start=18):
        sheet_name = _sheet_safe(s['label'])
        c_label = cover.cell(row=row_i, column=2, value=s['label'])
        c_label.font = f_text
        c_label.hyperlink = f"#'{sheet_name}'!A1"
        _write_int(cover, row_i, 3, s['matches'], font=f_text)
        _write_int(cover, row_i, 4, s['green'], font=f_text)
        _write_int(cover, row_i, 5, s['yellow'], font=f_text)
        _write_int(cover, row_i, 6, s['unr'], font=f_text)
        cover.cell(row=row_i, column=7, value=s['top_sym']).font = f_bold
        cover.cell(row=row_i, column=7).alignment = _TXT_ALIGN_LEFT
        _write_score(cover, row_i, 8, s['top_eta'], font=f_text)
        for c in range(2, 9):
            cover.cell(row=row_i, column=c).border = Border(
                bottom=Side(style='thin', color=RULE))

    cover.sheet_view.showGridLines = False

    # === Density tab — top names by archetype_count ===
    density_sheet = wb.create_sheet('Density')
    df_density = df.copy()
    df_density['_arch_n'] = df_density[arch_cols].fillna(0).astype(int).sum(axis=1)
    df_density_top = df_density.sort_values(
        ['_arch_n', sort_col], ascending=[False, False]
    ).head(args.n).reset_index(drop=True)
    _write_archetype_table(density_sheet, df_density_top,
                            "Cross-Archetype Density (top by archetype_count)",
                            n_total, sort_col)

    # === Per-archetype tabs ===
    for s in arch_summary:
        col = s['arch_col']
        sub_df = df[df[col].fillna(0).astype(int) == 1].copy()
        sub_df = sub_df.nlargest(args.n, sort_col).reset_index(drop=True)
        if sub_df.empty:
            continue
        sheet_name = _sheet_safe(s['label'])
        ws = wb.create_sheet(sheet_name)
        _write_archetype_table(ws, sub_df, s['label'], n_total, sort_col)

    wb.save(args.out)
    print(f'wrote {args.out}: {len(wb.worksheets)} sheets (Cover + Density + '
          f'{len(arch_summary)} archetypes)', file=sys.stderr)


if __name__ == '__main__':
    main()
