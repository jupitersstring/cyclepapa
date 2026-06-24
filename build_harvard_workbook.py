"""Build a Harvard-aesthetic Excel workbook for the top-N multibagger candidates.

Aesthetic conventions:
  - Harvard Crimson accent (#A51C30) on banners, header bars, footers.
  - Serif typography (Garamond / Cambria / Georgia fallback chain) for body.
  - Sans (Helvetica / Arial) for numeric data.
  - Light-grey rule lines, no heavy gridlines.
  - Each top-N name gets a one-pager: thesis, scoring decomposition, key
    metrics, risk flags, qual verdict, references.

Inputs:
  asymmetry_global.csv
  qualitative_aligned_green.csv
  qualitative_red_avoid.csv
  qualitative_extended_verdicts.csv
  <country>_yartseva.csv  (for intrinsic-discount inputs)
  archetype_tags.csv (optional)

Output:
  asymmetry_harvard_workbook.xlsx
"""
from __future__ import annotations
import argparse
import glob
import os
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Color, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- Monochrome palette + single-font specification --------------------------
# Per user request: one font size only, B&W with tasteful highlighting.
# Differentiation comes from BOLD / ITALIC / BORDER WEIGHT and from
# subtle grey row-fills, never from color or size.

INK = "000000"          # primary text
DARK_GREY = "404040"    # secondary text (table headers etc.)
MUTED = "707070"        # labels
RULE = "B0B0B0"         # hairlines
LIGHT_GREY = "E8E8E8"   # banner / header strip background
PALE_GREY = "F2F2F2"    # alternating-row / mild highlight
WHITE = "FFFFFF"

# Verdict highlighting — monochrome with subtle textural difference
# (border weight) rather than fill color.
GREEN_BG = LIGHT_GREY   # darkest fill - high conviction
YELLOW_BG = PALE_GREY   # lighter fill - caution
RED_BG = WHITE          # no fill - excluded
GRAY_BG = WHITE         # unresearched - no highlight

# Aliases kept for back-compat with other builders that import these
# names. New code should reference INK / DARK_GREY / etc.
CRIMSON = INK
CRIMSON_DARK = DARK_GREY

# Single font face + single size — bold/italic + border weight are the
# only typographic dimensions. Cambria is the closest widely-available
# Harvard-style serif (Microsoft's bundled approximation of Adobe Garamond /
# Sabon), renders cleanly in Excel on Windows / Mac / LibreOffice.
FONT_NAME = "Cambria"
FONT_SIZE = 10
SERIF = FONT_NAME
SANS = FONT_NAME
MONO = FONT_NAME


def _font(size=10, bold=False, color=INK, name=SERIF, italic=False):
    """Build a Font with the project's single-size monochrome spec.

    `size` argument is IGNORED — every call uses FONT_SIZE (10pt). Caller
    can still ask for bold/italic to create hierarchy. Color is forced
    to one of INK / DARK_GREY / MUTED — incoming non-monochrome colors
    are clamped to INK.
    """
    monochrome_palette = {INK, DARK_GREY, MUTED, "FFFFFF", "000000", "FFFFFFFF"}
    if color not in monochrome_palette and not color.upper().startswith("FF"):
        color = INK
    return Font(name=FONT_NAME, size=FONT_SIZE, bold=bold,
                color=color, italic=italic)


def _fill(rgb):
    return PatternFill(fill_type="solid", fgColor=rgb)


def _border(color=RULE, top=None, bottom=None, left=None, right=None):
    sides = {}
    for k, w in [("top", top), ("bottom", bottom), ("left", left), ("right", right)]:
        if w:
            sides[k] = Side(style=w, color=color)
    return Border(**sides)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# --- Data loading -------------------------------------------------------------
def load_verdicts() -> pd.DataFrame:
    frames = []
    for path, default in [
        ('qualitative_aligned_green.csv', 'GREEN'),
        ('qualitative_red_avoid.csv', 'RED'),
        ('qualitative_extended_verdicts.csv', None),
    ]:
        if not os.path.exists(path):
            continue
        try:
            d = pd.read_csv(path)
        except pd.errors.ParserError:
            d = pd.read_csv(path, engine='python', on_bad_lines='skip', quoting=3)
        if 'verdict' not in d.columns and default is not None:
            d['verdict'] = default
        keep = [c for c in ['symbol', 'verdict', 'why', 'why_avoid', 'thesis'] if c in d.columns]
        d = d[keep].copy()
        tc = next((c for c in ['why', 'why_avoid', 'thesis'] if c in d.columns), None)
        d['thesis'] = d[tc] if tc else ''
        frames.append(d[['symbol', 'verdict', 'thesis']])
    if not frames:
        return pd.DataFrame(columns=['symbol', 'verdict', 'thesis'])
    return pd.concat(frames, ignore_index=True).drop_duplicates('symbol', keep='last')


def load_quant() -> pd.DataFrame:
    df = pd.read_csv('asymmetry_global.csv').drop_duplicates('symbol')
    extra_cols = [
        'symbol', 'net_cash_pct_mcap', 'ncav_pct_mcap', 'cash_pct_ev',
        'not_priced_in_score', 'revenue_ttm', 'balance_sheet_date',
        'ev_ebit', 'ev_ebitda', 'p_e', 'p_s', 'roce', 'fcf_yield',
        'sales_yoy', 'ebitda_yoy', 'fcf_yoy',
    ]
    csvs = sorted(set(
        glob.glob('*_yartseva.csv')
        + glob.glob('us_nano_micro_small_yartseva.csv')
        + glob.glob('italian_yartseva.csv')
    ))
    frames = []
    for f in csvs:
        try:
            d = pd.read_csv(f, usecols=lambda c: c in extra_cols)
        except Exception:
            continue
        if 'symbol' in d.columns:
            frames.append(d)
    if frames:
        extra = pd.concat(frames, ignore_index=True).drop_duplicates('symbol', keep='first')
        mc = ['symbol'] + [c for c in extra.columns if c != 'symbol' and c not in df.columns]
        df = df.merge(extra[mc], on='symbol', how='left')
    return df


def compute_scores(df: pd.DataFrame, verdicts: pd.DataFrame) -> pd.DataFrame:
    # asymmetry_global may already carry a verdict column (post-enrichment).
    # Drop it before merging so the fresh verdicts file always wins.
    df = df.drop(columns=[c for c in ('verdict', 'thesis') if c in df.columns])
    df = df.merge(verdicts, on='symbol', how='left')
    df['verdict'] = df['verdict'].fillna('UNRESEARCHED')
    df['thesis'] = df['thesis'].fillna('')

    soft_mult = {'GREEN': 1.10, 'YELLOW': 0.85, 'RED': 0.40}
    strict_mult = {'GREEN': 1.30, 'YELLOW': 0.70, 'RED': 0.0, 'UNRESEARCHED': 0.85}
    df['qual_mult'] = df['verdict'].map(soft_mult).fillna(1.0)
    df['strict_mult'] = df['verdict'].map(strict_mult).fillna(0.85)

    def col(c, d=0.0):
        return df[c].fillna(d) if c in df.columns else pd.Series(d, index=df.index)

    def c01(s):
        return s.clip(0, 1).fillna(0)

    nc = c01(col('net_cash_pct_mcap'))
    ncav = c01(col('ncav_pct_mcap'))
    sub_book = c01(1.0 - col('pb', 2.0).clip(lower=0.01))
    cash_ev = c01((col('cash_pct_ev') - 1.0).clip(0, 2) / 2.0)
    npi = c01(col('not_priced_in_score'))
    df['intrinsic_discount'] = (
        0.30 * nc + 0.20 * ncav + 0.20 * sub_book + 0.15 * cash_ev + 0.15 * npi
    )
    boost = (1.0 + (df['intrinsic_discount'] - 0.25)).clip(0.5, 1.5)

    mom = col('momentum_12m').clip(-0.5, None)
    pr = pd.Series(1.0, index=df.index)
    mid = (mom > 0.30) & (mom <= 1.0)
    hi = (mom > 1.0) & (mom <= 3.0)
    ex = mom > 3.0
    pr.loc[mid] = 1.0 - (mom[mid] - 0.30) / 0.70 * 0.25
    pr.loc[hi] = 0.75 - (mom[hi] - 1.0) / 2.0 * 0.30
    pr.loc[ex] = 0.40
    df['post_rally_factor'] = pr.round(3)

    df['entry_today_asymmetry'] = df['asymmetry_score'] * boost * df['qual_mult'] * pr
    df['entry_today_upside'] = df['upside_score'] * boost * df['qual_mult'] * pr
    return df


# --- Helpers for sheet rendering ---------------------------------------------
# Number formats per the project's Harvard-style spec:
#   - Raw $ and millions integer: `#,##0;(#,##0);"–"`
#   - Percentages / pp / growth / margins: `#,##0.0;(#,##0.0);"–"` on a pre-
#     scaled (×100) value so 0.215 is entered as 21.5
#   - Ratios / scores / prices: `#,##0.00;(#,##0.00);"–"`
# All numbers right-aligned; em-dash for empty cells. Negatives in
# parentheses (financial-statement convention).
EM_DASH = "–"
FMT_INT_RAW = '#,##0;(#,##0);"–"'           # raw dollars + millions integer
FMT_ONE = '#,##0.0;(#,##0.0);"–"'           # percentages / pp / one-decimal numbers
FMT_TWO = '#,##0.00;(#,##0.00);"–"'         # ratios / scores / prices

_NUM_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_NUM_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_TXT_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def _write_num(ws, row, col, value, fmt, scale=1.0, align=_NUM_ALIGN_RIGHT, font=None):
    """Write a numeric cell with the right format + alignment.

    `value` may be None / NaN — we then write an em-dash string and keep
    right-alignment, since Excel's number-format text-section
    (positive;negative;zero;text) doesn't fire on blank cells.
    `scale=100.0` for percentages stored as decimals (0.215 -> 21.5).
    """
    cell = ws.cell(row=row, column=col)
    if value is None or (isinstance(value, float) and pd.isna(value)):
        cell.value = EM_DASH
        cell.alignment = align
    else:
        try:
            cell.value = float(value) * scale
            cell.number_format = fmt
        except (TypeError, ValueError):
            cell.value = EM_DASH
        cell.alignment = align
    if font is not None:
        cell.font = font
    return cell


def _write_money(ws, row, col, value, font=None):
    """Raw USD value: $25,093,996,544 / (123) / – (em-dash on missing)."""
    return _write_num(ws, row, col, value, FMT_INT_RAW, font=font)


def _write_money_m(ws, row, col, value, font=None):
    """Millions: same format, but the caller has already divided by 1e6."""
    return _write_num(ws, row, col, value, FMT_INT_RAW, font=font)


def _write_pct(ws, row, col, value, font=None):
    """Percentage stored as a decimal (0.215 -> 21.5). One decimal place."""
    return _write_num(ws, row, col, value, FMT_ONE, scale=100.0, font=font)


def _write_ratio(ws, row, col, value, font=None):
    """Two-decimal ratio / multiple (12.54x). The 'x' suffix is conveyed by
    the column header — the cell shows just the number for clean sorting."""
    return _write_num(ws, row, col, value, FMT_TWO, font=font)


def _write_score(ws, row, col, value, font=None):
    """Two-decimal composite score (asymmetry, yartseva, inflection, etc.)."""
    return _write_num(ws, row, col, value, FMT_TWO, font=font)


def _write_int(ws, row, col, value, font=None):
    return _write_num(ws, row, col, value, FMT_INT_RAW, font=font)


def _set_col_widths(ws: Worksheet, widths: dict):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _crimson_banner(ws: Worksheet, row: int, text: str, height: int = 30, span_cols: int = 6):
    """Monochrome banner: bold black text on light grey strip, thick rule
    underneath. Name preserved for back-compat with non-Harvard
    workbooks that call this helper."""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=True, color=INK)
    cell.fill = _fill(LIGHT_GREY)
    cell.alignment = _align(h="left", v="center")
    for c in range(1, span_cols + 1):
        ws.cell(row=row, column=c).fill = _fill(LIGHT_GREY)
        ws.cell(row=row, column=c).border = _border(color=INK, bottom="medium")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = height


def _section_rule(ws: Worksheet, row: int, text: str, span_cols: int = 6):
    """Bold uppercase label with a thin black underline. No fill."""
    cell = ws.cell(row=row, column=1, value=text.upper())
    cell.font = _font(bold=True, color=INK)
    cell.alignment = _align()
    for c in range(1, span_cols + 1):
        ws.cell(row=row, column=c).border = _border(color=INK, bottom="thin")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    ws.row_dimensions[row].height = 18


def _kv_row(ws: Worksheet, row: int, key: str, value, value_col_span: int = 5):
    ws.cell(row=row, column=1, value=key).font = _font(size=10, bold=True, color=MUTED, name=SANS)
    ws.cell(row=row, column=1).alignment = _align()
    vcell = ws.cell(row=row, column=2, value=value)
    vcell.font = _font(size=10, name=SERIF)
    vcell.alignment = _align(wrap=True)
    if value_col_span > 1:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + value_col_span)


def _verdict_badge(ws: Worksheet, row: int, col: int, verdict: str):
    """Monochrome verdict badge: full word for clarity, fill density
    encodes conviction. GREEN = medium grey fill, YELLOW = pale grey,
    RED = bold-italic on white (drawn-through), UNRESEARCHED = plain."""
    fill_map = {
        'GREEN':        LIGHT_GREY,
        'YELLOW':       PALE_GREY,
        'RED':          WHITE,
        'UNRESEARCHED': WHITE,
    }
    bg = fill_map.get(verdict, WHITE)
    cell = ws.cell(row=row, column=col, value=verdict)
    cell.font = _font(bold=True, italic=(verdict == 'RED'), color=INK)
    cell.fill = _fill(bg)
    cell.alignment = _align(h="center")
    cell.border = _border(color=RULE, top="thin", bottom="thin", left="thin", right="thin")


# --- Cover sheet --------------------------------------------------------------
def build_cover(ws: Worksheet, n_total: int, n_green: int, n_yellow: int,
                n_unresearched: int, top_n: int,
                extra_headlines: list[tuple[str, str, str]] | None = None):
    """Cover sheet: monochrome single-font with a HEADLINE FIGURES strip up
    top. All text at FONT_SIZE; differentiation via BOLD / ITALIC /
    fill / border weight.

    extra_headlines: optional list of (label, value, sub) tiles for the
    builder to inject above the standard four.
    """
    _set_col_widths(ws, {1: 4, 2: 22, 3: 22, 4: 22, 5: 22, 6: 22, 7: 22, 8: 4})

    # Title strip — bold caps on light grey, thin black rule under.
    title = ws.cell(row=2, column=2, value="ASYMMETRY")
    title.font = _font(bold=True, color=INK)
    title.alignment = _align(h="left", v="center")
    for c in range(2, 8):
        ws.cell(row=2, column=c).fill = _fill(LIGHT_GREY)
    for c in range(2, 8):
        ws.cell(row=2, column=c).border = _border(color=INK, bottom="thin")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 22

    # Subtitle (italic, plain background)
    s = ws.cell(row=3, column=2,
                value=f"Top {top_n} multibagger candidates  ·  quantitative + qualitative workbook")
    s.font = _font(italic=True, color=INK)
    s.alignment = _align(h="left", v="center")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws.row_dimensions[3].height = 18

    # Edition / framework note
    d = ws.cell(row=5, column=2,
                value="Asymmetry framework  ·  Yartseva-aligned upside  ·  Graham downside floor  ·  As of 24 June 2026")
    d.font = _font(italic=True, color=MUTED)
    d.alignment = _align(h="left", v="center")
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    # --- HEADLINE FIGURES strip --- (rows 7-9)
    reviewed = n_green + n_yellow + (n_total - n_unresearched - n_green - n_yellow)
    pct_reviewed = (reviewed / n_total * 100) if n_total else 0
    standard = [
        ("UNIVERSE",      f"{n_total:,}",       "names ranked"),
        ("TOP-N",         f"{top_n}",           "highest-conviction"),
        ("VERDICTS",      f"{reviewed:,}",      f"({pct_reviewed:.1f}% reviewed)"),
        ("GREEN",         f"{n_green}",         "high-conviction"),
        ("YELLOW",        f"{n_yellow}",        "risk-flagged"),
        ("UNRESEARCHED",  f"{n_unresearched:,}", "no thesis yet"),
    ]
    tiles = (extra_headlines or []) + standard
    tiles = tiles[:6]  # cap at 6 (we have 6 inner cols)

    # Header row of tiles (labels)
    for i, (lbl, _, _) in enumerate(tiles):
        col = 2 + i
        c = ws.cell(row=7, column=col, value=lbl)
        c.font = _font(bold=True, color=MUTED)
        c.alignment = _align(h="center")
        c.border = _border(color=INK, top="thin", left="thin", right="thin")
        c.fill = _fill(PALE_GREY)
    # Value row — emphasis via bold + tall row + thicker borders
    for i, (_, val, _) in enumerate(tiles):
        col = 2 + i
        c = ws.cell(row=8, column=col, value=val)
        c.font = _font(bold=True, color=INK)
        c.alignment = _align(h="center")
        c.border = _border(color=INK, left="thin", right="thin")
    ws.row_dimensions[8].height = 28
    # Sub row
    for i, (_, _, sub) in enumerate(tiles):
        col = 2 + i
        c = ws.cell(row=9, column=col, value=sub)
        c.font = _font(italic=True, color=MUTED)
        c.alignment = _align(h="center")
        c.border = _border(color=INK, bottom="thin", left="thin", right="thin")

    # --- Abstract block ---
    _section_rule(ws, 12, "Abstract", span_cols=7)
    abstract = (
        f"This workbook documents the top-{top_n} names from a global universe of "
        f"{n_total:,} equities ranked by an entry-today asymmetry score. The upside leg "
        "weights factors Anna Yartseva (CAFE WP 33, 2025) finds predictive of 10-bagger "
        "outcomes — FCF yield, book-to-market, small size, profitability level, asset-"
        "growth gate, contra-momentum. The downside floor leg combines Graham net-net, "
        "cash > EV, sub-book and negative-EV signals. Each candidate has been qualitatively "
        "reviewed and assigned a verdict of GREEN (high-conviction), YELLOW (risk-flagged), "
        "RED (avoid) or UNRESEARCHED. Sheets: Cover, Methodology, Index, per-name pages, "
        "Coverage, References."
    )
    a = ws.cell(row=13, column=2, value=abstract)
    a.font = _font(color=INK)
    a.alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=13, start_column=2, end_row=17, end_column=7)
    for r in range(13, 18):
        ws.row_dimensions[r].height = 18

    # Footer note
    f = ws.cell(row=20, column=2,
                value="Prepared for internal allocation use. Not investment advice.")
    f.font = _font(italic=True, color=MUTED)
    ws.merge_cells(start_row=20, start_column=2, end_row=20, end_column=7)
    for c in range(2, 8):
        ws.cell(row=20, column=c).border = _border(color=INK, top="thin")


# --- Methodology sheet --------------------------------------------------------
def build_methodology(ws: Worksheet):
    _set_col_widths(ws, {1: 4, 2: 22, 3: 60, 4: 4})
    _crimson_banner(ws, 1, "  Methodology", span_cols=4)
    body = [
        ("Universe",
         f"{15944:,} equities aggregated from financedatabase across 40+ country ISO codes, "
         "filtered to mcap ≥ $10M to exclude untradeable nano-caps."),
        ("Asymmetry score",
         "Composite of inflection (Yartseva growth-flip + first-positive prints), valuation "
         "floor (sub-book, NCAV, net-cash, P/E vs growth), and quality (ROCE, cash conversion, "
         "leverage). Each leg is rank-normalised within universe before weighting."),
        ("Intrinsic discount",
         "0.30·net_cash/mcap + 0.20·NCAV/mcap + 0.20·(1−P/B) + 0.15·cash_pct_EV + 0.15·"
         "not_priced_in. Drives a 0.5–1.5x multiplier on the asymmetry score."),
        ("Post-rally factor",
         "Smooth demotion (not exclusion) of names already up 30–300%+ over 12m; floor 0.40 at "
         "300%+. Keeps in-flight multibaggers visible but ranked behind un-run setups."),
        ("Qualitative multiplier",
         "GREEN +10% (soft) / +30% (strict per-region) · YELLOW −15% / −30% · RED −60% / "
         "EXCLUDED · UNRESEARCHED no-change (soft) / −15% (strict)."),
        ("Entry-today asymmetry",
         "asymmetry_score × intrinsic_boost × qual_mult × post_rally_factor. The primary "
         "ranking column in this workbook."),
        ("Verdicts",
         "Assigned by forensic per-name research: business activity, governance flags, recent "
         "filings, dilution history, sector dynamics. GREEN = plausible 3–5x on 3–5y; YELLOW = "
         "operates but material risk; RED = shell / fraud / serial dilution / wound down."),
    ]
    row = 3
    for k, v in body:
        ws.cell(row=row, column=2, value=k).font = _font(size=11, bold=True, color=CRIMSON_DARK, name=SERIF)
        ws.cell(row=row, column=2).alignment = _align(v="top")
        ws.cell(row=row, column=3, value=v).font = _font(size=10, name=SERIF)
        ws.cell(row=row, column=3).alignment = _align(wrap=True, v="top")
        ws.row_dimensions[row].height = 48
        row += 1

    _section_rule(ws, row + 1, "Notes on data sources", span_cols=4)
    notes = (
        "Fundamentals: yfinance quarterly + annual statements (income, cash flow, balance sheet); "
        "TTM rollups, semi-annual fallback for European smid-caps. Universe: financedatabase. "
        "Price/momentum: Yahoo Finance close prices. Qualitative research: company filings, "
        "regulator notices, press releases, broker research — see References."
    )
    ws.cell(row=row + 3, column=2, value=notes).font = _font(size=10, italic=True, color=MUTED, name=SERIF)
    ws.cell(row=row + 3, column=2).alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=row + 3, start_column=2, end_row=row + 5, end_column=3)
    for r in range(row + 3, row + 6):
        ws.row_dimensions[r].height = 22


# --- Index sheet --------------------------------------------------------------
def build_index(ws: Worksheet, top_df: pd.DataFrame):
    _set_col_widths(ws, {1: 4, 2: 5, 3: 14, 4: 38, 5: 8, 6: 18, 7: 10, 8: 14, 9: 4})
    _crimson_banner(ws, 1, "  Index of Names", span_cols=9)

    hdrs = ["#", "Ticker", "Company", "Cntry", "Sector", "Verdict", "Score"]
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        c.alignment = _align(h="left" if i in (3, 4, 6) else "center")
        c.border = _border(color=CRIMSON_DARK, bottom="medium")
    ws.row_dimensions[3].height = 22

    for i, (_, r) in enumerate(top_df.iterrows(), start=1):
        row = 3 + i
        ws.cell(row=row, column=2, value=i).font = _font(size=10, color=MUTED, name=SERIF)
        ws.cell(row=row, column=2).alignment = _align(h="center")

        # Ticker with hyperlink to its sheet
        sheet_name = _sheet_name_for(r['symbol'], i)
        link_cell = ws.cell(row=row, column=3, value=r['symbol'])
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        link_cell.alignment = _align()

        name = (r.get('name') or '')[:60]
        ws.cell(row=row, column=4, value=name).font = _font(size=10, name=SERIF)
        ws.cell(row=row, column=4).alignment = _align(wrap=False)

        ws.cell(row=row, column=5, value=r.get('src', '')).font = _font(size=10, name=SANS)
        ws.cell(row=row, column=5).alignment = _align(h="center")
        ws.cell(row=row, column=6, value=(r.get('sector') or '')).font = _font(size=10, name=SANS, color=MUTED)
        ws.cell(row=row, column=6).alignment = _align()

        _verdict_badge(ws, row, 7, r.get('verdict', 'UNRESEARCHED'))

        _write_score(ws, row, 8, r.get('entry_today_asymmetry'),
                     font=_font(size=10, name=MONO))

        # Row rule (very light)
        for c in range(2, 9):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 18


# --- Per-name sheet -----------------------------------------------------------
def _sheet_name_for(symbol: str, rank: int) -> str:
    safe = ''.join(ch if ch.isalnum() or ch in '_-' else '_' for ch in str(symbol))[:24]
    return f"{rank:02d}_{safe}"[:31]


def build_name_sheet(ws: Worksheet, rank: int, r: pd.Series):
    _set_col_widths(ws, {1: 4, 2: 22, 3: 26, 4: 22, 5: 26, 6: 4})

    # ── Header strip: bold black-on-grey, thin black underline
    sym = str(r.get('symbol', ''))
    name = str(r.get('name', '') or '')
    cell = ws.cell(row=1, column=1, value=f"  #{rank:02d}  ·  {sym}  ·  {name}")
    cell.font = _font(bold=True, color=INK)
    cell.fill = _fill(LIGHT_GREY)
    cell.alignment = _align(h="left", v="center")
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = _fill(LIGHT_GREY)
        ws.cell(row=1, column=c).border = _border(color=INK, bottom="thin")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 22

    # ── HEADLINE FIGURES row — single-font, bold values, light grey fill
    ws.cell(row=3, column=2, value="VERDICT").font = _font(bold=True, color=MUTED)
    _verdict_badge(ws, 3, 3, r.get('verdict', 'UNRESEARCHED'))
    ws.cell(row=3, column=4, value="ENTRY-TODAY SCORE").font = _font(bold=True, color=MUTED)
    sc = _write_score(ws, 3, 5, r.get('entry_today_asymmetry'),
                      font=_font(bold=True, color=INK))
    sc.alignment = _NUM_ALIGN_CENTER
    sc.fill = _fill(PALE_GREY)
    sc.border = _border(color=INK, top="thin", bottom="thin", left="thin", right="thin")

    # ── Snapshot section
    # Text values use _kv_row (label/text pairs). Numeric values
    # (market cap, revenue) go via _write_money so they appear as
    # right-aligned numbers with comma grouping and parens-on-negatives.
    row = 5
    _section_rule(ws, row, "Snapshot", span_cols=5); row += 1

    def _text_kv(row, label, value):
        ws.cell(row=row, column=1, value=label).font = _font(size=10, bold=True, color=MUTED, name=SANS)
        vcell = ws.cell(row=row, column=2, value=value or EM_DASH)
        vcell.font = _font(size=10, name=SERIF)
        vcell.alignment = _TXT_ALIGN_LEFT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)

    def _money_kv(row, label, value):
        # Label in col 1; numeric value in col 2 (no merge — keeps the
        # number right-aligned without dragging it across the row).
        ws.cell(row=row, column=1, value=label).font = _font(size=10, bold=True, color=MUTED, name=SANS)
        _write_money(ws, row, 2, value, font=_font(size=10, name=MONO))

    _text_kv(row, "Country", r.get('src') or '');                                row += 1
    _text_kv(row, "Sector", r.get('sector') or '');                              row += 1
    _text_kv(row, "Industry", r.get('industry') or '');                          row += 1
    _text_kv(row, "Bucket", r.get('market_cap_bucket') or '');                   row += 1
    _money_kv(row, "Market cap (USD)", r.get('market_cap_usd') or r.get('market_cap'));  row += 1
    _money_kv(row, "Revenue TTM (USD)", r.get('revenue_ttm_usd') or r.get('revenue_ttm')); row += 1
    _text_kv(row, "As-of", str(r.get('balance_sheet_date') or ''));              row += 1
    for r2 in range(5, row):
        ws.row_dimensions[r2].height = 16

    # ── Investment thesis
    row += 1
    _section_rule(ws, row, "Investment thesis", span_cols=5); row += 1
    thesis = str(r.get('thesis', '') or 'Unresearched. No qualitative thesis on file; ranking driven entirely by quantitative signals (see Score decomposition below).')
    t = ws.cell(row=row, column=2, value=thesis)
    t.font = _font(size=11, name=SERIF, italic=True)
    t.alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=5)
    for r2 in range(row, row + 4):
        ws.row_dimensions[r2].height = 22
    row += 5

    # ── Score decomposition
    # Each leg = (label, value, kind). kind is one of:
    #   "score"  - two-decimal composite [0..1]
    #   "ratio"  - two-decimal multiplier
    #   "int"    - integer (e.g. cluster count out of 7)
    _section_rule(ws, row, "Score decomposition", span_cols=5); row += 1
    legs = [
        ("Asymmetry (raw)",      r.get('asymmetry_score'),       "score"),
        ("Upside leg",           r.get('upside_score'),           "score"),
        ("Downside floor",       r.get('downside_floor_score'),   "score"),
        ("Yartseva composite",   r.get('yartseva_score'),         "score"),
        ("Berezin growth",       r.get('berezin_score'),          "score"),
        ("Intrinsic discount",   r.get('intrinsic_discount'),     "score"),
        ("Qual multiplier (x)",  r.get('qual_mult') or 1.0,       "ratio"),
        ("Post-rally factor (x)", r.get('post_rally_factor') or 1.0, "ratio"),
        ("Cluster signals (of 7)", int(r.get('cluster_n') or 0), "int"),
    ]
    label_font = _font(size=9, bold=True, color=MUTED, name=SANS)
    num_font = _font(size=10, name=MONO)
    half = (len(legs) + 1) // 2
    for i in range(half):
        left = legs[i]
        right = legs[i + half] if i + half < len(legs) else None
        # Left column: label (text) | value (numeric)
        c2 = ws.cell(row=row, column=2, value=left[0])
        c2.font = label_font
        c2.alignment = _TXT_ALIGN_LEFT
        _write_left = {"score": _write_score, "ratio": _write_ratio, "int": _write_int}[left[2]]
        _write_left(ws, row, 3, left[1], font=num_font)
        if right:
            c4 = ws.cell(row=row, column=4, value=right[0])
            c4.font = label_font
            c4.alignment = _TXT_ALIGN_LEFT
            _write_right = {"score": _write_score, "ratio": _write_ratio, "int": _write_int}[right[2]]
            _write_right(ws, row, 5, right[1], font=num_font)
        for c in (2, 3, 4, 5):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

    # ── Valuation & balance sheet
    # kind: "pct" (decimal -> *100, one decimal place) or "ratio" (two decimals)
    _section_rule(ws, row, "Valuation & balance sheet", span_cols=5); row += 1
    sales_yoy_val = r.get('sales_yoy') if pd.notna(r.get('sales_yoy')) else r.get('rev_yoy')
    vals = [
        ("P/B",                       r.get('pb'),                       "ratio"),
        ("EV / EBITDA",               r.get('ev_ebitda'),                "ratio"),
        ("EV / EBIT",                 r.get('ev_ebit'),                  "ratio"),
        ("P/E",                       r.get('p_e'),                      "ratio"),
        ("P/S",                       r.get('p_s'),                      "ratio"),
        ("FCF yield (%)",             r.get('fcf_yield'),                "pct"),
        ("Net cash / mcap (%)",       r.get('net_cash_pct_mcap'),        "pct"),
        ("Cash / EV (x)",             r.get('cash_pct_ev'),              "ratio"),
        ("NCAV / mcap (%)",           r.get('ncav_pct_mcap'),            "pct"),
        ("Net debt / EBITDA (x)",     r.get('net_debt_ebitda'),          "ratio"),
        ("ROCE (%)",                  r.get('roce'),                     "pct"),
        ("EBITDA margin (%)",         r.get('ebitda_margin'),            "pct"),
        ("Insider ownership (%)",     r.get('insider_ownership_pct'),    "pct"),
        ("Rev 3y CAGR (%)",           r.get('rev_3y_cagr'),              "pct"),
        ("Rev YoY (TTM, %)",          sales_yoy_val,                     "pct"),
        ("12m price momentum (%)",    r.get('momentum_12m'),             "pct"),
    ]
    writers = {"pct": _write_pct, "ratio": _write_ratio}
    half = (len(vals) + 1) // 2
    for i in range(half):
        left = vals[i]
        right = vals[i + half] if i + half < len(vals) else None
        c2 = ws.cell(row=row, column=2, value=left[0])
        c2.font = label_font
        c2.alignment = _TXT_ALIGN_LEFT
        writers[left[2]](ws, row, 3, left[1], font=num_font)
        if right:
            c4 = ws.cell(row=row, column=4, value=right[0])
            c4.font = label_font
            c4.alignment = _TXT_ALIGN_LEFT
            writers[right[2]](ws, row, 5, right[1], font=num_font)
        for c in (2, 3, 4, 5):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

    # ── Flag panel
    flags = []
    if r.get('cash_gt_ev_flag') == 1:
        flags.append("Cash > EV  (downside floor in place)")
    if r.get('graham_net_net_flag') == 1:
        flags.append("Graham net-net  (mcap < NCAV)")
    if r.get('pew_negative_ev_flag') == 1:
        flags.append("Negative enterprise value")
    if pd.notna(r.get('pb')) and 0 < r.get('pb', 99) < 1.0:
        flags.append(f"Sub-book  (P/B {float(r['pb']):.2f})")
    if pd.notna(r.get('insider_ownership_pct')) and r.get('insider_ownership_pct', 0) >= 0.30:
        flags.append(f"Insider-aligned  ({float(r['insider_ownership_pct']) * 100:.0f}% insider)")
    if int(r.get('cluster_n') or 0) >= 4:
        flags.append(f"Cluster stack  ({int(r['cluster_n'])} of 7 inflection signals)")
    if flags:
        _section_rule(ws, row, "Signals firing", span_cols=5); row += 1
        for f in flags:
            c = ws.cell(row=row, column=2, value="·  " + f)
            c.font = _font(size=10, name=SERIF)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            ws.row_dimensions[row].height = 16
            row += 1

    # ── Notes
    notes_text = str(r.get('notes', '') or '')
    if notes_text:
        row += 1
        _section_rule(ws, row, "Notes", span_cols=5); row += 1
        n = ws.cell(row=row, column=2, value=notes_text)
        n.font = _font(size=9, name=SERIF, color=MUTED, italic=True)
        n.alignment = _align(wrap=True, v="top")
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=5)
        for r2 in range(row, row + 3):
            ws.row_dimensions[r2].height = 18
        row += 4

    # Footer hairline
    for c in range(1, 7):
        ws.cell(row=row + 1, column=c).border = _border(color=CRIMSON, top="medium")
    ws.cell(row=row + 2, column=2,
            value=f"Asymmetry Workbook  ·  Sheet {rank:02d} of {get_top_n()}").font = _font(
        size=8, italic=True, color=MUTED, name=SERIF)


_TOP_N_BOX = {'n': 50}


def get_top_n():
    return _TOP_N_BOX['n']


# --- Coverage sheet -----------------------------------------------------------
def build_coverage(ws: Worksheet, full_df: pd.DataFrame, top_df: pd.DataFrame):
    _set_col_widths(ws, {1: 4, 2: 22, 3: 12, 4: 10, 5: 10, 6: 10, 7: 14, 8: 12, 9: 4})
    _crimson_banner(ws, 1, "  Coverage audit", span_cols=9)

    REGIONS = {
        'NorthAmerica':   {'US', 'CA'},
        'LatinAmerica':   {'BR', 'MX', 'CL', 'AR'},
        'EU_Core':        {'UK', 'DE', 'FR', 'NL', 'BE', 'CH', 'IE', 'IT', 'AT'},
        'EU_Nordics':     {'SE', 'NO', 'DK', 'FI', 'IS'},
        'EU_Periphery':   {'ES', 'GR', 'PT'},
        'EU_CEE_Baltics': {'CZ', 'HU', 'EE', 'LV', 'LT', 'PL', 'RO'},
        'Asia_Developed': {'JP', 'KR', 'TW', 'HK', 'SG', 'AU', 'NZ'},
        'Asia_Emerging':  {'IN', 'ID', 'TH', 'MY', 'CN'},
        'MEA':            {'TR', 'ZA', 'IL', 'SA'},
    }

    # Block 1: universe-level verdict mix
    _section_rule(ws, 3, "Universe-level verdict mix", span_cols=8)
    hdrs = ["", "Universe", "GREEN", "YELLOW", "RED", "UNRESEARCHED", "% covered"]
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(row=4, column=i, value=h)
        c.font = _font(size=9, bold=True, color=CRIMSON_DARK, name=SANS)
        c.alignment = _align(h="center")
        c.border = _border(color=CRIMSON_DARK, bottom="thin")

    # Universe row: text label in col 2, integer counts in cols 3-6, pct in col 7
    total = len(full_df)
    g = int((full_df['verdict'] == 'GREEN').sum())
    y = int((full_df['verdict'] == 'YELLOW').sum())
    rd = int((full_df['verdict'] == 'RED').sum())
    u = int((full_df['verdict'] == 'UNRESEARCHED').sum())
    pct = (total - u) / total if total else 0.0

    mono_b = _font(size=10, name=MONO, bold=True)
    mono = _font(size=10, name=MONO)
    label_b = _font(size=10, name=SANS, bold=True)

    lbl = ws.cell(row=5, column=2, value="All")
    lbl.font = label_b
    lbl.alignment = _TXT_ALIGN_LEFT
    _write_int(ws, 5, 3, total, font=mono_b)
    _write_int(ws, 5, 4, g, font=mono)
    _write_int(ws, 5, 5, y, font=mono)
    _write_int(ws, 5, 6, rd, font=mono)
    _write_int(ws, 5, 7, u, font=mono)
    _write_pct(ws, 5, 8, pct, font=mono)  # uses ×100; one decimal place

    # Block 2: top-50 verdict mix
    _section_rule(ws, 7, "Top-50 verdict mix", span_cols=8)
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(row=8, column=i, value=h)
        c.font = _font(size=9, bold=True, color=CRIMSON_DARK, name=SANS)
        c.alignment = _align(h="center")
        c.border = _border(color=CRIMSON_DARK, bottom="thin")
    tg = int((top_df['verdict'] == 'GREEN').sum())
    ty = int((top_df['verdict'] == 'YELLOW').sum())
    tr = int((top_df['verdict'] == 'RED').sum())
    tu = int((top_df['verdict'] == 'UNRESEARCHED').sum())
    tpct = (len(top_df) - tu) / len(top_df) if len(top_df) else 0.0
    lbl = ws.cell(row=9, column=2, value="Top-50")
    lbl.font = label_b
    lbl.alignment = _TXT_ALIGN_LEFT
    _write_int(ws, 9, 3, len(top_df), font=mono_b)
    _write_int(ws, 9, 4, tg, font=mono)
    _write_int(ws, 9, 5, ty, font=mono)
    _write_int(ws, 9, 6, tr, font=mono)
    _write_int(ws, 9, 7, tu, font=mono)
    _write_pct(ws, 9, 8, tpct, font=mono)

    # Block 3: per-region top-10 verdict mix (qualitative coverage in regional top-10s)
    _section_rule(ws, 11, "Per-region top-10 verdict mix (qualitative coverage)", span_cols=8)
    rhdrs = ["Region", "Names", "GREEN", "YELLOW", "RED", "UNRESEARCHED", "% covered"]
    for i, h in enumerate(rhdrs, start=2):
        c = ws.cell(row=12, column=i, value=h)
        c.font = _font(size=9, bold=True, color=CRIMSON_DARK, name=SANS)
        c.alignment = _align(h="center")
        c.border = _border(color=CRIMSON_DARK, bottom="thin")

    row = 13
    for region, cs in REGIONS.items():
        sub = full_df[full_df['src'].isin(cs) & (full_df['verdict'] != 'RED')]\
            .sort_values('entry_today_asymmetry', ascending=False).head(10)
        if sub.empty:
            continue
        gg = int((sub['verdict'] == 'GREEN').sum())
        yy = int((sub['verdict'] == 'YELLOW').sum())
        uu = int((sub['verdict'] == 'UNRESEARCHED').sum())
        pp = (len(sub) - uu) / len(sub) if len(sub) else 0.0
        lbl = ws.cell(row=row, column=2, value=region)
        lbl.font = label_b
        lbl.alignment = _TXT_ALIGN_LEFT
        _write_int(ws, row, 3, len(sub), font=mono_b)
        _write_int(ws, row, 4, gg, font=mono)
        _write_int(ws, row, 5, yy, font=mono)
        _write_int(ws, row, 6, 0, font=mono)
        _write_int(ws, row, 7, uu, font=mono)
        _write_pct(ws, row, 8, pp, font=mono)
        for c in range(2, 9):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        row += 1

    # Notes
    row += 2
    _section_rule(ws, row, "Notes on coverage", span_cols=7); row += 1
    note = (
        "UNRESEARCHED is a default state, not exclusion: every name in the universe receives a "
        "score and competes for top-N slots with a 0.85x penalty vs. GREEN's 1.30x. RED is the "
        "only verdict that drops a name from per-region lists. The coverage fix run for this "
        "edition targeted the UNRESEARCHED entries within both the top-50 global and each "
        "region's top-10."
    )
    n = ws.cell(row=row, column=2, value=note)
    n.font = _font(size=10, italic=True, color=MUTED, name=SERIF)
    n.alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=7)
    for r in range(row, row + 4):
        ws.row_dimensions[r].height = 18


# --- References sheet ---------------------------------------------------------
def build_references(ws: Worksheet, refs: list[tuple[str, str]]):
    _set_col_widths(ws, {1: 4, 2: 80, 3: 4})
    _crimson_banner(ws, 1, "  References", span_cols=3)

    intro = (
        "Sources consulted for the qualitative diligence layer of this workbook. Items are "
        "listed alphabetically by author/issuer in Harvard author-date style."
    )
    ws.cell(row=3, column=2, value=intro).font = _font(size=10, italic=True, color=MUTED, name=SERIF)
    ws.cell(row=3, column=2).alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)

    for i, (key, full_ref) in enumerate(refs):
        row = 6 + i
        ws.cell(row=row, column=2, value=f"{key}  {full_ref}").font = _font(size=10, name=SERIF)
        ws.cell(row=row, column=2).alignment = _align(wrap=True, v="top")
        ws.row_dimensions[row].height = 30


# --- Main ---------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--top-n', type=int, default=50)
    ap.add_argument('--out', default='asymmetry_harvard_workbook.xlsx')
    ap.add_argument('--min-mcap', type=float, default=10_000_000)
    args = ap.parse_args()

    _TOP_N_BOX['n'] = args.top_n

    print(f'Loading data...', file=sys.stderr)
    quant = load_quant()
    verdicts = load_verdicts()
    print(f'  quant rows: {len(quant)}', file=sys.stderr)
    print(f'  verdicts:   {len(verdicts)}', file=sys.stderr)

    df = compute_scores(quant, verdicts)
    df = df[df['market_cap'].fillna(0) >= args.min_mcap]
    top = df.sort_values('entry_today_asymmetry', ascending=False).head(args.top_n).copy()
    top = top.reset_index(drop=True)
    print(f'  top-{args.top_n} verdicts: '
          f'GREEN {(top.verdict=="GREEN").sum()}, '
          f'YELLOW {(top.verdict=="YELLOW").sum()}, '
          f'RED {(top.verdict=="RED").sum()}, '
          f'UNRESEARCHED {(top.verdict=="UNRESEARCHED").sum()}', file=sys.stderr)

    # Counts
    n_total = len(df)
    n_green = (df['verdict'] == 'GREEN').sum()
    n_yellow = (df['verdict'] == 'YELLOW').sum()
    n_unresearched = (df['verdict'] == 'UNRESEARCHED').sum()

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    build_cover(cover, n_total, n_green, n_yellow, n_unresearched, args.top_n)

    method = wb.create_sheet("Methodology")
    build_methodology(method)

    idx = wb.create_sheet("Index")
    build_index(idx, top)

    for i, (_, row) in enumerate(top.iterrows(), start=1):
        sheet_name = _sheet_name_for(row['symbol'], i)
        ws = wb.create_sheet(sheet_name)
        build_name_sheet(ws, i, row)

    cov = wb.create_sheet("Coverage")
    build_coverage(cov, df, top)

    refs = [
        ("Yartseva, A. (2025)",
         "The Alchemy of Multibagger Stocks: An empirical investigation of factors that "
         "drive outperformance in the stock market. CAFE Working Paper 33, "
         "Birmingham City University. https://www.open-access.bcu.ac.uk/16180/"),
        ("Alta Fox Capital (2020)",
         "Makings of a Multibagger. Summer Intern Class Project. "
         "https://www.altafoxcapital.com/s/Makings-of-a-MultiBagger.pdf"),
        ("Mayer, C. (2015)",
         "100 Baggers: Stocks That Return 100-to-1 and How to Find Them. "
         "Laissez Faire Books."),
        ("Phelps, T. W. (1972)",
         "100 to 1 in the Stock Market. McGraw-Hill, New York."),
        ("Graham, B. (1934)", "Security Analysis. McGraw-Hill, New York."),
        ("Greenblatt, J. (2006)",
         "The Little Book That Beats the Market. Wiley."),
        ("Mauboussin, M. J. (2017)",
         "The Base Rate Book: Integrating the Past to Better Anticipate the Future. "
         "Credit Suisse."),
        ("Russo, T. (2014)",
         "Global Value: How to Spot Bubbles, Avoid Market Crashes, and Earn Big Returns "
         "in the Stock Market. Graham & Doddsville interview series."),
        ("Cassel, I. (MicroCapClub)",
         "Top 10 Things in Micro Cap Investing. https://microcapclub.com/"),
        ("SEC EDGAR (2026)",
         "XBRL company-facts API. https://data.sec.gov/api/xbrl/companyfacts/, "
         "accessed 24 June 2026."),
        ("yfinance (2026)",
         "Yahoo! Finance Python wrapper. https://github.com/ranaroussi/yfinance, "
         "accessed 20 June 2026."),
        ("financedatabase (2026)",
         "Open securities database. https://github.com/JerBouma/FinanceDatabase, "
         "accessed 20 June 2026."),
        ("Internal house frameworks",
         "The Pew negative-EV screen (pew_archetype.py) and Berezin / Stockcoach "
         "microcap deep-value composite (yartseva_db.py) are internal naming "
         "conventions for hand-rolled archetypes adapted from public deep-value / "
         "MicroCapClub heritage — not citations to published frameworks. The "
         "'archetype taxonomy' is internal; the joinyellowbrick.com platform's "
         "Multibagger Monitor inspired the cluster-A through G naming but is not "
         "the source of the specific tag definitions."),
    ]
    refs_sheet = wb.create_sheet("References")
    build_references(refs_sheet, refs)

    # Hide default tab-color-style by setting tab colors per group
    cover.sheet_view.showGridLines = False
    method.sheet_view.showGridLines = False
    idx.sheet_view.showGridLines = False
    cov.sheet_view.showGridLines = False
    refs_sheet.sheet_view.showGridLines = False
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        if ws.title in ("Cover", "Methodology", "Index", "Coverage", "References"):
            ws.sheet_properties.tabColor = Color(rgb=DARK_GREY)
        else:
            ws.sheet_properties.tabColor = Color(rgb=MUTED)

    wb.save(args.out)
    print(f'wrote {args.out}', file=sys.stderr)


if __name__ == '__main__':
    main()
