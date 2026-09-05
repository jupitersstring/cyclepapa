"""Per-country Top-N-by-archetype book.

The global archetype book ranks across the whole universe, so cheap
high-scoring markets (Korea, Japan smalls, Indonesia...) dominate every
sheet. This book partitions by country first: one sheet per market, and
within it the top N names for EACH archetype, so every market's best
representatives of every pattern are visible side by side.

Reuses load_data + labels from build_archetype_book and the Harvard
styling from build_harvard_workbook. Sections are stacked vertically per
archetype, ordered by match count within the country; rows sorted by
entry_today_asymmetry. Includes the Lynch multiples (PEGY, EV/EBITDA-GY)
as columns.

Output: country_archetype_book.xlsx
"""
from __future__ import annotations
import argparse
import sys

import pandas as pd

from build_harvard_workbook import (
    Workbook,
    INK, MUTED, RULE,
    _font,
    _crimson_banner, _section_rule, _verdict_badge,
    _write_money, _write_pct, _write_score, _write_int,
    _NUM_ALIGN_RIGHT, _NUM_ALIGN_CENTER, _TXT_ALIGN_LEFT,
)
from build_archetype_book import load_data, ARCHETYPE_LABELS, _sheet_safe
from openpyxl.styles import Border, Side
from otc_flag import add_otc_mode_arg, apply_otc_mode

SORT_COL = 'entry_confirmed'

HEADERS = ['#', 'Ticker', 'Name', 'Sector', 'Bucket', 'Mcap (USD)',
           'Verdict', 'ETA', 'Asym', 'EV/EBITDA', 'P/E', 'PEGY', 'EV-GY',
           'FCF yld %', 'ROCE %', 'ND/EBITDA', 'Mom 12m %', 'Arch #', 'P/S', 'P/B']
N_COLS = len(HEADERS)
WIDTHS = {1: 4, 2: 12, 3: 34, 4: 16, 5: 11, 6: 15, 7: 12, 8: 7, 9: 7,
          10: 9, 11: 8, 12: 7, 13: 7, 14: 9, 15: 8, 16: 10, 17: 10, 18: 7, 19: 7, 20: 7}


def _write_country_sheet(ws, cdf, country, arch_cols, n_top):
    f_bold = _font(bold=True, color=INK)
    f_bold_muted = _font(bold=True, color=MUTED)
    f_text = _font(color=INK)
    f_text_muted = _font(color=MUTED)
    f_italic_muted = _font(italic=True, color=MUTED)

    for col, w in WIDTHS.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    t = ws.cell(row=2, column=1,
                value=f"{country} — Top {n_top} per archetype   "
                      f"({len(cdf):,} eligible names)")
    t.font = f_bold
    t.alignment = _TXT_ALIGN_LEFT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
    ws.row_dimensions[2].height = 22
    for c in range(1, N_COLS + 1):
        ws.cell(row=3, column=c).border = Border(bottom=Side(style='thin', color=INK))
    ws.row_dimensions[3].height = 4

    # Archetypes ordered by in-country match count, skip empty ones
    counts = [(col, int(cdf[col].fillna(0).sum())) for col in arch_cols]
    counts = [(col, n) for col, n in counts if n > 0]
    counts.sort(key=lambda x: -x[1])

    row = 5
    for col, n_match in counts:
        label = ARCHETYPE_LABELS.get(col, col)
        sub = (cdf[cdf[col].fillna(0) == 1]
               .sort_values(SORT_COL, ascending=False, na_position='last')
               .head(n_top))
        _section_rule(ws, row, f"{label}   —   {n_match:,} matches in {country}",
                      span_cols=N_COLS)
        row += 1
        for i, h in enumerate(HEADERS, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = f_bold_muted
            c.alignment = (_TXT_ALIGN_LEFT if i in (2, 3, 4) else
                           _NUM_ALIGN_CENTER if i in (5, 7) else
                           _NUM_ALIGN_RIGHT)
        row += 1
        for rank, (_, r) in enumerate(sub.iterrows(), start=1):
            _write_int(ws, row, 1, rank, font=f_text_muted)
            ws.cell(row=row, column=2, value=r['symbol']).font = f_bold
            ws.cell(row=row, column=2).alignment = _TXT_ALIGN_LEFT
            _t = lambda v: '' if pd.isna(v) else str(v)
            ws.cell(row=row, column=3, value=_t(r.get('name'))[:48]).font = f_text
            ws.cell(row=row, column=3).alignment = _TXT_ALIGN_LEFT
            ws.cell(row=row, column=4, value=_t(r.get('sector'))[:20]).font = f_text_muted
            ws.cell(row=row, column=4).alignment = _TXT_ALIGN_LEFT
            ws.cell(row=row, column=5, value=_t(r.get('market_cap_bucket'))).font = f_text_muted
            ws.cell(row=row, column=5).alignment = _NUM_ALIGN_CENTER
            _write_money(ws, row, 6, r.get('market_cap'), font=f_text)
            _verdict_badge(ws, row, 7, r.get('verdict', 'UNRESEARCHED'))
            _write_score(ws, row, 8, r.get('entry_today_asymmetry'), font=f_bold)
            _write_score(ws, row, 9, r.get('asymmetry_score'), font=f_text)
            _write_score(ws, row, 10, r.get('ev_ebitda'), font=f_text)
            _write_score(ws, row, 11, r.get('p_e'), font=f_text)
            _write_score(ws, row, 12, r.get('pegy'), font=f_text)
            _write_score(ws, row, 13, r.get('ev_ebitda_gy'), font=f_text)
            _write_pct(ws, row, 14, r.get('fcf_yield'), font=f_text)
            _write_pct(ws, row, 15, r.get('roce'), font=f_text)
            _write_score(ws, row, 16, r.get('net_debt_ebitda'), font=f_text)
            _write_pct(ws, row, 17, r.get('momentum_12m'), font=f_text)
            _write_int(ws, row, 18,
                       int(r['archetype_count']) if pd.notna(r.get('archetype_count')) else 0,
                       font=f_text_muted)
            _write_score(ws, row, 19, r.get('p_s'), font=f_text)
            _write_score(ws, row, 20, r.get('pb'), font=f_text)
            for c in range(1, N_COLS + 1):
                ws.cell(row=row, column=c).border = Border(
                    bottom=Side(style='thin', color=RULE))
            ws.row_dimensions[row].height = 15
            row += 1
        row += 2  # gap between archetype sections
    ws.sheet_view.showGridLines = False


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--n', type=int, default=5,
                    help='top N per archetype per country')
    ap.add_argument('--min-names', type=int, default=25,
                    help='minimum eligible names for a country to get a sheet')
    ap.add_argument('--out', default='country_archetype_book.xlsx')
    add_otc_mode_arg(ap)
    args = ap.parse_args()

    df, arch_cols = load_data()
    df = apply_otc_mode(df, args.otc_mode)
    df['src'] = df['src'].fillna('').astype(str).str.upper()
    print(f'  {len(df):,} eligible rows, {len(arch_cols)} archetypes',
          file=sys.stderr)

    wb = Workbook()
    cover = wb.active
    cover.title = 'Cover'
    for col, w in {1: 8, 2: 22, 3: 12, 4: 14, 5: 30}.items():
        cover.column_dimensions[cover.cell(row=1, column=col).column_letter].width = w
    _crimson_banner(cover, 2, 'COUNTRY x ARCHETYPE', span_cols=5)
    sub = cover.cell(row=4, column=1,
                     value=f'Top {args.n} names per archetype within each market '
                           f'— so no single country dominates the ranking')
    sub.font = _font(italic=True, color=MUTED)
    sub.alignment = _TXT_ALIGN_LEFT
    f_bold_muted = _font(bold=True, color=MUTED)
    f_text = _font(color=INK)
    f_text_muted = _font(color=MUTED)

    order = (df.groupby('src').size().sort_values(ascending=False))
    countries = [c for c, n in order.items()
                 if c and n >= args.min_names]

    _section_rule(cover, 7, 'Markets', span_cols=5)
    for i, h in enumerate(['#', 'Market', 'Names', 'Multi-arch', 'Top name by ETA'], start=1):
        cover.cell(row=8, column=i, value=h).font = f_bold_muted
    r = 9
    for i, ctry in enumerate(countries, start=1):
        cdf = df[df['src'] == ctry]
        _write_int(cover, r, 1, i, font=f_text_muted)
        cover.cell(row=r, column=2, value=ctry).font = f_text
        _write_int(cover, r, 3, len(cdf), font=f_text)
        _write_int(cover, r, 4, int((cdf['archetype_count'].fillna(0) >= 2).sum()),
                   font=f_text)
        top = cdf.sort_values(SORT_COL, ascending=False, na_position='last').head(1)
        cover.cell(row=r, column=5,
                   value=(top.iloc[0]['symbol'] if len(top) else '—')).font = f_text_muted
        r += 1

    for ctry in countries:
        cdf = df[df['src'] == ctry]
        ws = wb.create_sheet(_sheet_safe(ctry))
        _write_country_sheet(ws, cdf, ctry, arch_cols, args.n)
        print(f'  {ctry}: {len(cdf):,} names', file=sys.stderr)

    cover.sheet_view.showGridLines = False
    wb.save(args.out)
    print(f'wrote {args.out}  ({1 + len(countries)} sheets: Cover + '
          f'{len(countries)} countries)', file=sys.stderr)


if __name__ == '__main__':
    main()
