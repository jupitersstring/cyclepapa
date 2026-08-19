"""Emit a flat 'top N per country' table for at-a-glance review.

Uses the same entry_today_asymmetry computation as the Harvard workbook:
asymmetry_score * intrinsic_boost * qual_multiplier * post_rally_factor.

Default: top 10 per country, min mcap $10M USD, RED excluded.

Output: top_n_by_country.csv and a Harvard-styled XLSX one-pager.
"""
from __future__ import annotations
import argparse
import glob
import os
import sys

import numpy as np
import pandas as pd


def load_verdicts() -> pd.DataFrame:
    frames = []
    for path, default in [
        ('qualitative_aligned_green.csv', 'GREEN'),
        ('qualitative_red_avoid.csv', 'RED'),
        ('qualitative_extended_verdicts.csv', None),
    ]:
        if not os.path.exists(path):
            continue
        try:
            d = pd.read_csv(path)
        except pd.errors.ParserError:
            d = pd.read_csv(path, engine='python', on_bad_lines='skip', quoting=3)
        if 'verdict' not in d.columns and default is not None:
            d['verdict'] = default
        keep = [c for c in ['symbol', 'verdict', 'why', 'why_avoid', 'thesis'] if c in d.columns]
        d = d[keep].copy()
        tc = next((c for c in ['why', 'why_avoid', 'thesis'] if c in d.columns), None)
        d['thesis'] = d[tc] if tc else ''
        frames.append(d[['symbol', 'verdict', 'thesis']])
    if not frames:
        return pd.DataFrame(columns=['symbol', 'verdict', 'thesis'])
    return pd.concat(frames, ignore_index=True).drop_duplicates('symbol', keep='last')


def load_quant() -> pd.DataFrame:
    df = pd.read_csv('asymmetry_global.csv').drop_duplicates('symbol')
    extra_cols = [
        'symbol', 'net_cash_pct_mcap', 'ncav_pct_mcap', 'cash_pct_ev',
        'not_priced_in_score', 'inflection_flag', 'inflection_score',
        # Headline valuation columns surfaced into the per-country tables
        'ev_ebitda', 'ev_ebit', 'p_e', 'p_s', 'pb', 'p_tb',
        'fcf_yield', 'roce', 'net_debt_ebitda',
        'ebitda_margin', 'gross_margin', 'insider_ownership_pct',
        'tangible_equity_pct',
    ]
    frames = []
    for f in sorted(set(glob.glob('*_yartseva.csv'))):
        try:
            d = pd.read_csv(f, usecols=lambda c: c in extra_cols)
        except Exception:
            continue
        if 'symbol' in d.columns:
            frames.append(d)
    if frames:
        extra = pd.concat(frames, ignore_index=True).drop_duplicates('symbol', keep='first')
        mc = ['symbol'] + [c for c in extra.columns if c != 'symbol' and c not in df.columns]
        df = df.merge(extra[mc], on='symbol', how='left')
    return df


def compute_eta(df: pd.DataFrame, verdicts: pd.DataFrame) -> pd.DataFrame:
    # asymmetry_global may carry a verdict/thesis column from the
    # enrich_asymmetry_global post-pass. Drop those before merging the
    # fresh verdicts file (which is the most up-to-date source).
    df = df.drop(columns=[c for c in ('verdict', 'thesis') if c in df.columns])
    df = df.merge(verdicts, on='symbol', how='left')
    df['verdict'] = df['verdict'].fillna('UNRESEARCHED')
    df['thesis'] = df['thesis'].fillna('')

    soft = {'GREEN': 1.10, 'YELLOW': 0.85, 'RED': 0.40}
    df['qual_mult'] = df['verdict'].map(soft).fillna(1.0)

    def col(c, d=0.0):
        return df[c].fillna(d) if c in df.columns else pd.Series(d, index=df.index)

    def c01(s):
        return s.clip(0, 1).fillna(0)

    nc = c01(col('net_cash_pct_mcap'))
    ncav = c01(col('ncav_pct_mcap'))
    sub_book = c01(1.0 - col('pb', 2.0).clip(lower=0.01))
    cash_ev = c01((col('cash_pct_ev') - 1.0).clip(0, 2) / 2.0)
    npi = c01(col('not_priced_in_score'))
    df['intrinsic_discount'] = (
        0.30 * nc + 0.20 * ncav + 0.20 * sub_book + 0.15 * cash_ev + 0.15 * npi
    )
    boost = (1.0 + (df['intrinsic_discount'] - 0.25)).clip(0.5, 1.5)

    mom = col('momentum_12m').clip(-0.5, None)
    pr = pd.Series(1.0, index=df.index)
    mid = (mom > 0.30) & (mom <= 1.0)
    hi = (mom > 1.0) & (mom <= 3.0)
    ex = mom > 3.0
    pr.loc[mid] = 1.0 - (mom[mid] - 0.30) / 0.70 * 0.25
    pr.loc[hi] = 0.75 - (mom[hi] - 1.0) / 2.0 * 0.30
    pr.loc[ex] = 0.40
    df['post_rally_factor'] = pr.round(3)
    # Multi-measure confirmation upweight (pool-preserving, <=30%): float names
    # up where independent accounting measures + insider alignment agree — the
    # same "upweight where measures agree" treatment the archetype books apply
    # via entry_confirmed. Ranking only; raw asymmetry_score is still shown.
    _cfo = pd.to_numeric(df.get('confirm_overall'), errors='coerce').fillna(0.0)
    _bbs = pd.to_numeric(df.get('buyback_score'), errors='coerce').fillna(0.0)
    df['confirm_mult'] = 1.0 + 0.20 * _cfo + 0.10 * _bbs
    df['entry_today_asymmetry'] = (
        df['asymmetry_score'] * boost * df['qual_mult'] * pr * df['confirm_mult']
    )

    # Parallel entry-today score for the inflection style. Note: we do NOT
    # apply the post-rally factor here — momentum / breakout names should
    # not be penalised for trending up; that's the whole point of the
    # composite. Verdict multiplier + confirmation upweight still apply.
    if 'inflection_asymmetry_score' in df.columns:
        df['entry_today_inflection'] = (
            df['inflection_asymmetry_score'] * boost * df['qual_mult'] * df['confirm_mult']
        )
    else:
        df['entry_today_inflection'] = np.nan
    return df


COUNTRY_NAMES = {
    'US': 'United States', 'CA': 'Canada', 'UK': 'United Kingdom',
    'DE': 'Germany', 'FR': 'France', 'IT': 'Italy', 'NL': 'Netherlands',
    'BE': 'Belgium', 'CH': 'Switzerland', 'IE': 'Ireland', 'AT': 'Austria',
    'SE': 'Sweden', 'NO': 'Norway', 'DK': 'Denmark', 'FI': 'Finland', 'IS': 'Iceland',
    'ES': 'Spain', 'PT': 'Portugal', 'GR': 'Greece',
    'CZ': 'Czechia', 'HU': 'Hungary', 'PL': 'Poland', 'RO': 'Romania',
    'EE': 'Estonia', 'LV': 'Latvia', 'LT': 'Lithuania',
    'JP': 'Japan', 'KR': 'Korea', 'TW': 'Taiwan', 'HK': 'Hong Kong',
    'CN': 'China', 'SG': 'Singapore', 'AU': 'Australia', 'NZ': 'New Zealand',
    'IN': 'India', 'ID': 'Indonesia', 'TH': 'Thailand', 'MY': 'Malaysia',
    'BR': 'Brazil', 'MX': 'Mexico', 'CL': 'Chile', 'AR': 'Argentina',
    'TR': 'Turkey', 'ZA': 'South Africa', 'IL': 'Israel', 'SA': 'Saudi Arabia',
}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--n', type=int, default=10, help='top N per country')
    ap.add_argument('--min-mcap', type=float, default=10_000_000)
    ap.add_argument('--exclude-red', action='store_true', default=True)
    ap.add_argument('--include-red', dest='exclude_red', action='store_false')
    ap.add_argument('--sort-by', default='asymmetry',
                    choices=['asymmetry', 'inflection'],
                    help='rank within each country by value-and-contra '
                         'asymmetry or by inflection asymmetry')
    ap.add_argument('--out-csv', default='top_n_by_country.csv')
    ap.add_argument('--out-xlsx', default='top_n_by_country.xlsx')
    args = ap.parse_args()

    print('loading...', file=sys.stderr)
    quant = load_quant()
    verdicts = load_verdicts()
    df = compute_eta(quant, verdicts)
    # USD-normalise market cap so the min-mcap gate is comparable across markets
    if 'market_cap_usd' in df.columns:
        df['market_cap'] = (pd.to_numeric(df['market_cap_usd'], errors='coerce')
                            .fillna(pd.to_numeric(df['market_cap'], errors='coerce')))
    df = df[df['market_cap'].fillna(0) >= args.min_mcap]
    if args.exclude_red:
        df = df[df['verdict'] != 'RED']
    print(f'  {len(df):,} rows after filters, sort by {args.sort_by}', file=sys.stderr)

    sort_col = ('entry_today_asymmetry' if args.sort_by == 'asymmetry'
                else 'entry_today_inflection')
    # Rank within each country by the chosen score
    df['country_rank'] = (
        df.sort_values(sort_col, ascending=False)
          .groupby('src')
          .cumcount() + 1
    )
    top = df[df['country_rank'] <= args.n].copy()
    top = top.sort_values(['src', 'country_rank'])

    # Friendly country name
    top['country_name'] = top['src'].map(COUNTRY_NAMES).fillna(top['src'])

    out_cols = [
        'country_name', 'src', 'country_rank',
        'symbol', 'name', 'sector', 'industry', 'market_cap_bucket',
        'market_cap', 'verdict',
        'entry_today_asymmetry', 'entry_today_inflection',
        'asymmetry_score', 'inflection_asymmetry_score',
        'yartseva_score', 'inflection_score', 'inflection_flag',
        'berezin_score',
        'intrinsic_discount', 'cluster_n', 'confirm_overall',
        # Headline valuation set
        'ev_ebitda', 'ev_ebit', 'p_e', 'p_s', 'pb', 'p_tb',
        'fcf_yield', 'roce', 'net_debt_ebitda',
        'ebitda_margin', 'gross_margin', 'insider_ownership_pct',
        'momentum_12m',
        'thesis',
    ]
    out_cols = [c for c in out_cols if c in top.columns]
    out = top[out_cols].copy()
    out.to_csv(args.out_csv, index=False)
    print(f'  wrote {args.out_csv}: {len(out):,} rows across '
          f'{out["src"].nunique()} countries', file=sys.stderr)

    # XLSX with per-country tabs + cover + index. `df` is the FULL universe
    # post-filter (for headline figures); `out` is the top-N table.
    _write_xlsx(out, args.out_xlsx, args.n, full_df=df, sort_col=sort_col)


def _write_xlsx(out: pd.DataFrame, path: str, n: int, full_df=None, sort_col='entry_today_asymmetry'):
    """Harvard-style workbook with PER-COUNTRY TABS.

    Layout:
      Cover    headline figures for the whole universe + ranking method
      Index    one row per country with summary stats + hyperlinks
      <CC>     one tab per country with country-specific headline figures
               then the top-N table for that country

    `out`     is the cross-country top-N concatenation (one row per
              top-N name per country).
    `full_df` (optional) is the pre-top-N universe used for
              country-level headlines (universe size, GREEN counts, etc.).
              Falls back to `out` if not provided.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Monochrome palette — single font + grey-only highlights
    INK = '000000'
    DARK_GREY = '404040'
    MUTED = '707070'
    RULE = 'B0B0B0'
    LIGHT_GREY = 'E8E8E8'
    PALE_GREY = 'F2F2F2'
    WHITE = 'FFFFFF'
    CRIMSON = INK
    CRIMSON_DARK = DARK_GREY
    GREEN_BG = LIGHT_GREY
    YELLOW_BG = PALE_GREY
    GRAY_BG = WHITE
    FONT_NAME = 'Cambria'   # Harvard-style serif default
    FONT_SIZE = 10

    EM_DASH = '–'
    FMT_INT_RAW = '#,##0;(#,##0);"–"'
    FMT_ONE = '#,##0.0;(#,##0.0);"–"'
    FMT_TWO = '#,##0.00;(#,##0.00);"–"'

    A_LEFT = Alignment(horizontal='left', vertical='center')
    A_RIGHT = Alignment(horizontal='right', vertical='center')
    A_CENTER = Alignment(horizontal='center', vertical='center')

    wb = Workbook()
    if full_df is None:
        full_df = out

    # Shared font registry (Cambria 10pt, single size)
    f_text = Font(name=FONT_NAME, size=FONT_SIZE)
    f_text_muted = Font(name=FONT_NAME, size=FONT_SIZE, color=MUTED)
    f_text_italic = Font(name=FONT_NAME, size=FONT_SIZE, italic=True)
    f_text_italic_muted = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color=MUTED)
    f_bold = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    f_bold_muted = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=MUTED)
    f_red_verdict = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=True)

    def _cell_borders(ws, row, col, **sides):
        b_args = {}
        for side, style in sides.items():
            if style:
                b_args[side] = Side(style=style[0], color=style[1])
        ws.cell(row=row, column=col).border = Border(**b_args)

    def _put_num(ws, row, col, value, fmt, scale=1.0, font=None, align=A_RIGHT):
        cell = ws.cell(row=row, column=col)
        if value is None or (isinstance(value, float) and pd.isna(value)):
            cell.value = EM_DASH
        else:
            try:
                cell.value = float(value) * scale
                cell.number_format = fmt
            except (TypeError, ValueError):
                cell.value = EM_DASH
        cell.alignment = align
        if font is not None:
            cell.font = font

    def _put_int(ws, row, col, value, font=None):
        _put_num(ws, row, col, value, FMT_INT_RAW, font=font)

    def _put_money(ws, row, col, value, font=None):
        _put_num(ws, row, col, value, FMT_INT_RAW, font=font)

    def _put_score(ws, row, col, value, font=None):
        _put_num(ws, row, col, value, FMT_TWO, font=font)

    def _put_pct(ws, row, col, value, font=None):
        # value stored as fraction (e.g. 0.25) — display 25.0
        _put_num(ws, row, col, value, FMT_ONE, scale=100.0, font=font)

    def _put_text(ws, row, col, value, font=None, align=A_LEFT):
        cell = ws.cell(row=row, column=col,
                       value=value if value not in (None, '') else EM_DASH)
        cell.alignment = align
        if font is not None:
            cell.font = font

    def _verdict_marker(ws, row, col, verdict):
        """Text-only verdict marker — no fills, no borders. Matches the
        Harvard workbook spec."""
        cell = ws.cell(row=row, column=col, value=verdict)
        if verdict == 'GREEN':
            cell.font = f_bold
        elif verdict == 'RED':
            cell.font = f_red_verdict
        elif verdict == 'UNRESEARCHED':
            cell.font = f_text_italic_muted
        else:
            cell.font = f_text
        cell.alignment = A_CENTER

    def _section_label(ws, row, text, span_cols):
        cell = ws.cell(row=row, column=1, value=text.upper())
        cell.font = f_bold
        cell.alignment = A_LEFT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
        ws.row_dimensions[row].height = 16

    # Table column layout (post-headline-valuation addition):
    #   1 #             2 Ticker        3 Name          4 Sector
    #   5 Bucket        6 Mcap (USD)    7 Verdict       8 ETA
    #   9 Asym         10 EV/EBITDA    11 P/E         12 P/B
    #  13 FCF yld %    14 ROIC %       15 ND/EBITDA   16 EBITDA margin %
    #  17 Mom 12m %    18 Yartseva     19 Cluster
    N_COLS = 20

    def _write_table_row(ws, row, r, cols=N_COLS):
        """Write one country-rank row with valuation headline columns."""
        _put_int(ws, row, 1, int(r['country_rank']), font=f_text_muted)
        _put_text(ws, row, 2, r['symbol'], font=f_bold)
        _put_text(ws, row, 3, str(r.get('name') or '')[:50], font=f_text)
        _put_text(ws, row, 4, str(r.get('sector') or ''), font=f_text_muted)
        _put_text(ws, row, 5, str(r.get('market_cap_bucket') or ''), font=f_text_muted,
                  align=A_CENTER)
        _put_money(ws, row, 6, r.get('market_cap'), font=f_text)
        _verdict_marker(ws, row, 7, r['verdict'])
        _put_score(ws, row, 8, r.get('entry_today_asymmetry'), font=f_bold)
        _put_score(ws, row, 9, r.get('asymmetry_score'), font=f_text)

        # Headline valuation block (cols 10-17)
        _put_score(ws, row, 10, r.get('ev_ebitda'), font=f_text)
        _put_score(ws, row, 11, r.get('p_e'), font=f_text)
        _put_score(ws, row, 12, r.get('pb'), font=f_text)
        _put_pct(ws, row, 13, r.get('fcf_yield'), font=f_text)
        _put_pct(ws, row, 14, r.get('roce'), font=f_text)
        _put_score(ws, row, 15, r.get('net_debt_ebitda'), font=f_text)
        _put_pct(ws, row, 16, r.get('ebitda_margin'), font=f_text)
        _put_pct(ws, row, 17, r.get('momentum_12m'), font=f_text)

        _put_score(ws, row, 18, r.get('yartseva_score'), font=f_text_muted)
        _cn = r.get('cluster_n')
        _put_int(ws, row, 19, int(_cn) if pd.notna(_cn) else 0, font=f_text_muted)
        _put_score(ws, row, 20, r.get('confirm_overall'), font=f_text_muted)

        # Faint hairline under each row
        for cidx in range(1, cols + 1):
            ws.cell(row=row, column=cidx).border = Border(
                bottom=Side(style='thin', color=RULE))

    def _write_table_header(ws, row):
        headers = ['#', 'Ticker', 'Name', 'Sector', 'Bucket',
                   'Mcap (USD)', 'Verdict', 'ETA', 'Asym',
                   'EV/EBITDA', 'P/E', 'P/B',
                   'FCF yld %', 'ROIC %', 'ND/EBITDA', 'EBITDA m %',
                   'Mom 12m %', 'Yartseva', 'Cluster', 'Confirm']
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = f_bold_muted
            c.alignment = (A_LEFT if i in (2, 3, 4) else
                           A_CENTER if i in (5, 7) else
                           A_RIGHT)
        # Thin black rule between header and first data row
        for i in range(1, len(headers) + 1):
            ws.cell(row=row + 1, column=i).border = Border(
                top=Side(style='thin', color=INK))

    def _common_col_widths(ws):
        widths = {
            1: 4, 2: 11, 3: 32, 4: 16, 5: 11, 6: 17,
            7: 12, 8: 8, 9: 8,
            10: 10, 11: 8, 12: 8, 13: 10, 14: 9, 15: 10, 16: 10, 17: 11,
            18: 9, 19: 8, 20: 9,
        }
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

    # ===== COVER SHEET =====
    ws = wb.active
    ws.title = 'Cover'
    ws.column_dimensions['A'].width = 6
    for col_letter, w in [('B', 24), ('C', 24), ('D', 24), ('E', 24), ('F', 24), ('G', 24)]:
        ws.column_dimensions[col_letter].width = w
    ws.column_dimensions['H'].width = 6

    sort_label = 'Entry-today asymmetry (value + contra)' if sort_col == 'entry_today_asymmetry' \
        else 'Entry-today inflection (breakout + 52w-high)'

    # Masthead
    title = ws.cell(row=3, column=2,
                    value=f"Top {n} per Country  —  {sort_label}")
    title.font = f_bold
    title.alignment = A_LEFT
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)
    ws.row_dimensions[3].height = 22
    # Rule under masthead
    for c in range(2, 8):
        ws.cell(row=4, column=c).border = Border(bottom=Side(style='thin', color=INK))
    ws.row_dimensions[4].height = 4

    sub = ws.cell(row=5, column=2,
                  value="One tab per country, each with country-specific headline figures + a ranked top-N table")
    sub.font = f_text_italic
    sub.alignment = A_LEFT
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    note = ws.cell(row=7, column=2,
                   value="Yartseva-aligned upside  ·  Graham downside floor  ·  EDGAR XBRL ground truth  ·  As of 24 June 2026")
    note.font = f_text_italic_muted
    note.alignment = A_LEFT
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=7)

    # Headline figures — universe-wide
    _section_label(ws, 9, "Headline figures", span_cols=7)

    n_universe = len(full_df)
    n_countries = full_df['src'].nunique()
    g_total = int((full_df['verdict'] == 'GREEN').sum())
    y_total = int((full_df['verdict'] == 'YELLOW').sum())
    u_total = int((full_df['verdict'] == 'UNRESEARCHED').sum())

    tiles = [
        ("UNIVERSE", f"{n_universe:,}", "names ranked"),
        ("COUNTRIES", f"{n_countries}", "with at least one name"),
        ("TOP-N", f"{n}", "per country"),
        ("GREEN", f"{g_total}", "high conviction"),
        ("YELLOW", f"{y_total}", "risk-flagged"),
        ("UNRESEARCHED", f"{u_total:,}", "no thesis yet"),
    ]
    for i, (lbl, val, sub_lbl) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=10, column=col, value=lbl).font = f_text_italic_muted
        ws.cell(row=10, column=col).alignment = A_LEFT
        ws.cell(row=11, column=col, value=val).font = f_bold
        ws.cell(row=11, column=col).alignment = A_LEFT
        ws.cell(row=12, column=col, value=sub_lbl).font = f_text_italic_muted
        ws.cell(row=12, column=col).alignment = A_LEFT
    ws.row_dimensions[11].height = 24
    for c in range(2, 8):
        ws.cell(row=13, column=c).border = Border(top=Side(style='thin', color=INK))
    ws.row_dimensions[13].height = 4

    # Index of countries below headline figures (rows 15+)
    _section_label(ws, 15, "Index by country", span_cols=7)

    # Index columns: Country | ISO | n (universe) | n (NMS) | GREEN | top scorer | top ETA
    idx_headers = ['Country', 'ISO', 'Total', 'NMS', 'GREEN', 'Top scorer', 'Top ETA']
    for i, h in enumerate(idx_headers, start=2):
        c = ws.cell(row=16, column=i, value=h)
        c.font = f_bold_muted
        c.alignment = A_LEFT if i in (2, 7) else A_RIGHT
    for c in range(2, 9):
        ws.cell(row=17, column=c).border = Border(top=Side(style='thin', color=INK))

    # Build per-country index rows
    nms_buckets = {'Nano Cap', 'Micro Cap', 'Small Cap'}
    country_rows = []
    for src_code in sorted(full_df['src'].dropna().unique()):
        sub = full_df[full_df['src'] == src_code]
        if sub.empty:
            continue
        cname = (out[out.src == src_code]['country_name'].iloc[0]
                 if not out[out.src == src_code].empty
                 else src_code)
        country_rows.append({
            'src': src_code,
            'country_name': cname,
            'total': len(sub),
            'nms': int(sub['market_cap_bucket'].isin(nms_buckets).sum()),
            'green': int((sub['verdict'] == 'GREEN').sum()),
            'top_sym': sub.sort_values(sort_col, ascending=False).iloc[0]['symbol']
            if sub.sort_values(sort_col, ascending=False)[sort_col].notna().any() else '',
            'top_eta': sub[sort_col].max() if sub[sort_col].notna().any() else None,
        })
    country_rows.sort(key=lambda r: -(r['top_eta'] or 0))

    row_i = 18
    for cr in country_rows:
        sheet_name = _country_sheet_name(cr['src'])
        # Country name + hyperlink
        c_name_cell = ws.cell(row=row_i, column=2, value=cr['country_name'])
        c_name_cell.font = f_text
        c_name_cell.alignment = A_LEFT
        c_name_cell.hyperlink = f"#'{sheet_name}'!A1"
        _put_text(ws, row_i, 3, cr['src'], font=f_text_muted, align=A_CENTER)
        _put_int(ws, row_i, 4, cr['total'], font=f_text)
        _put_int(ws, row_i, 5, cr['nms'], font=f_text)
        _put_int(ws, row_i, 6, cr['green'], font=f_text)
        _put_text(ws, row_i, 7, cr['top_sym'], font=f_bold, align=A_LEFT)
        _put_score(ws, row_i, 8, cr['top_eta'], font=f_text)
        # Faint hairline
        for c in range(2, 9):
            ws.cell(row=row_i, column=c).border = Border(
                bottom=Side(style='thin', color=RULE))
        row_i += 1

    ws.sheet_view.showGridLines = False

    # ===== PER-COUNTRY TABS =====
    for cr in country_rows:
        src_code = cr['src']
        country_name = cr['country_name']
        sub_full = full_df[full_df['src'] == src_code]
        sub_top = out[out['src'] == src_code]

        sheet_name = _country_sheet_name(src_code)
        sheet = wb.create_sheet(sheet_name)
        _common_col_widths(sheet)

        # Masthead
        t = sheet.cell(row=2, column=1,
                       value=f"{country_name}  ({src_code})")
        t.font = f_bold
        t.alignment = A_LEFT
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
        sheet.row_dimensions[2].height = 22

        for c in range(1, 20):
            sheet.cell(row=3, column=c).border = Border(bottom=Side(style='thin', color=INK))
        sheet.row_dimensions[3].height = 4

        # Headline figures (country-specific)
        _section_label(sheet, 5, "Headline figures", span_cols=19)

        n_country_total = len(sub_full)
        n_country_nms = int(sub_full['market_cap_bucket'].isin(nms_buckets).sum())
        n_country_green = int((sub_full['verdict'] == 'GREEN').sum())
        n_country_yellow = int((sub_full['verdict'] == 'YELLOW').sum())
        n_country_unr = int((sub_full['verdict'] == 'UNRESEARCHED').sum())
        mean_asym = sub_full['asymmetry_score'].mean() if 'asymmetry_score' in sub_full.columns else None
        mean_eta = sub_full[sort_col].mean() if sort_col in sub_full.columns else None
        top_eta = cr['top_eta']
        top_sym = cr['top_sym']

        country_tiles = [
            ("UNIVERSE",     f"{n_country_total:,}", "names ranked"),
            ("NMS",          f"{n_country_nms:,}",   "sub-$2B bucket"),
            ("GREEN",        f"{n_country_green}",   "high conviction"),
            ("YELLOW",       f"{n_country_yellow}",  "risk-flagged"),
            ("UNRESEARCHED", f"{n_country_unr}",     "no thesis"),
            ("TOP SCORER",   top_sym,                "by ETA"),
        ]
        for i, (lbl, val, sub_lbl) in enumerate(country_tiles):
            col = 1 + i * 2  # tiles span 2 cols each
            cl = sheet.cell(row=6, column=col, value=lbl)
            cl.font = f_text_italic_muted
            cl.alignment = A_LEFT
            cv = sheet.cell(row=7, column=col, value=val)
            cv.font = f_bold
            cv.alignment = A_LEFT
            cs = sheet.cell(row=8, column=col, value=sub_lbl)
            cs.font = f_text_italic_muted
            cs.alignment = A_LEFT
        sheet.row_dimensions[7].height = 22

        # Second row of tiles (right side: scores)
        score_tiles = [
            ("MEAN ASYMMETRY", mean_asym, "across universe"),
            ("MEAN ETA", mean_eta, "verdict-weighted"),
            ("TOP ETA", top_eta, f"({top_sym})"),
        ]
        for i, (lbl, val, sub_lbl) in enumerate(score_tiles):
            col = 1 + i * 2
            sheet.cell(row=10, column=col, value=lbl).font = f_text_italic_muted
            sheet.cell(row=10, column=col).alignment = A_LEFT
            _put_score(sheet, 11, col, val, font=f_bold)
            sheet.cell(row=11, column=col).alignment = A_LEFT
            sheet.cell(row=12, column=col, value=sub_lbl).font = f_text_italic_muted
            sheet.cell(row=12, column=col).alignment = A_LEFT
        sheet.row_dimensions[11].height = 22

        # Thin rule under headlines
        for c in range(1, 20):
            sheet.cell(row=13, column=c).border = Border(top=Side(style='thin', color=INK))
        sheet.row_dimensions[13].height = 4

        # Top-N table for this country
        _section_label(sheet, 15, f"Top {len(sub_top)}  —  ranked by {sort_label}", span_cols=19)
        _write_table_header(sheet, 16)
        sheet.row_dimensions[16].height = 18
        # Re-rank within the local frame (in case some rows were dropped)
        sub_top = sub_top.sort_values('country_rank')
        for r_idx, (_, r) in enumerate(sub_top.iterrows(), start=18):
            _write_table_row(sheet, r_idx, r, 11)
            sheet.row_dimensions[r_idx].height = 16

        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = 'A18'
        # QoL: sortable/filterable table (header row 17 → last data row)
        from openpyxl.utils import get_column_letter as _gcl
        if sheet.max_row >= 18:
            sheet.auto_filter.ref = f"A17:{_gcl(sheet.max_column)}{sheet.max_row}"

    wb.save(path)
    from harvard_style import sanitize_nan_text
    sanitize_nan_text(path)
    print(f'  wrote {path}  ({len(wb.worksheets)} sheets: Cover + {len(country_rows)} countries)',
          file=sys.stderr)


def _country_sheet_name(src_code: str) -> str:
    """Safe Excel sheet name (max 31 chars, no slashes etc.)."""
    s = ''.join(ch if ch.isalnum() else '_' for ch in str(src_code))
    return s[:31]


if __name__ == '__main__':
    main()
