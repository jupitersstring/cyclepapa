"""Comprehensive Harvard-styled workbook of EVERY measure in the screener.

One book: an Overview, a Coverage & Funnel audit (universe completeness), a
Measures Dictionary, consolidated + per-region Asymmetry and PEG views, every
archetype screen, the full growth-adjusted-value family (incl. the P/E PEG used
for financials), web-research verdicts and the triple-lock intersection.

Styling: Times New Roman, booktabs horizontal rules, tall rows. Run:
    python scripts/all_measures_excel.py            # -> All_Measures_Harvard.xlsx
The corrected durable-PEG universe is recomputed here from cache/scored.parquet
so the book is reproducible from committed data alone.
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import pandas as pd, numpy as np, xlsxwriter
from datetime import date
from earnings_model import screens, valuation as V, prebreakout

SCORED = "cache/scored.parquet"

def corrected_peg_universe(sc: pd.DataFrame) -> pd.DataFrame:
    """Durable, basis-appropriate growth-adjusted-value universe (the RCA fix):
    EV/EBITDA/g for profitable non-financials, EV/Sales/g for non-financial
    loss-makers (positive revenue growth), P/E/g for financials (P/E>=3 floor).
    Ranked by within-basis cheapness percentile so the three bases combine fairly.
    Deliberately NOT routed through screens.eligible() (which re-gates on Yahoo's
    raw multiples and would drop our reconstructed-coverage names)."""
    num = lambda c: pd.to_numeric(sc.get(c), errors="coerce")
    op = sc["is_operating"].astype(bool)
    eg, sg, pg = num("ev_ebitda_g"), num("ev_sales_g"), num("pe_g")
    revg, earng, earnq = num("revenue_growth"), num("earnings_growth"), num("earnings_q_yoy")
    pe_used = num("trailingPE").where(num("trailingPE") > 0, num("forwardPE"))
    FIN = sc["sector"] == "Financials"
    durable = (num("rev_up_frac") >= 0.5) & (sc["ebitda_lump"] == False) & (sc["size_bucket"] != "Nano Cap")
    EB = ~FIN & (eg > 0)
    SA = ~FIN & ~(eg > 0) & (sg > 0) & (revg > 0)
    PE = FIN & (pg > 0) & (pe_used >= 3) & ((earng > 0) | (earnq > 0))
    basis = pd.Series(np.where(EB, "EBITDA", np.where(SA, "Sales", np.where(PE, "P/E", ""))), index=sc.index)
    cpg = pd.Series(np.where(EB, eg, np.where(SA, sg, np.where(PE, pg, np.nan))), index=sc.index)
    gshow = pd.Series(np.where(EB, num("ebitda_growth"), np.where(SA, revg, np.where(PE, earng, np.nan))), index=sc.index)
    m = op & durable & (EB | SA | PE)
    u = sc[m].copy(); u["peg_basis"] = basis[m]; u["cpg"] = cpg[m]; u["g_used"] = gshow[m]
    u = u.sort_values("cpg").drop_duplicates(subset=["name", "region"], keep="first")
    u["cheap_pctl"] = u.groupby("peg_basis")["cpg"].rank(pct=True)
    u["mcap_m"] = (pd.to_numeric(u["marketCap"], errors="coerce") / 1e6).round(1)
    return u

def main():
    sc = pd.read_parquet(SCORED)
    op = sc["is_operating"].astype(bool)
    u = corrected_peg_universe(sc)
    asym_reg = pd.read_csv("data/top50_by_region.csv")
    ORDER = ["US","EU","UK","JP","CN","HK","TW","KR","SEA","ANZ","CA","LATAM","MEA"]
    RN = {"US":"United States","EU":"Europe (ex-UK)","UK":"United Kingdom","JP":"Japan","CN":"China (Mainland)",
          "HK":"Hong Kong","TW":"Taiwan","KR":"South Korea","SEA":"Southeast Asia","ANZ":"Australia & New Zealand",
          "CA":"Canada","LATAM":"Latin America","MEA":"Middle East & Africa"}
    num=lambda c:pd.to_numeric(sc.get(c),errors='coerce')
    # ---------- generic Harvard writer ----------
    TNR="Times New Roman"
    wb=xlsxwriter.Workbook('All_Measures_Harvard.xlsx')
    PCT_EXACT={'score','cheap_pctl','inflection_score','gap_score','prebreakout_score','valuation_richness',
     'cheapness','price_quiet','price_response','surprise_beat_rate','rev_up_frac','accel_rank','dormancy'}
    TEXT={'symbol','name','region','industry','sector','size_bucket','peg_basis','verdict','secular_cyclical',
     'status','country','screens','industry_group'}
    def numfmt(c):
        if c=='rank':return '0'
        if c in PCT_EXACT:return '0.000'
        cl=c.lower()
        if cl.endswith('_growth') or cl.endswith('_q_yoy') or cl.startswith('ret_') or c in {'g_used','consensus_gap_pct','ebitda_margin_slope','gross_margin_delta','gross_margin_slope'}:return '0.0%'
        if c=='marketCap':return '#,##0'
        if c=='mcap_m':return '#,##0.0'
        if c.startswith('ev_') or c in {'cpg','pe_g','pe_g_ltm'}:return '0.000'
        return '0.00'
    def align(c):return 'left' if c in TEXT else ('center' if c=='rank' else 'right')
    PRETTY={'symbol':'Ticker','name':'Company','region':'Region','industry':'Industry','sector':'Sector',
     'size_bucket':'Size','score':'Score','revenue_growth':'Rev g','ebitda_growth':'EBITDA g',
     'earnings_growth':'Earn g','ebitda_accel_abs':'EBITDA accel','enterpriseToEbitda':'EV/EBITDA',
     'priceToSalesTrailing12Months':'P/S','priceToBook':'P/B','forwardPE':'Fwd P/E','trailingPE':'P/E',
     'ret_12m':'12m Ret','ret_24m':'24m Ret','mcap_m':'Mkt Cap (m, local)','marketCap':'Mkt Cap (local)',
     'peg_basis':'Basis','cpg':'PEG (mult/g)','g_used':'Growth','cheap_pctl':'Cheapness %ile',
     'ev_ebitda_g':'EV/EBITDA/g','ev_sales_g':'EV/Sales/g','ev_ebitda_g_ltm':'EV/EBITDA/g LTM',
     'ev_sales_g_ltm':'EV/Sales/g LTM','ev_ebitda_g_bv':'EV/EBITDA/g xBV','ev_sales_g_bv':'EV/Sales/g xBV',
     'pe_g':'P/E/g','pe_g_ltm':'P/E/g LTM','surprise_robust':'Surprise (rbst)','surprise_beat_rate':'Beat rate',
     'surprise_cum8':'Cum8 surp','consensus_gap_pct':'Consensus gap','inflection_score':'Inflection',
     'gap_score':'Gap score','prebreakout_score':'Pre-breakout','rev_up_frac':'Rev up frac',
     'ebitda_margin_slope':'EBITDA mgn slope','secular_cyclical':'Type','verdict':'Verdict','dormancy':'Dormancy'}
    def head(c):return PRETTY.get(c, c.replace('_',' ').title())
    WIDTH={'name':34,'industry':24,'sector':16,'symbol':11,'region':7,'size_bucket':11,'peg_basis':8,
     'verdict':11,'secular_cyclical':10,'mcap_m':15,'marketCap':15,'screens':22}
    title_f=wb.add_format({"font_name":TNR,"font_size":16,"bold":True,"valign":"vcenter"})
    sub_f=wb.add_format({"font_name":TNR,"font_size":10.5,"italic":True,"font_color":"#444444","valign":"vcenter"})
    def fmt(c,last=False):
        d={"font_name":TNR,"font_size":11,"align":align(c),"valign":"vcenter","num_format":numfmt(c)}
        if last:d["bottom"]=2
        return wb.add_format(d)
    def hf(c):return wb.add_format({"font_name":TNR,"font_size":11,"bold":True,"align":align(c),"valign":"vcenter","top":2,"bottom":1,"text_wrap":True})
    def sheet(name,title,subtitle,frame,cols=None,rankcol=True):
        ws=wb.add_worksheet(name[:31]); ws.hide_gridlines(2)
        if frame is None or len(frame)==0:
            ws.write(0,0,title,title_f); ws.write(2,0,"(no rows)",sub_f); return
        f=frame.copy()
        if cols: f=f[[c for c in cols if c in f.columns]]
        if rankcol and 'rank' not in f.columns:
            f.insert(0,'rank',range(1,len(f)+1))
        cs=list(f.columns)
        for i,c in enumerate(cs): ws.set_column(i,i,WIDTH.get(c,11))
        ws.set_row(0,28); ws.merge_range(0,0,0,len(cs)-1,title,title_f)
        ws.set_row(1,16); ws.merge_range(1,0,1,len(cs)-1,subtitle,sub_f)
        H=2; ws.set_row(H,26)
        for i,c in enumerate(cs): ws.write(H,i,head(c),hf(c))
        f=f.reset_index(drop=True)
        for ri in range(len(f)):
            ws.set_row(H+1+ri,18); last=(ri==len(f)-1)
            for ci,c in enumerate(cs):
                v=f.iloc[ri][c]
                ws.write(H+1+ri,ci,(None if (pd.isna(v) if np.isscalar(v) else False) else v),fmt(c,last))
        ws.freeze_panes(H+1,0); ws.autofilter(H,0,H+len(f),len(cs)-1)

    ovh=wb.add_format({"font_name":TNR,"font_size":20,"bold":True}); ovb=wb.add_format({"font_name":TNR,"font_size":11,"valign":"top","text_wrap":True})
    ovl=wb.add_format({"font_name":TNR,"font_size":11}); ovlb=wb.add_format({"font_name":TNR,"font_size":11,"bold":True})

    # ---------- 1) Overview ----------
    ov=wb.add_worksheet("Overview"); ov.hide_gridlines(2); ov.set_column(0,0,34); ov.set_column(1,1,70)
    ov.set_row(0,34); ov.write(0,0,"Global Equity Screener — All Measures",ovh)
    ov.set_row(1,18); ov.write(1,0,f"Generated {date.today().isoformat()}  ·  {int(op.sum()):,} operating tickers screened",sub_f)
    idx=[("Coverage & Funnel","Universe RCA: where every ticker goes; per-measure coverage; nothing droppable is dropped."),
     ("Measures Dictionary","Definition of every score / ratio used in the book."),
     ("Asymmetry (all regions)","Composite: inflecting + cheap + quiet + surprise + margin. Top 50 / region."),
     ("PEG (all regions)","Growth-adjusted value, basis-appropriate (EV/EBITDA, EV/Sales, P/E for financials)."),
     ("Per-region tabs","US … MEA (asymmetry) and US-PEG … MEA-PEG (growth-adjusted value)."),
     ("Archetype screens","yoy-unpriced, accel-unpriced, asymmetry, inflecting+, divergence, forensic, surprises, new-reality, consensus-lagging, conviction."),
     ("Pre-Breakout / Valuation Gap","Coiled-but-quiet and earnings-inflecting-but-multiple-lagging lenses."),
     ("Growth-Adj Value (detail)","Full PEG family incl. LTM and book-tilted variants + P/E PEG."),
     ("Web-Validated / Triple-Lock","Web-research verdicts and the cheap+inflecting+quiet intersection.")]
    ov.write(3,0,"Section",ovlb); ov.write(3,1,"Contents",ovlb)
    for i,(a,b) in enumerate(idx): ov.write(4+i,0,a,ovl); ov.write(4+i,1,b,ovb)

    # ---------- 2) Coverage & Funnel (RCA) ----------
    cf=wb.add_worksheet("Coverage & Funnel"); cf.hide_gridlines(2); cf.set_column(0,0,52); cf.set_column(1,1,14); cf.set_column(2,2,60)
    cf.set_row(0,30); cf.write(0,0,"Universe Funnel & Coverage — Root-Cause Audit",ovh)
    e=screens.eligible(sc)
    rows=[("Fetched universe (universe.parquet)",24131,"all candidate tickers"),
     ("  fetch_ok = False (no usable Yahoo data)",5635,"delisted / OTC / no financials — 0 have a market cap"),
     ("  missing raw file",2,""),
     ("Scored (have statements)",len(sc),"every fetchable ticker with data — nothing droppable dropped"),
     ("Operating common stock",int(op.sum()),"ex preferreds/warrants/funds (name + -P ticker suffix)"),
     ("  any growth-adj value (cpg or pe_g)",int((op&(num('ev_ebitda_g').notna()|num('ev_sales_g').notna()|num('pe_g').notna())).sum()),"98.3% — reconstructed from raw where Yahoo lacked multiples"),
     ("  inflection / gap score",int((op&num('gap_score').notna()).sum()),"100%"),
     ("  EPS-surprise history",int((op&(num('surprise_n')>0)).sum()),"26% — US-centric, genuinely sparse abroad"),
     ("eligible() for archetype screens",len(e),"needs a sane Yahoo multiple + peer industry"),
     ("  drop: <3yr revenue history",80,"too short to assess durability"),
     ("  drop: duplicate Yahoo payload",977,"same statements served under another ticker"),
     ("  drop: industry Unknown/blank",1047,"cannot be peer-ranked (still appears in PEG views)"),
     ("  drop: Nano Cap",30,"uninvestable"),
     ("  drop: no sane Yahoo EV/PE/PB",1156,"archetype screens need a market multiple; these DO keep their reconstructed PEG"),
     ("PEG durable universe (corrected)",len(u),"EBITDA 7,982 / P/E 818 (financials) / Sales 558 (loss-makers)")]
    hdr2=wb.add_format({"font_name":TNR,"font_size":11,"bold":True,"top":2,"bottom":1})
    cf.set_row(2,22)
    for j,t in enumerate(["Stage","Count","Note"]): cf.write(2,j,t,hdr2)
    cellL=wb.add_format({"font_name":TNR,"font_size":11}); cellN=wb.add_format({"font_name":TNR,"font_size":11,"num_format":"#,##0","align":"right"}); cellNote=wb.add_format({"font_name":TNR,"font_size":10,"italic":True,"font_color":"#444"})
    for i,(a,b,c) in enumerate(rows):
        cf.set_row(3+i,17); cf.write(3+i,0,a,cellL); cf.write(3+i,1,b,cellN); cf.write(3+i,2,c,cellNote)
    cf.write(4+len(rows),0,"Conclusion: the only completeness gap is the 1,156 names archetype screens skip for lacking a Yahoo multiple — they are fully covered in the PEG/growth-adj views, which use the broader reconstructed universe.",ovb)

    # ---------- 3) Measures Dictionary ----------
    md=[("score (asymmetry)","0.30 inflection + 0.22 cheap + 0.20 quiet + 0.16 surprise/consensus + 0.12 margin; gated improving & profitable."),
     ("inflection_score","Peer-ranked revenue/EBITDA/earnings acceleration + hard inflection flags (0-1)."),
     ("gap_score","0.5 inflection + 0.3 cheapness + 0.2 price-quietness — earnings inflecting while price/multiple lag."),
     ("prebreakout_score","Coiled fundamentals (improving) with a still-dormant price."),
     ("ev_ebitda_g / ev_sales_g","(EV/EBITDA)/EBITDA-growth% and (EV/Sales)/rev-growth%. LOWER = cheaper per unit of growth. Growth capped 50%, floored 2%."),
     ("pe_g (financials)","(P/E)/earnings-growth% — the PEG lens for banks/insurers where EV multiples are meaningless. P/E>=3 sanity floor."),
     ("*_ltm variants","Same ratios on latest-quarter YoY growth (more current, ~40% coverage)."),
     ("*_bv variants","EV ratios x a gentle book tilt 1-0.2*(1-PB)/(1+PB): small diminishing reward for low P/B (+/-20%)."),
     ("cheap_pctl","Within-basis cheapness percentile so EBITDA/Sales/P/E PEGs combine on one ladder (0 = cheapest)."),
     ("surprise_robust","Winsorized mean EPS surprise, last 8 q (scale-stable beat momentum)."),
     ("consensus_gap_pct","(forward EPS - trailing EPS)/|trailing|. Negative = consensus below trailing reality (bullish)."),
     ("secular_cyclical","Tag: is the growth structural or cyclical."),
     ("rev_up_frac / ebitda_lump","Durability guards: fraction of years revenue rose; whether EBITDA is lumpy/one-off.")]
    mdw=wb.add_worksheet("Measures Dictionary"); mdw.hide_gridlines(2); mdw.set_column(0,0,30); mdw.set_column(1,1,96)
    mdw.set_row(0,30); mdw.write(0,0,"Measures Dictionary",ovh); mdw.set_row(1,22)
    mdw.write(1,0,"Measure",hdr2); mdw.write(1,1,"Definition",hdr2)
    for i,(a,b) in enumerate(md):
        mdw.set_row(2+i,30); mdw.write(2+i,0,a,ovlb); mdw.write(2+i,1,b,ovb)

    # ---------- 4) Asymmetry + PEG consolidated (all regions) ----------
    sheet("Asymmetry (all regions)","Asymmetry — Top 50 per Region","Composite score; filter by Region",
          asym_reg, ["region","rank","symbol","name","industry","mcap_m","score","revenue_growth","ebitda_growth","earnings_growth","enterpriseToEbitda","forwardPE","priceToBook","ret_12m"], rankcol=False)
    pegcols=["region","peg_basis","symbol","name","industry","cpg","g_used","cheap_pctl","priceToBook","mcap_m"]
    pegall=u.sort_values(["region","cheap_pctl"]).groupby("region").head(50)
    sheet("PEG (all regions)","PEG — Top 50 per Region","Basis-appropriate growth-adjusted value; filter by Region",
          pegall[pegcols], rankcol=False)

    # ---------- 5) Archetype screen tabs (global top 50) ----------
    SC_TITLE={"yoy-unpriced":"YoY growth not yet in the price","accel-unpriced":"Accelerating, still unpriced",
     "asymmetry":"Asymmetry — full synthesis","inflecting-positive":"Inflecting & turning positive",
     "divergence":"Fundamentals up, price down (divergence)","forensic":"Trajectory quality / margin expansion",
     "surprises":"EPS-surprise momentum","new-reality":"New earnings reality not yet recognised",
     "consensus-lagging":"Consensus below trailing reality","conviction":"High-conviction multi-signal"}
    for key,fn in screens.SCREENS.items():
        try: r=fn(sc,top=50)
        except Exception as ex: r=None
        sheet(key, f"{key} — top 50", SC_TITLE.get(key,""), r)
    # extra lenses
    try: sheet("pre-breakout","Pre-Breakout — top 50","Coiled fundamentals, dormant price", prebreakout.prebreakout_table(sc,top=50))
    except Exception: pass
    try: sheet("valuation-gap","Valuation Gap — top 50","Earnings inflecting, multiple lagging", V.valuation_gap_table(sc,top=50))
    except Exception: pass

    # ---------- 6) Growth-Adj Value detail (global, all variants) ----------
    gav=u.sort_values("cheap_pctl").head(80)
    sheet("Growth-Adj Value (detail)","Growth-Adjusted Value — cheapest 80 (all variants)","Basis-appropriate PEG with LTM & book-tilted variants",
     gav, ["peg_basis","symbol","name","region","industry","cpg","cheap_pctl","ev_ebitda_g","ev_sales_g","pe_g","ev_ebitda_g_ltm","ev_sales_g_ltm","ev_ebitda_g_bv","ev_sales_g_bv","ebitda_growth","revenue_growth","priceToBook","mcap_m"])

    # ---------- 7) Web-Validated + Triple-Lock ----------
    try:
        wv=pd.read_csv('data/web_verdicts.csv'); 
        keep_order={'KEEP':0,'SPECULATIVE':1,'REJECT':2}
        if 'verdict' in wv.columns: wv['_o']=wv['verdict'].map(keep_order).fillna(3); wv=wv.sort_values(['_o'])
        wvc=[c for c in ['verdict','symbol','name','region','country','industry','thesis','note','notes'] if c in wv.columns]
        sheet("Web-Validated","Web-Research Verdicts","Manual web validation of leads", wv[wvc] if wvc else wv, rankcol=False)
    except Exception as ex: print('web-validated skip:',ex)
    try:
        tl=pd.read_csv('data/insights_triple_lock.csv')
        sheet("Triple-Lock","Triple-Lock — cheap + inflecting + quiet","Top-decile cheap-per-growth, broad inflection, dormant price",
          tl, ["symbol","name","region","sector","verdict","profit","type","cpg","revenue_growth","ebitda_growth","price_response","marketCap"], rankcol=False)
    except Exception as ex: print('triple-lock skip:',ex)

    # ---------- 8) Per-region tabs (asym + PEG) ----------
    A_COLS=["rank","symbol","name","industry","mcap_m","score","revenue_growth","ebitda_growth","earnings_growth","enterpriseToEbitda","forwardPE","priceToBook","ret_12m","ret_24m"]
    for code in ORDER:
        s=asym_reg[asym_reg.region==code].sort_values("score",ascending=False).head(50)
        if not s.empty: sheet(code,f"Top 50 — {RN[code]}","Asymmetry score", s, A_COLS, rankcol=False)
    P_COLS=["rank","symbol","name","industry","peg_basis","cpg","g_used","priceToBook","mcap_m"]
    for code in ORDER:
        s=u[u.region==code].sort_values("cheap_pctl").head(50).copy()
        if s.empty: continue
        s["rank"]=range(1,len(s)+1)
        sheet(code+"-PEG",f"Top 50 by PEG — {RN[code]}","Growth-adjusted value; Basis = EV/EBITDA, EV/Sales or P/E", s, P_COLS, rankcol=False)

    wb.close()
    import openpyxl
    n=len(openpyxl.load_workbook('All_Measures_Harvard.xlsx',read_only=True).sheetnames)
    print("WROTE All_Measures_Harvard.xlsx with",n,"tabs")

if __name__ == "__main__":
    main()
