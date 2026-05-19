"""Parse the three research markdown files and build a single Excel workbook
with one row per (fund, category, position) plus per-fund summary sheets."""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_FILES = {
    'US Activists – Tier 1': '/tmp/research_us_activists.md',
    'US Activists – Tier 2': '/tmp/research_us_activists_more.md',
    'US Activists – Tier 3 (small-cap)': '/tmp/research_activists_t3.md',
    'Small-cap / Multibagger / Special Sits – Tier 1': '/tmp/research_smallcap_specialsits.md',
    'Small-cap / Multibagger / Special Sits – Tier 2': '/tmp/research_smallcap_tier2.md',
    'Small-cap / Deep Value – Tier 3': '/tmp/research_smallcap_value_t3.md',
    'International Activist + Distressed': '/tmp/research_intl_distressed.md',
    'International Activist – Tier 2': '/tmp/research_intl_more.md',
    'Distressed / Event-Driven – Tier 2': '/tmp/research_distressed_eventdriven.md',
    'Japan + Distressed – Tier 3': '/tmp/research_japan_distressed_t3.md',
    'Tiger Cubs / L/S Legends': '/tmp/research_tiger_cubs.md',
    'Legendary Family Offices': '/tmp/research_legendary_fos.md',
    'Value & Multi-Strat Legends': '/tmp/research_value_multistrat.md',
    'European + Asia – Tier 2': '/tmp/research_europe_asia_t2.md',
    'Biotech Specialists': '/tmp/research_biotech_specialists.md',
    'Smaller Activists / Special Sits Tier 4': '/tmp/research_smaller_activists.md',
    'Global Quality + EM Specialists': '/tmp/research_quality_em.md',
    'Mega Multi-Strats / Quants': '/tmp/research_megamulti.md',
    'Microcap-Tactical + International T4': '/tmp/research_microcap_intl_t4.md',
    'Family-Office / Individual Filers': '/tmp/research_family_office_filers.md',
    'Small Concentrated Activists T4': '/tmp/research_small_concentrated_t4.md',
    'Skin-in-Game / Fat-Pitch Legends': '/tmp/research_skin_game_A.md',
    'Concentrated Value / Quality Compounders': '/tmp/research_value_B.md',
    'Elite Shorts + Concentrated Value': '/tmp/research_shorts_value_C.md',
    'Warrant Specialists (SPAC + Mining + Tail)': '/tmp/research_warrants.md',
}

# Category headings to detect in each fund block
CAT_PATTERNS = [
    ('1', 'Highest conviction positions (recent adds)', re.compile(r'###\s*\(1\)')),
    ('2', '>=5% / threshold disclosures', re.compile(r'###\s*\(2\)')),
    ('3', 'New positions sized large', re.compile(r'###\s*\(3\)')),
    ('4', 'Existing positions materially increased', re.compile(r'###\s*\(4\)')),
    ('N', 'Notes', re.compile(r'###\s*Notes')),
]


def parse_markdown(path):
    """Return list of dicts: {fund, sources, sections: {cat_key: rows or text}}."""
    with open(path) as f:
        text = f.read()

    # Split by H2 (fund) blocks
    parts = re.split(r'(?m)^##\s+(?!#)', text)
    funds = []
    for p in parts:
        if not p.strip() or p.lstrip().startswith('Cross-fund') or p.lstrip().startswith('Summary') or \
           p.lstrip().startswith('Important caveats') or p.lstrip().startswith('Key activist filings'):
            continue
        # First line is fund name
        lines = p.strip().splitlines()
        fund_name = lines[0].strip()
        body = '\n'.join(lines[1:])

        # Sources block
        sm = re.search(r'\*\*Sources used:\*\*\s*(.+?)(?=\n\n|\n###|$)', body, re.S)
        sources = sm.group(1).strip() if sm else ''

        # Sections
        sections = {}
        # Locate H3 sections
        section_positions = []
        for key, label, pat in CAT_PATTERNS:
            m = pat.search(body)
            if m:
                section_positions.append((m.start(), key, label))
        section_positions.sort()

        for i, (pos, key, label) in enumerate(section_positions):
            end = section_positions[i+1][0] if i+1 < len(section_positions) else len(body)
            block = body[pos:end]
            # strip the heading line itself
            block = block.split('\n', 1)[1] if '\n' in block else ''
            sections[key] = (label, block.strip())

        funds.append({
            'name': fund_name,
            'sources': sources,
            'sections': sections,
        })
    return funds


def parse_table(block):
    """Parse a Markdown pipe-table; return (headers, rows) or (None, None) if not a table.
    Handles tables that may be interspersed with prose lines."""
    if not block:
        return None, None
    lines = [ln for ln in block.splitlines() if ln.strip()]
    # find header
    hdr_idx = None
    for i, ln in enumerate(lines):
        if ln.strip().startswith('|') and i+1 < len(lines) and re.match(r'^\|\s*[-:|\s]+$', lines[i+1].strip()):
            hdr_idx = i
            break
    if hdr_idx is None:
        return None, None
    headers = [c.strip() for c in lines[hdr_idx].strip().strip('|').split('|')]
    rows = []
    for ln in lines[hdr_idx+2:]:
        if not ln.strip().startswith('|'):
            break
        cells = [c.strip() for c in ln.strip().strip('|').split('|')]
        # Pad / trim to header length
        if len(cells) < len(headers):
            cells = cells + [''] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[:len(headers)]
        # Skip placeholder rows that contain only "..." or are empty
        if all(c.strip() in ('...', '') for c in cells):
            continue
        # Skip "Not found" rows that snuck into the table
        if cells[0].lower().startswith('_not found'):
            continue
        rows.append(cells)
    return headers, rows


def parse_prose(block):
    """Return the prose text (italics block / notes / 'Not found' messages) stripped of tables."""
    if not block:
        return ''
    # remove pipe-table lines
    out = []
    in_table = False
    for ln in block.splitlines():
        if ln.strip().startswith('|'):
            in_table = True
            continue
        if in_table and not ln.strip():
            in_table = False
        if not in_table:
            out.append(ln)
    return '\n'.join(out).strip()


# Collect data
group_funds = {g: parse_markdown(p) for g, p in MD_FILES.items()}

# ------- Build workbook -------
wb = Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E78')
group_fill = PatternFill('solid', fgColor='305496')
fund_fill = PatternFill('solid', fgColor='D9E1F2')
cat_fill = PatternFill('solid', fgColor='FFF2CC')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def border_all(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap


# --- Sheet 1: Cover / Methodology ---
ws = wb.active
ws.title = 'Cover'
ws.append(['Recent Investor Activity — Funds from your saved-search universe'])
ws.append(['Compiled', '2026-05-11'])
ws.append(['Window', 'Nov 2025 – early May 2026 (last ~6 months of filings)'])
ws.append(['Universe scope', '30 funds across multibagger / small-cap / activist / special-sits / deep value buckets'])
ws.append([])
ws.append(['Source file with original saved searches', 'pb_searches_by_theme.xlsx (this branch)'])
ws.append([])
ws.append(['Categories'])
ws.append(['(1) Highest conviction positions (recent adds)', 'Top portfolio weights where the fund was a net buyer in the most recent quarterly filings'])
ws.append(['(2) >=5% disclosures (last 6 months)', '13D / 13G / 13D-A / TR-1 / AMF / HKEX / Japan large-shareholding filings disclosing stakes above local thresholds'])
ws.append(['(3) New positions sized large', 'Positions newly initiated in Q4 2025 (or Q1 2026 where available) at meaningful weight'])
ws.append(['(4) Existing positions materially increased', 'Pre-existing positions where share count increased materially (>25%) or weight moved up'])
ws.append([])
ws.append(['Caveats'])
ws.append(['Q1 2026 13Fs (Mar-end) not yet due / filed at compile date; deadline 15 May 2026'])
ws.append(['13F captures US-listed equities & options only — understates non-US books of Cevian, Oasis, Petrus, etc.'])
ws.append(['Some funds (Indaba, Mudrick, Knighthead, King Street, Davidson Kempner) have far larger credit/private books not visible in 13F'])
ws.append(['Where a category is empty for a fund: "Not found in public filings within window" (left blank in row tables)'])

for cell in ws['A']:
    cell.alignment = wrap
ws['A1'].font = Font(bold=True, size=14)
for r in (8, 13):
    ws.cell(row=r, column=1).font = Font(bold=True)
autosize(ws, [55, 90])


# --- Build a master 'All Activity' sheet ---
master = wb.create_sheet('All Activity')
master.append(['Group', 'Fund', 'Category', 'Ticker / Company / Detail', 'Detail 1', 'Detail 2', 'Detail 3', 'Source / Filing Date / Notes'])
style_header(master, 8)

cat_map = {
    '1': '(1) Highest conviction (recent adds)',
    '2': '(2) >=5% disclosures (last 6m)',
    '3': '(3) New positions sized large',
    '4': '(4) Existing positions materially increased',
}


def squash_to_5(headers, rows):
    """Map arbitrary table to 5 columns: Ticker, Detail1, Detail2, Detail3, Source."""
    out = []
    # Identify columns: ticker/company is col 0, last col is source, middle cols compressed
    for r in rows:
        if len(r) >= 5:
            ticker = r[0]
            source = r[-1]
            details = r[1:-1]
        elif len(r) == 4:
            ticker = r[0]
            source = r[-1]
            details = r[1:-1]
        else:
            ticker = r[0] if r else ''
            source = r[-1] if len(r) > 1 else ''
            details = r[1:-1] if len(r) > 2 else []
        # Pad details to 3
        details = (details + ['', '', ''])[:3]
        out.append([ticker, *details, source])
    return out


for group, funds in group_funds.items():
    for f in funds:
        for cat_key in ('1', '2', '3', '4'):
            if cat_key not in f['sections']:
                continue
            label, block = f['sections'][cat_key]
            headers, rows = parse_table(block)
            prose = parse_prose(block)
            cat_label = cat_map[cat_key]
            if rows:
                compact = squash_to_5(headers, rows)
                for cr in compact:
                    master.append([group, f['name'], cat_label, *cr])
            elif prose:
                master.append([group, f['name'], cat_label, '', '', '', '', prose])
border_all(master)
master.freeze_panes = 'A2'
autosize(master, [22, 32, 32, 36, 30, 30, 22, 70])


# --- One sheet per fund ---
def safe_name(s):
    s = re.sub(r'[:\\/\?\*\[\]]', ' ', s)
    return s[:31]


for group, funds in group_funds.items():
    for f in funds:
        ws = wb.create_sheet(safe_name(f['name']))
        # Header row 1: fund name banner
        ws.append([f['name']])
        ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
        ws['A1'].fill = group_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        # Group row
        ws.append([f'Group: {group}'])
        ws['A2'].font = Font(italic=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        # Sources
        if f['sources']:
            ws.append(['Sources:'])
            ws[f'A{ws.max_row}'].font = Font(bold=True)
            for line in [s.strip() for s in f['sources'].replace('; ', '\n').splitlines() if s.strip()]:
                ws.append([line])
        ws.append([])

        for cat_key in ('1', '2', '3', '4'):
            if cat_key not in f['sections']:
                continue
            label, block = f['sections'][cat_key]
            ws.append([f'({cat_key}) {label}'])
            cat_row = ws.max_row
            cell = ws.cell(row=cat_row, column=1)
            cell.font = Font(bold=True, size=11)
            cell.fill = cat_fill
            ws.merge_cells(start_row=cat_row, start_column=1, end_row=cat_row, end_column=6)

            headers, rows = parse_table(block)
            prose = parse_prose(block)
            if rows:
                ws.append(headers)
                hdr_row = ws.max_row
                for c in ws[hdr_row]:
                    c.font = Font(bold=True, color='FFFFFF')
                    c.fill = header_fill
                    c.alignment = Alignment(wrap_text=True, vertical='center')
                for r in rows:
                    ws.append(r)
            if prose:
                ws.append([prose])
                pr = ws.max_row
                cell = ws.cell(row=pr, column=1)
                cell.font = Font(italic=True, color='595959')
                ws.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=6)
            ws.append([])

        # Notes section
        if 'N' in f['sections']:
            label, block = f['sections']['N']
            ws.append(['Notes & active campaigns'])
            nrow = ws.max_row
            ws.cell(row=nrow, column=1).font = Font(bold=True, size=11)
            ws.cell(row=nrow, column=1).fill = cat_fill
            ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=6)
            for line in block.splitlines():
                if line.strip():
                    ws.append([line.strip()])

        # styling
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if cell.alignment is None or not cell.alignment.wrap_text:
                    cell.alignment = wrap
        autosize(ws, [16, 32, 14, 14, 14, 70])
        ws.freeze_panes = 'A3'


# --- Index sheet ---
index = wb.create_sheet('Index', 1)
index.append(['Group', 'Fund', 'Per-fund sheet'])
style_header(index, 3)
for group, funds in group_funds.items():
    for f in funds:
        sheet_name = safe_name(f['name'])
        index.append([group, f['name'], sheet_name])
        # Create hyperlink to the sheet
        link_cell = index.cell(row=index.max_row, column=3)
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = Font(color='0000FF', underline='single')
border_all(index)
index.freeze_panes = 'A2'
autosize(index, [38, 38, 36])

out = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'
wb.save(out)
print(f"Wrote {out}")
print(f"Sheets: {len(wb.sheetnames)}")
print(f"Funds covered: {sum(len(v) for v in group_funds.values())}")
print(f"All Activity rows: {master.max_row - 1}")
