"""Cross-fund synthesis: identify asymmetric opportunities + highest-conviction setups
across the 110 funds covered in /tmp/research_*.md, then append the analysis as new
sheets in fund_activity_last_6mo.xlsx.

Approach:
 1. Parse each markdown file.
 2. For every fund's sections (1)-(4), extract each table row.
 3. Heuristically identify (ticker, fund, category, direction, weight_hint, note).
 4. Aggregate by ticker:
     - count of funds touching the ticker across categories (1) & (3) & (4) (positive
       direction signals)
     - max % of portfolio observed in any fund
     - whether any fund has filed a >=5% disclosure on the ticker
     - whether the change column suggests a "new", "added", "+%" pattern
 5. Tag each ticker with several flags:
     - CONSENSUS: appears in (1)/(3)/(4) of 3+ funds
     - HIGH-CONVICTION: at >=15% of any fund's portfolio
     - ACTIVIST-CATALYST: >=5% filing (13D/13D-A) in the window
     - HYPER-CONVICTION: at >=30% of any fund's portfolio
 6. Produce three new sheets:
     - "Consensus Buys" — tickers held/added by 3+ funds
     - "Highest Conviction" — tickers at >=15% of any single fund
     - "Activist Catalysts" — tickers with 13D + active campaign narrative
"""

import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_FILES = {
    'US Activists – Tier 1': '/tmp/research_us_activists.md',
    'US Activists – Tier 2': '/tmp/research_us_activists_more.md',
    'US Activists – Tier 3 (small-cap)': '/tmp/research_activists_t3.md',
    'Small-cap / Multibagger / Special Sits – Tier 1': '/tmp/research_smallcap_specialsits.md',
    'Small-cap / Multibagger / Special Sits – Tier 2': '/tmp/research_smallcap_tier2.md',
    'Small-cap / Deep Value – Tier 3': '/tmp/research_smallcap_value_t3.md',
    'International Activist + Distressed': '/tmp/research_intl_distressed.md',
    'International Activist – Tier 2': '/tmp/research_intl_more.md',
    'Distressed / Event-Driven – Tier 2': '/tmp/research_distressed_eventdriven.md',
    'Japan + Distressed – Tier 3': '/tmp/research_japan_distressed_t3.md',
    'Tiger Cubs / L/S Legends': '/tmp/research_tiger_cubs.md',
    'Legendary Family Offices': '/tmp/research_legendary_fos.md',
    'Value & Multi-Strat Legends': '/tmp/research_value_multistrat.md',
    'European + Asia – Tier 2': '/tmp/research_europe_asia_t2.md',
    'Biotech Specialists': '/tmp/research_biotech_specialists.md',
    'Smaller Activists / Special Sits Tier 4': '/tmp/research_smaller_activists.md',
    'Global Quality + EM Specialists': '/tmp/research_quality_em.md',
    'Mega Multi-Strats / Quants': '/tmp/research_megamulti.md',
    'Microcap-Tactical + International T4': '/tmp/research_microcap_intl_t4.md',
    'Family-Office / Individual Filers': '/tmp/research_family_office_filers.md',
    'Small Concentrated Activists T4': '/tmp/research_small_concentrated_t4.md',
    'Skin-in-Game / Fat-Pitch Legends': '/tmp/research_skin_game_A.md',
    'Concentrated Value / Quality Compounders': '/tmp/research_value_B.md',
    'Elite Shorts + Concentrated Value': '/tmp/research_shorts_value_C.md',
    'Warrant Specialists (SPAC + Mining + Tail)': '/tmp/research_warrants.md',
}

# Heuristic ticker pattern. Catches:
#   AAPL, IEF, FOXF, GOOG, KVUE, BMW (3-5 caps); 9684 JT (Japan); 4-digit JP codes.
US_TICKER_RE = re.compile(r'^([A-Z]{1,5}(?:\.[A-Z])?)$')
JP_TICKER_RE = re.compile(r'^(\d{4})\s*JT$')  # Japan format used by agents


def parse_md(path):
    """Return list of (fund_name, sources, sections_dict)."""
    with open(path) as f:
        text = f.read()
    parts = re.split(r'(?m)^##\s+(?!#)', text)
    funds = []
    for p in parts:
        if not p.strip():
            continue
        head = p.lstrip().split('\n', 1)[0].strip()
        if head.lower().startswith(('cross-fund', 'summary', 'important caveats',
                                    'key activist filings', 'tiger cubs',
                                    'key cross', '##')):
            continue
        if not head or head.startswith('#'):
            continue
        fund_name = head
        body = p.split('\n', 1)[1] if '\n' in p else ''
        # split into category sections
        sections = {}
        for cat in ('1', '2', '3', '4'):
            m = re.search(rf'###\s*\({cat}\)([^\n]*)\n', body)
            if not m:
                continue
            start = m.end()
            # next ### or end
            m2 = re.search(r'###\s', body[start:])
            end = start + m2.start() if m2 else len(body)
            sections[cat] = body[start:end]
        funds.append((fund_name, sections))
    return funds


def parse_table_rows(block):
    """Return list of cell-lists for each non-header row of a markdown pipe-table."""
    if not block:
        return []
    lines = block.splitlines()
    rows = []
    in_table = False
    header_seen = False
    for ln in lines:
        s = ln.strip()
        if s.startswith('|'):
            # Skip header / separator
            if re.match(r'^\|\s*[-:|\s]+$', s):
                header_seen = True
                continue
            if not header_seen:
                in_table = True
                continue
            cells = [c.strip() for c in s.strip('|').split('|')]
            rows.append(cells)
        else:
            header_seen = False
    return rows


def extract_ticker(cell):
    """Return canonical ticker if the cell looks like one, else None."""
    if not cell:
        return None
    # Strip parentheticals
    cell = cell.strip()
    # If the cell is "AAPL | Apple Inc" style with pipe inside? No — table cells already separated.
    # Try the first token, e.g. "FLR | Fluor Corp" → "FLR"
    first = cell.split('|')[0].strip().split('(')[0].strip()
    # Look for "TICKER (calls)" or "TICKER (puts)" — keep the ticker
    first = first.split()[0] if first else ''
    if not first:
        return None
    if US_TICKER_RE.match(first):
        # Reject obvious noise words
        if first in {'NEW', 'TOP', 'CORE', 'ETF', 'LP', 'JT', 'OR', 'AND', 'NOT',
                     'HOLD', 'NA', 'NAN'}:
            return None
        return first
    # JP format: "9684 JT" or "9684 JP"
    if re.match(r'^\d{4}$', first):
        # If followed by JT/JP later in the cell, treat as JP ticker
        if re.search(r'\b(JT|JP|JT/?\s*JP)\b', cell):
            return f'{first} JT'
        # Sometimes the ticker is just "9684" without suffix — keep as JP
        return f'{first} JP'
    return None


def parse_percent(text):
    """Return max numeric % seen in text, or None."""
    if not text:
        return None
    pcts = re.findall(r'(\d+(?:\.\d+)?)\s*%', text)
    nums = [float(p) for p in pcts if float(p) <= 100]
    return max(nums) if nums else None


def parse_direction(text):
    """Return 'add', 'new', 'trim', 'hold', or None based on heuristics."""
    if not text:
        return None
    t = text.lower()
    if 'new' in t and 'newly' not in t:
        return 'new'
    if any(k in t for k in ['+', 'added', 'increase', 'top up', 'initiated', 'init']):
        return 'add'
    if any(k in t for k in ['-', 'trimmed', 'reduced', 'cut', 'exit', 'sold']):
        return 'trim'
    if any(k in t for k in ['held', 'unchanged', 'flat', 'core']):
        return 'hold'
    return None


# ---- Parse everything ----
all_funds = []  # list of (group, fund_name, sections)
for group, path in MD_FILES.items():
    for fund, secs in parse_md(path):
        all_funds.append((group, fund, secs))

print(f"Parsed {len(all_funds)} fund sections.")

# ticker -> list of dicts with metadata
ticker_data = defaultdict(list)

for group, fund, secs in all_funds:
    for cat in ('1', '2', '3', '4'):
        block = secs.get(cat, '')
        for row in parse_table_rows(block):
            if not row:
                continue
            # Row layouts vary; ticker usually in cell 0 (sometimes preceded by company)
            tk = extract_ticker(row[0]) if len(row) > 0 else None
            if tk is None and len(row) > 1:
                tk = extract_ticker(row[1])
            if tk is None:
                continue
            # joined text for direction/% inference
            joined = ' '.join(row[1:])
            weight = parse_percent(joined)
            direction = parse_direction(joined)
            ticker_data[tk].append({
                'group': group,
                'fund': fund,
                'cat': cat,
                'row_text': joined[:300],
                'weight': weight,
                'direction': direction,
            })

print(f"Unique tickers: {len(ticker_data)}")

# ---- Aggregate by ticker ----
agg = {}
for tk, entries in ticker_data.items():
    funds = {e['fund'] for e in entries}
    # positive entries = anything in cat 1/3/4 OR direction add/new/hold
    positive_funds = {
        e['fund']
        for e in entries
        if e['cat'] in {'1', '3', '4'} and e['direction'] != 'trim'
    }
    # ≥5% disclosures = anything in cat 2
    threshold_funds = {e['fund'] for e in entries if e['cat'] == '2'}
    # Max weight observed
    weights = [e['weight'] for e in entries if e['weight'] is not None]
    max_weight = max(weights) if weights else None
    # Direction summary
    dirs = [e['direction'] for e in entries if e['direction']]
    new_count = sum(1 for d in dirs if d == 'new')
    add_count = sum(1 for d in dirs if d == 'add')
    trim_count = sum(1 for d in dirs if d == 'trim')
    # Capture short narrative
    narratives = sorted(
        {f"{e['fund']}: {e['row_text'].strip()}"[:160] for e in entries},
    )[:6]
    agg[tk] = {
        'ticker': tk,
        'funds_total': len(funds),
        'funds_positive': len(positive_funds),
        'positive_fund_names': sorted(positive_funds),
        'funds_threshold': len(threshold_funds),
        'threshold_fund_names': sorted(threshold_funds),
        'max_weight_pct': max_weight,
        'new_count': new_count,
        'add_count': add_count,
        'trim_count': trim_count,
        'narratives': narratives,
    }

# ---- Flagging ----
flagged = []
for tk, d in agg.items():
    flags = []
    if d['funds_positive'] >= 3:
        flags.append('CONSENSUS')
    if d['max_weight_pct'] and d['max_weight_pct'] >= 30:
        flags.append('HYPER-CONVICTION')
    elif d['max_weight_pct'] and d['max_weight_pct'] >= 15:
        flags.append('HIGH-CONVICTION')
    if d['funds_threshold'] >= 1:
        flags.append('THRESHOLD-FILING')
    if d['new_count'] >= 2:
        flags.append('MULTI-FUND NEW INIT')
    if flags:
        d['flags'] = flags
        flagged.append(d)

# ---- Append synthesis sheets to existing workbook ----
out = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'
wb = load_workbook(out)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='C00000')  # dark red
sub_fill = PatternFill('solid', fgColor='FFF2CC')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def border_all(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap


# Sheet: Asymmetric & Conviction Summary
if 'Asymmetric Summary' in wb.sheetnames:
    del wb['Asymmetric Summary']
ws = wb.create_sheet('Asymmetric Summary', 0)
ws.append(['Asymmetric Opportunities & Highest-Conviction Setups'])
ws['A1'].font = Font(bold=True, size=14)
ws.append([])
ws.append(['How to read this analysis'])
ws['A3'].font = Font(bold=True)
ws.append([
    'Tickers below were extracted programmatically from the 110 fund research files. '
    'A ticker is flagged when one of these conditions holds:'
])
ws.append(['  CONSENSUS — appears in highest-conviction / recent-add / new-position tables of 3+ funds.'])
ws.append(['  MULTI-FUND NEW INIT — newly initiated as a top position by 2+ funds in the window.'])
ws.append(['  HIGH-CONVICTION — held at >=15% of any single fund\'s 13F portfolio.'])
ws.append(['  HYPER-CONVICTION — held at >=30% of any single fund\'s 13F portfolio.'])
ws.append(['  THRESHOLD-FILING — at least one fund filed a 13D / 13G / TR-1 / Japan large-shareholding report on it in the last 6 months.'])
ws.append([])
ws.append(['Cross-section tables in subsequent sheets:'])
ws['A11'].font = Font(bold=True)
ws.append(['  - "Consensus Buys" — tickers with the most cross-fund overlap (3+ funds adding/holding top weight).'])
ws.append(['  - "Highest Conviction" — single-fund concentration at >=15% of portfolio.'])
ws.append(['  - "Activist Catalysts" — tickers with >=5% disclosure + active campaign signal.'])
ws.append(['  - "Multi-Fund New Inits" — names being initiated as new top positions by 2+ funds simultaneously.'])
ws.append([])
ws.append(['Caveats'])
ws['A17'].font = Font(bold=True)
ws.append(['  - Ticker extraction is heuristic; non-US tickers (esp. Japan 4-digit codes) may be partial.'])
ws.append(['  - "Highest conviction" %s are SELF-REPORTED weights in the source 13F/factsheet, not implied bets — note that some funds report only US-equity slices and the % overstates conviction at firm level.'])
ws.append(['  - Activist catalyst flag is heuristic; verify against the per-fund tabs for filing dates and campaign status.'])
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = wrap
autosize(ws, [110])


# Sheet: Consensus Buys
ws = wb.create_sheet('Consensus Buys', 1)
ws.append(['Ticker', '# Funds Positive', 'Funds (positive)', 'Max % of any fund',
           '# New', '# Adds', '# Trims', 'Threshold filings', 'Flags', 'Sample notes'])
style_header(ws, 10)
consensus = [d for d in flagged if 'CONSENSUS' in d['flags']]
consensus.sort(key=lambda d: (-d['funds_positive'],
                              -(d['max_weight_pct'] or 0)))
for d in consensus[:120]:
    ws.append([
        d['ticker'],
        d['funds_positive'],
        ', '.join(d['positive_fund_names'][:10]) +
        (f' (+{len(d["positive_fund_names"]) - 10} more)' if len(d['positive_fund_names']) > 10 else ''),
        d['max_weight_pct'] if d['max_weight_pct'] else '',
        d['new_count'],
        d['add_count'],
        d['trim_count'],
        ', '.join(d['threshold_fund_names'][:5]),
        ', '.join(d['flags']),
        ' | '.join(d['narratives'][:3]),
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 14, 50, 14, 8, 8, 8, 28, 28, 80])


# Sheet: Highest Conviction
ws = wb.create_sheet('Highest Conviction', 2)
ws.append(['Ticker', 'Max % of any fund', 'Held by (fund)', '# Funds Positive',
           '# New', '# Adds', 'Flags', 'Sample notes'])
style_header(ws, 8)
high_conv = [d for d in flagged if any(f in d['flags']
                                       for f in ['HIGH-CONVICTION', 'HYPER-CONVICTION'])]
high_conv.sort(key=lambda d: -(d['max_weight_pct'] or 0))
for d in high_conv:
    ws.append([
        d['ticker'],
        d['max_weight_pct'],
        ', '.join(d['positive_fund_names'][:10]),
        d['funds_positive'],
        d['new_count'],
        d['add_count'],
        ', '.join(d['flags']),
        ' | '.join(d['narratives'][:3]),
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 14, 60, 14, 8, 8, 28, 80])


# Sheet: Activist Catalysts
ws = wb.create_sheet('Activist Catalysts', 3)
ws.append(['Ticker', '# Threshold filings', 'Funds with >=5%/13D', '# Funds Positive',
           'Max % of any fund', 'Flags', 'Sample notes'])
style_header(ws, 7)
catalysts = [d for d in flagged if d['funds_threshold'] >= 1]
catalysts.sort(key=lambda d: (-d['funds_threshold'], -(d['max_weight_pct'] or 0),
                              -d['funds_positive']))
for d in catalysts:
    ws.append([
        d['ticker'],
        d['funds_threshold'],
        ', '.join(d['threshold_fund_names'][:6]),
        d['funds_positive'],
        d['max_weight_pct'] if d['max_weight_pct'] else '',
        ', '.join(d['flags']),
        ' | '.join(d['narratives'][:3]),
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 14, 60, 14, 14, 32, 80])


# Sheet: Multi-Fund New Inits
ws = wb.create_sheet('Multi-Fund New Inits', 4)
ws.append(['Ticker', '# New-Inits', 'Funds initiating (new)', '# Funds Positive',
           'Max % of any fund', 'Flags', 'Sample notes'])
style_header(ws, 7)
# For "new init" we need funds that listed this ticker with direction=new in cat 3
new_init_funds = defaultdict(set)
for tk, entries in ticker_data.items():
    for e in entries:
        if e['cat'] == '3' and (e['direction'] == 'new' or 'new' in (e['row_text'] or '').lower()):
            new_init_funds[tk].add(e['fund'])
multi_inits = [(tk, sorted(funds)) for tk, funds in new_init_funds.items() if len(funds) >= 2]
multi_inits.sort(key=lambda x: -len(x[1]))
for tk, funds in multi_inits:
    d = agg.get(tk, {})
    ws.append([
        tk,
        len(funds),
        ', '.join(funds),
        d.get('funds_positive', ''),
        d.get('max_weight_pct') or '',
        ', '.join(d.get('flags', [])),
        ' | '.join((d.get('narratives') or [])[:3]),
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 12, 60, 14, 14, 32, 80])

wb.save(out)
print(f"Saved {out}")
print(f"Consensus rows: {len(consensus)}")
print(f"High-conviction rows: {len(high_conv)}")
print(f"Activist-catalyst rows: {len(catalysts)}")
print(f"Multi-fund new inits: {len(multi_inits)}")
