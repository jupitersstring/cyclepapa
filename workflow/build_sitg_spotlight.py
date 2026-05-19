"""Build 'SITG Smallcap Spotlight' sheet: for each high-SITG smallcap fund, surface
the highest-conviction current position + recent activity from per-fund tabs."""
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB_PATH = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'

# Fund key (in workbook) -> SITG profile
FUND_PROFILES = [
    # (fund_key, blurb, sitg_score, typical_size)
    ('Abrams Capital', 'David Abrams; closed fund; "Buffett of hedge funds"', 5, 'Mid/Large'),
    ('Mithaq Capital', 'Turki AlRajhi own Saudi family money', 5, 'Small/Micro'),
    ('MILFAM', 'Neil Subin (Lloyd Miller III estate); permanent capital', 5, 'Micro'),
    ('Bradley L. Radoff', 'BLR Partners + Radoff Family Foundation own capital', 5, 'Micro'),
    ('Braden M. Leonard', 'BML Investment Partners; own money', 5, 'Micro/Smallbio'),
    ('Peter H. Kamin', '3K Limited Partnership; family money; illiquid takedowns', 5, 'Micro'),
    ('22NW', '22NW Fund LP; Aron Englander own money', 5, 'Micro'),
    ('Maran Capital', 'Dan Roller; sole PM own money', 5, 'Micro'),
    ('Findell Capital', 'Brian Finn; HFJ Tomorrow Titan 2023', 4, 'Smallcap'),
    ('Funicular Funds', 'Jacob Ma-Weaver / Cable Car own money', 5, 'Micro'),
    ('Voss Capital', 'Travis Cocke concentrated activist ~$500M', 4, 'Smallcap'),
    ('Bandera Partners', 'Wellington / Bylinsky activist', 4, 'Smallcap'),
    ('Greenhaven Road', 'Scott Miller own money; ~$200-400M', 4, 'Smallcap'),
    ('Praesidium', 'Sammons / Hudson activist', 4, 'Smallcap'),
    ('Alta Fox Capital', 'Connor Haley own money; ~$400M', 4, 'Smallcap'),
    ('Stilwell Value', 'Joseph Stilwell own money; CEF + small bank', 5, 'Micro'),
    ('Star Equity Fund', 'Jeffrey Eberwein own money; smallcap', 5, 'Micro'),
    ('Veradace Partners', 'Vezendan + Conlin concentrated activist', 5, 'Smallcap'),
    ('FrontFour Capital', 'Loukas / Lorber / George concentrated', 4, 'Smallcap'),
    ('Privet Fund Management', 'concentrated microcap activist', 5, 'Micro'),
    ('Cove Street Capital', 'Jeff Bronchick concentrated value', 4, 'Smallcap'),
    ('Cannell Capital', 'Carlo Cannell "PE in microcap"; ~$850M', 5, 'Micro'),
    ('Engaged Capital', 'Glenn Welling; 39+ campaigns; ex-Relational', 4, 'Mid/Small'),
    ('Caligan Partners', 'David Johnson 100% healthcare; ex-Carlyle', 4, 'Smallcap'),
    ('Engine Capital', 'Arnaud Ajdler activist', 4, 'Smallcap'),
    ('Hestia Capital', 'Kurt Wolf; original GameStop activist 2012', 4, 'Micro'),
    ('Punch Card', 'Norbert Lou; ultra-concentrated 5 positions', 5, 'Mid/Large'),
    ('Robotti & Co', 'Bob Robotti; concentrated industrials/energy/EM', 5, 'Smallcap'),
    ('Fairholme', 'Bruce Berkowitz; JOE 80% lifetime conviction', 5, 'Mid'),
    ('Horizon Kinetics', 'Stahl team / Bregman post-Stahl death; TPL', 5, 'Micro/Mid'),
    ('Greenhaven Associates', 'Edgar Wachenheim III; concentrated quality', 4, 'Mid/Large'),
    ('Southeastern Asset', 'Mason Hawkins / Longleaf; concentrated', 4, 'Mid'),
    ('Brave Warrior', 'Glenn Greenberg own money; ex-Chieftain', 4, 'Mid/Large'),
    ('Indaba Capital', 'Derek Schrier ex-Farallon concentrated', 4, 'Smallcap'),
    ('JCP Investment', 'small concentrated activist', 4, 'Micro'),
    ('Aurelius Capital', 'Mark Brodsky ex-Elliott; distressed', 4, 'Mid/Small'),
    ('Polygon Investment', 'biotech concentrated', 4, 'Smallcap'),
    ('Harbert Discovery', 'smallcap activist', 4, 'Smallcap'),
    ('Barington Capital', 'James Mitarotonda activism since 2000', 4, 'Smallcap'),
]

# Load workbook
wb = load_workbook(WB_PATH)
all_ws = wb['All Activity']
rows = list(all_ws.iter_rows(values_only=True))

# Index All Activity by fund (substring match)
def fund_match(fund_name_in_row, key):
    return key.lower() in (fund_name_in_row or '').lower()

# For each fund, find their top positions (cat 1 = highest conviction), recent adds (cat 3, 4), activist filings (cat 2)
fund_summaries = []
for fund_key, blurb, sitg_score, size_typ in FUND_PROFILES:
    # Find all rows for this fund
    fund_rows = [r for r in rows[1:] if fund_match(r[1] or '', fund_key)]
    if not fund_rows:
        continue

    # Extract top conviction (cat 1), recent adds (3/4), activist (2)
    cat1 = [r for r in fund_rows if '(1)' in str(r[2] or '')]
    cat2 = [r for r in fund_rows if '(2)' in str(r[2] or '')]
    cat3 = [r for r in fund_rows if '(3)' in str(r[2] or '')]
    cat4 = [r for r in fund_rows if '(4)' in str(r[2] or '')]

    def fmt_row(r):
        """Format a row picking the meaningful pieces."""
        # Most fund-row layouts: col 3 = index/ticker, col 4 = ticker/issuer, col 5+ = details
        parts = [str(c) for c in (r[3], r[4], r[5], r[6]) if c is not None and str(c).strip() and str(c).lower() != 'none']
        return ' | '.join(parts)[:120]

    top_pos = '; '.join([fmt_row(r) for r in cat1[:3]])
    recent_activist = '; '.join([fmt_row(r) for r in cat2[:2]])
    new_init = '; '.join([fmt_row(r) for r in cat3[:2]])
    recent_add = '; '.join([fmt_row(r) for r in cat4[:2]])

    fund_summaries.append({
        'fund_key': fund_key,
        'blurb': blurb,
        'sitg_score': sitg_score,
        'size': size_typ,
        'total_rows': len(fund_rows),
        'top_position': top_pos,
        'recent_activist_filing': recent_activist,
        'new_initiations': new_init,
        'recent_adds': recent_add,
    })

print(f"Built summaries for {len(fund_summaries)} funds")
for fs in fund_summaries[:3]:
    print(f"\n{fs['fund_key']}:")
    print(f"  top: {fs['top_position'][:100]}")
    print(f"  activist: {fs['recent_activist_filing'][:100]}")

# ============ Build the sheet ============
SHEET_NAME = 'SITG Smallcap Spotlight'
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

# Position at top, after 13D Sweep
target_idx = 1
current_idx = wb.sheetnames.index(SHEET_NAME)
if current_idx != target_idx:
    wb.move_sheet(SHEET_NAME, offset=target_idx - current_idx)

# Styles
HEADER_FILL = PatternFill(start_color='006100', end_color='006100', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SITG5_FILL = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')  # gold = SITG 5
SITG4_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # light gold = SITG 4
MICRO_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # green = Micro
SMALL_FILL = PatternFill(start_color='DEEBF7', end_color='DEEBF7', fill_type='solid')  # blue = Smallcap

# Title rows
ws.cell(row=1, column=1, value="SITG SMALLCAP SPOTLIGHT — Concentrated, Skin-in-the-Game Fund Cohort")
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='006100')
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

ws.cell(row=2, column=1, value="SITG Score: 5 = PM-owned firm, eats own cooking, no outside capital. 4 = concentrated activist, founder-led. Size: Micro = <$2B mkt cap targets; Smallcap = <$10B")
ws.cell(row=2, column=1).font = Font(italic=True, size=10)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

HEADER = ['SITG', 'Size', 'Fund', 'Profile', 'Top Conviction Positions (Cat 1)', 'Recent >=5% / Activist Filings (Cat 2)', 'New Initiations (Cat 3)', 'Recent Adds (Cat 4)']
for col, h in enumerate(HEADER, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Sort: SITG 5 first then 4; within score, Micro first
def sort_key(fs):
    size_order = {'Micro': 0, 'Micro/Small': 1, 'Smallcap': 2, 'Small/Micro': 1,
                  'Micro/Smallbio': 1, 'Mid/Small': 3, 'Smallbio': 2, 'Mid': 4, 'Mid/Large': 5}
    size_rank = size_order.get(fs['size'], 6)
    return (-fs['sitg_score'], size_rank, fs['fund_key'])

fund_summaries.sort(key=sort_key)

# Write rows
for i, fs in enumerate(fund_summaries, start=5):
    full = [
        fs['sitg_score'],
        fs['size'],
        fs['fund_key'],
        fs['blurb'],
        fs['top_position'] or '_(no cat 1 data in workbook)_',
        fs['recent_activist_filing'] or '_(none in window)_',
        fs['new_initiations'] or '_(none in window)_',
        fs['recent_adds'] or '_(none in window)_',
    ]
    for col, v in enumerate(full, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = Font(size=9)
    # Color by SITG
    sitg_fill = SITG5_FILL if fs['sitg_score'] == 5 else SITG4_FILL
    ws.cell(row=i, column=1).fill = sitg_fill
    # Color size column
    if 'Micro' in fs['size']:
        ws.cell(row=i, column=2).fill = MICRO_FILL
    elif 'Small' in fs['size']:
        ws.cell(row=i, column=2).fill = SMALL_FILL

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 60
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 45
ws.column_dimensions['H'].width = 45

# Row heights
for i in range(5, 5 + len(fund_summaries)):
    ws.row_dimensions[i].height = 60

ws.freeze_panes = 'D5'
ws.auto_filter.ref = f'A4:H{len(fund_summaries)+4}'

wb.save(WB_PATH)

print(f"\nWrote '{SHEET_NAME}' with {len(fund_summaries)} funds")
sitg5 = sum(1 for f in fund_summaries if f['sitg_score'] == 5)
sitg4 = sum(1 for f in fund_summaries if f['sitg_score'] == 4)
micro = sum(1 for f in fund_summaries if 'Micro' in f['size'])
print(f"  SITG 5: {sitg5}")
print(f"  SITG 4: {sitg4}")
print(f"  Micro-cap focus: {micro}")
