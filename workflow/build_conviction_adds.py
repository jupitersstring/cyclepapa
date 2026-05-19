"""Add two new sheets to fund_activity_last_6mo.xlsx:
   - 'Conviction Adds': add events (new init or material +)
   - 'Micro-Cap Conviction Adds': filtered to small/micro-cap

Heuristic: parse cat (3) and (4) tables for every fund. Direction signal = 'new' or '+%'
or '+$' in the narrative. Aggregate by ticker.

Mega/large-cap exclusion list lets us produce a focused micro-cap subset."""
import re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MD_FILES = [
    '/tmp/research_us_activists.md',
    '/tmp/research_us_activists_more.md',
    '/tmp/research_activists_t3.md',
    '/tmp/research_smallcap_specialsits.md',
    '/tmp/research_smallcap_tier2.md',
    '/tmp/research_smallcap_value_t3.md',
    '/tmp/research_intl_distressed.md',
    '/tmp/research_intl_more.md',
    '/tmp/research_distressed_eventdriven.md',
    '/tmp/research_japan_distressed_t3.md',
    '/tmp/research_tiger_cubs.md',
    '/tmp/research_legendary_fos.md',
    '/tmp/research_value_multistrat.md',
    '/tmp/research_europe_asia_t2.md',
    '/tmp/research_biotech_specialists.md',
    '/tmp/research_smaller_activists.md',
    '/tmp/research_quality_em.md',
    '/tmp/research_megamulti.md',
    '/tmp/research_microcap_intl_t4.md',
    '/tmp/research_family_office_filers.md',
    '/tmp/research_small_concentrated_t4.md',
    '/tmp/research_skin_game_A.md',
    '/tmp/research_value_B.md',
    '/tmp/research_shorts_value_C.md',
    '/tmp/research_warrants.md',
]

# Mega/large-cap exclusion list — anything *not* in this list (and not obviously ETF/index)
# is treated as potentially small/mid/micro.
MEGA_LARGE = {
    'AMZN', 'MSFT', 'GOOGL', 'GOOG', 'NVDA', 'META', 'AAPL', 'AVGO', 'TSM', 'TSLA',
    'BRK.A', 'BRK.B', 'COST', 'NFLX', 'INTU', 'MA', 'V', 'PYPL', 'ASML', 'AMD',
    'BSX', 'EW', 'BLK', 'MCO', 'SPGI', 'JPM', 'BAC', 'WFC', 'GS', 'C', 'MS',
    'ORCL', 'CRM', 'ADBE', 'MSCI', 'NOW', 'ANET', 'CSCO', 'IBM',
    'DIS', 'KO', 'PEP', 'WMT', 'HD', 'MCD', 'NKE', 'SBUX', 'CMG', 'BKNG',
    'ABBV', 'JNJ', 'PFE', 'MRK', 'LLY', 'ABT', 'TMO', 'DHR', 'UNH', 'CVS',
    'CMCSA', 'T', 'VZ', 'CHTR',
    'AXP', 'AIG', 'MET', 'ALL', 'TRV', 'PNC', 'SCHW',
    'CVX', 'XOM', 'OXY', 'COP',
    'PG', 'KHC', 'DUK', 'NEE', 'NRG',
    'AMGN', 'GILD', 'BMY',
    'F', 'GM',
    'INTC', 'MU', 'QCOM', 'TXN',
    'BABA', 'PDD', 'JD', 'BIDU',
    'COR', 'CI', 'HUM', 'ELV', 'CNC',
    'CRH', 'BN', 'BUR', 'SE',
    'GE', 'GEV', 'CAT', 'LMT', 'BA', 'UPS',
    'WSM', 'WYNN', 'MGM', 'CCL', 'RCL', 'NCLH',
    'UBER', 'LYFT', 'ABNB', 'DASH', 'CART',
    'CP', 'CNI', 'UNP', 'NSC',
    'PLTR', 'SHOP', 'SPOT', 'BX', 'KKR',
    'AAL', 'DAL', 'LUV',
    'WBD', 'PARA', 'FOX', 'FOXA',
    'PSX', 'MPC', 'MPLX', 'MGM', 'MAR', 'HLT',
    'IT',
    # ETFs / indices treated as macro hedges, not micro
    'SPY', 'QQQ', 'IWM', 'IEF', 'TLT', 'VOO', 'IVV', 'IJR', 'IEMG', 'XLF',
    'GLD', 'IAU', 'GDX', 'XLP', 'XLE', 'XLB', 'XLK', 'XLY', 'VUG', 'IWF',
    'VB', 'FNDA', 'VOOG', 'REMX', 'EWA', 'EUFN', 'KWEB', 'EWY', 'EWZ',
    'EEM', 'EFA', 'IWF', 'FXI', 'HYG', 'AGG', 'TIP', 'IBIT', 'IBIT', 'RSP',
    'XBI', 'NEM', 'VAL', 'TLT', 'HYG', 'SLV',
}

# Mid-cap names that aren't really micro/small (avoid showing them in micro tab)
MID_LARGE = {
    'HHH', 'KVUE', 'GE', 'GEV', 'TKO', 'CVNA', 'RKT', 'COIN',
    'BABA', 'JD', 'PDD', 'BIDU',
    'JBL', 'INTA', 'FROG',
    'KMB', 'NSC', 'MS', 'JEF', 'C',
    'TKO', 'WBD', 'TGT',
    'WHR', 'NRG', 'FLUT', 'PFGC', 'WYNN', 'WPM', 'FNV', 'RGLD', 'PAAS',
    'KGC', 'NEM', 'TFPM',
    'CMG', 'CEG', 'GMS', 'BLD',
    'AGCO', 'AAP', 'NFG',
    'GD', 'NOC', 'RTX', 'BA', 'LMT',
    'TDG', 'PEP', 'PG', 'CB', 'COF',
    'JHG', 'JANUSCAP',
    'JYNT',  # JYNT is small-cap actually — leave it
    'TEVA', 'BHC', 'MASI',  # mid-cap, but include if user wants
    'CHE',
    'HTZ', 'AAL', 'DAL', 'LUV', 'EVTL',
    'NUVL',
    'RVMD', 'MDGL', 'KRYS',  # biotech mid-caps
    'PSO', 'SNN',  # ADRs
    'GPOR',
    'BALY',
    'IEP', 'CVI', 'CTRI',
    'PCG', 'TLN', 'VST',
    'TSAT', 'SBGI', 'MNRO', 'GATX', 'NOBH',
    'MATW', 'BL', 'BILL', 'M', 'VSCO', 'GIL',
    'GFF', 'TSN', 'CTSH', 'MMYT',
    'EAF', 'EFRA',
    'CGON', 'INSM',  # mid-cap biotech
    'CCO',  # Clear Channel Outdoor (~$1B)
}

MAX_MICRO_HINTS = {'micro', 'pipe', 'preclinical', 'phase 1', 'phase 2', 'small-cap',
                   'small cap', 'biotech', 'sub-', 'reorg'}


def parse_md(path):
    """Return [(fund_name, sections_dict)] for one md file."""
    with open(path) as f:
        text = f.read()
    parts = re.split(r'(?m)^##\s+(?!#)', text)
    funds = []
    for p in parts:
        if not p.strip():
            continue
        head = p.lstrip().split('\n', 1)[0].strip()
        if not head or head.startswith('#'):
            continue
        if head.lower().startswith(('cross', 'summary', 'important', 'key', 'closing',
                                    'tier ', 'notes', 'observations', 'note')):
            continue
        body = p.split('\n', 1)[1] if '\n' in p else ''
        sections = {}
        for cat in ('1', '2', '3', '4'):
            m = re.search(rf'###\s*\({cat}\)[^\n]*\n', body)
            if not m:
                continue
            start = m.end()
            m2 = re.search(r'###\s', body[start:])
            end = start + m2.start() if m2 else len(body)
            sections[cat] = body[start:end]
        funds.append((head, sections))
    return funds


def parse_table_rows(block):
    if not block:
        return []
    rows = []
    header_seen = False
    for ln in block.splitlines():
        s = ln.strip()
        if not s.startswith('|'):
            header_seen = False
            continue
        if re.match(r'^\|\s*[-:|\s]+$', s):
            header_seen = True
            continue
        if not header_seen:
            continue
        cells = [c.strip() for c in s.strip('|').split('|')]
        if any(c.lower().startswith(('_not found', '_no ', '_specific', '_data', '_top-10'))
               for c in cells):
            continue
        rows.append(cells)
    return rows


US_TICKER = re.compile(r'^([A-Z]{1,5}(?:\.[A-Z])?)$')
JP_TICKER = re.compile(r'^(\d{4})\s*(JT|JP)?$')


def extract_ticker(cell):
    if not cell:
        return None
    first = cell.split('|')[0].strip().split('(')[0].strip().split()
    first = first[0] if first else ''
    if not first:
        return None
    if US_TICKER.match(first) and first not in {'NEW', 'TOP', 'HOLD', 'NA', 'NAN',
                                                 'OR', 'AND', 'NOT', 'NEXT', 'BUY',
                                                 'CORE'}:
        return first
    m = JP_TICKER.match(first)
    if m:
        return f'{m.group(1)} JT'
    return None


def parse_dollar_hint(text):
    """Return numeric value from things like '+$1.5B', '$310M', etc., in $M."""
    if not text:
        return None
    m = re.search(r'\$([0-9.]+)\s*B', text, re.I)
    if m:
        return float(m.group(1)) * 1000
    m = re.search(r'\$([0-9.]+)\s*M', text, re.I)
    if m:
        return float(m.group(1))
    return None


def parse_pct_hint(text):
    if not text:
        return None
    m = re.search(r'\+(\d+(?:\.\d+)?)\s*%', text)
    return float(m.group(1)) if m else None


# ---- Walk all files, collect add events ----
adds = []  # list of dicts
for path in MD_FILES:
    for fund, secs in parse_md(path):
        for cat in ('3', '4'):
            block = secs.get(cat, '')
            for row in parse_table_rows(block):
                ticker = extract_ticker(row[0]) if row else None
                if ticker is None and len(row) > 1:
                    ticker = extract_ticker(row[1])
                if not ticker:
                    continue
                joined = ' '.join(row[1:])
                # is this a real add?
                joined_l = joined.lower()
                is_new = bool(re.search(r'\bnew\b', joined_l)) and 'newly' not in joined_l
                is_add = ('+' in joined or 'added' in joined_l or 'increase' in joined_l
                          or 'top up' in joined_l)
                is_initiated = 'init' in joined_l
                if not (is_new or is_add or is_initiated):
                    continue
                # is this not a trim?
                if 'trim' in joined_l or '-$' in joined or 'reduced' in joined_l:
                    if not (is_new and ('+' in joined or 'added' in joined_l)):
                        continue
                dollar = parse_dollar_hint(joined)
                pct = parse_pct_hint(joined)
                adds.append({
                    'ticker': ticker,
                    'fund': fund,
                    'cat': cat,
                    'narrative': joined[:280],
                    'is_new': is_new,
                    'dollar_m': dollar,
                    'pct_add': pct,
                })

print(f"Captured {len(adds)} add events.")

# ---- Aggregate per ticker ----
by_ticker = defaultdict(list)
for a in adds:
    by_ticker[a['ticker']].append(a)

# Score each ticker
ticker_rows = []
for tk, evts in by_ticker.items():
    funds = sorted({e['fund'] for e in evts})
    new_initiations = sum(1 for e in evts if e['is_new'])
    max_dollar = max([e['dollar_m'] for e in evts if e['dollar_m']] or [0])
    sum_dollar = sum([e['dollar_m'] for e in evts if e['dollar_m']] or [0])
    max_pct = max([e['pct_add'] for e in evts if e['pct_add']] or [0])
    narratives = sorted({f"{e['fund']}: {e['narrative']}" for e in evts})[:6]
    ticker_rows.append({
        'ticker': tk,
        'n_funds': len(funds),
        'n_new': new_initiations,
        'max_dollar_m': max_dollar,
        'sum_dollar_m': sum_dollar,
        'max_pct_add': max_pct,
        'funds': funds,
        'narratives': narratives,
    })

# Rank
ticker_rows.sort(key=lambda d: (-d['n_funds'], -d['sum_dollar_m'], -d['max_pct_add']))

# ---- Append to workbook ----
out = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'
wb = load_workbook(out)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E78')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def border_all(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = border
            cell.alignment = wrap


# Conviction Adds sheet — all caps, ranked
for sheet in ('Conviction Adds', 'Micro-Cap Conviction Adds'):
    if sheet in wb.sheetnames:
        del wb[sheet]

ws = wb.create_sheet('Conviction Adds', 3)
ws.append(['Ticker', '# Funds Adding', '# New Initiations', 'Sum $ Adds (M)',
           'Max $ Add (M)', 'Max % Add', 'Funds', 'Sample narratives'])
style_header(ws, 8)
for t in ticker_rows[:180]:
    ws.append([
        t['ticker'],
        t['n_funds'],
        t['n_new'],
        round(t['sum_dollar_m'], 1) if t['sum_dollar_m'] else '',
        round(t['max_dollar_m'], 1) if t['max_dollar_m'] else '',
        t['max_pct_add'] if t['max_pct_add'] else '',
        ', '.join(t['funds'][:8]) + (f' (+{len(t["funds"]) - 8} more)' if len(t['funds']) > 8 else ''),
        ' | '.join(t['narratives'][:3])[:500],
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 14, 14, 16, 16, 12, 60, 80])

# Micro-Cap Conviction Adds — filter
microcap_rows = [t for t in ticker_rows
                 if t['ticker'] not in MEGA_LARGE
                 and t['ticker'] not in MID_LARGE
                 and not t['ticker'].endswith(' JT')]  # JP tickers handled separately
ws = wb.create_sheet('Micro-Cap Conviction Adds', 4)
ws.append(['Ticker', '# Funds Adding', '# New Initiations', 'Sum $ Adds (M)',
           'Max $ Add (M)', 'Max % Add', 'Funds (smart money)', 'Sample narratives'])
style_header(ws, 8)
for t in microcap_rows[:180]:
    ws.append([
        t['ticker'],
        t['n_funds'],
        t['n_new'],
        round(t['sum_dollar_m'], 1) if t['sum_dollar_m'] else '',
        round(t['max_dollar_m'], 1) if t['max_dollar_m'] else '',
        t['max_pct_add'] if t['max_pct_add'] else '',
        ', '.join(t['funds'][:8]) + (f' (+{len(t["funds"]) - 8} more)' if len(t['funds']) > 8 else ''),
        ' | '.join(t['narratives'][:3])[:500],
    ])
border_all(ws)
ws.freeze_panes = 'A2'
autosize(ws, [10, 14, 14, 16, 16, 12, 60, 80])

wb.save(out)
print(f"Saved {out}")
print(f"Conviction Adds rows: {len(ticker_rows)}")
print(f"Micro-cap subset rows: {len(microcap_rows)}")
print()
print("Top 25 conviction adds:")
for t in ticker_rows[:25]:
    funds_short = ', '.join(t['funds'][:4])
    print(f"  {t['ticker']:8s}  {t['n_funds']:2d} funds  ${t['sum_dollar_m']:>6.0f}M sum   ({funds_short}...)")
print()
print("Top 25 micro/small-cap conviction adds:")
for t in microcap_rows[:25]:
    funds_short = ', '.join(t['funds'][:4])
    print(f"  {t['ticker']:8s}  {t['n_funds']:2d} funds  ${t['sum_dollar_m']:>6.0f}M sum   ({funds_short}...)")
