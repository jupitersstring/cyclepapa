"""Categorize PitchBook saved searches by theme/style and build a tidy Excel output."""
import json
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open('/tmp/searches.json') as f:
    data = json.load(f)


def classify(name: str, criteria: str | None):
    """Return (theme, sub_theme) for a saved search."""
    n = (name or '').lower()
    c = (criteria or '').lower()

    # ---- Distressed / Special Situations (long bias / short bias variants) ----
    if 'boutique distressed' in n or 'boutique distresed' in n:
        if 'short bias' in n:
            return ('Distressed & Special Situations', 'Boutique Distressed – Short Bias')
        if 'long bias' in n:
            if 'control path' in n:
                return ('Distressed & Special Situations', 'Boutique Distressed – Long Bias: Control Path')
            if 'spin-off' in n or 'spin off' in n or 'stub' in n:
                return ('Distressed & Special Situations', 'Boutique Distressed – Long Bias: Spin-off / Stub')
            if 'value unlock' in n or 'insider floor' in n:
                return ('Distressed & Special Situations', 'Boutique Distressed – Long Bias: Value Unlock / Insider Floor')
            return ('Distressed & Special Situations', 'Boutique Distressed – Long Bias')
        if 'pipe' in n:
            return ('Distressed & Special Situations', 'Boutique Distressed – PIPE Activity')
        if 'public' in n:
            return ('Distressed & Special Situations', 'Boutique Distressed – Public Deals')
        return ('Distressed & Special Situations', 'Boutique Distressed – General')
    if 'distress' in n or 'distressed value' in n:
        return ('Distressed & Special Situations', 'Distressed Value / Deep Value')
    if 'deep value' in n:
        return ('Distressed & Special Situations', 'Deep Value Investors')
    if 'bankrupt' in n:
        return ('Distressed & Special Situations', 'Bankrupt / Inactive Listed')
    if 'turnaround' in n:
        return ('Distressed & Special Situations', 'Turnaround – Profitable & Growing')
    if 'special sits' in n:
        return ('Distressed & Special Situations', 'Special Situations – People')

    # ---- Activists ----
    if 'activist' in n:
        return ('Activists', 'Activist Investors')

    # ---- PIPEs ----
    if 'pipe' in n:
        if 'service provider' in n:
            return ('PIPEs', 'PIPE Service Providers')
        if 'people' in n:
            return ('PIPEs', 'PIPE People Screeners')
        if 'growing' in n:
            return ('PIPEs', 'PIPE – Growing Businesses')
        if 'young' in n:
            return ('PIPEs', 'PIPE – Young Businesses')
        if 'inflecting' in n:
            return ('PIPEs', 'PIPE – Inflecting Businesses')
        if 'emerging market' in n or 'em ' in n or n.startswith('em '):
            return ('PIPEs', 'PIPE – Emerging Markets')
        if 'emerging technologies' in n:
            return ('PIPEs', 'PIPE – Emerging Technologies')
        if 'europe' in n:
            return ('PIPEs', 'PIPE Screener – Europe')
        if 'global' in n or 'broad' in n or 'tight' in n or 'smid' in n:
            return ('PIPEs', 'PIPE Screeners – Broad / Global')
        if 'very quickly growing' in n:
            return ('PIPEs', 'PIPE – Very Quickly Growing')
        if 'all pipes' in n or 'activity' in n:
            return ('PIPEs', 'PIPE Activity (All)')
        return ('PIPEs', 'PIPE – General')

    # ---- P2Ps / Take-Privates ----
    if 'p2p' in n:
        if 'pipe' in n:
            return ('Take-Privates (P2P)', 'P2P & PIPE Combined')
        if 'completed' in n:
            return ('Take-Privates (P2P)', 'Completed P2Ps')
        if 'announced' in n:
            return ('Take-Privates (P2P)', 'Announced P2Ps')
        if 'non' in n:
            return ('Take-Privates (P2P)', 'Non-Completed P2Ps')
        return ('Take-Privates (P2P)', 'All P2Ps')

    # ---- MBOs ----
    if 'mbo' in n:
        return ('Take-Privates (P2P)', 'Listed MBOs')

    # ---- Spin-offs / Spinouts ----
    if 'spin-off' in n or 'spinout' in n:
        if 'profitable' in n:
            return ('Spin-offs / Spinouts', 'Profitable Spinouts')
        if 'fast growing' in n:
            return ('Spin-offs / Spinouts', 'Fast Growing Spinouts')
        return ('Spin-offs / Spinouts', 'Growing Spin-offs / Spinouts')

    # ---- Roll-ups ----
    if 'roll up' in n or 'roll-up' in n or 'roll ups' in n:
        if 'quickly' in n:
            return ('Roll-ups', 'Quickly Growing Roll-ups')
        return ('Roll-ups', 'Growing Roll-ups')

    # ---- Government / SWF / Public Agency ----
    if 'sovereign' in n or 'swf' in n or 'game of thrones' in n:
        return ('Government / SWF', 'Sovereign Wealth Fund Screener')
    if 'government' in n or 'goverment' in n or 'gov owned' in n or 'public agency' in n or 'senator' in n:
        if 'senator' in n:
            return ('Government / SWF', 'Senator Positions')
        if 'listed' in n or 'invested' in n or 'lp' in n:
            return ('Government / SWF', 'Government Investors in Listed Entities')
        if 'agencies' in n:
            return ('Government / SWF', 'Government Agencies')
        if 'gov owned' in n:
            return ('Government / SWF', 'Government-Owned Businesses')
        return ('Government / SWF', 'Government / Public Agency Activity')

    # ---- VC / Venture ----
    if n.startswith('vc ') or ' vc ' in (' ' + n + ' ') or 'venture' in n:
        if 'board' in n:
            return ('VC / Venture', 'VC Board Seats')
        if 'mega' in n:
            return ('VC / Venture', 'VC Fund Search – Mega')
        if 'selected' in n:
            return ('VC / Venture', 'Selected VC Funds')
        return ('VC / Venture', 'VC – General')
    if 'formerly vc' in n or 'vc adj' in n or 'vc / vc' in n:
        return ('VC / Venture', 'Listed Formerly VC-Backed')
    if 'formerly accelerator' in n or 'angel backed' in n:
        return ('VC / Venture', 'Listed Formerly Accelerator / Angel-Backed')
    if 'ex sequioa' in n or 'ex sequoia' in n:
        return ('VC / Venture', 'Ex-Sequoia (alumni screen)')

    # ---- Biotech / Longevity ----
    if 'biotech' in n:
        return ('Biotech / Longevity', 'Biotech Investors')
    if 'longevity' in n:
        return ('Biotech / Longevity', 'Longevity Funds & Deals')

    # ---- Emerging Markets / Asia / Geography ----
    if 'em and natural resources' in n:
        return ('Geography & Region', 'EM & Natural Resources – SMID')
    if 'emerging market' in n and 'pipe' not in n:
        return ('Geography & Region', 'Emerging Markets – General')
    if 'em ipos' in n or 'em pipes' in n:
        return ('Geography & Region', 'EM IPOs / PIPEs')
    if 'asia' in n:
        return ('Geography & Region', 'Asia Investors – Deep Value')
    if 'ex us cheap growing' in n:
        return ('Geography & Region', 'Ex-US Cheap & Growing (incl. SimilarWeb)')
    if 'us cheap growing' in n or ('us ' in n and 'cheap growing' in n):
        return ('Geography & Region', 'US Cheap & Growing (incl. SimilarWeb)')
    if 'europe' in n and 'strategic' in n:
        return ('Geography & Region', 'Europe – Strategic Acquirers')
    if 'europe' in n and 'fund' in n:
        return ('Geography & Region', 'Europe – Fund / Buyout Universe')
    if 'malta' in n:
        return ('Geography & Region', 'Malta')
    if 'panpapers' in n:
        return ('Geography & Region', 'Panama Papers / Offshore')

    # ---- Funds (Buyout / Closed / Open / Mid-Mkt) ----
    if 'funds monitor' in n or 'open funds' in n or 'funds with raise' in n or \
       'fund universe' in n or 'closed funds' in n or 'recent fund raises' in n or \
       'investors that have raised' in n or 'tighter search of funds' in n or \
       'european funds buyout' in n or 'selected funds' in n or 'funds screener' in n:
        return ('Funds & Fundraising', 'Fund / LP Screeners')

    # ---- Dividends ----
    if 'divi' in n or 'dividend' in n:
        return ('Income / Dividends', 'Dividend Growers & Yielders')
    if 'low leverage high divi' in n:
        return ('Income / Dividends', 'Low Leverage / High Dividend')

    # ---- Share Buybacks ----
    if 'buyback' in n or 'buy back' in n or 'share repurchase' in n or 'repurchase' in n:
        if 'inflecting' in n:
            return ('Share Buybacks & Repurchases', 'Buybacks – Inflecting Businesses')
        if 'young' in n:
            return ('Share Buybacks & Repurchases', 'Buybacks – Young / Young & Growing')
        if 'fast growing' in n:
            return ('Share Buybacks & Repurchases', 'Buybacks – Fast Growing (15%/30%)')
        if 'gdp' in n:
            return ('Share Buybacks & Repurchases', 'Buybacks – GDP+ Growers')
        if 'growing' in n or '7%' in n:
            return ('Share Buybacks & Repurchases', 'Buybacks – Growing Businesses')
        if 'debt' in n:
            return ('Share Buybacks & Repurchases', 'Share Repurchases – Debt Deals')
        return ('Share Buybacks & Repurchases', 'Share Buybacks / Repurchases – General')

    # ---- Convertibles / Warrants ----
    if 'warrant' in n or 'convertible' in n or 'converted transaction' in n:
        if 'equity and debt' in n:
            return ('Convertibles & Warrants', 'Converted – Equity & Debt Package')
        if 'profitable' in n:
            return ('Convertibles & Warrants', 'Recently Converted – Profitable')
        if 'no debt' in n:
            return ('Convertibles & Warrants', 'Converted Transactions & No Debt')
        return ('Convertibles & Warrants', 'Warrants & Convertibles – General')

    # ---- Refinancing / Debt / Lenders ----
    if 'refi' in n or 'refinanc' in n:
        return ('Debt / Refinancings', 'Refinancings (Senior / Other)')
    if 'ucc filing' in n:
        return ('Debt / Refinancings', 'UCC Filings')
    if 'tlb' in n or 'term loan' in n or 'unitranche' in n:
        return ('Debt / Refinancings', 'Term Loans / Unitranche')
    if 'debt & lender' in n or 'debt and lender' in n:
        return ('Debt / Refinancings', 'Debt & Lenders')
    if 'acquisition financing' in n:
        return ('Debt / Refinancings', 'Acquisition Financings')

    # ---- IPOs ----
    if 'ipo' in n:
        if 'distressed' in n:
            return ('IPOs', 'Distressed Value IPOs')
        if 'cheapish' in n or 'cheap' in n:
            return ('IPOs', 'Cheapish IPOs (growing)')
        return ('IPOs', 'IPO Activity')

    # ---- Corporate Actions / Exits / Divestitures ----
    if 'corporate action' in n:
        return ('Corporate Actions & Exits', 'Corporate Actions – Weekly')
    if 'divestiture' in n:
        return ('Corporate Actions & Exits', 'Corporate Divestitures')
    if 'exits' in n or 'pe exits' in n:
        return ('Corporate Actions & Exits', 'PE Exits')
    if 'deals last' in n:
        return ('Corporate Actions & Exits', 'Recent Deals (L90D)')

    # ---- Web growth / SimilarWeb ----
    if 'web growth' in n or 'web visitors' in n or 'similarweb' in n or 'simweb' in n or 'webgrra' in n:
        if 'size multiple' in n:
            return ('Web Traffic & SimilarWeb', 'Web Growth + Size Multiple 100%')
        return ('Web Traffic & SimilarWeb', 'Web / Similar Web Growth Rate')
    if 'size multiple' in n:
        return ('Web Traffic & SimilarWeb', 'Size Multiple + Growth 100%')

    # ---- Cheap / Valuation Multiples ----
    if 'under 3x' in n:
        return ('Cheap Valuation Multiples', 'Under 3x (zero / low debt)')
    if 'under 4x' in n:
        return ('Cheap Valuation Multiples', 'Under 4x (zero debt, growing)')
    if 'under 8x' in n:
        return ('Cheap Valuation Multiples', 'Under 8x (zero debt, growing)')
    if 'under 10x' in n:
        return ('Cheap Valuation Multiples', 'Under 10x (growing EBITDA & revs)')
    if '3 fives' in n:
        return ('Cheap Valuation Multiples', '3 Fives – Rev / EBITDA / PE Multiple')
    if '4 fives' in n:
        return ('Cheap Valuation Multiples', '4 Fives – + 5% Dividend')

    # ---- Growth (revenue / EBITDA) ----
    if 'ebitda growth' in n:
        if '30%' in n:
            return ('Growth Screens', 'EBITDA Growth – 30% LTM')
        if '60%' in n:
            return ('Growth Screens', 'EBITDA Growth – 60% LTM')
        return ('Growth Screens', 'EBITDA Growth – General')
    if 'cheap and growing' in n or 'cheap growing' in n or 'cheapish' in n or 'cheap growing pftibale' in n:
        return ('Growth Screens', 'Cheap & Growing (Profitable)')
    if 'scalers high growth' in n:
        return ('Growth Screens', 'Scalers – High Growth')
    if '50% growth' in n:
        return ('Growth Screens', '50% Revenue Growth (LTM / since 2024)')
    if 'just profitable' in n:
        return ('Growth Screens', 'Just Profitable EBITDA')

    # ---- Selected Verticals / Emerging Technologies ----
    if 'selected verticals' in n:
        return ('Sector / Vertical', 'Quickly Growing Selected Verticals')
    if 'emerging technologies' in n:
        return ('Sector / Vertical', 'Profitable / Growing Emerging Technologies')

    # ---- Multi-bagger / Small cap ----
    if 'multibagger' in n or 'small cap investor' in n or 'smidcap' in n:
        if '30% ebitda' in n:
            return ('Multi-bagger / Small Cap', 'Multibagger – 30% EBITDA Growth')
        if 'profitable and cheap' in n:
            return ('Multi-bagger / Small Cap', 'Multibagger – Profitable & Cheap')
        if 'growing' in n:
            return ('Multi-bagger / Small Cap', 'Multibagger – Growing')
        if 'board' in n:
            return ('Multi-bagger / Small Cap', 'Multibagger – Board Positions')
        return ('Multi-bagger / Small Cap', 'Multibagger Monitor')

    # ---- People / Service Providers / CEOs ----
    if 'broad search of people' in n:
        return ('People & Service Providers', 'Broad Search of People')
    if 'ceo' in n or 'ceos' in n:
        return ('People & Service Providers', 'CEOs / VCs / Others')
    if 'service provider' in n:
        return ('People & Service Providers', 'Service Providers')
    if 'catalyst driven' in n:
        return ('People & Service Providers', 'Catalyst-Driven Investors')

    # ---- Family Office / HNWI / Hedge Fund ----
    if 'family office' in n or 'hnwi' in n:
        return ('Investor Type', 'Family Office / HNWI Activity')
    if 'hedge fund' in n:
        return ('Investor Type', 'Hedge Fund Owned Equities')
    if 'quants' in n:
        return ('Investor Type', 'Quant Investors')
    if 'pe backed listed' in n:
        return ('Investor Type', 'PE-Backed Listed Businesses')

    # ---- Keywords / Privatizations / Capital Structure ----
    if 'privatization' in n:
        return ('Keyword Searches', 'Privatizations')
    if 'capital structure' in n:
        return ('Keyword Searches', 'Capital Structure Changes')
    if 'ageing portfolio' in n:
        return ('Keyword Searches', 'Ageing Portfolio – Profitable Cos')

    # ---- VIP / Misc ----
    if 'vip screen' in n:
        return ('Other / Misc', 'VIP Screen')
    if 'untitled' in n:
        return ('Other / Misc', 'Untitled / Unlabelled')

    return ('Other / Misc', 'Uncategorised')


# Classify all
classified = []
themes = defaultdict(lambda: defaultdict(list))
for d in data:
    theme, sub = classify(d['name'], d['criteria'])
    classified.append({**d, 'theme': theme, 'sub_theme': sub})
    themes[theme][sub].append(d)

# ---- Suggested additions: gaps that would round out the collection ----
suggested = [
    # Valuation extensions
    ('Cheap Valuation Multiples', 'Under 5x (0 debt, growing) – missing tier between 4x and 8x',
     'Logical gap between existing Under 4x and Under 8x screens'),
    ('Cheap Valuation Multiples', '5 fives – rev / EBITDA / PE / div / FCF yield',
     'Extension of 3-fives / 4-fives series adding FCF yield criteria'),
    ('Cheap Valuation Multiples', 'EV/Sales under 1x, growing >20%',
     'Sales-multiple equivalent of the EV/EBITDA screens'),
    # Growth tier completeness
    ('Growth Screens', '40% Revenue Growth LTM (& cheap)',
     'Fills tier between 30% EBITDA growth and 50% revenue growth'),
    ('Growth Screens', '100% Revenue Growth LTM (hyper-growth)',
     'Top-end of growth series – mirrors web growth 100% screens'),
    ('Growth Screens', 'EBITDA Growth – 100% LTM',
     'Top tier of EBITDA growth series'),
    # PIPE gaps
    ('PIPEs', 'PIPE – Healthcare / Biotech specific',
     'Combines existing biotech-investor and PIPE themes'),
    ('PIPEs', 'PIPE – Industrials / Manufacturing specific',
     'Sector cut of PIPE universe not currently represented'),
    ('PIPEs', 'PIPE – Profitable & Cheap (cross with valuation screen)',
     'Combine PIPE flow with valuation discipline'),
    # Take-private gaps
    ('Take-Privates (P2P)', 'Failed / Withdrawn P2Ps',
     'Counterpart to Completed / Announced / Non-completed P2Ps'),
    ('Take-Privates (P2P)', 'P2P – Sponsor-backed only',
     'Filter take-private universe to PE-led deals'),
    ('Take-Privates (P2P)', 'P2P – Management-led (MBO) – Europe only',
     'Regional cut of existing Listed MBOs screen'),
    # Distressed / Special Situations gaps
    ('Distressed & Special Situations', 'Chapter 11 emergences / post-reorg equities',
     'Counterpart to Bankrupt businesses listed'),
    ('Distressed & Special Situations', 'Liability management exercises (LMEs)',
     'Distinct from refinancings – out-of-court restructurings'),
    ('Distressed & Special Situations', 'Net-net / sub-NCAV (Graham deep value)',
     'Classic deep value screen complement to existing Distressed Value'),
    # Activists
    ('Activists', 'Activist – Recently filed 13D/13G',
     'Triggered filings list to complement Activists & Squeezers'),
    ('Activists', 'Activist – Proxy fight / consent solicitation',
     'Specific catalyst sub-set of activist universe'),
    # Government / SWF
    ('Government / SWF', 'SWF co-investments with PE / VC',
     'Cross-section of SWF screener and fund screener'),
    ('Government / SWF', 'GIC / Mubadala / PIF deal flow LTM',
     'Targeted SWF tracking – mirrors Senator Positions style'),
    # Buybacks
    ('Share Buybacks & Repurchases', 'Buybacks – Insider buying alongside (double signal)',
     'Combine repurchase activity with insider open-market buys'),
    ('Share Buybacks & Repurchases', 'Tender offer / Dutch auction buybacks',
     'Specific buyback mechanic not isolated in current set'),
    # Web traffic
    ('Web Traffic & SimilarWeb', 'Web traffic decelerating – short / pair-trade ideas',
     'Counterpart to existing high-web-growth screens'),
    # Convertibles
    ('Convertibles & Warrants', 'Convertible refinancings approaching maturity',
     'Catalyst-driven complement to existing converted transactions'),
    # Geography
    ('Geography & Region', 'Japan – cheap & growing (TSE reform beneficiaries)',
     'Japan corporate governance reform is a hot theme not isolated here'),
    ('Geography & Region', 'India / SE Asia – cheap & growing',
     'EM cut not currently represented separately'),
    ('Geography & Region', 'LatAm – cheap & growing',
     'EM cut not currently represented separately'),
    # Funds
    ('Funds & Fundraising', 'First-time fund managers (Fund I) – global',
     'Fundraising cut focused on emerging managers'),
    ('Funds & Fundraising', 'Secondaries / GP-led continuation funds',
     'Fast-growing fund category not represented'),
    # Sector
    ('Sector / Vertical', 'Energy transition / critical minerals – cheap & growing',
     'Thematic cut complementing EM & Natural Resources'),
    ('Sector / Vertical', 'AI infrastructure / data-centre adjacent – cheap & growing',
     'Hot 2025/26 theme not isolated in current set'),
    ('Sector / Vertical', 'Defence / dual-use – cheap & growing',
     'Sector tailwind theme not in current set'),
    # Income / Yield
    ('Income / Dividends', 'High FCF yield (>10%) & buyback combined',
     'FCF-based yield screen complementing dividend screens'),
    # Multi-bagger
    ('Multi-bagger / Small Cap', 'Micro-cap (< $250m mkt cap) cheap & growing',
     'Smaller-cap tier than existing SMID multibagger screens'),
]

# ---- Build the workbook ----
wb = Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E78')
sub_fill = PatternFill('solid', fgColor='D9E1F2')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# Sheet 1: Summary
ws = wb.active
ws.title = 'Summary'
ws.append(['Theme', 'Sub-theme', 'Saved Searches in Set', 'Suggested Additions'])
style_header(ws, 4)

theme_order = sorted(themes.keys())
sugg_by_theme = defaultdict(int)
for theme, sub, _ in suggested:
    sugg_by_theme[theme] += 1

for theme in theme_order:
    subs = themes[theme]
    total = sum(len(v) for v in subs.values())
    ws.append([theme, f'{len(subs)} sub-themes', total, sugg_by_theme.get(theme, 0)])
ws.append(['', '', '', ''])
ws.append(['TOTAL', '', sum(len(d) for sub in themes.values() for d in sub.values()), len(suggested)])
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = wrap
# Bold totals row
last = ws.max_row
for c in ws[last]:
    c.font = Font(bold=True)
autosize(ws, [42, 28, 22, 22])
ws.freeze_panes = 'A2'

# Sheet 2: All searches (master with theme columns)
ws = wb.create_sheet('All Searches by Theme')
ws.append(['Theme', 'Sub-theme', 'Search ID', 'Saved Search Name', 'Criteria Summary'])
style_header(ws, 5)
rows_sorted = sorted(classified, key=lambda x: (x['theme'], x['sub_theme'], x['name'] or ''))
for r in rows_sorted:
    ws.append([r['theme'], r['sub_theme'], r['id'], r['name'], r['criteria']])
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = wrap
autosize(ws, [32, 42, 14, 56, 80])
ws.freeze_panes = 'A2'

# Sheet 3+: One sheet per theme
def safe_sheet_name(name):
    # Excel sheet names: max 31 chars, no : \ / ? * [ ]
    bad = re.compile(r'[:\\/\?\*\[\]]')
    s = bad.sub(' ', name)
    return s[:31]

for theme in theme_order:
    sheet = wb.create_sheet(safe_sheet_name(theme))
    sheet.append(['Sub-theme', 'Search ID', 'Saved Search Name', 'Criteria Summary'])
    style_header(sheet, 4)
    # Section rows per sub-theme
    subs_sorted = sorted(themes[theme].items())
    for sub, items in subs_sorted:
        # Sub-theme banner
        sheet.append([sub, '', '', ''])
        banner_row = sheet.max_row
        for c in range(1, 5):
            cell = sheet.cell(row=banner_row, column=c)
            cell.fill = sub_fill
            cell.font = Font(bold=True)
            cell.border = border
        # Items
        for it in sorted(items, key=lambda x: x['name'] or ''):
            sheet.append(['', it['id'], it['name'], it['criteria']])
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = wrap
    autosize(sheet, [42, 14, 56, 80])
    sheet.freeze_panes = 'A2'

# Final sheet: Suggested additions / gaps
ws = wb.create_sheet('Suggested Additions')
ws.append(['Theme', 'Suggested New Saved Search', 'Rationale (why it complements the set)'])
style_header(ws, 3)
for theme, sub, why in suggested:
    ws.append([theme, sub, why])
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = wrap
autosize(ws, [32, 60, 70])
ws.freeze_panes = 'A2'

out = '/home/user/cyclepapa/pb_searches_by_theme.xlsx'
wb.save(out)
print(f"Wrote {out}")

# Quick stats
print(f"\nThemes: {len(theme_order)}")
for t in theme_order:
    sub_count = len(themes[t])
    item_count = sum(len(v) for v in themes[t].values())
    print(f"  {t}: {sub_count} sub-themes, {item_count} searches")
print(f"\nSuggested additions: {len(suggested)}")
uncat = [r for r in classified if r['sub_theme'] == 'Uncategorised']
print(f"\nUncategorised: {len(uncat)}")
for u in uncat:
    print(f"  - {u['name']}")
