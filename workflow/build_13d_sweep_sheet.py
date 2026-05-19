"""Build a '13D Sweep — Last 14 Days' synthesis sheet from /tmp/sweep_13d_may_2026.md."""
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB_PATH = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'

# Each row: (date, filer, ticker, issuer, stake/change, action, signal_strength, source_url)
# signal_strength: 5 = NEW 13D, 4 = 13D/A escalation/agreement, 3 = NEW 13G high pct, 2 = 13G/A reaffirm, 1 = 13G/A reduction
ROWS = [
    # Section 1: NEW 13D filings (5)
    ('2026-05-14', 'STARTEEPO Invest (Frantisek Bostl)', 'XRX', 'Xerox Holdings', '5.15% (6.74M shs)', 'NEW 13D — reserves right to engage on board/strategy', 5, 'https://www.stocktitan.net/sec-filings/XRX/schedule-13d-xerox-holdings-corp-major-shareholder-acquisition-5-afea9e558c0b.html'),
    ('2026-05-14', 'River Global Investors (UK)', 'MXCT', 'MaxCyte', '6.1% (6.53M shs)', 'NEW 13D — passive intent in 13D wrapper', 5, 'https://www.stocktitan.net/sec-filings/MXCT/schedule-13d-maxcyte-inc-major-shareholder-acquisition-5-3881b5979146.html'),
    ('2026-05-14', 'William X. Kerber III (insider)', 'PDYN', 'Palladyne AI', '5.0% (2.37M shs)', 'NEW 13D — insider crossing threshold', 5, ''),

    # Section 2: 13D/A escalations (4)
    ('2026-05-13', 'RA Capital Healthcare Fund', 'ARTV', 'Artiva Biotherapeutics', '→ 33.7% (16.38M shs)', '13D/A UP — $100M new $ deployed (largest concentrated bet in window)', 5, 'https://www.stocktitan.net/sec-filings/ARTV/schedule-13d-a-artiva-biotherapeutics-inc-amended-major-shareholder-r-560ba2a26b9d.html'),
    ('2026-05-13', 'Hill Path Capital (Scott Ross)', 'PRKS', 'United Parks & Resorts', '→ 57.9%', '13D/A — buyback drove % up', 4, ''),
    ('2026-05-13', 'CHS US Investments / CHS group', 'RENT', 'Rent the Runway', '→ 59.8% (19.98M shs)', '13D/A — controlling sponsor disclosure', 4, ''),
    ('2026-05-13', 'Gateway Runway / Nexus Capital', 'RENT', 'Rent the Runway', '12.8% (4.27M shs)', '13D/A — CEO Hyman resign trigger', 4, ''),
    ('2026-05-13', 'V3 Holding / Vavilovs (Bitfury)', 'CIFR', 'Cipher Mining', '15.0% (61.3M)', '13D/A — $100M prepaid forward entered', 4, ''),
    ('2026-05-13', 'Weichai Power HK / Shandong Heavy', 'BLDP', 'Ballard Power', '→ 13.02% (39.25M)', '13D/A DOWN — 2 board seats lost', 3, ''),
    ('2026-05-13', 'Cantor Fitzgerald (Brandon Lutnick)', 'SATL', 'Satellogic', '→ 5.7% (7.86M shs)', '13D/A DOWN — open-market sales', 2, ''),
    ('2026-05-12', 'JANA Partners', 'ALKT', 'Alkami Technology', '→ 4.99% (5.34M shs)', '13D/A DOWN — deliberately to engage privately', 4, 'https://www.stocktitan.net/sec-filings/ALKT/schedule-13d-a-alkami-technology-inc-amended-major-shareholder-report-e72e08c8ba5f.html'),
    ('2026-05-12', 'MPM BioVentures / UBS Oncology', 'HOWL', 'Werewolf Therapeutics', '→ 3.8% (Gadicke aggr.)', '13D/A DOWN — sales $0.62-$0.77', 2, ''),
    ('2026-05-12', 'Catherine Spear (Co-Founder)', 'FIGS', 'FIGS Inc', '15.0% (27.67M Cl A)', '13D/A — dual-class composite disclosure', 3, ''),
    ('2026-05-12', 'Heather Hasson (Co-Founder)', 'FIGS', 'FIGS Inc', '→ 8.4% (14.42M)', '13D/A — tax-withholding sale', 2, ''),
    ('2026-05-11', 'Onex Corp + Apollo consortium', 'EEX', 'Emerald Holding', '93.0% (184M shs)', '13D/A — going-private merger', 5, ''),
    ('2026-05-11', 'Alexandra Seros (Ulloa family trust)', 'EVC', 'Entravision', '13.61% (11.25M)', '13D/A — post-Walter Ulloa estate', 3, ''),
    ('2026-05-08', 'Starboard Value (Jeff Smith)', 'BLMN', "Bloomin' Brands", '4.9% direct + 9.3% w/ swaps', '13D/A — DOWN direct, kept econ via swap', 4, 'https://www.stocktitan.net/sec-filings/BLMN/schedule-13d-a-bloomin-brands-inc-amended-major-shareholder-report-d9250ac0b052.html'),
    ('2026-05-08', 'Pershing Square (Bill Ackman)', 'QSR', 'Restaurant Brands Intl', '7.8% (27.14M)', '13D/A — reorg add PSUS as reporting', 3, ''),
    ('2026-05-07', 'Pershing Square', 'PSUS', 'Pershing Square USA Ltd', '100% Series A pref + 4M common', 'NEW 13D — anchor at IPO close', 5, 'https://www.stocktitan.net/sec-filings/PSUS/schedule-13d-pershing-square-usa-ltd-major-shareholder-acquisition-5-dd2533660e1a.html'),
    ('2026-05-06', 'Starboard Value', 'MTCH', 'Match Group', '4.6% (10.83M) + swaps', '13D/A — derivative structure clarified', 4, ''),
    ('2026-05-06', 'Fairfax Financial (Prem Watsa)', 'BB', 'BlackBerry', '→ 4.5% (26.26M shs)', '13D/A DOWN — fell below 5%', 2, ''),
    ('2026-05-04', 'Brett Blundy / BBRC Intl', 'VSCO', "Victoria's Secret", '13.0% (~10.3M)', '13D/A flat + preliminary proxy filed', 5, ''),

    # 13D/As: Asia/founder disclosures
    ('2026-05-15', 'Ahishay Sardes (Co-Founder)', 'REE', 'REE Automotive', '6.14% Cl A / 26.9% voting', '13D/A UP — RSU vesting', 3, ''),
    ('2026-05-15', 'Nio Capital / Abundant Grace', 'UXIN', 'Uxin Ltd', '16.2% / 14.4% / 7.4% / 0.9%', '13D/A — share distribution restructure', 3, ''),
    ('2026-05-14', 'Redmile Group (Jeremy Green)', 'ATRA', 'Atara Biotherapeutics', '9.9% (capped by blocker)', '13D/A UP', 3, ''),

    # Section 3: NEW 13G filings (>=5% passive)
    ('2026-05-15', 'Anson Funds Management', 'LAES', 'SEALSQ Corp', '5.6% (12.46M shs)', 'NEW 13G — passive', 3, ''),
    ('2026-05-15', 'Nantahala Capital Management', 'PTON', 'Peloton Interactive', '7.31% (29.94M shs)', 'NEW 13G — passive', 3, 'https://www.stocktitan.net/sec-filings/PTON/schedule-13g-peloton-interactive-inc-passive-investment-disclosure-5-b0961729bebe.html'),
    ('2026-05-14', 'Aristeia Capital', 'MEVO', 'M Evo Global Acquisition II', '6.11% (1.83M Units)', 'NEW 13G — SPAC arb', 3, ''),
    ('2026-05-14', 'Bank of America', 'AD', 'Array Digital Infrastructure', '5.7% (3.06M shs)', 'NEW 13G', 2, ''),
    ('2026-05-13', 'Capital World Investors (Cap Group)', 'CTVA', 'Corteva', '5.5% (36.86M shs)', 'NEW 13G', 2, ''),
    ('2026-05-13', 'Aristeia Capital', 'PAAC', 'Proem Acquisition Corp I', '5.66% (1.04M Units)', 'NEW 13G — SPAC arb', 3, ''),
    ('2026-05-12', 'State Street', 'NKE', 'NIKE', '5.0% (59.59M shs)', 'NEW 13G — passive', 2, ''),
    ('2026-05-15', 'AQR Capital Management', 'TXRH', 'Texas Roadhouse', '5.49% (3.62M shs)', 'NEW 13G — reporting', 2, ''),

    # Section 4: 13G/A material amendments
    ('2026-05-15', 'Western Digital', 'SNDK', 'SanDisk Corp', '→ 0.7% (1.04M)', '13G/A DOWN — distribution wind-down', 1, ''),
    ('2026-05-15', 'Logos Global Management', 'PTN', 'Palatin Technologies', '9.9% (incl warrants)', '13G/A — post reverse-split', 3, ''),
    ('2026-05-15', 'Michael Cannon-Brookes (founder)', 'TEAM', 'Atlassian', '22.7% Cl A / 42.7% voting', '13G/A — founder control', 3, ''),
    ('2026-05-15', 'T. Rowe Price Associates', 'VNOM', 'Viper Energy', '→ 4.7% (9.11M shs)', '13G/A DOWN below 5%', 1, ''),
    ('2026-05-15', 'Armistice Capital (Steven Boyd)', 'AUTL', 'Autolus Therapeutics', '5.94% (15.8M ADS)', '13G/A reaffirm', 2, ''),
    ('2026-05-15', 'AQR Capital Management', 'PNFP', 'Pinnacle Financial Partners', '→ 2.64% (3.99M shs)', '13G/A DOWN sub-5%', 1, ''),
    ('2026-05-15', 'RTW Investments (Roderick Wong)', 'REPL', 'Replimune', '8.9% (7.36M shs)', '13G/A reaffirm', 2, ''),
    ('2026-05-15', 'Paradigm BioCapital', 'EYPT', 'EyePoint', '→ 3.8% / 3.3%', '13G/A DOWN below 5%', 1, ''),
    ('2026-05-14', 'Point72 Asset Management', 'TNXP', 'Tonix Pharmaceuticals', '6.9% (1.10M shs)', '13G/A reporting', 2, ''),
    ('2026-05-13', 'Bank of Montreal', 'BNS', 'Bank of Nova Scotia', '13.02% (161.56M)', '13G/A reporting', 2, ''),
    ('2026-05-13', 'Oaktree Capital', 'CBL', 'CBL & Associates Properties', '→ 2.65% (820k shs)', '13G/A DOWN below 5%', 1, ''),
    ('2026-05-08', 'Millennium Management (Englander)', 'TPG', 'TPG Inc', '→ 4.3% (6.65M Cl A)', '13G/A DOWN sub-5%', 1, ''),
    ('2026-05-07', 'JSTX Holdings', 'BTE', 'Baytex Energy', '5.29% (38.14M)', '13G/A reporting', 2, ''),
    ('2026-05-04', 'Nantahala Capital Mgmt', 'FOSL', 'Fossil Group', '9.99% (5.88M)', '13G/A reporting (capped)', 2, ''),

    # Section 5: Activist letters / PX14A6G / campaigns
    ('2026-05-04', 'Brett Blundy / BBRC (13%)', 'VSCO', "Victoria's Secret", 'PROXY — vote AGAINST Chair James + Dir. Naficy at 2026 AGM', 'CAMPAIGN — preliminary proxy', 5, ''),
    ('2026-05-03', 'Impactive Capital (Lauren Taylor Wolfe)', 'WEX', 'WEX Inc', 'SETTLEMENT — 3 nominees (Adams/Alemany/Wolfe) added; board to 11; Chair/CEO split', 'CAMPAIGN WIN — settlement', 5, 'https://www.stocktitan.net/sec-filings/WEX/dfan14a-wex-inc-sec-filing-d9eccdaa7796.html'),
    ('~2026-05-13', 'Saba Capital', 'ASA', 'ASA Gold & Precious Metals', 'BUYOUT PROPOSAL — BDC-aligned transaction; 31.9% stake', 'CAMPAIGN — non-binding', 5, ''),
    ('2026-04-24 → 5/2', 'Radoff-JEC (Bradley Radoff + Michael Torok)', 'SEER', 'Seer Inc', '$2.35/sh + CVR cash buyout proposal; WHITE proxy slate intended', 'CAMPAIGN — buyout + slate', 5, ''),
    ('2026-05-12', 'Diana Shipping', 'GNK', 'Genco Shipping', 'HOSTILE TENDER $23.50/sh; rejected by Genco board 5/15', 'CAMPAIGN — hostile rejected', 4, ''),
    ('Apr-May 2026', 'Elliott Investment Mgmt', 'LSEG (UK)', 'London Stock Exchange Group', 'Push for portfolio review + £5B buyback', 'CAMPAIGN — UK', 4, ''),
    ('ongoing', 'Veradace Partners (Vezendan/Conlin)', 'RPAY', 'Repay Holdings', '8.6%; nominated Vezendan + W. Jacobs for 2026 AGM', 'CAMPAIGN — slate', 4, ''),
    ('Mar 24 amend', 'Trian + General Catalyst (Jupiter)', 'JHG', 'Janus Henderson', 'Merger price RAISED $49 → $52/share', 'CAMPAIGN WIN', 5, ''),
    ('pre-window', 'Two Seas Capital', 'CORZ', 'Core Scientific', '1st independent director deadline extended to 5/30/2026', 'COOPERATION', 3, ''),
    ('2026-05-09', 'Apollo / Onex', 'EEX', 'Emerald Holding', 'Merger agreement + >90% written consent same day', 'CAMPAIGN WIN — go-private', 5, ''),
    ('~2026-05-13', 'Biogen', 'APLS', 'Apellis Pharmaceuticals', 'Tender expired 11:59 PM 5/13; merger closed', 'CAMPAIGN COMPLETE', 4, ''),
]

# Header
HEADER = ['Date', 'Filer', 'Ticker', 'Issuer', 'Stake / Change', 'Action / Type', 'Signal', 'In Workbook?', 'Source URL']

wb = load_workbook(WB_PATH)

# Existing fund sheets (for in_workbook check)
def in_workbook_check(filer_name):
    if not filer_name:
        return ''
    # Strip suffixes
    fl = filer_name.lower()
    for skip in ['llc', 'lp', 'l.p.', 'inc', 'ltd', 'limited', 'capital', 'partners', 'asset', 'management', 'mgmt', 'group', 'fund', 'advisors', 'advisers', 'investments', 'investment', '/', '(', ')']:
        fl = fl.replace(skip, ' ')
    fl = ' '.join(fl.split())
    if len(fl) < 4:
        return 'No'
    # Tokens to match
    tokens = [t for t in fl.split() if len(t) >= 4]
    if not tokens:
        return 'No'
    first_token = tokens[0]
    for s in wb.sheetnames:
        sn = s.lower()
        if first_token in sn:
            return 'Yes'
    return 'No'

# Create sheet
SHEET_NAME = '13D Sweep Last 14 Days'
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

# Position near other synthesis tabs (top of book)
target_idx = 0  # move to position 0 (very top)
current_idx = wb.sheetnames.index(SHEET_NAME)
if current_idx != target_idx:
    wb.move_sheet(SHEET_NAME, offset=target_idx - current_idx)

# Styles
HEADER_FILL = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
SIGNAL_FILLS = {
    5: PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),  # gold - new 13D
    4: PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid'),  # light gold
    3: PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # very light
    2: PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),  # green
    1: PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),  # pale red - reductions
}
IN_WB_FONT = Font(bold=True, color='006100')

# Title row
ws.cell(row=1, column=1, value=f"13D / 13G SWEEP — Last 14 Days (May 2-16, 2026) — Refreshed 2026-05-16")
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color='C00000')
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

# Subtitle row
ws.cell(row=2, column=1, value=f"Signal: 5=NEW 13D | 4=13D/A escalation/campaign | 3=NEW 13G high pct | 2=13G/A reaffirm | 1=13G/A reduction. Sorted by date desc, then signal desc.")
ws.cell(row=2, column=1).font = Font(italic=True, size=10)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

# Header (row 4)
for col, h in enumerate(HEADER, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Sort rows: by date desc, then signal desc
def sort_key(r):
    date = r[0] or ''
    # Push '~' and 'ongoing' / 'pre-window' to bottom
    sortable_date = date if date and date[0].isdigit() else '1900-01-01'
    return (sortable_date, r[6])

ROWS_SORTED = sorted(ROWS, key=sort_key, reverse=True)

# Write rows
new_13d_count = 0
campaign_count = 0
in_wb_count = 0
for i, row in enumerate(ROWS_SORTED, start=5):
    in_wb = in_workbook_check(row[1])
    full = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], in_wb, row[7]]
    fill = SIGNAL_FILLS.get(row[6], None)
    for col, v in enumerate(full, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = Font(size=10)
        if fill:
            c.fill = fill
    # Highlight in-workbook filer
    if in_wb == 'Yes':
        ws.cell(row=i, column=8).font = IN_WB_FONT
        in_wb_count += 1
    if row[6] == 5:
        if 'NEW 13D' in row[5]:
            new_13d_count += 1
        if 'CAMPAIGN' in row[5] or 'PROXY' in row[5] or 'BUYOUT' in row[5] or 'SETTLEMENT' in row[5]:
            campaign_count += 1

# Column widths
ws.column_dimensions['A'].width = 13
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 9
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 32
ws.column_dimensions['F'].width = 50
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 13
ws.column_dimensions['I'].width = 60

# Freeze panes
ws.freeze_panes = 'A5'

# Auto-filter
ws.auto_filter.ref = f'A4:I{len(ROWS_SORTED)+4}'

wb.save(WB_PATH)
print(f"Wrote '{SHEET_NAME}' sheet with {len(ROWS_SORTED)} rows")
print(f"  NEW 13Ds: {new_13d_count}")
print(f"  Active campaigns/settlements: {campaign_count}")
print(f"  Filers already in workbook: {in_wb_count}")
