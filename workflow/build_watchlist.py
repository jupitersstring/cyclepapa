"""Add a 'User-Suggested Watchlist' sheet to fund_activity_last_6mo.xlsx that catalogs
the ~150 hidden microcap managers from the user's research dump. Each row tagged
with whether the fund is already covered in the workbook's per-fund tabs.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Pull the list of per-fund sheets currently in workbook to mark "Already in Workbook"
wb = load_workbook('/home/user/cyclepapa/fund_activity_last_6mo.xlsx')
already_sheets = {s.lower() for s in wb.sheetnames}


def in_workbook(name):
    """Heuristic: does any sheet name contain this fund?"""
    n = name.lower()
    # First strip parentheticals / "the" prefix
    short = n.split('(')[0].split('/')[0].strip()
    # Try a few keyword tokens
    tokens = [w for w in short.split() if len(w) > 3]
    for s in already_sheets:
        s2 = s.lower()
        # all top-3 tokens must appear in some sheet name
        match_count = sum(1 for t in tokens[:3] if t in s2)
        if match_count >= min(2, len(tokens)):
            return 'Yes'
    return 'No'


# Catalog rows: (Tier, Fund/Filer, Key Person(s), Location, Style/AUM, Notable holdings/campaigns)
ROWS = [
    # === TIER 1: Individual / family-office filers ===
    ('Tier 1: Individual / FO filers', 'MILFAM LLC (Lloyd Miller III estate)', 'Neil Subin', 'USA', 'Microcap 13D, sub-13F. CIK 0001029454. Manages 19+ Miller family trusts.', 'Scully Royalty (SRL) 13.0% — nominated 5 directors Dec 2025 AGM. Historical: Straight Path 10x+, Spark Networks, Centrus Energy, Transworld'),
    ('Tier 1: Individual / FO filers', 'Peter H. Kamin / 3K LP', 'Peter Kamin (ex-ValueAct co-founder 2000-2011)', 'USA', 'No 13F. 24+ 13D filings via 3K LP + Trusts.', 'Chairman: Tile Shop (TTS), Calloway\'s Nursery (CLWY), Rand Worldwide. Director: Psychemedics (PMD), MAM Software'),
    ('Tier 1: Individual / FO filers', '22NW Fund LP', 'Aron English, Bryson Hirai-Hadley', 'Seattle WA', 'CIK 0001640809. ~$237M 13F. Files w/ 22NW Inc.', 'Culp (CULP) 14.7%; Stoneridge (SRI, English on board Feb 2026); Lifecore (LFCR) 12.4%; DIRTT 29.3%; Farmer Bros (FARM) 9.3%'),
    ('Tier 1: Individual / FO filers', 'Bradley Radoff / BLR Partners + Radoff Family Foundation', 'Bradley Radoff (ex-Citadel/Third Point); often joint w/ Joshua Schechter', 'Houston TX', 'CIK 0001380585. Olshan Frome Wolosky counsel.', 'VAALCO Energy (EGY) board; CDI Corp (sold 2017); Support.com (SPRT) ~50x squeeze 2021; EMCORE; Newpark Resources 6.0%'),
    ('Tier 1: Individual / FO filers', 'James C. Pappas / JCP Investment Management', 'James Pappas (Pappas Restaurants heir)', 'Houston TX', '$252M ADV / $126M 13F. Restaurant/consumer activist.', 'Red Robin (RRGB) 14.7% + 2 board seats + $8.3M PIPE; Cheesecake (CAKE) 2%; Denny\'s (DENN) 1.6%; Tandy Leather (TLF) board; UNFI'),
    ('Tier 1: Individual / FO filers', 'Mark E. Schwarz / Newcastle Capital Management', 'Mark Schwarz', 'Dallas TX', 'CIK 0001212993. Founded 1993. Schwarz 2012 Family Trust + Newcastle Partners.', 'Chair: Wilhelmina (WHLM), Rave Restaurant (RAVE), Hallmark Financial (HALL, 24+ 13D/A amends), Bell Industries'),
    ('Tier 1: Individual / FO filers', 'Eric Singer / VIEX Capital Advisors', 'Eric Singer', 'USA', 'Small-cap tech activist specialist.', 'A10 Networks (ATEN) Chair; Quantum (QMCO) proxy win; Mattersight (sold to NICE); ESIO (acquired by MKS); Support.com'),
    ('Tier 1: Individual / FO filers', 'Charles L. Frischer', 'Charles Frischer (ex-Zephyr Mgmt)', 'Seattle WA', 'CIK ~0001632304. Individual + Libby Frischer Family Partnership + IRAs.', 'RAIT Financial Trust 9.8% (2018); Regional Health (RHE) preferreds; Presidential Realty (PDNLB) 7.2%'),
    ('Tier 1: Individual / FO filers', 'Jeffrey Eberwein / Star Equity Fund (formerly Lone Star Value)', 'Jeffrey Eberwein', 'USA', 'CIK ~0001597264. CAVEAT: SEC sanctions 2017 ($90K), 2020 ($25K).', 'Digirad/Star Equity (STRR) control; Hudson Global (HSON) Chair; ATRM merger. Partnerships w/ Engaged, JCP, Heartland'),
    ('Tier 1: Individual / FO filers', 'J. Daniel Plants / Voce Capital Management', 'J. Daniel Plants (ex-Sullivan & Cromwell / Goldman M&A / JPM)', 'USA', '~$57.8M 13F. Concentrated ~12 longs.', 'Argo (ARGO) board → sold to Brookfield; Natus (BABY) 2018 proxy → sold to ArchiMed; Air Methods (AIRM) → KKR take-private; ITG 46% IRR'),
    ('Tier 1: Individual / FO filers', 'Joseph Stilwell / Stilwell Group', 'Joseph Stilwell', 'USA', 'CIK 0001034941. ~$275M ADV. 60+ 13D filings since 2000 — most prolific community-bank/thrift activist.', 'Ottawa Bancorp (OTTW) 15.4% standstill Feb 2026; Kingsway (KFS); Provident; Catalyst; IF Bancorp; Sound Financial; Central Plains'),
    ('Tier 1: Individual / FO filers', 'Braden M. Leonard / BML Investment Partners', 'Braden Leonard', 'Zionsville IN', 'CIK 0001373604. NO 13F. 40+ 13D + 187+ 13G filings.', 'Adverum (ADVM) 13D May 2025; Forte Biosciences (FBRX) 13D May 2024; Matrixx Initiatives (taken private)'),
    ('Tier 1: Individual / FO filers', 'Eric Shahinian / Camac Partners', 'Eric Shahinian (ex-Kingstown)', 'NYC (350 Park Ave)', 'CIK 0001516478. ~$205M 13F. Founded 2011.', 'Cedar Realty (CDR) 13D Nov 2020 — pushed for sale → merged w/ Wheeler; AmBase (ABCP); TuSimple (TSP) 2024; Forte Biosciences 2022'),
    ('Tier 1: Individual / FO filers', 'Ronald L. Chez', 'Ronald Chez', 'Chicago (1524 N. Astor)', 'Individual + IRAs + Chez Family Foundation.', 'Repligen (RGEN) 9.2% in 2012 → ~50x multibagger ($4→$200+); Ironclad Performance Wear (ICPW) board → sold 2017; OptimizeRx (OPRX) 7.9% 2019'),
    ('Tier 1: Individual / FO filers', 'Salvatore Muoio / S. Muoio & Co.', 'Sal Muoio', 'NYC', 'Deep-value microcap activist. Net-cash / liquidation plays.', 'MathStar (MATH) liquidation campaign 2008. Style: passes 5% urging wind-down'),
    ('Tier 1: Individual / FO filers', 'Ephraim Fields / Echo Lake Capital', 'Ephraim Fields (ex-DLJ/Wasserstein Perella)', 'NYC (501 Madison)', 'Individual filer. Net-cash specialist.', 'Arotech (ARTX) 8.1% 2015 (w/ Cruiser/Rosenbloom); NeuroMetrix (NURO) 2024 urging liquidation @ 70% net-cash discount; Edgewater; Tix Corp'),
    ('Tier 1: Individual / FO filers', 'Roger Lipton / RHL Associates LP', 'Roger Lipton (Liptonfinancial.com)', 'USA', 'CIK 0001071193. Restaurant specialist.', 'Boston Restaurant Associates (BRA) 21.76%'),
    ('Tier 1: Individual / FO filers', 'Lawrence J. Goldstein / Santa Monica Partners LP', 'Lawrence Goldstein', 'NYC', 'NY family partnership since 1982. Annual letters 1982-2021.', 'LandBridge (LB) active on earnings calls'),
    ('Tier 1: Individual / FO filers', 'Tim Eriksen / Eriksen Capital Management', 'Tim Eriksen', 'Custer WA (8695 Glendale)', 'CIK 0001616134. SMA business. Separate from Cedar Creek.', 'Solitron Devices (SODI) 6 amends since 2014; Nocopi (NNUP); Sitestar/Enterprise Diversified'),
    ('Tier 1: Individual / FO filers', 'Robert L. Chapman / Chapman Capital', 'Bob Chapman', 'El Segundo CA', '17 13Ds since 1997, ~30% annualized historic. Now dormant.', 'USA Detergents (sold 6mo); Vitesse Semiconductor; Embarcadero Technologies; Glenayre; Agile Software'),
    ('Tier 1: Individual / FO filers', 'Andrew Wiederhorn / Fog Cutter Holdings', 'Andrew Wiederhorn', 'USA', 'Files multiple 13Ds.', 'FAT Brands (FAT) — controls'),
    ('Tier 1: Individual / FO filers', 'Gregory Fortunoff (Jaftex Corp)', 'Gregory + family (Laurie, Scott, Jill, Darren Gerstenblatt)', 'USA', 'CIK 0001212542.', 'FAT Brands 6.7%+ (9.9% waiver); Kingstone Companies (KINS) 3.6%'),
    ('Tier 1: Individual / FO filers', 'Will Wyatt / Galloway Capital Partners', 'Will Wyatt', 'USA', 'CIK 0001892207. Microcap activist.', 'WW International (WW) 2025 distress; PodcastOne (PODC) Apr 2025'),
    ('Tier 1: Individual / FO filers', 'Christopher P. Mittleman / Mittleman Brothers', 'Chris Mittleman', 'NYC', 'Concentrated deep-value.', 'LodgeNet; Spectrum Brands; Revlon (13G); Aimia (Canada, settled 2023)'),
    ('Tier 1: Individual / FO filers', 'Daniel Lewis / Orange Capital Ventures', 'Daniel Lewis', 'West Palm Beach FL', 'CIK 0001426756. Distressed/Canadian-orphan.', 'Bellatrix Exploration; Pinnacle Entertainment (sold to Penn); PHH Corporation'),
    ('Tier 1: Individual / FO filers', 'Joseph A. De Perio / Clinton Relational Opportunity', 'Joe De Perio', 'NYC', 'Nominee on Clinton Group small-cap slates.', 'Wet Seal 2012 consent solicitation; EveryWare Global; A. Schulman'),
    ('Tier 1: Individual / FO filers', 'Mithaq Capital SPC', 'Turki AlRajhi (Saudi family office)', 'Riyadh, Saudi Arabia', 'CIK 0001980117. Multi-billion family office.', 'The Children\'s Place (PLCE) 55-56% in 2024 + $90M unsecured financing + nominated 11 directors — PLCE +162%'),
    ('Tier 1: Individual / FO filers', 'Ned L. Sherwood', 'Ned Sherwood', 'USA', 'Individual filer.', 'Navios (NMM) 13D Nov 2024'),
    ('Tier 1: Individual / FO filers', 'Knighted Pastures LLC', 'n/d', 'USA', 'Family-office trust filer (Snowball-listed).', 'Multiple microcap 13Ds'),
    ('Tier 1: Individual / FO filers', 'Saunders Family Trust', 'n/d', 'USA', 'Family-office trust filer (Snowball-listed).', 'Multiple microcap 13Ds'),
    ('Tier 1: Individual / FO filers', 'Irrevocable Larson Family Investment Trust', 'n/d', 'USA', 'Family-office trust filer (Snowball-listed).', 'Multiple microcap 13Ds'),
    ('Tier 1: Individual / FO filers', 'Hoak Public Equities', 'J.B. Hoak', 'USA', 'Family-office filer (Snowball-listed).', 'Multiple microcap 13Ds'),
    # === TIER 2: Small concentrated funds w/ active 13D ===
    ('Tier 2: Small concentrated activists', 'Hestia Capital Management', 'Kurt Wolf', 'Adams Twp PA', 'CIK 0001540979 / 0001456565. ~$200-300M AUM.', 'First activist at GameStop 2019; Pitney Bowes (PBI) won 4 board seats 2023, +80% 2024, Wolf named CEO May 2025; Edgewater (EDGW) sold to Alithya'),
    ('Tier 2: Small concentrated activists', 'Maran Capital Management / Maran Partners Fund', 'Dan Roller (ex-Scopus/Avesta/Impala)', 'Denver CO', 'CIK 0001722429. ~$50-100M. 2021 +50% net; 5yr ~19% net CAGR.', 'Pure Cycle (PCYO) — board seat Jan 14 2026 @ 14.7%; Lead Independent Dir at Horizon Kinetics (HKHC); Scott\'s Liquid Gold proxy win 2021'),
    ('Tier 2: Small concentrated activists', 'Findell Capital Management', 'Brian Finn', 'NYC', 'CIK 0001814465. ~$75M (2023). +141% net Jul 2019–Apr 2023 vs R2000 +21%.', 'Oportun (OPRT) activist letter — board changes won; Cutera (CUTR). HFJ Tomorrow\'s Titans 2023'),
    ('Tier 2: Small concentrated activists', 'Funicular Funds / Cable Car Capital', 'Jacob Ma-Weaver', 'San Francisco', 'CIK 0001737186. 17.1% net annualized since Nov 2013.', 'Dune Acquisition; Synlogic; Rice Acquisition; ARCA biopharma; AtlasClear; Wheeler REIT preferred'),
    ('Tier 2: Small concentrated activists', 'Outerbridge Capital Management', 'Rory Wallace', 'NYC', 'Olshan client.', 'Allot Ltd (ALLT) multiple 13D/A; Barnes & Noble Education (BNED) 9.9%'),
    ('Tier 2: Small concentrated activists', 'Marlton Partners / ATG Fund II', 'James C. Elbaor', 'Chicago', 'CIK 0001613672.', '180 Degree Capital (TURN) Sep 2024 proxy — 4 nominees (Elbaor, Gliksberg, Morris, Greenberg). Also files on UK microcaps'),
    ('Tier 2: Small concentrated activists', 'Veradace Partners', 'Alexander Vezendan, John Conlin', 'Dallas TX', 'CIK 0001772351.', 'Repay (RPAY) 8.6% Class A — nominated Vezendan + Jacobs for 2026; Scientific Industries (SCND) 13.56%; opposed CSU/KUBRA acquisition Apr 2026'),
    ('Tier 2: Small concentrated activists', 'Kanen Wealth Management / Philotimo Fund', 'David Kanen', 'Coconut Creek FL', 'CIK 0001681614 / 0001688522. ~$337M ADV.', 'The RealReal (REAL) 13D Oct 2024 at 5.4M sh'),
    ('Tier 2: Small concentrated activists', 'Red Oak Partners', 'David Sandberg', 'Boca Raton FL', 'CIK 0001275914. Founded 2003.', 'GEE Group (JOB) 8.7%; Global Crossing Airlines (JETMF) 20.3% Oct 2024 13D/A; Educational Development historical'),
    ('Tier 2: Small concentrated activists', 'Pleasant Lake Partners / Fund 1 Investments', 'Jonathan Lennon', 'USA', 'CIK 0001959730.', 'BJ\'s Restaurants (BJRI) 13D Jan 2024 @ 9.99% pushing for sale'),
    ('Tier 2: Small concentrated activists', 'Steamboat Capital Partners', 'Parsa Kiai', 'NYC', 'CIK 0001631791. ~$371M.', 'CompoSecure (CMPO); International Money Express (IMXI)'),
    ('Tier 2: Small concentrated activists', 'Whetstone Capital Advisors', 'David Atterbury', 'Shawnee Mission KS', 'CIK 0001553700.', 'OptimizeRx (OPRX) 13D/A Mar 2025'),
    ('Tier 2: Small concentrated activists', 'Settian Capital / PAVP Family Office Fund', 'Karl + Kristofer Segerberg', 'NYC', 'Family office fund.', 'Westwood Holdings (WHG) 13D Aug 2025 @ 5.0% — broad activist agenda (M&A, AI cost cuts)'),
    ('Tier 2: Small concentrated activists', 'Forager Capital Management', 'n/d', 'USA', 'CIK 0001802986. 11 13D, 20 13G, 25 Form 4 filings.', 'Distinct from Australia\'s Steve Johnson Forager Funds'),
    ('Tier 2: Small concentrated activists', 'Northern Right Capital Management', 'Matthew Drapkin', 'Palm Beach Gardens FL', 'CIK 0001346543. Post-2017 successor to Becker Drapkin. ~$380M.', 'Great Elm Capital Corp (GECC) 7.6% Dec 2024'),
    ('Tier 2: Small concentrated activists', 'Ananym Capital Management', 'Charlie Penner (ex-JANA/Engine No. 1)', 'USA', 'Founded 2024. Emerging small/mid-cap activist.', 'n/d'),
    ('Tier 2: Small concentrated activists', '325 Capital LLC', 'Bob Lynch, John Marchisi (ex-Discovery Group)', 'NYC', 'Concentrated microcap.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Doma Perpetual Capital', 'Pedro Escudero', 'USA', 'Snowball-listed activist 13D/G filer.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Lynx1 Capital', 'Weston Nichols', 'USA', 'Healthcare-focused. Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Heard Capital', 'William Heard', 'Chicago', 'Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Whitefort Capital Management', 'Salanic / Kessman', 'NYC', 'Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Kent Lake PR', 'Benjamin Natter', 'USA', 'Real-estate focused. Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Sieve Capital', 'n/d', 'USA', 'Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Strategic Value Investors', 'n/d', 'USA', 'Bank-focused. Snowball-listed.', 'n/d'),
    ('Tier 2: Small concentrated activists', 'Snow Park Capital Partners', 'Jeffrey Pierce', 'USA', 'REIT-only activist. Sub-$200M post-2020.', 'FelCor Lodging +2,800 bps alpha (sold to RLJ $1.18B); Ashford Hospitality Prime +5,100 bps; Monogram Residential +2,100 bps; Front Yard Residential proxy'),
    ('Tier 2: Small concentrated activists', 'Driver Management Company', 'J. Abbott Cooper', 'USA', 'Pure-play community-bank activist. ~$50-100M.', 'First United $9.8M settlement 2021; Republic First; Codorus Valley; First of Long Island; AmeriServ (ASRV) 8.6%'),
    ('Tier 2: Small concentrated activists', 'Cruiser Capital Advisors', 'Keith Rosenbloom', 'USA', '~$90-100M. Industrial/financial sector focus.', 'American Vanguard (AVD) — won 3 board seats; SunOpta'),
    ('Tier 2: Small concentrated activists', 'FrontFour Capital Group', 'Stephen Loukas, David Lorber, Zachary George (ex-Pirate)', 'USA', '~$200-290M. Hedgeweek 2011 Best Event-Driven.', 'Granite REIT — won 3 board seats; ILG (timeshare); ClubCorp strategic review; OM Group proxy'),
    ('Tier 2: Small concentrated activists', 'Discovery Group I', 'Daniel Donoghue, Michael Murphy', 'Chicago', 'CIK 0001372183. Founded 2002.', 'Invuity (sold to Stryker); Entellus Medical (Stryker 2017); Amber Road; ClickSoftware'),
    ('Tier 2: Small concentrated activists', 'Crescendo Partners', 'Eric Rosenfeld', 'USA', '~$200M. 20+ board seats in 11+ years.', 'Aeropostale, Topps, Charming Shoppes, Imvescor, StarTek, Destination Maternity. Mentored Ajdler (Engine Capital)'),
    ('Tier 2: Small concentrated activists', 'PW Partners', 'Patrick Walsh', 'USA', 'Restaurant/consumer activist.', 'BJ\'s Restaurants board 2014-2022; Del Taco board; Famous Dave\'s board'),
    ('Tier 2: Small concentrated activists', 'Sidus Investment Management', 'Michael Barone, Alfred Tobia', 'USA', '$52M ADV. IT L/S.', 'SEC IAPD #160406'),
    ('Tier 2: Small concentrated activists', '180 Degree Capital Corp (TURN)', 'Kevin Rendino (ex-BlackRock)', 'USA', 'Publicly traded closed-end activist. ~$55-80M NAV.', 'Portfolio +225% Q4 2016 to present vs Russell Microcap +35%. Comscore, Synchronoss, Synacor, Potbelly, Lantronix. Merging w/ Mount Logan Capital'),
    ('Tier 2: Small concentrated activists', 'Roumell Asset Management', 'James Roumell', 'USA', '$95M ADV. 55 historical 13D + 71 13G.', 'Olo, Opera, Liberty Energy'),
    ('Tier 2: Small concentrated activists', 'North Run Capital LP', 'n/d', 'Boston', '11 13D + 99 13G + 76 13F filings since 2007.', 'PowerFleet (PWFL), Sequans (SQNS), LENSAR (LNSR), Airgain — all sub-$500M'),
    ('Tier 2: Small concentrated activists', 'Marathon Partners Equity Management', 'Mario Cibelli (ex-Robotti/GAMCO)', 'NYC', '~$139M 13F. Founded 1997.', 'Discovered Netflix @ $10/sh in 2004; first Blockbuster DVD-volume study'),
    ('Tier 2: Small concentrated activists', 'Cove Street Capital', 'Jeffrey Bronchick', 'USA', '~$316-364M. Cove Street Small Cap Value Fund (CSCSX).', 'IAC; Ecovyst; ViaSat'),
    ('Tier 2: Small concentrated activists', 'Engine Capital', 'Arnaud Ajdler (ex-Crescendo)', 'USA', '$1.26B ADV / $495M 13F. Top 10 = 76.7%.', 'Harvard Bioscience (HBIO) 13D 2019 +56%; Lyft 2025'),
    ('Tier 2: Small concentrated activists', 'Praesidium Investment Management', 'Kevin Oram, Peter Uddo', 'USA', '~$678M. Top 10 = 95.07%.', 'PTC largest holding'),
    ('Tier 2: Small concentrated activists', 'Permian Investment Partners', 'Alex Duran, Scott Hendrickson (Brahman alumni)', 'USA', '~$1B 13F. ~11 positions. Top 10 = 96.15%.', 'Western Europe focus'),
    ('Tier 2: Small concentrated activists', 'Quogue Capital LLC', 'Wayne P. Rothbaum', 'USA', 'Family office since 2001. CAVEAT: 2008 SEC Rule 105 fine.', 'Co-founded Acerta Pharma → sold to AstraZeneca for $4B (acalabrutinib). Concentrated biotech'),
    # === TIER 3: Concentrated long-only managers from podcast/substack circuit ===
    ('Tier 3: Concentrated long-only / emerging mgrs', 'McIntyre Partnerships', 'Chris McIntyre', 'NYC', '~$75M. 17% gross/13% net since inception vs R2000V 7%.', 'Concentrated SMID value 5-10 names'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Askeladden Capital', 'Samir Patel', 'Dallas/Ft. Worth', 'Long-only microcap value ~20 positions. 186% cumulative 2016-19 vs S&P1000 +59%.', 'MIXT activist letter'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Kingdom Capital Advisors', 'David Bastian', 'Fairfax VA', '<=20 names, up to 25% single position. ~22% net since Jan 2022 vs R2000 ~5%.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', '1 Main Capital', 'Yaron Naymark', 'Westport CT', 'Concentrated long-biased value. 20.9% annualized net since Feb 2018.', 'Limbach (LMB) multibagger'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Choice Equities', 'Mitchell Scott', 'Raleigh NC', '10-15 longs + modest short. 16% IRR since Jan 2017.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Crossroads Capital', 'Ryan O\'Connor', 'Kansas City', 'L/S concentrated Buffett-partnership model. ~14.7% net 5yr.', 'Grandfather was original Buffett LP'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Praetorian Capital', 'Harris "Kuppy" Kupperman', 'USA', 'Inflection investing, thematic, commodities. ~$317M.', 'Triple-digit returns 2020 & 2021'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Andvari Associates', 'Douglas Ott', 'Atlanta', '5-10 quality compounders.', 'Serial-acquirer focus (CSU, Topicus, Lumine)'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Saber Capital', 'John Huber', 'Fuquay-Varina NC', '~$57M 13F. Buffett-partnership fees (no mgmt fee).', 'Base Hit Investing blog'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'North Peak Capital', 'Jeremy + Michael Kahan', 'NYC', 'Grew from <$5M to ~$1.7B. +45.8% in 2017.', '"PE in public markets" concentrated SMID'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Old West Investment Mgmt', 'Joe Boskovich Sr./Jr.', 'LA', 'Owner-operator value, uranium/gold/critical minerals.', 'Long-only SMA +45-51% net 2021; Opportunity Fund +66%'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Intelligent Fanatics Capital', 'Ian Cassel, Michael Liu', 'Lancaster PA', '6-10 microcaps US/CA/UK/AU. ~60 families.', 'Founder of MicroCapClub'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Blueprint Capital Mgmt', 'Neil Cataldi, Jason Revland', 'Philadelphia', '12-15 microcap tech/consumer. Launched 2012.', 'Ex-options market-maker'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Rhizome Partners', 'Bill Chen', 'NYC', 'Real-estate value, REITs, hard assets.', 'ALX deep dive; Latticework 2024 multifamily REIT pitch'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Right Tail Capital', 'Jeremy Kokemor', 'Richmond VA', 'Long-only quality compounders.', 'Ex-T. Rowe/TSW; CDW, NSIT'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Curreen Capital', 'Christian Ryther', 'Brooklyn', 'Global small-cap value, spinoffs. $50K min.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'LVS Advisory', 'Luis Sanchez', 'NY/FL', 'Event-driven defensive + concentrated growth. Growth Portfolio +81% vs SPX 41% Jan 2020-22.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Merion Road Capital', 'Aaron Sallen', 'Miami', 'Small-cap special situations / busted-deal arb.', 'Monarch Cement, Bel Fuse, Rocky Brands'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Poplar Point Capital', 'Jad Fakhry', 'Burlingame CA', 'Microcap + liquidation/merger arb, 3-5yr.', 'Monarch Cement (MCEM), Armanino Foods (AMNF), Gyrodyne 13D'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Monimus Capital', 'Brian Bellinger', 'NJ', 'L/S SMID + microcap. +31% net inception-Sep 2021.', 'Spun out of Raging Capital'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'TMR Capital', 'Ted Rosenthal', 'NYC', 'L/S SMID, 20%+ IRR target. 18%+ net since Oct 2019, no down years.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Blue Outlier Capital', 'Ryan Rahinsky', 'Tampa', 'Value + LEAP options on special sits.', 'No mgmt fee; perf-only'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Rogue Funds', 'Jacob Rowe', 'Raleigh NC', 'Distressed/special sits. ~75%+ since May 2023 launch.', 'SEZL 7x in 6 months'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Lake Cornelia', 'Judd Arnold', 'Minnesota', 'Inflection investing, oncology. 211.4% net 3/1/2020-11/30/2020.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Half Moon Capital', 'Eric DeLamarter', 'Larchmont NY', 'Deep-value L/S small/mid + special sits.', 'Tile Shop, BlueLinx; Columbia G&D'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Rangeley Capital', 'Chris DeMuth Jr.', 'USA', 'Event-driven, SPACs, M&A/antitrust arb.', 'Yet Another Value Blog\'s Andrew Walker is PM'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Far View Capital', 'Brad Hathaway', 'USA', 'Global small/mid special situations.', 'Radisson Hospitality squeeze-out; DSGR; Naked Wines'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Worm Capital', 'Arne Alsin', 'La Jolla CA', 'Concentrated tech/innovation L/S. ~$108-225M.', '~70% Tesla / ~25% Spotify'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Plough Penny Partners', 'Judson B. Traphagen', 'NYC', 'Long-biased growth, 10-20 longs. +51% gross 2024.', 'Traphagen family office'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Greenwood Investors', 'Steven D. Wood', 'Greensboro NC', '$116.5M (Mar 2025). Concentrated global "constructivist".', 'Builders Fund I'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'SVN Capital', 'Shreekkanth Viswanathan', 'Chicago', 'Global value, 10-15 names, perf-only fees.', 'HEICO during COVID'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Saguaro Capital', 'James Falbe', 'USA', 'AI + value intersection.', 'Compounders Podcast guest'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Recurve Capital', 'Aaron Chan', 'USA', 'Long-biased TMT/consumer.', '47.7% Carvana position; ex-$1.2B family office'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Adestella Investment Mgmt', 'Andrew Jakubowski', 'Grand Rapids MI', 'L/S behavioral + fundamental. +93.3% net 2020.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Carlson Ridge Capital', 'Denver Smith, Sam Jurrens', 'Denver', 'Activist concentrated.', 'XPEL ~$1.50→$60 (2018-21); 22% annualized since Oct 2015'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Roubaix Capital', 'Chris Hillary', 'Denver', 'SMID L/S fundamental. ~$165-200M; <50% net exposure.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Beech Capital', 'Melvin Beech', 'UK', '5-7yr horizon, no macro. 17%+ p.a. on longest strategies.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'City Different Investments', 'Connor Browne, Rob MacDonald', 'Santa Fe NM', 'Domestic equities. Ex-Thornburg duo.', 'Launched 2021'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Tourlite Capital', 'Jeffrey G. Cherkin', 'NYC', 'Public equities. Founded 2022; sub-$300M.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Pernas Research/Capital', 'Deiya Pernas', 'USA', 'Microcap fundamental research. Planet MicroCap regular.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Hudson 215 Capital', 'Jason Hirschman', 'USA', 'Microcap value.', '@EightTrack180'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Circle City Capital', 'Colin King (@valuedontlie)', 'Indianapolis', 'HoldCo + activist longs.', 'TEVA, TAP, HBI deleveraging plays'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Night Watch Investment Mgmt', 'Roderick van Zuylen', 'USA', 'Special situations.', 'Marex thesis'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Avory & Co.', 'Sean Emory', 'USA', 'Tech-tilted growth.', 'Clear Secure thesis'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Summers Value Partners', 'Andy Summers', 'USA', 'Long-only equity.', 'Theravance (TBPH)'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Olesen Value Fund', 'Christian Olesen', 'USA', 'International deep value.', 'UK homebuilders'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Lowell Capital', 'Jim + Abby Zimmerman', 'USA', '"Fort Knox balance sheet" microcaps.', 'Boring niche leaders'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'T&T Capital', 'Tim Travis', 'USA', 'Deep-value REITs/BDCs/microcaps + options.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Carriage House Fund', 'Will Cleary', 'USA', 'Concentrated.', 'FTAI Aviation deep work'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Korwell Capital', 'Kenny Chan', 'USA', '"Phil Fisher at Graham prices". Founder 23 yo, ex-NYU Stern.', 'AAP, TRBR.V'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Bossert Capital', 'Alex Bossert', 'Minneapolis', 'Concentrated long-term.', 'Co/Investor Club Top 50'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'PFH Capital', 'Thomas Bachrach', 'USA', 'Off-radar international/EM.', 'High insider ownership'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Eagle Point Capital', 'Matt + Dan', 'USA', 'Microcap newsletter + concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Headwaters Capital', 'Christopher Godfrey', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Foster Point Capital', 'Tony Capeloto', 'USA', 'Microcap.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Sea Otter Advisors', 'Peter Smyth', 'USA', 'SPAC/special situations microcap.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Jenga Investment Partners', 'Dede Eyesan', 'USA', 'Global multibagger research.', '"Global Outperformers"'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Plural Investing', 'Chris Waller', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Colarion Partners', 'Sam Haskell', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Asheville Capital', 'Jake Barfield', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', '1035 Capital', 'Chris Abbott', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Ace River Capital', 'Nicholas D\'Agnillo', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Halvio Capital', 'Anonymous PM', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Oberon Asset Management', 'Kevin Tracey', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'DV Capital Partners', 'Jarratt Davis', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Jackson Peak Capital', 'Patrick O\'Brien', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Matrice Capital', 'Andrew Marasco', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'MJG Capital', 'Matt Geiger', 'USA', 'Mining/resources microcap.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Phoenician Capital', 'John Khabbaz', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'J&S Capital', 'John Davenport', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Silver Beech Capital', 'James Hollier', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'River Oaks Capital', 'Whit Huguley', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'MCJ Capital Partners', 'Carter Johnson', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Fairlight Capital', 'Andrew Martin', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Ballina Capital', 'Kevin Durkin', 'USA', 'Concentrated.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Dakota Value Funds', 'Scott Reardon', 'USA', 'Value.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'N.A.S. Capital', 'Nat Stewart', 'Portsmouth NH', 'Microcap value + reflexivity.', 'Winmark 100x example; "Stock Picking" Substack'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Boole Microcap Fund', 'Chris Olson', 'USA', 'Cigar-butt/sub-liquidation $10-25M cap.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Holland Advisors', 'Andrew Hollingworth', 'UK', '"Economies-scaled-shared" concentrated quality.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Eschler Asset Management', 'Theron de Ris', 'Europe', 'Family-run businesses value.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Howay Investments', 'David Collins', 'UK', 'High-quality entrepreneurial.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Sophon Capital Research', 'Franco Chomnalez', 'USA', 'Microcap research consultancy.', 'Sophon Microcap Atlas'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'GeoInvesting', 'Maj Soueidan', 'USA', 'Microcap multibagger discovery.', 'Substack "Microcap Investing Cliff Notes"'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Small Cap Discoveries', 'Paul Andreola', 'Canada', 'Canadian microcap growth.', 'Also CEO of NameSilo'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Crystal Rock Capital', 'Michael Melby pedigree', 'Chicago', 'Microcap.', 'Where Gate City\'s Melby trained'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'First Wilshire Securities', 'Scott Hood (Fred Astman founder)', 'USA', 'Microcap since 1977. ~20% annualized 3 decades.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Auer Growth Fund (AUERX)', 'Robert Auer', 'USA', '"Cheap growth" microcap mutual fund. ~$125-130M.', '~45% microcaps'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Acuitas Investments', 'Doug Porter', 'USA', 'Microcap-focused PM.', 'Planet MicroCap regular'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Tributary Capital Mgmt', 'Mark Wynegar, Donald Radtke', 'Omaha', 'Small Company Fund (FOSCX/FOSBX).', 'First National Bank subsidiary'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'O\'Keefe Stevens Advisory', 'n/d', 'Rochester NY', 'Value, $385M 13F.', 'NVDA, FNMA winners 2024'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Whetstone Capital / Three Arch / CUSH', 'n/d', 'Kansas City', 'Value.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Boyar Asset Management', 'Jon Boyar', 'NYC', '"Forgotten 40" small-cap value.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'HD Capital / Capital H', 'Harley Grosser', 'Australia', 'ASX small-cap.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Pier Capital', 'Alexander Yakirevich', 'Stamford CT', 'Small-cap. 17.03% 10-yr annualized.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Greenbrier Partners', 'n/d', 'Dallas TX', 'Concentrated. ~$842M 13F; 91.92% top 10.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Kayak Investment Partners', 'n/d', 'San Francisco', 'Concentrated. 75.36% 1-yr / 31.02% 3-yr 2020.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Cypress Funds', 'n/d', 'LA', '16 holdings, 71.76% top 10. 57.52% 1-yr 2020.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Oakmont Corp', 'n/d', 'LA', '27 holdings, 90.51% top 10.', 'Filings since 2001'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Ancient Art / Teton Capital', 'n/d', 'Austin', '~$650M 13F; 20 holdings; 86.84% top 10. 41.34% 1-yr 2020.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Newtyn Management', 'Eugene Dozortsev', 'NYC', 'Concentrated value.', '13F since 2012'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Kehlet Capital Management', 'Michael Wright', 'Austin TX', 'Concentrated microcap compounders. Ex-Bares Capital.', 'WRAP, SLP'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Merewether Capital', 'Luke Winchester', 'Newcastle AU', 'Aussie microcap Inception Fund. Hard close at AUD 50m.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Hurdle Rate Unit Trust', 'Tristan Waine', 'Australia', 'Professional/financial services. Wound down to SMSF.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Arquitos Capital / Willow Oak', 'Steven Kiel', 'USA', 'Concentrated; multi-strat platform.', 'Controls Enterprise Diversified'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Bares Capital Management', 'Brian Bares', 'Austin', 'Concentrated micro/small-cap SMA. ~$2-3B.', '"Smaller of Two Evils" author'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Value Creators Capital', 'Kevin Fogarty', 'Pennsylvania', 'Concentrated quality. Spun out of DuPont Capital 2023.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Eight Track Capital', 'EightTrack180', 'USA', 'Undiscovered high-ROIC.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Holdco Asset Management', 'Vik Ghei, Misha Zaitzeff', 'NYC', 'Broad mandate; community banks. >$2B regulatory AUM.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Uzo Capital', '"Jerome" (anonymous)', 'USA', 'Global public equities + music royalties.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Tao Value', 'Tao Wang (anonymous)', 'USA', 'VIC-style blog/SMA.', 'taovalue.net'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Compounder Fund', 'Ser Jing Chong, Jeremy Chia', 'Singapore', 'Concentrated global compounders.', ''),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Lamassu Holdings LLC', 'n/d', 'USA', 'Net-cash microcap activist.', 'Ditech Networks (DITC) demand-letter activist'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Nery Capital Partners', 'n/d', 'USA', 'Net-cash/liquidation activist.', 'InFocus (INFS) 9.8-12.2% strategic alts'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Maiden Financial', 'Gwen Hofmeyr', 'USA', 'Value/microcap research.', 'Planet MicroCap guest'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Rosen Partnership', 'Jason Kirsch', 'USA', '"Active Value Strategy".', 'Planet MicroCap'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Fin Capital Management', 'Ben Finser', 'USA', 'International microcaps.', 'Planet MicroCap'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Bison Interests', 'Josh Young', 'USA', 'Oil & gas concentrated long.', 'Planet MicroCap'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Ridgewood Investments', 'Sam Namiri', 'USA', 'Microcap "dark stocks" tax-loss.', 'Planet MicroCap'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Aganju Capital', 'Tolu Bukola', 'USA', 'Long thesis work.', 'DKNG/prediction markets'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Focus Capital Advisers', 'Mordechai', 'USA', 'International deep value.', 'Valeura Energy thesis'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Militia Capital', 'David Orr', 'USA', 'Off-beaten-path.', 'Former pro poker player'),
    ('Tier 3: Concentrated long-only / emerging mgrs', 'Uncommon Profits', 'Luke Wolgram', 'USA', 'Microcap fundamental + technical.', 'Early XPEL/LEAT; TipRanks #1 2021'),
    # === TIER 4: International activists with US 13D/G filings ===
    ('Tier 4: International activists', 'Effissimo Capital Management', 'Imai, Yamaguchi, Kosaka (ex-Murakami)', 'Singapore', '~$10B+. Files via ECM Feeder Fund 2 LP.', 'Toshiba ADR (TOSYY); Kawasaki Kisen ~$4B excess profit; Dai-Ichi Life ~$1B'),
    ('Tier 4: International activists', '3D Investment Partners', 'Kanya Hasegawa', 'Singapore', 'SEC CRD 282411.', 'Toshiba 7.2%, Toho HD ~22%, Sapporo HD active campaign'),
    ('Tier 4: International activists', 'Strategic Capital Inc.', 'Tsuyoshi Maruki', 'Tokyo', '~¥9.7B. Columbia endowment is LP.', 'Japan Securities Finance, Chori, Tosho Printing, Daidoh'),
    ('Tier 4: International activists', 'Misaki Capital', 'Yasunori Nakagami, Takeo Aso', 'Japan', '~¥100B "Japan Constructivism" ~10 positions.', ''),
    ('Tier 4: International activists', 'Symphony Financial Partners', 'David Baran, Kazuhiko Shibata', 'Singapore/Tokyo', '~$1B.', ''),
    ('Tier 4: International activists', 'Oasis Management', 'Seth Fischer', 'Hong Kong', 'CIK 1317904.', 'Nam Tai Property (NTP) SC 13D Apr 2022'),
    ('Tier 4: International activists', 'Asset Value Investors / AVI Japan Opportunity Trust', 'Joe Bauernfreund', 'London', 'SEC-registered IA.', 'Deep-value asset-backed Japanese small caps'),
    ('Tier 4: International activists', 'Nippon Active Value Fund / Rising Sun Mgmt', 'James Rosenwald III (Dalton affiliate)', 'London', 'Co-files w/ Dalton.', 'Nissan Tokyo Sales 5.02%; Fuji Media; Toyo Suisan; Aska Pharma'),
    ('Tier 4: International activists', 'Cevian Capital', 'Gardell, Förberg', 'Sweden/Jersey', '~$16B.', 'Autoliv (ALV) 6.9% via swaps; Ericsson (ERIC) 4.94%; Pearson (PSO) 18.14% (~$1.03B)'),
    ('Tier 4: International activists', 'Petrus Advisers Ltd', 'Klaus Umek, Till Hufnagel', 'London/Bratislava', 'SEC CRD 324991; CIK 1711611. 10+ 13D filings.', 'Distinct from Petrus Capital (Perot FO)'),
    ('Tier 4: International activists', 'Teleios Capital Partners', 'Kuzniar, Epstein', 'Zug, Switzerland', 'CIK 1690451; $1.15B.', 'SodaStream 4.5%; Quanex Building Products 13D/A 2025'),
    ('Tier 4: International activists', 'Bluebell Capital Partners', 'Bivona, Taricco, Trapani', 'London/Italy', '~$200M. "One share, one campaign".', 'BlackRock proxy materials AGM May 2024'),
    ('Tier 4: International activists', 'Active Ownership Capital', 'Schuhbauer, Röhrig, Schempp', 'Luxembourg', 'German-speaking countries focus.', 'Gerresheimer 5.3%+ Aug 2025'),
    ('Tier 4: International activists', 'Toscafund', 'Martin Hughes', 'London', 'Historical 13D/G filings.', 'Phoenix Cos; WaMu 105.5M sh 2008; Akazoo ADR'),
    ('Tier 4: International activists', 'Naya Capital', 'Masroor Siddiqui', 'London', '~$1.2B 13F.', ''),
    ('Tier 4: International activists', 'Pelham Capital', 'Ross Turner (ex-Lansdowne)', 'London', '~$166M 13F. Small/mid cap L/S.', ''),
    ('Tier 4: International activists', 'Polygon Global Partners', 'Reade Griffith, Patrick Dear', 'London', 'CIK 1511306.', 'Nightstar Therapeutics (NITE) 13D Apr 2019'),
    ('Tier 4: International activists', 'Amber Capital', 'Joseph Oughourlian', 'France/UK', 'Historical 13G on Pride International.', ''),
    ('Tier 4: International activists', 'Sterling Strategic Value', 'n/d', 'Luxembourg SICAV-RAIF', '13F/13D/G filings since 2021.', ''),
    ('Tier 4: International activists', 'K2 & Associates', 'Shawn Kimel', 'Toronto', 'Multi-strat w/ US 13D/G.', ''),
    ('Tier 4: International activists', 'West Face Capital', 'Gregory Boland', 'Toronto', 'Confirmed US 13D filer.', 'Gran Tierra Energy (GTE) 9.8% Apr 2015; SunOpta (STKL) 6.1%; Dune Energy 15.4%; ACE Aviation'),
    ('Tier 4: International activists', 'Ewing Morris', 'John Ewing, Darcy Morris', 'Toronto', '~$290M.', 'Cedar Realty Trust group 13D Feb 2021'),
    ('Tier 4: International activists', 'Goodwood Inc.', 'Peter Puccetti', 'Toronto', 'CIK 1297339.', 'COSCIENS Biopharma 13D 2024; Westaim Corp historical'),
    ('Tier 4: International activists', 'Smoothwater Capital', 'Garfield Mitchell, Stephen Griggs', 'Toronto', 'Primarily TSX.', 'Genesis Land 22%'),
    ('Tier 4: International activists', 'Penderfund Capital', 'n/d', 'Vancouver', '34+ 13F, 2 13D, 3 13G filings.', ''),
    ('Tier 4: International activists', 'Burgundy Asset Management', 'n/d', 'Toronto', 'CIK 1315868. ~$9.65B.', ''),
    ('Tier 4: International activists', 'Donville Kent Asset Management', 'Jason Donville, Jesse Gamble', 'Toronto', '~CAD$76M ROE-focused growth.', ''),
    ('Tier 4: International activists', 'White Falcon Capital', 'Balkar Sivia', 'Ontario', 'Ex-Burgundy/McElvaine.', ''),
    ('Tier 4: International activists', 'Caldwell Investment Management', 'Brendan + Thomas Caldwell', 'Toronto', '', ''),
    ('Tier 4: International activists', 'Lightwater Partners', 'n/d', 'Toronto', '', ''),
    ('Tier 4: International activists', 'JC Clark Limited', 'n/d', 'Toronto', '', ''),
    ('Tier 4: International activists', 'Sandon Capital', 'Gabriel Radzyminski', 'Sydney', 'ASX-listed LIC SNC. 50 campaigns since 2008.', 'Southern Cross Austereo 5.05%; up to 15% intl allowed'),
    ('Tier 4: International activists', 'VGI Partners / Regal Partners', 'n/d', 'Sydney', 'CIK 1577774. $1.27B portfolio, 46 US 13F holdings.', ''),
    ('Tier 4: International activists', 'L1 Capital Global Opportunities', 'n/d', 'Melbourne/Cayman', '', 'Sunshine Biopharma 13D/G; SEALSQ Corp 13G/A Feb 2025'),
    ('Tier 4: International activists', 'Senjin Capital', 'James Halse (ex-Platinum)', 'Australia', '', ''),
    ('Tier 4: International activists', 'DMX Asset Management', 'Steven McCarthy', 'Australia', '', ''),
    ('Tier 4: International activists', 'Ganes Focused Value Fund', 'Wayne Jones', 'Australia', '', ''),
    ('Tier 4: International activists', 'Value Base Ltd.', 'n/d', 'Israel', 'The 13D Report\'s non-13F activist profiles.', ''),
    ('Tier 4: International activists', 'City of London Investment Management', 'n/d', 'UK', 'Closed-end fund CEF discount activist.', 'First Israel Fund, Aberdeen Israel Fund, multi-country'),
    ('Tier 4: International activists', 'Vltava Fund', 'Daniel Gladiš', 'Czech Republic', 'Frequent MOI Global speaker.', ''),
    ('Tier 4: International activists', 'Palm Harbour Capital', 'Peter Smith', 'UK', '', ''),
    ('Tier 4: International activists', 'Quercus Fund', 'Diego B. Milano', 'USA', '', ''),
    ('Tier 4: International activists', 'Edelweiss Capital Research / Anker Capital', 'Javier Pérez Álvarez', 'USA', '', ''),
    ('Tier 4: International activists', 'Kathmandu Partners', 'Vincent Lo', 'USA', '', ''),
]


# Build sheet
if 'User-Suggested Watchlist' in wb.sheetnames:
    del wb['User-Suggested Watchlist']

ws = wb.create_sheet('User-Suggested Watchlist', 5)
ws.append(['Tier', 'Fund / Filer', 'Key Person(s)', 'Location', 'Style / AUM / Notes',
           'Notable Holdings / Campaigns', 'Already in Workbook?'])

# Style header
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E78')
thin = Side(border_style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical='top')
for col in range(1, 8):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border

for tier, fund, person, loc, style, notable in ROWS:
    in_wb = in_workbook(fund)
    ws.append([tier, fund, person, loc, style, notable, in_wb])

# Border and wrap
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = wrap
    # Highlight rows that are NEW (not in workbook) in pale yellow
    last_cell = row[-1]
    if last_cell.value == 'No':
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='FFF2CC')

ws.freeze_panes = 'A2'
widths = [32, 44, 32, 18, 50, 80, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save('/home/user/cyclepapa/fund_activity_last_6mo.xlsx')
already_count = sum(1 for r in ROWS if in_workbook(r[1]) == 'Yes')
print(f"Wrote 'User-Suggested Watchlist' sheet with {len(ROWS)} rows.")
print(f"Already in workbook: {already_count}")
print(f"New to research: {len(ROWS) - already_count}")
