"""Build 'Watchlist v2 — Warrants + Hidden Superinvestors' sheet."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WB_PATH = '/home/user/cyclepapa/fund_activity_last_6mo.xlsx'

# Each row: (fund_name, manager_or_key_person, category, sub_category, edge_or_track_record_note, filing_visibility, geography)
ROWS = [
    # ============ GAP 1A: VIC-to-fund + SumZero top analysts ============
    ('Alta Fox Capital', 'Connor Haley', 'VIC-to-Fund', 'Small/Micro L/S Activist', '~$400M; Collectors Univ, Hasbro activist campaigns', '13F + 13D', 'US'),
    ('ADW Capital', 'Adam Wyden', 'VIC-to-Fund', 'Concentrated US/Canada/Europe SMC', '~$405-500M; PAR Tech, AutoCanada; active VIC poster', '13F', 'US/Intl'),
    ('Gate City Capital', 'Michael Melby', 'VIC-to-Fund', 'Deep-value smallcap (CLOSED)', '$216M; ~91% top-10; 22.4% net annualized since 2011', 'Closed; ADV only', 'US'),
    ('Cobia Capital', 'Jeff Meyers', 'VIC-to-Fund', 'Smallcap tech value L/S', 'SumZero #1 all-time (26% median across 34 ideas); SILC 1,300%', 'Limited', 'US'),
    ('Park Lane Family Office', 'Henrik Heffermehl (Oslo)', 'VIC-to-Fund', 'Nordic smallcap', 'SumZero top across multiple cats; Norway family money', 'None (FO)', 'Nordic'),
    ('Argand Capital Advisers', 'John Rolfe', 'VIC-to-Fund', 'One-man shop', 'SumZero top-3 globally (50 ideas, 39.32% annualized avg)', 'Limited', 'US'),
    ('Caro-Kann Capital', 'Artem Fokin', 'VIC-to-Fund', 'Concentrated EM smallcap', '3-yr SumZero winner 2017-19; Burford Muddy Waters rebuttal; Kaspi.kz', '13F', 'EM/US'),
    ('Praetorian Capital', 'Harris "Kuppy" Kupperman', 'VIC-to-Fund', 'Commodity inflection + SPAC warrants', '$347M; VIC handle hkup881 since 2002; SHLL/NKLA warrant wins', '13F + ADV', 'US'),
    ('Bireme Capital', 'Tindell & Ballentine', 'VIC-to-Fund', 'Concentrated value', '+388% net since 2016', '13F + ADV', 'US'),
    ('Laughing Water Capital', 'Matt Sweeney', 'VIC-to-Fund', 'Concentrated 10-15', '29% annualized net since 2016 per letters', 'Limited', 'US'),
    ('1 Main Capital', 'Yaron Naymark', 'VIC-to-Fund', 'Concentrated quality growth', '+180% net since 2018', 'Limited', 'US'),
    ('Greenwood Investors', 'Steven Wood', 'VIC-to-Fund', 'Founder/owner-led + activist', '$116.5M; 2025 +52.5% vs MSCI ACWI +22.1%', '13F + ADV', 'US/Intl'),
    ('Arquitos Capital', 'Steven Kiel', 'VIC-to-Fund', 'NOL/special situations', '21.3% annualized since 2012', '13F + ADV', 'US'),
    ('Saga Partners', 'Joe Frankenfield', 'VIC-to-Fund', 'Concentrated', '34% annualized since 2017 per letters (Q2 2022 -56% drawdown)', 'Limited', 'US'),

    # Substack/FinTwit fund managers
    ('Yet Another Value Fund / Rangeley', 'Andrew Walker / Chris DeMuth', 'VIC-to-Fund', 'Special situations / event', 'YAVP podcast; concentrated event-driven', '13F + ADV', 'US'),
    ('Saber Capital', 'John Huber', 'Substack', 'Concentrated value', '$117M; long-form investor letters', '13F + ADV', 'US'),
    ('Woodlock House Family Capital', 'Chris Mayer', 'Substack', '100-baggers / quality compounders', '~$200M; "100 Baggers" book author', '13F + ADV', 'US'),
    ('Sullimar Capital', 'Bill Brewster', 'Substack', 'Concentrated value', 'Active podcast presence', 'Limited', 'US'),
    ('Focused Compounding', 'Kuhn / Gannon', 'Substack', 'Quality value', 'Podcast + Substack', 'Limited', 'US'),
    ('Kingdom Capital Advisors', 'Bastian / Edwards', 'Substack', 'Microcap value', 'Active VIC/Substack', 'Limited', 'US'),
    ('Greystone Capital', 'Adam Wilk', 'Substack', 'Microcap value', 'Idea Brunch guest', 'Limited', 'US'),
    ('Alluvial Capital', 'Dave Waters', 'Substack', 'Microcap deep value', '~$100M; +122% since 2017 vs +57% R-Microcap', '13F + ADV', 'US'),
    ('Askeladden Capital', 'Samir Patel', 'Substack', '<$300M caps', 'Microcap focus; idea brunch', 'Limited', 'US'),
    ('LVS Advisory', 'Luis Sanchez', 'Substack', 'Concentrated value', 'Idea Brunch', 'Limited', 'US'),
    ('Choice Equities Fund', 'Mitchell Scott', 'Substack', 'Concentrated smallcap', '16% IRR since 2017 vs 6% R2000', '13F + ADV', 'US'),
    ('Andvari Associates', 'Doug Ott', 'Substack', 'Quality', 'Podcast', 'Limited', 'US'),
    ('Blue Outlier Capital', 'Ryan Rahinsky', 'Substack', 'LEAPs-driven', 'Substack', 'Limited', 'US'),
    ('Spree Capital', 'Thatcher Martin', 'Substack', 'Concentrated', 'Idea Brunch', 'Limited', 'US'),
    ('Mile 26 Capital', 'Navi Hehar', 'Substack', 'Concentrated', 'Idea Brunch', 'Limited', 'US'),
    ('Tiburon Holdings', 'Peter Lupoff', 'Substack', 'Activist/event', 'Multiple 13Ds', '13F + 13D', 'US'),
    ('Elmrox Investment Group', 'Daniel Lawrence', 'Substack', 'Concentrated', 'Idea Brunch', 'Limited', 'US'),
    ('Massif Capital', 'Will Thomson', 'Substack', 'Real assets L/S', '14.5% annualized since 2016', '13F + ADV', 'US'),

    # Anonymous VIC handles
    ('Claar Advisors', 'Gary Claar (ex-JANA)', 'VIC-to-Fund', 'Concentrated value', 'VIC handle "gary9"; JANA co-founder', '13F + ADV', 'US'),
    ('Mangrove Partners', 'Nathaniel August', 'VIC-to-Fund', 'Activist event', 'VIC handle "nha855"', '13F + 13D', 'US'),
    ('Mittleman Investment Management', 'Chris Mittleman', 'VIC-to-Fund', 'Global value unconstrained', 'VIC handle "mimval"', '13F + ADV', 'US/Intl'),

    # ============ GAP 1B: Podcast-profiled concentrated managers ============
    ('CAS Investment Partners (Sosin)', 'Cliff Sosin', 'Podcast Profiled', 'Ultra-concentrated 4-8', '~$1.7-2B; ~35% compounded since 2012; ~50% Carvana; ILTB feature', '13F + ADV', 'US'),
    ('Edenbrook Capital', 'Jonathan Brolin', 'Podcast Profiled', 'PE-style public smallcap', '~$400M; Capital Allocators feature', '13F + ADV', 'US'),
    ('ShawSpring Partners', 'Dennis Hong', 'Podcast Profiled', 'Long-only concentrated 5-10', '~$1.5B; ex-Yale/Altimeter', '13F + ADV', 'US/Intl'),
    ('Hayden Capital', 'Fred Liu', 'Podcast Profiled', 'Emerging Asia compounders', '~$200M; SE Limited + Amazon = 2 ten-baggers', '13F + ADV', 'Asia/US'),
    ('Plural Investing', 'Chris Waller', 'Podcast Profiled', 'Concentrated 7-8 scuttlebutt', 'TerraVest multibagger; MOI/V:AH regular', 'Limited', 'US'),
    ('Rhizome Partners', 'Bill Chen', 'Podcast Profiled', 'Concentrated REIT/real estate', '23.7% net 2020; Griffin Industrial; ALX thesis', '13F + ADV', 'US'),
    ('Tsai Capital', 'Christopher Tsai', 'Podcast Profiled', 'Concentrated quality growth', '~$100M; running since 1997', '13F + ADV', 'US'),
    ('Recurve Capital', 'Aaron Chan', 'Podcast Profiled', 'TMT + consumer concentrated', 'Large Carvana position', '13F + ADV', 'US'),
    ('Bonhoeffer Fund', 'Keith Smith (Willow Oak)', 'Podcast Profiled', 'Global owner-operator', '28.1% annualized 15-yr pre-fund', '13F + ADV', 'Global'),
    ('Heller House', 'Marcelo Lima', 'Podcast Profiled', '"Day 1" software', 'Acquirers Podcast', '13F + ADV', 'US'),
    ('TCW New Americas Premier Equities', 'Joseph Shaposhnik', 'Podcast Profiled', 'Concentrated mid/large recurring', 'Heavy Constellation Software', '13F + ADV (mutual fund)', 'US/Intl'),
    ('Lake Cornelia Research', 'Judd Arnold', 'Podcast Profiled', 'Inflection investing', 'TOI oncology rollup; YAVP', 'Limited', 'US'),
    ('Dorsey Asset Management', 'Pat Dorsey', 'Podcast Profiled', 'Wide-moat 9-15 names', '$1.1B', '13F + ADV', 'US'),
    ('Semper Augustus', 'Chris Bloomstran', 'Podcast Profiled', 'Quality value', '~$900M+; ILTB/V:AH/Talking Billions regular', '13F + ADV', 'US'),
    ('Olesen Value Fund', 'Christian Olesen', 'Podcast Profiled', 'UK homebuilders/global obscure', '~$50M; 19%+ annualized through 2014; multiple 13Ds', '13F + 13D', 'US/UK'),

    # MOI Best Ideas presenters
    ('Punch & Associates', 'Howard Punch', 'MOI Presenter', 'Smallcap value', '$1.7B', '13F + ADV', 'US'),
    ('Brennan Asset Management', 'Brennan', 'MOI Presenter', 'Concentrated', 'MOI presenter', '13F + ADV', 'US'),
    ('GDS Investments', 'Glenn Surowiec', 'MOI Presenter', 'Concentrated value', 'MOI presenter', '13F + ADV', 'US'),
    ('Stacey Muirhead Capital', 'Muirhead', 'MOI Presenter', 'Concentrated value (Canada)', 'MOI presenter', 'Limited', 'Canada'),
    ('Norwood Capital', 'Hoeveler', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Pledge Capital', 'Edward Chang', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Papyrus Capital', 'Sacheti', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('MPK Partners', 'Mike Kruger', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Clayton Partners', 'Alex Gates', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Bretton Fund', 'Stephen Dodson', 'MOI Presenter', 'Concentrated value', 'MOI presenter', '13F + ADV', 'US'),
    ('Aquitania Capital', 'Karlin', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Moerus Capital', 'Wadhwaney / Campagna', 'MOI Presenter', 'Global deep value', '~$200M; Lipper Global Small/Mid award', '13F + ADV', 'Global'),
    ('Robotti & Co (Schwartz)', 'Bob Robotti / Isaac Schwartz', 'MOI Presenter', 'Kazakhstan/industrial', 'Already in WB (Robotti)', 'IN WORKBOOK', 'US/EM'),
    ('Shareholder Value Mgmt AG', 'Frank Fischer', 'MOI Presenter', 'Germany concentrated', 'Frankfurt-listed', 'BaFin', 'Germany'),
    ('Santa Lucia AM', 'Florian Weidinger', 'MOI Presenter', 'Asia frontier', 'MOI/Asian frontier', 'Limited', 'Asia frontier'),
    ('BrightGate Capital', 'López Bernardo', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'Spain'),
    ('Pernas Research', 'Deiya Pernas', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),
    ('Carriage House Fund', 'Will Cleary', 'MOI Presenter', 'Concentrated (FTAI)', 'MOI presenter', 'Limited', 'US'),
    ('Auxier AM', 'Jeff Auxier', 'MOI Presenter', 'Value', 'Auxier Focus mutual fund', '13F + ADV', 'US'),
    ('Summers Value Partners', 'Andy Summers', 'MOI Presenter', 'Concentrated', 'MOI presenter', 'Limited', 'US'),

    # International / Asian podcast managers
    ('Holland Advisors', 'Andrew Hollingworth', 'Intl Podcast', 'UK quality value', 'Long-form letters', 'Limited', 'UK'),
    ('Parkway Capital', 'Dan Rupp', 'Intl Podcast', 'Asia smallcap', 'Idea Brunch', 'Limited', 'Asia'),
    ('Kold Investments', 'Simon Kold (Copenhagen)', 'Intl Podcast', 'Concentrated value', 'Idea Brunch', 'Limited', 'Nordic'),
    ('Blue Infinitas', 'Bogumil Baranowski', 'Intl Podcast', 'Concentrated', 'Podcasts', 'Limited', 'US/Intl'),
    ('River Oaks Capital', 'Whit Huguley', 'Intl Podcast', 'Microcap value', 'Idea Brunch', 'Limited', 'US'),
    ('Langdon Equity Partners', 'Greg Dean', 'Intl Podcast', 'Toronto smallcap', 'Talking Billions', '13F + ADV', 'Canada'),
    ('Albert Bridge Capital', 'Drew Dickson', 'Intl Podcast', 'European long-only', 'London-based', 'Limited', 'Europe'),
    ('Seawolf Capital', 'Daniel / Collins', 'Intl Podcast', 'Energy/cyclicals', 'Talking Billions', '13F + ADV', 'US'),
    ('Prevatt Capital', 'Jonathan Tepper', 'Intl Podcast', 'Concentrated value (Bahamas)', 'FT "Britain\'s answer to Buffett"', 'Limited', 'Bahamas'),
    ('Kopernik Global Investors', 'Dave Iben', 'Intl Podcast', 'Global contrarian deep value', '$6B; uranium/EM/commodities', 'IN WORKBOOK', 'Global'),

    # ============ GAP 1C: International / EM concentrated value ============
    ('azValor AM', 'Guzmán de Lázaro & Bernad', 'Spanish Value', 'Deep value commodities', '~€2.5-3B; 11.2% annualized since 2003 vs 8.7%', 'CNMV', 'Spain'),
    ('Horos AM', 'Javier Ruiz, Martín, Rodríguez', 'Spanish Value', 'Small/forgotten markets', '12.4% annualized at Metagestión; "few fishermen"', 'CNMV', 'Spain'),
    ('Magallanes Value Investors', 'Iván Martín', 'Spanish Value', 'European concentrated', '~€2.2B; 2023 Citywire Best Europe Eq Mgr', 'CNMV', 'Spain/EU'),
    ('Valentum AM', 'Luis de Blas / Jesús Domínguez', 'Spanish Value', 'European concentrated', '11.97%/yr since 2014 vs 5.77% MSCI Europe', 'CNMV', 'Spain'),
    ('Equam Capital', 'Muñoz / Larraz', 'Spanish Value', 'Global value UCITS', '~€55-80M; Lux UCITS', 'CNMV/CSSF', 'Spain'),

    ('RV Capital / Business Owner Fund', 'Robert Vinall (Switzerland)', 'European Value', '13 concentrated stocks', '~$380M; ~18.5% annualized since 2006', '13F + FINMA', 'Switzerland'),
    ('Findlay Park Partners', 'Kingsley / Findlay', 'European Value', 'US equities run from London', '£11B; +1,200pp vs R1000 since launch', 'UK FCA', 'UK'),
    ('Comgest', 'Team-based', 'European Value', 'Quality growth multi-region', '~$30B+', 'AMF + others', 'France'),
    ('Aubrey Capital Management', 'Andrew Dalrymple', 'European Value', 'EM consumption growth', '~$1B; 9% annualized since 2015 vs 4.7% MSCI EM', 'UK FCA', 'UK/EM'),
    ('Slater Investments', 'Mark Slater', 'European Value', 'UK smallcap PEG', 'FT top-10 most consistent UK manager', 'UK FCA', 'UK'),
    ('Tellworth (Premier Miton)', 'Marriage / Warren', 'European Value', 'UK smallcap', '+41.4% since 2018 vs +30.1% Numis SC + AIM', 'UK FCA', 'UK'),

    ('Turtle Creek Asset Management', 'Brenton, Cole, Hebel', 'Canadian Value', 'N. American midcap "synthetic PE"', '~$4B USD; ~20% annualized over 21 years', '13F + ADV', 'Canada'),
    ('Giverny Capital', 'François Rochon', 'Canadian Value', 'Quality compounders', '~$1.8B; ~15% annualized 30+ years', '13F + ADV', 'Canada'),
    ('EdgePoint', 'Bousada / MacDonald / Farmer', 'Canadian Value', 'Concentrated global', '~C$30B; ~16.4% Global Portfolio inception', 'Public funds', 'Canada'),
    ('Pender Fund Capital', 'David Barr / Felix Narhi', 'Canadian Value', 'Smallcap PE-style', '~C$2B+; Pender Small Cap Lipper winner', 'Public funds', 'Canada'),
    ('Mawer Investment Management', 'Viswanathan / Mo', 'Canadian Value', 'New Canada small cap', 'Calgary-based; multi-billion', 'Public funds', 'Canada'),
    ('Burgundy Asset Management', 'Tony Arrell', 'Canadian Value', 'US/Cdn smallcap + EM + Asia', 'Multi-billion; multiple sub-mandates', 'Limited', 'Canada'),
    ('Donville Kent', 'Jason Donville / Jesse Gamble', 'Canadian Value', 'High-ROE Cdn compounders', 'Newsletter + funds', 'Limited', 'Canada'),
    ('Lester Asset Management', 'Stephen Takacsy', 'Canadian Value', 'All-cap Canadian', 'C$360M', 'Limited', 'Canada'),

    ('SPARX Group', 'Shuhei Abe', 'Japan Value', 'Japan smallcap pioneer', 'Founded 1989; sub-advises Hennessy Japan Small Cap', 'Public', 'Japan'),
    ('Indus Capital Partners', 'Smith / Gill / Pinkel', 'Asia Pacific', 'Long/short Asia + Japan', '$3.5-6.2B; Capital Allocators feature', 'Limited', 'Asia'),
    ('Genesis IM', 'Christopher Ellyatt', 'EM Specialist', 'Institutional EM', 'London-based', 'UK FCA', 'EM'),
    ('East Capital', 'Jacob Grapengiesser (Stockholm)', 'EM/Frontier', 'EM + frontier; Balkans', 'Lipper award Balkans fund', 'Limited', 'EM'),
    ('Asia Frontier Capital', 'Thomas Hugger', 'Frontier', 'Vietnam/Iraq/Uzbekistan', 'HK-based', 'Limited', 'Frontier Asia'),
    ('Pangolin Asia Fund', 'James Hay (KL)', 'Asia Smallcap', 'Deep Asian smallcap value', '8.6% annualized', 'Limited', 'SE Asia'),
    ('African Lions Fund', 'Tim Staermose', 'Frontier', 'Sub-Saharan Africa', '40%+ since inception', 'Limited', 'Africa'),

    # ============ GAP 1D: Turnaround / Special Sits / PE-in-public-markets ============
    ('Solus Alternative AM', 'Chris Pucillo', 'Distressed Special Sits', 'Bristow Group top holding', '~$3.3B; ~28 13Ds', '13F + 13D', 'US'),
    ('Centerbridge Partners', 'Aronson / Gallogly', 'Distressed Special Sits', 'Extended Stay, OTG, PacWest', '~$38-56B', 'Limited', 'US'),
    ('Anchorage Capital Advisors', 'Baron / Gournay', 'Distressed Special Sits', 'MGM Studios, J.Crew', '~$26B; 2022 spin from Kevin Ulrich', 'Limited', 'US'),

    ('Mantle Ridge', 'Paul Hilal (ex-Pershing)', 'Activist', 'Concentrated activist', '~$4.6B; CSX, Aramark, Dollar Tree, Air Products 2025', '13F + 13D', 'US'),
    ('Impactive Capital', 'Wolfe / Asmar (ex-Blue Harbour)', 'Activist', '8-12 concentrated', '~$3B; Asbury, MAR Vacations, Concentrix, Clarivate, Etsy', '13F + 13D', 'US'),
    ('HG Vora', 'Parag Vora (ex-Silver Point/GS)', 'Activist', 'Concentrated event', 'Penn proxy 2024-25 (2 board); Ryder $4.4B bid 2024', '13F + 13D', 'US'),
    ('Coliseum Capital', 'Adam Gray & Chris Shackelton', 'Activist', 'PE-like concentrated', '~$2.5B; Sonos, GMS, MasterCraft, ModivCare, Blue Bird', '13F + 13D', 'US'),
    ('Sessa Capital', 'John Petry (ex-Gotham/Greenblatt)', 'Activist', 'Concentrated event', '~$4.2B; Ashford Hospitality Prime', '13F + 13D', 'US'),
    ('Cruiser Capital', 'Keith Rosenbloom', 'Activist', 'Industrials/financials governance', '~$91-107M', '13F + 13D', 'US'),
    ('Permian Investment Partners', 'Duran / Hendrickson / Swain', 'Activist', 'W. Europe concentrated', '~$2.2B; Aramark largest; HFM Europe HF of Year 2012', '13F + ADV', 'US/EU'),
    ('180 Degree Capital (TURN)', 'Kevin Rendino', 'Public BDC Activist', 'Microcap activist BDC', 'Public BDC; Synacor, TheStreet', 'NASDAQ-listed', 'US'),

    ('Boyar Value Group', 'Mark & Jonathan Boyar', 'Spin-off / Hidden Asset', '"Forgotten Forty" since 1975', '43% of researched co. historically taken out at premium', '13F + ADV', 'US'),
    ('Spin-Off Advisors', 'Joe Cornell', 'Spin-off Research', 'Spin-off research', 'Research signal, not fund', 'Research only', 'US'),
    ('Stock Spinoff Investing', 'Rich Howe', 'Spin-off Research', 'Newsletter', 'Subscribers include Gator, Plural, Curreen', 'Newsletter', 'US'),
    ('The Edge Consulting Group', 'Jim Osman', 'Spin-off Research', 'Institutional spin-off', 'Research signal', 'Research only', 'UK/US'),

    # ============ GAP 2A: SPAC WARRANT SPECIALISTS ============
    ('Periscope Capital', 'Jamie Wise / Stephen Church', 'WARRANT SPECIALIST', 'ONLY fund w/ explicit Warrant Arb sleeve', '~$454M 13F (peak $3.59B SPACs 2021); Unlimited Podcast', '13F + ADV (Toronto)', 'Canada'),
    ('Polar Asset Management', 'Paul Sabourin', 'WARRANT SPECIALIST', 'SPAC warrants itemized in 13F', '~$5.6B; historic 13Fs explicitly itemize SPAC warrants', '13F + ADV', 'Canada'),
    ('Glazer Capital', 'Paul Glazer / Mark Ort', 'WARRANT SPECIALIST', 'II called him "true King of SPACs"', '~$2B (peak $4.82B SPACs); now Ares subsidiary', '13F + ADV', 'US'),
    ('Radcliffe Capital', 'Steve Katznelson', 'WARRANT SPECIALIST', 'Historic itemized warrant 13Fs', '$3.04B SPACs 2021', '13F + ADV', 'US'),
    ('RiverNorth (SPCZ ETF)', 'Patrick Galley', 'WARRANT SPECIALIST', 'SPCZ explicit warrants/rights mandate', 'Enhanced Pre-Merger SPAC ETF', 'Public ETF', 'US'),
    ('Westchester Capital (MERFX)', 'Behren / Shannon', 'WARRANT SPECIALIST', 'Prospectus permits SPAC commons + warrants', 'Public mutual fund', '13F + Public', 'US'),
    ('Karpus Investment Management', 'City of London sub', 'WARRANT SPECIALIST', 'Pre-acquisition SPAC arb (ADV)', '$3.5B+', '13F + ADV', 'US'),
    ('Bulldog Investors', 'Goldstein/Dakos/Samuels/Das', 'WARRANT SPECIALIST', 'Coined "Bulldog provisions"', 'Runs SPE CEF', '13F + 13D', 'US'),

    # SPAC arb participants
    ('Aristeia Capital', '-', 'SPAC Arb', 'Major SPAC arb', '$3.78B SPACs 2021', '13F', 'US'),
    ('Linden Advisors', '-', 'SPAC Arb', 'Major SPAC arb', '~$10B', '13F + ADV', 'US'),
    ('HGC Investment Mgmt', 'Kallir / Lindros', 'SPAC Arb', 'Canadian market-neutral', 'Best Market Neutral Cdn award', 'ADV', 'Canada'),
    ('AQR Arbitrage', '-', 'SPAC Arb', 'AQR Diversified Arb + 1,400-SPAC database', 'Quant-driven', 'IN WORKBOOK (AQR)', 'US'),
    ('Shaolin Capital', '-', 'SPAC Arb', 'Major SPAC arb', '-', '13F', 'US'),
    ('Weiss Asset Mgmt', '-', 'SPAC Arb', 'Major SPAC arb', '-', '13F', 'US'),
    ('Castle Creek Arbitrage', '-', 'SPAC Arb', 'Major SPAC arb', '-', '13F', 'US'),
    ('Highbridge Capital', '-', 'SPAC Arb', 'Major SPAC arb', '-', '13F', 'US'),
    ('Yakira Capital', '-', 'SPAC Arb', 'Major SPAC arb', '-', '13F', 'US'),
    ('ATW Partners', '-', 'SPAC Arb', 'Microcap warrant 13Gs', '-', '13G', 'US'),

    # Individual voices
    ('Whitney Tilson', '-', 'SPAC Warrant Voice', 'Original T2 SPAC Fund (Oct 2008)', '~$5M doubled in <1yr', 'Newsletter', 'US'),
    ('Empire SPAC Investor', 'Enrique Abeyta', 'SPAC Warrant Voice', 'Newsletter editor', 'Newsletter', 'Newsletter', 'US'),
    ('Accelerate Arbitrage Fund (ARB.TO)', 'Julian Klymochko', 'SPAC Warrant Voice', '"Art of SPAC Arbitrage" framework', 'TSX-listed', 'Public', 'Canada'),
    ('Morgan Creek-Exos SPAC ETF', 'Mark Yusko', 'SPAC Warrant Voice', 'Public ETF', '-', 'Public', 'US'),

    # ============ GAP 2B: MINING + INTERNATIONAL WARRANTS ============
    ('Eric Sprott (personal — 2176423 Ont)', 'Eric Sprott', 'MINING WARRANT', 'Lead-orders nearly every junior PP', 'Hycroft $60M (33%), Chesapeake (17.9%), Goldgroup (12.5%+11.25M warr)', 'SEDAR EWR', 'Canada'),
    ('Sprott Inc. / Sprott US Holdings', 'Rick Rule / Whitney George', 'MINING WARRANT', 'RED PP LP dedicated junior PP+warrants', '~$31B AUM; public TSX/NYSE', 'Public + 13F', 'Canada/US'),
    ('Crescat Capital (CPM Fund)', 'Smith / Costa / Hennigh', 'MINING WARRANT', '"Friendly activist" PIPE', 'Goliath Resources 17.48% / 19.82% diluted', 'SEDAR EWR', 'US'),
    ('Equity Management Associates', 'Lawrence Lepard', 'MINING WARRANT', 'Gold-and-miners partnership since 2008', 'Amarillo Gold director', 'Limited', 'US'),
    ('Adrian Day Asset Management', 'Adrian Day', 'MINING WARRANT', 'Sub-adviser EuroPacific Gold Fund', '78% avg return all closed positions in newsletter', 'Newsletter', 'US'),
    ('U.S. Global Investors', 'Frank Holmes', 'MINING WARRANT', 'USERX (first no-load gold), UNWPX, PSPFX, GOAU', 'Mining Journal Mgr of Year 2006/2016', '13F + ETF', 'US'),
    ('Tocqueville Gold Fund', 'Hathaway (now Sprott) / Groh', 'MINING WARRANT', 'Hathaway moved to Sprott', 'Public fund', 'Public', 'US'),
    ('Goehring & Rozencwajg', 'Leigh Goehring / Adam Rozencwajg', 'MINING WARRANT', 'Ex-Chilton Global Nat Resources $5B+', 'New Global Resources Trust May 2025', '13F + ADV', 'US'),
    ('Casey Research / Intl Speculator', 'Casey / James / Forest / Katusa', 'MINING WARRANT', '"8 Ps" framework includes warrants', 'Newsletter', 'Newsletter', 'US'),
    ('Exploration Insights', 'Brent Cook / Joe Mazumdar', 'MINING WARRANT', 'Geologist-run; advises funds', 'Newsletter', 'Newsletter', 'US'),
    ('Palisades Goldcorp', 'Collin Kettell', 'MINING WARRANT', 'Modeled on Sprott\'s PP/warrant strategy', 'Runs Palisades Gold Radio podcast', 'TSXV', 'Canada'),
    ('ThreeD Capital', 'Sheldon Inwentash', 'MINING WARRANT', 'Successor to Pinetree (peak C$1B+/400 names)', 'TSXV-listed', 'TSXV', 'Canada'),
    ('Bob Moriarty (321gold.com)', 'Bob Moriarty', 'MINING WARRANT', 'Newsletter / commentary', '321gold.com', 'Newsletter', 'US'),

    # Reorg warrant recipients
    ('Avenue Capital Group', 'Marc Lasry / Sonia Gardner', 'REORG WARRANT', 'Reorg warrant recipient', 'Already in WB (Avenue)', 'IN WORKBOOK', 'US'),
    ('Marathon Asset Mgmt', '-', 'REORG WARRANT', 'Reorg warrant recipient', 'Already in WB (Marathon)', 'IN WORKBOOK', 'US'),
    ('Berkshire Hathaway (cornerstone)', 'Warren Buffett', 'REORG WARRANT', 'GS warrants 2008 (+$3.1B), BAC 2011 (+$12B)', 'Legendary case study; already in WB', 'IN WORKBOOK', 'US'),

    # Universa / convexity
    ('Universa Investments', 'Mark Spitznagel (Taleb-advised)', 'TAIL CONVEXITY', 'Tail-risk hedge / put options', '~$20B; 2020 returns +3,612% / +4,144%', 'Limited', 'US'),

    # International warrant markets
    ('Lion Selection Group (LSX.AX)', '-', 'INTL WARRANT', 'Junior miner LIC since 1997', '"Lion Mining Clock"', 'ASX', 'Australia'),
    ('Tribeca Global Natural Resources (TGF.AX)', '-', 'INTL WARRANT', 'Co-invested w/ Sprott in Hycroft', 'ASX-listed', 'ASX', 'Australia'),
    ('Regal Funds Mgmt / Regal Partners', 'Phil King', 'INTL WARRANT', '>$15B', 'ASX-listed (RPL)', 'ASX', 'Australia'),
    ('Taurus Funds Management', '-', 'INTL WARRANT', 'Junior miner financier', '-', 'Limited', 'Australia'),
    ('L1 Capital + L1 Gold Fund (LGF)', 'Lamm / Landau', 'INTL WARRANT', 'L1 Long Short (LSF) + new Gold Fund', 'ASX-listed', 'ASX', 'Australia'),

    ('The McHattie Group / Warrants Alert', 'Andrew McHattie', 'WARRANT DATA', 'Longest-running warrant pub (37+ years)', 'UK; since 1989', 'Newsletter', 'UK'),
    ('QuotedData', '-', 'WARRANT DATA', 'Zeros/Warrants/Subscription Shares monthly', '-', 'Research', 'UK'),

    # Warrant databases / primary research
    ('CommonStockWarrants.com', 'Dudley Pierce Baker', 'WARRANT DATA', 'Primary US+Canada warrant database', 'NDM.WT.B 10x; CIFRW +4184%; USARW +4997%; TMCWW +7429%', 'Database', 'US/Canada'),
    ('CanadianWarrants.com', '-', 'WARRANT DATA', 'Detailed Cdn warrant terms', '-', 'Database', 'Canada'),
    ('HKEX Derivative Warrant Database', '-', 'WARRANT DATA', 'One of largest globally', 'UBS, GS, Macquarie, HSBC, SG (28 yrs), Citi, JPM issuers', 'HKEX', 'HK'),
    ('SEDAR+', '-', 'WARRANT DATA', 'Canadian Early Warning Reports', 'Includes warrants in >10%', 'Government', 'Canada'),
    ('ADVFN London Covered Warrants', '-', 'WARRANT DATA', 'UK covered warrants page', '-', 'Database', 'UK'),

    # Rights / hidden-asset research
    ('PAA Research', 'Brad Safalow', 'HIDDEN ASSET RESEARCH', '~$6.8M revenue; ex-JPM lev fin', '-', 'Research', 'US'),
]

# Header
HEADER = ['Fund / Vehicle', 'Manager / Key Person', 'Category', 'Sub-category / Edge', 'Track Record / Edge Note', 'Filing Visibility', 'Geography', 'Already in Workbook?']

# Open workbook
wb = load_workbook(WB_PATH)

# Existing fund sheets (for in_workbook detection)
def in_workbook_check(fund_name):
    """Detect by fuzzy match against sheet names."""
    if not fund_name:
        return 'No'
    if 'IN WORKBOOK' in (fund_name or '') or 'Already in WB' in (fund_name or ''):
        return 'Yes'
    fl = fund_name.lower()
    # Strip common suffixes
    for skip in ['llc', 'lp', 'l.p.', 'inc', 'ltd', 'limited', 'capital', 'partners', 'asset', 'management', 'mgmt', 'advisers', 'advisors', 'group', 'fund', '&', 'co']:
        fl = fl.replace(skip, ' ')
    fl = ' '.join(fl.split())
    if len(fl) < 4:
        return 'No'
    for s in wb.sheetnames:
        if fl in s.lower():
            return 'Yes'
    # Also check first 8 chars exactly
    first_token = fund_name.split('(')[0].split('/')[0].strip()
    if first_token:
        ft = first_token.lower()
        for s in wb.sheetnames:
            if ft[:max(6, len(ft)//2)] in s.lower() and len(ft) >= 6:
                return 'Yes (partial match)'
    return 'No'

# Create sheet
SHEET_NAME = 'Watchlist v2 Warrants+Hidden'
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)

# Position it after User-Suggested Watchlist
target_idx = None
for i, s in enumerate(wb.sheetnames):
    if s == 'User-Suggested Watchlist':
        target_idx = i + 1
        break
if target_idx is not None:
    wb.move_sheet(SHEET_NAME, offset=target_idx - wb.sheetnames.index(SHEET_NAME))

# Styles
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
WARRANT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
IN_WB_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
NEW_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

# Write header
for col, h in enumerate(HEADER, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Write rows
new_count = 0
in_wb_count = 0
warrant_count = 0
for i, row in enumerate(ROWS, start=2):
    in_wb = in_workbook_check(row[0])
    full = list(row) + [in_wb]
    for col, v in enumerate(full, start=1):
        c = ws.cell(row=i, column=col, value=v)
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        c.font = Font(size=10)
    # Color
    if 'WARRANT' in (row[2] or '').upper() or 'SPAC' in (row[2] or '').upper():
        warrant_count += 1
        for col in range(1, len(HEADER)+1):
            ws.cell(row=i, column=col).fill = WARRANT_FILL
    if in_wb.startswith('Yes'):
        in_wb_count += 1
        ws.cell(row=i, column=8).fill = IN_WB_FILL
    else:
        new_count += 1
        ws.cell(row=i, column=8).fill = NEW_FILL

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 60
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 18

# Freeze top row + first column
ws.freeze_panes = 'B2'

# Filter
ws.auto_filter.ref = f'A1:H{len(ROWS)+1}'

wb.save(WB_PATH)
print(f"Wrote '{SHEET_NAME}' sheet with {len(ROWS)} rows")
print(f"  In workbook: {in_wb_count}")
print(f"  New: {new_count}")
print(f"  Warrant-related (yellow): {warrant_count}")
EOF