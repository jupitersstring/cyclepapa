"""Nano / Micro / Small-cap book.

Same Harvard aesthetic as build_harvard_workbook.py, but:
  - Universe filtered to financedatabase Nano/Micro/Small Cap buckets
    (mcap < ~$2B in the original FDB bucketisation).
  - Global Top 100 (vs 50) by entry_today_asymmetry.
  - 9 per-region Top 25 sheets so smaller regions aren't crowded out
    by US/UK/JP names in the global table.
  - Methodology page rewritten around the size-effect literature:
    Yartseva's small-EV result (37.7% annual excess at EV<$250M vs
    9.7% large-cap), Alta Fox's 84% of multibaggers ending below $2B,
    Mayer 100-Baggers, Phelps 100-to-1.

Output: asymmetry_nms_book.xlsx
"""
from __future__ import annotations
import argparse
import sys
from collections import OrderedDict

import pandas as pd

import build_harvard_workbook as bhw
from build_harvard_workbook import (
    Workbook, Color,
    INK, DARK_GREY, MUTED, RULE, LIGHT_GREY, PALE_GREY, WHITE,
    CRIMSON, CRIMSON_DARK, SERIF, SANS, MONO,
    EM_DASH, FMT_INT_RAW, FMT_ONE, FMT_TWO,
    _font, _fill, _border, _align,
    _set_col_widths, _crimson_banner, _section_rule, _verdict_badge,
    _write_money, _write_pct, _write_ratio, _write_score, _write_int,
    _NUM_ALIGN_RIGHT, _NUM_ALIGN_CENTER, _TXT_ALIGN_LEFT,
    _sheet_name_for, build_methodology, build_name_sheet,
    load_quant, load_verdicts, compute_scores,
)


NMS_BUCKETS = {'Nano Cap', 'Micro Cap', 'Small Cap'}

REGIONS = OrderedDict([
    ('North America',   ('NorthAmerica',   {'US', 'CA'})),
    ('Latin America',   ('LatinAmerica',   {'BR', 'MX', 'CL', 'AR'})),
    ('EU Core',         ('EU_Core',        {'UK', 'DE', 'FR', 'NL', 'BE', 'CH', 'IE', 'IT', 'AT'})),
    ('EU Nordics',      ('EU_Nordics',     {'SE', 'NO', 'DK', 'FI', 'IS'})),
    ('EU Periphery',    ('EU_Periphery',   {'ES', 'GR', 'PT'})),
    ('EU CEE / Baltics', ('EU_CEE',        {'CZ', 'HU', 'EE', 'LV', 'LT', 'PL', 'RO'})),
    ('Asia Developed',  ('Asia_Dev',       {'JP', 'KR', 'TW', 'HK', 'SG', 'AU', 'NZ'})),
    ('Asia Emerging',   ('Asia_Em',        {'IN', 'ID', 'TH', 'MY', 'CN'})),
    ('MEA',             ('MEA',            {'TR', 'ZA', 'IL', 'SA'})),
])


# --- Cover ----------------------------------------------------------------
def build_nms_cover(ws, n_universe: int, n_nms: int, top_n: int, n_regions: int,
                    n_green: int, n_yellow: int, n_unresearched: int):
    _set_col_widths(ws, {1: 4, 2: 28, 3: 28, 4: 28, 5: 28, 6: 4})

    # Heavy crimson top stripe
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = _fill(LIGHT_GREY)
    ws.row_dimensions[1].height = 8

    t = ws.cell(row=3, column=2, value="NANO · MICRO · SMALL")
    t.font = _font(size=36, bold=True, color=CRIMSON, name=SERIF)
    t.alignment = _align(h="left", v="center")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    ws.row_dimensions[3].height = 50

    s = ws.cell(row=4, column=2,
                value=("A Multibagger Field Guide for Sub-$2B Equities  ·  "
                       "Top 100 Global  +  Top 25 per Region"))
    s.font = _font(size=12, italic=True, color=bhw.INK, name=SERIF)
    s.alignment = _align(h="left", v="center")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    ws.row_dimensions[4].height = 22

    for c in range(2, 6):
        ws.cell(row=5, column=c).border = _border(color=CRIMSON, bottom="medium")
    ws.row_dimensions[5].height = 6

    d = ws.cell(row=7, column=2,
                value=("Compiled from the asymmetry framework (Yartseva-aligned upside, "
                       "Graham downside floor)  ·  Nano / Micro / Small-Cap buckets  ·  "
                       "As of 24 June 2026"))
    d.font = _font(size=10, color=MUTED, name=SANS, italic=True)
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=5)

    # Abstract
    _section_rule(ws, 11, "Abstract", span_cols=5)
    abstract = (
        "This book documents the top-100 nano/micro/small-cap candidates from a "
        f"{n_universe:,}-equity global universe, restricted to the {n_nms:,} names "
        "tagged by financedatabase as Nano Cap, Micro Cap or Small Cap (broadly, "
        "market caps below ~$2B). Yartseva (2025) finds the EV<$250M segment "
        "delivered 37.7% annualised excess returns over 2009–2024 vs 9.7% for "
        "large-caps — a ~28-point edge that survives controls for valuation, "
        "profitability and macro factors. Alta Fox (2020) corroborates: 84% of "
        "their 350%+ TSR 5-year sample ended below $2B. We rank within the size-"
        "filtered universe using the same entry-today asymmetry score that drives "
        "the master Harvard workbook, then carve out per-region Top 25 tabs so "
        "smaller geographies aren't crowded out by US/UK/JP names."
    )
    a = ws.cell(row=12, column=2, value=abstract)
    a.font = _font(size=11, name=SERIF)
    a.alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=12, start_column=2, end_row=17, end_column=5)
    for r in range(12, 18):
        ws.row_dimensions[r].height = 22

    # Tiles
    _section_rule(ws, 20, "Coverage", span_cols=5)
    tiles = [
        ("UNIVERSE",       f"{n_universe:,}", "names total"),
        ("NMS SUB-UNIVERSE", f"{n_nms:,}",    "after size filter"),
        ("GREEN",          f"{n_green}",      "high-conviction"),
        ("REGIONS",        f"{n_regions}",    "per-region tabs"),
    ]
    for i, (lbl, val, sub) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=22, column=col, value=lbl).font = _font(size=9, bold=True, color=MUTED, name=SANS)
        ws.cell(row=22, column=col).alignment = _align(h="center")
        ws.cell(row=23, column=col, value=val).font = _font(size=22, bold=True, color=CRIMSON, name=SERIF)
        ws.cell(row=23, column=col).alignment = _align(h="center")
        ws.cell(row=24, column=col, value=sub).font = _font(size=9, italic=True, color=MUTED, name=SERIF)
        ws.cell(row=24, column=col).alignment = _align(h="center")
        for r in (22, 23, 24):
            ws.cell(row=r, column=col).border = _border(
                color=RULE,
                top="thin" if r == 22 else None,
                bottom="thin" if r == 24 else None,
                left="thin", right="thin",
            )
    ws.row_dimensions[23].height = 32

    ws.cell(row=30, column=2,
            value="Prepared for internal allocation use. Not investment advice.").font = _font(
        size=9, italic=True, color=MUTED, name=SERIF)
    ws.merge_cells(start_row=30, start_column=2, end_row=30, end_column=5)

    for c in range(1, 7):
        ws.cell(row=33, column=c).fill = _fill(LIGHT_GREY)
    ws.row_dimensions[33].height = 6


# --- Methodology specific to NMS ------------------------------------------
def build_nms_methodology(ws):
    _set_col_widths(ws, {1: 4, 2: 24, 3: 60, 4: 4})
    _crimson_banner(ws, 1, "  Methodology — Nano/Micro/Small focus", span_cols=4)
    body = [
        ("Why size matters",
         "Yartseva (CAFE WP 33, 2025): firms with EV<$250M delivered 37.7% "
         "annualised excess returns 2009-2024, vs 9.7% large-cap. Effect "
         "survives controls for value, profitability and macro. Alta Fox "
         "(2020): 84% of 350%+ TSR 5-year set ended below $2B. Mayer "
         "(100 Baggers, 2015): high-ROIC small-caps with long reinvestment "
         "runways are the empirical archetype."),
        ("Universe filter",
         "financedatabase market_cap_bucket in {Nano Cap, Micro Cap, "
         "Small Cap}, additionally floored at mcap >= $10M to exclude "
         "untradeable shells. RED-verdicted names dropped."),
        ("Ranking",
         "Same entry-today asymmetry score as the master Harvard book: "
         "asymmetry_score x intrinsic_boost x qual_multiplier x "
         "post_rally_factor. Geometric-mean asymmetry (sqrt(upside x floor)) "
         "ensures both legs must fire."),
        ("Per-region breakouts",
         "9 region top-25 sheets (NA / LatAm / EU Core / EU Nordics / EU "
         "Periphery / EU CEE / Asia Developed / Asia Emerging / MEA). "
         "Surfaces local leaders that the global top-100 would otherwise "
         "miss in under-covered geographies."),
        ("Verdicts",
         "GREEN = plausible 3-5x on 3-5 year view; YELLOW = operating "
         "business with material risk; RED = excluded; UNRESEARCHED = "
         "ranking driven by quant only."),
        ("Per-name pages",
         "Top 100 global candidates each get a one-pager: verdict badge, "
         "snapshot, investment thesis, score decomposition, valuation & "
         "balance-sheet ratios, signal flags, notes."),
    ]
    row = 3
    for k, v in body:
        ws.cell(row=row, column=2, value=k).font = _font(size=11, bold=True, color=CRIMSON_DARK, name=SERIF)
        ws.cell(row=row, column=2).alignment = _align(v="top")
        ws.cell(row=row, column=3, value=v).font = _font(size=10, name=SERIF)
        ws.cell(row=row, column=3).alignment = _align(wrap=True, v="top")
        ws.row_dimensions[row].height = 60
        row += 1


# --- Index sheet (top-100 by entry_today_asymmetry) ----------------------
def build_nms_index(ws, top_df: pd.DataFrame):
    _set_col_widths(ws, {1: 4, 2: 5, 3: 14, 4: 38, 5: 8, 6: 18, 7: 12, 8: 14, 9: 10, 10: 14, 11: 4})
    _crimson_banner(ws, 1, "  Global Top 100 — entry-today asymmetry", span_cols=10)

    hdrs = ["#", "Ticker", "Company", "Cntry", "Sector", "Bucket", "Mcap (loc)", "Verdict", "ETA"]
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        if i in (4,):
            c.alignment = _align(h="left")
        elif i in (3, 6):
            c.alignment = _align(h="left")
        elif i in (5,):
            c.alignment = _align(h="center")
        elif i in (8,):
            c.alignment = _align(h="center")
        elif i >= 9:
            c.alignment = _align(h="right")
        else:
            c.alignment = _align(h="center")
        c.border = _border(color=CRIMSON_DARK, bottom="medium")
    ws.row_dimensions[3].height = 22

    mono = _font(size=10, name=MONO)
    serif = _font(size=10, name=SERIF)
    sans = _font(size=10, name=SANS)
    sans_muted = _font(size=10, name=SANS, color=MUTED)

    for i, (_, r) in enumerate(top_df.iterrows(), start=1):
        row = 3 + i
        c2 = ws.cell(row=row, column=2, value=i)
        c2.font = _font(size=10, color=MUTED, name=SERIF)
        c2.alignment = _NUM_ALIGN_CENTER

        sheet_name = _sheet_name_for(r['symbol'], i)
        link_cell = ws.cell(row=row, column=3, value=r['symbol'])
        link_cell.hyperlink = f"#'{sheet_name}'!A1"
        link_cell.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        link_cell.alignment = _TXT_ALIGN_LEFT

        name = (r.get('name') or '')[:60]
        nm_cell = ws.cell(row=row, column=4, value=name)
        nm_cell.font = serif
        nm_cell.alignment = _TXT_ALIGN_LEFT

        ws.cell(row=row, column=5, value=r.get('src', '')).font = sans
        ws.cell(row=row, column=5).alignment = _align(h="center")
        ws.cell(row=row, column=6, value=(r.get('sector') or '')).font = sans_muted
        ws.cell(row=row, column=6).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=row, column=7, value=(r.get('market_cap_bucket') or '')).font = sans_muted
        ws.cell(row=row, column=7).alignment = _align(h="center")

        _write_money(ws, row, 8, r.get('market_cap'), font=mono)
        _verdict_badge(ws, row, 9, r.get('verdict', 'UNRESEARCHED'))
        _write_score(ws, row, 10, r.get('entry_today_asymmetry'), font=mono)

        for c in range(2, 11):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 17


# --- Per-region top-25 sheet ----------------------------------------------
def build_region_sheet(ws, region_label: str, sub_df: pd.DataFrame):
    _set_col_widths(ws, {1: 4, 2: 5, 3: 14, 4: 38, 5: 8, 6: 18, 7: 12, 8: 14, 9: 10, 10: 10, 11: 4})
    _crimson_banner(ws, 1, f"  {region_label} — Top 25 NMS", span_cols=10)

    hdrs = ["#", "Ticker", "Company", "Cntry", "Sector", "Bucket", "Mcap (loc)", "Verdict", "ETA"]
    for i, h in enumerate(hdrs, start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        if i in (3, 4, 6):
            c.alignment = _align(h="left")
        elif i in (5, 7, 8):
            c.alignment = _align(h="center")
        else:
            c.alignment = _align(h="right")
        c.border = _border(color=CRIMSON_DARK, bottom="medium")
    ws.row_dimensions[3].height = 22

    mono = _font(size=10, name=MONO)
    serif = _font(size=10, name=SERIF)
    sans = _font(size=10, name=SANS)
    sans_muted = _font(size=10, name=SANS, color=MUTED)

    for i, (_, r) in enumerate(sub_df.iterrows(), start=1):
        row = 3 + i
        c2 = ws.cell(row=row, column=2, value=i)
        c2.font = _font(size=10, color=MUTED, name=SERIF)
        c2.alignment = _NUM_ALIGN_CENTER

        ws.cell(row=row, column=3, value=r['symbol']).font = _font(size=10, bold=True, name=SANS)
        ws.cell(row=row, column=3).alignment = _TXT_ALIGN_LEFT

        ws.cell(row=row, column=4, value=(r.get('name') or '')[:60]).font = serif
        ws.cell(row=row, column=4).alignment = _TXT_ALIGN_LEFT

        ws.cell(row=row, column=5, value=r.get('src', '')).font = sans
        ws.cell(row=row, column=5).alignment = _align(h="center")
        ws.cell(row=row, column=6, value=(r.get('sector') or '')).font = sans_muted
        ws.cell(row=row, column=6).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=row, column=7, value=(r.get('market_cap_bucket') or '')).font = sans_muted
        ws.cell(row=row, column=7).alignment = _align(h="center")

        _write_money(ws, row, 8, r.get('market_cap'), font=mono)
        _verdict_badge(ws, row, 9, r.get('verdict', 'UNRESEARCHED'))
        _write_score(ws, row, 10, r.get('entry_today_asymmetry'), font=mono)

        for c in range(2, 11):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 17

    # QoL: freeze the header + sortable/filterable table (header row 3)
    ws.freeze_panes = 'A4'
    if ws.max_row >= 4:
        from openpyxl.utils import get_column_letter as _gcl
        ws.auto_filter.ref = f"B3:{_gcl(ws.max_column)}{ws.max_row}"


# --- Main ------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--top-n', type=int, default=100, help='global top-N')
    ap.add_argument('--per-region-n', type=int, default=25, help='top-N per region')
    ap.add_argument('--out', default='asymmetry_nms_book.xlsx')
    ap.add_argument('--no-bucket-filter', action='store_true',
                    help='surface the entire universe (no NMS bucket filter)')
    ap.add_argument('--min-mcap', type=float, default=10_000_000)
    args = ap.parse_args()

    bhw._TOP_N_BOX['n'] = args.top_n

    print('loading...', file=sys.stderr)
    quant = load_quant()
    verdicts = load_verdicts()
    df = compute_scores(quant, verdicts)

    n_universe = len(df)

    # Universe gate: mcap floor + exclude RED. NMS bucket filter is the
    # default ("NMS book") but can be turned off via --no-bucket-filter
    # to surface the entire universe.
    if getattr(args, 'no_bucket_filter', False):
        nms = df[(df['market_cap'].fillna(0) >= args.min_mcap)
                 & (df['verdict'] != 'RED')].copy()
        print(f'  universe: {n_universe:,}  (no bucket filter)  '
              f'eligible: {len(nms):,}', file=sys.stderr)
    else:
        nms = df[df['market_cap_bucket'].isin(NMS_BUCKETS)
                 & (df['market_cap'].fillna(0) >= args.min_mcap)
                 & (df['verdict'] != 'RED')].copy()
        print(f'  universe: {n_universe:,}, NMS sub-universe: {len(nms):,}',
              file=sys.stderr)
    n_nms = len(nms)

    # Global Top N
    top = nms.sort_values('entry_today_asymmetry', ascending=False).head(args.top_n).copy().reset_index(drop=True)
    n_green = int((top['verdict'] == 'GREEN').sum())
    n_yellow = int((top['verdict'] == 'YELLOW').sum())
    n_unresearched = int((top['verdict'] == 'UNRESEARCHED').sum())
    print(f'  top-{args.top_n} verdict mix: GREEN={n_green}, YELLOW={n_yellow}, '
          f'UNRESEARCHED={n_unresearched}', file=sys.stderr)

    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    build_nms_cover(cover, n_universe, n_nms, args.top_n, len(REGIONS),
                    n_green, n_yellow, n_unresearched)

    method = wb.create_sheet("Methodology")
    build_nms_methodology(method)

    idx = wb.create_sheet("Global_Top_100")
    build_nms_index(idx, top)

    # Per-region top-N sheets
    for label, (slug, countries) in REGIONS.items():
        sub = nms[nms['src'].isin(countries)].sort_values(
            'entry_today_asymmetry', ascending=False
        ).head(args.per_region_n)
        if sub.empty:
            continue
        # Excel sheet name limit 31 chars
        sn = f"R_{slug}"[:31]
        ws = wb.create_sheet(sn)
        build_region_sheet(ws, label, sub)

    # Per-name sheets (top 100 only — same as Harvard build)
    for i, (_, row) in enumerate(top.iterrows(), start=1):
        sheet_name = _sheet_name_for(row['symbol'], i)
        ws = wb.create_sheet(sheet_name)
        build_name_sheet(ws, i, row)

    # Coverage + References (reuse master builder helpers)
    cov = wb.create_sheet("Coverage")
    bhw.build_coverage(cov, df, top)

    refs_sheet = wb.create_sheet("References")
    refs = [
        ("Yartseva, A. (2025)",
         "The Alchemy of Multibagger Stocks: An empirical investigation. "
         "CAFE Working Paper 33, Birmingham City University. "
         "https://www.open-access.bcu.ac.uk/16180/"),
        ("Alta Fox Capital (2020)",
         "Makings of a Multibagger. Summer Intern Class Project. "
         "https://www.altafoxcapital.com/s/Makings-of-a-MultiBagger.pdf"),
        ("Mayer, C. (2015)",
         "100 Baggers: Stocks That Return 100-to-1 and How to Find Them. "
         "Laissez Faire Books."),
        ("Phelps, T. W. (1972)",
         "100 to 1 in the Stock Market. McGraw-Hill, New York."),
        ("Greenblatt, J. (2006)",
         "The Little Book That Beats the Market. Wiley."),
        ("Dorsey, P. (2008)",
         "The Little Book That Builds Wealth: The Knockout Formula for "
         "Finding Great Investments. Wiley."),
        ("Cassel, I. (MicroCapClub)",
         "Top 10 Things in Micro Cap Investing. https://microcapclub.com/"),
        ("financedatabase (2026)",
         "Open securities database. github.com/JerBouma/FinanceDatabase."),
        ("yfinance (2026)",
         "Yahoo! Finance Python wrapper. github.com/ranaroussi/yfinance."),
    ]
    bhw.build_references(refs_sheet, refs)

    # Tab colors + hide gridlines
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        if ws.title.startswith("R_"):
            ws.sheet_properties.tabColor = Color(rgb=MUTED)
        elif ws.title in ("Cover", "Methodology", "Global_Top_100",
                          "Coverage", "References"):
            ws.sheet_properties.tabColor = Color(rgb=DARK_GREY)

    wb.save(args.out)
    print(f'wrote {args.out}: {len(wb.worksheets)} sheets', file=sys.stderr)


if __name__ == '__main__':
    main()
