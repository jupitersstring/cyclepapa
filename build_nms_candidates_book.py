"""NMS multibagger candidates book — Harvard-formatted XLSX.

Three tiers, each ranked by entry-today asymmetry:
  STRICT     — archetype_count >= 2 AND cluster_n >= 3 AND asymmetry >= 0.40
  STRONG     — archetype_count >= 1 AND cluster_n >= 3 AND asymmetry >= 0.35
  CANDIDATE  — archetype_count >= 1 OR (cluster_n >= 4 AND asymmetry >= 0.35)

Universe = NMS only (Nano + Micro + Small Cap buckets, mcap >= $10M, RED excluded).

Output: nms_multibagger_candidates.xlsx
"""
from __future__ import annotations
import os
import sys

import pandas as pd

import build_harvard_workbook as bhw
from build_harvard_workbook import (
    Workbook, Color,
    INK, DARK_GREY, MUTED, RULE, LIGHT_GREY, PALE_GREY, WHITE,
    CRIMSON, CRIMSON_DARK, SERIF, SANS, MONO,
    _font, _fill, _border, _align,
    _set_col_widths, _crimson_banner, _section_rule, _verdict_badge,
    _write_money, _write_pct, _write_ratio, _write_score, _write_int,
    _NUM_ALIGN_RIGHT, _NUM_ALIGN_CENTER, _TXT_ALIGN_LEFT,
)


def load_candidates() -> pd.DataFrame:
    return pd.read_csv('nms_multibagger_candidates.csv')


def build_cover(ws, df: pd.DataFrame):
    _set_col_widths(ws, {1: 4, 2: 28, 3: 28, 4: 28, 5: 28, 6: 4})

    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = _fill(LIGHT_GREY)
    ws.row_dimensions[1].height = 8

    t = ws.cell(row=3, column=2, value="MULTIBAGGER CANDIDATES")
    t.font = _font(size=32, bold=True, color=CRIMSON, name=SERIF)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
    ws.row_dimensions[3].height = 46

    s = ws.cell(row=4, column=2,
                value="Nano · Micro · Small-Cap Universe  ·  Archetype × Asymmetry tiers")
    s.font = _font(size=12, italic=True, color=bhw.INK, name=SERIF)
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    ws.row_dimensions[4].height = 22

    for c in range(2, 6):
        ws.cell(row=5, column=c).border = _border(color=CRIMSON, bottom="medium")
    ws.row_dimensions[5].height = 6

    n_strict = int((df.tier == 'STRICT').sum())
    n_strong = int((df.tier == 'STRONG').sum())
    n_cand = int((df.tier == 'CANDIDATE').sum())
    n_green = int((df.verdict == 'GREEN').sum())

    _section_rule(ws, 8, "Abstract", span_cols=5)
    abstract = (
        f"This shortlist filters the {n_strict + n_strong + n_cand:,} multibagger "
        "candidates in the NMS sub-universe (mcap < ~$2B, RED-verdicted names "
        "excluded). Each name is assigned a conviction tier based on its "
        "archetype-count, inflection-cluster density and asymmetry score. "
        f"STRICT ({n_strict}) = 2+ archetypes firing AND 3+ inflection signals "
        f"AND asymmetry >= 0.40. STRONG ({n_strong}) = 1+ archetype AND 3+ "
        f"inflection signals AND asymmetry >= 0.35. CANDIDATE ({n_cand}) = the "
        "broader funnel - 1+ archetype or 4+ inflection signals at asymmetry "
        ">= 0.35. Within each tier, names are ranked by entry-today asymmetry "
        "(asymmetry_score x qual_multiplier x post-rally factor)."
    )
    a = ws.cell(row=9, column=2, value=abstract)
    a.font = _font(size=11, name=SERIF)
    a.alignment = _align(wrap=True, v="top")
    ws.merge_cells(start_row=9, start_column=2, end_row=14, end_column=5)
    for r in range(9, 15):
        ws.row_dimensions[r].height = 22

    _section_rule(ws, 17, "Coverage", span_cols=5)
    tiles = [
        ("STRICT",    f"{n_strict:,}",  "high conviction"),
        ("STRONG",    f"{n_strong:,}",  "watchlist"),
        ("CANDIDATE", f"{n_cand:,}",    "broader funnel"),
        ("GREEN",     f"{n_green:,}",   "qualitative-aligned"),
    ]
    for i, (lbl, val, sub) in enumerate(tiles):
        col = 2 + i
        ws.cell(row=19, column=col, value=lbl).font = _font(size=9, bold=True, color=MUTED, name=SANS)
        ws.cell(row=19, column=col).alignment = _align(h="center")
        ws.cell(row=20, column=col, value=val).font = _font(size=22, bold=True, color=CRIMSON, name=SERIF)
        ws.cell(row=20, column=col).alignment = _align(h="center")
        ws.cell(row=21, column=col, value=sub).font = _font(size=9, italic=True, color=MUTED, name=SERIF)
        ws.cell(row=21, column=col).alignment = _align(h="center")
        for r in (19, 20, 21):
            ws.cell(row=r, column=col).border = _border(
                color=RULE,
                top="thin" if r == 19 else None,
                bottom="thin" if r == 21 else None,
                left="thin", right="thin",
            )
    ws.row_dimensions[20].height = 30

    for c in range(1, 7):
        ws.cell(row=27, column=c).fill = _fill(LIGHT_GREY)
    ws.row_dimensions[27].height = 6


def build_methodology(ws):
    _set_col_widths(ws, {1: 4, 2: 22, 3: 70, 4: 4})
    _crimson_banner(ws, 1, "  Tier methodology", span_cols=4)
    body = [
        ("STRICT",
         "archetype_count >= 2 AND cluster_n >= 3 AND asymmetry_score >= 0.40. "
         "Reads as: at least two distinct multibagger archetypes firing on the "
         "same name (e.g. DiscountedVehicle + CapitalDiscipline), at least three "
         "of seven inflection signals on, and a geometric-mean asymmetry above "
         "the 60th percentile. Highest conviction within the funnel."),
        ("STRONG",
         "archetype_count >= 1 AND cluster_n >= 3 AND asymmetry_score >= 0.35. "
         "One archetype + meaningful inflection density + above-median asymmetry. "
         "Names worth shortlisting for diligence; many UNRESEARCHED here."),
        ("CANDIDATE",
         "archetype_count >= 1 OR (cluster_n >= 4 AND asymmetry_score >= 0.35). "
         "The broader funnel for screening. Use to surface names that haven't "
         "made the cut on the tighter tiers but still register meaningful "
         "structural multibagger signals."),
        ("Archetypes (from archetype_tags.py)",
         "NarrativeLag (flat price + printed inflection) · FixedCost+DemandShock "
         "(heavy-asset sector with accel + margin expansion) · DiscountedVehicle "
         "(sub-book + cash > EV) · CapitalDiscipline (insider-aligned + lightly "
         "levered + durable margin + not re-rated) · RegimeCyclical (cyclical "
         "down 20%+ with inflection) · DeadOption (down 40%+ but FCF positive "
         "and lightly levered) · KPIThreshold (first-positive print + margin/ROCE "
         "confirming) · BlindSpot (blind-spot geography + small mcap) · "
         "MicroActivistInflect (microcap + inflection + clean BS + cheap)."),
        ("Inflection cluster (cluster_n, 7 signals)",
         "cheap_under_7x · first_positive (FCF/EBITDA/CFO/NI) · roce_inflect · "
         "fcf_eta_4q · growth_inflect (rev/EBITDA/FCF YoY) · accel_sales > 5pp · "
         "not_priced_in > 10pp."),
        ("ETA (entry-today)",
         "asymmetry_score x qual_mult (GREEN 1.10 / YELLOW 0.85 / RED 0.40) x "
         "post_rally_factor (smooth demotion of names already up >30% over 12m, "
         "floor 0.40 at +300%+). Used to rank within each tier."),
    ]
    row = 3
    for k, v in body:
        ws.cell(row=row, column=2, value=k).font = _font(size=11, bold=True, color=CRIMSON_DARK, name=SERIF)
        ws.cell(row=row, column=2).alignment = _align(v="top")
        ws.cell(row=row, column=3, value=v).font = _font(size=10, name=SERIF)
        ws.cell(row=row, column=3).alignment = _align(wrap=True, v="top")
        ws.row_dimensions[row].height = 64
        row += 1


def build_tier_sheet(ws, tier_name: str, df: pd.DataFrame):
    """Render a tier as a sortable table with Harvard number formatting."""
    # Cols: # | Tier | Ticker | Company | Cntry | Sector | Bucket | Mcap | Verdict |
    #       Arch# | Cluster | Archetype tags | Asym | ETA
    _set_col_widths(ws, {1: 4, 2: 5, 3: 8, 4: 12, 5: 36, 6: 6, 7: 16, 8: 10,
                         9: 14, 10: 10, 11: 7, 12: 9, 13: 50, 14: 10, 15: 10, 16: 4})
    _crimson_banner(ws, 1, f"  {tier_name} — {len(df):,} names", span_cols=15)

    headers = ['#', 'Tier', 'Ticker', 'Company', 'Cntry', 'Sector', 'Bucket',
               'Mcap (loc)', 'Verdict', 'Arch#', 'Cluster', 'Archetypes', 'Asym', 'ETA']
    for i, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=i, value=h)
        c.font = _font(size=10, bold=True, color=CRIMSON_DARK, name=SANS)
        if i in (4, 5, 7, 13):
            c.alignment = _align(h="left")
        elif i in (3, 6, 9):
            c.alignment = _align(h="center")
        else:
            c.alignment = _align(h="right") if i >= 8 else _align(h="center")
        c.border = _border(color=CRIMSON_DARK, bottom="medium")
    ws.row_dimensions[3].height = 22

    mono = _font(size=9, name=MONO)
    serif = _font(size=10, name=SERIF)
    sans = _font(size=10, name=SANS)
    sans_muted = _font(size=10, name=SANS, color=MUTED)

    for i, (_, r) in enumerate(df.iterrows(), start=1):
        row = 3 + i

        c2 = ws.cell(row=row, column=2, value=i)
        c2.font = _font(size=9, color=MUTED, name=SERIF)
        c2.alignment = _NUM_ALIGN_CENTER

        # Tier badge
        tcell = ws.cell(row=row, column=3, value=r['tier'])
        tcell.font = _font(size=9, bold=True, color=CRIMSON_DARK, name=SANS)
        tcell.alignment = _NUM_ALIGN_CENTER

        ws.cell(row=row, column=4, value=r['symbol']).font = _font(size=10, bold=True, name=SANS)
        ws.cell(row=row, column=4).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=row, column=5, value=str(r.get('name') or '')[:60]).font = serif
        ws.cell(row=row, column=5).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=row, column=6, value=r.get('src', '')).font = sans
        ws.cell(row=row, column=6).alignment = _NUM_ALIGN_CENTER
        ws.cell(row=row, column=7, value=str(r.get('sector') or '')).font = sans_muted
        ws.cell(row=row, column=7).alignment = _TXT_ALIGN_LEFT
        ws.cell(row=row, column=8, value=str(r.get('market_cap_bucket') or '')).font = sans_muted
        ws.cell(row=row, column=8).alignment = _NUM_ALIGN_CENTER

        _write_money(ws, row, 9, r.get('market_cap'), font=mono)
        _verdict_badge(ws, row, 10, r.get('verdict', 'UNRESEARCHED'))
        _write_int(ws, row, 11, int(r.get('archetype_count') or 0), font=mono)
        _write_int(ws, row, 12, int(r.get('cluster_n') or 0), font=mono)

        # Archetypes string (truncate to 70 chars)
        tags = str(r.get('archetype_tags_str') or '')[:70]
        ws.cell(row=row, column=13, value=tags).font = _font(size=8, name=SANS, color=MUTED)
        ws.cell(row=row, column=13).alignment = _TXT_ALIGN_LEFT

        _write_score(ws, row, 14, r.get('asymmetry_score'), font=mono)
        _write_score(ws, row, 15, r.get('eta'), font=mono)

        for c in range(2, 16):
            ws.cell(row=row, column=c).border = _border(color=RULE, bottom="thin")
        ws.row_dimensions[row].height = 16

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A4'
    # QoL: sortable/filterable table (header row 3 → last data row)
    from openpyxl.utils import get_column_letter as _gcl
    if ws.max_row >= 4:
        ws.auto_filter.ref = f"B3:{_gcl(ws.max_column)}{ws.max_row}"


def main():
    df = load_candidates()
    print(f'loaded {len(df):,} candidates', file=sys.stderr)

    wb = Workbook()

    cover = wb.active
    cover.title = "Cover"
    build_cover(cover, df)

    method = wb.create_sheet("Methodology")
    build_methodology(method)

    strict = df[df.tier == 'STRICT'].sort_values('eta', ascending=False)
    strong = df[df.tier == 'STRONG'].sort_values('eta', ascending=False)
    cand = df[df.tier == 'CANDIDATE'].sort_values('eta', ascending=False).head(500)

    s1 = wb.create_sheet("STRICT")
    build_tier_sheet(s1, "STRICT — high conviction", strict)

    s2 = wb.create_sheet("STRONG")
    build_tier_sheet(s2, "STRONG — watchlist", strong)

    s3 = wb.create_sheet("CANDIDATE_Top500")
    build_tier_sheet(s3, "CANDIDATE — broader funnel (top 500 by ETA)", cand)

    # Tab colors + hide gridlines
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = Color(rgb=CRIMSON)

    out = 'nms_multibagger_candidates.xlsx'
    wb.save(out)
    print(f'wrote {out}: {len(wb.worksheets)} sheets', file=sys.stderr)
    print(f'  STRICT: {len(strict):,}', file=sys.stderr)
    print(f'  STRONG: {len(strong):,}', file=sys.stderr)
    print(f'  CANDIDATE: {len(cand):,} (top 500 of {(df.tier=="CANDIDATE").sum():,})', file=sys.stderr)


if __name__ == '__main__':
    main()
